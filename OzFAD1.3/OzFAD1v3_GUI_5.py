# -*- coding: UTF-8 -*-
# #!/usr/bin/python3

#  TO TURN THIS FILE INTO A .EXE FILE RUN IN TERMINAL: pyinstaller --onefile -w OzFAD1v3_GUI_5.py
# EXE file will execute correctly, only when OzFAD_logo_HR.png is in the same folder as the exe file


#Jan Philipp Menzel 
#Notes: Testing GUI
import math
import datetime
import tkinter as tk
from tkinter import ttk
from PIL import ImageTk, Image
import webbrowser
import subprocess
import sys
import pandas as pd
import openpyxl
from openpyxl import Workbook

# colours 0 and 1 are colours of OzFAD logo
colouroptions=['#FFF2CC','#FD9D3D','blanched almond','snow','bisque','old lace','antique white','linen','papaya whip']

#from tkinter import *
#from tkinter.ttk import *

zframe=tk.Tk(className='~ OzFAD: Ozone-enabled Fatty Acid Discovery ~')	# window and name of the window
zframe.geometry('850x800+20+20')					# size of window
zframe['bg']='white'  #'blue' #'white'						# background colour of window

# add OzFAD logo to top left of window
canvas=tk.Canvas(zframe, width=100, height=100, bg='white', highlightthickness=0)
canvas.place(x=20, y=10, height=100, width=100)
ozfadlogo=(Image.open('OzFAD_logo_HR.png'))
resized_ozfadlogo=ozfadlogo.resize((98, 98), Image.ANTIALIAS)
new_ozfadlogo=ImageTk.PhotoImage(resized_ozfadlogo)
canvas.create_image(0, 0, anchor='nw', image=new_ozfadlogo)
#
ozfadlogo=tk.PhotoImage(file='OzFAD_logo_HR.png')		# set OzFAD logo as icon for window
zframe.iconphoto(False, ozfadlogo)						# set OzFAD logo as icon for window
#

# add arrows
aht=15	#15
awd=31	#30
xarrow=125
ac=tk.Canvas(zframe, width=awd, height=aht, bg='white', highlightthickness=0)
ac.place(x=xarrow, y=125, height=aht, width=awd)
ozfadarrow=(Image.open('OzFAD_gui_arrow_1.png'))
resized_ozfadarrow=ozfadarrow.resize((awd, aht), Image.ANTIALIAS)
anew_ozfadarrow=ImageTk.PhotoImage(resized_ozfadarrow)
ac.create_image(0, 0, anchor='nw', image=anew_ozfadarrow)
xarrow=xarrow+140

bc=tk.Canvas(zframe, width=awd, height=aht, bg='white', highlightthickness=0)
bc.place(x=xarrow, y=125, height=aht, width=awd)
ozfadarrow=(Image.open('OzFAD_gui_arrow_1.png'))
resized_ozfadarrow=ozfadarrow.resize((awd, aht), Image.ANTIALIAS)
bnew_ozfadarrow=ImageTk.PhotoImage(resized_ozfadarrow)
bc.create_image(0, 0, anchor='nw', image=bnew_ozfadarrow)
xarrow=xarrow+140

cc=tk.Canvas(zframe, width=awd, height=aht, bg='white', highlightthickness=0)
cc.place(x=xarrow, y=125, height=aht, width=awd)
ozfadarrow=(Image.open('OzFAD_gui_arrow_1.png'))
resized_ozfadarrow=ozfadarrow.resize((awd, aht), Image.ANTIALIAS)
cnew_ozfadarrow=ImageTk.PhotoImage(resized_ozfadarrow)
cc.create_image(0, 0, anchor='nw', image=cnew_ozfadarrow)
xarrow=xarrow+140

dc=tk.Canvas(zframe, width=awd, height=aht, bg='white', highlightthickness=0)
dc.place(x=xarrow, y=125, height=aht, width=awd)
ozfadarrow=(Image.open('OzFAD_gui_arrow_1.png'))
resized_ozfadarrow=ozfadarrow.resize((awd, aht), Image.ANTIALIAS)
dnew_ozfadarrow=ImageTk.PhotoImage(resized_ozfadarrow)
dc.create_image(0, 0, anchor='nw', image=dnew_ozfadarrow)
xarrow=xarrow+140

ec=tk.Canvas(zframe, width=awd, height=aht, bg='white', highlightthickness=0)
ec.place(x=xarrow, y=125, height=aht, width=awd)
ozfadarrow=(Image.open('OzFAD_gui_arrow_1.png'))
resized_ozfadarrow=ozfadarrow.resize((awd, aht), Image.ANTIALIAS)
enew_ozfadarrow=ImageTk.PhotoImage(resized_ozfadarrow)
ec.create_image(0, 0, anchor='nw', image=enew_ozfadarrow)

#lower arrow from Direct Infusion to Summary Table
fc=tk.Canvas(zframe, width=awd, height=aht, bg='white', highlightthickness=0)
fc.place(x=125, y=170, height=aht, width=awd)
ozfadarrow=(Image.open('OzFAD_gui_arrow_1.png'))
resized_ozfadarrow=ozfadarrow.resize((awd, aht), Image.ANTIALIAS)
fnew_ozfadarrow=ImageTk.PhotoImage(resized_ozfadarrow)
fc.create_image(0, 0, anchor='nw', image=fnew_ozfadarrow)

# wide arrow
ht=20
wd=720-140
gc=tk.Canvas(zframe, width=wd, height=ht, bg='white', highlightthickness=0)
gc.place(x=200, y=144, height=ht, width=wd)
ozfadarroww=(Image.open('OzFAD_gui_arrow_wide.png'))
resized_ozfadarroww=ozfadarroww.resize((wd, ht), Image.ANTIALIAS)
gnew_ozfadarrow=ImageTk.PhotoImage(resized_ozfadarroww)
gc.create_image(0, 0, anchor='nw', image=gnew_ozfadarrow)

# begin non variable part to be displayed always
def callback(url):
	webbrowser.open_new_tab(url)
lbl1t='OzFAD - Ozone-enabled Fatty Acid Discovery based on LC-OzID-MS/MS.'
lbl2t='This workflow employs the Skyline Mass Spectrometry Environment. For instructions, follow the link to the associated publication below.'
lbl3t='Created by: Dr. Jan Philipp Menzel, Mass Spectrometry Development Laboratory, Queensland University of Technology, 2022.' # add hyperlink
lbl4t='https://www.biorxiv.org/content/10.1101/2022.10.24.513604v1'	# Hyperlink to OzFAD preprint
lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'))
lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 8'))
lbl3=tk.Label(text=lbl3t, fg='black', bg='white', width=len(lbl3t), height=1, font=('Helvetica 8')) #, anchor='w')
lbl4=tk.Label(text=lbl4t, fg='blue', bg='white', width=len(lbl4t), height=1, font=('Helvetica 8 underline'), anchor='w', cursor='hand2')
lbl1.place(x=125, y=10, height=20, width=710)
lbl2.place(x=125, y=35, height=20, width=710)
lbl3.place(x=125, y=60, height=20, width=710)
lbl4.place(x=300, y=85, height=20, width=317)
lbl4.bind('<Button-1>', lambda e: callback('https://www.biorxiv.org/content/10.1101/2022.10.24.513604v1'))		# Hyperlink to OzFAD preprint
# end non variable part to be displayed always

derivlist=['AMPP', 'IAMP', 'NMPA', 'NMPE', 'MDPE', 'NEPE', 'EDPE', 'NPPE', 'PLPC', 'PLPE']

canh=640	# height of canvas, which hides other fields when switching between steps
cany=200	# y position of canves, which hides other fields when switching between steps

yposd=200	# y position of description
yposc=315	# y position of checklist
yposi=480	# y position of input

########################################################################################################################################### STEP 1  # 

#begin build page swap buttons and initialize different frames for the steps of the workflow
def frame1():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0) #bg='white'
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# Begin STEP 1 TARGETS ##################################################################################################################################
	# Begin variable part to be displayed for step 1

	lbl1t='Step 1: Initial analysis of DIA LC-OzID-MS data (precursor analysis) and target list creation.'
	lbl2t='  This step enables calculation of a target list for data-dependent acquisition (LC-OzID-MS/MS).'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'), anchor='w')
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  There is enough free disk space available, recommended is at least 10 GB.'
	lbl7t='  2.  The maximum retention time in the Transition Settings in the Skyline file template.sky is set according to the analysis.'
	lbl8t='  3.  The dataset (DIA raw data) is located in DIA_current_LCMS_dataset.'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	lbl8=tk.Label(text=lbl8t, fg='black', bg='white', width=len(lbl8t), height=1, font=('Helvetica 10'), anchor='w')
	lbl5.place(x=20, y=yposc, height=20, width=800)
	lbl6.place(x=20, y=yposc+25, height=20, width=800)
	lbl7.place(x=20, y=yposc+50, height=20, width=800)
	lbl8.place(x=20, y=yposc+75, height=20, width=800)
	
	dt=datetime.datetime.now()
	dt=str(dt)
	today=dt[0]+dt[1]+dt[2]+dt[3]+'_'+dt[5]+dt[6]+'_'+dt[8]+dt[9]+'_DIA_targets_'
	lbl9t='Enter the identifier for the sample:  '#+today
	lbl9t2=today
	lbl9=tk.Label(text=lbl9t, fg='black', bg='white', width=len(lbl9t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl92=tk.Label(text=lbl9t2, fg='black', bg='white', width=len(lbl9t2), height=1, font=('Helvetica 10'), anchor='w')
	lbl9.place(x=20, y=yposi, height=20, width=290)
	lbl92.place(x=250, y=yposi, height=20, width=160)	#OK
	# enter identifier
	entryid=tk.Entry(zframe, width=40, bg='white')
	entryid.focus_set()
	entryid.place(x=415, y=yposi, height=20, width=320) #OK

	#yposderiv=450
	#yposad=yposderiv+30
	lbl10t='Select the derivatization agent:'
	lbl10=tk.Label(text=lbl10t, fg='black', bg='white', width=len(lbl10t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl10.place(x=20, y=yposi+25, height=20, width=210)	#OK
	# Radiobuttons for common derivatization agents and alternative entry option

	def otherinput():
		# begin build alternative entry option
		entryoflc=tk.Entry(zframe, width=40, bg='white')
		entryoflc.focus_set()
		entryoflc.place(x=310, y=yposi+60, height=20, width=45) #OK
		#oflc='XXXX'

		def ask_sumformula():

			yposad=yposi+60
			oflc=entryoflc.get()
			other=str(oflc)
			if other=='':
				ok=1
			else:
				#print(other)
				other=other[0]+other[1]+other[2]+other[3]	#OK

			lbl=tk.Label(text='C', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=430, y=yposad, height=20, width=15)
			entrync=tk.Entry(zframe, width=40, bg='white')
			entrync.focus_set()
			entrync.place(x=447, y=yposad, height=20, width=20) #OK

			lbl=tk.Label(text='H', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=470, y=yposad, height=20, width=15)
			entrynh=tk.Entry(zframe, width=40, bg='white')
			entrynh.focus_set()
			entrynh.place(x=487, y=yposad, height=20, width=20) #OK

			lbl=tk.Label(text='N', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=510, y=yposad, height=20, width=15)
			entrynn=tk.Entry(zframe, width=40, bg='white')
			entrynn.focus_set()
			entrynn.place(x=527, y=yposad, height=20, width=20) #OK

			lbl=tk.Label(text='O', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=550, y=yposad, height=20, width=15)
			entryno=tk.Entry(zframe, width=40, bg='white')
			entryno.focus_set()
			entryno.place(x=567, y=yposad, height=20, width=20) #OK

			lbl=tk.Label(text='P', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=590, y=yposad, height=20, width=15)
			entrynp=tk.Entry(zframe, width=40, bg='white')
			entrynp.focus_set()
			entrynp.place(x=607, y=yposad, height=20, width=20) #OK

			lbl=tk.Label(text='I', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=630, y=yposad, height=20, width=15)
			entryni=tk.Entry(zframe, width=40, bg='white')
			entryni.focus_set()
			entryni.place(x=647, y=yposad, height=20, width=20) #OK

			lbl=tk.Label(text='D', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=670, y=yposad, height=20, width=15)
			entrynd=tk.Entry(zframe, width=40, bg='white')
			entrynd.focus_set()
			entrynd.place(x=687, y=yposad, height=20, width=20) #OK

		buttonlabel='Confirm'  #'Run' #
		ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=ask_sumformula).place(x=360, y=yposi+60, height=22, width=60)
		# end build alternative entry option

	park=0
	if park==0:
		i=0
		k=0
		m=0
		flc=tk.StringVar()
		values={'AMPP':'AMPP', 'IAMP':'IAMP', 'NMPA':'NMPA', 'NMPE':'NMPE', 'MDPE':'MDPE', 'NEPE':'NEPE', 'EDPE':'EDPE', 'NPPE':'NPPE', 'PLPC':'PLPC', 'PLPE':'PLPE', 'Other: ':'Other'}
		for (text, value) in values.items():
			if i>6:
				k=700
				m=30
			if i==10:
				tk.Radiobutton(zframe, text=text, variable=flc, value=value, bg=colouroptions[0], command=otherinput).place(x=240+i*70-k, y=yposi+30+m, height=20, width=60) 	# OK bg='white' 'lightgrey'
			else:
				tk.Radiobutton(zframe, text=text, variable=flc, value=value, bg=colouroptions[0]).place(x=240+i*70-k, y=yposi+30+m, height=20, width=60) 	# OK bg='white'
			i=i+1


	yposderiv=yposi+25
	lbl11t='Enter maximum retention time:'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl11.place(x=20, y=yposderiv+60, height=20, width=210)
	entryrt=tk.Entry(zframe, width=40, bg='white')
	entryrt.focus_set()
	entryrt.place(x=240, y=yposderiv+60, height=20, width=35) #OK
	lbl11t='min'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'))
	lbl11.place(x=280, y=yposderiv+60, height=20, width=25)

	lbl11t='Limit analysis to fatty acids with'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl11.place(x=20, y=yposderiv+90, height=20, width=210)
	entrycmin=tk.Entry(zframe, width=40, bg='white')
	entrycmin.focus_set()
	entrycmin.place(x=240, y=yposderiv+90, height=20, width=25) #OK
	lbl11t='up to'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'))
	lbl11.place(x=270, y=yposderiv+90, height=20, width=33)
	entrycmax=tk.Entry(zframe, width=40, bg='white')
	entrycmax.focus_set()
	entrycmax.place(x=310, y=yposderiv+90, height=20, width=25) #OK
	lbl11t='carbon atoms. (Min 4 to max 40, recommended is 12 to 24)'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'))
	lbl11.place(x=340, y=yposderiv+90, height=20, width=370)

	lbl11t='Enter maximum m/z error:'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold bold'), anchor='w')
	lbl11.place(x=20, y=yposderiv+120, height=20, width=210)
	entrymzr=tk.Entry(zframe, width=40, bg='white')
	entrymzr.focus_set()
	entrymzr.place(x=240, y=yposderiv+120, height=20, width=25) #OK
	lbl11t='ppm (Recommended: 60)'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'))
	lbl11.place(x=270, y=yposderiv+120, height=20, width=160)

	lbl11t='Enter intensity threshold for detection of precursor peaks:'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl11.place(x=20, y=yposderiv+150, height=20, width=375)
	entrythr=tk.Entry(zframe, width=40, bg='white')
	entrythr.focus_set()
	entrythr.place(x=405, y=yposderiv+150, height=20, width=35) #OK
	lbl11t='(Recommended: 150)'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'))
	lbl11.place(x=445, y=yposderiv+150, height=20, width=160)


	def run_targets():
		global entry
		rawidentifier=entryid.get()
		fourlettcode=flc.get()
		lastexrt=entryrt.get()
		minlenfa=entrycmin.get()
		maxlenfa=entrycmax.get()
		mzcutoff=entrymzr.get()
		largeareathreshold=entrythr.get()
		dt=datetime.datetime.now()
		dt=str(dt)
		today=dt[0]+dt[1]+dt[2]+dt[3]+'_'+dt[5]+dt[6]+'_'+dt[8]+dt[9]+'_DIA_targets_'+fourlettcode+'_'
		identifier=today+rawidentifier
		identifier=str(identifier)
		fourlettcode=str(fourlettcode)
		lastexrt=str(lastexrt)
		minlenfa=str(minlenfa)
		maxlenfa=str(maxlenfa)
		mzcutoff=str(mzcutoff)
		largeareathreshold=str(largeareathreshold)
		if fourlettcode=='AMPP':
			cderiv=12
			hderiv=12
			dderiv=0
			nderiv=2
			oderiv=0
			pderiv=0
			ideriv=0
		elif fourlettcode=='NMPA':
			cderiv=7
			hderiv=10
			dderiv=0
			nderiv=2
			oderiv=0
			pderiv=0
			ideriv=0
		elif fourlettcode=='NMPE':
			cderiv=7
			hderiv=9
			dderiv=0
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='MDPE':
			cderiv=7
			hderiv=6
			dderiv=3
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='NEPE':
			cderiv=8
			hderiv=11
			dderiv=0
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='EDPE':
			cderiv=6
			hderiv=6
			dderiv=5
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='NPPE':
			cderiv=9
			hderiv=13
			dderiv=0
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='IAMP':
			cderiv=12
			hderiv=11
			dderiv=0
			nderiv=2
			oderiv=0
			pderiv=0
			ideriv=1
		elif fourlettcode=='PLPC':
			cderiv=8
			hderiv=20
			dderiv=0
			nderiv=1
			oderiv=6
			pderiv=1
			ideriv=0
		elif fourlettcode=='PLPE':
			cderiv=5
			hderiv=14
			dderiv=0
			nderiv=1
			oderiv=6
			pderiv=1
			ideriv=0
		else:
			cderiv=str(entrync.get())
			hderiv=str(entrynh.get())
			dderiv=str(entrynd.get())
			nderiv=str(entrynn.get())
			oderiv=str(entryno.get())
			pderiv=str(entrynp.get())
			ideriv=str(entryni.get())


		arglist=[]
		arglist.append(identifier)
		arglist.append(fourlettcode)
		arglist.append(lastexrt)
		arglist.append(minlenfa)
		arglist.append(maxlenfa)
		arglist.append(mzcutoff)
		arglist.append(largeareathreshold)
		#label.configure(text='Calculating ...')
		#label9.configure(text=identifier)

		# writing in OzFAD_workflow_parameters.xlsx 
		# begin save workflow parameters in csv file (to be used in following python steps during automated workflow)
		rettimecutoff=lastexrt
		productareathreshold=100
		nchunks=1
		rtlimitation=2
		mostwanted=1
		transtest=0
		runprecheck=1
		workflow=1
		workflowidentifier=identifier
		transferindexlist=['Four letter code of derivatization agent', 'C', 'H', 'D', 'N', 'O', 'P', 'I', 'min FA length', 'max FA length', 'm/z filter tolerance [ppm]', 'max RT [min]', 'Precursor area threshold', 'Workflow mode [1-Slow/full; 2-Fast/limited]', 'Product area threshold', 'Number of XIC files', 'RT limitation [1-Y; 0-N]', 'Library inclusion [1-Y; 0-N]', '-2H transition [1-Y; 0-N]', 'Pre-check [1-Y; 0-N]', 'Identifier']
		transferlist=[]
		transferlist.append(fourlettcode)	#0
		transferlist.append(cderiv)
		transferlist.append(hderiv)
		transferlist.append(dderiv)
		transferlist.append(nderiv)
		transferlist.append(oderiv)	#5
		transferlist.append(pderiv)
		transferlist.append(ideriv)
		transferlist.append(minlenfa)
		transferlist.append(maxlenfa)
		transferlist.append(mzcutoff)	#10
		transferlist.append(rettimecutoff)
		transferlist.append(largeareathreshold)
		transferlist.append(workflow)
		transferlist.append(productareathreshold)
		transferlist.append(nchunks)	#15
		transferlist.append(rtlimitation)
		transferlist.append(mostwanted)
		transferlist.append(transtest)
		transferlist.append(runprecheck)
		transferlist.append(workflowidentifier)	#20
		#print(transferlist)
		wb = Workbook(write_only=True)
		ws = wb.create_sheet()
		wb.save('OzFAD1_workflow_parameters.xlsx')
		wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
		ws=wb.active
		tli=0
		while tli<len(transferindexlist):
			ws.cell(row=tli+1, column=1).value=transferindexlist[tli]
			ws.cell(row=tli+1, column=2).value=transferlist[tli]
			tli=tli+1

		wb.save('OzFAD1_workflow_parameters.xlsx')
		# end save workflow parameters in excel file (to be used in following python steps during automated workflow)

		# THIS WORKS LOCALLY
		#subprocess.call([r'C:\Users\menzel2\OzFAD1.3\OzFAD1_1_DIA_Targets.bat', identifier]) #, fourlettcode, lastexrt, minlenfa, maxlenfa, mzcutoff, largeareathreshold])		# 
		# RELATIVE PATH
		subprocess.call([r'.\OzFAD1_1_DIA_Targets.bat', identifier]) # OK
		#end initiate running batch file and send values to batch file for running correctly

	buttonlabel='Build Target List'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_targets).place(x=320, y=yposi+230, height=25, width=120)

	label9=tk.Label(zframe, text='', font=('Helvetica 10 bold'), bg='white') #OK	#display complete identifier
	label9.place(x=180, y=790, height=20, width=400)
	# End STEP 1 TARGETS ##################################################################################################################################

########################################################################################################################################### STEP 1  # 
########################################################################################################################################### STEP 2  # # 

def frame2():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0)	# bg='white'
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# build variable labels and buttons
	# Begin variable part to be displayed for step 1

	# Begin STEP 2 DISCOVERY ##################################################################################################################################
	lbl1t='Step 2: Discovery analysis of DDA LC-OzID-MS/MS data. This step can be performed in three modes.'
	lbl2t='   - Full discovery: A fully exhaustive search of all double bond positions and patterns.'
	lbl3t='   - Library-only: A search for all fatty acid species listed in the associated excel library jpm_fa_lib.xlsx.'
	lbl4t='   - Streamlined: A fully exhaustive search for all fatty acids with up to 3 double bonds, otherwise library-based.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'), anchor='w')
	lbl3=tk.Label(text=lbl3t, fg='black', bg='white', width=len(lbl3t), height=1, font=('Helvetica 10'), anchor='w')
	lbl4=tk.Label(text=lbl4t, fg='black', bg='white', width=len(lbl4t), height=1, font=('Helvetica 10'), anchor='w')
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)
	lbl3.place(x=20, y=yposd+50, height=20, width=700)
	lbl4.place(x=20, y=yposd+75, height=20, width=743)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  There is enough free disk space available, recommended is at least 10 GB.'
	lbl7t='  2.  The maximum retention time in the Transition Settings in the Skyline file template.sky is set according to the analysis.'
	lbl8t='  3.  The dataset (DIA raw data) is located in DIA_current_LCMS_dataset.'
	lbl9t='  4.  The dataset (DDA raw data) is located in DDA_current_LCMS_dataset.'
	lbl10t='  5.  The file OzFAD1_DDA_targetlist.txt and OzFAD1_workflow_parameters.xlsx and is located in OzFAD1.3.'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	lbl8=tk.Label(text=lbl8t, fg='black', bg='white', width=len(lbl8t), height=1, font=('Helvetica 10'), anchor='w')
	lbl9=tk.Label(text=lbl9t, fg='black', bg='white', width=len(lbl9t), height=1, font=('Helvetica 10'), anchor='w')
	lbl10=tk.Label(text=lbl10t, fg='black', bg='white', width=len(lbl10t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=820)
	lbl6.place(x=20, y=yposc+25, height=20, width=820)
	lbl7.place(x=20, y=yposc+50, height=20, width=820)
	lbl8.place(x=20, y=yposc+75, height=20, width=820)
	lbl9.place(x=20, y=yposc+100, height=20, width=820)
	lbl10.place(x=20, y=yposc+125, height=20, width=820)

	yposid=yposi#+25

	dt=datetime.datetime.now()
	dt=str(dt)
	today=dt[0]+dt[1]+dt[2]+dt[3]+'_'+dt[5]+dt[6]+'_'+dt[8]+dt[9]+'_DDA_discovery_'
	lbl9t='Enter the identifier for the sample:  '#+today
	lbl9t2=today
	lbl9=tk.Label(text=lbl9t, fg='black', bg='white', width=len(lbl9t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl92=tk.Label(text=lbl9t2, fg='black', bg='white', width=len(lbl9t2), height=1, font=('Helvetica 10'))
	lbl9.place(x=20, y=yposid, height=20, width=290)
	lbl92.place(x=250, y=yposid, height=20, width=180)	#OK
	# enter identifier
	entryidd=tk.Entry(zframe, width=40, bg='white')
	entryidd.focus_set()
	entryidd.place(x=430, y=yposid, height=20, width=320) #OK

	lbl10t='Select the workflow mode:'
	lbl10=tk.Label(text=lbl10t, fg='black', bg='white', width=len(lbl10t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl10.place(x=20, y=yposid+25, height=20, width=200)	#OK
	# Radiobuttons for common derivatization agents and alternative entry option
	park=0
	if park==0:
		i=0
		k=0
		m=0
		dlvl=tk.StringVar()
		values={'Fully exhaustive search':'0', 'Library-only search':'2', 'Streamlined search':'1'}
		for (text, value) in values.items():
			if i>6:
				k=700
				m=30
			tk.Radiobutton(zframe, text=text, variable=dlvl, value=value, bg=colouroptions[0]).place(x=255+i*170-k, y=yposid+25+m, height=20, width=160) 	# OK
			i=i+1

	def run_discovery():
		global entry
		rawidentifierd=entryidd.get()
		dlevel=str(dlvl.get())
		dt=datetime.datetime.now()
		dt=str(dt)
		today=dt[0]+dt[1]+dt[2]+dt[3]+'_'+dt[5]+dt[6]+'_'+dt[8]+dt[9]+'_DDA_discovery_' #+fourlettcode+'_'
		identifierd=today+rawidentifierd
		# begin run analysis - run batch file in subprocess and pass identifier
		#subprocess.call([r'C:\Users\menzel2\OzFAD1.3\OzFAD1_2_DDA_Discovery.bat', identifierd, dlevel])
		subprocess.call([r'.\OzFAD1_2_DDA_Discovery.bat', identifierd, dlevel])
		# end run analysis - run batch file in subprocess and pass identifier

	buttonlabel='Discover fatty acids'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_discovery).place(x=320, y=yposid+75, height=25, width=150)

########################################################################################################################################### STEP 2  # # 
########################################################################################################################################### STEP 3  # # # 

def frame3():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=640, bg='white', highlightthickness=0)  # bg='white'
	canvas2.place(x=0, y=cany, width=850, height=640)
	# build variable labels and buttons

	# Begin STEP 3 S/N ##################################################################################################################################
	lbl1t='Step 3: Determination of S/N values for selected MS/MS spectra.'
	lbl2t='  This algorithm reads an excel file with MS/MS spectra (Skyline export), combines spectra if applicable and determines S/N values.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'), anchor='w')
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  The selected MS/MS spectra are in file OzFAD1_2_MSMS_input.xlsx. in folder OzFAD1.3'
	lbl7t='  2.  Spectra (Exported from Skyline MS view), scaled to full view - m/z = 100 to 1200, are in fields A1, D1, G1, ...'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=800)
	lbl6.place(x=20, y=yposc+25, height=20, width=800)
	lbl7.place(x=20, y=yposc+50, height=20, width=800)

	def run_sn():
		# begin run analysis - run batch file in subprocess and pass identifier
		#subprocess.call([sys.executable, 'OzFAD1_MSMS_SN.py'])	#works only, if python script in OzFAD1.3 #BACKUP
		#subprocess.call([sys.executable, 'C:/Users/menzel2/OzFAD1.3/OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_MSMS_SN.py'])	#OK, ONLY LOCALLY
		#subprocess.call([sys.executable, './OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_MSMS_SN.py'])	## OK ???
		subprocess.call([r'.\OzFAD1_3_MSMS_SN.bat'])
		# end run analysis - run batch file in subprocess and pass identifier

	buttonlabel='Determine S/N'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_sn).place(x=320, y=yposc+100, height=25, width=140)

########################################################################################################################################### STEP 3  # # #  
########################################################################################################################################### STEP 4  # # #  # 

def frame4():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0)  # bg='white'
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# build variable labels and buttons

	# Begin STEP 4 DDA to DIA ##################################################################################################################################
	lbl1t='Step 4: Analysis of LC-OzID-MS data for quantification.'
	lbl2t='  This step creates a Skyline file with the identified fatty acids for quantification based on DIA LC-OzID-MS.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'), anchor='w')
	lbl1.place(x=20, y=yposd, height=20, width=350)
	lbl2.place(x=20, y=yposd+25, height=20, width=650)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  There is enough free disk space available, recommended is at least 10 GB.'
	lbl7t='  2.  The maximum retention time in the Transition Settings in the Skyline file template.sky is set according to the analysis.'
	lbl8t='  3.  The dataset (DIA raw data) is located in DIA_current_LCMS_dataset.'
	lbl9t='  4.  The dataset (DDA raw data) is located in DDA_current_LCMS_dataset.'
	lbl11t='  5.  The file skyl_report_dda_found.csv and OzFAD1_workflow_parameters.xlsx is located in OzFAD1.3.'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	lbl8=tk.Label(text=lbl8t, fg='black', bg='white', width=len(lbl8t), height=1, font=('Helvetica 10'), anchor='w')
	lbl9=tk.Label(text=lbl9t, fg='black', bg='white', width=len(lbl9t), height=1, font=('Helvetica 10'), anchor='w')
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=800)
	lbl6.place(x=20, y=yposc+25, height=20, width=800)
	lbl7.place(x=20, y=yposc+50, height=20, width=800)
	lbl8.place(x=20, y=yposc+75, height=20, width=800)
	lbl9.place(x=20, y=yposc+100, height=20, width=800)
	lbl11.place(x=20, y=yposc+125, height=20, width=800)

	dt=datetime.datetime.now()
	dt=str(dt)
	today=dt[0]+dt[1]+dt[2]+dt[3]+'_'+dt[5]+dt[6]+'_'+dt[8]+dt[9]+'_DDA_to_DIA_'
	lbl9t='Enter the identifier for the sample:'
	lbl9t2=today
	lbl9=tk.Label(text=lbl9t, fg='black', bg='white', width=len(lbl9t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl92=tk.Label(text=lbl9t2, fg='black', bg='white', width=len(lbl9t2), height=1, font=('Helvetica 10'))
	lbl9.place(x=20, y=yposi, height=20, width=225)
	lbl92.place(x=250, y=yposi, height=20, width=180)	#OK
	# enter identifier
	entryida=tk.Entry(zframe, width=40, bg='white')
	entryida.focus_set()
	entryida.place(x=420, y=yposi, height=20, width=320) #OK

	def dda_dia():
		global entry
		rawidentifiera=entryida.get()
		dt=datetime.datetime.now()
		dt=str(dt)
		today=dt[0]+dt[1]+dt[2]+dt[3]+'_'+dt[5]+dt[6]+'_'+dt[8]+dt[9]+'_DDA_to_DIA_'#+fourlettcode+'_'
		identifiera=today+rawidentifiera
		# begin run analysis - run batch file in subprocess and pass identifier
		#subprocess.call([r'C:\Users\menzel2\OzFAD1.3\OzFAD1_3_DDA_to_DIA.bat', identifiera])	# OK, LOCALLY
		subprocess.call([r'.\OzFAD1_4_DDA_to_DIA.bat', identifiera])
		# end run analysis - run batch file in subprocess and pass identifier

	buttonlabel='Generate DIA Skyline file for quantification'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=dda_dia).place(x=250, y=yposi+50, height=25, width=280)

########################################################################################################################################### STEP 4  # # #  # 
########################################################################################################################################### STEP 5  # # #  # #

def frame5():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0)
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# build variable labels and buttons

	# Begin STEP 4 Deconvolution ##################################################################################################################################
	lbl1t='Step 5: Deconvolution of precursor extracted ion chromatograms of LC-OzID-MS data for quantification.'
	lbl2t='  This step creates an excel file for quantification based on deconvoluted precursor XICs and OzID product ion XICs of DIA LC-OzID-MS data.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'))
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  There is enough free disk space available, recommended is at least 10 GB.'
	lbl7t='  2.  The maximum retention time in the Transition Settings in the Skyline file template.sky is set according to the analysis.'
	lbl8t='  3.  The dataset (DIA raw data) is located in DIA_current_LCMS_dataset.'
	lbl9t='  4.  The dataset (DDA raw data) is located in DDA_current_LCMS_dataset.'
	lbl11t='  5.  The files skyl_report_dia_int.csv, skyl_report_dia_xic.tsv and OzFAD1_workflow_parameters are located in OzFAD1.3.'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	lbl8=tk.Label(text=lbl8t, fg='black', bg='white', width=len(lbl8t), height=1, font=('Helvetica 10'), anchor='w')
	lbl9=tk.Label(text=lbl9t, fg='black', bg='white', width=len(lbl9t), height=1, font=('Helvetica 10'), anchor='w')
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=810)
	lbl6.place(x=20, y=yposc+25, height=20, width=810)
	lbl7.place(x=20, y=yposc+50, height=20, width=810)
	lbl8.place(x=20, y=yposc+75, height=20, width=810)
	lbl9.place(x=20, y=yposc+100, height=20, width=810)
	lbl11.place(x=20, y=yposc+125, height=20, width=810)

	dt=datetime.datetime.now()
	dt=str(dt)
	today=dt[0]+dt[1]+dt[2]+dt[3]+'_'+dt[5]+dt[6]+'_'+dt[8]+dt[9]+'_DIA_deconvolution_'
	lbl9t='Enter the identifier for the sample:  '
	lbl9t2=today
	lbl9=tk.Label(text=lbl9t, fg='black', bg='white', width=len(lbl9t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl92=tk.Label(text=lbl9t2, fg='black', bg='white', width=len(lbl9t2), height=1, font=('Helvetica 10'))
	lbl9.place(x=20, y=yposi, height=20, width=290)
	lbl92.place(x=250, y=yposi, height=20, width=220)	#OK
	# enter identifier
	entryidc=tk.Entry(zframe, width=40, bg='white')
	entryidc.focus_set()
	entryidc.place(x=460, y=yposi, height=20, width=320) #OK

	def run_decon():
		global entry
		rawidentifierc=entryidc.get()
		dt=datetime.datetime.now()
		dt=str(dt)
		today=dt[0]+dt[1]+dt[2]+dt[3]+'_'+dt[5]+dt[6]+'_'+dt[8]+dt[9]+'_DIA_deconvolution_'#+fourlettcode+'_'
		identifierc=today+rawidentifierc
		# begin run analysis - run batch file in subprocess and pass identifier
		#subprocess.call([r'C:\Users\menzel2\OzFAD1.3\OzFAD1_4_Deconvolution.bat', identifierc]) #OK LOCALLY
		subprocess.call([r'.\OzFAD1_5_Deconvolution.bat', identifierc]) 
		# end run analysis - run batch file in subprocess and pass identifier

	buttonlabel='Generate excel file for deconvolution'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_decon).place(x=270, y=yposi+50, height=25, width=260)

########################################################################################################################################### STEP 5  # # #  # # 
########################################################################################################################################### STEP 6  # # #  # # #

def frame6():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0)  # bg='white'
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# build variable labels and buttons

	# Begin STEP 3 S/N ##################################################################################################################################
	lbl1t='Step 6: Preliminary Plot and Table.'
	lbl2t='  This algorithm reads excel files with deconvolution parameters and quantification data, and generates a preliminary plot and table.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'))
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  The file OzFAD1_4_input_DIA_Q.xlsx and OzFAD1_4_DIA_deconv.xlsx are in folder OzFAD1.3'
	#lbl7t='  2.  Spectrum, scaled to full view - m/z = 100 to 1200, is in field A5.'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	#lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=800)
	lbl6.place(x=20, y=yposc+25, height=20, width=800)
	#lbl7.place(x=20, y=285, height=20, width=800)

	lbl10t='Select the derivatization agent:'
	lbl10=tk.Label(text=lbl10t, fg='black', bg='white', width=len(lbl10t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl10.place(x=20, y=yposi, height=20, width=210)	#OK
	# Radiobuttons for common derivatization agents and alternative entry option

	def otherinput():
		# begin build alternative entry option
		entryoflc=tk.Entry(zframe, width=40, bg='white')
		entryoflc.focus_set()
		entryoflc.place(x=310, y=yposi+30, height=20, width=45) #OK
		#oflc='XXXX'

		def ask_sumformula():
			oflc=entryoflc.get()
			other=str(oflc)
			if other=='':
				ok=1
			else:
				#print(other)
				other=other[0]+other[1]+other[2]+other[3]	#OK

			lbl=tk.Label(text='C', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=430, y=yposi+30, height=20, width=15)
			entrync=tk.Entry(zframe, width=40, bg='white')
			entrync.focus_set()
			entrync.place(x=447, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='H', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=470, y=yposi+30, height=20, width=15)
			entrynh=tk.Entry(zframe, width=40, bg='white')
			entrynh.focus_set()
			entrynh.place(x=487, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='N', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=510, y=yposi+30, height=20, width=15)
			entrynn=tk.Entry(zframe, width=40, bg='white')
			entrynn.focus_set()
			entrynn.place(x=527, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='O', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=550, y=yposi+30, height=20, width=15)
			entryno=tk.Entry(zframe, width=40, bg='white')
			entryno.focus_set()
			entryno.place(x=567, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='P', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=590, y=yposi+30, height=20, width=15)
			entrynp=tk.Entry(zframe, width=40, bg='white')
			entrynp.focus_set()
			entrynp.place(x=607, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='I', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=630, y=yposi+30, height=20, width=15)
			entryni=tk.Entry(zframe, width=40, bg='white')
			entryni.focus_set()
			entryni.place(x=647, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='D', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=670, y=yposi+30, height=20, width=15)
			entrynd=tk.Entry(zframe, width=40, bg='white')
			entrynd.focus_set()
			entrynd.place(x=687, y=yposi+30, height=20, width=20) #OK

		buttonlabel='Confirm'  #'Run' #
		ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=ask_sumformula).place(x=360, y=yposi+30, height=22, width=60)
		# end build alternative entry option

	i=0
	k=0
	m=0
	flc=tk.StringVar()
	values={'AMPP':'AMPP', 'IAMP':'IAMP', 'NMPA':'NMPA', 'NMPE':'NMPE', 'MDPE':'MDPE', 'NEPE':'NEPE', 'EDPE':'EDPE', 'NPPE':'NPPE', 'PLPC':'PLPC', 'PLPE':'PLPE', 'Other: ':'Other'}
	for (text, value) in values.items():
		if i>6:
			k=700
			m=30
		if i==10:
			tk.Radiobutton(zframe, text=text, variable=flc, value=value, bg=colouroptions[0], command=otherinput).place(x=240+i*70-k, y=yposi+m, height=20, width=60) 	# OK bg='white'
		else:
			tk.Radiobutton(zframe, text=text, variable=flc, value=value, bg=colouroptions[0]).place(x=240+i*70-k, y=yposi+m, height=20, width=60) 	# OK bg='white'
		i=i+1

	def run_pt():
		# begin run analysis - run batch file in subprocess and pass identifier
		fourlettcode=str(flc.get())
		if fourlettcode=='AMPP':
			cderiv=12
			hderiv=12
			dderiv=0
			nderiv=2
			oderiv=0
			pderiv=0
			ideriv=0
		elif fourlettcode=='NMPA':
			cderiv=7
			hderiv=10
			dderiv=0
			nderiv=2
			oderiv=0
			pderiv=0
			ideriv=0
		elif fourlettcode=='NMPE':
			cderiv=7
			hderiv=9
			dderiv=0
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='MDPE':
			cderiv=7
			hderiv=6
			dderiv=3
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='NEPE':
			cderiv=8
			hderiv=11
			dderiv=0
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='EDPE':
			cderiv=6
			hderiv=6
			dderiv=5
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='NPPE':
			cderiv=9
			hderiv=13
			dderiv=0
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='IAMP':
			cderiv=12
			hderiv=11
			dderiv=0
			nderiv=2
			oderiv=0
			pderiv=0
			ideriv=1
		elif fourlettcode=='PLPC':
			cderiv=8
			hderiv=20
			dderiv=0
			nderiv=1
			oderiv=6
			pderiv=1
			ideriv=0
		elif fourlettcode=='PLPE':
			cderiv=5
			hderiv=14
			dderiv=0
			nderiv=1
			oderiv=6
			pderiv=1
			ideriv=0
		else:
			cderiv=str(entrync.get())
			hderiv=str(entrynh.get())
			dderiv=str(entrynd.get())
			nderiv=str(entrynn.get())
			oderiv=str(entryno.get())
			pderiv=str(entrynp.get())
			ideriv=str(entryni.get())

		cderiv=str(cderiv)
		hderiv=str(hderiv)
		dderiv=str(dderiv)
		nderiv=str(nderiv)
		oderiv=str(oderiv)
		pderiv=str(pderiv)
		ideriv=str(ideriv)

		arglist=[]
		arglist.append(fourlettcode)
		arglist.append(str(cderiv))
		arglist.append(str(hderiv))
		arglist.append(str(dderiv))
		arglist.append(str(nderiv))
		arglist.append(str(oderiv))
		arglist.append(str(pderiv))
		arglist.append(str(ideriv))
		#print(arglist)

		#subprocess.call([sys.executable, 'OzFAD1_MSMS_SN.py'])	#works only, if python script in OzFAD1.3
		
		#OK LOCALLY
		#subprocess.call([sys.executable, 'C:/Users/menzel2/OzFAD1.3/OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_5_Plot_Table.py', fourlettcode, cderiv, hderiv, dderiv, nderiv, oderiv, pderiv, ideriv])	#OK
		
		subprocess.call([sys.executable, './OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_5_Plot_Table.py', fourlettcode, cderiv, hderiv, dderiv, nderiv, oderiv, pderiv, ideriv])	#OK
		subprocess.call([r'.\OzFAD1_6_Plot_Table.bat', fourlettcode, cderiv, hderiv, dderiv, nderiv, oderiv, pderiv, ideriv])
		# end run analysis - run batch file in subprocess and pass identifier

	buttonlabel='Generate preliminary Plot and Table'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_pt).place(x=260, y=yposi+80, height=25, width=270)

########################################################################################################################################### STEP 6  # # #  # # #
########################################################################################################################################### STEP 7  # # #  # # #  # 

def frame7():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0)  # bg='white'
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# build variable labels and buttons

	# Begin STEP 3 S/N ##################################################################################################################################
	lbl1t='Step 9: Replicate Plot.'
	lbl2t='  This algorithm reads three excel files with relative quantification data, and generates a replicate barchart.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'), anchor='w')
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  The files OzFAD1_5_plot_table_rep1.xlsx ...rep2.xlsx and ...rep3.xlsx are in folder OzFAD1.3'
	#lbl7t='  2.  Spectrum, scaled to full view - m/z = 100 to 1200, is in field A5.'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	#lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=800)
	lbl6.place(x=20, y=yposc+25, height=20, width=800)
	#lbl7.place(x=20, y=285, height=20, width=800)

	def run_rep_bc():
		#subprocess.call([sys.executable, 'C:/Users/menzel2/OzFAD1.3/OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_6_Replicate_plot.py'])	#OK LOCALLY
		#subprocess.call([sys.executable, './OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_6_Replicate_plot.py'])
		subprocess.call([r'.\OzFAD1_7_Replicate_Barchart.bat'])

	buttonlabel='Generate Replicate Bar Chart'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_rep_bc).place(x=280, y=yposi+25, height=25, width=230)

########################################################################################################################################### STEP 7  # # #  # # #  # 
########################################################################################################################################### STEP 8  # # #  # # #  # #

def frame8():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0)  # bg='white'
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# build variable labels and buttons

	# Begin STEP 3 S/N ##################################################################################################################################
	lbl1t='Step 7: Analysis of direct infusion mass spectra.'
	lbl2t='  This algorithm reads an excel file with a direct infusion spectrum (MassLynx export), and determines isotope corrected integration values.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'))
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  The selected spectrum is in file OzFAD1_raw_ms_di.xlsx. in folder OzFAD1.3'
	lbl7t='  2.  Spectrum, scaled to full view - m/z = 100 to 1200, is in field A5.'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=800)
	lbl6.place(x=20, y=yposc+25, height=20, width=800)
	lbl7.place(x=20, y=yposc+50, height=20, width=800)

	lbl10t='Select the derivatization agent:'
	lbl10=tk.Label(text=lbl10t, fg='black', bg='white', width=len(lbl10t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl10.place(x=20, y=yposi, height=20, width=210)	#OK
	# Radiobuttons for common derivatization agents and alternative entry option

	def otherinput():
		# begin build alternative entry option
		entryoflc=tk.Entry(zframe, width=40, bg='white')
		entryoflc.focus_set()
		entryoflc.place(x=310, y=yposi+30, height=20, width=45) #OK
		#oflc='XXXX'

		def ask_sumformula():
			oflc=entryoflc.get()
			other=str(oflc)
			if other=='':
				ok=1
			else:
				#print(other)
				other=other[0]+other[1]+other[2]+other[3]	#OK

			lbl=tk.Label(text='C', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=430, y=yposi+30, height=20, width=15)
			entrync=tk.Entry(zframe, width=40, bg='white')
			entrync.focus_set()
			entrync.place(x=447, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='H', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=470, y=yposi+30, height=20, width=15)
			entrynh=tk.Entry(zframe, width=40, bg='white')
			entrynh.focus_set()
			entrynh.place(x=487, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='N', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=510, y=350, height=20, width=15)
			entrynn=tk.Entry(zframe, width=40, bg='white')
			entrynn.focus_set()
			entrynn.place(x=527, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='O', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=550, y=yposi+30, height=20, width=15)
			entryno=tk.Entry(zframe, width=40, bg='white')
			entryno.focus_set()
			entryno.place(x=567, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='P', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=590, y=yposi+30, height=20, width=15)
			entrynp=tk.Entry(zframe, width=40, bg='white')
			entrynp.focus_set()
			entrynp.place(x=607, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='I', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=630, y=yposi+30, height=20, width=15)
			entryni=tk.Entry(zframe, width=40, bg='white')
			entryni.focus_set()
			entryni.place(x=647, y=yposi+30, height=20, width=20) #OK

			lbl=tk.Label(text='D', fg='black', bg='white', width=20, height=1, font=('Helvetica 10'), anchor='w')
			lbl.place(x=670, y=yposi+30, height=20, width=15)
			entrynd=tk.Entry(zframe, width=40, bg='white')
			entrynd.focus_set()
			entrynd.place(x=687, y=yposi+30, height=20, width=20) #OK

		buttonlabel='Confirm'  #'Run' #
		ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=ask_sumformula).place(x=360, y=yposi+30, height=22, width=60)
		# end build alternative entry option

	i=0
	k=0
	m=0
	flc=tk.StringVar()
	values={'AMPP':'AMPP', 'IAMP':'IAMP', 'NMPA':'NMPA', 'NMPE':'NMPE', 'MDPE':'MDPE', 'NEPE':'NEPE', 'EDPE':'EDPE', 'NPPE':'NPPE', 'PLPC':'PLPC', 'PLPE':'PLPE', 'Other: ':'Other'}
	for (text, value) in values.items():
		if i>6:
			k=700
			m=30
		if i==10:
			tk.Radiobutton(zframe, text=text, variable=flc, value=value, bg=colouroptions[0], command=otherinput).place(x=240+i*70-k, y=yposi+m, height=20, width=60) 	# OK bg='white'
		else:
			tk.Radiobutton(zframe, text=text, variable=flc, value=value, bg=colouroptions[0]).place(x=240+i*70-k, y=yposi+m, height=20, width=60) 	# OK bg='white'
		i=i+1

	lbl11t='Limit analysis to fatty acids with'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl11.place(x=20, y=yposi+60, height=20, width=210)
	entrycmin=tk.Entry(zframe, width=40, bg='white')
	entrycmin.focus_set()
	entrycmin.place(x=240, y=yposi+60, height=20, width=25) #OK
	lbl11t='up to'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'))
	lbl11.place(x=270, y=yposi+60, height=20, width=33)
	entrycmax=tk.Entry(zframe, width=40, bg='white')
	entrycmax.focus_set()
	entrycmax.place(x=308, y=yposi+60, height=20, width=25) #OK
	lbl11t='carbon atoms. (Min 4 to max 40, recommended is 12 to 24)'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'))
	lbl11.place(x=345, y=yposi+60, height=20, width=370)

	lbl11t='Enter minimum intensity cutoff (recommended start value is 3000):'
	lbl11=tk.Label(text=lbl11t, fg='black', bg='white', width=len(lbl11t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl11.place(x=20, y=yposi+90, height=20, width=430)
	entryico=tk.Entry(zframe, width=40, bg='white')
	entryico.focus_set()
	entryico.place(x=450, y=yposi+90, height=20, width=35) #OK

	lbl10t='Analysis of 37 mix:'
	lbl10=tk.Label(text=lbl10t, fg='black', bg='white', width=len(lbl10t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl10.place(x=20, y=yposi+120, height=20, width=210)	#OK
	# Radiobuttons for FAME 37 mix
	i=0
	k=0
	m=0
	ftm=tk.StringVar()
	valuesftm={'Yes':'1', 'No':'0'}
	for (text, value) in valuesftm.items():
		if i>6:
			k=700
			m=30
		if i==10:
			tk.Radiobutton(zframe, text=text, variable=ftm, value=value, bg=colouroptions[0], command=otherinput).place(x=240+i*70-k, y=yposi+120+m, height=20, width=60) 	# OK bg='white'
		else:
			tk.Radiobutton(zframe, text=text, variable=ftm, value=value, bg=colouroptions[0]).place(x=240+i*70-k, y=yposi+120+m, height=20, width=60) 	# OK bg='white'
		i=i+1

	def run_di():
		# begin run analysis - run batch file in subprocess and pass identifier
		fourlettcode=str(flc.get())
		minlenfa=entrycmin.get()
		maxlenfa=entrycmax.get()
		intcutoff=entryico.get()
		#intcutoff=str(intcutoff)
		ftmix=str(ftm.get())
		minlenfa=str(minlenfa)
		maxlenfa=str(maxlenfa)
		intcutoff=str(intcutoff)
		if fourlettcode=='AMPP':
			cderiv=12
			hderiv=12
			dderiv=0
			nderiv=2
			oderiv=0
			pderiv=0
			ideriv=0
		elif fourlettcode=='NMPA':
			cderiv=7
			hderiv=10
			dderiv=0
			nderiv=2
			oderiv=0
			pderiv=0
			ideriv=0
		elif fourlettcode=='NMPE':
			cderiv=7
			hderiv=9
			dderiv=0
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='MDPE':
			cderiv=7
			hderiv=6
			dderiv=3
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='NEPE':
			cderiv=8
			hderiv=11
			dderiv=0
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='EDPE':
			cderiv=6
			hderiv=6
			dderiv=5
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='NPPE':
			cderiv=9
			hderiv=13
			dderiv=0
			nderiv=1
			oderiv=1
			pderiv=0
			ideriv=0
		elif fourlettcode=='IAMP':
			cderiv=12
			hderiv=11
			dderiv=0
			nderiv=2
			oderiv=0
			pderiv=0
			ideriv=1
		elif fourlettcode=='PLPC':
			cderiv=8
			hderiv=20
			dderiv=0
			nderiv=1
			oderiv=6
			pderiv=1
			ideriv=0
		elif fourlettcode=='PLPE':
			cderiv=5
			hderiv=14
			dderiv=0
			nderiv=1
			oderiv=6
			pderiv=1
			ideriv=0
		else:
			cderiv=str(entrync.get())
			hderiv=str(entrynh.get())
			dderiv=str(entrynd.get())
			nderiv=str(entrynn.get())
			oderiv=str(entryno.get())
			pderiv=str(entrynp.get())
			ideriv=str(entryni.get())

		cderiv=str(cderiv)
		hderiv=str(hderiv)
		dderiv=str(dderiv)
		nderiv=str(nderiv)
		oderiv=str(oderiv)
		pderiv=str(pderiv)
		ideriv=str(ideriv)

		arglist=[]
		arglist.append(fourlettcode)
		arglist.append(intcutoff)
		arglist.append(minlenfa)
		arglist.append(maxlenfa)
		arglist.append(ftmix)
		arglist.append(str(cderiv))
		arglist.append(str(hderiv))
		arglist.append(str(dderiv))
		arglist.append(str(nderiv))
		arglist.append(str(oderiv))
		arglist.append(str(pderiv))
		arglist.append(str(ideriv))

		#print('arglist:')
		#print(arglist)

		#subprocess.call([sys.executable, 'OzFAD1_MSMS_SN.py'])	#works only, if python script in OzFAD1.3

		# OK LOCALLY
		#subprocess.call([sys.executable, 'C:/Users/menzel2/OzFAD1.3/OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_Direct_Infusion.py', fourlettcode, intcutoff, minlenfa, maxlenfa, ftmix, cderiv, hderiv, dderiv, nderiv, oderiv, pderiv, ideriv])	#OK
		
		#subprocess.call([sys.executable, './OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_Direct_Infusion.py', fourlettcode, intcutoff, minlenfa, maxlenfa, ftmix, cderiv, hderiv, dderiv, nderiv, oderiv, pderiv, ideriv])	#OK
		subprocess.call([r'.\OzFAD1_8_Direct_Infusion.bat', fourlettcode, intcutoff, minlenfa, maxlenfa, ftmix, cderiv, hderiv, dderiv, nderiv, oderiv, pderiv, ideriv])
		# end run analysis - run batch file in subprocess and pass identifier

	buttonlabel='Analyse Direct Infusion MS'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_di).place(x=290, y=yposi+170, height=25, width=230)

########################################################################################################################################### STEP 8  # # #  # # #  # # 
########################################################################################################################################### STEP 9  # # #  # # #  # # #  

def frame9():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0)  # bg='white'
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# build variable labels and buttons

	# Begin STEP 3 S/N ##################################################################################################################################
	lbl1t='Step 8: Assemble a Summary and Table.'
	lbl2t='  This algorithm creates a Summary Table from three fatty acid analysis excel files (relative quantification data, three replicates)' 
	lbl3t='  and one excel file with FA species and absolute quantification values.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'), anchor='w')
	lbl3=tk.Label(text=lbl3t, fg='black', bg='white', width=len(lbl3t), height=1, font=('Helvetica 10'), anchor='w')
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)
	lbl3.place(x=20, y=yposd+50, height=20, width=810)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  The file OzFAD1_abs_quant_rep1_2_3.xlsx (abs. quantification data from direct infusion) is in folder OzFAD1.3'
	lbl7t='     --> FA (e.g., 15:1) in A3, A4, ... / rep1 data abs in nmol mL-1 in B3, B4, ... / rep2 C / rep3 D / / FA for Table F3, F4, ....'
	lbl8t='  2.  The files OzFAD1_5_plot_table_rep1.xlsx, ...rep2.xlsx and ...rep3.xlsx (rel. quantification data) are in folder OzFAD1.3'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	lbl8=tk.Label(text=lbl8t, fg='black', bg='white', width=len(lbl8t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=800)
	lbl6.place(x=20, y=yposc+25, height=20, width=800)
	lbl7.place(x=20, y=yposc+50, height=20, width=800)
	lbl8.place(x=20, y=yposc+75, height=20, width=800)

	def run_sum_tab():
		#OK LOCALLY
		#subprocess.call([sys.executable, 'C:/Users/menzel2/OzFAD1.3/OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_7_Summary_Table.py'])	#OK
		#subprocess.call([sys.executable, './OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_7_Summary_Table.py'])	#OK
		subprocess.call([r'.\OzFAD1_9_Summary_Table.bat'])


	buttonlabel='Generate Summary Table'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_sum_tab).place(x=300, y=yposc+150, height=25, width=210)

########################################################################################################################################### STEP 9   # # #  # # #  # # #  
########################################################################################################################################### STEP 10  # # #  # # #  # # #  # 

def frame10():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0)  # bg='white'
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# build variable labels and buttons

	# Begin STEP 3 S/N ##################################################################################################################################
	lbl1t='Step 10: Calculation of P values and data for heatmap.'
	lbl2t='  This algorithm reads six excel files with relative quantification data, and calculates fold-change and P values.'
	lbl3t='  The output data can be displayed as a volcano plot and a heatmap.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'), anchor='w')
	lbl3=tk.Label(text=lbl3t, fg='black', bg='white', width=len(lbl3t), height=1, font=('Helvetica 10'), anchor='w')
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)
	lbl3.place(x=20, y=yposd+50, height=20, width=810)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  The files OzFAD1_5_plot_table_rep1_d1.xlsx to ...rep3_d2.xlsx are in folder OzFAD1.3'
	#lbl7t='  2.  Spectrum, scaled to full view - m/z = 100 to 1200, is in field A5.'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	#lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=800)
	lbl6.place(x=20, y=yposc+25, height=20, width=800)
	#lbl7.place(x=20, y=305, height=20, width=800)	

	def run_volcano():
		#subprocess.call([sys.executable, 'C:/Users/menzel2/OzFAD1.3/OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_P_value_heatmap_data.py'])	#OK
		#subprocess.call([sys.executable, './OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_P_value_heatmap_data.py'])
		subprocess.call([r'.\OzFAD1_10_Volcano.bat'])

	buttonlabel='Generate data for Volcano Plot and Heat Map'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_volcano).place(x=240, y=yposc+100, height=25, width=295)

########################################################################################################################################### STEP 10  # # #  # # #  # # #  #
########################################################################################################################################### STEP 11  # # #  # # #  # # #  # #

def frame11():
	# hide previous variable labels and buttons behind canvas
	canvas2=tk.Canvas(zframe, width=850, height=canh, bg='white', highlightthickness=0)  # bg='white'
	canvas2.place(x=0, y=cany, width=850, height=canh)
	# build variable labels and buttons

	# Begin STEP 3 S/N ##################################################################################################################################
	lbl1t='Step 11: Generation of a Venn diagram inspired barchart.'
	lbl2t='  This algorithm reads an excel file with data on literature findings, and generates a Venn-diagram inspired barchart.'
	lbl1=tk.Label(text=lbl1t, fg='black', bg='white', width=len(lbl1t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl2=tk.Label(text=lbl2t, fg='black', bg='white', width=len(lbl2t), height=1, font=('Helvetica 10'), anchor='w')
	lbl1.place(x=20, y=yposd, height=20, width=810)
	lbl2.place(x=20, y=yposd+25, height=20, width=810)

	lbl5t='Before running this analysis step, ensure:'
	lbl6t='  1.  The file vennbar_chart_maker_input.xlsx is in folder OzFAD1.3'
	#lbl7t='  2.  Spectrum, scaled to full view - m/z = 100 to 1200, is in field A5.'
	lbl5=tk.Label(text=lbl5t, fg='black', bg='white', width=len(lbl5t), height=1, font=('Helvetica 10 bold'), anchor='w')
	lbl6=tk.Label(text=lbl6t, fg='black', bg='white', width=len(lbl6t), height=1, font=('Helvetica 10'), anchor='w')
	#lbl7=tk.Label(text=lbl7t, fg='black', bg='white', width=len(lbl7t), height=1, font=('Helvetica 10'), anchor='w')
	################### EDIT !!!!
	lbl5.place(x=20, y=yposc, height=20, width=800)
	lbl6.place(x=20, y=yposc+25, height=20, width=800)
	#lbl7.place(x=20, y=285, height=20, width=800)

	def run_venn():
		#subprocess.call([sys.executable, 'C:/Users/menzel2/OzFAD1.3/OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_Venn_barchart.py'])	#OK
		subprocess.call([r'.\OzFAD1_11_Venn_Barchart.bat'])
		#subprocess.call([sys.executable, './OzFAD1_black_box/OzFAD1_py/OzFAD_py_tools/OzFAD1_Venn_barchart.py']) #NOT OK

	buttonlabel='Generate Venn Bar Chart'  #'Run' #
	ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=run_venn).place(x=305, y=yposc+100, height=25, width=200)

########################################################################################################################################### STEP 11  # # #  # # #  # # #  # #

# colours 0 and 1 are colours of OzFAD logo
colouroptions=['#FFF2CC','#FD9D3D','blanched almond','snow','bisque','old lace','antique white','linen','papaya whip']

style = ttk.Style()
style.theme_use('alt')
style.configure('TButton', background = colouroptions[0], foreground = 'black', width = 20, borderwidth=1, focusthickness=3, focuscolor='none')
style.map('TButton', background=[('active', colouroptions[1])])

# Begin Buttons for selection of step of workflow
bwidth=100
yub=120

buttonlabel='Targets'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame1).place(x=20, y=yub, height=25, width=bwidth)
buttonlabel='Discovery'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame2).place(x=160, y=yub, height=25, width=bwidth)
buttonlabel='S/N'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame3).place(x=300, y=yub, height=25, width=bwidth)
buttonlabel='DDA to DIA'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame4).place(x=440, y=yub, height=25, width=bwidth)
buttonlabel='Deconvolution'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame5).place(x=580, y=yub, height=25, width=bwidth)
buttonlabel='Plot and Table'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame6).place(x=720, y=yub, height=25, width=bwidth)

ylb=165

# xlb=[20, 195, 370, 545, 720]	# not preferred
xlb=[20, 160, 300, 440, 580]

buttonlabel='Replicate Plot'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame7).place(x=xlb[2], y=ylb, height=25, width=bwidth)
buttonlabel='Direct Infusion'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame8).place(x=xlb[0], y=ylb, height=25, width=bwidth)
buttonlabel='Summary Table'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame9).place(x=xlb[1], y=ylb, height=25, width=bwidth)
buttonlabel='P val. & Heatmap'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame10).place(x=xlb[3], y=ylb, height=25, width=bwidth)
buttonlabel='Venn Bar Chart'  ##
ttk.Button(zframe, text=buttonlabel, width=len(buttonlabel)+2, command=frame11).place(x=xlb[4], y=ylb, height=25, width=bwidth)
# End Buttons for selection of step of workflow

zframe.mainloop()


