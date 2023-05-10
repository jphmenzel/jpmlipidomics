# -*- coding: UTF-8 -*-

#Jan Philipp Menzel 
# Program: Calculate data for p value heatmap from OzFAD rep plot input excel file
#created: 2022 05 12
#Notes: Reads excel files of 2 sets of 3 replicates (6 excel files) of input files for rep plot of OzFAD workflow
#Notes: Calculates p values for each comparison of species (or groups of species in each field), ready to be plotted as a heatmap in origin
import math
import datetime
import pandas as pd
import scipy
from scipy import stats
import openpyxl
from openpyxl import Workbook
import pingouin
from pingouin import ttest
import statistics

gui=1   # default set 1
statprint=0     #set 1 only when troubleshooting

if gui==0:
    print('This code will create an excel file containing data that can be displayed as the P values of comparison of')
    print('  two fatty acid analysis datasets from the OzFAD1 workflow.')
    print('Please ensure that the data is entered correctly into the files:')
    print('Six input files with fatty acid data are required, which need to be named as follows:')
    print('OzFAD1_5_plot_table_rep1_d1.xlsx')
    print('OzFAD1_5_plot_table_rep2_d1.xlsx')
    print('OzFAD1_5_plot_table_rep3_d1.xlsx')
    print('OzFAD1_5_plot_table_rep1_d2.xlsx')
    print('OzFAD1_5_plot_table_rep2_d2.xlsx')
    print('OzFAD1_5_plot_table_rep3_d2.xlsx')
    print('Each three replicates of the two analyses are compared to each other and P values as well as fold change values are calculated.')

#begin read all 6 excel files into lists of lists (read relative isomer abundance data by FA and db array from sheet final_barchart)
tnrep=3 # total number of replicates is 3
xlc=tnrep*2 # 6 excel files in total
superlist=[]    # list of list of lists (contains all 6 slists, which contain lists for rows of raw input data) ## superlist[dataset 0 to 5][row index][column index]

cds=1
while cds<(xlc+1):
    if cds==1:
        wb=openpyxl.load_workbook('OzFAD1_5_plot_table_rep1_d1.xlsx')
        ws=wb['final_barchart']
    elif cds==2:
        wb=openpyxl.load_workbook('OzFAD1_5_plot_table_rep2_d1.xlsx')
        ws=wb['final_barchart']
        #print('Check1')
        #quit()
    elif cds==3:
        wb=openpyxl.load_workbook('OzFAD1_5_plot_table_rep3_d1.xlsx')
        ws=wb['final_barchart']
    elif cds==4:
        wb=openpyxl.load_workbook('OzFAD1_5_plot_table_rep1_d2.xlsx')
        ws=wb['final_barchart']
    elif cds==5:
        wb=openpyxl.load_workbook('OzFAD1_5_plot_table_rep2_d2.xlsx')
        ws=wb['final_barchart']
    elif cds==6:
        wb=openpyxl.load_workbook('OzFAD1_5_plot_table_rep3_d2.xlsx')
        ws=wb['final_barchart']
    rowlist=[]  #list with row entries (contains different db information for one FA group, e.g., 15:1)
    slist=[]    #list of rowlists (whole dataset for one sample, one individual replicate)
    r=1
    c=1
    go=1
    while go==1:
        c=1
        v=ws.cell(row=r, column=c)
        v=v.value
        v=str(v)
        #print(v)
        #vn=ws.cell(row=r+1, column=c)
        #vn=vn.value
        if v is None:
            go=0
        elif v=='None':
            go=0
        else:
            rowlist=[]  #list with row entries (contains different db information for one FA group, e.g., 15:1)
            c=1
            gor=1
            while gor==1:
                vc=ws.cell(row=r, column=c)
                vc=vc.value
                if vc is None:
                    gor=0
                else:
                    rowlist.append(vc)
                c=c+1
            slist.append(rowlist)
            #print(rowlist)
        r=r+1
    superlist.append(slist)
    #print(slist)
    cds=cds+1
if gui==0:
    print('Reading of datasets complete, superlist is generated.')
#print(superlist)
#print(superlist[1][27][6])
#print(len(superlist[0]))
#print(len(superlist[1]))
#print(len(superlist[2]))
#print(len(superlist[3]))
#print(len(superlist[4]))
#print(len(superlist[5]))
#begin check datasets in superlist
ok=0
ok2=0
if len(superlist[0])==len(superlist[1]):
    if len(superlist[2])==len(superlist[1]):
        ok=1
    else:
        print('Check input datasets. Data may be inconsistent (different number of rows in excel files 2 3).')
else:
    print('Check input datasets. Data may be inconsistent (different number of rows in excel files 1 2).')
if len(superlist[3])==len(superlist[4]):
    if len(superlist[3])==len(superlist[5]):
        ok2=1
    else:
        print('Check input datasets. Data may be inconsistent (different number of rows in excel files 4 6).')
else:
    print('Check input datasets. Data may be inconsistent (different number of rows in excel files 4 5).')
if ok==0:
    quit()
elif ok2==0:
    quit()
#end check datasets in superlist
#begin insert empty rows in superlist to complete datasets and align rows of FA
if gui==0:
    print('Datasets must align, otherwise p values will be calculated wrongly!')
#create code to automate insertion of rows...
#end insert empty rows in superlist to complete datasets and align rows of FA
#end read all 6 excel files into lists of lists (read relative isomer abundance data by FA and db array from sheet final_barchart)

#begin calculate p values 
tlist=[]
ctlist=[]
plist=[]
cplist=[]
cpinglist=[]
pinglist=[]

meand1list=[]
stdevd1list=[]
meand2list=[]
stdevd2list=[]

doflist=[]
confintlowlist=[]
confintuplist=[]
ttslist=[]

r=1
c=1
while r<len(superlist[0]):
    c=1
    ctlist=[]
    cplist=[]
    cpinglist=[]
    cmeand1list=[]
    cstdevd1list=[]
    cmeand2list=[]
    cstdevd2list=[]
    while c<len(superlist[0][0]):
        cd1list=[]
        cd2list=[]
        cds=0
        while cds<3:
            cd1=float(superlist[cds][r][c])
            cd1list.append(cd1)
            cds=cds+1
        while cds<6:
            cd2=float(superlist[cds][r][c])
            cd2list.append(cd2)
            cds=cds+1

        if cd1list[0]==cd1list[1]==cd1list[2]==0:
            t='1'
            p='1'
            pingpl=[1]
        elif cd2list[0]==cd2list[1]==cd2list[2]==0:
            t='1'
            p='1'
            pingpl=[1]
        else:
            t,p=stats.ttest_ind(cd1list, cd2list, equal_var=False)
            pingp=ttest(cd1list, cd2list, correction=True)          # using pingouin module to get reporting values for p value calculation
            pingpl=pingp.values.tolist()
            if statprint==1:
                print(pingp)
                print('..............')
                print(pingpl)
                print(pingpl[0][3])
                print(pingpl[0][0])
                print(p)
                print(t)
            t=str(t)
            p=str(p)

        cmeand1=statistics.mean(cd1list)
        cmeand2=statistics.mean(cd2list)
        cstdevd1=statistics.stdev(cd1list)
        cstdevd2=statistics.stdev(cd2list)
            
        if t=='nan':
            t=str(1.0)
        if p=='nan':
            p=str(1.0)
        ctlist.append(t)
        cplist.append(p)
        cpinglist.append(pingpl)
        cmeand1list.append(cmeand1)
        cstdevd1list.append(cstdevd1)
        cmeand2list.append(cmeand2)
        cstdevd2list.append(cstdevd2)
        c=c+1
    tlist.append(ctlist)        # tlist is list of t statistic for each p value calculation
    plist.append(cplist)
    pinglist.append(cpinglist)      # pinglist contains all reporting values for p value calculation statistics in lists indexed as tlist and plist   (pinglist CHECK)
    meand1list.append(cmeand1list)
    stdevd1list.append(cstdevd1list)
    meand2list.append(cmeand2list)
    stdevd2list.append(cstdevd2list)
    r=r+1
#print(plist)
#print(plist[10][15])





# use https://ethanweed.github.io/pythonbook/05.02-ttest.html
# pingouin module to generate all info from welshs ttest






    #if str(superlist[0][r][0])==str(superlist[1][r][0]):
    #    if str(superlist[0][r][0])==str(superlist[2][r][0]):
    #        if str(superlist[0][r][0])==str(superlist[3][r][0]):

# begin calculate mean fold changes between datasets
foldlist=[]
cfoldlist=[]
r=1
c=1
while r<len(superlist[0]):
    c=1
    cfoldlist=[]
    while c<len(superlist[0][0]):
        cablist=[]
        cds=0
        while cds<6:
            cab=float(superlist[cds][r][c])
            cablist.append(cab)
            cds=cds+1
        if 0 in cablist:
            cfold=0
        else:
            cfold=((cablist[3]+cablist[4]+cablist[5])/3)/((cablist[0]+cablist[1]+cablist[2])/3) # calculate fold change of mean values of samples of the two groups
        cfoldlist.append(cfold)
        c=c+1
    foldlist.append(cfoldlist)
    r=r+1
#print(foldlist)
#print(foldlist[10][15]) # [10][15] in MCF7 LNCaP LNCaP_SCD-1i is 14:1n-5cis
#print(foldlist[11][25]) # [11][25] in MCF7 LNCaP LNCaP_SCD-1i is 16:1n-7cis


# end calculate mean fold changes between datasets

#end calculate p values
writeoutput=1
if writeoutput==1:
    #begin write p values in output excel file 
    wb = Workbook(write_only=True)
    wb.save('jpmlipidomics_p_values_for_heatmap.xlsx')
    wb=openpyxl.load_workbook('jpmlipidomics_p_values_for_heatmap.xlsx')
    ws=wb.active
    ws.title='Data for Heatmaps'
    ws=wb['Data for Heatmaps']
    ndblist=[]

    c=0
    while c<len(superlist[0][0]):   #write top row with db assignment
        if 'NMI, E' in str(superlist[0][0][c]):     # correct entries for final plot
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx+1:
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'branched)'
                dx=dx+1
        elif 'NMI, Z' in str(superlist[0][0][c]):
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx+1:
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'NMI)'
                dx=dx+1
        elif 'Bu' in str(superlist[0][0][c]):
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx+1:
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'NMI (Bu))'
                dx=dx+1
        elif 'Me, E' in str(superlist[0][0][c]):
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx-1: #previously +1
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'t' #'trans)' #'trans (E))'
                dx=dx+1
        elif 'Me, Z' in str(superlist[0][0][c]):
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx-1: #previously +1
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'c'  #'cis)' #'cis (Z))'
                dx=dx+1
        else:
            ndblabel=str(superlist[0][0][c])
        ndblist.append(ndblabel)
        ws.cell(row=1, column=c+1).value=ndblabel  # write row label for p value heatmap   
        ws.cell(row=1+4+len(superlist[0]), column=c+1).value=ndblabel  # write row label for fold change heatmap    
        c=c+1
    r=0
    while r<len(superlist[0]):      
        ws.cell(row=r+1, column=1).value=superlist[0][r][0]     #write column assignment for p value heatmap (FA)
        ws.cell(row=r+5+len(superlist[0]), column=1).value=superlist[0][r][0]     #write column assignment for fold-change heatmap (FA)
        r=r+1
    r=0
    while r<len(plist):
        c=0
        while c<len(plist[r]):
            ws.cell(row=r+2, column=c+2).value=float(plist[r][c])   #write p values
            ws.cell(row=r+6+len(superlist[0]), column=c+2).value=float(foldlist[r][c])   #write fold change values
            c=c+1
        r=r+1
    ws.cell(row=2+len(superlist[0]), column=2).value='Above: P values for heatmap'   #write Note
    ws.cell(row=3+len(superlist[0]), column=2).value='Below: Fold change values for heatmap'   #write Note

    wb.save('jpmlipidomics_p_values_for_heatmap.xlsx')
    #print('P values are saved in excel file jpmlipidomics_p_values_for_heatmap.xlsx')
    #end write p values and fold change values in output excel file 
#print('Done.')

#print(ndblist)
#print('##########')
#print(foldlist[0])
#print(superlist[0][1][0])
#quit()

# begin assemble data for volcano plot
# use plist and foldlist and superlist indices to build list for volcano plot data
vfalist=[]  #list of fatty acid isomer labels for volcano plot
vfclist=[]  #list of fold change values for volcano plot
vpvlist=[]  #list of P values for volcano plot
vlogfclist=[]  #list of LOG fold change values for volcano plot
vlogpvlist=[]  #list of LOG P values for volcano plot
wmeand1list=[]
wmeand2list=[]
wstdevd1list=[]
wstdevd2list=[]


si=0    # rows in p value table, e.g. 12:1 species, 14:1 species ...
sii=0   # columns in p value table, e.g. n-7(Z) species, n-7(E) species ...
while si<(len(foldlist)):
    sii=0
    while sii<(len(foldlist[0])):
        if float(foldlist[si][sii])>0:
            if float(plist[si][sii])<1:
                #found a data point for volcano plot
                cvfa=str(superlist[0][si+1][0])   #+str(superlist[0][0][sii])
                # read top row of p value data
                ndblbl=ndblist[sii+1]
                cvfa=cvfa+str(ndblbl)
                vfalist.append(cvfa)
                cvfc=float(foldlist[si][sii])
                vfclist.append(cvfc)
                cvpv=float(plist[si][sii])
                vpvlist.append(cvpv)

                # get data of p value calculation and write into output lists           # FIX LIST INDEXING  !!!!!!
                if pinglist[si][sii][0]==1:
                    ok=1
                    doflist.append(1)
                    confintlowlist.append(1)
                    confintuplist.append(1)
                    ttslist.append(1)
                    wmeand1list.append(1)
                    wmeand2list.append(1)
                    wstdevd1list.append(1)
                    wstdevd2list.append(1)

                else:
                    if statprint==1:
                        print('plist:               $$$$$$$$$$$$$$$$$$$$$$$$$$$$')
                        print(plist)
                        print(' $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$')
                        print('pinglist:               ########################')
                        print(pinglist)
                        print(' ###############################################')
                        print(pinglist[si][sii][0])
                        print('ttest statistic:')
                        print(pinglist[si][sii][0][0])  # ttest statistic 
                        print('dof:')
                        print(pinglist[si][sii][0][1])  # degrees of freedom
                        print('CI_95%:')
                        print(pinglist[si][sii][0][4][0])  # confidence interval 95%, lower limit
                        print(pinglist[si][sii][0][4][1])  # confidence interval 95%, upper limit

                    doflist.append(pinglist[si][sii][0][1])                 # CHECK INDEXING TO ACCESS CORRECT VALUES
                    confintlowlist.append(pinglist[si][sii][0][4][0])
                    confintuplist.append(pinglist[si][sii][0][4][1])
                    ttslist.append(pinglist[si][sii][0][0])
                    wmeand1list.append(meand1list[si][sii])
                    wmeand2list.append(meand2list[si][sii])
                    wstdevd1list.append(stdevd1list[si][sii])
                    wstdevd2list.append(stdevd2list[si][sii])

                if statprint==1:
                    print('doflist:')
                    print(doflist)
                    print('confintlowlist:')
                    print(confintlowlist)
                    print('confintuplist:')
                    print(confintuplist)
                    print('ttslist:')
                    print(ttslist)
                
                #if sii==1:
                #    quit()
                #elif si==1:
                #    quit()

                cvlogfc=math.log2(cvfc)
                cvlogpv=math.log10(cvpv)
                cvlogpv=-1*cvlogpv
                vlogfclist.append(cvlogfc)
                vlogpvlist.append(cvlogpv)
        sii=sii+1
    si=si+1

wb=openpyxl.load_workbook('jpmlipidomics_p_values_for_heatmap.xlsx')
wb.create_sheet(title='Data for Volcano Plot')
wsvp=wb['Data for Volcano Plot']

colshift1=4
colshift2=8

wsvp.cell(row=1, column=1).value='FA isomer'    #write top row

wsvp.cell(row=1, column=2).value='Mean (d1)'    #write top row
wsvp.cell(row=1, column=3).value='Standard deviation (d1)'    #write top row
wsvp.cell(row=1, column=4).value='Mean (d2)'    #write top row
wsvp.cell(row=1, column=5).value='Standard deviation (d2)'    #write top row

wsvp.cell(row=1, column=2+colshift1).value='Fold change'    #write top row
wsvp.cell(row=1, column=3+colshift1).value='P value'    #write top row

wsvp.cell(row=1, column=8).value='t test statistic'    #write top row
wsvp.cell(row=1, column=9).value='degrees of freedom'    #write top row
wsvp.cell(row=1, column=10).value='confidence interval 95%, lower limit'    #write top row
wsvp.cell(row=1, column=11).value='confidence interval 95%, upper limit'    #write top row

wsvp.cell(row=1, column=5+colshift2).value='FA isomer'    #write top row
wsvp.cell(row=1, column=6+colshift2).value='log2 fold change'    #write top row
wsvp.cell(row=1, column=7+colshift2).value='="-log10 p-value"'    #write top row
wsvp.cell(row=1, column=8+colshift2).value='="-log10 p-value"'    #write top row

wsvp.cell(row=1, column=13+colshift2).value='="line p=0.05"'    #write top row
wsvp.cell(row=1, column=13+colshift2).value='="line fold ch. *0.5'    #write top row
wsvp.cell(row=1, column=13+colshift2).value='="line fold ch. *2"'    #write top row
wsvp.cell(row=1, column=13+colshift2).value='="line fold ch. *0"'    #write top row

wsvp.cell(row=1, column=13+colshift2).value='="n-2"'    #write top row
wsvp.cell(row=1, column=14+colshift2).value='="n-3"'    #write top row
wsvp.cell(row=1, column=15+colshift2).value='="n-4"'    #write top row
wsvp.cell(row=1, column=16+colshift2).value='="n-5"'    #write top row
wsvp.cell(row=1, column=17+colshift2).value='="n-6"'    #write top row
wsvp.cell(row=1, column=18+colshift2).value='="n-7"'    #write top row
wsvp.cell(row=1, column=19+colshift2).value='="n-8"'    #write top row
wsvp.cell(row=1, column=20+colshift2).value='="n-9"'    #write top row
wsvp.cell(row=1, column=21+colshift2).value='="n-10"'    #write top row
wsvp.cell(row=1, column=22+colshift2).value='="n-11"'    #write top row
wsvp.cell(row=1, column=23+colshift2).value='="n-12"'    #write top row
wsvp.cell(row=1, column=24+colshift2).value='="n-13"'    #write top row
wsvp.cell(row=1, column=25+colshift2).value='="n-14"'    #write top row
wsvp.cell(row=1, column=26+colshift2).value='="n-15"'    #write top row
wsvp.cell(row=1, column=27+colshift2).value='="n-16"'    #write top row

vol=2
while vol<(len(vfalist)+2):
    wsvp.cell(row=vol, column=1).value=vfalist[vol-2]    #write fatty acid labels for volcano plot data

    wsvp.cell(row=vol, column=2).value=wmeand1list[vol-2]    #write mean d1 values for table
    wsvp.cell(row=vol, column=3).value=wstdevd1list[vol-2]    #write stdev d1 values for table
    wsvp.cell(row=vol, column=4).value=wmeand2list[vol-2]    #write mean d2 values for table
    wsvp.cell(row=vol, column=5).value=wstdevd2list[vol-2]    #write stdev d2 values for table

    wsvp.cell(row=vol, column=2+colshift1).value=vfclist[vol-2]    #write fold change values for table
    wsvp.cell(row=vol, column=3+colshift1).value=vpvlist[vol-2]    #write p values for table 

    wsvp.cell(row=vol, column=4+colshift1).value=ttslist[vol-2]    #write ttest statistic values for table
    wsvp.cell(row=vol, column=5+colshift1).value=doflist[vol-2]    #write dof values for table
    wsvp.cell(row=vol, column=6+colshift1).value=confintlowlist[vol-2]    #write ci values lower limit for table
    wsvp.cell(row=vol, column=7+colshift1).value=confintuplist[vol-2]    #write ci values upper limit for table
    

    wsvp.cell(row=vol, column=5+colshift2).value=vfalist[vol-2]    #write fatty acid labels for volcano plot data
    wsvp.cell(row=vol, column=6+colshift2).value=vlogfclist[vol-2]    #write fatty acid labels for volcano plot data
    if abs(float(vlogfclist[vol-2]))<1:
        wsvp.cell(row=vol, column=8+colshift2).value=vlogpvlist[vol-2]    #write significant, little changing FA datapoints for volcano plot data
    elif float(vlogpvlist[vol-2])<1.30103:
        wsvp.cell(row=vol, column=8+colshift2).value=vlogpvlist[vol-2]    #write non significant, highly changing FA datapoints for volcano plot data
    else:
        wsvp.cell(row=vol, column=7+colshift2).value=vlogpvlist[vol-2]    #write significant and highly changing FA datapoints for volcano plot data
        if len(vfalist[vol-2])==8:
            cndb=int(vfalist[vol-2][6])
        elif len(vfalist[vol-2])==9:
            cndb=int(vfalist[vol-2][6]+vfalist[vol-2][7])
        elif len(vfalist[vol-2])>9:
            if (vfalist[vol-2][7])==' ':
                cndb=int(vfalist[vol-2][6])
            elif (vfalist[vol-2][8])==' ':
                cndb=int(vfalist[vol-2][6]+vfalist[vol-2][7])
            else:
                cndb=int(vfalist[vol-2][6])
        wsvp.cell(row=vol, column=7+4+cndb+colshift2).value=vlogpvlist[vol-2]    #write significant and highly changing FA datapoints for volcano plot data (bubbles)
    vol=vol+1

wsvp.cell(row=(len(vfalist)+2), column=6+colshift2).value=-10    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+2), column=9+colshift2).value=1.30103    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+3), column=6+colshift2).value=10    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+3), column=9+colshift2).value=1.30103    #write values for border lines to denote statistical significance and high changes (> twofold)

wsvp.cell(row=(len(vfalist)+4), column=6+colshift2).value=-1.0000000001    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+4), column=10+colshift2).value=-1    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+5), column=6+colshift2).value=-1    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+5), column=10+colshift2).value=10    #write values for border lines to denote statistical significance and high changes (> twofold)

wsvp.cell(row=(len(vfalist)+6), column=6+colshift2).value=1   #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+6), column=11+colshift2).value=-1    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+7), column=6+colshift2).value=1.0000000001    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+7), column=11+colshift2).value=10    #write values for border lines to denote statistical significance and high changes (> twofold)

wsvp.cell(row=(len(vfalist)+8), column=6+colshift2).value=0    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+8), column=12+colshift2).value=-1    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+9), column=6+colshift2).value=0.0000000001    #write values for border lines to denote statistical significance and high changes (> twofold)
wsvp.cell(row=(len(vfalist)+9), column=12+colshift2).value=10    #write values for border lines to denote statistical significance and high changes (> twofold)

wb.save('jpmlipidomics_p_values_for_heatmap.xlsx')





if gui==0:
    print('P values, fold change values and data for volcano plot are saved in excel file jpmlipidomics_p_values_for_heatmap.xlsx')
# end assemble data for volcano plot

