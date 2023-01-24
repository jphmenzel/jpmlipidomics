# -*- coding: UTF-8 -*-

#Jan Philipp Menzel 
#created: 2022
#Notes: Generate data of fatty acid isomer relative and absolute abundances incl. standard deviation and COV as workflow output (replicate data of 3 replicates)
import math
import datetime
import pandas as pd
import matplotlib.cm as cm
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook
import statistics
import sigfig
from sigfig import round
# ask for number of replicates
gui=1

if gui==0:
    print('This program is part of the OzFAD1 workflow and creates a Summary Table from three fatty acid analysis excel files (relative quantification data, three replicates) and one excel file with FA species and absolute quantification values.')
    print('The data needs to be in files with the following filenames:')
    print('OzFAD1_abs_quant_rep1_2_3.xlsx')
    print('FA (e.g., 15:1) in A3, A4, ... / rep1 data abs in nmol mL-1 in B3, B4, ... / rep2 C / rep3 D / / FA for Table F3, F4, ...')
    print('OzFAD1_5_plot_table_rep1.xlsx')
    print('OzFAD1_5_plot_table_rep2.xlsx')
    print('OzFAD1_5_plot_table_rep3.xlsx')
# begin read excel files into dataframes and calculate average values

wb1=openpyxl.load_workbook('OzFAD1_5_plot_table_rep1.xlsx', data_only=True)
sheetr1=wb1['final_barchart']
wb2=openpyxl.load_workbook('OzFAD1_5_plot_table_rep2.xlsx', data_only=True)
sheetr2=wb2['final_barchart']
wb3=openpyxl.load_workbook('OzFAD1_5_plot_table_rep3.xlsx', data_only=True)
sheetr3=wb3['final_barchart']
wb4=openpyxl.load_workbook('OzFAD1_abs_quant_rep1_2_3.xlsx', data_only=True)
sheetabs=wb4['Sheet1']
# begin determine dfcolumns
if gui==0:
    print('Excel files with replicates are read. Calculating ...')
dfcolumns=[] # list with label for n position
cc=2
go=1
while go==1:
    e=sheetr1.cell(row=1, column=cc)
    e=e.value
    if e is None:
        go=0
        e=str(e)
    elif str(e)=='':
        go=0
    else:
        dfcolumns.append(e)
    cc=cc+1
# end determine dfcolumns
#print(len(dfcolumns))
#print(dfcolumns)

# begin assemble lists from excel files
abslistx=[]
abslistr1=[]
abslistr2=[]
abslistr3=[]
fatlist=[]
rep1list=[]
rep2list=[]
rep3list=[]

r=3
go=1
while go==1:
    e=sheetabs.cell(row=r, column=1)
    e=e.value
    f=sheetabs.cell(row=r, column=2)
    f=f.value
    g=sheetabs.cell(row=r, column=3)
    g=g.value
    h=sheetabs.cell(row=r, column=4)
    h=h.value
    if e is None:
        go=0
        e=str(e)
    elif str(e)=='':
        go=0
    else:
        e=str(e)
        if len(e)==9:
            e=e[5:]
            #print(e)
        f=str(f)
        g=str(g)
        h=str(h)
        abslistx.append(e)
        abslistr1.append(f)
        abslistr2.append(g)
        abslistr3.append(h)
    r=r+1

#print(abslistx)
#print(abslistr1)
#print(abslistr2)
#print(abslistr3)

r=3
go=1
while go==1:
    e=sheetabs.cell(row=r, column=6)
    e=e.value
    if e is None:
        go=0
        e=str(e)
    elif str(e)=='':
        go=0
    else:
        e=str(e)
        fatlist.append(e)
    r=r+1

#print(fatlist)


r=1
c=1
gor=1
goc=1
while gor==1:
    e=sheetr1.cell(row=r, column=c)
    e=e.value
    if e is None:
        gor=0
        e=str(e)
    elif str(e)=='':
        gor=0
    else:
        e=str(e)
        cfalist=[]
        c=1
        goc=1
        while goc==1:
            e=sheetr1.cell(row=r, column=c)
            e=e.value
            if e is None:
                goc=0
                e=str(e)
            elif str(e)=='':
                goc=0
            else:
                cfalist.append(e)
            c=c+1
        rep1list.append(cfalist)
        c=1
    r=r+1
#print(rep1list)

r=1
c=1
gor=1
goc=1
while gor==1:
    e=sheetr2.cell(row=r, column=c)
    e=e.value
    if e is None:
        gor=0
        e=str(e)
    elif str(e)=='':
        gor=0
    else:
        e=str(e)
        cfalist=[]
        c=1
        goc=1
        while goc==1:
            e=sheetr2.cell(row=r, column=c)
            e=e.value
            if e is None:
                goc=0
                e=str(e)
            elif str(e)=='':
                goc=0
            else:
                cfalist.append(e)
            c=c+1
        rep2list.append(cfalist)
        c=1
    r=r+1
#print(rep2list)

r=1
c=1
gor=1
goc=1
while gor==1:
    e=sheetr3.cell(row=r, column=c)
    e=e.value
    if e is None:
        gor=0
        e=str(e)
    elif str(e)=='':
        gor=0
    else:
        e=str(e)
        cfalist=[]
        c=1
        goc=1
        while goc==1:
            e=sheetr3.cell(row=r, column=c)
            e=e.value
            if e is None:
                goc=0
                e=str(e)
            elif str(e)=='':
                goc=0
            else:
                cfalist.append(e)
            c=c+1
        rep3list.append(cfalist)
        c=1
    r=r+1
#print(rep3list)

#####################################  All lists from input are built
meanrel=[]
stdevrel=[]
covrel=[]
meanabs=[]
stdevabs=[]
covabs=[]
relr1=[]
relr2=[]
relr3=[]
absr1=[]
absr2=[]
absr3=[]
omeanrel=[]
ostdevrel=[]
omeanabs=[]
ostdevabs=[]

ma=0
while ma<len(fatlist):
    meanrel.append('_')
    stdevrel.append('_')
    covrel.append('_')
    meanabs.append('_')
    stdevabs.append('_')
    covabs.append('_')
    relr1.append('_')
    relr2.append('_')
    relr3.append('_')
    absr1.append('_')
    absr2.append('_')
    absr3.append('_')
    omeanrel.append('_')
    ostdevrel.append('_')
    omeanabs.append('_')
    ostdevabs.append('_')
    ma=ma+1

fati=0
while fati<len(fatlist):
    if fatlist[fati][4]=='0':
        #SatFA
        ax=1
        while ax<len(abslistx):
            if str(abslistx[ax]) in str(fatlist[fati]):
                cmeanabs=(float(abslistr1[ax])+float(abslistr2[ax])+float(abslistr3[ax]))/3
                if cmeanabs<1:
                    meanabs[fati]=round(cmeanabs, 2)
                else:
                    meanabs[fati]=round(cmeanabs, 3)
                cstdevabs=statistics.stdev([float(abslistr1[ax]),float(abslistr2[ax]),float(abslistr3[ax])])
                stdevabs[fati]=round(cstdevabs, 1)
                if cmeanabs==0:
                    ok=1
                else:
                    ccovabs=cstdevabs/cmeanabs
                    covabs[fati]=round(ccovabs, 2)
                absr1[fati]=round(float(abslistr1[ax]), 3)
                absr2[fati]=round(float(abslistr2[ax]), 3)
                absr3[fati]=round(float(abslistr3[ax]), 3)
            ax=ax+1
    elif fatlist[fati][4]=='1':
        #MUFA
        #print(fatlist[fati])
        #print('checkpoint1')
        cfa=str(fatlist[fati][1])+str(fatlist[fati][2])+str(fatlist[fati][3])+str(fatlist[fati][4])     # e.g. 12:1
        if len(fatlist[fati])==8:
            nfa=str(fatlist[fati][7])
            ezfa=1
        elif str(fatlist[fati][len(fatlist[fati])-2])=='E':
            if str(fatlist[fati][8])=='_':
                nfa=str(fatlist[fati][7])
            elif str(fatlist[fati][9])=='_':
                nfa=str(fatlist[fati][7])+str(fatlist[fati][8])
            ezfa=2
        elif len(fatlist[fati])==9:
            nfa=str(fatlist[fati][7])+str(fatlist[fati][8])
            ezfa=1
        elif str(fatlist[fati][len(fatlist[fati])-2])=='b':
            if str(fatlist[fati][8])=='_':
                nfa=str(fatlist[fati][7])
            elif str(fatlist[fati][9])=='_':
                nfa=str(fatlist[fati][7])+str(fatlist[fati][8])
            ezfa=3
        elif str(fatlist[fati][len(fatlist[fati])-2])=='B':
            if str(fatlist[fati][8])=='_':
                nfa=str(fatlist[fati][7])
            elif str(fatlist[fati][9])=='_':
                nfa=str(fatlist[fati][7])+str(fatlist[fati][8])
            ezfa=3
        else:
            nfa=str(fatlist[fati][7])+str(fatlist[fati][8])
            ezfa=2
        nfa=int(nfa)
        dbix=((nfa-1)*5)-3
        if ezfa==2:
            dbix=dbix+1
        elif ezfa==3:
            dbix=dbix+4
        dbix=dbix-1
        # dbix is index in array with rel. abundance values / index in repxlist[FA_index][dbix]
        ax=0
        while ax<len(abslistx):
            if cfa==str(abslistx[ax]):
                cabsr1=float(abslistr1[ax])
                cabsr2=float(abslistr2[ax])
                cabsr3=float(abslistr3[ax])
            ax=ax+1
        fai=1
        while fai<len(rep1list):
            if cfa==str(rep1list[fai][0]):
                crelr1=float(rep1list[fai][dbix])
                crelr2=float(rep2list[fai][dbix])
                crelr3=float(rep3list[fai][dbix])
            fai=fai+1
        # all input for this MUFA assembled
        if ((crelr1+crelr2+crelr3)/3)<1:
            meanrel[fati]=round((crelr1+crelr2+crelr3)/3, 2)
        else:
            meanrel[fati]=round((crelr1+crelr2+crelr3)/3, 3)
        stdevrel[fati]=round(statistics.stdev([crelr1,crelr2,crelr3]), 1)
        if meanrel[fati]==0:
            ok=1
        else:
            covrel[fati]=round(stdevrel[fati]/meanrel[fati], 2)
        if (((cabsr1*0.01*crelr1)+(cabsr2*0.01*crelr2)+(cabsr3*0.01*crelr3))/3)<1:
            meanabs[fati]=round(((cabsr1*0.01*crelr1)+(cabsr2*0.01*crelr2)+(cabsr3*0.01*crelr3))/3, 2)
        else:
            meanabs[fati]=round(((cabsr1*0.01*crelr1)+(cabsr2*0.01*crelr2)+(cabsr3*0.01*crelr3))/3, 3)
        stdevabs[fati]=round(statistics.stdev([(cabsr1*0.01*crelr1),(cabsr2*0.01*crelr2),(cabsr3*0.01*crelr3)]), 1)
        if meanabs[fati]==0:
            ok=1
        else:
            covabs[fati]=round(stdevabs[fati]/meanabs[fati], 2)
        relr1[fati]=round(crelr1, 3)
        relr2[fati]=round(crelr2, 3)
        relr3[fati]=round(crelr3, 3)
        absr1[fati]=round((cabsr1*0.01*crelr1), 3)
        absr2[fati]=round((cabsr2*0.01*crelr2), 3)
        absr3[fati]=round((cabsr3*0.01*crelr3), 3)
        #print('checkpoint2')
    else:
        #PUFA
        skip=0
        cfa=str(fatlist[fati][1])+str(fatlist[fati][2])+str(fatlist[fati][3])+str(fatlist[fati][4])     # e.g. 14:2 (22:6)
        fai=1
        while fai<len(rep1list):
            if cfa==str(rep1list[fai][0]):
                if 100 in rep1list[fai]:
                    if 100 in rep2list[fai]:
                        if 100 in rep3list[fai]:
                            #Only 1 isomer within isomer family, use abs values as is
                            skip=1
                            ax=1
                            while ax<len(abslistx):
                                if str(abslistx[ax]) in str(fatlist[fati]):
                                    cmeanabs=(float(abslistr1[ax])+float(abslistr2[ax])+float(abslistr3[ax]))/3
                                    if (cmeanabs)<1:
                                        meanabs[fati]=round(cmeanabs, 2)
                                    else:
                                        meanabs[fati]=round(cmeanabs, 3)
                                    cstdevabs=statistics.stdev([float(abslistr1[ax]),float(abslistr2[ax]),float(abslistr3[ax])])
                                    stdevabs[fati]=round(cstdevabs, 1)
                                    relr1[fati]='100'
                                    relr2[fati]='100'
                                    relr3[fati]='100'
                                    absr1[fati]=round(abslistr1[ax], 3)
                                    absr2[fati]=round(abslistr2[ax], 3)
                                    absr3[fati]=round(abslistr3[ax], 3)
                                ax=ax+1
            fai=fai+1
        
        if skip==0: # PUFA with multiple isomers in isomer group
            #print(fatlist[fati])
            if str(fatlist[fati][8])==',':
                nfa=str(fatlist[fati][7])
            elif str(fatlist[fati][9])==',':
                nfa=str(fatlist[fati][7])+str(fatlist[fati][8])
            if str(fatlist[fati][len(fatlist[fati])-2])=='E':
                ezfa=2
            else:
                ezfa=1
            cfanlist=[]  #list with n positions of current FA
            gof=1
            add=0
            fi=7
            #print(fatlist[fati])
            while gof==1:
                if fi==(len(fatlist[fati])-3):
                    cn=str(fatlist[fati][fi])
                    cnn=str(fatlist[fati][fi+2])
                    add=2
                    gof=0
                elif fi==(len(fatlist[fati])-2):
                    cn=str(fatlist[fati][fi])+str(fatlist[fati][fi+1])
                    add=1
                    gof=0
                else:
                    if str(fatlist[fati][fi+1])==',':
                        cn=str(fatlist[fati][fi])
                        fi=fi+2
                        add=1
                    elif str(fatlist[fati][fi+1])=='_':
                        cn=str(fatlist[fati][fi])
                        add=1
                        gof=0
                    elif str(fatlist[fati][fi+2])=='_':
                        cn=str(fatlist[fati][fi])+str(fatlist[fati][fi+1])
                        add=1
                        fi=fi+3
                        gof=0
                    elif str(fatlist[fati][fi+2])==',':
                        cn=str(fatlist[fati][fi])+str(fatlist[fati][fi+1])
                        add=1
                        fi=fi+3
                if add==1:
                    cn=int(cn)
                    cfanlist.append(cn)
                elif add==2:
                    cn=int(cn)
                    cnn=int(cnn)
                    cfanlist.append(cn)
                    cfanlist.append(cnn)
            #print(fatlist[fati])
            #print(cfanlist)
            # determined doublebond positions, now determine if Bu interrupted or NMI or methylene interrupted
            nfa=cfanlist[0]
            dbix=((nfa-1)*5)-3
            dn=0
            if len(cfanlist)==2:
                if cfanlist[1]-cfanlist[0]==6:
                    # Bu
                    dbix=dbix+2
                    dn=1
            if dn==0:
                cni=0
                mei=1
                while cni<(len(cfanlist)-1):
                    if cfanlist[cni+1]-cfanlist[cni]==3:
                        mei=mei
                    else:
                        mei=0
                    cni=cni+1
                if mei==1:
                    #Me
                    if str(fatlist[fati][len(fatlist[fati])-2])=='E':
                        #E
                        dbix=dbix+1
                else:
                    dbix=dbix+3 #NMI
            dbix=dbix-1
            #print(dbix)
            # next gather data and run statistics
            ax=1
            while ax<len(abslistx):
                if cfa==str(abslistx[ax]):
                    cabsr1=float(abslistr1[ax])
                    cabsr2=float(abslistr2[ax])
                    cabsr3=float(abslistr3[ax])
                ax=ax+1
            fai=1
            while fai<len(rep1list):
                if cfa==str(rep1list[fai][0]):
                    crelr1=float(rep1list[fai][dbix])
                    crelr2=float(rep2list[fai][dbix])
                    crelr3=float(rep3list[fai][dbix])
                fai=fai+1
            # all input for this PUFA assembled
            if ((crelr1+crelr2+crelr3)/3)<1:
                meanrel[fati]=round((crelr1+crelr2+crelr3)/3, 2)
            else:
                meanrel[fati]=round((crelr1+crelr2+crelr3)/3, 3)
            stdevrel[fati]=round(statistics.stdev([crelr1,crelr2,crelr3]), 1)
            if meanrel[fati]==0:
                ok=1
            else:
                covrel[fati]=round(stdevrel[fati]/meanrel[fati], 2)
            if (((cabsr1*0.01*crelr1)+(cabsr2*0.01*crelr2)+(cabsr3*0.01*crelr3))/3)<1:
                meanabs[fati]=round(((cabsr1*0.01*crelr1)+(cabsr2*0.01*crelr2)+(cabsr3*0.01*crelr3))/3, 2)
            else:
                meanabs[fati]=round(((cabsr1*0.01*crelr1)+(cabsr2*0.01*crelr2)+(cabsr3*0.01*crelr3))/3, 3)
            stdevabs[fati]=round(statistics.stdev([(cabsr1*0.01*crelr1),(cabsr2*0.01*crelr2),(cabsr3*0.01*crelr3)]), 1)
            if meanabs[fati]==0:
                ok=1
            else:
                covabs[fati]=round(stdevabs[fati]/meanabs[fati], 2)
            relr1[fati]=round(crelr1, 3)
            relr2[fati]=round(crelr2, 3)
            relr3[fati]=round(crelr3, 3)
            absr1[fati]=round((cabsr1*0.01*crelr1), 3)
            absr2[fati]=round((cabsr2*0.01*crelr2), 3)
            absr3[fati]=round((cabsr3*0.01*crelr3), 3)

        ok=1
    fati=fati+1

#print(meanabs)
#print(stdevabs)



wbout = Workbook()  #write_only=True)
wsout = wbout.active

toprow=['Fatty acid species', 'mean relative isomer quantity / %', 'Standard deviation rel. isomer quantity', 'COV rel. isomer quantity', 
'mean absolute isomer quantity', 'Standard deviation abs. isomer quantity', 'COV abs. isomer quantity', 
'rel. isomer quantity, replicate 1', 'rel. isomer quantity, repl. 2', 'rel. isomer quantity, repl. 3', 
'abs. fatty acid quantity, replicate 1', 'abs. fatty acid quantity, repl. 2', 'abs. fatty acid quantity, repl. 3', '', 'Rel. isomer quantity', 'Abs. isomer quantity']

tp=0
while tp<len(toprow):
    wsout.cell(row=1, column=tp+1).value=toprow[tp]
    tp=tp+1

r=0
while r<(len(fatlist)):
    omeanrel[r]=meanrel[r]
    ostdevrel[r]=stdevrel[r]
    omeanabs[r]=meanabs[r]
    ostdevabs[r]=stdevabs[r]
    # begin round mean to adjusted number of significant digits depending on standard deviation
    extraround=1
    if extraround==1:
        if str(stdevrel[r])=='_':
            gorn=0
        else:
            gorn=1
        if str(meanrel[r])=='_':
            gorn=0
        else:
            gorn=gorn
        if gorn==1:
            if 0.000999<float(stdevrel[r]):
                if float(stdevrel[r])<0.01:
                    if 0.001<float(meanrel[r]):
                        if float(meanrel[r])<0.01:
                            meanrel[r]=str(round(float(meanrel[r]), 1))
            if 0.00999<float(stdevrel[r]):
                if float(stdevrel[r])<0.1:
                    if 0.01<float(meanrel[r]):
                        if float(meanrel[r])<0.1:
                            meanrel[r]=str(round(float(meanrel[r]), 1))
            if 0.0999<float(stdevrel[r]):
                if float(stdevrel[r])<1:
                    if 0.1<float(meanrel[r]):
                        if float(meanrel[r])<1:
                            meanrel[r]=str(round(float(meanrel[r]), 1))
            if 0.0999<float(stdevrel[r]):
                if float(stdevrel[r])<1:
                    if 1<float(meanrel[r]):
                        if float(meanrel[r])<10:
                            meanrel[r]=str(round(float(meanrel[r]), 2))
            if 0.999<float(stdevrel[r]):
                if float(stdevrel[r])<10:
                    if 10<float(meanrel[r]):
                        if float(meanrel[r])<100:
                            meanrel[r]=str(round(float(meanrel[r]), 2))
            if 0.999<float(stdevrel[r]):
                if float(stdevrel[r])<10:
                    if 1<float(meanrel[r]):
                        if float(meanrel[r])<10:
                            meanrel[r]=str(round(float(meanrel[r]), 1))
            if 9.99<float(stdevrel[r]):
                if float(stdevrel[r])<100:
                    if 100<float(meanrel[r]):
                        if float(meanrel[r])<1000:
                            meanrel[r]=str(round(float(meanrel[r]), 2))
            if 99.9<float(stdevrel[r]):
                if float(stdevrel[r])<1000:
                    if 1000<float(meanrel[r]):
                        if float(meanrel[r])<10000:
                            meanrel[r]=str(round(float(meanrel[r]), 2))
            if 9.99<float(stdevrel[r]):
                if float(stdevrel[r])<100:
                    if 10<float(meanrel[r]):
                        if float(meanrel[r])<100:
                            meanrel[r]=str(round(float(meanrel[r]), 1))

        if str(stdevabs[r])=='_':
            gorn=0
        else:
            gorn=1
        if str(meanabs[r])=='_':
            gorn=0
        else:
            gorn=gorn
        if gorn==1:
            if 0.000999<float(stdevabs[r]):
                if float(stdevabs[r])<0.01:
                    if 0.001<float(meanabs[r]):
                        if float(meanabs[r])<0.01:
                            meanabs[r]=str(round(float(meanabs[r]), 1))
            if 0.00999<float(stdevabs[r]):
                if float(stdevabs[r])<0.1:
                    if 0.01<float(meanabs[r]):
                        if float(meanabs[r])<0.1:
                            meanabs[r]=str(round(float(meanabs[r]), 1))
            if 0.0999<float(stdevabs[r]):
                if float(stdevabs[r])<1:
                    if 0.1<float(meanabs[r]):
                        if float(meanabs[r])<1:
                            meanabs[r]=str(round(float(meanabs[r]), 1))
            if 0.0999<float(stdevabs[r]):
                if float(stdevabs[r])<1:
                    if 1<float(meanabs[r]):
                        if float(meanabs[r])<10:
                            meanabs[r]=str(round(float(meanabs[r]), 2))
            if 0.999<float(stdevabs[r]):
                if float(stdevabs[r])<10:
                    if 10<float(meanabs[r]):
                        if float(meanabs[r])<100:
                            meanabs[r]=str(round(float(meanabs[r]), 2))
            if 0.999<float(stdevabs[r]):
                if float(stdevabs[r])<10:
                    if 1<float(meanabs[r]):
                        if float(meanabs[r])<10:
                            meanabs[r]=str(round(float(meanabs[r]), 1))
            if 9.99<float(stdevabs[r]):
                if float(stdevabs[r])<100:
                    if 100<float(meanabs[r]):
                        if float(meanabs[r])<1000:
                            meanabs[r]=str(round(float(meanabs[r]), 2))
            if 99.9<float(stdevabs[r]):
                if float(stdevabs[r])<1000:
                    if 1000<float(meanabs[r]):
                        if float(meanabs[r])<10000:
                            meanabs[r]=str(round(float(meanabs[r]), 2))
            if 9.99<float(stdevabs[r]):
                if float(stdevabs[r])<100:
                    if 10<float(meanabs[r]):
                        if float(meanabs[r])<100:
                            meanabs[r]=str(round(float(meanabs[r]), 1))
    # end round mean to adjusted number of significant digits depending on standard deviation
    stdevabs[r]=str(stdevabs[r])
    meanabs[r]=str(meanabs[r])
    stdevrel[r]=str(stdevrel[r])
    meanrel[r]=str(meanrel[r])
    #begin remove .0 where appropriate
    remz=1
    if remz==1:
        if str(meanrel[r][len(meanrel[r])-1])=='0':
            if str(meanrel[r][len(meanrel[r])-2])=='.':
                if 0.1<float(stdevrel[r]):
                    if float(stdevrel[r])<1:
                        ok=1
                    else:
                        meanrel[r]=meanrel[r][:-2]
                else:
                    meanrel[r]=meanrel[r][:-2]
        if str(stdevrel[r][len(stdevrel[r])-1])=='0':
            if str(stdevrel[r][len(stdevrel[r])-2])=='.':
                stdevrel[r]=stdevrel[r][:-2]
        if str(meanabs[r][len(meanabs[r])-1])=='0':
            if str(meanabs[r][len(meanabs[r])-2])=='.':
                #print(meanabs[r])
                if 0.1<float(stdevabs[r]):
                    if float(stdevabs[r])<1:
                        ok=1
                    else:
                        meanabs[r]=meanabs[r][:-2]
                else:
                    meanabs[r]=meanabs[r][:-2]
                #print(meanabs[r])
        if str(stdevabs[r][len(stdevabs[r])-1])=='0':
            if str(stdevabs[r][len(stdevabs[r])-2])=='.':
                stdevabs[r]=stdevabs[r][:-2]
    #end remove .0 where appropriate
    # begin round COV to 1 significant digit
    if str(covrel[r])=='_':
        ok=1
    else:
        covrel[r]=str(round(float(covrel[r]), 1))
    if str(covabs[r])=='_':
        ok=1
    else:
        covabs[r]=str(round(float(covabs[r]), 1))
    # end round COV to 1 significant digit
    wsout.cell(row=r+2, column=1).value=fatlist[r]
    wsout.cell(row=r+2, column=2).value=omeanrel[r]
    wsout.cell(row=r+2, column=3).value=ostdevrel[r]
    wsout.cell(row=r+2, column=4).value=covrel[r]
    wsout.cell(row=r+2, column=5).value=omeanabs[r]
    wsout.cell(row=r+2, column=6).value=ostdevabs[r]
    wsout.cell(row=r+2, column=7).value=covabs[r]
    wsout.cell(row=r+2, column=8).value=relr1[r]
    wsout.cell(row=r+2, column=9).value=relr2[r]
    wsout.cell(row=r+2, column=10).value=relr3[r]
    wsout.cell(row=r+2, column=11).value=absr1[r]
    wsout.cell(row=r+2, column=12).value=absr2[r]
    wsout.cell(row=r+2, column=13).value=absr3[r]
    if str(meanrel[r])=='_':
        wsout.cell(row=r+2, column=15).value=str(meanrel[r])
    else:
        wsout.cell(row=r+2, column=15).value=str(meanrel[r])+' '+'\u00B1'+' '+str(stdevrel[r])

    wsout.cell(row=r+2, column=16).value=str(meanabs[r])+' '+'\u00B1'+' '+str(stdevabs[r])
    r=r+1
wbout.save('Summary_FA_quantification_data.xlsx')
if gui==0:
    print('The data, including Standard deviations and coefficients of variation are saved in Summary_FA_quantification_data.xlsx.')


# Enter conditions for rounding of mean values depending on stdev (no higher precision than stdev allows)

