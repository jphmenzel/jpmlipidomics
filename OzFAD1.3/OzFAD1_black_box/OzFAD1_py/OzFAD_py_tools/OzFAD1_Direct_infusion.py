# -*- coding: UTF-8 -*-

# Jan Philipp Menzel 
# goal: quantitative evaluation of direct infusion derivatized fatty acid mass spectra.
# Notes: imports ms data from excel. makes list with theoretical masses of searched species. numerically integrates species. 
# Loop for sequential spectra evaluation (analysis and subtraction of Process Blank).

import math
import sys
import openpyxl
from openpyxl import Workbook
import pandas as pd
import datetime
#import brainpy
#from brainpy import isotopic_variants
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.styles.colors import RGB
################ DATABASE ## Source: Internetchemie.info
#isotope=["12C", "13C", "1H", "2H", "14N", "15N", "16O", "17O", "18O"] ["19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#imass=[12.00000, 13.00335, 1.0078250322, 2.01410, 14.003074004, 15.00011, 15.99491462, 16.99913, 17.99916]  #[18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
ia=[0.9893, 0.0107, 0.999885, 0.000115, 0.99636, 0.00364, 0.99747, 0.00038, 0.00205] #[100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
#isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
######################################################################
gui=1
if gui==0:
	print('Direct infusion mass spectrum of sample must be in: OzFAD1_raw_ms_di.xlsx') # and Process Blank in: jpmlipidomics_raw_ms_di_process_blank.xlsx (Spectrum starts in row 9.)')

procblank=0		#If 0, Process Blank will not be subtracted
mzmin=305.0						# define m/z region of interest
mzmax=900  #570.0


if gui==0:

	minlenfa=12
	maxlenfa=30
	cutoff=3000
	#famemix=0
	minlenfa=eval(input('Number of C in shortest FA chain?'))
	maxlenfa=eval(input('Number of C in longest FA chain?'))
	#cutoff=eval(input('Enter cutoff for relevant integrals (recommended: start with low value, e.g. 3000, then rerun script with higher value, e.g. 50000 or 200000) :'))
	famemix=eval(input('Is this a spectrum of the derivatized FAME 37 mix? Yes: 1 | No: 0 :'))
	#procblank=eval(input('Subtract Process Blank? Yes: 1 | No: 0 :'))

	# begin determine derivatization agent
	fourlettcode=input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP, PLPC, PLPE) :')
	#fourlettcode='AMPP'
	#fourlettcode='IAMP'
	#print('Edit code, line 49, if derivatization agent not AMPP! Thanks.')
	fourlettcode=str(fourlettcode)
	if fourlettcode=='AMPP':
		cderiv=12
		hderiv=12
		dderiv=0
		nderiv=2
		oderiv=0
		pderiv=0
		ideriv=0
	elif fourlettcode=='NMPA':
		cderiv=7
		hderiv=10
		dderiv=0
		nderiv=2
		oderiv=0
		pderiv=0
		ideriv=0
	elif fourlettcode=='NMPE':
		cderiv=7
		hderiv=9
		dderiv=0
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='MDPE':
		cderiv=7
		hderiv=6
		dderiv=3
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='NEPE':
		cderiv=8
		hderiv=11
		dderiv=0
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='EDPE':
		cderiv=6
		hderiv=6
		dderiv=5
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='NPPE':
		cderiv=9
		hderiv=13
		dderiv=0
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='IAMP':
		cderiv=12
		hderiv=11
		dderiv=0
		nderiv=2
		oderiv=0
		pderiv=0
		ideriv=1
	elif fourlettcode=='PLPC':
		cderiv=8
		hderiv=20
		dderiv=0
		nderiv=1
		oderiv=6
		pderiv=1
		ideriv=0
	elif fourlettcode=='PLPE':
		cderiv=5
		hderiv=14
		dderiv=0
		nderiv=1
		oderiv=6
		pderiv=1
		ideriv=0
	else:
		cderiv=eval(input('Number of C atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		hderiv=eval(input('Number of H atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		dderiv=eval(input('Number of D atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		nderiv=eval(input('Number of N atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		oderiv=eval(input('Number of O atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		pderiv=eval(input('Number of P atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		ideriv=eval(input('Number of I atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	# end determine derivatization agent
else:
	fourlettcode=str(sys.argv[1])	#retrieve arguments from gui
	#print('received:')
	#print(fourlettcode)
	cutoff=float(sys.argv[2])
	minlenfa=int(sys.argv[3])
	maxlenfa=int(sys.argv[4])
	famemix=int(sys.argv[5])
	cderiv=int(sys.argv[6])
	hderiv=int(sys.argv[7])
	dderiv=int(sys.argv[8])
	nderiv=int(sys.argv[9])
	oderiv=int(sys.argv[10])
	pderiv=int(sys.argv[11])
	ideriv=int(sys.argv[12])


	# get other values !!!

#if procblank==1:
	#print('Spectrum for Process Blank needs to be located in file jpmlipidomics_raw_ms_di_process_blank.xlsx')


explistx=[]
explisty=[]

if procblank==1:
	spectrumall=3
else:
	spectrumall=2
#############		# begin evaluation  #######
spectrum=1
while spectrum<spectrumall:
	if spectrum==1:
		if gui==0:
			print('Mass spectrum of Sample is being evaluated ...')
	elif spectrum==2:
		print('Mass spectrum of Process Blank is being evaluated ...')
	explistx=[]
	explisty=[]
	#############		# build explistx and explisty (full experimental list from 450 to 850)
	if spectrum==1:
		wb=openpyxl.load_workbook('OzFAD1_raw_ms_di.xlsx')		# load excel file from home folder
		sheetms=wb['raw_ms']
	elif spectrum==2:
		wb=openpyxl.load_workbook('OzFAD1_raw_ms_di_process_blank.xlsx')		# load excel file from home folder
		sheetms=wb['raw_ms']

	specx=1
	specy=specx+1
	xn=9
	xs=sheetms.cell(row=xn, column=specx)		# go through excel file to mzmin
	x=xs.value
	ys=sheetms.cell(row=xn, column=specy)
	y=ys.value
	while x<mzmin:
		xs=sheetms.cell(row=xn, column=specx)
		x=xs.value
		ys=sheetms.cell(row=xn, column=specy)
		y=ys.value
		xn=xn+1

	while x<mzmax:								# go through excel file, append to lists until mzmax
		explistx.append(x)
		explisty.append(y)
		xs=sheetms.cell(row=xn, column=specx)
		x=xs.value
		ys=sheetms.cell(row=xn, column=specy)
		y=ys.value
		if x is None:
			x=mzmax+1
		xn=xn+1
	#print('length of list of m/z values within range of interest and list of intensities in range:')
	#print(len(explistx))
	#print(len(explisty))
	# begin calculate 100 pt avg (over data without datapoints > 100000), add each point below it to a background list, and subtract from explisty, if value neg set zero
	hptlist=[]
	hptavglist=[]
	hptavgxlist=[]
	hexplist=[]
	r=0
	while r<(len(explisty)):
		if len(hptlist)<100:
			cuta=20000
			if explisty[r]<cuta:
				cy=explisty[r]
			else:
				if len(hptlist)>1:
					cy=min(hptlist)
				else:
					cy=0
			hptlist.append(cy)
			hexplist.append(explisty[r])
		else:
			avg=sum(hptlist)/len(hptlist)
			cut=min(hexplist)
			if cut<(cuta/4):
				cut=cuta
			else:
				cut=cut*2
			hptavglist.append(avg)
			avgx=r-50
			hptavgxlist.append(avgx)
			if explisty[r]<cut:
				cy=explisty[r]
			else:
				cy=min(hptlist)
			hptlist.append(cy)
			hexplist.append(explisty[r])
			del hptlist[0]
			del hexplist[0]
		r=r+1
	#print(len(hptavglist))
	bgylist=[]
	r=0
	t=0
	while r<(len(explisty)):
		if explistx[r]<hptavgxlist[0]:
			bgylist.append(hptavglist[0])
		elif explistx[r]>hptavgxlist[len(hptavgxlist)-1]:
			bgylist.append(hptavglist[len(hptavglist)-1])
		else:
			if t<(len(hptavglist)):
				bgylist.append(hptavglist[t])
			else:
				bgylist.append(hptavglist[len(hptavglist)-1])
			#print(hptavglist[t])
			t=t+1
		r=r+1
	# hptavglist is baseline
	#print(len(bgylist))
	explistyc=[] 	# baseline corrected spectrum
	r=0
	while r<(len(explisty)):
		corr=explisty[r]-bgylist[r]
		if corr<0:
			corr=0
		explistyc.append(corr)
		r=r+1
	# end calculate 100 pt avg, add each point below it to a background list, and subtract from explisty, if value neg set zero
	# begin save baseline corrected data
	savethis=spectrum #0 # spectrum, if original spectrum shall be saved. 1 if Process Blank spectrum shall be saved
	if savethis==1:
		wb = Workbook() #write_only=True)
		ws=wb.active
		r=0
		kr=2
		while r<len(explistyc):
			ws.cell(row=kr, column=1).value=explistx[r]	# write mz values
			ws.cell(row=kr, column=2).value=explisty[r]	# write original values
			ws.cell(row=kr, column=3).value=explistyc[r]	# write baseline corrected values
			ws.cell(row=kr, column=4).value=bgylist[r]	# write baseline values
			r=r+1
			kr=kr+1
		wb.save('ms_baseline_corrected.xlsx')
		#quit()
	# end save baseline corrected data

	# begin apply mz (lockmass) correction based on AMPP_16:0
	#lockmassa=423.3375	# AMPP_16:0
	#lockmassb=451.3688	# AMPP_18:0

	# begin sum formula dependent calculation of lockmasslist
	#isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
	#imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
	palmiticacidmz=imass[2]*(cderiv+16)+imass[0]*(hderiv+31)+imass[1]*(dderiv)+imass[3]*(nderiv)+imass[4]*(oderiv+1)+imass[5]*(pderiv)+imass[10]*(ideriv)
	stearicacidmz=imass[2]*(cderiv+18)+imass[0]*(hderiv+35)+imass[1]*(dderiv)+imass[3]*(nderiv)+imass[4]*(oderiv+1)+imass[5]*(pderiv)+imass[10]*(ideriv)
	if gui==0:
		print('palmiticacidmz')
		print(palmiticacidmz)
		print('stearicacidmz')
		print(stearicacidmz)
	lockmasslist=[]
	lockmasslist.append(palmiticacidmz)
	lockmasslist.append(stearicacidmz)
	# end sum formula dependent calculation of lockmasslist
	#print(lockmasslist)
	shiftlist=[]
	fwhmlist=[]
	lmk=0
	while lmk<2:
		k=0
		while explistx[k]<lockmasslist[lmk]:
			k=k+1
		k=k-1
		#print('k')
		#print(k)
		#print(explistyc[k])
		if explistyc[k]>0:
			if explistyc[k-1]<explistyc[k]:	# theoretical peak before experimental peak maximum
				check=1
				check2=0
				while check>0:
					if explistyc[k-1]>0:
						if explistyc[k-1]<explistyc[k]:
							check=1
						else:
							check=0
					else:
						check=0
						check2=1
					if check==1:
						k=k+1
					else:
						k=k
				k=k-1
				if check2==1:
					k=k+1
				# peak mz identified, begin calc lockmass correction
				shift=lockmasslist[lmk]-explistx[k]
				shiftlist.append(shift)
				# end calc lockmass correction
			elif explistyc[k-1]>explistyc[k]:	# theoretical peak after experimental peak maximum
				check=1
				while check>0:
					if explistyc[k-1]>0:
						if explistyc[k-1]>explistyc[k]:
							check=1
						else:
							check=0
					else:
						check=0
					if check==1:
						k=k-1
					else:
						k=k
				k=k
				# peak mz identified, begin calc lockmass correction
				shift=lockmasslist[lmk]-explistx[k]
				shiftlist.append(shift)
				# end calc lockmass correction
			else:
				k=k
			# begin determine FWHM
			#print(shiftlist)
			#print(explistyc[k])
			fwi=k
			while explistyc[fwi]>(0.5*explistyc[k]):
				fwi=fwi-1
			lfwi=fwi
			fwi=k
			while explistyc[fwi]>(0.5*explistyc[k]):
				fwi=fwi+1
			ufwi=fwi
			#print(fwi)
			#print(lfwi)
			#print(ufwi)
			fwhm=(explistx[ufwi-1]+((explistx[ufwi]-explistx[ufwi-1])*((explistyc[ufwi-1]-(0.5*explistyc[k]))/(explistyc[ufwi-1]-explistyc[ufwi]))))-(explistx[lfwi+1]-((explistx[lfwi+1]-explistx[lfwi])*((explistyc[lfwi+1]-(0.5*explistyc[k]))/(explistyc[lfwi+1]-explistyc[lfwi]))))
			fwhmlist.append(fwhm)
			# end determine FWHM 
		lmk=lmk+1
	#print(shiftlist)
	#print(fwhmlist)
	if spectrum==1:
		normfwhm=(fwhmlist[0]+fwhmlist[1])/2
	# begin apply lockmass correction
	avgshift=(shiftlist[0]+shiftlist[1])/2
	p=0
	while p<(len(explistx)):
		explistx[p]=explistx[p]+avgshift
		p=p+1
	# end apply lockmass correction
	# end apply mz correction based on AMPP_16:0 and 18:0
	#print(normfwhm)
	########## begin read transitionlist for direct injection analysis (precursor only)
	excelread=0
	if excelread==1:
		trdf=pd.read_csv('jpmlipidomics_precursor_di.csv')
		toprowx=[trdf.columns.values.tolist()]
		toprow=toprowx[0]
		trdf=trdf.transpose()
		writelist=trdf.values.tolist()
		ki=len(writelist[0])
	#print('Number of rows in jpmlipidomics_precursor_di.csv (Transition list containing AMPP FA without rt variation) : %d' % ki)
	########## end read transitionlist for direct injection analysis (precursor only)

	# begin calculate writelist in case of other derivatization agent than AMPP - independent from jpmlipidomics_precursor_di.csv
	# begin make productname list prodnmwrite
	minlenfa=int(minlenfa)
	clenfa=minlenfa
	maxlenfa=int(maxlenfa)
	prodnmwrite=[]
	cdb=1
	while cdb<7:		#unsaturated FA precursor
		clenfa=minlenfa
		if cdb==2:
			if clenfa<5:
				clenfa=5
		elif cdb==3:
			if clenfa<7:
				clenfa=7
		elif cdb==4:
			if clenfa<9:
				clenfa=9
		elif cdb==5:
			if clenfa<11:
				clenfa=11
		elif cdb==6:
			if clenfa<13:
				clenfa=13
		while clenfa<(maxlenfa+1):
			if clenfa<10:
				prodnm=fourlettcode+'_0'+str(clenfa)+':'+str(cdb)
			else:
				prodnm=fourlettcode+'_'+str(clenfa)+':'+str(cdb)
			prodnmwrite.append(prodnm)
			clenfa=clenfa+1
		cdb=cdb+1
	clenfa=minlenfa		# saturated FA precursor
	while clenfa<(maxlenfa+1):
		if clenfa<10:
			prodnm=fourlettcode+'_0'+str(clenfa)+':'+'0'
		else:
			prodnm=fourlettcode+'_'+str(clenfa)+':'+'0'
		prodnmwrite.append(prodnm)
		clenfa=clenfa+1
	# end make productname list prodnmwrite
	# begin make product formula list prodformulawrite
	prodformulawrite=[]
	prodmzwrite=[]
	mpf=0
	while mpf<(len(prodnmwrite)):
		pc=cderiv+int(prodnmwrite[mpf][6])+(10*(int(prodnmwrite[mpf][5])))
		ph=hderiv+(((int(prodnmwrite[mpf][6])+(10*(int(prodnmwrite[mpf][5]))))-2)*2)+3-(2*(int(prodnmwrite[mpf][8])))
		po=oderiv+1
		prodfm='C'+str(pc)+'H'+str(ph)+"H'"+str(dderiv)+'N'+str(nderiv)+'O'+str(po)+'P'+str(pderiv)+'I'+str(ideriv)
		prodformulawrite.append(prodfm)
		prodmz=imass[2]*pc+imass[0]*ph+imass[1]*dderiv+imass[3]*nderiv+imass[4]*po+imass[5]*pderiv+imass[10]*ideriv-imass[8]
		prodmzwrite.append(prodmz)
		mpf=mpf+1
	# end make product formula list prodformulawrite
	# begin append values to writelist for deuterated palmitic and stearic acid standards
	#if spectrum==1:
		#print('Mass spectrum of Sample is being evaluated ...')
	dpalmiticprodnm=str('AMPP')+'_d31_'+str('16')+':'+str(0)
	dstearicprodnm=str('AMPP')+'_d35_'+str('18')+':'+str(0)
	prodnmwrite.append(dpalmiticprodnm)
	prodnmwrite.append(dstearicprodnm)
	prodfm='C'+str(28)+'H'+str(12)+"H'"+str(31)+'N'+str(2)+'O'+str(1)+'P'+str(0)+'I'+str(0)
	prodformulawrite.append(prodfm)
	prodfm='C'+str(30)+'H'+str(12)+"H'"+str(35)+'N'+str(2)+'O'+str(1)+'P'+str(0)+'I'+str(0)
	prodformulawrite.append(prodfm)
	prodmz=imass[2]*28+imass[0]*12+imass[1]*31+imass[3]*2+imass[4]*1+imass[5]*0+imass[10]*0-imass[8]
	prodmzwrite.append(prodmz)
	prodmz=imass[2]*30+imass[0]*12+imass[1]*35+imass[3]*2+imass[4]*1+imass[5]*0+imass[10]*0-imass[8]
	prodmzwrite.append(prodmz)
	#elif spectrum==2:
		#print('Mass spectrum of Process Blank is being evaluated ...')
	
	# end append values to writelist for deuterated palmitic and stearic acid standards
	writelist=[]
	writelist.append(prodnmwrite)
	writelist.append(prodformulawrite)
	writelist.append(prodmzwrite)
	#print(writelist)
	#quit()
	# end calculate writelist in case of other derivatization agent than AMPP

	################	# start identify and integrate peaks, make integrallists
	integrallisty=[]
	integrallistx=[]
	integrallistz=[]
	warninglist=[]
	i=0
	while i<len(writelist[0]):			# go through theoretical list of peaks systematically
		k=0
		while writelist[2][i]>explistx[k]:	# go to m/z value of theoretical peak in experimental list
			k=k+1
		xpeak=explistx[k]
		npeak=writelist[0][i]
	
		integral=0				# identify beginning of peak
		if explistyc[k]>0:
			if explistyc[k-1]<explistyc[k]:	# theoretical peak before experimental peak maximum
				check=1
				while check>0:
					if explistyc[k-1]>0:
						if explistyc[k-1]<explistyc[k]:
							check=1
						else:
							check=0
					else:
						check=0
					if check==1:
						k=k-1
					else:
						k=k
				checkint=1			# start integrate peak
				stop=0
				peakdetect=0
				while checkint>0:
					if explistyc[k+1]>0:
						if explistyc[k+1]>explistyc[k]:
							if stop==0:
								checkint=1
								integral=integral+((explistyc[k]+explistyc[k+1])/2)*(explistx[k+1]-explistx[k])
								k=k+1
							else:
								checkint=0
						else:
							if peakdetect==0:
								peakmax=explistx[k]
								peakdetect=1
							stop=1
							checkint=1
							integral=integral+((explistyc[k]+explistyc[k+1])/2)*(explistx[k+1]-explistx[k])
							k=k+1   # end integrate peak
					else:
						checkint=0
				if abs(peakmax-writelist[2][i])>(1.5*normfwhm):
					# begin integrate within limits of +- 1 FWHM
					mzb=writelist[2][i]-normfwhm
					mze=writelist[2][i]+normfwhm
					integral=0
					k=0
					while mzb>explistx[k]:	# go to m/z value of theoretical peak minus normfwhm in experimental list
						k=k+1
					while mze>explistx[k]:	# integrate to m/z value of theoretical peak plu normfwhm in experimental list
						integral=integral+((explistyc[k]+explistyc[k+1])/2)*(explistx[k+1]-explistx[k])
						k=k+1
					# end integrate within limits of +- 1 FWHM
					integrallisty.append(integral)
					integrallistx.append(xpeak)
					integrallistz.append(npeak)
					specname=str(writelist[0][i][5])+str(writelist[0][i][6])+str(writelist[0][i][7])+str(writelist[0][i][8])
					warningtext='Integration: theor. m/z +- '+str(round(normfwhm,4))
					warninglist.append(warningtext)
					#print('The peak of species %s is close to a major contaminant - integration limits were arbitrarily set as +- 1 FWHM and the respective integration needs to be checked manually.' % specname)
				else:
					integrallisty.append(integral)
					integrallistx.append(xpeak)
					integrallistz.append(npeak)
					warninglist.append('')
			else:
				check=1		# theoretical peak after experimental peak maximum
				while check>0:
					if explisty[k+1]>0:
						if explistyc[k+1]<explistyc[k]:
							check=1
						else:
							check=0
					else:
						check=0
					if check==1:
						k=k+1
					else:
						k=k
				checkint=1			# start integrate peak
				stop=0
				peakdetect=0
				while checkint>0:
					if explistyc[k-1]>0:
						if explistyc[k-1]>explistyc[k]:
							if stop==0:
								checkint=1
								integral=integral+((explistyc[k]+explistyc[k-1])/2)*(explistx[k]-explistx[k-1])
								k=k-1
							else:
								checkint=0
						else:
							if peakdetect==0:
								peakmax=explistx[k]
								peakdetect=1
							stop=1
							checkint=1
							integral=integral+((explistyc[k]+explistyc[k-1])/2)*(explistx[k]-explistx[k-1])
							k=k-1   # end integrate peak
					else:
						checkint=0
				if abs(peakmax-writelist[2][i])>(1.5*normfwhm):
					# begin integrate within limits of +- 1 FWHM
					mzb=writelist[2][i]-normfwhm
					mze=writelist[2][i]+normfwhm
					integral=0
					k=0
					while mzb>explistx[k]:	# go to m/z value of theoretical peak minus normfwhm in experimental list
						k=k+1
					while mze>explistx[k]:	# integrate to m/z value of theoretical peak plu normfwhm in experimental list
						integral=integral+((explistyc[k]+explistyc[k+1])/2)*(explistx[k+1]-explistx[k])
						k=k+1
					# end integrate within limits of +- 1 FWHM
					integrallisty.append(integral)
					integrallistx.append(xpeak)
					integrallistz.append(npeak)
					specname=str(writelist[0][i][5])+str(writelist[0][i][6])+str(writelist[0][i][7])+str(writelist[0][i][8])
					warningtext='Integration: theor. m/z +- '+str(round(normfwhm,4))
					warninglist.append(warningtext)
					#print('The peak of species %s is close to a major contaminant - integration limits were arbitrarily set as +- 1 FWHM and the respective integration needs to be checked manually.' % specname)
				else:
					integrallisty.append(integral)
					integrallistx.append(xpeak)
					integrallistz.append(npeak)
					warninglist.append('')
		else:
			integrallisty.append(integral)
			integrallistx.append(xpeak)
			integrallistz.append(npeak)
			warninglist.append('')
		i=i+1						# all relevant peaks integrated
	#print(integrallistz)
	#print(integrallistx)
	#print(integrallisty)
	################	# end identify and integrate peaks 
	#quit()
	# begin overlap correction (e.g. +2 isomer of 20:5 subtracted from monoisotopic ion of 20:4 ...)
	ndb=6
	k=0
	while ndb>0:
		while k<(len(writelist[0])):
			if writelist[0][k][8]==ndb:
				# begin calculate +2 isotope integral contribution
				comment=0		# set comment=0, when using brainpy module to calculate +2 isotope, else if AMPP can use comment=1
				if comment==1:
					cn=10*int(writelist[1][k][1])+int(writelist[1][k][2])
					hn=10*int(writelist[1][k][4])+int(writelist[1][k][5])
					nn=2
					on=1
					# sum formula is defined
					monoiso=ia[0]**cn*ia[2]**hn*ia[4]**2*ia[6]**1		# fraction of the isotopic pattern being the monoisotopic ion
					sa=ia[0]**cn*ia[1]**0*ia[2]**hn*ia[3]**0*ia[4]**2*ia[5]**0*ia[6]**0*ia[7]**0*ia[8]**1	# all combinations of isotopes contributing to +2 peak (fractions)
					sb=ia[0]**cn*ia[1]**0*ia[2]**hn*ia[3]**0*ia[4]**1*ia[5]**1*ia[6]**0*ia[7]**1*ia[8]**0
					sc=ia[0]**cn*ia[1]**0*ia[2]**(hn-1)*ia[3]**1*ia[4]**2*ia[5]**0*ia[6]**0*ia[7]**1*ia[8]**0
					sd=ia[0]**(cn-1)*ia[1]**1*ia[2]**hn*ia[3]**0*ia[4]**2*ia[5]**0*ia[6]**0*ia[7]**1*ia[8]**0
					se=ia[0]**cn*ia[1]**0*ia[2]**hn*ia[3]**0*ia[4]**0*ia[5]**2*ia[6]**1*ia[7]**0*ia[8]**0
					sf=ia[0]**cn*ia[1]**0*ia[2]**(hn-1)*ia[3]**1*ia[4]**1*ia[5]**1*ia[6]**1*ia[7]**0*ia[8]**0
					sg=ia[0]**(cn-1)*ia[1]**1*ia[2]**hn*ia[3]**0*ia[4]**1*ia[5]**1*ia[6]**1*ia[7]**0*ia[8]**0
					sh=ia[0]**cn*ia[1]**0*ia[2]**(hn-2)*ia[3]**2*ia[4]**2*ia[5]**0*ia[6]**1*ia[7]**0*ia[8]**0
					si=ia[0]**(cn-1)*ia[1]**1*ia[2]**(hn-1)*ia[3]**1*ia[4]**2*ia[5]**0*ia[6]**1*ia[7]**0*ia[8]**0
					sj=ia[0]**(cn-2)*ia[1]**2*ia[2]**hn*ia[3]**0*ia[4]**2*ia[5]**0*ia[6]**1*ia[7]**0*ia[8]**0
					fpt=(sa+sb+sc+sd+se+sf+sg+sh+si+sj)/monoiso		# monoisotopic ion integral * fpt = +2 ion
					ptcontrib=integrallisty[k]*fpt					# contribution of k species (+2 isotope) to m species
				else:
					# begin read precursor sum formula and edit product sum formula
					e=str(writelist[7][k]) #get formula from csv
					clist=[]
					hlist=[]
					dlist=[]
					nlist=[]
					olist=[]
					plist=[]
					ilist=[]
					i=0
					ca=0
					ha=0
					da=0
					na=0
					oa=0
					pa=0
					ia=0
					while i<len(e):
						if e[i]=='H':
							if e[i+1]=="'":
								ha=0
							else:
								ca=0
						#if e[i]=="'":
						#	ha=0		
						if e[i]=='N':
							ha=0
							da=0
						if e[i]=='O':
							ha=0
							da=0
							na=0
						if e[i]=='P':
							ha=0
							da=0
							na=0
							oa=0
						if e[i]=='I':
							ha=0
							da=0
							na=0
							oa=0
							pa=0
						if ca==1:
							clist.append(e[i])
						if ha==1:
							hlist.append(e[i])
						if da==1:
							dlist.append(e[i])
						if na==1:
							nlist.append(e[i])
						if oa==1:
							olist.append(e[i])
						if pa==1:
							plist.append(e[i])
						if ia==1:
							ilist.append(e[i])
						if e[i]=='C':
							ca=1
						if e[i]=='H':
							if e[i+1]=="'":
								ca=0
								ha=0
								da=1
								i=i+1
							else:
								ca=0
								ha=1
						#if e[i]=="'":
						#	ca=0
						#	ha=0
						#	da=1		
						if e[i]=='N':
							ha=0
							da=0
							na=1
							if e[i+1]=='O':
								nlist.append('1')
								na=0
						if e[i]=='O':
							ha=0
							da=0
							na=0
							oa=1
							if (i+1)<len(e):
								if e[i+1]=='P':
									olist.append('1')
									oa=0
							else:
								olist.append('1')
								oa=0					
						if e[i]=='P':
							da=0
							na=0
							oa=0
							pa=1
							if (i+1)<len(e):
								if e[i+1]=='I':
									plist.append('1')
									pa=0
							else:
								plist.append('1')
								pa=0
						if e[i]=='I':
							da=0
							na=0
							oa=0
							pa=0
							ia=1
							if i==(len(e)-1):
								ilist.append('1')
								ia=0
						i=i+1
					#print(clist)
					#print(hlist)
					#print(dlist)
					#print(nlist)
					#print(olist)
					#print(plist)
					if len(clist)==0:
						cn=0
					if len(hlist)==0:
						hn=0
					if len(dlist)==0:
						dn=0	
					if len(nlist)==0:
						nn=0
					if len(olist)==0:
						on=0
					if len(plist)==0:
						pn=0
					if len(ilist)==0:
						iodon=0
					if len(clist)==1:
						cn=int(clist[0])
					if len(clist)==2:
						cn=10*int(clist[0])+int(clist[1])
					if len(clist)==3:
						cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
					if len(hlist)==1:
						hn=int(hlist[0])
					if len(hlist)==2:
						hn=10*int(hlist[0])+int(hlist[1])
					if len(hlist)==3:
						hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
					if len(dlist)==1:
						dn=int(dlist[0])
					if len(dlist)==2:
						dn=10*int(dlist[0])+int(dlist[1])
					if len(dlist)==3:
						dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
					if len(nlist)==1:
						nn=int(nlist[0])
					if len(nlist)==2:
						nn=10*int(nlist[0])+int(nlist[1])
					if len(nlist)==3:
						nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
					if len(olist)==1:
						on=int(olist[0])
					if len(olist)==2:
						on=10*int(olist[0])+int(olist[1])
					if len(olist)==3:
						on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
					if len(plist)==1:
						pn=int(plist[0])
					if len(plist)==2:
						pn=10*int(plist[0])+int(plist[1])
					if len(plist)==3:
						pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])	
					if len(ilist)==1:
						iodon=int(ilist[0])
					if len(ilist)==2:
						iodon=10*int(ilist[0])+int(ilist[1])
					if len(ilist)==3:
						iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		
					# end read precursor sum formula
					isopatternmz=[]
					isopatternint=[]
					transition={'H':hn-1, 'C':cn, 'O':on, 'N':nn, 'P':pn, 'I':iodon}
					theoretical_isotopic_cluster=isotopic_variants(transition, npeaks=5, charge=1)
					for peak in theoretical_isotopic_cluster:
						#print(peak.mz)
						cpkmz=float(peak.mz)
						isopatternmz.append(cpkmz)
						#print(peak.intensity)
						cpkint=float(peak.intensity)
						isopatternint.append(cpkint)
					#print(isopatternmz)
					#print(isopatternint)
					fpt=isopatternint[2]/isopatternint[0]		# monoisotopic ion integral * fpt = +2 ion
					ptcontrib=integrallisty[k]*fpt				# contribution of k species (+2 isotope) to m species

				# end calculate +2 isotope integral contribution
				m=0
				while m<(len(writelist[0])):
					csp=str(writelist[0][k][5])+str(writelist[0][k][6])+str(writelist[0][k][7])+str(writelist[0][k][8])
					ssp=str(writelist[0][m][5])+str(writelist[0][m][6])+str(writelist[0][m][7])+str(int(writelist[0][m][8])+1)
					if csp==ssp:
						# begin subtract +2 isotope integral contribution from m species
						integrallisty[m]=integrallisty[m]-ptcontrib
						if integrallisty[m]<0:
							integrallisty[m]=0
						# end subtract +2 isotope integral contribution from m species
					m=m+1
			k=k+1
		ndb=ndb-1
	# end overlap correction (e.g. +2 isomer of 20:5 subtracted from monoisotopic ion of 20:4 ...)
	# begin general isotope correction from monoisotopic ion to whole isotopic distribution
	k=0
	while k<(len(writelist[0])):
		cn=10*int(writelist[1][k][1])+int(writelist[1][k][2])
		hn=10*int(writelist[1][k][4])+int(writelist[1][k][5])
		nn=2
		on=1
		monoiso=ia[0]**cn*ia[2]**hn*ia[4]**2*ia[6]**1		# fraction of the isotopic pattern being the monoisotopic ion
		integrallisty[k]=integrallisty[k]/monoiso
		k=k+1
	# end general isotope correction from monoisotopic ion to whole isotopic distribution
	# begin filter out species with low integral
	#cutoff=50000

	#if spectrum==2:		## Only activate if certain that cutoff should be applied to PB !!!
	#	cutoff=eval(input('Use cutoff for PB?'))
	#	#cutoff=3000

	fintlisty=[]
	fintlistn=[]
	k=0
	while k<(len(writelist[0])):
		if integrallisty[k]<cutoff:
			k=k
		else:
			fintlisty.append(integrallisty[k])
			i=1
			cfint=str(writelist[0][k][0])
			while i<(len(writelist[0][k])): #9:
				cfint=str(cfint)+str(writelist[0][k][i])
				i=i+1
			fintlistn.append(cfint)
		k=k+1
	# end filter out species with low integral	
	# begin normalize selected integrals
	nintlisty=[]
	sumint=sum(fintlisty)
	k=0
	while k<(len(fintlisty)):
		cnorm=fintlisty[k]/sumint
		nintlisty.append(cnorm)
		k=k+1
	# end normalize selected integrals
	# begin add data for expected normalized mol% in FAME 37mix
	if spectrum==2:
		famemix=0
	if famemix==1:
		allfamespecieslist=['14:1', '15:1', '16:1', '17:1', '18:1', '20:1', '22:1', '24:1', '18:2', '20:2', '22:2', '18:3', '20:3', '20:4', '20:5', '22:6', '04:0', '06:0', '08:0', '10:0', '11:0', '12:0', '13:0', '14:0', '15:0', '16:0', '17:0', '18:0', '20:0', '21:0', '22:0', '23:0', '24:0']
		allfamemolplist=[2.06, 1.94, 1.84, 1.75, 5.01, 1.52, 1.4, 1.3, 3.36, 1.53, 1.41, 3.38, 3.08, 1.55, 1.56, 1.44, 9.69, 7.60, 6.25, 5.31, 2.47, 4.62, 2.17, 4.08, 1.93, 5.49, 1.74, 3.31, 3.03, 1.45, 2.79, 1.34, 2.59]
		famespecieslist=[]
		famemolplist=[]
		k=0
		while k<(len(fintlistn)):
			csp=str(fintlistn[k][5])+str(fintlistn[k][6])+str(fintlistn[k][7])+str(fintlistn[k][8])
			fsi=0
			found=0
			while fsi<(len(allfamespecieslist)):
				if allfamespecieslist[fsi]==csp:
					famespecieslist.append(allfamespecieslist[fsi])
					famemolplist.append(allfamemolplist[fsi])
					found=1
				else:
					found=found
				fsi=fsi+1
			if found==0:
				famespecieslist.append(csp)
				famemolplist.append(0)
			k=k+1
		k=0
		sumfmpl=sum(famemolplist)
		while k<(len(famemolplist)):
			famemolplist[k]=famemolplist[k]/sumfmpl
			k=k+1
	# end add data for expected normalized mol% in FAME 37mix
	# begin save data 
	if spectrum==1:
		wb = Workbook() #write_only=True)
		#ws = wb.create_sheet()
		ws=wb.active
		toprow=['m/z', 'Original intensity', 'Baseline corrected intensity', 'Baseline', '_', 'All AMPP_FA', 'All Integrals (Isotope corrected)', 'All m/z (found)', 'Interference', 'AMPP_FA (found)', 'Integral (Isotope corrected)', 'Normalized integral (Isotope corrected)', 'Expected normalized amount in FAME 37mix']
		c=0
		while c<(len(toprow)):
			wc=c+1
			ws.cell(row=1, column=wc).value=toprow[c]	#write top row
			c=c+1
		r=0
		kr=2
		while r<len(explistyc):
			ws.cell(row=kr, column=1).value=explistx[r]		# write mz values
			ws.cell(row=kr, column=2).value=explisty[r]		# write original values
			ws.cell(row=kr, column=3).value=explistyc[r]	# write baseline corrected values
			ws.cell(row=kr, column=4).value=bgylist[r]		# write baseline values
			r=r+1
			kr=kr+1
		r=0
		kr=2
		while r<len(writelist[0]):
			if (len(writelist[0][r]))==9:
				species=fourlettcode+str('_')+str(writelist[0][r][5])+str(writelist[0][r][6])+str(writelist[0][r][7])+str(writelist[0][r][8])
			else:
				species=writelist[0][r]
			ws.cell(row=kr, column=6).value=species		# write list of species (full)
			ws.cell(row=kr, column=7).value=integrallisty[r]	# write integrals of species (full)
			ws.cell(row=kr, column=8).value=integrallistx[r]	# write found m/z (m/z corrected) of species (full)
			ws.cell(row=kr, column=9).value=warninglist[r]	#write warnings
			r=r+1
			kr=kr+1
		r=0
		kr=2
		lenfintlistn=len(fintlistn)
		while r<len(fintlistn):
			dlt=1						#################################################################### shortened name for better labelling in graph
			if dlt==1:
				if 'AMPP' in str(fintlistn[r]):
					sh=5
					#print('#######')
					nfai=str()
					while sh<(len(fintlistn[r])):
						nfai=nfai+str(fintlistn[r][sh])
						sh=sh+1
					fintlistn[r]=nfai
			ws.cell(row=kr, column=10).value=fintlistn[r]		# write list of species (found)
			ws.cell(row=kr, column=11).value=fintlisty[r]		# write integrals of species (found)
			r=r+1
			kr=kr+1
		if famemix==1:
			r=0
			kr=2
			while r<len(famemolplist):
				ws.cell(row=kr, column=12).value=nintlisty[r]		# write normalized integral of species (found)				
				ws.cell(row=kr, column=13).value=famemolplist[r]		# write normalized expected mol percent in FAME 37 mix
				r=r+1
				kr=kr+1

		wb.save('ms_di_fa_integrals.xlsx')
	# end save data
	if procblank==0:
		if gui==0:
			print('Done. Data is saved in ms_di_fa_integrals.xlsx.')
		quit()
	# begin subtract process blank
	if spectrum==2:
		#read file containing data prior to process blank subtraction, substract process blank, then save
		uncwb=openpyxl.load_workbook('ms_di_fa_integrals.xlsx')		# load excel file from home folder
		uncws=uncwb.active
		uncws.title='data'
		toprowpb=['_', 'AMPP_FA (found)', 'Integral after PB subtr.', 'Norm. integral after PB subtr.', 'Expected norm. amount in FAME 37mix', '_', 'PB: m/z', 'PB: Original MS', 'PB: Baseline corrected', 'PB: Baseline', '_', 'PB: AMPP_FA', 'PB: Integral', 'PB: m/z found', 'Interference', 'PB: AMPP_FA found', 'PB: Integral']
		c=14
		while c<(len(toprowpb)+14):
			pbc=c-14
			uncws.cell(row=1, column=c).value=toprowpb[pbc]	#write top row
			c=c+1
		newintlist=[]
		# begin determine dstcorrect
		kr=2
		go=1
		while go==1:
			csps=uncws.cell(row=kr, column=10)
			csps=csps.value
			if csps is None:
				go=0
			csps=str(csps)
			if csps=='AMPP_d31_16:0':
				cintpa=uncws.cell(row=kr, column=11)
				cintpa=cintpa.value
			elif csps=='d31_16:0':
				cintpa=uncws.cell(row=kr, column=11)
				cintpa=cintpa.value
			elif csps=='AMPP_d35_18:0':
				cintsa=uncws.cell(row=kr, column=11)
				cintsa=cintsa.value
				go=0
			elif csps=='d35_18:0':
				cintsa=uncws.cell(row=kr, column=11)
				cintsa=cintsa.value
				go=0
			kr=kr+1
		#pbintpa=fintlisty[len(fintlisty)-2]
		pbintsa=fintlisty[len(fintlisty)-1]
		dstcorrect=cintsa/pbintsa  # Process blank subtraction based on 18:0_d35 only #((cintpa/pbintpa)+(cintsa/pbintsa))/2
		# end determine dstcorrect
		r=0
		kr=2
		while r<(lenfintlistn):
			csps=uncws.cell(row=kr, column=10)
			csps=csps.value
			csps=str(csps)
			cexpmolp=uncws.cell(row=kr, column=13)
			cexpmolp=cexpmolp.value
			q=0
			found=0
			while q<(len(fintlistn)):
				if csps==str(fintlistn[q]):
					cint=uncws.cell(row=kr, column=11)
					cint=cint.value
					cint=float(cint)
					if csps=='AMPP_19:0':
						nint=cint-(fintlisty[q]*dstcorrect) 	# activate, if deuterated FA is used as internal standard for both sample and Process blank
					elif csps=='AMPP_d31_16:0':
						nint=cint
					elif csps=='d31_16:0':
						nint=cint
					else:
						nint=cint-(fintlisty[q]*dstcorrect)		# SUBTRACTION OF PROCESS BLANK, dstcorrect is factor calc. from ratio of deuterated standards in sample and Process blank
					if nint<0:
						nint=0
					uncws.cell(row=kr, column=15).value=csps	#write AMPP_FA (found)
					uncws.cell(row=kr, column=16).value=nint	#write new process blank subtracted data
					newintlist.append(nint)
					uncws.cell(row=kr, column=18).value=cexpmolp	#write expected normalized amount in FAME 37mix
					found=1
				else:
					found=found
				q=q+1
			if found==0:
				cint=uncws.cell(row=kr, column=11)
				cint=cint.value
				cint=float(cint)
				nint=cint
				uncws.cell(row=kr, column=15).value=csps	#write AMPP_FA (found)
				uncws.cell(row=kr, column=16).value=nint	#write new process blank subtracted data
				newintlist.append(nint)
				uncws.cell(row=kr, column=18).value=cexpmolp	#write expected normalized amount in FAME 37mix
			r=r+1
			kr=kr+1
		nsumint=sum(newintlist)
		#normintlist=[]
		r=0
		kr=2
		while r<(len(newintlist)): #(lenfintlistn+2):
			normint=newintlist[r]/nsumint
			#normintlist.append(normint)
			uncws.cell(row=kr, column=17).value=normint	#write new process blank subtracted data
			r=r+1
			kr=kr+1
		r=0
		kr=2
		while r<len(explistyc):
			uncws.cell(row=kr, column=20).value=explistx[r]		# write mz values
			uncws.cell(row=kr, column=21).value=explisty[r]		# write original values
			uncws.cell(row=kr, column=22).value=explistyc[r]	# write baseline corrected values
			uncws.cell(row=kr, column=23).value=bgylist[r]		# write baseline values
			r=r+1
			kr=kr+1
		r=0
		kr=2
		while r<len(writelist[0]):
			species=fourlettcode+str('_')+str(writelist[0][r][5])+str(writelist[0][r][6])+str(writelist[0][r][7])+str(writelist[0][r][8])
			uncws.cell(row=kr, column=25).value=species		# write list of species (full)
			uncws.cell(row=kr, column=26).value=integrallisty[r]	# write integrals of species (full)
			uncws.cell(row=kr, column=27).value=integrallistx[r]	# write found m/z (m/z corrected) of species (full)
			uncws.cell(row=kr, column=28).value=warninglist[r]	#write warnings
			r=r+1
			kr=kr+1
		r=0
		kr=2
		while r<len(fintlistn):
			uncws.cell(row=kr, column=29).value=fintlistn[r]		# write list of species (found)
			uncws.cell(row=kr, column=30).value=fintlisty[r]		# write integrals of species (found)
			r=r+1
			kr=kr+1
		#uncwb.save('ms_di_ampp_fa_integrals_pb_subtracted.xlsx')
		#print('Done. Data is saved in ms_di_ampp_fa_integrals_pb_subtracted.xlsx.')
		# end subtract process blank
		# begin plot final results (bar chart) in new sheet
		ws=uncwb.create_sheet()
		ws.title='plot'
		#uncwb.save('ms_di_ampp_fa_integrals_pb_subtracted.xlsx')

		chart1 = BarChart()
		chart1.type = "col"
		chart1.style = 10
		chart1.title = 'Direct infusion analysis of '+fourlettcode+' derivatized fatty acids'
		chart1.y_axis.title = 'normalized abundance'
		mr=lenfintlistn+1
		data = Reference(uncws, min_col=17, min_row=1, max_row=mr, max_col=18)
		cats = Reference(uncws, min_col=15, min_row=2, max_row=mr)
		chart1.add_data(data, titles_from_data=True)
		chart1.set_categories(cats)
		chart1.shape = 4
		colorschemebarchart=['black', 'green']
		#series=Series(cats, data, title_from_data=True)
		#series.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[1])
		testing=1
		if testing==1:
			kc=0
			while kc<lenfintlistn:
				s=chart1.series[0]   #define datapoint in column
				pt=DataPoint(idx=kc)     #define which column
				pt.graphicalProperties.solidFill=ColorChoice(prstClr=colorschemebarchart[0])
				pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[0])
				s.dPt.append(pt)
				s=chart1.series[1]   #define datapoint in column
				pt=DataPoint(idx=kc)     #define which column
				pt.graphicalProperties.solidFill=ColorChoice(prstClr=colorschemebarchart[1])
				pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[1])
				s.dPt.append(pt)
				kc=kc+1

		ws.add_chart(chart1, "B2")
		uncwb.save('ms_di_fa_integrals_pb_subtracted.xlsx')
		if gui==0:
			print('Done. Data is saved in ms_di_fa_integrals_pb_subtracted.xlsx.')
		# end plot final results (bar chart) in new sheet
	#############		# end evaluation  #######
	spectrum=spectrum+1

