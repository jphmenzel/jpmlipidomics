# -*- coding: UTF-8 -*-

# Jan Philipp Menzel, 2021 / 2022
# Goal: read csv file containing data from last skyline routine, 
# save data as xlsx file with plot and data for replicate plot
## NOTES: STAGE 5 for LIPIDOMICS WORKFLOW OzFAD1. 
##			flags overlap of precursor species
##			corrects areas for isotopic pattern using sum formula of species
##			saves data both in csv file and xlsx file


import math
import sys
import openpyxl
import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.styles.colors import RGB
from openpyxl.chart.axis import ChartLines
import requests
import urllib.request
import time
from bs4 import BeautifulSoup
beforeall=datetime.datetime.now()
################ DATABASE ## Source: Internetchemie.info
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.7, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
################
gui=1

if gui==0:
	print('This program generates the fatty acid analysis output of the OzFAD1 workflow including a preliminary barchart.')
	print('Three output files of this program of three replicates can be used to create a replicate plot by the python script OzFAD1_6_Replicate_plot.')
	print('Before proceeding, ensure that the files OzFAD1_4_input_DIA_Q.xlsx and OzFAD1_4_DIA_deconv.xlsx are in the correct directory.')

#selectiontype=eval(input('Generate Transition Results based on m/z error and retention time cutoff only (0) or based on strict selection criteria (1)? : '))
# begin determine derivatization group sum formula

if gui==0:
	fourlettcode=input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP, PLPC, PLPE): ')
	fourlettcode=str(fourlettcode)
	if fourlettcode=='AMPP':
		cderiv=12
		hderiv=12
		dderiv=0
		nderiv=2
		oderiv=0
		pderiv=0
		ideriv=0
	elif fourlettcode=='NMPA':
		cderiv=7
		hderiv=10
		dderiv=0
		nderiv=2
		oderiv=0
		pderiv=0
		ideriv=0
	elif fourlettcode=='NMPE':
		cderiv=7
		hderiv=9
		dderiv=0
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='MDPE':
		cderiv=7
		hderiv=6
		dderiv=3
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='NEPE':
		cderiv=8
		hderiv=11
		dderiv=0
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='EDPE':
		cderiv=6
		hderiv=6
		dderiv=5
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='NPPE':
		cderiv=9
		hderiv=13
		dderiv=0
		nderiv=1
		oderiv=1
		pderiv=0
		ideriv=0
	elif fourlettcode=='IAMP':
		cderiv=12
		hderiv=11
		dderiv=0
		nderiv=2
		oderiv=0
		pderiv=0
		ideriv=1
	elif fourlettcode=='PLPC':
		cderiv=8
		hderiv=20
		dderiv=0
		nderiv=1
		oderiv=6
		pderiv=1
		ideriv=0
	elif fourlettcode=='PLPE':
		cderiv=5
		hderiv=14
		dderiv=0
		nderiv=1
		oderiv=6
		pderiv=1
		ideriv=0
	else:
		cderiv=eval(input('Number of C atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		hderiv=eval(input('Number of H atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		dderiv=eval(input('Number of D atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		nderiv=eval(input('Number of N atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		oderiv=eval(input('Number of O atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		pderiv=eval(input('Number of P atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
		ideriv=eval(input('Number of I atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	# end determine derivatization group sum formula
else:
	fourlettcode=sys.argv[1]	#retrieve arguments from gui
	cderiv=int(sys.argv[2])
	hderiv=int(sys.argv[3])
	dderiv=int(sys.argv[4])
	nderiv=int(sys.argv[5])
	oderiv=int(sys.argv[6])
	pderiv=int(sys.argv[7])
	ideriv=int(sys.argv[8])


manualfilter=1 
selectiontype=1
#rettimecutoff=17.8
#precareathreshold=2000	#applies to precursor
#prodareathreshold=1000	#applies to products of loss
#abundance=[99.9885, 98.93, 99.636, 99.7, 94.93] not updated

#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists

wbdecon=openpyxl.load_workbook('OzFAD1_4_DIA_deconv.xlsx', data_only=True)
sheetnum=len(wbdecon.sheetnames)
#print(sheetnum)
#quit()
#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists
############################################################################################################################################
# begin read precursor_results
shortlistfa=[]	#list of unsaturated FA species in precursorresults
allshortlistfa=[]
shortlistarea=[]	#list of areas representing sum of product fragments associated to precursor
wb=openpyxl.load_workbook('OzFAD1_4_input_DIA_Q.xlsx')
sheet=wb['precursor_results']
toprow=['Precursor name', 'Product name', 'Fatty acid', 'RT / min', 'Precursor area', 'Sum of product areas', 'Systematic name', 'LipidMAPS ID']
c=1
while c<9:
	sheet.cell(row=1, column=c).value=toprow[c-1]
	c=c+1
#print('Writing precursor results is in progress.')
shortlistfa=[]
shortlistarea=[]
tli=1
go=1
while go==1:
	cprecname=sheet.cell(row=tli+1, column=2)
	cprecname=cprecname.value
	cprecfaname=sheet.cell(row=tli+1, column=1)
	cprecfaname=cprecfaname.value
	if cprecname is None:
		go=0
	elif cprecname=='nan':
		go=0
	else:
		cprecname=str(cprecname)
		if 'OVERLAPPING' in cprecname:
			#get fa (eg 16_1 from 16:1), search other excel sheet, get deconvoluted area and calc final area for quantification
			faspec='DC_'+str(cprecname[5])+str(cprecname[6])+'_'+str(cprecname[8])
			#print('faspec')
			#print(faspec)
			csheet=wbdecon[faspec]
			rtnum=csheet.cell(row=1, column=12)
			rtnum=rtnum.value
			rtnum=int(rtnum)
			cfaname=sheet.cell(row=tli+1, column=1)
			cfaname=cfaname.value
			cfaname=str(cfaname)
			gos=1
			gon=0
			while gos==1:
				# get current species from decon list
				cdeconn=csheet.cell(row=1, column=17+gon)
				cdeconn=cdeconn.value
				#print('cdeconn')
				#print(cdeconn)
				if cdeconn is None:
					gos=0
				elif cdeconn=='nan':
					gos=1
				else:
					cdeconn=str(cdeconn)
					if cfaname==cdeconn:
						#cgaussint=csheet.cell(row=rtnum+5, column=17+gon)
						#cgaussint=cgaussint.value
						#cgaussint=float(cgaussint) # can't access result of formula from excel this way, calc here:
						cgaussint=0
						csearch='1_('+str(cdeconn)+')'
						cspn=2
						cfind=csheet.cell(row=cspn, column=1)
						cfind=cfind.value
						#cfind=str(cfind)
						gofind=1
						while gofind==1:
							cfind=csheet.cell(row=cspn, column=1)
							cfind=cfind.value
							#cfind=str(cfind)
							if cfind is None:
								gofind=0
							elif cfind=='nan':
								gofind=0
							else:
								if csearch==str(cfind):
									crc=cspn
							cspn=cspn+1


						ccamp=csheet.cell(row=crc, column=2)
						ccamp=ccamp.value
						ccamp=float(ccamp)
						ccc=csheet.cell(row=4, column=2)
						ccc=ccc.value
						ccc=float(ccc)
						cgaussint=ccamp*ccc*(2*3.1415926)**0.5
						cgaussint=cgaussint #/0.0038		# convert from integral to value that compares to OzID areas and precursor areas
						# begin get Sum of numerically integrated OzID Product XICs
						osheet=wbdecon['Summary_OzID_Integrals']
						oi=2
						goo=1
						while goo==1:
							cfa=osheet.cell(row=oi, column=1)
							cfa=cfa.value
							if cfa is None:
								goo=0
							elif str(cfa)=='nan':
								goo=0
							else:
								cfa=str(cfa)
								if cfa==str(cdeconn):
									cprodarea=osheet.cell(row=oi, column=2)
									cprodarea=cprodarea.value
									cprodarea=float(cprodarea)
							oi=oi+1	
						
						# end get Sum of numerically integrated OzID Product XICs
						cprodarea=cprodarea*0.0038
						#print('cdeconn')
						#print(cdeconn)
						#print('cprodarea')
						#print(cprodarea)
						sheet.cell(row=1, column=9).value='Deconvoluted_Precursor_Integral'
						sheet.cell(row=1, column=10).value='Sum_OzID+Precursor_XIC_Integrals'
						sheet.cell(row=1, column=11).value='Sum_OzID_XIC_Integrals'
						sheet.cell(row=tli+1, column=11).value=cprodarea
						sheet.cell(row=tli+1, column=9).value=cgaussint
						cquant=cprodarea+cgaussint
						sheet.cell(row=tli+1, column=10).value=cquant
						#print('cgaussint')
						#print(cgaussint)
						k=5	# for AMPP k=5 
						label=str('')
						while k<(len(cprecname)):
							label=label+str(cprecname[k])
							k=k+1
						# e.g. 16:1_n-5_6.7
						shortlistfa.append(label)
						shortlistarea.append(cquant) #cgaussint)
				gon=gon+1
				


			# write results in excel lsheet precursor_results and add do shortlistfa and shortlistareas
		else:
			# deconvoluted area is precursor area
			#cprecarea=sheet.cell(row=tli+1, column=5)
			#cprecarea=cprecarea.value
			#cprecarea=float(cprecarea)

			# begin get Sum of numerically integrated OzID Product XICs
			cprodarea=0
			osheet=wbdecon['Summary_OzID_Integrals']
			oi=2
			goo=1
			while goo==1:
				cfa=osheet.cell(row=oi, column=1)
				cfa=cfa.value
				if cfa is None:
					goo=0
				elif str(cfa)=='nan':
					goo=0
				else:
					cfa=str(cfa)
					if cfa==str(cprecfaname):
						cprodarea=osheet.cell(row=oi, column=2)
						cprodarea=cprodarea.value
						cprodarea=float(cprodarea)
						cprecarea=osheet.cell(row=oi, column=3)
						cprecarea=cprecarea.value
						cprecarea=float(cprecarea)
				oi=oi+1	
			# end get Sum of numerically integrated OzID Product XICs
			cprodarea=cprodarea*0.0038
			cprecarea=cprecarea*0.0038
			#print('cprecfaname')
			#print(cprecfaname)
			#print('cprodarea')
			#print(cprodarea)
			cquant=cprodarea+cprecarea
			sheet.cell(row=1, column=11).value='Sum_OzID_XIC_Integrals'
			sheet.cell(row=tli+1, column=11).value=cprodarea
			sheet.cell(row=1, column=10).value='Sum_OzID+Precursor_XIC_Integrals'
			sheet.cell(row=tli+1, column=10).value=cquant
			k=5	# for AMPP k=5 
			label=str('')
			while k<(len(cprecname)):
				label=label+str(cprecname[k])
				k=k+1
			# e.g. 16:1_n-5_6.7
			shortlistfa.append(label)
			shortlistarea.append(cquant)
			# # write results in excel lsheet precursor_results and add do shortlistfa and shortlistareas
			# 	
	tli=tli+1
wb.save('OzFAD1_5_plot_table.xlsx')

oldref=0		######################################################## old stuff for reference
if oldref==1:
	r=2
	rw=2
	fragareasum=0
	while r<(len(mlistname)+2):
		findloss=prodname[r-2].find('loss')
		if findloss>0:
			fragareasum=fragareasum+area[r-2]
		else:
			wfragareasum=fragareasum
			if str(prodname[r-2][(len(prodname[r-2])-1)])=='r':
				fragareasum=0
		if str(prodname[r-2][(len(prodname[r-2])-1)])=='r':		#unsaturated FA, precursor
			sheet.cell(row=rw, column=1).value=precname[r-2]
			sheet.cell(row=rw, column=2).value=prodname[r-2]
			k=5	# for AMPP k=5 
			label=str('')
			while k<(len(precname[r-2])):
				label=label+str(precname[r-2][k])
				k=k+1
			sheet.cell(row=rw, column=3).value=label	#column C in excel file, e.g. 16:1_n-5_6.7
			shortlistfa.append(label)
			allshortlistfa.append(label)
			sheet.cell(row=rw, column=4).value=rettime[r-2]
			sheet.cell(row=rw, column=5).value=area[r-2]
			sheet.cell(row=rw, column=6).value=wfragareasum		#Sum of product transition areas
			shortlistarea.append(wfragareasum)
			rw=rw+1
		elif len(prodname[r-2])==9:		#saturated FA
			sheet.cell(row=rw, column=1).value=precname[r-2]
			sheet.cell(row=rw, column=2).value=prodname[r-2]
			k=5	# for AMPP k=5 
			label=str('')
			while k<(len(precname[r-2])):
				label=label+str(precname[r-2][k])
				k=k+1
			sheet.cell(row=rw, column=3).value=label
			allshortlistfa.append(label)	
			sheet.cell(row=rw, column=4).value=rettime[r-2]
			sheet.cell(row=rw, column=5).value=area[r-2]
			rw=rw+1
		r=r+1
	wb.save('jpmlipidomics_vpw20_10_selected_final.xlsx')
# end write precursor_results

###########################################################################################################################################################################
###########################################################################################################################################################################
# begin write results_summary sheet
wb=openpyxl.load_workbook('OzFAD1_5_plot_table.xlsx')
sheetname='results_summary'
ws = wb.create_sheet(sheetname)
sheetsummary=wb[sheetname]			# sheetsummary is results_summary
sheet=wb['precursor_results']		# sheet is precursor_results
#toprow=['Fatty acid', 'FA shorthand', 'Systematic name', 'LipidMAPS ID', 'Common name', 'RT / min', 'Relative isomer abundance / %', 'abundance nmol g-1']
toprow=['Fatty acid (n-x)', 'FA shorthand', 'Systematic name', 'LipidMAPS ID', 'Common Name', 'RT / min', 'Relative isomer abundance / %', 'Abundance nmol g-1']
c=1
while c<8:
	sheetsummary.cell(row=1, column=c).value=toprow[c-1]
	c=c+1
fanlist=[]	# Fatty acid species written in column C (3) in precursor_results is being read into list fanlist
rr=2
go=1
while go==1:
	fan=sheet.cell(row=rr, column=3)
	fan=fan.value
	if fan is None:
		go=0
	elif str(fan)=='nan':
		go=0
	else:
		fan=str(fan)
		fn=len(fan)-1
		gon=1
		while gon==1:
			fn=len(fan)-1
			if fan[fn]=='_':
				gon=0
				fan=fan[:-1]
			else:
				fan=fan[:-1]
		fanlist.append(fan)
	rr=rr+1
# do not change fanlist from now (containing e.g. 14:1_n-7)
# begin go through fanlist to generate fanclist, if entry only one or first of multiple, do not specify cis or trans, if not first of multiple, specify as trans (E)

# To add: check for three entries, if found, specify as branched, cis and trans, respectively!!		!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

fanclist=[]
rf=0
while rf<len(fanlist):
	rfi=0
	ffound=0
	while rfi<len(fanlist):
		if rf==rfi:
			rf=rf
		else:
			if fanlist[rf]==fanlist[rfi]:
				if rf>rfi:
					cfan=fanlist[rf]
					if int(cfan[3])>1:
						# PUFA
						fanc='C'
						ci=0
						while ci<len(cfan):
							if cfan[ci]=='_':
								ci=ci+1
							else:
								if ci>8:
									if cfan[ci]=='n':
										ci=ci+1
									elif cfan[ci]=='-':
										ci=ci+1
										fanc=fanc+','
									else:
										fanc=fanc+str(cfan[ci])
										ci=ci+1
								else:
									fanc=fanc+str(cfan[ci])
									ci=ci+1
						fanc=fanc+'_(E)'
					else:
						fanc='C'
						ci=0
						while ci<len(cfan):
							if cfan[ci]=='_':
								ci=ci+1
							else:
								fanc=fanc+str(cfan[ci])
								ci=ci+1
						fanc=fanc+'_(E)'
					ffound=1
		rfi=rfi+1
	if ffound==0:
		cfan=fanlist[rf]
		if int(cfan[3])>1:
			# PUFA
			fanc='C'
			ci=0
			while ci<len(cfan):
				if cfan[ci]=='_':
					ci=ci+1
				else:
					if ci>8:
						if cfan[ci]=='n':
							ci=ci+1
						elif cfan[ci]=='-':
							ci=ci+1
							fanc=fanc+','
						else:
							fanc=fanc+str(cfan[ci])
							ci=ci+1
					else:
						fanc=fanc+str(cfan[ci])
						ci=ci+1
		else:
			fanc='C'
			ci=0
			while ci<len(cfan):
				if cfan[ci]=='_':
					ci=ci+1
				else:
					fanc=fanc+str(cfan[ci])
					ci=ci+1
	fanclist.append(fanc)
	rf=rf+1
#print('fanclist:')
#print(fanclist)
lfancl=len(fanclist)
#print(lfancl)
# end go through fanlist to generate fanclist, if entry only one or first of multiple, do not specify cis or trans, if not first of multiple, specify as trans (E)
fashorthand=[]
fasysname=[]
lipidmapsid=[]
commonnamelist=[]
rtlist=[]	# RT / min
relab=[]	# relative isomer abundance
rf=0
while rf<len(fanclist):
	# begin get RT 
	crt=sheet.cell(row=rf+2, column=4)
	crt=crt.value
	crt=str(crt)
	rtlist.append(crt)
	# end get RT
	cfas='' #'FA ' # fatty acid shorthand
	cfas=cfas+str(fanclist[rf][1])+str(fanclist[rf][2])+str(fanclist[rf][3])+str(fanclist[rf][4])+'('
	if int(fanclist[rf][4])==1:
		#get delta position of MUFA
		fasys=sheet.cell(row=rf+2, column=7)
		fasys=fasys.value
		fasys=str(fasys)
		cy=0
		go=1
		while go==1:
			if str(fasys[cy])=='Z':
				# begin change systematic name to trans MUFA
				cyc=0
				efasys=str()
				while cyc<cy:
					efasys=efasys+str(fasys[cyc])
					cyc=cyc+1
				if str(fanclist[rf][len(fanclist[rf])-2])=='E':
					efasys=efasys+'E'
				else:
					efasys=efasys+'Z'
				cyc=cy+1
				while cyc<len(fasys):
					efasys=efasys+str(fasys[cyc])
					cyc=cyc+1
				fasys=efasys	# end change systematic name to trans MUFA
			elif str(fasys[cy])=='-':
				go=0
			else:
				cfas=cfas+str(fasys[cy])
			cy=cy+1
		if str(fanclist[rf][len(fanclist[rf])-2])=='E':
			cfas=cfas+'E)'
		else:
			cfas=cfas+'Z)'
		fashorthand.append(cfas)
		fasysname.append(fasys)
	elif int(fanclist[rf][4])==0:
		go=go
		fasys=sheet.cell(row=rf+2, column=7)
		fasys=fasys.value
		fasys=str(fasys)
		cfas=str(fasys)
		fashorthand.append(cfas)
		fasysname.append(fasys)	
	else:
		#get delta position of PUFA
		fasys=sheet.cell(row=rf+2, column=7)
		fasys=fasys.value
		fasys=str(fasys)
		#print('fasys')
		#print(fasys)
		cy=0
		go=1
		while go==1:
			if str(fasys[cy])=='Z':
				# begin change systematic name to trans PUFA, by default, if trans PUFA, highest delta position is labelled E
				cyc=0
				efasys=str()
				while cyc<cy:
					efasys=efasys+str(fasys[cyc])
					cyc=cyc+1
				if str(fanclist[rf][len(fanclist[rf])-2])=='E':
					if str(fasys[cy+1])=='-':
						efasys=efasys+'E'
					else:
						efasys=efasys+'Z'
				else:
					efasys=efasys+'Z'
				cyc=cy+1
				while cyc<len(fasys):
					efasys=efasys+str(fasys[cyc])
					cyc=cyc+1
				fasys=efasys	# end change systematic name to trans MUFA
			elif str(fasys[cy])=='-':
				go=0
			elif fasys=='hexadecanoic acid':
				go=0
			else:
				cfas=cfas+str(fasys[cy])
			cy=cy+1
		if str(fanclist[rf][len(fanclist[rf])-2])=='E':
			cfas=cfas+')'
		else:
			cfas=cfas+')'
		fashorthand.append(cfas)
		fasysname.append(fasys)	
	rf=rf+1


# write results in summary sheet
fw=0
while fw<len(fanclist):
	sheetsummary.cell(row=fw+2, column=1).value=fanclist[fw]
	sheetsummary.cell(row=fw+2, column=2).value=fashorthand[fw]
	sheetsummary.cell(row=fw+2, column=3).value=fasysname[fw]
	#sheetsummary.cell(row=fw+2, column=4).value=lipidmapsid[fw]	#written later
	#sheetsummary.cell(row=fw+2, column=5).value=commonname[fw]		#not functional
	sheetsummary.cell(row=fw+2, column=6).value=rtlist[fw]
	#sheetsummary.cell(row=fw+2, column=7).value=relab[fw]			#written later
	fw=fw+1

wb.save('OzFAD1_5_plot_table.xlsx')

# end write results_summary sheet
###########################################################################################################################################################################
###########################################################################################################################################################################
############################################################################################################################################
# begin write final_barchart
#print('shortlistfa:')
#print(shortlistfa)
lslfa=len(shortlistfa)
#print(lslfa)
#print(shortlistarea)
wb=openpyxl.load_workbook('OzFAD1_5_plot_table.xlsx')
#sheetname='final_barchart'
#ws = wb.create_sheet(sheetname)
sheetbc=wb['final_barchart']
#sheetprecr=wb['precursorresults']
toprow=['FA', 'n-2 (Me, Z)', 'n-2 (Me, E)', 'n-2 (Bu)', 'n-2 (NMI)', 'n-2 (Branched)', 
'n-3 (Me, Z)', 'n-3 (Me, E)', 'n-3 (Bu)', 'n-3 (NMI)', 'n-3 (Branched)',
'n-4 (Me, Z)', 'n-4 (Me, E)', 'n-4 (Bu)', 'n-4 (NMI)', 'n-4 (Branched)',
'n-5 (Me, Z)', 'n-5 (Me, E)', 'n-5 (Bu)', 'n-5 (NMI)', 'n-5 (Branched)',
'n-6 (Me, Z)', 'n-6 (Me, E)', 'n-6 (Bu)', 'n-6 (NMI)', 'n-6 (Branched)',
'n-7 (Me, Z)', 'n-7 (Me, E)', 'n-7 (Bu)', 'n-7 (NMI)', 'n-7 (Branched)',
'n-8 (Me, Z)', 'n-8 (Me, E)', 'n-8 (Bu)', 'n-8 (NMI)', 'n-8 (Branched)',
'n-9 (Me, Z)', 'n-9 (Me, E)', 'n-9 (Bu)', 'n-9 (NMI)', 'n-9 (Branched)',
'n-10 (Me, Z)', 'n-10 (Me, E)', 'n-10 (Bu)', 'n-10 (NMI)', 'n-10 (Branched)',
'n-11 (Me, Z)', 'n-11 (Me, E)', 'n-11 (Bu)', 'n-11 (NMI)', 'n-11 (Branched)',
'n-12 (Me, Z)', 'n-12 (Me, E)', 'n-12 (Bu)', 'n-12 (NMI)', 'n-12 (Branched)',
'n-13 (Me, Z)', 'n-13 (Me, E)', 'n-13 (Bu)', 'n-13 (NMI)', 'n-13 (Branched)',
'n-14 (Me, Z)', 'n-14 (Me, E)', 'n-14 (Bu)', 'n-14 (NMI)', 'n-14 (Branched)',
'n-15 (Me, Z)', 'n-15 (Me, E)', 'n-15 (Bu)', 'n-15 (NMI)', 'n-15 (Branched)',
'n-16 (Me, Z)', 'n-16 (Me, E)', 'n-16 (Bu)', 'n-16 (NMI)', 'n-16 (Branched)']

# 'n-4 (Me)', 'n-4 (Bu)', 'n-4 (Other)', 'n-5 (Me)', 'n-5 (Bu)', 'n-5 (Other)', 'n-6 (Me)', 'n-6 (Bu)', 'n-6 (Other)',
#'n-7 (Me)', 'n-7 (Bu)', 'n-7 (Other)', 'n-8 (Me)', 'n-8 (Bu)', 'n-8 (Other)', 'n-9 (Me)', 'n-9 (Bu)', 'n-9 (Other)', 'n-10 (Me)', 'n-10 (Bu)', 'n-10 (Other)', 'n-11 (Me)', 
#'n-11 (Bu)', 'n-11 (Other)', 'n-12 (Me)', 'n-12 (Bu)', 'n-12 (Other)', 'n-13 (Me)', 'n-13 (Bu)', 'n-13 (Other)', 'n-14 (Me)', 'n-14 (Bu)', 'n-14 (Other)', 'n-15 (Me)', 'n-15 (Bu)', 
#'n-15 (Other)']
ebclist=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
ibclist=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
c=1
while c<(len(toprow)+1):
	sheetbc.cell(row=1, column=c).value=toprow[c-1]
	c=c+1
barchartlist=[]
#ibclist=ebclist
if shortlistfa[0]=='16:0d31_precursor':
	i=1
else:
	i=0

if len(shortlistfa)>0:
	if shortlistfa[i][8]=='_':		#determine n position (dbn) of first FA
		dbn=int(shortlistfa[i][7])
	elif shortlistfa[i][9]=='_':
		dbn=10*int(shortlistfa[i][7])+int(shortlistfa[i][8])
	dbn=dbn-1
	ibci=(5*dbn)-4
	ibclist[ibci]=shortlistarea[i]
	cfa=str(shortlistfa[i][0])+str(shortlistfa[i][1])+str(shortlistfa[i][2])+str(shortlistfa[i][3])
	ibclist[0]=cfa
	barchartlist.append(ibclist)
	#print(barchartlist)
	#quit()
	satfaonly=0
else:
	dbn=0
	satfaonly=1

if shortlistfa[0]=='16:0d31_precursor':
	i=1
else:
	i=0
while i<(len(shortlistfa)):
	cdbs=[]	#determine if species Me or Bu spaced or monounsat
	use=0
	if int(shortlistfa[i][3])==1:
		use=1
	elif int(shortlistfa[i][3])==0:
		use=0
	else:
		fi=4		#determine if polyunsaturated FA Me/Bu spaced
		while fi<(len(shortlistfa[i])):
			if str(shortlistfa[i][fi])=='n':
				if str(shortlistfa[i][fi+3])=='_':
					cdb=int(shortlistfa[i][fi+2])
					cdbs.append(cdb)
				elif str(shortlistfa[i][fi+4])=='_':
					cdb=10*int(shortlistfa[i][fi+2])+int(shortlistfa[i][fi+3])
					cdb=int(cdb)
					cdbs.append(cdb)
			fi=fi+1
		dbi=0		# e.g. cdbs=[6, 9, 15]
		check=1
		while dbi<(len(cdbs)-1):
			if int(abs((cdbs[dbi])-(cdbs[dbi+1])) % 3)==0:	# Me or Bu or Hept spaced
				check=check
			else:
				check=0
			dbi=dbi+1
		if check==1:
			use=1
			cat=0
	if use==1:
		cat=0
		cfa=str(shortlistfa[i][0])+str(shortlistfa[i][1])+str(shortlistfa[i][2])+str(shortlistfa[i][3])
		if (len(cdbs))>1:
			if (cdbs[1]-cdbs[0])==6:
				cat=2 # Bu
			elif (cdbs[1]-cdbs[0])==9:
				cat=3 # Other
			elif (cdbs[1]-cdbs[0])==3:
				if (len(cdbs))==3:
					if (cdbs[2]-cdbs[1])==3:
						cat=0
					else:
						cat=3
				else:
					cat=0					
			else:
				cat=0	
	else:
		cfa=str(shortlistfa[i][0])+str(shortlistfa[i][1])+str(shortlistfa[i][2])+str(shortlistfa[i][3])
		cat=3 #dbn=dbn+26
	if cat==0:
		if str(fanclist[i][len(fanclist[i])-2])=='E':
			cat=1		# Me (E)
	if cat==3:
		if str(fanclist[i][len(fanclist[i])-2])=='E':
			cat=4		# NMI (E)		
	#use=1
	if use==1:
		#print(cat)
		#print(cfa)
		#print(shortlistfa[i])
		#print(barchartlist)
		bcli=0
		found=0
		while bcli<(len(barchartlist)):
			if barchartlist[bcli][0]==cfa:
				if shortlistfa[i][8]=='_':
					dbn=int(shortlistfa[i][7])
				else:
					dbn=10*int(shortlistfa[i][7])+int(shortlistfa[i][8])
					dbn=int(dbn)
				#print(dbn)
				dbn=dbn-1
				if cat==0:
					dbn=(5*dbn)-4
				elif cat==1:
					dbn=(5*dbn)-3
				elif cat==2:
					dbn=5*dbn-2
				elif cat==3:
					dbn=(5*dbn)-1
				elif cat==4:
					dbn=5*dbn
				barchartlist[bcli][dbn]=float(barchartlist[bcli][dbn])+float(shortlistarea[i])		#
				found=1
			else:
				found=found
			bcli=bcli+1
		if found==0:
			ibclist=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
			ibclist[0]=cfa
			if shortlistfa[i][8]=='_':
				dbn=int(shortlistfa[i][7])
			elif shortlistfa[i][9]=='_':
				dbn=10*int(shortlistfa[i][7])+int(shortlistfa[i][8])
			dbn=dbn-1
			if cat==0:
				dbn=(5*dbn)-4
			elif cat==1:
				dbn=(5*dbn)-3
			elif cat==2:
				dbn=5*dbn-2
			elif cat==3:
				dbn=(5*dbn)-1
			elif cat==4:
				dbn=5*dbn
			ibci=dbn
			ibclist[ibci]=shortlistarea[i]
			barchartlist.append(ibclist)
			#print(barchartlist)
	i=i+1
#convert results to percentages
#print(barchartlist)
pbarchartlist=[]		#results for barchart in percentages
bcl=0
while bcl<(len(barchartlist)):
	cfal=1
	csum=0
	while cfal<(len(barchartlist[bcl])):
		csum=csum+barchartlist[bcl][cfal]
		cfal=cfal+1
	pbcl=[]
	pbcl.append(barchartlist[bcl][0])
	cpi=1
	while cpi<(len(barchartlist[bcl])):
		cp=(barchartlist[bcl][cpi]/csum)*100
		pbcl.append(cp)
		cpi=cpi+1
	pbarchartlist.append(pbcl)
	bcl=bcl+1
#print(pbarchartlist)

## begin sort pbarchartlist by odd and even FA
spbarchartlist=[]	#sorted pbarchartlist
oddlist=[1,3,5,7,9]
evenlist=[0,2,4,6,8]
sp=0
while sp<(len(pbarchartlist)):
	if int(pbarchartlist[sp][0][1]) in oddlist:
		spbarchartlist.append(pbarchartlist[sp])
	sp=sp+1
sp=0
while sp<(len(pbarchartlist)):
	if int(pbarchartlist[sp][0][1]) in evenlist:
		spbarchartlist.append(pbarchartlist[sp])
	sp=sp+1
## end sort pbarchartlist by odd and even FA
#print(spbarchartlist)

#write results in pbarchartlist in excel file
r=2
while r<(len(spbarchartlist)+2):
	c=1
	while c<(len(spbarchartlist[r-2])+1):
		sheetbc.cell(row=r, column=c).value=spbarchartlist[r-2][c-1]
		c=c+1
	r=r+1
#begin create bar chart in excel sheet
assigned=[]
assigned.append(toprow)
#aslist=['index', 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2] # 0 is Me; 1 is Bu; 2 is Other
aslist=['index', 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4] # 0 is Me; 1 is Bu; 2 is Other
ias=0
while ias<(len(spbarchartlist)):
	assigned.append(aslist)
	ias=ias+1

mr=len(spbarchartlist)+1
if satfaonly==1:
	print('Only saturated fatty acids were found, results are saved in jpm_lipidomics_vpw11_5_final_output.xlsx')
	quit()
else:
	mc=len(spbarchartlist[0])
chart1 = BarChart()
chart1.type = "col"
chart1.style = 12
chart1.grouping = "stacked"
chart1.overlap = 100
chart1.title = "Fatty acids: Methylene and butylene interrupted and unusual double bond positions"
chart1.y_axis.title = 'Percentage of FA'
#chart1.x_axis.title = 'FA'
chart1.y_axis.scaling.max = 100   
data = Reference(sheetbc, min_col=2, min_row=1, max_row=mr, max_col=mc)
cats = Reference(sheetbc, min_col=1, min_row=2, max_row=mr)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4

gridoff=0
if gridoff==1:
	# begin turn majorGridlines off (setting colour white: FFFFFF)
	chart1.y_axis.majorGridlines.spPr = GraphicalProperties(noFill = 'True')
	chart1.y_axis.majorGridlines.spPr.ln = LineProperties(solidFill = 'FFFFFF')
	chart1.x_axis.majorGridlines = ChartLines()
	chart1.x_axis.majorGridlines.spPr = GraphicalProperties(noFill = 'True')
	chart1.x_axis.majorGridlines.spPr.ln = LineProperties(solidFill = 'FFFFFF')
	# end turn majorGridlines off (setting colour white: FFFFFF)

#colours=["0066cc", "660066", "003366", "ccffff", "800080", "00ccff", "993300", "99cc00", "000080", "ff00ff", "ff0000", "808000"] 
#colours=["4f81bd", "7f7f7f", "ffc000", "c4bd97", "8064a2", "4bacc6", "c0504d", "9bbb59", "1f497d", "f05eb5", "ff0000", "808000", "003300", 
#"0066cc", "660066", "003366", "ccffff", "800080", "00ccff", "993300", "99cc00", "000080", "ff00ff", "ff0000", "707000", "002200", 
#"0055cc", "550066", "003355", "00ffff", "700070", "00cc00", "883300", "88cc00", "000070", "0000ff", "ff0011", "606000", "001100"]
#cv=0
#while cv<len(colours):
#	s = chart1.series[cv]
#	s.graphicalProperties.line.solidFill = colours[cv]
#	s.graphicalProperties.solidFill = colours[cv]
#	cv=cv+1

#begin test new barchart with patterns
colors=['black', 'cornflowerBlue', 'lightGray', 'magenta', 'gold', 'mediumPurple', 'deepSkyBlue', 'sienna', 'limeGreen', 'yellow', 'orange', 'red', 'dkOliveGreen', 'cyan', 'blue']
colorschemebarchart=[]
cl=0
while cl<(len(colors)):
	csbc=str(colors[cl])
	colorschemebarchart.append(csbc)
	colorschemebarchart.append(csbc)
	colorschemebarchart.append(csbc)
	colorschemebarchart.append(csbc)
	colorschemebarchart.append(csbc)
	cl=cl+1

stbr=0
while stbr<(len(assigned[0])-1):
    clm=0
    while clm<(len(assigned)-1):      # 6 is number of columns in barchart (16:1, 17:1, ...)
        if assigned[clm+1][stbr+1]==0:
            s=chart1.series[stbr]   #define datapoint in column ([0] is n-2; [1] is n-3 ...)
            pt=DataPoint(idx=clm)     #define which column (e.g. idx=0 is 16:1; idx=1 is 17:1 ...)
            pt.graphicalProperties.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            s.dPt.append(pt)
        elif assigned[clm+1][stbr+1]==1:
            s=chart1.series[stbr]   #define datapoint in column ([0] is n-2; [1] is n-3 ...)
            pt=DataPoint(idx=clm)     #define which column (e.g. idx=0 is 16:1; idx=1 is 17:1 ...)
            fill=PatternFillProperties(prst='dkDnDiag')
            fill.foreground=ColorChoice(prstClr='black')
            fill.background=ColorChoice(prstClr=colorschemebarchart[stbr])
            pt.graphicalProperties.pattFill=fill
            pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            s.dPt.append(pt)
        elif assigned[clm+1][stbr+1]==2:
            s=chart1.series[stbr]   #define datapoint in column ([0] is n-2; [1] is n-3 ...)
            pt=DataPoint(idx=clm)     #define which column (e.g. idx=0 is 16:1; idx=1 is 17:1 ...)
            fill=PatternFillProperties(prst='dkUpDiag')
            fill.foreground=ColorChoice(prstClr='black')
            fill.background=ColorChoice(prstClr=colorschemebarchart[stbr])
            pt.graphicalProperties.pattFill=fill
            pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            s.dPt.append(pt)
        elif assigned[clm+1][stbr+1]==3:
            s=chart1.series[stbr]   #define datapoint in column ([0] is n-2; [1] is n-3 ...)
            pt=DataPoint(idx=clm)     #define which column (e.g. idx=0 is 16:1; idx=1 is 17:1 ...)
            fill=PatternFillProperties(prst='dkVert')
            fill.foreground=ColorChoice(prstClr='black')
            fill.background=ColorChoice(prstClr=colorschemebarchart[stbr])
            pt.graphicalProperties.pattFill=fill
            pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            s.dPt.append(pt)
        else:
            s=chart1.series[stbr]   #define datapoint in column ([0] is n-2; [1] is n-3 ...)
            pt=DataPoint(idx=clm)     #define which column (e.g. idx=0 is 16:1; idx=1 is 17:1 ...)
            fill=PatternFillProperties(prst='smCheck')
            fill.foreground=ColorChoice(prstClr='black')
            fill.background=ColorChoice(prstClr=colorschemebarchart[stbr])
            pt.graphicalProperties.pattFill=fill
            pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            s.dPt.append(pt)
        clm=clm+1
    stbr=stbr+1
chart1.legend=None
#end test new barchart with patterns
chartposition=str('A')+str(len(spbarchartlist)+3)
sheetbc.add_chart(chart1, chartposition)
#end create bar chart in excel sheet

#begin create chart with legend
legendrows = [] 
lralist=['Categories', 'n-2', 'n-3', 'n-4', 'n-5', 'n-6', 'n-7', 'n-8', 'n-9', 'n-10', 'n-11', 'n-12', 'n-13', 'n-14', 'n-15', 'n-16', 'cis (Z)', 'trans (E)', 'NMI (Bu)', 'NMI', 'Branched']
legendrows.append(lralist)
lrlist=['FA', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
i=0
while i<(len(spbarchartlist)):
	fa=spbarchartlist[i][0]
	#lrlist[0]=fa
	lrblist=[]
	lrblist.append(fa)
	k=0
	while k<(len(lrlist)-1):
		lrblist.append(0)
		k=k+1
	legendrows.append(lrblist)
	i=i+1

r=70+len(spbarchartlist)
rinit=r
while r<(len(legendrows)+rinit):
	c=1
	while c<(len(legendrows[r-rinit])+1):
		sheetbc.cell(row=r, column=c).value=legendrows[r-rinit][c-1]
		c=c+1
	r=r+1
minr=rinit #+1
mr=r-1
mc=c-1
chart2 = BarChart()
chart2.type = "col"
chart2.style = 12
chart2.grouping = "stacked"
chart2.overlap = 100
chart2.title = chart1.title
chart2.y_axis.title = chart1.y_axis.title
#chart1.x_axis.title = 'Percentage of FA'
chart2.y_axis.scaling.max = 100   #############
data = Reference(sheetbc, min_col=2, min_row=minr, max_row=mr, max_col=mc)
cats = Reference(sheetbc, min_col=1, min_row=minr+1, max_row=mr)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.shape = 4

#colorscheme=['cornflowerBlue', 'lightGray', 'magenta', 'gold', 'mediumPurple', 'deepSkyBlue', 'sienna', 'limeGreen', 'lightYellow', 'orange', 'red', 'dkOliveGreen', 'ltCyan', 'white']
cl=0
while cl<15:
    sa = chart2.series[cl]
    fill=PatternFillProperties(prst='dkUpDiag')
    fill.foreground=ColorChoice(prstClr=colors[cl])
    fill.background=ColorChoice(prstClr=colors[cl])
    sa.graphicalProperties.pattFill=fill
    sa.graphicalProperties.line.solidFill=ColorChoice(prstClr=colors[cl])
    cl=cl+1

sa = chart2.series[15]							# legend label Me (Z)
fill=PatternFillProperties(prst='dkUpDiag')
fill.foreground=ColorChoice(prstClr='white')
fill.background=ColorChoice(prstClr='white')
sa.graphicalProperties.pattFill=fill
sa.graphicalProperties.line.solidFill=ColorChoice(prstClr='black')

sa = chart2.series[16]							# legend label Me (E)
fill=PatternFillProperties(prst='dkDnDiag')
fill.foreground=ColorChoice(prstClr='black')
fill.background=ColorChoice(prstClr='white')
sa.graphicalProperties.pattFill=fill
sa.graphicalProperties.line.solidFill=ColorChoice(prstClr='white')

sa = chart2.series[17]							# legend label Bu (Z)
fill=PatternFillProperties(prst='dkUpDiag')
fill.foreground=ColorChoice(prstClr='black')
fill.background=ColorChoice(prstClr='white')
sa.graphicalProperties.pattFill=fill
sa.graphicalProperties.line.solidFill=ColorChoice(prstClr='white')

sa = chart2.series[18]							# legend label NMI (Z)
fill=PatternFillProperties(prst='dkVert')
fill.foreground=ColorChoice(prstClr='black')
fill.background=ColorChoice(prstClr='white')
sa.graphicalProperties.pattFill=fill
sa.graphicalProperties.line.solidFill=ColorChoice(prstClr='white')

sa = chart2.series[19]							# legend label NMI (E)
fill=PatternFillProperties(prst='smCheck')			# smGrid is horizontal and vertical lines as grid, smCheck is chessboard
fill.foreground=ColorChoice(prstClr='black')
fill.background=ColorChoice(prstClr='white')
sa.graphicalProperties.pattFill=fill
sa.graphicalProperties.line.solidFill=ColorChoice(prstClr='white')

chartposition=str('A')+str(35+len(spbarchartlist))
sheetbc.add_chart(chart2, chartposition)
#end create chart with legend
after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_10_'+fourlettcode+'_'
filename='OzFAD1_5_plot_table.xlsx'	# today+
wb.save(filename)
####################################


# SWITCH for troubleshooting to terminate before LIPIDMAPS LOOKUP
#quit()
#################################################################


## begin write remaining results summary
wb=openpyxl.load_workbook('OzFAD1_5_plot_table.xlsx')
sheetbc=wb['final_barchart']
sheetsummary=wb['results_summary']
ozidbc=wb['ozid_barchart']
#print(spbarchartlist)
#quit()

rf=0
while rf<len(fanclist):
	# begin get relative isomer abundance
	cfaa=str()
	cfaa=cfaa+str(fanclist[rf][1])+str(fanclist[rf][2])+str(fanclist[rf][3])+str(fanclist[rf][4])	#e.g. 14:1
	pi=0
	while pi<len(spbarchartlist):
		if cfaa==spbarchartlist[pi][0]:
			if int(fanclist[rf][4])==1:
				#MUFA
				efa=0
				if len(fanclist[rf])==8:
					cfaan=int(fanclist[rf][7])	# db position (Z FA)
				elif len(fanclist[rf])==9:
					cfaan=10*int(fanclist[rf][7])+int(fanclist[rf][8])	# db position (Z FA)
				elif len(fanclist[rf])==12:
					cfaan=int(fanclist[rf][7])	# db position (E FA)
					efa=1
				elif len(fanclist[rf])==13:
					cfaan=10*int(fanclist[rf][7])+int(fanclist[rf][8])	# db position (E FA)
					efa=1
				if efa==1:
					pipos=2+(5*(cfaan-2))	# E FA
				else:
					pipos=1+(5*(cfaan-2))	# Z FA
				crelab=float(spbarchartlist[pi][pipos])
				if crelab==100:
					crelab=round(crelab,0)
				elif crelab>10:
					crelab=round(crelab,0)
				elif crelab>1:
					crelab=round(crelab,1)
				elif crelab>0.1:
					crelab=round(crelab,2)
				else:
					crelab=round(crelab,3)
				relab.append(crelab)
			else:
				#PUFA
				# get first db pos
				if fanclist[rf][8]==',':
					cfaan=int(fanclist[rf][7])	# db position
					ii=9
				elif fanclist[rf][9]==',':
					cfaan=10*int(fanclist[rf][7])+int(fanclist[rf][8])	# db position
					ii=10
				pipos=1+(5*(cfaan-2))
				#check db pattern, Me, Bu or other
				dbpatterna=0
				dbpatternb=0
				dbpatternc=0
				dbpatternd=0
				dbpatterne=0
				while ii<len(fanclist[rf]):
					if ii==len(fanclist[rf])-1:
						ti=1
					elif ii==len(fanclist[rf])-2:
						ti=2
					else:
						if str(fanclist[rf][ii+1])==',':
							ti=1
						elif str(fanclist[rf][ii+2])==',':
							ti=2
						elif str(fanclist[rf][ii+1])=='_':
							ti=1
						elif str(fanclist[rf][ii+2])=='_':
							ti=2
					if ti==1:
						if int(fanclist[rf][ii])-cfaan==3:
							dbpatterna=1
						elif int(fanclist[rf][ii])-cfaan==6:
							dbpatternb=1
						else:
							dbpatternc=1
						cfaan=int(fanclist[rf][ii])
						ii=ii+2	
						if ii<len(fanclist[rf]):
							if str(fanclist[rf][ii])=='(':
								ii=ii+3
					elif ti==2:
						if (10*int(fanclist[rf][ii])+int(fanclist[rf][ii+1]))-cfaan==3:
							if str(fanclist[rf][len(fanclist[rf])-2])=='E':
								dbpatternb=1
							else:
								dbpatterna=1
						elif (10*int(fanclist[rf][ii])+int(fanclist[rf][ii+1]))-cfaan==6:
							dbpatternc=1
						else:
							if str(fanclist[rf][len(fanclist[rf])-2])=='E':
								dbpatterne=1
							else:	
								dbpatternd=1
						cfaan=(10*int(fanclist[rf][ii])+int(fanclist[rf][ii+1]))
						ii=ii+3
						if ii<len(fanclist[rf]):
							if str(fanclist[rf][ii])=='(':
								ii=ii+3
				if dbpatternd==1:
					pipos=pipos+3
				elif dbpatterne==1:
					pipos=pipos+4
				elif dbpatternc==1:
					if dbpatterna==0:
						pipos=pipos+2
					else:
						pipos=pipos+3
				elif dbpatterna==1:
					pipos=pipos
				elif dbpatternb==1:
					pipos=pipos+1
				crelab=float(pbarchartlist[pi][pipos])
				if crelab==100:
					crelab=round(crelab,0)
				elif crelab>10:
					crelab=round(crelab,0)
				elif crelab>1:
					crelab=round(crelab,1)
				elif crelab>0.1:
					crelab=round(crelab,2)
				else:
					crelab=round(crelab,3)
				relab.append(crelab)
		pi=pi+1
	if str(fanclist[rf][4])==0:
		relab.append('N/A')
	# end get relative isomer abundance

	rf=rf+1


wbld=openpyxl.load_workbook('LIPID_MAPS_local_data.xlsx')
wsld=wbld['local_data']
addlocal=[]

#begin get lipidmaps ID and common name
#fasysname contains updated systematic names
fullsearch=len(fasysname)
lipidmapsidlist=[]
linklist=[]
fsi=0
while fsi<fullsearch:
	searchfor=str(fasysname[fsi])

	# begin check excel file LIPID_MAPS_local_data.xlsx for IDs of found species # to speed up Database lookup and only search for species not saved locally
	# later build excel file LIPID_MAPS_local_data.xlsx, if ID found that is not saved locally
	localid=0
	rd=2
	gol=1
	while gol==1:
		dl=wsld.cell(row=rd, column=1)
		dl=dl.value
		#dl=str(dl)
		if dl is None:
			gol=0
		elif dl=='nan':
			gol=0
		elif 'acid' in dl:
			if searchfor==dl:
				localid=1
				# get data saved locally for this species
				dls=wsld.cell(row=rd, column=2)
				dls=dls.value
				dls=str(dls)
				lipidmapsidlist.append(dls)
				dlc=wsld.cell(row=rd, column=3)
				dlc=dlc.value
				dlc=str(dlc)
				commonnamelist.append(dlc)
				dll=wsld.cell(row=rd, column=4)
				dll=dll.value
				dll=str(dll)
				linklist.append(dll)
				gol=0
		else:
			gol=0
		rd=rd+1

	# end check excel file LIPID_MAPS_local_data.xlsx for IDs of found species # to speed up Database lookup and only search for species not saved locally
	

	if localid==0:
		# begin extract LipidMAPS ID
		#searchfor='6Z,9Z-octadecadienoic+acid'
		#print(searchfor)
		#searchfor='6Z,9Z-tetradecadienoic+acid'
		#searchfor='8Z,11Z-eicosadienoic+acid'
		urlpart='https://www.lipidmaps.org/search/quicksearch.php?Name='
		url=urlpart+searchfor
		#url='https://www.lipidmaps.org/search/quicksearch.php?Name=9Z-octadecenoic+acid'
		response = requests.get(url)
		#print(response)
		soup = BeautifulSoup(response.text, 'html.parser')
		results = soup.find_all('a')
		#print(len(results))
		k=0
		idfound=0
		cnidfound=0
		while k<len(results):
			one_a_tag = soup.findAll('a')[k]
			link = one_a_tag['href']
			#print(link)
			slink=str(link)
			finder=slink.find('LMID')
			if finder>0:
				idfound=1
				#print(link)
				#print(finder)
				extractid=str()
				eid=finder
				go=0
				while eid<(len(link)-1):
					if link[eid]=='=':
						go=1
					if go==1:	
						extractid=extractid+str(link[eid+1])
					eid=eid+1			
			k=k+1
		
		if idfound==1:
			print('LipidMAPS ID is:')
			print(extractid)
			lipidmapsidlist.append(extractid)
			clink='https://www.lipidmaps.org/databases/lmsd/'
			clink=clink+extractid+'?LMID='+extractid
			linklist.append(clink)

			# begin get common name and append to list commonname
			commonname=str()
			#cnfound=0
			#print('LipidMAPS ID is:')
			#print(extractid)
			#begin retrieve common name
			urlcfa='https://www.lipidmaps.org/databases/lmsd/'+str(extractid)+'?LMID='+str(extractid)
			#print(urlcfa)
			responsecfa = requests.get(urlcfa)
			#print(responsecfa.content)
			soup = BeautifulSoup(responsecfa.text, 'html.parser')
			results = soup.find_all('div')
			#print(results)
			#soup=BeautifulSoup(responsecfa.content,'html5lib')
			#print(soup.prettify())
			k=0
			idfound=0
			while k<len(results):
				divtag = soup.findAll('div')[k]
				#print(divtag)
				#link = divtag['class']
				#print(link)
				foundcn=0
				slink=str(divtag)
				finder=slink.find('Common Name')
				if finder>0:
					idfound=1
					#print(link)
					#print('-------')
					#print(finder)
					#print(k)
					#print('Common Name found !!!')
					#print(divtag)
					#print('-------')
					sdivtag=str(divtag)

					cnextractid=str()
					cni=0
					cnk=0
					while cni<len(sdivtag):
						cnj=cni
						cmn=str()
						while cnj<cni+11:
							cmn=cmn+str(sdivtag[cnj])
							cnj=cnj+1
						if cmn=='Common Name':
							#print('---####----')
							cnk=cni
							cni=len(sdivtag)
						cni=cni+1
					cna=cnk
					while cnk<len(sdivtag)-11:
						cnj=cnk
						cmn=str()
						while cnj<cnk+11:
							cmn=cmn+str(sdivtag[cnj])
							cnj=cnj+1
						if cmn=='3 lg:mt-0">':
							#print('*************************')
							foundcn=1
							cna=cnk+17
							cnk=len(sdivtag)
						cnk=cnk+1
					cnr=cna-1
					gor=1
					while gor==1:
						if cnr<len(sdivtag):
							if str(sdivtag[cnr])==' ':
								if str(sdivtag[cnr+1])==' ':
									gor=0
								else:
									cnextractid=cnextractid+str(sdivtag[cnr])
							elif str(sdivtag[cnr])=='<':
								gor=0
							elif cnr>cnk+100:
								gor=0
							else:
								cnextractid=cnextractid+str(sdivtag[cnr])
						else:
							gor=0
						cnr=cnr+1
					cnextractid=cnextractid[:-1]
					#print('##################')
					#print(cnextractid)
					if foundcn==1:
						commonname=cnextractid
				k=k+1
			commonnamelist.append(commonname)
			dlocal=[]
			dlocal.append(str(fasysname[fsi]))
			dlocal.append(extractid)
			dlocal.append(commonname)
			dlocal.append(clink)
			addlocal.append(dlocal)
			# end get common name and append to commonnamelist
		else:
			print('Fatty acid not found in LIPID MAPS:')
			print(fasysname[fsi])
			lipidmapsidlist.append('Not found in LIPID MAPS.')
			linklist.append('_')
			commonnamelist.append('_')
		# end extract LipidMAPS ID
		time.sleep(0.0005)		# waiting period in between LipidMAPS database lookup to prevent server overload (previously set to 0.05 or 0.0005 without problems)
	fsi=fsi+1
#end get lipidmaps ID and common name

# begin save lipidmaps IDs etc that are now found but are not yet in the local database
#print('checkpoint1')
dw=2
gol=1
while gol==1:
	dadd=0
	dl=wsld.cell(row=dw, column=1)	
	dl=dl.value
	#dl=str(dl)
	if dl is None:
		dadd=1
	elif dl=='nan':
		dadd=1
	else:
		dadd=0
	if dadd==1:
		dww=0
		while dww<(len(addlocal)):
			wsld.cell(row=dw, column=1).value=addlocal[dww][0]
			wsld.cell(row=dw, column=2).value=addlocal[dww][1]
			wsld.cell(row=dw, column=3).value=addlocal[dww][2]
			wsld.cell(row=dw, column=4).value=addlocal[dww][3]
			dww=dww+1
			dw=dw+1
		gol=0
	dw=dw+1

wbld.save('LIPID_MAPS_local_data.xlsx')
# end save lipidmaps IDs etc that are now found but are not yet in the local database
#print('checkpoint2')

#print(relab)
# write results in summary sheet
if fanclist[0]=='C16:0d31':
	#print('len(fanclist)')
	#print(len(fanclist))
	#print('len(lipidmapsidlist)')
	#print(len(lipidmapsidlist))
	#print('len(linklist)')
	#print(len(linklist))
	#print('len(commonnamelist)')
	#print(len(commonnamelist))
	#print('len(relab)')
	#print(len(relab))
	if len(relab)<len(lipidmapsidlist):
		relab.insert(0,100)
	fw=0
	afw=0
else:
	afw=0
	fw=0
while len(relab)<len(fanclist):
	relab.append('N/A')



while fw<(len(fanclist)):
	#print(fw)
	sheetsummary.cell(row=fw+2+afw, column=4).value=lipidmapsidlist[fw]
	if linklist[fw]=='_':
		fw=fw
	else:
		sheetsummary.cell(row=fw+2+afw, column=4).hyperlink=linklist[fw]
	sheetsummary.cell(row=fw+2+afw, column=5).value=commonnamelist[fw]		# web scraping common name from LipidMaps database  functional here
	sheetsummary.cell(row=fw+2+afw, column=7).value=relab[fw]
	fw=fw+1

#wb.remove_sheet(ozidbc)
wb.remove(ozidbc)

wb.save('OzFAD1_5_plot_table.xlsx')


## end write remaining results summary
afterall=datetime.datetime.now()
dt=afterall-beforeall
#print('Calculation time(h:mm:s) is:')
#print(dt)
if gui==0:
	print('Calculation time (h:mm:ss) is: %s' % dt)
	print('Final results are saved as OzFAD1_5_plot_table.xlsx')	# yyyy_mm_dd
	print('The excel file contains four worksheets:')
	print(' - The first contains a report of the transitions from Skyline. This report can be changed into a transition list.')
	print(' - The second contains a results summary including preliminary integrals.')
	print(' - The third contains a table of relative isomer quantities, which are the data for the barchart shown beneath the table.')
	print('    To complete the barchart shown, copy the legend from the "empty" barchart below into the barchart showing the data.')
	print(' - The fourth contains a final results summary including fatty acid shorthand notations, systematic names, Lipid MAPS IDs, common names, retention times and relative isomer abundance values.')

#end write bar chart data and save in worksheet
###########################################################################################################################################################################
###########################################################################################################################################################################
