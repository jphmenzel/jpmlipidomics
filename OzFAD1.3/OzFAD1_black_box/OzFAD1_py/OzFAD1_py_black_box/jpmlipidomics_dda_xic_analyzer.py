# -*- coding: UTF-8 -*-

#Jan Philipp Menzel 
#Notes: Reads chromatograms file from DDA Skyline export (tsv) and reads target list (txt) to generate excel plottable XICs of each transition from DDA Skyline file.
# Conducts peak detection and assembles filtered transition list for Skyline
# Deletes datapoints that do not belong to the FA species that is targeted
import math
import openpyxl
import pandas as pd
import datetime
import statistics
import csv
import sys
from openpyxl import Workbook

checkup=0 #1 	# set 1 to display all additional info while running code, clean run: checkup=0
#csv.field_size_limit(sys.maxsize)
maxInt=sys.maxsize
while True:
	try:
		csv.field_size_limit(maxInt)
		break
	except OverflowError:
		maxInt=int(maxInt/10)

beforeall=datetime.datetime.now()

segmentsize=500	# min number of entries in xic report and transitions report to be processed at once (once functional change to 500, then increase to test if advantageous)

convertfile=1		# set 0 for troubleshooting (run this python script on its own using csv file), 1 is default value for running workflow through batch file
if convertfile==1:
	# begin convert tsv file generated from Skyline runner to csv file # BEGIN EXTRACT INTENSITIES
	try:
	    with open(r'skyl_xic_report_dda_analyzer.tsv', 'r', newline='\n') as in_f, \
	         open(r'skyl_xic_report_dda_analyzer_intensities.csv', 'w', newline='\n') as out_f:
	        reader = csv.reader(in_f, delimiter='\t')
	        writer = csv.writer(out_f, delimiter=',')
	        for li in reader:
	            try:
	                writer.writerow([li[0], li[1], li[2], li[3], li[4], li[5], li[6], li[7], li[9]])
	            except IndexError:  # Prevent errors on blank lines.
	                pass
	except IOError as err:
	    print(err)
	# end convert tsv file generated from Skyline runner to csv file
	# begin delete double quotes from generated csv file
	with open('skyl_xic_report_dda_analyzer_intensities.csv', "r+", encoding="utf-8") as csv_file:
	    content = csv_file.read()
	with open('skyl_xic_report_dda_analyzer_intensities.csv', "w+", encoding="utf-8") as csv_file:
	    csv_file.write(content.replace('"', ''))
	# end delete double quotes from generated csv file # END EXTRACT INTENSITIES
	# begin convert float values for intensities to integers to reduce file size
	tempdf=pd.read_csv('skyl_xic_report_dda_analyzer_intensities.csv', header=None, skiprows=1)
	templist=tempdf.values.tolist()
	tcol=0
	trow=0
	while trow<(len(templist)):		# replaces content of first column (FileName) with int(0)
		templist[trow][tcol]=int(0)
		trow=trow+1
	tcol=2
	trow=0
	while trow<(len(templist)):		# replaces content of third column (Precursorcharge) with int(0)
		templist[trow][tcol]=int(0)
		trow=trow+1
	tcol=6
	trow=0
	while trow<(len(templist)):		# replaces content of seventh column (IsotopeLabel) with int(0)
		templist[trow][tcol]=int(0)
		trow=trow+1
	tcol=9
	trow=0
	while trow<(len(templist)):		# converts intensities to integers
		tcol=9
		while tcol<(len(templist[0])):
			templist[trow][tcol]=float(templist[trow][tcol])
			templist[trow][tcol]=round(templist[trow][tcol], 0)
			templist[trow][tcol]=int(templist[trow][tcol])
			tcol=tcol+1
		trow=trow+1
	xicintlist=templist
	tempconvdf=pd.DataFrame(templist)
	filename='skyl_xic_report_dda_analyzer_intensities.csv'
	tempconvdf.to_csv(filename, index=False)
	templist=[]
	tempconvdf=pd.DataFrame(templist)
	tempdf=pd.DataFrame(templist)
	# end convert float values for intensities to integers to reduce file size
if convertfile==1:
	# begin convert tsv file generated from Skyline runner to csv file # BEGIN EXTRACT TIMES
	try:
	    with open(r'skyl_xic_report_dda_analyzer.tsv', 'r', newline='\n') as in_f, \
	         open(r'skyl_xic_report_dda_analyzer_times.csv', 'w', newline='\n') as out_f:
	        reader = csv.reader(in_f, delimiter='\t')
	        writer = csv.writer(out_f, delimiter=',')
	        for li in reader:
	            try:
	                writer.writerow([li[8]])
	            except IndexError:  # Prevent errors on blank lines.
	                pass
	except IOError as err:
	    print(err)
	# end convert tsv file generated from Skyline runner to csv file
	# begin delete double quotes from generated csv file
	with open('skyl_xic_report_dda_analyzer_times.csv', "r+", encoding="utf-8") as csv_file:
	    content = csv_file.read()
	with open('skyl_xic_report_dda_analyzer_times.csv', "w+", encoding="utf-8") as csv_file:
	    csv_file.write(content.replace('"', ''))
	# end delete double quotes from generated csv file
	xictimesdf=pd.read_csv('skyl_xic_report_dda_analyzer_times.csv', header=None, skiprows=1, nrows=1)
	xictimeslistfromdf=xictimesdf.values.tolist()
	xictimeslist=xictimeslistfromdf[0]		# contains times of XICs
	aa=0
	while aa<8:
		xictimeslist.insert(0, 0)
		aa=aa+1
	# begin save times of XICs in csv file
	xictimesconvdf=pd.DataFrame(xictimeslist).transpose()
	filename='skyl_xic_report_dda_analyzer_times.csv'
	xictimesconvdf.to_csv(filename, index=False)
	# end save times of XICs in csv file
	xictimesdf=pd.DataFrame(templist)
	xictimeslistfromdf=[]
	#print(xictimesdf)
	#print('list:')
	#print(xictimeslist)
	#print(len(xictimeslist))
	# END EXTRACT TIMES
###################################################################################################################
# end get chromatograms from Skyline  # times are in xictimeslist 	# intensities are in xicintlist #########
###################################################################################################################
if checkup==1:
	print('First FA (XIC):')
	print(xicintlist[0][1])
	print('First RT of first FA (XIC):')
	print(xictimeslist[8])
	print('First Intensity value of first FA (XIC):')
	print(xicintlist[0][8])
# begin read target list
# begin get target list from txt file
rawtarget=open('OzFAD1_dda_targetlist.txt','r')
rawtargetlist=rawtarget.readlines()
#print(rawtargetlist)
mztargetlist=[]
rttargetlist=[]
rdi=0
while rdi<(len(rawtargetlist)):
	cstr=''
	rdii=0
	while rdii<10:
		stri=str(rawtargetlist[rdi][rdii])
		cstr=cstr+stri
		rdii=rdii+1
	cmz=float(cstr)
	mztargetlist.append(cmz)
	cstr=''
	rdii=33
	while rdii<43:
		stri=str(rawtargetlist[rdi][rdii])
		cstr=cstr+stri
		rdii=rdii+1
	if cstr[len(cstr)-1]==',':
		cstr=cstr[:-1]
	crt=float(cstr)/60
	rttargetlist.append(crt)
	rdi=rdi+1
lmzt=len(mztargetlist)
print('Number of targets in targetlist: %d' % lmzt)
if checkup==1:
	print('First target mz:')
	print(mztargetlist[0])
	print('First target RT (min):')
	print(rttargetlist[0])
#quit()
###################################################################################################################
# end get target list from txt file
###################################################################################################################
#cleanxiclist=[] #not used
cxicrt=[]	# clean xic list rt 
cxicint=[]	# clean xic list int
donelist=[]
xfi=0
while xfi<len(xicintlist):
	cfa=str(xicintlist[xfi][1][5])+str(xicintlist[xfi][1][6])+str(xicintlist[xfi][1][7])+str(xicintlist[xfi][1][8])		# current FA, e.g. 14:1
	if '16:0d31' in xicintlist[xfi][1]:
		ok=1
	elif int(cfa[3])==0:
		ok=1
	else:
		if checkup==1:
			print('Analyzing FA:')
			print(xicintlist[xfi][1])
		if int(cfa[3])>0:				#if unsaturated FA, add db to cfa, then check donelist for it (e.g. 14:1_n-5 or 18:2_n-6_n-9)
			dusi=1
			hc=0
			while dusi<(int(cfa[3])+1):
				if str(xicintlist[xfi][1][9+hc+(dusi*4)])=='_':
					cfa=cfa+str(xicintlist[xfi][1][5+hc+(dusi*4)])+str(xicintlist[xfi][1][6+hc+(dusi*4)])+str(xicintlist[xfi][1][7+hc+(dusi*4)])+str(xicintlist[xfi][1][8+hc+(dusi*4)])
				elif str(xicintlist[xfi][1][10+hc+(dusi*4)])=='_':
					cfa=cfa+str(xicintlist[xfi][1][5+hc+(dusi*4)])+str(xicintlist[xfi][1][6+hc+(dusi*4)])+str(xicintlist[xfi][1][7+hc+(dusi*4)])+str(xicintlist[xfi][1][8+hc+(dusi*4)])+str(xicintlist[xfi][1][9+hc+(dusi*4)])
					hc=hc+1
				dusi=dusi+1
		if checkup==1:
			print('FA:')
			print(cfa)

		if cfa in donelist:
			ok=1
		else:
			donelist.append(cfa)								# OK, proceeding here, if a new species is added to donelist that was not processed before
			ccxicrt=[]	# current clean xic list rt values
			ccxicint=[]	# current clean xic list int values (precursor)
			ccxicozidint=[]	# current clean xic list int values (products)
			ccxicrt.append('X '+cfa)	
			ccxicint.append('Y '+cfa)	
			# go through targetlist and identify the targets relevant to this FA species - use the RT to find the associated XIC intensity values for this target and 
			# add to clean XIC lists
			precfind=0
			ozcount=0
			while precfind==0:
				if 'precursor' in xicintlist[xfi][4]:
					precfind=1
				else:
					ozcount=ozcount+1
					xfi=xfi+1
			fxozcount=ozcount		# number of OzID product ions
			cxicozidint=[]
			cxicozidint.append('Y '+cfa+' OzID')
			while ozcount>0:
				ccxicozidint.append(['Y '+cfa+' OzID'])  # each appended list needs to be identified separately, rather than pointing to same list
				ozcount=ozcount-1
			ozcount=fxozcount
			if checkup==2:
				print(ccxicozidint)
				print('^^^^^^^')
			cmzprec=float(xicintlist[xfi][3])			# current precursor mz
			tli=0
			while tli<len(rttargetlist):
				if abs(float(mztargetlist[tli])-cmzprec)<0.2:		# if target and current precursor match in mz (same FA group)
					rti=8
					if checkup==2:
						print('###################################################')
					while rti<len(xictimeslist):
						if abs(float((rttargetlist[tli]))-float(xictimeslist[rti]))<0.009: # tolerance to capture actual targeting vs. targets in targetlist 
							if xicintlist[xfi][rti]>150:
								if checkup==2:
									print(cfa)
									print('RT (targetlist):')
									print(rttargetlist[tli])
									print('RT (XIC):')
									print(xictimeslist[rti])
									print('Intensity precursor:')
									print(xicintlist[xfi][rti])	
									print('-------------------------')		
								# assigned a target and xic values to use for clean XIC
								# decide, if scan actually targeted or not	# delete if both neighboring entries are more than 3x the current, 
								# also delete two neighboring, if their neighbors are more than 3x
								ign=0
								if xicintlist[xfi][rti-1]>(3*float(xicintlist[xfi][rti])):
									if xicintlist[xfi][rti+1]>(3*float(xicintlist[xfi][rti])):
										ign=1	# scan not from this target 
								if xicintlist[xfi][rti-1]>(3*float(xicintlist[xfi][rti])):
									if xicintlist[xfi][rti+2]>(3*float(xicintlist[xfi][rti])):
										ign=1	# scan not from this target 
								if xicintlist[xfi][rti-2]>(3*float(xicintlist[xfi][rti])):
									if xicintlist[xfi][rti+1]>(3*float(xicintlist[xfi][rti])):
										ign=1	# scan not from this target 
								if ign==0:
									ccxicrt.append(xictimeslist[rti])	
									ccxicint.append(xicintlist[xfi][rti])
									# add XIC data for OzID product ions !!!
									ozc=1
									while ozcount>0:
										adding=int(xicintlist[xfi-ozc][rti])
										tempcl=ccxicozidint[ozcount-1]
										tempcl.append(adding)		## OK
										ccxicozidint[ozcount-1]=tempcl
										#print('ADDED')
										#print(xicintlist[xfi-ozc][rti])
										#print(ccxicozidint)
										ozc=ozc+1
										ozcount=ozcount-1
									ozcount=fxozcount

									#print(ccxicozidint)
									#quit()	
									if checkup==1:
										if '16:1_n-7' in xicintlist[xfi][1]:
											#print('###')
											if rttargetlist[tli]>7.35:
												if rttargetlist[tli]<7.46:
													print(cfa)
													print('RT (targetlist):')
													print(rttargetlist[tli])
													print('RT (XIC):')
													print(xictimeslist[rti])
													print('Intensity precursor:')
													print(xicintlist[xfi][rti])	
													rterror=rttargetlist[tli]-xictimeslist[rti]
													print(rterror)
													print('-------------------------')					
						rti=rti+1
				tli=tli+1
			#quit()
			cxicrt.append(ccxicrt)
			cxicint.append(ccxicint)
			ozc=0
			while ozcount>0:
				cxicrt.append(ccxicrt)
				cxicint.append(ccxicozidint[ozc]) #len(ccxicozidint)-1-ozc])		#
				ozc=ozc+1
				ozcount=ozcount-1
			ozcount=fxozcount


		ok=1
	xfi=xfi+1
	if checkup==2:
		if len(donelist)>2:
			################################################
			xfi=len(xicintlist)
			################################################


if checkup==1:
	print(donelist)
	print('$$$$$$$$$$')
	print(cxicrt)		## rt 
	print(cxicint)		## int
	print('$$$$$$$$$$')
	#print(ozcount)
	print(ccxicrt)
	print(ccxicint)
	print(ccxicozidint)
	#quit()

# begin transcribe lists into new lists that can be edited
lxrt=len(cxicrt)
lxint=len(cxicint)
if lxrt==lxint:
	ok=1
else:
	print('lists need to be reviewed')
txicrt=[]
txicint=[]
for i in range(lxrt):
	txicrt.append([])
	txicint.append([])
spi=0	#index to iterate through spectra
while spi<len(cxicrt):
	xi=0
	while xi<len(cxicrt[spi]):
		txicrt[spi].append(cxicrt[spi][xi])
		txicint[spi].append(cxicint[spi][xi])
		xi=xi+1
	spi=spi+1
# end transcribe lists into new lists that can be edited	# OK


# begin go through XICs to remove duplicate datapoints 
spi=0	#index to iterate through spectra
while spi<len(txicrt):
	xi=1
	while xi<len(txicrt[spi]):
		xsi=xi+1
		while xsi<len(txicrt[spi]):
			if xsi>xi:
				if txicrt[spi][xi]==txicrt[spi][xsi]:
					if txicint[spi][xi]==txicint[spi][xsi]:
						# found duplicate entry, delete this one (xsi)
						del txicrt[spi][xsi]
						del txicint[spi][xsi]
						xsi=xsi-1
			xsi=xsi+1
		xi=xi+1
	spi=spi+1
# end go through XICs to remove duplicate datapoints	# OK

rawcheckup=0
if rawcheckup==1:
	#begin save clean xic lists in excel file	# OK
	wb = Workbook(write_only=True)
	ws = wb.create_sheet()
	wb.save('Clean_DDA_XICs.xlsx')
	wb=openpyxl.load_workbook('Clean_DDA_XICs.xlsx')
	ws=wb.active
	cxi=0
	cxii=1
	colx=1
	while cxi<len(txicrt):
		cndb=int(txicrt[cxi][0][5])	# number of double bonds in current species
		xi=0
		while xi<len(txicrt[cxi]):
			#ws.cell(row=xi+1, column=(cxi+1)*2).value=cxicrt[cxi][xi]
			#ws.cell(row=xi+1, column=((cxi+1)*2)+1).value=cxicint[cxi][xi]
			ws.cell(row=xi+1, column=(colx)).value=txicrt[cxi][xi]
			ws.cell(row=xi+1, column=(colx+1)).value=txicint[cxi][xi]
			xi=xi+1
		colx=colx+2
		cxi=cxi+1
		idb=0
		while idb<cndb:
			xi=0
			while xi<len(txicrt[cxi]):
				ws.cell(row=xi+1, column=(colx)).value=txicrt[cxi][xi]	# aldehyde
				ws.cell(row=xi+1, column=(colx+1)).value=txicint[cxi][xi]
				ws.cell(row=xi+1, column=(colx+2)).value=txicrt[cxi+1][xi]	# criegee
				ws.cell(row=xi+1, column=(colx+3)).value=txicint[cxi+1][xi]
				xi=xi+1
			colx=colx+4
			cxi=cxi+2
			idb=idb+1
	wb.save('Clean_DDA_XICs.xlsx')
	print('Clean XIC data from DDA acquisition are saved in Clean_DDA_XICs.xlsx')


# begin derive peaks, get RTs and build dECL plot, make assignments and build transition list with assignments and exact RT for Skyline !!!
# begin run 5 point smoothing of all XICs and save in another file
allfpsxicrt=[]
allfpsxicint=[]
spi=0
while spi<len(txicrt):
	fpsxicrt=[]
	fpsxicint=[]
	fpsxicrt.append(txicrt[spi][0])			# copy spectrum info into first list entry
	fpsxicint.append(txicint[spi][0])		# copy spectrum info into first list entry
	if len(txicrt[spi])>1:
		fpsxicrt.append(txicrt[spi][1])		# copy first datapoint
		fpsxicint.append(txicint[spi][1]) 	# copy first datapoint
	if len(txicrt[spi])>2:
		fpsxicrt.append(txicrt[spi][2])		# copy second datapoint
		fpsxicint.append(txicint[spi][2])	# copy second datapoint
	nrt=3
	while nrt<(len(txicrt[spi])-2):	# RT
		fpsxicrt.append(txicrt[spi][nrt])
		nrt=nrt+1
	nrt=3
	while nrt<(len(txicint[spi])-2):	# average of 3 and 5 point smoothed intensity of XIC
		tpsn=((txicint[spi][nrt-1]+txicint[spi][nrt]+txicint[spi][nrt+1])/3)
		fpsn=((txicint[spi][nrt-2]+txicint[spi][nrt-1]+txicint[spi][nrt]+txicint[spi][nrt+1]+txicint[spi][nrt+2])/5)
		tfspn=(tpsn+fpsn)/2
		fpsxicint.append(tfspn)
		nrt=nrt+1
	if len(txicrt[spi])==4:
		fpsxicrt.append(txicrt[spi][len(txicrt[spi])-1])
		fpsxicint.append(txicint[spi][len(txicint[spi])-1])
	elif len(txicrt[spi])>4:
		fpsxicrt.append(txicrt[spi][len(txicrt[spi])-2])
		fpsxicrt.append(txicrt[spi][len(txicrt[spi])-1])
		fpsxicint.append(txicint[spi][len(txicint[spi])-2])
		fpsxicint.append(txicint[spi][len(txicint[spi])-1])

	allfpsxicrt.append(fpsxicrt)
	allfpsxicint.append(fpsxicint)
	spi=spi+1

#begin save clean five point smoothed xic lists in excel file	# OK
wb = Workbook(write_only=True)
ws = wb.create_sheet()
wb.save('Clean_DDA_XICs_smoothed.xlsx')
wb=openpyxl.load_workbook('Clean_DDA_XICs_smoothed.xlsx')
ws=wb.active
cxi=0
cxii=1

colx=1
while cxi<len(allfpsxicrt):
	cndb=int(allfpsxicrt[cxi][0][5])	# number of double bonds in current species
	xi=0
	while xi<len(allfpsxicrt[cxi]):
		#ws.cell(row=xi+1, column=(cxi+1)*2).value=cxicrt[cxi][xi]
		#ws.cell(row=xi+1, column=((cxi+1)*2)+1).value=cxicint[cxi][xi]
		ws.cell(row=xi+1, column=(colx)).value=allfpsxicrt[cxi][xi]
		ws.cell(row=xi+1, column=(colx+1)).value=allfpsxicint[cxi][xi]
		xi=xi+1
	colx=colx+2
	cxi=cxi+1
	idb=0
	while idb<cndb:
		xi=0
		while xi<len(allfpsxicrt[cxi]):
			ws.cell(row=xi+1, column=(colx)).value=allfpsxicrt[cxi][xi]	# aldehyde
			ws.cell(row=xi+1, column=(colx+1)).value=allfpsxicint[cxi][xi]
			ws.cell(row=xi+1, column=(colx+2)).value=allfpsxicrt[cxi+1][xi]	# criegee
			ws.cell(row=xi+1, column=(colx+3)).value=allfpsxicint[cxi+1][xi]
			xi=xi+1
		colx=colx+4
		cxi=cxi+2
		idb=idb+1

wb.save('Clean_DDA_XICs_smoothed.xlsx')
print('Clean XIC data from DDA acquisition are saved in Clean_DDA_XICs_smoothed.xlsx')
# end run (average of 3 and 5 point) smoothing of all XICs and save in another file

#quit()

# begin derive peak positions with exact RT (save as peak list), make assignment for cis trans and branched based on dECL plot, make dECL plot and build transition list for Skyline
# begin build lists with FA species as empty peak lists
lxrt=len(cxicrt)
lxint=len(cxicint)
peakrt=[]
for i in range(lxrt):
	peakrt.append([])
spi=0	#index to iterate through spectra
while spi<len(cxicrt):
	peakrt[spi].append(allfpsxicrt[spi][0])
	spi=spi+1

# end build lists with FA species as empty peak lists (only FA species as only entry of each list)
#print(len(cxicrt))
#print(len(peakrt))
#print(peakrt)
# begin find peaks
spi=0
while spi<len(allfpsxicrt):
	if 'OzID' in allfpsxicint[spi][0]:
		spi=spi
	elif len(allfpsxicrt[spi])==1:
		spi=spi
	else:
		#print('Analyzing:')
		#print(allfpsxicint[spi][0])
		# analyse this FA
		noz=2*int(allfpsxicrt[spi][0][5])	# number of OzID XICs
		cmaxrtlist=[]
		cnoz=1
		while cnoz<(noz+1):
			f=allfpsxicint[spi+cnoz]
			g=f[1:]
			#print(g)
			cmaxint_index=(g.index(max(g)))+1
			cmaxrt=float(allfpsxicrt[spi+cnoz][cmaxint_index])
			cmaxrtlist.append(cmaxrt)
			cnoz=cnoz+1
		# average the RT values of the max int of the OzID product peaks
		avgmaxrt=sum(cmaxrtlist)/len(cmaxrtlist)
		peakrt[spi].append(avgmaxrt)
		
	spi=spi+1

#print('#########################################$$$$$$$$$$$$$$$$$')
#print('peakrt')
#print(peakrt)
# Retention times of features with max OzID product intensity are saved - begin building dECL plot and assigning EZ isomerism and branching (after: detect other peaks and assign)

# write module to extract RTs of SatFA and determince 2nd order polynomial fit and expression			!!!!!!!!!!!!!!!!!!!!!!!!!
apoly=0.04115	# value for MCF7 in OzFAD paper
bpoly=0.79914	# value for MCF7 in OzFAD paper
cpoly=6.50734	# value for MCF7 in OzFAD paper

mufalist=[]
mufadecl=[]
mufacndb=[]
mufaez=[]
mufart=[]

spi=0
while spi<len(peakrt):
	if len(peakrt[spi])==1:
		spi=spi+1
	else:
		ecl=((apoly*peakrt[spi][1]*peakrt[spi][1])+(bpoly*peakrt[spi][1])+cpoly)
		decl=ecl-((10*int(peakrt[spi][0][2]))+int(peakrt[spi][0][3]))
		peakrt[spi].append(decl)
		# get first double bond position
		if len(peakrt[spi][0])==10:
			cndb=int(peakrt[spi][0][9])
		elif len(peakrt[spi][0])>10:
			if str(peakrt[spi][0][10])=='_':
				cndb=int(peakrt[spi][0][9])
			else:
				cndb=10*int(peakrt[spi][0][9])+int(peakrt[spi][0][10])
		else:
			print('Check')
			cndb=0
		peakrt[spi].append(cndb)
		if int(peakrt[spi][0][5])==1:	#MUFA	
			cfa=peakrt[spi][0][2:]
			mufalist.append(cfa)
			mufacndb.append(cndb)
			mufadecl.append(decl)
			cet=str(peakrt[spi][0][2])+str(peakrt[spi][0][3])+str(peakrt[spi][0][4])+str(peakrt[spi][0][5])+' Z'
			mufaez.append(cet)
			fart=float(peakrt[spi][1])
			mufart.append(fart)
		spi=spi+1
if checkup==1:
	print('peakrt')
	print(peakrt)

# begin write data (and sort) in excel file as dECL plot input (MUFA)
# sort mufalists
mfs=[]
mfdb=[]
mfdecl=[]
mfez=[]
mfart=[]

mf=0
while mf<len(mufalist):
	cfal=[]
	cfal.append(mufalist[mf])
	cdbl=[]
	cdbl.append(mufacndb[mf])
	cfi=[]
	cfi.append(mf)
	mfi=mf+1
	mfik=mfi
	while mfik<len(mufaez):
		if mufaez[mfi]==mufaez[mf]:	# get all entries of current FA and their first db
			cfal.append(mufalist[mfi])
			cdbl.append(mufacndb[mfi])
			cfi.append(mfi)
			mfi=mfi+1
			mfik=mfi
		else:
			mfik=len(mufaez)
	# enter entries according to cfal and cdbl into new sorted mflists
	while len(cfal)>0:
		addi=cdbl.index(min(cdbl))
		addj=cfi[addi]
		mfs.append(mufalist[addj])
		mfdb.append(mufacndb[addj])
		mfdecl.append(mufadecl[addj])
		mfez.append(mufaez[addj])
		mfart.append(mufart[addj])
		del cfal[addi]
		del cdbl[addi]
		del cfi[addi]
	mf=mfi+1
if checkup==1:
	print('########################################')
	print(mfs)
	print(mfdb)
	print(mfdecl)
	print(mfez)
	print(mfart)

#quit()

# begin write data into excel file for plotting (then develop automated assignment)
wb = Workbook(write_only=True)
ws = wb.create_sheet()
wb.save('MUFA_preliminary_dECL_plot.xlsx')
wb=openpyxl.load_workbook('MUFA_preliminary_dECL_plot.xlsx')
ws=wb.active

ws.cell(row=1, column=8).value='The data shown in this file are preliminary and are not to be used as is.'	# 

ws.cell(row=1, column=4).value='Parameters for 2nd order polynomial fit'	# 
ws.cell(row=2, column=3).value='a'	#
ws.cell(row=3, column=3).value='b'	#
ws.cell(row=4, column=3).value='c'	#
ws.cell(row=2, column=4).value=apoly	#
ws.cell(row=3, column=4).value=bpoly	#
ws.cell(row=4, column=4).value=cpoly	#

ws.cell(row=6, column=2).value='FA'	# 
ws.cell(row=6, column=3).value='RT'	# 
ws.cell(row=6, column=4).value='dECL'	# 
written=[]
ccol=4
fi=0
while fi<len(mfs):
	# write lines in excel file
	ws.cell(row=fi+7, column=2).value=mfs[fi]	# current FA
	ws.cell(row=fi+7, column=3).value=mfart[fi]	# RT
	ws.cell(row=fi+7, column=4).value=mfdecl[fi]	# dECL
	if mfez[fi] in written:
		ws.cell(row=fi+7, column=ccol).value=mfdb[fi]
	else:
		written.append(mfez[fi])
		ws.cell(row=6, column=ccol+1).value=str(mfez[fi])
		ws.cell(row=fi+7, column=ccol+1).value=mfdb[fi]	
		ccol=ccol+1
	fi=fi+1

# write data for transposed plot of dECL vs. db
ws.cell(row=6, column=ccol+4).value='FA'	# 
ws.cell(row=6, column=ccol+5).value='RT'	# 
ws.cell(row=6, column=ccol+6).value='DB & dECL'	# 
written=[]
ccol=ccol+6
fcol=ccol-2
fi=0
while fi<len(mfs):
	# write lines in excel file
	ws.cell(row=fi+7, column=fcol).value=mfs[fi]	# current FA
	ws.cell(row=fi+7, column=fcol+1).value=mfart[fi]	# RT
	ws.cell(row=fi+7, column=fcol+2).value=mfdb[fi]	# db
	if mfez[fi] in written:
		ws.cell(row=fi+7, column=ccol).value=mfdecl[fi]
	else:
		written.append(mfez[fi])
		ws.cell(row=6, column=ccol+1).value=str(mfez[fi])
		ws.cell(row=fi+7, column=ccol+1).value=mfdecl[fi]	
		ccol=ccol+1
	fi=fi+1

wb.save('MUFA_preliminary_dECL_plot.xlsx')
print('Data for the preliminary dECL plot are saved in MUFA_preliminary_dECL_plot.xlsx')
# end write data (and sort) in excel file as dECL plot input (MUFA)

# begin detect other peaks, all data from smoothed XICs are in allfpsxicrt and allfpsxicint
if checkup==1:
	print(allfpsxicrt)
	print(allfpsxicint)

newpeakrt=[]
newpeakfa=[]
newpeakinta=[]
newpeakintb=[]
newpeakint=[]
spi=0
while spi<len(allfpsxicrt):
	newpeakrta=[]
	newpeakrtb=[]
	xi=1
	while xi<(len(allfpsxicint[spi+1])-2):
		if float(allfpsxicint[spi+1][xi])<float(allfpsxicint[spi+1][xi+1]):
			if float(allfpsxicint[spi+1][xi+1])>float(allfpsxicint[spi+1][xi+2]):
				if abs(float(allfpsxicrt[spi+1][xi+1])-float(peakrt[spi][1]))>0.07:
					#maximum other than previously detected peak found here based on one OzID product ion XIC
					newpeakrta.append(float(allfpsxicrt[spi+1][xi+1]))
					newpeakinta.append(float(allfpsxicint[spi+1][xi+1]))
		if float(allfpsxicint[spi+2][xi])<float(allfpsxicint[spi+2][xi+1]):
			if float(allfpsxicint[spi+2][xi+1])>float(allfpsxicint[spi+2][xi+2]):
				if abs(float(allfpsxicrt[spi+2][xi+1])-float(peakrt[spi][1]))>0.07:
					#maximum other than previously detected peak found here based on other OzID product ion XIC
					newpeakrtb.append(float(allfpsxicrt[spi+2][xi+1]))
					newpeakintb.append(float(allfpsxicint[spi+2][xi+1]))
		xi=xi+1
	np=0
	while np<len(newpeakrta):
		npi=0
		while npi<len(newpeakrtb):
			if abs(newpeakrta[np]-newpeakrtb[npi])<0.05:
				#matching maxima --> another peak detected
				newpeakrt.append(float((newpeakrta[np]+newpeakrtb[npi])/2))
				newpeakfa.append(str(allfpsxicrt[spi][0][2:]))
				newpeakint.append(float((newpeakinta[np]+newpeakintb[npi])/2))
			npi=npi+1
		np=np+1
	spi=spi+1
	go=1
	while go==1:
		if 'OzID' in allfpsxicint[spi][0]:
			spi=spi+1
			if spi>(len(allfpsxicint)-1):
				go=0
		else:
			go=0

#print('------------------ new peak ------------')
#print(newpeakfa)
#print(newpeakrt)
#print(len(newpeakfa))
# allow up to two peak detections before and after each detected peak
# delete instances of peak detections that are within 0.05 min of each other (same peak detected)
np=0
while np<len(newpeakfa):
	npi=0
	while npi<len(newpeakfa):
		if np==npi:
			npi=npi+1
		else:
			#print(np)
			#print(npi)
			if newpeakfa[np]==newpeakfa[npi]:
				if abs(newpeakrt[np]-newpeakrt[npi])<0.05:
					if newpeakint[npi]<newpeakint[np]:
						del newpeakrt[npi]
						del newpeakfa[npi]
						del newpeakint[npi]
						npi=npi-1
						#np=np-1
					else:
						del newpeakrt[np]
						del newpeakfa[np]
						del newpeakint[np]
						#np=np-1
						npi=npi-1
			npi=npi+1
	np=np+1
if checkup==1:
	print('------------------ new peak ------------')
	print(newpeakfa)
	print(newpeakrt)
	print(len(newpeakrt))
# OK
#print(len(mfs))
#print(mfs)
# end derive peaks, get RTs and build preliminary dECL plot

# begin merge and sort peakrt and newpeakfa etc to build transition list with assignments for Skyline
# add all species in allidusfa/rt (unsorted)
allidusfa=[]	# FA, eg 12:1_n-7
allidusrt=[]	# RT, eg 4.556267
allidusdb=[]	# DB, eg 7
allidusfg=[]	# FA group eg 12:1
spi=0
while spi<len(peakrt):
	if len(peakrt[spi])==1:
		spi=spi+1
	else:
		allidusfa.append(peakrt[spi][0][2:])
		allidusrt.append(peakrt[spi][1])
		allidusdb.append(peakrt[spi][3])
		allidusfg.append(str(peakrt[spi][0][2])+str(peakrt[spi][0][3])+str(peakrt[spi][0][4])+str(peakrt[spi][0][5]))
		spi=spi+1
np=0
while np<len(newpeakfa):
	allidusfa.append(newpeakfa[np])
	allidusrt.append(newpeakrt[np])
	if len(newpeakfa[np])==8:
		allidusdb.append(int(newpeakfa[np][7]))
	elif len(newpeakfa[np])==9:
		allidusdb.append(int(10*int(newpeakfa[np][7])+int(newpeakfa[np][8])))
	elif str(newpeakfa[np][8])=='_':
		allidusdb.append(int(newpeakfa[np][7]))
	elif str(newpeakfa[np][9])=='_':
		allidusdb.append(int(10*int(newpeakfa[np][7])+int(newpeakfa[np][8])))
	allidusfg.append(str(newpeakfa[np][0])+str(newpeakfa[np][1])+str(newpeakfa[np][2])+str(newpeakfa[np][3]))
	np=np+1
if checkup==1:
	print('Unsorted merged lists of detected species:')		# OK
	print(allidusfa)
	print(allidusrt)
	print(allidusdb)
	print(allidusfg)
	print(len(allidusfa))
# sort all species into allidfa/rt						# OK
allidfa=[]	#sorted	
allidrt=[]
alliddb=[]	
allidfg=[]

mf=0
while mf<len(allidusfa):
	cfal=[]
	cfal.append(allidusfa[mf])
	cdbl=[]
	cdbl.append(allidusdb[mf])
	cfi=[]
	cfi.append(mf)
	mfi=mf+1
	mfik=mfi
	while mfik<len(allidusfa):
		if allidusfg[mfi]==allidusfg[mf]:	# get all entries of current FA group and their first db
			cfal.append(allidusfa[mfi])
			cdbl.append(allidusdb[mfi])
			cfi.append(mfi)
			mfi=mfi+1
			mfik=mfi
		else:
			mfi=mfi+1
			mfik=mfik+1 #len(allidusfg)
	# enter entries according to cfal and cdbl into new sorted mflists
	#print(cfal)
	while len(cfal)>0:
		addi=cdbl.index(min(cdbl))
		addj=cfi[addi]
		allidfa.append(allidusfa[addj])
		alliddb.append(allidusdb[addj])
		allidrt.append(allidusrt[addj])
		allidfg.append(allidusfg[addj])
		del cfal[addi]
		del cdbl[addi]
		del cfi[addi]

	if allidusfg[mf]=='17:1*':	# disabled checkup
		#print(cfal)
		#print(cdbl)
		print(allidfa)
		print(allidrt)
		print(alliddb)
		print(allidfg)
		print(len(allidfa))
		quit()

	mf=mf+1
	skip=0
	if len(allidusfg)>mf:
		if allidusfg[mf] in allidfg:
			skip=1
	while skip==1:
		mf=mf+1
		if len(allidusfg)>mf:
			if allidusfg[mf] in allidfg:
				skip=1
			else:
				skip=0
		else:
			skip=0
if checkup==0:
	print('Sorted merged lists of detected species:')		# OK
	print(allidfa)
	print(allidrt)
	#print(alliddb)
	#print(allidfg)
	#print(len(allidfa))
# all species sorted, begin build transition list for Skyline

# go through allidfa, for each entry build one block of transitions in transition list (using entries from report OzFAD1_2_DDA_found.csv) use RT from allidrt
# begin read transition list from previous run (csv) into lists
trdf=pd.read_csv('OzFAD1_2_DDA_found_preliminary.csv')
toprowx=[trdf.columns.values.tolist()]
toprow=toprowx[0]
trdf=trdf.transpose()
writelist=trdf.values.tolist()
ki=len(writelist[0])
# end read report csv into lists
#print(writelist)
#quit()

cmlistname=[]
cprecname=[]
cprecformula=[]
cprecadduct=[]
cprecmz=[]
cprecchrg=[]
cprodname=[]
cprodformula=[]
cprodadduct=[]
cprodmz=[]
cprodchrg=[]
cexplicitrt=[]
cexrtwindow=[]
cwritelist=[]

ai=0
while ai<len(allidfa):
	sr=0
	while sr<len(writelist[1]): #report col 2
		if allidfa[ai] in writelist[1][sr]:
			#print('found')
			#print(allidfa[ai])
			#quit()
			sj=sr
			while writelist[1][sj]==writelist[1][sr]:
				cmlistname.append(writelist[0][sj])
				if str(writelist[1][sj][-5])=='_':
					cprecname.append(str(writelist[1][sj][:-5])+'_'+str(round(allidrt[ai], 2)))
				elif str(writelist[1][sj][-4])=='_':
					cprecname.append(str(writelist[1][sj][:-4])+'_'+str(round(allidrt[ai], 2)))
				elif str(writelist[1][sj][-6])=='_':
					cprecname.append(str(writelist[1][sj][:-6])+'_'+str(round(allidrt[ai], 2)))
				cprecformula.append(writelist[2][sj])
				cprecadduct.append(writelist[3][sj])
				cprecmz.append(writelist[4][sj])
				cprecchrg.append(writelist[5][sj])
				cprodname.append(writelist[6][sj])
				cprodformula.append(writelist[7][sj])
				cprodadduct.append(writelist[8][sj])
				cprodmz.append(writelist[9][sj])
				cprodchrg.append(writelist[10][sj])
				cexplicitrt.append(round(allidrt[ai], 6))
				cexrtwindow.append(0.1)
				sj=sj+1
			sr=len(writelist[1])
		sr=sr+1
	ai=ai+1

cwritelist.append(cmlistname)
cwritelist.append(cprecname)
cwritelist.append(cprecformula)
cwritelist.append(cprecadduct)
cwritelist.append(cprecmz)
cwritelist.append(cprecchrg)
cwritelist.append(cprodname)
cwritelist.append(cprodformula)
cwritelist.append(cprodadduct)
cwritelist.append(cprodmz)
cwritelist.append(cprodchrg)
cwritelist.append(cexplicitrt)
cwritelist.append(cexrtwindow)

transitionresultsdf=pd.DataFrame(cwritelist).transpose()		#print('Transposed')
transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
filename='jpmlipidomics_dda_found.csv'
transitionresultsdf.to_csv(filename, index=False)
print('Transition list for Skyline is saved as jpmlipidomics_dda_found.csv')


# connect to batch files and implement loading into Skyline !!!!!!!!!!!!!! test on multiple datasets










		




