# -*- coding: UTF-8 -*-

# Jan Philipp Menzel jpm_lipidomics_vpw13_1_precursor_tr.py
# Goal: STEP 1. Generate transition list for Skyline containing precursors (intact derivatized fatty acids), either one of seven pre-defined derivatives or any new structure
## Notes: Derivative, positive fixed charge
## NOTES: VIRTUAL PRECURSOR - PrecursorName and PrecursorMz are artificially set +Xe (only column 3 and 5), fragment transitions correct including precursor
## NOTES: Virtual precursor forces Skyline to consider all transitions incl. real precursor (fragment in transition list), Skyline Setting: TransitionSettings-Filter-IonTypes-f 
import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
################ DATABASE ## Source: Internetchemie.info 
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.757, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
################
#default=eval(input('Run workflow with default parameters? (Yes: 1 / No: 0) (Derivatization agent: AMPP; Slow and full workflow including all FA; Apply retention time limitation; Use Fatty Acid Library to prevent filtering out important FA; Max RT = 17.5 min; FA: 12 - 24 C; Max m/z error: 10 ppm; Precursor peak area threshold: 3000; Product peak area threshold: 200.) :'))
default=0
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
###########
checkup=1
widestep=1	# step width for generating transitionsat varied RT
vxcut=0

# begin read workflow parameters
transferlist=[]
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
go=1
while go==1:
	tfe=ws.cell(row=tli+1, column=2)
	tfe=tfe.value
	if tli==21:
		go=0
	if tfe is None:
		go=0
	else:
		transferlist.append(tfe)
	tli=tli+1
########## 
fourlettcode=transferlist[0]		#input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
cderiv=int(transferlist[1])
hderiv=int(transferlist[2])
dderiv=int(transferlist[3])
nderiv=int(transferlist[4])
oderiv=int(transferlist[5])
pderiv=int(transferlist[6])
ideriv=int(transferlist[7])
minlenfa=int(transferlist[8]) #eval(input('Enter number of C atoms in shortest FA chain (e.g. 4) :'))
maxlenfa=int(transferlist[9]) #eval(input('Enter number of C atoms in longest FA chain (e.g. 24) :'))
mzcutoff=int(transferlist[10]) 		#eval(input('What is max mz [ppm] for positive identification of species? (e.g. 30) :'))
rettimecutoff=float(transferlist[11]) 		#eval(input('What is the maximum retention time [min]? (e.g. 11.9 or 17.8) :'))
areathreshold=int(transferlist[12])		#eval(input('What is the threshold for detected precursor peak area? (compare Skyline report file, e.g. 3000) :'))
workflowmode=int(transferlist[13])
rtlimitation=int(transferlist[16])
mostwanted=int(transferlist[17])
runprecheck=int(transferlist[19])
identifier=str(transferlist[20])
# end read workflow parameters
if runprecheck==1:
	# begin derive setpoint and slope from chromatograms of precheck analysis
	convertfile=0
	if convertfile==0:
		xictimesdf=pd.read_csv('skyl_xic_report_vpw20_0_times.csv', header=None, skiprows=1, nrows=1)
		xictimeslistfromdf=xictimesdf.values.tolist()
		xictimeslist=xictimeslistfromdf[0]		# contains times of XICs # actual list starts at index 8

	segtrdf=pd.read_csv('skyl_xic_report_vpw20_0_intensities.csv', skiprows=1, header=None, nrows=3, low_memory=False)
	allxiclist=segtrdf.values.tolist()
	#begin write trdf to csv file to check its contents 	# TROUBLESHOOTING
	#trdf.to_csv('skyl_xic_report_vpw16_3_troubleshooting.csv', index=False)
	#end write trdf to csv file to check its contents 	# TROUBLESHOOTING
	#trdf=trdf.transpose()#

	#print(allxiclist[0][1])		####################################### 
	kix=len(allxiclist)
	#print('Number of rows in segment in skyl_xic_report_vpw20_0_precheck.csv: %d' % kix)
	#begin determine columns in xic_results, determine length of XIC
	ci=len(allxiclist[0])
	xiclength=int((ci-8)) #/2
	#print('XIC length')
	#print(xiclength)
	#end determine columns in xic_results, determine length of XIC
	intalist=[]
	intblist=[]
	itl=8

	cutxic=(len(allxiclist[0]))
	if fourlettcode=='PLPC':
		cutxic=(len(allxiclist[0]))/3
	elif fourlettcode=='PLPE':
		cutxic=(len(allxiclist[0]))/3

	while itl<cutxic:
		intalist.append(int(allxiclist[0][itl]))
		intblist.append(int(allxiclist[1][itl]))
		if fourlettcode=='PLPC':
			if int(xictimeslist[itl])>10:
				cutxic=itl
		itl=itl+1
	maxinta=max(intalist)
	maxintb=max(intblist)
	#print('maxinta')
	#print(maxinta)
	#print('maxintb')
	#print(maxintb)
	itl=8
	while itl<(len(allxiclist[0])):
		if allxiclist[0][itl]==maxinta:
			ita=itl
		if allxiclist[1][itl]==maxintb:
			itb=itl
		itl=itl+1
	rtmaxa=xictimeslist[ita]
	rtmaxb=xictimeslist[itb]
	#print('rtmaxa')
	#print(rtmaxa)
	#print('rtmaxb')
	#print(rtmaxb)
	setpoint=(rtmaxa+rtmaxb)/2
	slope=(rtmaxb-rtmaxa)/2
	#print('Arrived here')
	# end derive setpoint and slope from chromatograms of precheck analysis
else:
	rtlimitation=2
	setpoint=8.14	#eval(input('Setpoint: '))
	slope=0.57		#eval(input('Slope: '))

predictedrtindex=[]	# list containing species identifier ['04:0', '05:0', ...'30:6', ... '40:6']
predictedrtexact=[]
predictedrtmin=[]	# list containing predicted RT range start for species identifier with same list index
predictedrtmax=[]	# list containing predicted RT range end for species identifier with same list index
predictionconstantsat=[-13.91, -13.8, -13.2, -11.9, -10.575, -9.2, -7.785, -6.6, -5.4, -4.208, -3.075, -2.02, -1, 0, 1, 1.94, 2.85, 3.68, 4.525, 5.34, 6.139, 6.92, 7.668, 8.4, 9.13, 9.84, 10.53, 11.2, 11.85, 12.48, 13.09, 13.716, 14.326, 14.92, 15.519, 16.06, 16.629]
predictionconstantmono=[1.43, 1.43, 1.43, 1.43, 1.43, 1.45, 1.47, 1.47, 1.46, 1.46, 1.47, 1.44, 1.43, 1.43, 1.43, 1.45, 1.45, 1.44, 1.42, 1.41, 1.4, 1.41, 1.4, 1.4, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39]
predictionconstantbis=[1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.2, 1.2, 1.21, 1.2, 1.21, 1.19, 1.17, 1.18, 1.18, 1.18, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17]
predictionconstanttri=[1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1, 1.025, 1.05, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025]
predictionconstanttetra=[0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72]
predictionconstantpenta=[1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04]
predictionconstanthexa=[0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73]
predictionconstant=[]
predictionconstant=predictionconstant+predictionconstantsat
predictionconstant=predictionconstant+predictionconstantmono
predictionconstant=predictionconstant+predictionconstantbis
predictionconstant=predictionconstant+predictionconstanttri
predictionconstant=predictionconstant+predictionconstanttetra
predictionconstant=predictionconstant+predictionconstantpenta
predictionconstant=predictionconstant+predictionconstanthexa
#print(predictionconstant[9])
pr=0
go=1
ndb=0
nc=4
while go==1:
	cfaid=str()
	if nc<10:
		cfaid=cfaid+'0'+str(nc)+':'+str(ndb)
	else:
		cfaid=cfaid+str(nc)+':'+str(ndb)
	predictedrtindex.append(cfaid)
	# begin make prediction
	if ndb==0:
		predicted=setpoint+(slope*(float(predictionconstant[pr])))
		if predicted<0.05:
			predicted=0.05
		if predicted>rettimecutoff:
			predicted=rettimecutoff-2.5
		predictedrtexact.append(predicted)
		if (0.15*predicted)>1.0:
			prtmin=predicted-(0.15*predicted)
			if prtmin<0.05:
				prtmin=0.05
			prtmax=predicted+(0.15*predicted)
		else:
			prtmin=predicted-(1.0)
			if prtmin<0.05:
				prtmin=0.05
			prtmax=predicted+(1.0)
		if prtmin>rettimecutoff:
			prtmin=rettimecutoff-5
		if prtmax>rettimecutoff:
			prtmax=rettimecutoff-0.05
		predictedrtmin.append(prtmin)
		predictedrtmax.append(prtmax)
	else:
		predicted=(float(predictedrtexact[pr-37]))-(slope*(float(predictionconstant[pr])))
		if predicted<0.05:
			predicted=0.05
		if predicted>rettimecutoff:
			predicted=rettimecutoff-2.5
		predictedrtexact.append(predicted)
		if (0.15*predicted)>1.0:
			prtmin=predicted-(0.15*predicted)
			if prtmin<0.05:
				prtmin=0.05
			prtmax=predicted+(0.15*predicted)
		else:
			prtmin=predicted-(1.0)
			if prtmin<0.05:
				prtmin=0.05
			prtmax=predicted+(1.0)
		if prtmin>rettimecutoff:
			prtmin=rettimecutoff-5
		if prtmax>rettimecutoff:
			prtmax=rettimecutoff-0.05
		predictedrtmin.append(prtmin)
		predictedrtmax.append(prtmax)
	# end make prediction
	if nc<40:
		go=1
	else:
		nc=3
		if ndb<6:
			ndb=ndb+1
		else:
			go=0
	nc=nc+1
	pr=pr+1
# end read info from workflow parameters and use, do not request new, make indexed RT prediction of range, calc +- 1 min or +- 15% of RT
print('Setpoint:')
print(setpoint)
print('Slope:')
print(slope)
###
#print(predictedrtindex)				### Write predicted range information in file to use in precursor analysis as RT limitation
#print(len(predictedrtindex))
#print(predictedrtexact)
#print(len(predictedrtexact))
#print(predictedrtmin)
#print(predictedrtmax)
#quit()

# begin write results of retention time range prediction in file
prpwritelist=[]
prpwritelist.append(predictedrtindex)
prpwritelist.append(predictedrtexact)
prpwritelist.append(predictedrtmin)
prpwritelist.append(predictedrtmax)
prpdf=pd.DataFrame(prpwritelist).transpose()
prpdf.columns=['FA', 'RT_predicted', 'RT_min_predicted', 'RT_max_predicted']
filename='Predicted_RT_range.csv'
prpdf.to_csv(filename, index=False)
print('List with predicted retention time ranges is saved as Predicted_RT_range.csv')

#prp=0
#while prp<len(predictedrtindex):
#	ok=1

# end write results of retention time range prediction in file



nchunks=1
# begin calculate monounsaturated precursors from input
beforeall=datetime.datetime.now()
print('Workflow is running ...')
lfa=maxlenfa-minlenfa+1
moleculegrouplist=[]
precursornamelist=[]
precursorformulalist=[]
precursoradductlist=[]
precursormzlist=[]
precursorchargelist=[]
productnamelist=[]
productformulalist=[]
productadductlist=[]
productmzlist=[]
productchargelist=[]
molg=fourlettcode+'_FA'
padd='[M]1+'
prdn='precursor'
li=0
while li<lfa:
	moleculegrouplist.append(molg)
	clfa=li+minlenfa
	currentlfa=str(clfa)
	if len(currentlfa)<2:
		currentlfa='0'+currentlfa
		currentlfa=str(currentlfa)
	pnm=fourlettcode+'_'+currentlfa+':1'
	precursornamelist.append(pnm)
	prf=''
	currentcderiv=0
	currenthderiv=0
	currentdderiv=0
	currentnderiv=0
	currentoderiv=0
	currentpderiv=0
	currentideriv=0
	if cderiv>0:
		currentcderiv=cderiv+int(currentlfa)
		prf=prf+'C'+str(currentcderiv)
	if hderiv>0:
		currenthderiv=hderiv+(2*int(currentlfa))-3
		prf=prf+'H'+str(currenthderiv)
	if dderiv>0:
		currentdderiv=dderiv
		prf=prf+"H'"+str(currentdderiv)
	if nderiv>0:
		currentnderiv=nderiv
		prf=prf+'N'+str(nderiv)
	if oderiv>(-1):
		currentoderiv=oderiv+1
		prf=prf+'O'+str(currentoderiv)
	if pderiv>0:
		currentpderiv=pderiv
		prf=prf+'P'+str(pderiv)
	if ideriv>0:
		currentideriv=ideriv
		prf=prf+'I'+str(ideriv)
	precursorformulalist.append(prf)
	productformulalist.append(prf)
	precursoradductlist.append(padd)
	prmz=imass[0]*currenthderiv+imass[1]*currentdderiv+imass[2]*currentcderiv+imass[3]*currentnderiv+imass[4]*currentoderiv+imass[5]*currentpderiv+imass[10]*currentideriv
	precursormzlist.append(prmz)
	productmzlist.append(prmz)
	precursorchargelist.append(1)
	productnamelist.append(prdn)
	productadductlist.append(padd)
	productchargelist.append(1)
	li=li+1
fareadlist=[]
fareadlist.append(moleculegrouplist)
fareadlist.append(precursornamelist)
fareadlist.append(precursorformulalist)
fareadlist.append(precursoradductlist)
fareadlist.append(precursormzlist)
fareadlist.append(precursorchargelist)
fareadlist.append(productnamelist)
fareadlist.append(productformulalist)
fareadlist.append(productadductlist)
fareadlist.append(productmzlist)
fareadlist.append(productchargelist)
# end calculate monounsaturated precursors from input
preconly=1
#minrt=0.0
#print(fareadlist[1])
#maxrt=eval(input('At which retention time [min] ends gradient? :'))
#explrt=(minrt+maxrt)/2
#explrtw=explrt-minrt
nspec=3	# number of species: precursor, aldehyde, criegee
# begin create empty lists
toprow=[]
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]		# double bond position list, contains strings with double bond position n-1, n-2 ...
dbindexlist=[]	# double bond index list, contains index for the double bond position that is closest to head group of FA
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
toprow=['Moleculegroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
ki=len(fareadlist[0])
satlist=fareadlist	#create lists for saturated FAs

writelist=[]
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
writelist.append(mlistname)
writelist.append(precname)
writelist.append(precformula)
writelist.append(precadduct)
writelist.append(precmz)
writelist.append(precchrg)
writelist.append(prodname)
writelist.append(prodformula)
writelist.append(prodadduct)
writelist.append(prodmz)
writelist.append(prodchrg)

#print('begin add saturated FAs')
osatlist=satlist
r=0
while r<(len(satlist[1])):
	#satlist[1][r][len(satlist[1][r])-1]='0'	
	precn=''
	p=0
	while p<(len(satlist[1][r])-1):
		precn=precn+satlist[1][r][p]
		p=p+1
	precn=precn+'0'					#PrecursorName to saturated FA
	satlist[1][r]=precn
	satlist[4][r]=float(satlist[4][r])+(2*imass[0])	#PrecursorMz to sat. FA
	satlist[9][r]=float(satlist[9][r])+(2*imass[0])	#ProductMz to sat. FA
	hcurr=(10*(int(satlist[2][r][4])))+(int(satlist[2][r][5]))
	hcurr=str(hcurr+2)
	p=0
	precf=''
	while p<4:
		precf=precf+satlist[2][r][p]
		p=p+1
	precf=precf+hcurr
	p=p+2
	while p<(len(satlist[2][r])):
		precf=precf+satlist[2][r][p]
		p=p+1
	satlist[2][r]=str(precf)	#PrecursorFormula is edited
	satlist[7][r]=str(precf)	#ProductFormula is edited
	r=r+1
pnm=0
while pnm<len(satlist[0]):
	satlist[6][pnm]=str(satlist[1][pnm])+'_precursor'
	pnm=pnm+1
clmn=0
while clmn<(len(satlist)):
	writelist[clmn]=writelist[clmn]+satlist[clmn]
	clmn=clmn+1
# end add saturated FAs

addunsat=1
if addunsat==1:
	dby=1
	while dby<7: #7:
		msatlist=osatlist
		#print('begin add unsaturated FAs')
		r=0
		while r<(len(osatlist[1])):
			#satlist[1][r][len(satlist[1][r])-1]='0'	
			precn=''
			p=0
			while p<(len(osatlist[1][r])-1):
				precn=precn+osatlist[1][r][p]
				p=p+1
			precn=precn+str(dby)					#PrecursorName to saturated FA
			msatlist[1][r]=precn
			msatlist[4][r]=float(osatlist[4][r])-(2*imass[0])	#PrecursorMz to sat. FA
			msatlist[9][r]=float(osatlist[9][r])-(2*imass[0])	#ProductMz to sat. FA
			hcurr=(10*(int(osatlist[2][r][4])))+(int(osatlist[2][r][5]))
			hcurr=str(hcurr-(2))
			p=0
			precf=''
			while p<4:
				precf=precf+satlist[2][r][p]
				p=p+1
			precf=precf+hcurr
			p=p+2
			while p<(len(osatlist[2][r])):
				precf=precf+osatlist[2][r][p]
				p=p+1
			msatlist[2][r]=str(precf)	#PrecursorFormula is edited
			msatlist[7][r]=str(precf)	#ProductFormula is edited
			r=r+1
		pnm=0
		while pnm<len(satlist[0]):
			msatlist[6][pnm]=str(osatlist[1][pnm])+'_precursor'
			pnm=pnm+1
		clmn=0
		while clmn<(len(satlist)):
			writelist[clmn]=writelist[clmn]+osatlist[clmn]
			clmn=clmn+1
		dby=dby+1
	# end add unsaturated FAs


##########################################################VIRTUAL#PRECURSOR###########################################################################
######################################################################################################################################################
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]		# content in lists is deleted to free up memory
# begin introduce virtual precursor
ki=len(writelist[0])

#print('Entries in excel file after all FA transitions are generated: %d' % ki)
virtualprecformula=[]
virtualprecmz=[]

vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]

r=0
while r<ki:
	e=str(writelist[2][r])#sheetinput.cell(row=r, column=3)	# PrecursorFormula
	ve=e+'Xe'
	virtualprecformula.append(ve) # virtual precursorFormula
	e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
	ve=e+(imass[9])		#add one Xe atom to generate virtual precursor
	virtualprecmz.append(ve)	# virtual precursormz
	e=str(writelist[0][r])#sheetinput.cell(row=r, column=1)
	vmlistname.append(e)
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)
	vprecname.append(e)
	e=str(writelist[3][r]) #sheetinput.cell(row=r, column=4)
	vprecadduct.append(e)
	e=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)
	vprecchrg.append(e)
	e=str(writelist[6][r]) #sheetinput.cell(row=r, column=7)
	vprodname.append(e)
	e=str(writelist[7][r]) #sheetinput.cell(row=r, column=8)
	vprodformula.append(e)
	e=str(writelist[8][r]) #sheetinput.cell(row=r, column=9)
	vprodadduct.append(e)
	e=float(writelist[9][r]) #sheetinput.cell(row=r, column=10)
	vprodmz.append(e)
	e=int(writelist[10][r]) #sheetinput.cell(row=r, column=11)
	vprodchrg.append(e)
	r=r+1
# begin save excel file with virtual precursor as csv file
toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
vwritelist=[]
vwritelist.append(vmlistname)
vwritelist.append(vprecname)
vwritelist.append(virtualprecformula)
vwritelist.append(vprecadduct)
vwritelist.append(virtualprecmz)
vwritelist.append(vprecchrg)
vwritelist.append(vprodname)
vwritelist.append(vprodformula)
vwritelist.append(vprodadduct)
vwritelist.append(vprodmz)
vwritelist.append(vprodchrg)
terminate=0
if terminate==1:
	print('writelist created')
	transitionresultsdf=pd.DataFrame(vwritelist).transpose()
	print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10]]
	print('Transposed and DataFrame created')
	transitionresultsdf.to_csv('jpmlipidomics_vpw20_virtual_precursor.csv', index=False)
	print('Transition list saved as jpmlipidomics_vpw20_virtual_precursor.csv')
# end save excel file with virtual precursor as csv file

#print('Saving data.')
#print('len(vmlistname)')
#print(len(vmlistname))
virtualprecformula=[]
virtualprecmz=[]
vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[] # test to delete lists to save up memory

if checkup==1:
	print('Transition list is modified with virtual precursor [M + Xe].')

#terminate=1
if terminate==1:
	quit()
if checkup==1:
	print('reduce and expand')
#################################################################################################################################################
############################# BEGIN REDUCE TO ONE ENTRY PER PRECURSOR AND EXPAND WITH VARIED EXPLICIT RETENTION TIME ############################
#################################################################################################################################################
#terminate=eval(input('Reduce transition list to one entry per precursor and expand with varied explicit retention time? Yes: 1 No: 0 ::'))
terminate=1
if terminate==0:
	quit()
else:
	#exrtstep=eval(input('Stepwidth for varied explicit retention time (e.g. 0.05 min)? ::'))
	#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
	normalexrtstep=0.027
	bigexrtstep=0.045
	if widestep==1:
		normalexrtstep=0.1
		bigexrtstep=0.2
	#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
# begin reduce and expand
ki=len(vwritelist[0])
# end count entries in excel file
vmlistname=[]
vprecname=[]
virtualprecformula=[]
vprecadduct=[]
virtualprecmz=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]
precrt=[]
rtwindow=[]
r=0
ki=ki
pos=1
while r<ki:
	if pos==1:
		if rtlimitation==1:
			# begin redefine firstexrt as cfirstexrt and lastexrt as clastexrt depending on chainlength of current FA
			e=(10*int(vwritelist[1][r][5]))+int(vwritelist[1][r][6])
			e=int(e)
			if e>15:
				if e<30:
					cfirstexrt=firstexrt+(((lastexrt-firstexrt)*(1/4))*(((e-15)/15)))
				else:
					cfirstexrt=firstexrt+(((lastexrt-firstexrt)*(1/4))+((lastexrt-firstexrt)*(1/4))*(((e-30)/10)))
				cfirstexrt=round(cfirstexrt, 3)
			else:
				cfirstexrt=firstexrt
			if e<20:
				if e<8:
					clastexrt=lastexrt-(((lastexrt-firstexrt)*(1/8))+((lastexrt-firstexrt)*(1/8))*(((8-e)/4)))
				else:
					clastexrt=lastexrt-(((lastexrt-firstexrt)*(1/8))*(((20-e)/12)))
				clastexrt=round(clastexrt, 3)
			else:
				clastexrt=lastexrt
			# end redefine firstexrt as cfirstexrt and lastexrt as clastexrt depending on chainlength of current FA
		elif rtlimitation==2:
			# begin set cfirstexrt according to predicted RT range start (predictedrtmin) and clastexrt / predictedrtmax // predictedrtindex
			cfaspec=str(vwritelist[1][r][5])+str(vwritelist[1][r][6])+str(vwritelist[1][r][7])+str(vwritelist[1][r][8])
			pr=0
			while pr<(len(predictedrtindex)):
				if cfaspec==str(predictedrtindex[pr]):
					cfirstexrt=predictedrtmin[pr]
					clastexrt=predictedrtmax[pr]
				pr=pr+1
			# end set cfirstexrt according to predicted RT range start (predictedrtmin) and clastexrt / predictedrtmax // predictedrtindex
		else:
			cfirstexrt=firstexrt #float(firstexrt)
			clastexrt=lastexrt #float(lastexrt)
			clastexrt=round(clastexrt, 3)
			cfirstexrt=round(cfirstexrt, 3)
		#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
		if int(vwritelist[1][r][8])>2:
			nca=10*int(vwritelist[1][r][5])+int(vwritelist[1][r][6])
			if nca>17:
				exrt=cfirstexrt+bigexrtstep
			else:
				exrt=cfirstexrt+normalexrtstep
		else:	
			exrt=cfirstexrt+normalexrtstep		# firstexrt is the first explicit retention time that is set to look for species (e.g. 1.50 min)
		#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
		exrtstr=str(exrt)
		while exrt<clastexrt:		# lastexrt is the last explicit retention time that is set to look for species (e.g. 18.00 min)
			ve=str(vwritelist[2][r]) #sheetinput.cell(row=r, column=3)	# PrecursorFormula
			virtualprecformula.append(ve) # virtual precursorFormula
			ve=float(vwritelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz	
			virtualprecmz.append(ve)	# virtual precursormz
			e=str(vwritelist[0][r]) #sheetinput.cell(row=r, column=1)
			vmlistname.append(e)
			e=str(vwritelist[1][r]) #sheetinput.cell(row=r, column=2)
			cm='_'
			cm=str(cm)
			exrtstr=str(round(exrt, 2))		#
			ee=e+cm+exrtstr			#
			vprecname.append(ee)
			precrt.append(exrt)
			e=str(vwritelist[3][r]) #sheetinput.cell(row=r, column=4)
			vprecadduct.append(e)
			e=int(vwritelist[5][r]) #sheetinput.cell(row=r, column=6)
			vprecchrg.append(e)
			e=str(vwritelist[6][r]) #sheetinput.cell(row=r, column=7)
			vprodname.append(e)
			e=str(vwritelist[7][r]) #sheetinput.cell(row=r, column=8)
			vprodformula.append(e)
			e=str(vwritelist[8][r]) #sheetinput.cell(row=r, column=9)
			vprodadduct.append(e)
			e=float(vwritelist[9][r]) #sheetinput.cell(row=r, column=10)
			vprodmz.append(e)
			e=int(vwritelist[10][r]) #sheetinput.cell(row=r, column=11)
			vprodchrg.append(e)	
			#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
			if int(vwritelist[1][r][8])>2:
				nca=10*int(vwritelist[1][r][5])+int(vwritelist[1][r][6])
				if nca>17:
					exrt=exrt+bigexrtstep
					exrtstep=bigexrtstep
				else:
					exrt=exrt+normalexrtstep
					exrtstep=normalexrtstep	
			else:	
				exrt=exrt+normalexrtstep
				exrtstep=normalexrtstep
			#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain	
			e=exrtstep*2
			rtwindow.append(e)
			#print(exrt)
			#print(clastexrt)
			exrt=clastexrt		##### deactivate this line to revert to RT variation (expansion of transition list by varied RT within predicted range) ######## SWITCH #######
	vprecmz=vwritelist[4][r] #sheetinput.cell(row=r, column=5)	# PrecursorMz
	r=r+1
	corr=0
	if r==ki:
		r=r-1
		corr=1
	e=vwritelist[4][r] #sheetinput.cell(row=r, column=5)	# PrecursorMz
	if e==vprecmz:
		pos=0
	else:
		pos=1
	r=r+corr
	
# begin remove H'0 from formula, where applicable
#print("Start removing H'0")
pfi=0
while pfi<len(virtualprecformula):
	if "H'0" in str(virtualprecformula[pfi]):
		#remove H'0 from formula
		ri=virtualprecformula[pfi].index("H'0")
		virtualprecformula[pfi]=virtualprecformula[pfi][0:ri:]+virtualprecformula[pfi][ri+3::]
	if "H'0" in str(vprodformula[pfi]):
		#remove H'0 from formula
		ri=vprodformula[pfi].index("H'0")
		vprodformula[pfi]=vprodformula[pfi][0:ri:]+vprodformula[pfi][ri+3::]
	pfi=pfi+1
#print("End removing H'0")
# end remove H'0 from formula, where applicable

# begin save to csv file
prt='PrecursorRT'
prt=str(prt)
toprow.append(prt)
rtw='PrecursorRTWindow'
rtw=str(rtw)
toprow.append(rtw)
ki=2+len(vmlistname)

vwritelist=[]
vwritelist.append(vmlistname)
vwritelist.append(vprecname)
vwritelist.append(virtualprecformula)
vwritelist.append(vprecadduct)
vwritelist.append(virtualprecmz)
vwritelist.append(vprecchrg)
vwritelist.append(vprodname)
vwritelist.append(vprodformula)
vwritelist.append(vprodadduct)
vwritelist.append(vprodmz)
vwritelist.append(vprodchrg)
vwritelist.append(precrt)
vwritelist.append(rtwindow)
#print('writelist created')
transitionresultsdf=pd.DataFrame(vwritelist).transpose()
#print('Transposed')
transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
#print('Transposed and DataFrame created')
after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_1_'+fourlettcode+'_'
#filename=today+'jpmlipidomics_vpw13_1_precursor.csv'
filename='jpmlipidomics_vpw20_1_precursor.csv'
transitionresultsdf.to_csv(filename, index=False)
afterall=datetime.datetime.now()
dt=afterall-beforeall
nrows=len(vmlistname)
#print('Transition list is saved as yyyy_mm_dd_1_xxxx_jpmlipidomics_vpw13_1_precursor.csv (%d rows)' % nrows)
print('Transition list is saved as jpmlipidomics_vpw20_1_precursor.csv (%d rows)' % nrows)
print('Calculation time (h:mm:ss) is:')
print(dt)
# end save to csv file
#################################################################################################################################################
############################### END REDUCE TO ONE ENTRY PER PRECURSOR AND EXPAND WITH VARIED EXPLICIT RETENTION TIME ############################
#################################################################################################################################################


