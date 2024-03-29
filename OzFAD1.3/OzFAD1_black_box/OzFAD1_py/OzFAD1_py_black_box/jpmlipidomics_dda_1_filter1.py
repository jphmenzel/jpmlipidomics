# -*- coding: UTF-8 -*-

# Jan Philipp Menzel 
# Goal: Filter transition list
## Notes: Derivative, positive fixed charge
## NOTES: VIRTUAL PRECURSOR - PrecursorName and PrecursorMz are artificially set +Xe (only column 3 and 5), fragment transitions correct including precursor
## NOTES: Virtual precursor forces Skyline to consider all transitions incl. real precursor (fragment in transition list), Skyline Setting: TransitionSettings-Filter-IonTypes-f 
import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
################ DATABASE ## Source: Internetchemie.info 
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.757, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
################
#default=eval(input('Run workflow with default parameters? (Yes: 1 / No: 0) (Derivatization agent: AMPP; Slow and full workflow including all FA; Apply retention time limitation; Use Fatty Acid Library to prevent filtering out important FA; Max RT = 17.5 min; FA: 12 - 24 C; Max m/z error: 10 ppm; Precursor peak area threshold: 3000; Product peak area threshold: 200.) :'))
default=0
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
###########

########### begin read workflow parameters
transferlist=[]
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
go=1
while go==1:
	tfe=ws.cell(row=tli+1, column=2)
	tfe=tfe.value
	if tli==21:
		go=0
	if tfe is None:
		go=0
	else:
		transferlist.append(tfe)
	tli=tli+1
########## 
fourlettcode=transferlist[0]		#input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
cderiv=int(transferlist[1])
hderiv=int(transferlist[2])
dderiv=int(transferlist[3])
nderiv=int(transferlist[4])
oderiv=int(transferlist[5])
pderiv=int(transferlist[6])
ideriv=int(transferlist[7])
minlenfa=int(transferlist[8])
maxlenfa=int(transferlist[9])
mzcutoff=int(transferlist[10]) 		#eval(input('What is max mz [ppm] for positive identification of species? (e.g. 30) :'))
rettimecutoff=float(transferlist[11]) 		#eval(input('What is the maximum retention time [min]? (e.g. 11.9 or 17.8) :'))
areathreshold=int(transferlist[12])		#eval(input('What is the threshold for detected precursor peak area? (compare Skyline report file, e.g. 3000) :'))
workflowmode=int(transferlist[13])
productareathreshold=int(transferlist[14])
nchunks=int(transferlist[15])
rtlimitation=int(transferlist[16])
mostwanted=int(transferlist[17])
transtest=int(transferlist[18])
runprecheck=int(transferlist[19])
workflowidentifier=str(transferlist[20])
########### end read workflow parameters
# begin adjust workflow parameters for DDA analysis
mzcutoff=60
precareathreshold=50
prodareathreshold=50
# end adjust workflow parameters for DDA analysis
# begin build DDA identifier (replace AI or DIA with DDA in identifier YYYY_MM_DD_DIA_SAMPLE_NAME)
wfidentifierdda=str()
wfi=0
while wfi<11:
	wfidentifierdda=wfidentifierdda+workflowidentifier[wfi]
	wfi=wfi+1
wfidentifierdda=wfidentifierdda+'DDA'
if str(workflowidentifier[13])=='_':
	wfi=13
elif str(workflowidentifier[14])=='_':
	wfi=14
while wfi<len(workflowidentifier):
	wfidentifierdda=wfidentifierdda+workflowidentifier[wfi]
	wfi=wfi+1
# end build DDA identifier (replace AI with DDA)

# begin calculate monounsaturated precursors from input
beforeall=datetime.datetime.now()
#print('Workflow is running ...')

# begin build mostwantedlist 
mostwantedlist=[]
wbfa=openpyxl.load_workbook('jpm_fa_lib.xlsx')
wsfa=wbfa.active
mwi=0
go=1
while go==1:
	tfe=wsfa.cell(row=mwi+3, column=1)
	tfe=tfe.value
	if tfe is None:
		go=0
	else:
		if int(tfe)<10:
			tfe='0'+str(tfe)
		mwfa=str(tfe)+':'
		tfe=wsfa.cell(row=mwi+3, column=2)
		tfe=tfe.value
		dbfa=int(tfe)
		if int(tfe)==0:
			mwfa=mwfa+str(tfe)
		else:
			mwfa=mwfa+str(tfe)+'_n-'
			dbfai=1
			while dbfai<(dbfa+1):
				tfe=wsfa.cell(row=mwi+3, column=dbfai+2)
				tfe=tfe.value
				if dbfai<(dbfa):
					mwfa=mwfa+str(tfe)+'_n-'
				else:
					mwfa=mwfa+str(tfe)
				dbfai=dbfai+1
		mostwantedlist.append(mwfa)
	mwi=mwi+1
# end build mostwantedlist
#print('Arrived here.#################################################################################################################################')
#quit()

# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
diagnostics=[]
#end create empty lists


#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists
trdf=pd.read_csv('skyl_report_dda_vpw20_0.csv')
toprowx=[trdf.columns.values.tolist()]
toprow=toprowx[0]
trdf=trdf.transpose()
writelist=trdf.values.tolist()
ki=len(writelist[0])
print('Number of rows in skyl_report_dda_vpw20_0.csv: %d' % ki)
#print(ki)
#####################################################################################################
# begin assign whether transitions diagnostic or non-diagnostic
r=0
while r<ki:
	t=r
	if int(writelist[1][t][8])>2:
		go=1
		ch=len(writelist[1][t])-1
		while go==1:
			if str(writelist[1][t][ch])=='n':
				go=0
				if str(writelist[1][t][ch+4])=='_':
					lastdb=10*int(writelist[1][t][ch+2])+int(writelist[1][t][ch+3])		# determine last db (closest to amide bond)
				else:
					lastdb=int(writelist[1][t][ch+2])
				if str(writelist[1][t][ch-4])=='-':	
					seclastdb=10*int(writelist[1][t][ch-3])+int(writelist[1][t][ch-2])	# determine second last db (second closest to amide bond)
				else:
					seclastdb=int(writelist[1][t][ch-2])
			else:
				go=1
			ch=ch-1
		if lastdb>13:
			if lastdb-seclastdb==3:
				if str(writelist[6][t][len(writelist[6][t])-1])=='r':
					diagnostics.append('diagnostic')
				else:
					if str(writelist[6][t][len(writelist[6][t])-4])=='n':
						if lastdb==10*int(writelist[6][t][len(writelist[6][t])-2])+int(writelist[6][t][len(writelist[6][t])-1]):
							diagnostics.append('non-diagnostic')
						else:
							diagnostics.append('diagnostic')
					else:
						if lastdb==int(writelist[6][t][len(writelist[6][t])-1]):
							diagnostics.append('non-diagnostic')
						else:
							diagnostics.append('diagnostic')
			else:
				diagnostics.append('diagnostic')
		else:
			diagnostics.append('diagnostic')
	else:
		diagnostics.append('diagnostic')
	r=r+1
# end assign whether transitions diagnostic or non-diagnostic
####################################################################################################
r=0
ki=ki
while r<ki:		# go through rows of excel file 
	currentmzerror=[]
	currentrettime=[]
	currentareas=[]
	currentprodmz=[]
	currentprodnm=[]
	e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# Precursorname		# begin determine which row to start (r) and to end (s)
	s=r+1
	st=0
	while st<1:
		if s>(len(writelist[1])-1):
			ne='stop_loop'
		else:
			ne=writelist[1][s] #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
	t=r
	pos=0
	fpos=0
	gpos=0
	prodmzcutoff=mzcutoff#*3.5
	while t<(s+1):
		e=str(writelist[11][t]) ## mzerror
		if e=='nan':
			if str(diagnostics[t])=='diagnostic':
				pos=1
		else:
			e=float(writelist[11][t]) ## mzerror
			currentmzerror.append(e)
			prodnm=str(writelist[6][t]) ## begin mzcutoff for precursors and transitions from five db with lowest n
			requireddb=5			# required number of double bonds to be tested for cutoff condition #################################################
			if prodnm[(len(prodnm)-1)]=='r':  # not enabled: use normal mzcutoff for any precursor
				if abs(e)>mzcutoff:
					if str(diagnostics[t])=='diagnostic':
						pos=pos
			elif (int(prodnm[8]))<(requireddb+1):	# use 2.5 times normal mzcutoff for any product transition from FA with less than 6 double bonds
				if abs(e)>(prodmzcutoff):
					if str(diagnostics[t])=='diagnostic':
						pos=1
			else:
				if prodnm[len(prodnm)-3]=='-':
					currentdb=10*(int(prodnm[len(prodnm)-2]))+(int(prodnm[len(prodnm)-1]))
				elif prodnm[len(prodnm)-2]=='-':
					currentdb=(int(prodnm[len(prodnm)-1]))
				else:
					if prodnm[13]=='_':
						currentdb=(int(prodnm[12]))
					else:
						currentdb=10*(int(prodnm[12]))+(int(prodnm[13]))
				searchdb=0
				go=0
				while go<requireddb: #requireddb
					if (str(prodnm[searchdb]))=='-':
						go=go+1
						if (str(prodnm[searchdb+3]))=='_':
							checkdb=(int(prodnm[searchdb+1])*10)+(int(prodnm[searchdb+2]))
						elif (str(prodnm[searchdb+2]))=='_':
							checkdb=(int(prodnm[searchdb+1]))
						if checkdb==currentdb:
							if go<3:
								if abs(e)>prodmzcutoff:
									if str(diagnostics[t])=='diagnostic':
										pos=1
							elif go<4:
								if abs(e)>(prodmzcutoff+3):
									if str(diagnostics[t])=='diagnostic':
										pos=1
							else:
								if abs(e)>(prodmzcutoff+10):	# use wider mzcutoff for any transition from FA with 6 double bonds from 5th or 6th db
									if str(diagnostics[t])=='diagnostic':
										pos=1
					searchdb=searchdb+1		# end mzcutoff for precursors and transitions from three db with lowest n

			f=float(writelist[17][t]) ## explicit retention time
			currentrettime.append(f)
			if f>rettimecutoff:
				fpos=0 #1
			g=float(writelist[13][t]) ## area index 13 (column 14) 
			#print(g)
			prodnm=str(writelist[6][t]) ## begin area cutoff for precursors and transitions from five db with lowest n
			currentprodnm.append(prodnm)
			requireddb=5			# required number of double bonds to be tested for cutoff condition #################################################
			if prodnm[(len(prodnm)-1)]=='r':
				if g<precareathreshold: 	# precursor area CUTOFF
					if mostwanted==0:
						gpos=1
					else:
						# begin test whether species to be cut is in the mostwantedlist (if it is, keep only if area higher than half the areacutoff)
						fashort=str()
						fas=5
						while fas<(len(writelist[1][t])-5):
							fashort=fashort+str(writelist[1][t][fas])
							fas=fas+1
						mwl=0
						gposcancel=0
						while mwl<(len(mostwantedlist)):
							if mostwantedlist[mwl]==fashort:
								gposcancel=1
								if g<(precareathreshold*0.5):		# use half the area threshold for mostwanted species
									gpos=1
								elif writelist[6][t][8]==0:			# use normal area threshold for all saturated FA
									gpos=1
							mwl=mwl+1
						if gposcancel==0:
							gpos=1
						# end test whether species to be cut is in the mostwantedlist
					#gpos=1
			elif (int(prodnm[8]))<(requireddb+1):
				if g<prodareathreshold:		# product area CUTOFF (for species with 1 or 2 or 3 db)
					if str(diagnostics[t])=='diagnostic':
						if mostwanted==0:
							gpos=1
						else:
							# begin test whether species to be cut is in the mostwantedlist
							fashort=str()
							fas=5
							while fas<(len(writelist[1][t])-5):
								fashort=fashort+str(writelist[1][t][fas])
								fas=fas+1
							mwl=0
							gposcancel=0
							while mwl<(len(mostwantedlist)):
								if mostwantedlist[mwl]==fashort:
									#print('!!! FOUND ONE !!!')
									#print(writelist[6][t])
									#print('!!! FOUND ONE !!!')
									gposcancel=1
								mwl=mwl+1
							if gposcancel==0:
								gpos=1
							# end test whether species to be cut is in the mostwantedlist
						#gpos=1
			else:
				if prodnm[len(prodnm)-3]=='-':
					currentdb=10*(int(prodnm[len(prodnm)-2]))+(int(prodnm[len(prodnm)-1]))
				elif prodnm[len(prodnm)-2]=='-':
					currentdb=(int(prodnm[len(prodnm)-1]))
				else:
					if prodnm[13]=='_':
						currentdb=(int(prodnm[12]))
					else:
						currentdb=10*(int(prodnm[12]))+(int(prodnm[13]))
				searchdb=0
				go=0
				while go<requireddb:
					if (str(prodnm[searchdb]))=='-':
						go=go+1
						if (str(prodnm[searchdb+3]))=='_':
							checkdb=(int(prodnm[searchdb+1])*10)+(int(prodnm[searchdb+2]))
						elif (str(prodnm[searchdb+2]))=='_':
							checkdb=(int(prodnm[searchdb+1]))
						if checkdb==currentdb:
							if g<(prodareathreshold/3):		# product area CUTOFF (past third db)
								if str(diagnostics[t])=='diagnostic':
									if mostwanted==0:
										gpos=1
									else:
										# begin test whether species to be cut is in the mostwantedlist
										fashort=str()
										fas=5
										while fas<(len(writelist[1][t])-5):
											fashort=fashort+str(writelist[1][t][fas])
											fas=fas+1
										mwl=0
										gposcancel=0
										while mwl<(len(mostwantedlist)):
											if mostwantedlist[mwl]==fashort:
												gposcancel=1
											mwl=mwl+1
										if gposcancel==0:
											gpos=1
										# end test whether species to be cut is in the mostwantedlist
					searchdb=searchdb+1		# end area cutoff for precursors and transitions from three db with lowest n
			#currentareas.append(g)
			g=float(writelist[9][t]) #sheetinput.cell(row=t, column=10)	# prodmz
			currentprodmz.append(g)
		t=t+1
	# begin use current area and prodmz to determine, whether current species has a precursor abundance that is lower than all fragments (criterium for exclusion)
	if len(currentareas)>1:		#only exclude unsaturated FAs, keep saturated FAs 
		ci=0
		call=0
		while call<1:
			ciprodmz=currentprodmz[ci]
			cj=0
			cc=0
			ct=1
			while cj<len(currentprodmz):
				cjprodmz=currentprodmz[cj]
				if ciprodmz>(cjprodmz-0.001):
					cc=1
				else:
					ct=0
				cj=cj+1
			if ct==1:
				call=1
				cmax=ci
			else:
				ci=ci+1		# species with highest mz is identified
		cmax=len(currentareas)-1		# assume that last species is precursor (disable this line, if not the case!!!)
		# cmax is index for species with highest mass (precursor)
		ciareas=currentareas[cmax]
		cj=0
		cc=0
		ct=1
		while cj<len(currentareas):
			cjareas=currentareas[cj]
			if ciareas<(cjareas+0.001):
				cc=1
			else:
				ct=0
			cj=cj+1
		if ct==1:
			if selectiontype==1:
				if mostwanted==0:
					gpos=1
				else:
					# begin test whether species to be cut is in the mostwantedlist
					fashort=str()
					fas=5
					while fas<(len(writelist[1][t])-5):
						fashort=fashort+str(writelist[1][t][fas])
						fas=fas+1
					mwl=0
					gposcancel=0
					while mwl<(len(mostwantedlist)):
						if mostwantedlist[mwl]==fashort:
							gposcancel=1
						mwl=mwl+1
					if gposcancel==0:
						gpos=1
					# end test whether species to be cut is in the mostwantedlist
			else:
				nochange=0
	# end use current area and prodmz to determine, whether current species has a precursor abundance that is lower than all fragments (criterium for exclusion)
	# begin use current area and productname to determine, whether current species has for any double bond a criegee area larger than respective aldehyde area
	# 		exclude, if species not in mostwantedlist, if mostwantedlist is in use, otherwise exclude
	cpi=0
	while cpi<(len(currentprodnm)):
		if str(currentprodnm[cpi]).find('aldehyde')==1:
			cpj=cpi
			while cpj<(len(currentprodnm)):
				if (str(currentprodnm[cpi][len(currentprodnm[cpi])-1])+str(currentprodnm[cpi][len(currentprodnm[cpi])-2]))==(str(currentprodnm[cpi][len(currentprodnm[cpi])-1])+str(currentprodnm[cpi][len(currentprodnm[cpi])-2])):
					if (float(currentareas[cpi])*1.05)<(float(currentareas[cpj])):
						if mostwanted==0:
							pos=0 #1
						else:
							# begin test whether species to be cut is in the mostwantedlist
							fashort=str()
							fas=5
							while fas<(len(writelist[1][t])-5):
								fashort=fashort+str(writelist[1][t][fas])
								fas=fas+1
							mwl=0
							gposcancel=0
							while mwl<(len(mostwantedlist)):
								if mostwantedlist[mwl]==fashort:
									gposcancel=1
								mwl=mwl+1
							if gposcancel==0:
								pos=0 #1
							# end test whether species to be cut is in the mostwantedlist
				cpj=cpj+1
		cpi=cpi+1
	# end use current area and productname to determine, whether current species has for any double bond a criegee area larger than respective aldehyde area
	if pos==1:
		r=s+1
	elif fpos==1:
		r=s+1
	elif gpos==1:
		r=s+1
	else:
		t=r
		while t<(s+1):
			#e=sheetinput.cell(row=t, column=19)	# 	begin determine whether duplicate is found, duplicate kept
			#e=e.value
			#rtstartcurrent=e
			#e=sheetinput.cell(row=t, column=20)	# 	
			#e=e.value
			#rtendcurrent=e
			#if ki==0:
			#	apos=1
			#else:
			#	apos=0
				#rtstartprevious=rtstart[(len(rtstart)-1)]
				#rtendprevious=rtend[(len(rtend)-1)]
				#if rtstartprevious==rtstartcurrent:
					#if rtendprevious==rtendcurrent:
						#apos=1			# 	duplicates removed, not kept for later stage of analysis with transitions of fragments
						#apos=0			# 	end determine whether duplicate is found, duplicate kept, all saved in next step
			apos=0
			if apos==0:
				e=writelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving in excel output
				mlistname.append(e)
				e=writelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
				precname.append(e)
				e=writelist[2][t] #sheetinput.cell(row=t, column=3)	# precname	
				precformula.append(e)
				e=writelist[3][t] #sheetinput.cell(row=t, column=4)	# precformula	
				precadduct.append(e)
				e=writelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
				precmz.append(e)
				e=writelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
				precchrg.append(e)
				e=writelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
				prodname.append(e)
				e=writelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
				prodformula.append(e)
				e=writelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
				prodadduct.append(e)
				e=writelist[9][t] #sheetinput.cell(row=t, column=10)	# 
				prodmz.append(e)
				e=writelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
				prodchrg.append(e)
				e=writelist[11][t] #sheetinput.cell(row=t, column=12)	# 	
				mzerror.append(e)
				e=writelist[12][t] #sheetinput.cell(row=t, column=13)	# 	
				rettime.append(e)
				e=writelist[13][t] #sheetinput.cell(row=t, column=14)	# 	
				area.append(e)
				e=writelist[14][t] #sheetinput.cell(row=t, column=15)	# 	convert to float does not work
				areanormalpercent.append(e)
				e=writelist[15][t] #sheetinput.cell(row=t, column=16)	# 	
				background.append(e)
				e=writelist[16][t] #sheetinput.cell(row=t, column=17)	# 	
				fwhm.append(e)
				e=writelist[17][t] #sheetinput.cell(row=t, column=18)	# 	
				explicitrt.append(e)
				e=writelist[18][t] #sheetinput.cell(row=t, column=19)	# 	
				rtstart.append(e)
				e=writelist[19][t] #sheetinput.cell(row=t, column=20)	# 	
				rtend.append(e)
			t=t+1
	r=s+1
# end read transition results and append suitable species to lists
swritelist=[]
swritelist.append(mlistname)
swritelist.append(precname)
swritelist.append(precformula)
swritelist.append(precadduct)
swritelist.append(precmz)
swritelist.append(precchrg)
swritelist.append(prodname)
swritelist.append(prodformula)
swritelist.append(prodadduct)
swritelist.append(prodmz)
swritelist.append(prodchrg)
swritelist.append(mzerror)
swritelist.append(rettime)
swritelist.append(area)
swritelist.append(areanormalpercent)
swritelist.append(background)
swritelist.append(fwhm)
swritelist.append(explicitrt)
swritelist.append(rtstart)
swritelist.append(rtend)

# first filter completed. Begin delete duplicates that are close to the duplicate with max integral
valprecname=[]
valprodname=[]
valexplicitrt=[]
valprodareasum=[]
ki=len(swritelist[0])
#print('Number of entries after first filter step:')
#print(ki)
r=0
while r<ki:		# go through rows of list 
	e=swritelist[1][r] # Precursorname		# begin determine which row to start (r) and to end (s)
	s=r+1
	st=0
	while st<1:
		if s>(len(swritelist[1])-1):
			ne='stop_loop'
		else:
			ne=swritelist[1][s] #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
	# get current sum of product areas
	t=r
	cprodareasum=0
	while t<s:
		cprodareasum=cprodareasum+float(swritelist[13][t])
		t=t+1
	valprodareasum.append(float(cprodareasum))
	valprecname.append(str(swritelist[1][t]))
	valprodname.append(str(swritelist[6][t]))
	valexplicitrt.append(float(swritelist[17][t]))
	r=s+1
# val lists are built as overview of species in list
# compare and delete duplicate species within max +- 0.07 min from swritelist
# determine duplicate with max prod area sum
droplist=[]		# species to be deleted later
rv=0
while rv<(len(valprecname)):
	kv=0
	comparelist=[]
	compareprodarea=[]
	compareexplicitrt=[]
	while kv<(len(valprecname)):
		if rv==kv:
			rv=rv
		else:
			if valprodname[rv]==valprodname[kv]:
				comparelist.append(valprecname[kv])
				compareprodarea.append(valprodareasum[kv])
				compareexplicitrt.append(valexplicitrt[kv])
		kv=kv+1
	if (len(comparelist))>0:
		if valprodareasum[rv]>((max(compareprodarea))-0.1):
			rv=rv
		else:
			# determine kv for max as tv
			tv=0
			while tv<(len(valprecname)):
				if valprodareasum[tv]>((max(compareprodarea))-0.1):
					tvmax=tv
				tv=tv+1
			if abs(float(valexplicitrt[rv])-float(valexplicitrt[tvmax]))<0.08:
				droplist.append(str(valprecname[rv])) #mark this rv for deletion
			elif abs(valprodareasum[rv]-valprodareasum[tvmax])<0.01:
				droplist.append(str(valprecname[rv])) #mark this rv for deletion
	rv=rv+1
# droplist contains precnames of species to be deleted

# add species to droplist that have same prod area

#print(droplist)
ldl=len(droplist)
#print(ldl)
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
#end create empty lists
r=0
ki=len(swritelist[0])
while r<ki:
	e=swritelist[1][r] # Precursorname		# begin determine which row to start (r) and to end (s)
	s=r+1
	st=0
	while st<1:
		if s>(len(swritelist[1])-1):
			ne='stop_loop'
		else:
			ne=swritelist[1][s] #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
	t=r
	q=0
	dpos=0
	while q<(len(droplist)):
		if str(swritelist[1][t])==str(droplist[q]):
			dpos=1#
		else:
			dpos=dpos
		q=q+1
	if dpos==1:
		r=s+1
	else:
		while t<s+1:
			e=swritelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving in excel output
			mlistname.append(e)
			e=swritelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
			precname.append(e)
			e=swritelist[2][t] #sheetinput.cell(row=t, column=3)	# precname	
			precformula.append(e)
			e=swritelist[3][t] #sheetinput.cell(row=t, column=4)	# precformula	
			precadduct.append(e)
			e=swritelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
			precmz.append(e)
			e=swritelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
			precchrg.append(e)
			e=swritelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
			prodname.append(e)
			e=swritelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
			prodformula.append(e)
			e=swritelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
			prodadduct.append(e)
			e=swritelist[9][t] #sheetinput.cell(row=t, column=10)	# 
			prodmz.append(e)
			e=swritelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
			prodchrg.append(e)
			e=swritelist[11][t] #sheetinput.cell(row=t, column=12)	# 	
			mzerror.append(e)
			e=swritelist[12][t] #sheetinput.cell(row=t, column=13)	# 	
			rettime.append(e)
			e=swritelist[13][t] #sheetinput.cell(row=t, column=14)	# 	
			area.append(e)
			e=swritelist[14][t] #sheetinput.cell(row=t, column=15)	# 	convert to float does not work
			areanormalpercent.append(e)
			e=swritelist[15][t] #sheetinput.cell(row=t, column=16)	# 	
			background.append(e)
			e=swritelist[16][t] #sheetinput.cell(row=t, column=17)	# 	
			fwhm.append(e)
			e=swritelist[17][t] #sheetinput.cell(row=t, column=18)	# 	
			explicitrt.append(e)
			e=swritelist[18][t] #sheetinput.cell(row=t, column=19)	# 	
			rtstart.append(e)
			e=swritelist[19][t] #sheetinput.cell(row=t, column=20)	# 	
			rtend.append(e)
			t=t+1
	r=s+1

swritelist=[]
swritelist.append(mlistname)
swritelist.append(precname)
swritelist.append(precformula)
swritelist.append(precadduct)
swritelist.append(precmz)
swritelist.append(precchrg)
swritelist.append(prodname)
swritelist.append(prodformula)
swritelist.append(prodadduct)
swritelist.append(prodmz)
swritelist.append(prodchrg)
swritelist.append(mzerror)
swritelist.append(rettime)
swritelist.append(area)
swritelist.append(areanormalpercent)
swritelist.append(background)
swritelist.append(fwhm)
swritelist.append(explicitrt)
swritelist.append(rtstart)
swritelist.append(rtend)

# End delete duplicates that are close to the duplicate with max integral

mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
exrt=[]
exrtwindow=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
precoverlap=[]
#end create empty lists
#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists

ki=len(swritelist[0])
print('Number of rows after filter step: %d' % ki)
keeplist=[]		# list with indexes r which are to be kept (the ones not included belong to lines that are part of duplicates)
r=0
e=str(swritelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname		# begin determine which row to start (r) and to end (s)
s=r+1
st=0
while st<1:
	if s>(len(swritelist[1])-1):
		st=1
		s=s-1
	else:
		ne=str(swritelist[1][s]) #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
while r<s+1:
	keeplist.append(r)	# attach first species to keeplist
	r=r+1

kl=0
t=0
while kl<(len(swritelist[0])): #(len(keeplist)):
	#t=keeplist[kl]
	e=swritelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving 
	mlistname.append(e)
	e=swritelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
	precname.append(e)
	e=swritelist[2][t] #sheetinput.cell(row=t, column=3)	# precformula	
	precformula.append(e)
	e=swritelist[3][t] #sheetinput.cell(row=t, column=4)	# precadduct	
	precadduct.append(e)
	e=swritelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
	precmz.append(e)
	e=swritelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
	precchrg.append(e)
	e=swritelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
	prodname.append(e)
	e=swritelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
	prodformula.append(e)
	e=swritelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
	prodadduct.append(e)
	e=swritelist[9][t] #sheetinput.cell(row=t, column=10)	# 
	prodmz.append(e)
	e=swritelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
	prodchrg.append(e)
	e=swritelist[11][t] #sheetinput.cell(row=t, column=12)	# 	
	mzerror.append(e)
	e=swritelist[12][t] #sheetinput.cell(row=t, column=13)	# 	
	rettime.append(e)
	e=swritelist[13][t] #sheetinput.cell(row=t, column=14)	# 	
	area.append(e)
	e=swritelist[14][t] #sheetinput.cell(row=t, column=15)	# 	
	areanormalpercent.append(e)
	e=swritelist[15][t] #sheetinput.cell(row=t, column=16)	# 	
	background.append(e)
	e=swritelist[16][t] #sheetinput.cell(row=t, column=17)	# 	
	fwhm.append(e)
	e=swritelist[17][t] #sheetinput.cell(row=t, column=18)	# 	
	explicitrt.append(e)
	e=swritelist[18][t] #sheetinput.cell(row=t, column=19)	# 	
	rtstart.append(e)
	e=swritelist[19][t] #sheetinput.cell(row=t, column=20)	# 	
	rtend.append(e)
	e=str(swritelist[1][t]) #sheetinput.cell(row=t, column=2)	# precname to get exrt
	ls=len(e)
	k=e[ls-1]
	go=1
	z=1
	while go==1:
		k=e[ls-z]
		if k=='_':
			go=0
		else:
			z=z+1
	x=z-2
	cexrt=e[ls-x]
	x=x-1
	while x>0:
		cexrt=cexrt+e[ls-x]
		x=x-1
	cexrt=str(cexrt)
	cexrt=float(cexrt)
	exrt.append(cexrt)	#exrt done
	exrtwindow.append(0.1)		################################# ENTER EXPLICIT RETENTION TIME WINDOW ##############################
	precoverlap.append('ok')
	kl=kl+1
	t=t+1


# begin remove H'0 from formula, where applicable
#print("Start removing H'0")
pfi=0
while pfi<len(precformula):
	if "H'0" in str(precformula[pfi]):
		#remove H'0 from formula
		ri=precformula[pfi].index("H'0")
		precformula[pfi]=precformula[pfi][0:ri:]+precformula[pfi][ri+3::]
	if "H'0" in str(prodformula[pfi]):
		#remove H'0 from formula
		ri=prodformula[pfi].index("H'0")
		prodformula[pfi]=prodformula[pfi][0:ri:]+prodformula[pfi][ri+3::]
	pfi=pfi+1
#print("End removing H'0")
# end remove H'0 from formula, where applicable


writelist=[]
writelist.append(mlistname)
writelist.append(precname)
writelist.append(precformula)
writelist.append(precadduct)
writelist.append(precmz)
writelist.append(precchrg)
writelist.append(prodname)
writelist.append(prodformula)
writelist.append(prodadduct)
writelist.append(prodmz)
writelist.append(prodchrg)
writelist.append(explicitrt)
writelist.append(exrtwindow)
#print('First filter step is complete.')
############################################################################################################################################################

toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 
		'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']
#print('swritelist created')
transitionresultsdf=pd.DataFrame(writelist).transpose()
#print('Transposed')
transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
#print('Transposed and DataFrame created')
after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_5_'+fourlettcode+'_'
filename='jpmlipidomics_dda_vpw20_1_filtered.csv'
transitionresultsdf.to_csv(filename, index=False)
print('Transition list is saved as jpmlipidomics_dda_vpw20_1_filtered.csv')
afterall=datetime.datetime.now()
dt=afterall-beforeall
print('Calculation time (h:mm:ss) is: %s' % dt)
#print('Calculation time (h:mm:ss) is:')
#print(dt)

