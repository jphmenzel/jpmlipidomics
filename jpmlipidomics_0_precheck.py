# -*- coding: UTF-8 -*-

# Jan Philipp Menzel jpm_lipidomics_vpw13_1_precursor_tr.py
# Goal: STEP 1. Generate transition list for Skyline containing derivatized palmitic and stearic acid
## NOTES: VIRTUAL PRECURSOR - PrecursorName and PrecursorMz are artificially set +Xe (only column 3 and 5), fragment transitions correct including precursor
## NOTES: Virtual precursor forces Skyline to consider all transitions incl. real precursor (fragment in transition list), Skyline Setting: TransitionSettings-Filter-IonTypes-f 
import math
import sys
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
################ DATABASE ## Source: Internetchemie.info 
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.757, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
################
#default=eval(input('Run workflow with default parameters? (Yes: 1 / No: 0) (Derivatization agent: AMPP; Slow and full workflow including all FA; Apply retention time limitation; Use Fatty Acid Library to prevent filtering out important FA; Max RT = 17.5 min; FA: 12 - 24 C; Max m/z error: 10 ppm; Precursor peak area threshold: 3000; Product peak area threshold: 200.) :'))
default=0
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
###########

# begin determine derivatization group sum formula
if default==1:
	fourlettcode='AMPP'
else:
	fourlettcode=input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
if fourlettcode=='AMPP':
	cderiv=12
	hderiv=12
	dderiv=0
	nderiv=2
	oderiv=0
	pderiv=0
	ideriv=0
elif fourlettcode=='NMPA':
	cderiv=7
	hderiv=10
	dderiv=0
	nderiv=2
	oderiv=0
	pderiv=0
	ideriv=0
elif fourlettcode=='NMPE':
	cderiv=7
	hderiv=9
	dderiv=0
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='MDPE':
	cderiv=7
	hderiv=6
	dderiv=3
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='NEPE':
	cderiv=8
	hderiv=11
	dderiv=0
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='EDPE':
	cderiv=6
	hderiv=6
	dderiv=5
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='NPPE':
	cderiv=9
	hderiv=13
	dderiv=0
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='IAMP':
	cderiv=12
	hderiv=11
	dderiv=0
	nderiv=2
	oderiv=0
	pderiv=0
	ideriv=1
else:
	cderiv=eval(input('Number of C atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	hderiv=eval(input('Number of H atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	dderiv=eval(input('Number of D atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	nderiv=eval(input('Number of N atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	oderiv=eval(input('Number of O atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	pderiv=eval(input('Number of P atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	ideriv=eval(input('Number of I atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
# end determine derivatization group sum formula
if default==1:		# default analysis parameters
	workflow=1
	rtlimitation=2
	mostwanted=1
	firstexrt=1.01
	lastexrt=17.5	# lastexrt is the last explicit retention time that is set to look for species (e.g. 18.00 min)
	minlenfa=12		#minlenfa=4	#shortest expected fatty acid
	maxlenfa=24		#maxlenfa=24 #40 #longest expected fatty acid
	mzcutoff=10
	largeareathreshold=3000
	productareathreshold=200
else:
	#workflow=eval(input('Run slow and full workflow (including all FA): 1 or quick and limited workflow (excluding FA with 5 or 6 double bonds): 2. Workflow: '))
	workflow=1
	rtlimitation=2 #eval(input('Apply retention time limitation (1; recommended when linear - Gelb - gradient is used) or no limitation (0; use when retention time trends unknown) :'))
	#mostwanted=eval(input('Use fatty acid library to prevent filtering out important fatty acids? (Apply library: 1; Pure De-Novo Search: 0): '))
	mostwanted=1
	transtest=0 #eval(input('Add -2H species to list? (Yes: 1; No: 0) :'))
	firstexrt=1.01	# 0.3 # firstexrt is the first explicit retention time that is set to look for species (e.g. 1.5 min)	
	lastexrt=eval(input('Enter maximum retention time [min] (at which FA are expected, e.g. 17): '))
	minlenfa=eval(input('Enter number of C atoms in shortest FA chain (at least 4; recommended: 12): '))
	maxlenfa=eval(input('Enter number of C atoms in longest FA chain (max. 30; recommended: 24): '))
	print('Next, enter parameters for transition results filtering:')
	mzcutoff=eval(input('What is max mz error [ppm] for positive identification of species? (e.g. 50): '))
	#largeareathreshold=eval(input('What is the threshold for detected precursor peak area? (compare Skyline report file, e.g. 250): '))
	largeareathreshold=250
	#productareathreshold=eval(input('What is the threshold for detected product peak area? (compare Skyline report file, e.g. 100): '))
	productareathreshold=100
	#runprecheck=eval(input('Run pre check based on palmitic acid and stearic acid (Retention time range prediction)? (Yes: 1; No: 0): '))
	runprecheck=1
	workflowidentifier=sys.argv[1]	# for individual run of this script use instead '23478635249' #
	workflowidentifier=str(workflowidentifier)
	#print(workflowidentifier)
	

beforeall=datetime.datetime.now()
print('Calculation begins at:')
print(beforeall)

print('Workflow is running ...')

nchunks=1
rettimecutoff=lastexrt #eval(input('What is the maximum retention time [min]? (e.g. 11.9 or 17.8) :'))
# begin save workflow parameters in csv file (to be used in following python steps during automated workflow)
transferindexlist=['Four letter code of derivatization agent', 'C', 'H', 'D', 'N', 'O', 'P', 'I', 'min FA length', 'max FA length', 'm/z filter tolerance [ppm]', 'max RT [min]', 'Precursor area threshold', 'Workflow mode [1-Slow/full; 2-Fast/limited]', 'Product area threshold', 'Number of XIC files', 'RT limitation [1-Y; 0-N]', 'Library inclusion [1-Y; 0-N]', '-2H transition [1-Y; 0-N]', 'Pre-check [1-Y; 0-N]', 'Identifier']
transferlist=[]
transferlist.append(fourlettcode)	#0
transferlist.append(cderiv)
transferlist.append(hderiv)
transferlist.append(dderiv)
transferlist.append(nderiv)
transferlist.append(oderiv)	#5
transferlist.append(pderiv)
transferlist.append(ideriv)
transferlist.append(minlenfa)
transferlist.append(maxlenfa)
transferlist.append(mzcutoff)	#10
transferlist.append(rettimecutoff)
transferlist.append(largeareathreshold)
transferlist.append(workflow)
transferlist.append(productareathreshold)
transferlist.append(nchunks)	#15
transferlist.append(rtlimitation)
transferlist.append(mostwanted)
transferlist.append(transtest)
transferlist.append(runprecheck)
transferlist.append(workflowidentifier)	#20
#print(transferlist)
wb = Workbook(write_only=True)
ws = wb.create_sheet()
wb.save('OzFAD1_workflow_parameters.xlsx')
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
while tli<len(transferindexlist):
	ws.cell(row=tli+1, column=1).value=transferindexlist[tli]
	ws.cell(row=tli+1, column=2).value=transferlist[tli]
	tli=tli+1

wb.save('OzFAD1_workflow_parameters.xlsx')
# end save workflow parameters in csv file (to be used in following python steps during automated workflow)
# begin calculate monounsaturated precursors from input

#quit()
# begin define workflow parameters for precheck only
minlenfa=16
maxlenfa=18
# end define workflow parameters for precheck only

lfa=maxlenfa-minlenfa+1
moleculegrouplist=[]
precursornamelist=[]
precursorformulalist=[]
precursoradductlist=[]
precursormzlist=[]
precursorchargelist=[]
productnamelist=[]
productformulalist=[]
productadductlist=[]
productmzlist=[]
productchargelist=[]
molg=fourlettcode+'_FA'
padd='[M]1+'
prdn='precursor'
li=0
while li<lfa:
	moleculegrouplist.append(molg)
	clfa=li+minlenfa
	currentlfa=str(clfa)
	if len(currentlfa)<2:
		currentlfa='0'+currentlfa
		currentlfa=str(currentlfa)
	pnm=fourlettcode+'_'+currentlfa+':1'
	precursornamelist.append(pnm)
	prf=''
	currentcderiv=0
	currenthderiv=0
	currentdderiv=0
	currentnderiv=0
	currentoderiv=0
	currentpderiv=0
	currentideriv=0
	if cderiv>0:
		currentcderiv=cderiv+int(currentlfa)
		prf=prf+'C'+str(currentcderiv)
	if hderiv>0:
		currenthderiv=hderiv+(2*int(currentlfa))-3
		prf=prf+'H'+str(currenthderiv)
	if dderiv>0:
		currentdderiv=dderiv
		prf=prf+"H'"+str(currentdderiv)
	if nderiv>0:
		currentnderiv=nderiv
		prf=prf+'N'+str(nderiv)
	if oderiv>(-1):
		currentoderiv=oderiv+1
		prf=prf+'O'+str(currentoderiv)
	if pderiv>0:
		currentpderiv=pderiv
		prf=prf+'P'+str(pderiv)
	if ideriv>0:
		currentideriv=ideriv
		prf=prf+'I'+str(ideriv)
	precursorformulalist.append(prf)
	productformulalist.append(prf)
	precursoradductlist.append(padd)
	prmz=imass[0]*currenthderiv+imass[1]*currentdderiv+imass[2]*currentcderiv+imass[3]*currentnderiv+imass[4]*currentoderiv+imass[5]*currentpderiv+imass[10]*currentideriv
	precursormzlist.append(prmz)
	productmzlist.append(prmz)
	precursorchargelist.append(1)
	productnamelist.append(prdn)
	productadductlist.append(padd)
	productchargelist.append(1)
	li=li+1
fareadlist=[]
fareadlist.append(moleculegrouplist)
fareadlist.append(precursornamelist)
fareadlist.append(precursorformulalist)
fareadlist.append(precursoradductlist)
fareadlist.append(precursormzlist)
fareadlist.append(precursorchargelist)
fareadlist.append(productnamelist)
fareadlist.append(productformulalist)
fareadlist.append(productadductlist)
fareadlist.append(productmzlist)
fareadlist.append(productchargelist)
# end calculate monounsaturated precursors from input
preconly=1
#minrt=0.0
#print(fareadlist[1])
#maxrt=eval(input('At which retention time [min] ends gradient? :'))
#explrt=(minrt+maxrt)/2
#explrtw=explrt-minrt
nspec=3	# number of species: precursor, aldehyde, criegee
# begin create empty lists
toprow=[]
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]		# double bond position list, contains strings with double bond position n-1, n-2 ...
dbindexlist=[]	# double bond index list, contains index for the double bond position that is closest to head group of FA
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
toprow=['Moleculegroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
ki=len(fareadlist[0])
satlist=fareadlist	#create lists for saturated FAs
#print(ki)
r=0 #2
#ki=ki+2
while r<ki:		#go through rows of excel file jpmlipidozidinput
	# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
	cchain=0
	e=str(fareadlist[1][r]) ## Precursorname
	i=len(e)-3
	add=int(e[i])
	cchain=cchain+add
	i=i-1
	add=int(e[i])
	if add>0:
		cchain=cchain+(10*add)
	else:
		cchain=cchain
		#print('Please check source code (determine number of C atoms in chain)')
	maxdbp=cchain-2
	# begin determine number of C atoms in chain, define highest possible double bond position
	kadd=nspec*maxdbp#+1		#(nspec=3 precursor and 2 products - aldehyde and crigee  - for each of maxdbp possible double bond positions)

	e=fareadlist[0][r] ## MoleculeGroup
	f=fareadlist[1][r] ## Precursorname
	g=fareadlist[2][r]  ##Precursorformula
	h=float(fareadlist[5][r])  ## PrecursorCharge
	k=0
	while k<kadd:
		mlistname.append(e)		# copied, no change
		precname.append(f)		# copied, no change
		precformula.append(g)	# copied, no change
		precchrg.append(h)		# copied, no change
		k=k+1

	e=fareadlist[3][r] ## Precursoradduct
	k=0
	while k<kadd:
		precadduct.append(e) 	# AMPP, precursor
		precadduct.append(e)	# AMPP, aldehyde product
		precadduct.append(e)	# AMPP, crigee product
		k=k+nspec

	e=float(fareadlist[4][r])  ## PrecursorMz
	k=0
	while k<kadd:
		precmz.append(e)	# precursor
		precmz.append(e)	# aldehyde
		precmz.append(e)	# crigee
		k=k+nspec

	e=fareadlist[1][r] ## Productname
	if e=='Chol':
		fragment='_'	
	else:	
		k=0
		csub=1
		while k<kadd:
			dbp='_n-'+str(csub)
			fragment='_precursor'
			ne=e+dbp+fragment
			prodname.append(ne)		# precursor
			dbl=[]	#begin save double bond position for later
			dbl.append(csub)
			dbindexlist.append(dbl)	#end save double bond position for later
			dbp='_n-'+str(csub)
			fragment='_aldehyde neutral loss'
			ne=e+dbp+fragment
			dbpi=0
			while dbpi<nspec:
				dbplist.append(dbp)
				dbpi=dbpi+1
			#
			if preconly==0:	
				prodname.append(ne)		# aldehyde
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				#prodname.append(ne)
				#prodname.append(ne)
				fragment='_criegee neutral loss'
				ne=e+dbp+fragment
				prodname.append(ne)		# crigee
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				#prodname.append(ne)
				#prodname.append(ne)
			if preconly==1:
				dbp='_n-'+str(csub)
				fragment='_dummy precursor a'
				ne=e+dbp+fragment
				prodname.append(ne)		# dummy precursor
				fragment='_dummy precursor b'
				ne=e+dbp+fragment
				prodname.append(ne)		# dummy precursor
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)
				dbindexlist.append(dbl)	#end save double bond position for later
			csub=csub+1
			k=k+nspec

	e=fareadlist[2][r] #
	# begin read precursor sum formula and edit product sum formula
	#print(e)
	#print(e[0])
	#print(e[1])
	#print(e[2])
	#print(e[3])
	#print(len(e))
	clist=[]
	hlist=[]
	dlist=[]
	nlist=[]
	olist=[]
	plist=[]
	ilist=[]
	i=0
	ca=0
	ha=0
	da=0
	na=0
	oa=0
	pa=0
	ia=0
	while i<len(e):
		if e[i]=='H':
			if e[i+1]=="'":
				ha=0
			else:
				ca=0
		#if e[i]=="'":
		#	ha=0		
		if e[i]=='N':
			ha=0
			da=0
		if e[i]=='O':
			ha=0
			da=0
			na=0
		if e[i]=='P':
			ha=0
			da=0
			na=0
			oa=0
		if e[i]=='I':
			ha=0
			da=0
			na=0
			oa=0
			pa=0
		if ca==1:
			clist.append(e[i])
		if ha==1:
			hlist.append(e[i])
		if da==1:
			dlist.append(e[i])
		if na==1:
			nlist.append(e[i])
		if oa==1:
			olist.append(e[i])
		if pa==1:
			plist.append(e[i])
		if ia==1:
			ilist.append(e[i])
		if e[i]=='C':
			ca=1
		if e[i]=='H':
			if e[i+1]=="'":
				ca=0
				ha=0
				da=1
				i=i+1
			else:
				ca=0
				ha=1
		#if e[i]=="'":
		#	ca=0
		#	ha=0
		#	da=1		
		if e[i]=='N':
			ha=0
			da=0
			na=1
			if e[i+1]=='O':
				nlist.append('1')
				na=0
		if e[i]=='O':
			ha=0
			da=0
			na=0
			oa=1
			if (i+1)<len(e):
				if e[i+1]=='P':
					olist.append('1')
					oa=0
			else:
				olist.append('1')
				oa=0					
		if e[i]=='P':
			da=0
			na=0
			oa=0
			pa=1
			if (i+1)<len(e):
				if e[i+1]=='I':
					plist.append('1')
					pa=0
			else:
				plist.append('1')
				pa=0
		if e[i]=='I':
			da=0
			na=0
			oa=0
			pa=0
			ia=1
			if i==(len(e)-1):
				ilist.append('1')
				ia=0
		i=i+1
	#print(clist)
	#print(hlist)
	#print(dlist)
	#print(nlist)
	#print(olist)
	#print(plist)
	if len(clist)==0:
		cn=0
	if len(hlist)==0:
		hn=0
	if len(dlist)==0:
		dn=0	
	if len(nlist)==0:
		nn=0
	if len(olist)==0:
		on=0
	if len(plist)==0:
		pn=0
	if len(ilist)==0:
		iodon=0
	if len(clist)==1:
		cn=int(clist[0])
	if len(clist)==2:
		cn=10*int(clist[0])+int(clist[1])
	if len(clist)==3:
		cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
	if len(hlist)==1:
		hn=int(hlist[0])
	if len(hlist)==2:
		hn=10*int(hlist[0])+int(hlist[1])
	if len(hlist)==3:
		hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
	if len(dlist)==1:
		dn=int(dlist[0])
	if len(dlist)==2:
		dn=10*int(dlist[0])+int(dlist[1])
	if len(dlist)==3:
		dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
	if len(nlist)==1:
		nn=int(nlist[0])
	if len(nlist)==2:
		nn=10*int(nlist[0])+int(nlist[1])
	if len(nlist)==3:
		nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
	if len(olist)==1:
		on=int(olist[0])
	if len(olist)==2:
		on=10*int(olist[0])+int(olist[1])
	if len(olist)==3:
		on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
	if len(plist)==1:
		pn=int(plist[0])
	if len(plist)==2:
		pn=10*int(plist[0])+int(plist[1])
	if len(plist)==3:
		pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])	
	if len(ilist)==1:
		iodon=int(ilist[0])
	if len(ilist)==2:
		iodon=10*int(ilist[0])+int(ilist[1])
	if len(ilist)==3:
		iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		
	# end read precursor sum formula
	#print(cn)
	e=fareadlist[0][r] ## begin calculate product sum formula
	if e=='SPLASH':
		e=e 		# begin calculate product sum formula for SPLASH
	else:
		k=0
		csub=1
		while k<kadd:
			cnp=cn-(csub)
			hnp=hn-(2*csub)
			onald=on+1
			oncrigee=on+2
			precursor='C'+str(cn)+'H'+str(hn)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
			productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
			productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
			prodformula.append(precursor)
			if preconly==0:
				prodformula.append(productaldehyde)
				#prodformula.append(productaldehyde)
				#prodformula.append(productaldehyde)
				prodformula.append(productcrigee)
				#prodformula.append(productcrigee)
				#prodformula.append(productcrigee)	# product formula is saved in list for current double bond position
			if preconly==1:
				prodformula.append(precursor)
				prodformula.append(precursor)
			precursormz=imass[0]*(hn)+imass[1]*(dn)+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
			prodmz.append(precursormz)
			if preconly==0:
				productmz=imass[0]*(hnp+0)+imass[1]*(dn)+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, AMPP
				prodmz.append(productmz)
				productmz=imass[0]*(hnp+0)+imass[1]*(dn)+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, AMPP
				prodmz.append(productmz)
			if preconly==1:
				prodmz.append(precursormz)
				prodmz.append(precursormz)
			csub=csub+1
			k=k+nspec			
	# end read precursor sum formula and edit product sum formula

	prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
	prodchrg=precchrg 	# ProductCharge

	r=r+1
	e=fareadlist[1][r-1] 	# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	if r>(len(fareadlist[0])-1):
		r=r+1
	else:
		e=fareadlist[1][r] 
		prec=e
		while prec==prevprec:
			r=r+1
			e=fareadlist[1][r-1] 
			prevprec=e
			e=fareadlist[1][r] 
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	
#end save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
writelist=[]
writelist.append(mlistname)
writelist.append(precname)
writelist.append(precformula)
writelist.append(precadduct)
writelist.append(precmz)
writelist.append(precchrg)
writelist.append(prodname)
writelist.append(prodformula)
writelist.append(prodadduct)
writelist.append(prodmz)
writelist.append(prodchrg)
#end save data in writelist

#print('All calculations for monounsaturated fatty acids are done.')

# begin add saturated FAs
r=0
while r<(len(satlist[1])):
	#satlist[1][r][len(satlist[1][r])-1]='0'	
	precn=''
	p=0
	while p<(len(satlist[1][r])-1):
		precn=precn+satlist[1][r][p]
		p=p+1
	precn=precn+'0'					#PrecursorName to saturated FA
	satlist[1][r]=precn
	satlist[4][r]=float(satlist[4][r])+(2*imass[0])	#PrecursorMz to sat. FA
	satlist[9][r]=float(satlist[9][r])+(2*imass[0])	#ProductMz to sat. FA
	hcurr=(10*(int(satlist[2][r][4])))+(int(satlist[2][r][5]))
	hcurr=str(hcurr+2)
	p=0
	precf=''
	while p<4:
		precf=precf+satlist[2][r][p]
		p=p+1
	precf=precf+hcurr
	p=p+2
	while p<(len(satlist[2][r])):
		precf=precf+satlist[2][r][p]
		p=p+1
	satlist[2][r]=str(precf)	#PrecursorFormula is edited
	satlist[7][r]=str(precf)	#ProductFormula is edited
	r=r+1
pnm=0
while pnm<len(satlist[0]):
	satlist[6][pnm]=str(satlist[1][pnm])+'_precursor'
	pnm=pnm+1
clmn=0
while clmn<(len(satlist)):
	writelist[clmn]=writelist[clmn]+satlist[clmn]
	clmn=clmn+1
# end add saturated FAs

##########################################################VIRTUAL#PRECURSOR###########################################################################
######################################################################################################################################################
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]		# content in lists is deleted to free up memory
# begin introduce virtual precursor
ki=len(writelist[0])

#print('Entries in excel file after all FA transitions are generated: %d' % ki)
virtualprecformula=[]
virtualprecmz=[]

vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]

r=0
ki=ki
while r<ki:
	go=0
	cfa=str(writelist[1][r][5])+str(writelist[1][r][6])+str(writelist[1][r][7])+str(writelist[1][r][8])
	if cfa=='16:0':
		go=1
	elif cfa=='18:0':
		go=1
	else:
		go=0
	if go==1:
		e=str(writelist[2][r])#sheetinput.cell(row=r, column=3)	# PrecursorFormula
		ve=e+'Xe'
		virtualprecformula.append(ve) # virtual precursorFormula
		e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
		ve=e+(imass[9])		#add one Xe atom to generate virtual precursor
		virtualprecmz.append(ve)	# virtual precursormz
		e=str(writelist[0][r])#sheetinput.cell(row=r, column=1)
		vmlistname.append(e)
		e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)
		vprecname.append(e)
		e=str(writelist[3][r]) #sheetinput.cell(row=r, column=4)
		vprecadduct.append(e)
		e=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)
		vprecchrg.append(e)
		e=str(writelist[6][r]) #sheetinput.cell(row=r, column=7)
		vprodname.append(e)
		e=str(writelist[7][r]) #sheetinput.cell(row=r, column=8)
		vprodformula.append(e)
		e=str(writelist[8][r]) #sheetinput.cell(row=r, column=9)
		vprodadduct.append(e)
		e=float(writelist[9][r]) #sheetinput.cell(row=r, column=10)
		vprodmz.append(e)
		e=int(writelist[10][r]) #sheetinput.cell(row=r, column=11)
		vprodchrg.append(e)
	r=r+1
# begin save excel file with virtual precursor as csv file
toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
vwritelist=[]
vwritelist.append(vmlistname)
vwritelist.append(vprecname)
vwritelist.append(virtualprecformula)
vwritelist.append(vprecadduct)
vwritelist.append(virtualprecmz)
vwritelist.append(vprecchrg)
vwritelist.append(vprodname)
vwritelist.append(vprodformula)
vwritelist.append(vprodadduct)
vwritelist.append(vprodmz)
vwritelist.append(vprodchrg)
terminate=0
if terminate==1:
	print('writelist created')
	transitionresultsdf=pd.DataFrame(vwritelist).transpose()
	print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10]]
	print('Transposed and DataFrame created')
	transitionresultsdf.to_csv('jpmlipidomics_vpw20_virtual_precursor.csv', index=False)
	print('Transition list saved as jpmlipidomics_vpw20_virtual_precursor.csv')
# end save excel file with virtual precursor as csv file

#print('Saving data.')
#print('len(vmlistname)')
#print(len(vmlistname))
virtualprecformula=[]
virtualprecmz=[]
vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[] # test to delete lists to save up memory

#print('Transition list is modified with virtual precursor [M + Xe].')

#terminate=1
if terminate==1:
	quit()

#################################################################################################################################################
############################# BEGIN REDUCE TO ONE ENTRY PER PRECURSOR AND EXPAND WITH VARIED EXPLICIT RETENTION TIME ############################
#################################################################################################################################################
#terminate=eval(input('Reduce transition list to one entry per precursor and expand with varied explicit retention time? Yes: 1 No: 0 ::'))
terminate=1
if terminate==0:
	quit()
else:
	#exrtstep=eval(input('Stepwidth for varied explicit retention time (e.g. 0.05 min)? ::'))
	#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
	normalexrtstep=0.027
	bigexrtstep=0.045
	#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
# begin reduce and expand
ki=len(vwritelist[0])
# end count entries in excel file
vmlistname=[]
vprecname=[]
virtualprecformula=[]
vprecadduct=[]
virtualprecmz=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]
precrt=[]
rtwindow=[]
r=0
ki=ki
pos=1
while r<ki:
	if pos==1:
		if rtlimitation==1:
			# begin redefine firstexrt as cfirstexrt and lastexrt as clastexrt depending on chainlength of current FA
			e=(10*int(vwritelist[1][r][5]))+int(vwritelist[1][r][6])
			e=int(e)
			if e>15:
				if e<30:
					cfirstexrt=firstexrt+(((lastexrt-firstexrt)*(1/4))*(((e-15)/15)))
				else:
					cfirstexrt=firstexrt+(((lastexrt-firstexrt)*(1/4))+((lastexrt-firstexrt)*(1/4))*(((e-30)/10)))
				cfirstexrt=round(cfirstexrt, 3)
			else:
				cfirstexrt=firstexrt
			if e<20:
				if e<8:
					clastexrt=lastexrt-(((lastexrt-firstexrt)*(1/8))+((lastexrt-firstexrt)*(1/8))*(((8-e)/4)))
				else:
					clastexrt=lastexrt-(((lastexrt-firstexrt)*(1/8))*(((20-e)/12)))
				clastexrt=round(clastexrt, 3)
			else:
				clastexrt=lastexrt
			# end redefine firstexrt as cfirstexrt and lastexrt as clastexrt depending on chainlength of current FA
		else:
			cfirstexrt=firstexrt #float(firstexrt)
			clastexrt=lastexrt #float(lastexrt)
			clastexrt=round(clastexrt, 3)
			cfirstexrt=round(cfirstexrt, 3)
		#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
		if int(vwritelist[1][r][8])>2:
			nca=10*int(vwritelist[1][r][5])+int(vwritelist[1][r][6])
			if nca>17:
				exrt=cfirstexrt+bigexrtstep
			else:
				exrt=cfirstexrt+normalexrtstep
		else:	
			exrt=cfirstexrt+normalexrtstep		# firstexrt is the first explicit retention time that is set to look for species (e.g. 1.50 min)
		#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
		exrtstr=str(exrt)
		while exrt<clastexrt:		# lastexrt is the last explicit retention time that is set to look for species (e.g. 18.00 min)
			ve=str(vwritelist[2][r]) #sheetinput.cell(row=r, column=3)	# PrecursorFormula
			virtualprecformula.append(ve) # virtual precursorFormula
			ve=float(vwritelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz	
			virtualprecmz.append(ve)	# virtual precursormz
			e=str(vwritelist[0][r]) #sheetinput.cell(row=r, column=1)
			vmlistname.append(e)
			e=str(vwritelist[1][r]) #sheetinput.cell(row=r, column=2)
			cm='_'
			cm=str(cm)
			exrtstr=str(round(exrt, 2))		#
			ee=e+cm+exrtstr			#
			vprecname.append(ee)
			precrt.append(exrt)
			e=str(vwritelist[3][r]) #sheetinput.cell(row=r, column=4)
			vprecadduct.append(e)
			e=int(vwritelist[5][r]) #sheetinput.cell(row=r, column=6)
			vprecchrg.append(e)
			e=str(vwritelist[6][r]) #sheetinput.cell(row=r, column=7)
			vprodname.append(e)
			e=str(vwritelist[7][r]) #sheetinput.cell(row=r, column=8)
			vprodformula.append(e)
			e=str(vwritelist[8][r]) #sheetinput.cell(row=r, column=9)
			vprodadduct.append(e)
			e=float(vwritelist[9][r]) #sheetinput.cell(row=r, column=10)
			vprodmz.append(e)
			e=int(vwritelist[10][r]) #sheetinput.cell(row=r, column=11)
			vprodchrg.append(e)	
			#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
			if int(vwritelist[1][r][8])>2:
				nca=10*int(vwritelist[1][r][5])+int(vwritelist[1][r][6])
				if nca>17:
					exrt=exrt+bigexrtstep
					exrtstep=bigexrtstep
				else:
					exrt=exrt+normalexrtstep
					exrtstep=normalexrtstep	
			else:	
				exrt=exrt+normalexrtstep
				exrtstep=normalexrtstep
			#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain	
			e=exrtstep*2
			rtwindow.append(e)
			if runprecheck==0:
				r=ki-1		# reduce entries in transition list to save time, if precheck not run
			exrt=clastexrt		################# disables varied exrt !!		##################################################
	vprecmz=vwritelist[4][r] #sheetinput.cell(row=r, column=5)	# PrecursorMz
	r=r+1
	corr=0
	if r==ki:
		r=r-1
		corr=1
	e=vwritelist[4][r] #sheetinput.cell(row=r, column=5)	# PrecursorMz
	if e==vprecmz:
		pos=0
	else:
		pos=1
	r=r+corr

# begin save to csv file
prt='PrecursorRT'
prt=str(prt)
toprow.append(prt)
rtw='PrecursorRTWindow'
rtw=str(rtw)
toprow.append(rtw)
ki=2+len(vmlistname)

vwritelist=[]
vwritelist.append(vmlistname)
vwritelist.append(vprecname)
vwritelist.append(virtualprecformula)
vwritelist.append(vprecadduct)
vwritelist.append(virtualprecmz)
vwritelist.append(vprecchrg)
vwritelist.append(vprodname)
vwritelist.append(vprodformula)
vwritelist.append(vprodadduct)
vwritelist.append(vprodmz)
vwritelist.append(vprodchrg)
#vwritelist.append(precrt)
#vwritelist.append(rtwindow)
#print('writelist created')
transitionresultsdf=pd.DataFrame(vwritelist).transpose()
#print('Transposed')
transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10]] #,toprow[11],toprow[12]]
#print('Transposed and DataFrame created')
after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_1_'+fourlettcode+'_'
#filename=today+'jpmlipidomics_vpw13_1_precursor.csv'
filename='jpmlipidomics_vpw20_0_precheck.csv'
transitionresultsdf.to_csv(filename, index=False)
afterall=datetime.datetime.now()
dt=afterall-beforeall
nrows=len(vmlistname)
#print('Transition list is saved as yyyy_mm_dd_1_xxxx_jpmlipidomics_vpw13_1_precursor.csv (%d rows)' % nrows)
print('Transition list is saved as jpmlipidomics_vpw20_0_precheck.csv (%d rows)' % nrows)
print('Calculation time (h:mm:ss) is:')
print(dt)
# end save to csv file
#################################################################################################################################################
############################### END REDUCE TO ONE ENTRY PER PRECURSOR AND EXPAND WITH VARIED EXPLICIT RETENTION TIME ############################
#################################################################################################################################################


