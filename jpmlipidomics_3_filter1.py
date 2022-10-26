# -*- coding: UTF-8 -*-

# Philipp Menzel lipidomics oz id transition list generate from excel list, use for skyline
#created: 09 07 2020
#modified: regularly until 25 08 2020 
# Goal: read excel file containing data for monounsaturated lipids without double bond info, add rows for OzID product ions, save in excel file
## Notes: work in progress for fatty acids AMPP derivatives (no other ionization), double bond position added to Precursorname after calculations but before saving in excel file 
## Notes: addition for saturated FAs, bisunsaturated FAs, added line for precursor, option for precursor-only transition list with dummy percursor.
## NOTES: STAGE 2 for LIPIDOMICS WORKFLOW AMPP VPW05. Here: read TransitionResults from Skyline for precursor only results with MinPeakFoundRatio=1.0. Select correct species based on
##		  manually defined selection rules. OUTPUT CONTAINS RT INFORMATION FOR PRECURSOR ONLY ANALYSIS
##		  strict selection criteria enabled, based on relative abundance of precursor to fragments
##			flag overlap of precursor species
##	DONE ## 
import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
import subprocess
beforeall=datetime.datetime.now()
################ DATABASE ## Source: Internetchemie.info
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.7, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
################
########### begin read workflow parameters insert
transferlist=[]
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
go=1
while go==1:
	tfe=ws.cell(row=tli+1, column=2)
	tfe=tfe.value
	if tli==17:
		go=0
	if tfe is None:
		go=0
	else:
		transferlist.append(tfe)
	tli=tli+1
########## 
fourlettcode=transferlist[0]		#input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
cderiv=transferlist[1]
hderiv=transferlist[2]
dderiv=transferlist[3]
nderiv=transferlist[4]
oderiv=transferlist[5]
pderiv=transferlist[6]
ideriv=transferlist[7]
mzcutoff=int(transferlist[10]) 		#eval(input('What is max mz [ppm] for positive identification of species? (e.g. 40) :'))
#rettimecutoff=eval(input('What is the maximum retention time [min] ? (e.g. 17.8) :')) 
rettimecutoff=float(transferlist[11]) 		#
prodareathreshold=int(transferlist[14]) #eval(input('What is the area threshold for products? (e.g. 300) :')) ##################################### ACTIVATE #############
precareathreshold=int(transferlist[12])		#	#applies to precursor
mostwanted=int(transferlist[17])
manualfilter=0 
selectiontype=1
########### end read workflow parameters

# begin build mostwantedlist 
mostwantedlist=[]
wbfa=openpyxl.load_workbook('jpm_fa_lib.xlsx')
wsfa=wbfa.active
mwi=0
go=1
while go==1:
	tfe=wsfa.cell(row=mwi+3, column=1)
	tfe=tfe.value
	if tfe is None:
		go=0
	else:
		if int(tfe)<10:
			tfe='0'+str(tfe)
		mwfa=str(tfe)+':'
		tfe=wsfa.cell(row=mwi+3, column=2)
		tfe=tfe.value
		dbfa=int(tfe)
		if int(tfe)==0:
			mwfa=mwfa+str(tfe)
		else:
			mwfa=mwfa+str(tfe)+'_n-'
			dbfai=1
			while dbfai<(dbfa+1):
				tfe=wsfa.cell(row=mwi+3, column=dbfai+2)
				tfe=tfe.value
				if dbfai<(dbfa):
					mwfa=mwfa+str(tfe)+'_n-'
				else:
					mwfa=mwfa+str(tfe)
				dbfai=dbfai+1
		mostwantedlist.append(mwfa)
	mwi=mwi+1
# end build mostwantedlist

#prodareathreshold=200	#applies to products of loss
#nspec=3	# number of species: precursor, aldehyde, crigee
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585]
#abundance=[99.9885, 98.93, 99.636, 99.7, 94.93] not updated
#ki=eval(input('Number of entries?'))
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
diagnostics=[]
#end create empty lists
#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists
trdf=pd.read_csv('skyl_report_vpw20_2.csv')
toprowx=[trdf.columns.values.tolist()]
toprow=toprowx[0]
trdf=trdf.transpose()
writelist=trdf.values.tolist()
ki=len(writelist[0])
print('Number of rows in skyl_report_vpw20_2.csv: %d' % ki)
#print(ki)
#####################################################################################################
# begin assign whether transitions diagnostic or non-diagnostic
r=0
while r<ki:
	t=r
	if int(writelist[1][t][8])>2:
		go=1
		ch=len(writelist[1][t])-1
		while go==1:
			if str(writelist[1][t][ch])=='n':
				go=0
				if str(writelist[1][t][ch+4])=='_':
					lastdb=10*int(writelist[1][t][ch+2])+int(writelist[1][t][ch+3])		# determine last db (closest to amide bond)
				else:
					lastdb=int(writelist[1][t][ch+2])
				if str(writelist[1][t][ch-4])=='-':	
					seclastdb=10*int(writelist[1][t][ch-3])+int(writelist[1][t][ch-2])	# determine second last db (second closest to amide bond)
				else:
					seclastdb=int(writelist[1][t][ch-2])
			else:
				go=1
			ch=ch-1
		if lastdb>13:
			if lastdb-seclastdb==3:
				if str(writelist[6][t][len(writelist[6][t])-1])=='r':
					diagnostics.append('diagnostic')
				else:
					if str(writelist[6][t][len(writelist[6][t])-4])=='n':
						if lastdb==10*int(writelist[6][t][len(writelist[6][t])-2])+int(writelist[6][t][len(writelist[6][t])-1]):
							diagnostics.append('non-diagnostic')
						else:
							diagnostics.append('diagnostic')
					else:
						if lastdb==int(writelist[6][t][len(writelist[6][t])-1]):
							diagnostics.append('non-diagnostic')
						else:
							diagnostics.append('diagnostic')
			else:
				diagnostics.append('diagnostic')
		else:
			diagnostics.append('diagnostic')
	else:
		diagnostics.append('diagnostic')
	r=r+1
# end assign whether transitions diagnostic or non-diagnostic
####################################################################################################
r=0
ki=ki
while r<ki:		# go through rows of excel file 
	currentmzerror=[]
	currentrettime=[]
	currentareas=[]
	currentprodmz=[]
	currentprodnm=[]
	e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# Precursorname		# begin determine which row to start (r) and to end (s)
	s=r+1
	st=0
	while st<1:
		if s>(len(writelist[1])-1):
			ne='stop_loop'
		else:
			ne=writelist[1][s] #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
	t=r
	pos=0
	fpos=0
	gpos=0
	prodmzcutoff=mzcutoff*3.5
	while t<(s+1):
		e=str(writelist[11][t]) ## mzerror
		if e=='nan':
			if str(diagnostics[t])=='diagnostic':
				pos=1
		else:
			e=float(writelist[11][t]) ## mzerror
			currentmzerror.append(e)
			prodnm=str(writelist[6][t]) ## begin mzcutoff for precursors and transitions from five db with lowest n
			requireddb=5			# required number of double bonds to be tested for cutoff condition #################################################
			if prodnm[(len(prodnm)-1)]=='r':  # not enabled: use normal mzcutoff for any precursor
				if abs(e)>mzcutoff:
					if str(diagnostics[t])=='diagnostic':
						pos=pos
			elif (int(prodnm[8]))<(requireddb+1):	# use 2.5 times normal mzcutoff for any product transition from FA with less than 6 double bonds
				if abs(e)>(prodmzcutoff):
					if str(diagnostics[t])=='diagnostic':
						pos=1
			else:
				if prodnm[len(prodnm)-3]=='-':
					currentdb=10*(int(prodnm[len(prodnm)-2]))+(int(prodnm[len(prodnm)-1]))
				elif prodnm[len(prodnm)-2]=='-':
					currentdb=(int(prodnm[len(prodnm)-1]))
				else:
					if prodnm[13]=='_':
						currentdb=(int(prodnm[12]))
					else:
						currentdb=10*(int(prodnm[12]))+(int(prodnm[13]))
				searchdb=0
				go=0
				while go<requireddb: #requireddb
					if (str(prodnm[searchdb]))=='-':
						go=go+1
						if (str(prodnm[searchdb+3]))=='_':
							checkdb=(int(prodnm[searchdb+1])*10)+(int(prodnm[searchdb+2]))
						elif (str(prodnm[searchdb+2]))=='_':
							checkdb=(int(prodnm[searchdb+1]))
						if checkdb==currentdb:
							if go<3:
								if abs(e)>prodmzcutoff:
									if str(diagnostics[t])=='diagnostic':
										pos=1
							elif go<4:
								if abs(e)>(prodmzcutoff+3):
									if str(diagnostics[t])=='diagnostic':
										pos=1
							else:
								if abs(e)>(prodmzcutoff+10):	# use wider mzcutoff for any transition from FA with 6 double bonds from 5th or 6th db
									if str(diagnostics[t])=='diagnostic':
										pos=1
					searchdb=searchdb+1		# end mzcutoff for precursors and transitions from three db with lowest n

			f=float(writelist[17][t]) ## explicit retention time
			currentrettime.append(f)
			if f>rettimecutoff:
				fpos=1
			g=float(writelist[13][t]) ## area index 13 (column 14) 
			#print(g)
			prodnm=str(writelist[6][t]) ## begin area cutoff for precursors and transitions from five db with lowest n
			currentprodnm.append(prodnm)
			requireddb=5			# required number of double bonds to be tested for cutoff condition #################################################
			if prodnm[(len(prodnm)-1)]=='r':
				if g<precareathreshold: 	# precursor area CUTOFF
					if mostwanted==0:
						gpos=1
					else:
						# begin test whether species to be cut is in the mostwantedlist (if it is, keep only if area higher than half the areacutoff)
						fashort=str()
						fas=5
						while fas<(len(writelist[1][t])-5):
							fashort=fashort+str(writelist[1][t][fas])
							fas=fas+1
						mwl=0
						gposcancel=0
						while mwl<(len(mostwantedlist)):
							if mostwantedlist[mwl]==fashort:
								gposcancel=1
								if g<(precareathreshold*0.5):		# use half the area threshold for mostwanted species
									gpos=1
								elif writelist[6][t][8]==0:			# use normal area threshold for all saturated FA
									gpos=1
							mwl=mwl+1
						if gposcancel==0:
							gpos=1
						# end test whether species to be cut is in the mostwantedlist
					#gpos=1
			elif (int(prodnm[8]))<(requireddb+1):
				if g<prodareathreshold:		# product area CUTOFF (for species with 1 or 2 or 3 db)
					if str(diagnostics[t])=='diagnostic':
						if mostwanted==0:
							gpos=1
						else:
							# begin test whether species to be cut is in the mostwantedlist
							fashort=str()
							fas=5
							while fas<(len(writelist[1][t])-5):
								fashort=fashort+str(writelist[1][t][fas])
								fas=fas+1
							mwl=0
							gposcancel=0
							while mwl<(len(mostwantedlist)):
								if mostwantedlist[mwl]==fashort:
									#print('!!! FOUND ONE !!!')
									#print(writelist[6][t])
									#print('!!! FOUND ONE !!!')
									gposcancel=1
								mwl=mwl+1
							if gposcancel==0:
								gpos=1
							# end test whether species to be cut is in the mostwantedlist
						#gpos=1
			else:
				if prodnm[len(prodnm)-3]=='-':
					currentdb=10*(int(prodnm[len(prodnm)-2]))+(int(prodnm[len(prodnm)-1]))
				elif prodnm[len(prodnm)-2]=='-':
					currentdb=(int(prodnm[len(prodnm)-1]))
				else:
					if prodnm[13]=='_':
						currentdb=(int(prodnm[12]))
					else:
						currentdb=10*(int(prodnm[12]))+(int(prodnm[13]))
				searchdb=0
				go=0
				while go<requireddb:
					if (str(prodnm[searchdb]))=='-':
						go=go+1
						if (str(prodnm[searchdb+3]))=='_':
							checkdb=(int(prodnm[searchdb+1])*10)+(int(prodnm[searchdb+2]))
						elif (str(prodnm[searchdb+2]))=='_':
							checkdb=(int(prodnm[searchdb+1]))
						if checkdb==currentdb:
							if g<(prodareathreshold/3):		# product area CUTOFF (past third db)
								if str(diagnostics[t])=='diagnostic':
									if mostwanted==0:
										gpos=1
									else:
										# begin test whether species to be cut is in the mostwantedlist
										fashort=str()
										fas=5
										while fas<(len(writelist[1][t])-5):
											fashort=fashort+str(writelist[1][t][fas])
											fas=fas+1
										mwl=0
										gposcancel=0
										while mwl<(len(mostwantedlist)):
											if mostwantedlist[mwl]==fashort:
												gposcancel=1
											mwl=mwl+1
										if gposcancel==0:
											gpos=1
										# end test whether species to be cut is in the mostwantedlist
					searchdb=searchdb+1		# end area cutoff for precursors and transitions from three db with lowest n
			#currentareas.append(g)
			g=float(writelist[9][t]) #sheetinput.cell(row=t, column=10)	# prodmz
			currentprodmz.append(g)
		t=t+1
	# begin use current area and prodmz to determine, whether current species has a precursor abundance that is lower than all fragments (criterium for exclusion)
	if len(currentareas)>1:		#only exclude unsaturated FAs, keep saturated FAs 
		ci=0
		call=0
		while call<1:
			ciprodmz=currentprodmz[ci]
			cj=0
			cc=0
			ct=1
			while cj<len(currentprodmz):
				cjprodmz=currentprodmz[cj]
				if ciprodmz>(cjprodmz-0.001):
					cc=1
				else:
					ct=0
				cj=cj+1
			if ct==1:
				call=1
				cmax=ci
			else:
				ci=ci+1		# species with highest mz is identified
		cmax=len(currentareas)-1		# assume that last species is precursor (disable this line, if not the case!!!)
		# cmax is index for species with highest mass (precursor)
		ciareas=currentareas[cmax]
		cj=0
		cc=0
		ct=1
		while cj<len(currentareas):
			cjareas=currentareas[cj]
			if ciareas<(cjareas+0.001):
				cc=1
			else:
				ct=0
			cj=cj+1
		if ct==1:
			if selectiontype==1:
				if mostwanted==0:
					gpos=1
				else:
					# begin test whether species to be cut is in the mostwantedlist
					fashort=str()
					fas=5
					while fas<(len(writelist[1][t])-5):
						fashort=fashort+str(writelist[1][t][fas])
						fas=fas+1
					mwl=0
					gposcancel=0
					while mwl<(len(mostwantedlist)):
						if mostwantedlist[mwl]==fashort:
							gposcancel=1
						mwl=mwl+1
					if gposcancel==0:
						gpos=1
					# end test whether species to be cut is in the mostwantedlist
			else:
				nochange=0
	# end use current area and prodmz to determine, whether current species has a precursor abundance that is lower than all fragments (criterium for exclusion)
	# begin subtract +2 of precursor from n-1 aldehyde transition
	# This module was not created, as the transitions slightly deviate in their exact mass, not allowing a clean subtraction 
	# end subtract +2 of precursor from n-1 aldehyde transition
	# begin use current area and productname to determine, whether current species has for any double bond a criegee area larger than respective aldehyde area
	# 		exclude, if species not in mostwantedlist, if mostwantedlist is in use, otherwise exclude
	cpi=0
	while cpi<(len(currentprodnm)):
		if str(currentprodnm[cpi]).find('aldehyde')==1:
			cpj=cpi
			while cpj<(len(currentprodnm)):
				if (str(currentprodnm[cpi][len(currentprodnm[cpi])-1])+str(currentprodnm[cpi][len(currentprodnm[cpi])-2]))==(str(currentprodnm[cpi][len(currentprodnm[cpi])-1])+str(currentprodnm[cpi][len(currentprodnm[cpi])-2])):
					if (float(currentareas[cpi])*1.05)<(float(currentareas[cpj])):
						if mostwanted==0:
							pos=1
						else:
							# begin test whether species to be cut is in the mostwantedlist
							fashort=str()
							fas=5
							while fas<(len(writelist[1][t])-5):
								fashort=fashort+str(writelist[1][t][fas])
								fas=fas+1
							mwl=0
							gposcancel=0
							while mwl<(len(mostwantedlist)):
								if mostwantedlist[mwl]==fashort:
									gposcancel=1
								mwl=mwl+1
							if gposcancel==0:
								pos=1
							# end test whether species to be cut is in the mostwantedlist
				cpj=cpj+1
		cpi=cpi+1
	# end use current area and productname to determine, whether current species has for any double bond a criegee area larger than respective aldehyde area
	if pos==1:
		r=s+1
	elif fpos==1:
		r=s+1
	elif gpos==1:
		r=s+1
	else:
		t=r
		while t<(s+1):
			#e=sheetinput.cell(row=t, column=19)	# 	begin determine whether duplicate is found, duplicate kept
			#e=e.value
			#rtstartcurrent=e
			#e=sheetinput.cell(row=t, column=20)	# 	
			#e=e.value
			#rtendcurrent=e
			#if ki==0:
			#	apos=1
			#else:
			#	apos=0
				#rtstartprevious=rtstart[(len(rtstart)-1)]
				#rtendprevious=rtend[(len(rtend)-1)]
				#if rtstartprevious==rtstartcurrent:
					#if rtendprevious==rtendcurrent:
						#apos=1			# 	duplicates removed, not kept for later stage of analysis with transitions of fragments
						#apos=0			# 	end determine whether duplicate is found, duplicate kept, all saved in next step
			apos=0
			if apos==0:
				e=writelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving in excel output
				mlistname.append(e)
				e=writelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
				precname.append(e)
				e=writelist[2][t] #sheetinput.cell(row=t, column=3)	# precname	
				precformula.append(e)
				e=writelist[3][t] #sheetinput.cell(row=t, column=4)	# precformula	
				precadduct.append(e)
				e=writelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
				precmz.append(e)
				e=writelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
				precchrg.append(e)
				e=writelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
				prodname.append(e)
				e=writelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
				prodformula.append(e)
				e=writelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
				prodadduct.append(e)
				e=writelist[9][t] #sheetinput.cell(row=t, column=10)	# 
				prodmz.append(e)
				e=writelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
				prodchrg.append(e)
				e=writelist[11][t] #sheetinput.cell(row=t, column=12)	# 	
				mzerror.append(e)
				e=writelist[12][t] #sheetinput.cell(row=t, column=13)	# 	
				rettime.append(e)
				e=writelist[13][t] #sheetinput.cell(row=t, column=14)	# 	
				area.append(e)
				e=writelist[14][t] #sheetinput.cell(row=t, column=15)	# 	convert to float does not work
				areanormalpercent.append(e)
				e=writelist[15][t] #sheetinput.cell(row=t, column=16)	# 	
				background.append(e)
				e=writelist[16][t] #sheetinput.cell(row=t, column=17)	# 	
				fwhm.append(e)
				e=writelist[17][t] #sheetinput.cell(row=t, column=18)	# 	
				explicitrt.append(e)
				e=writelist[18][t] #sheetinput.cell(row=t, column=19)	# 	
				rtstart.append(e)
				e=writelist[19][t] #sheetinput.cell(row=t, column=20)	# 	
				rtend.append(e)
			t=t+1
	r=s+1
# end read transition results and append suitable species to lists
swritelist=[]
swritelist.append(mlistname)
swritelist.append(precname)
swritelist.append(precformula)
swritelist.append(precadduct)
swritelist.append(precmz)
swritelist.append(precchrg)
swritelist.append(prodname)
swritelist.append(prodformula)
swritelist.append(prodadduct)
swritelist.append(prodmz)
swritelist.append(prodchrg)
swritelist.append(mzerror)
swritelist.append(rettime)
swritelist.append(area)
swritelist.append(areanormalpercent)
swritelist.append(background)
swritelist.append(fwhm)
swritelist.append(explicitrt)
swritelist.append(rtstart)
swritelist.append(rtend)

#print('First filter step is complete.')
#quit()
######################################################################################################################################################################
######################################################end first filter step###########################################################################################
######################################################################################################################################################################
######################################################################################################################################################################
######################################################################################################################################################################
#print('Please check that the excel file jpmlipidomics_transitionresults_input.xlsx contains appropriate values in the correct columns')
#selectiontype=eval(input('Generate Transition Results based on m/z error and retention time cutoff only (0) or based on strict selection criteria (1)? : '))
selectiontype=1
#if selectiontype==1:
	#print('Species are excluded also, if precursor is less abundant than both aldehyde and criegee.')
	#print('Not yet available.')
#mzcutoff=eval(input('What is max mz for positive identification of species? (e.g. 4.9) :'))
#rettimecutoff=eval(input('What is the maximum retention time [min], before wash? (e.g. 17.8) :'))
#nspec=3	# number of species: precursor, aldehyde, criegee
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
#abundance=[99.9885, 98.93, 99.636, 99.7, 94.93] not updated
#ki=eval(input('Number of entries?'))
# begin create empty lists
#toprow=[] # Is defined above, nothing changed
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
#end create empty lists
#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists

ki=len(swritelist[0])

#print('Number of rows after first filter step: %d' % ki)
#print(ki)
r=0
ki=ki
while r<ki:		# go through rows of swritelist
	currentmzerror=[]
	currentrettime=[]
	currentareas=[]
	currentprodmz=[]
	e=str(swritelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname		# begin determine which row to start (r) and to end (s)
	s=r+1
	st=0
	while st<1:
		if s>(len(swritelist[1])-1):
			s=s-1
			st=1
		else:
			ne=str(swritelist[1][s]) #sheetinput.cell(row=s, column=2)	# Precursorname
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
	#print('check')
	t=r
	pos=0
	fpos=0
	gpos=0
	while t<(s+1):
		e=float(swritelist[11][t]) #sheetinput.cell(row=t, column=12)	# mzerror
		currentmzerror.append(e)
		f=float(swritelist[12][t]) #sheetinput.cell(row=t, column=13)	# retention time
		currentrettime.append(f)
		if f>rettimecutoff:
			fpos=1
		g=float(swritelist[13][t]) #sheetinput.cell(row=t, column=14)	# area
		currentareas.append(g)
		g=float(swritelist[9][t]) #sheetinput.cell(row=t, column=10)	# prodmz
		currentprodmz.append(g)
		t=t+1
	# begin use current area and prodmz to determine, whether current species has a precursor abundance that is lower than all fragments (criterium for exclusion)
	if len(currentareas)>1:		#only exclude unsaturated FAs, keep saturated FAs 
		ci=0
		call=0
		while call<1:
			ciprodmz=currentprodmz[ci]
			cj=0
			cc=0
			ct=1
			while cj<len(currentprodmz):
				cjprodmz=currentprodmz[cj]
				if ciprodmz>(cjprodmz-0.001):
					cc=1
				else:
					ct=0
				cj=cj+1
			if ct==1:
				call=1
				cmax=ci
			else:
				ci=ci+1		# species with highest mz is identified
		cmax=len(currentareas)-1		# assume that last species is precursor (disable this line, if not the case!!!)
		# cmax is index for species with highest mass (precursor)
		ciareas=currentareas[cmax]
		cj=0
		cc=0
		ct=1
		while cj<len(currentareas):
			cjareas=currentareas[cj]
			if ciareas<(cjareas+0.001):
				cc=1
			else:
				ct=0
			cj=cj+1
		if ct==1:
			if selectiontype==1:
				if mostwanted==0:
					gpos=1
				else:
					# begin test whether species to be cut is in the mostwantedlist
					fashort=str()
					fas=5
					while fas<(len(writelist[1][t])-5):
						fashort=fashort+str(writelist[1][t][fas])
						fas=fas+1
					mwl=0
					gposcancel=0
					while mwl<(len(mostwantedlist)):
						if mostwantedlist[mwl]==fashort:
							gposcancel=1
						mwl=mwl+1
					if gposcancel==0:
						gpos=1
					# end test whether species to be cut is in the mostwantedlist
				#gpos=1
	# begin use current area and prodmz to determine, whether current species has a precursor abundance that is lower than all fragments (criterium for exclusion)
	if pos==1:
		r=s+1
	elif fpos==1:
		r=s+1
	elif gpos==1:
		r=s+1
	else:
		t=r
		while t<(s+1):
			e=swritelist[0][t] ## mlistname	# begin append rows of suitable species to lists for later saving in excel output
			mlistname.append(e)
			e=swritelist[1][t] ## precname	
			precname.append(e)
			e=swritelist[2][t] ## precformula	
			precformula.append(e)
			e=swritelist[3][t] ## precadduct
			precadduct.append(e)
			e=swritelist[4][t] ## precmz
			precmz.append(e)
			e=swritelist[5][t] ## precchrg
			precchrg.append(e)
			e=swritelist[6][t] ## prodname
			prodname.append(e)
			e=swritelist[7][t] ## prodformula
			prodformula.append(e)
			e=swritelist[8][t] ## prodadduct
			prodadduct.append(e)
			e=swritelist[9][t] ## prodmz
			prodmz.append(e)
			e=swritelist[10][t] ## prodchrg
			prodchrg.append(e)
			e=swritelist[11][t] ## 	
			mzerror.append(e)
			e=swritelist[12][t] ## 	
			rettime.append(e)
			e=swritelist[13][t] ## 	
			area.append(e)
			e=swritelist[14][t] ## 	
			areanormalpercent.append(e)
			e=swritelist[15][t] ## 	
			background.append(e)
			e=swritelist[16][t] ## 	
			fwhm.append(e)
			e=swritelist[17][t] ## 	
			explicitrt.append(e)
			e=swritelist[18][t] ## 	
			rtstart.append(e)
			e=swritelist[19][t] ## 	
			rtend.append(e)
			t=t+1
	r=s+1
# end read transition results and append suitable species to lists
swritelist=[]
swritelist.append(mlistname)
swritelist.append(precname)
swritelist.append(precformula)
swritelist.append(precadduct)
swritelist.append(precmz)
swritelist.append(precchrg)
swritelist.append(prodname)
swritelist.append(prodformula)
swritelist.append(prodadduct)
swritelist.append(prodmz)
swritelist.append(prodchrg)
swritelist.append(mzerror)
swritelist.append(rettime)
swritelist.append(area)
swritelist.append(areanormalpercent)
swritelist.append(background)
swritelist.append(fwhm)
swritelist.append(explicitrt)
swritelist.append(rtstart)
swritelist.append(rtend)

#print('Second filter step is done.')
#quit() ## check ok
#########################################################end second filter step#############################################################
######################################################begin remove duplicates, flag overlap#################################################
# begin create empty lists
#toprow=[]
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
exrt=[]
exrtwindow=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
precoverlap=[]
#end create empty lists
#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists

ki=len(swritelist[0])
print('Number of rows after filter step: %d' % ki)
keeplist=[]		# list with indexes r which are to be kept (the ones not included belong to lines that are part of duplicates)
r=0
e=str(swritelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname		# begin determine which row to start (r) and to end (s)
s=r+1
st=0
while st<1:
	if s>(len(swritelist[1])-1):
		st=1
		s=s-1
	else:
		ne=str(swritelist[1][s]) #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
while r<s+1:
	keeplist.append(r)	# attach first species to keeplist
	r=r+1
#while r<ki:
#	e=str(swritelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname		# begin determine which row to start (r) and to end (s)
#	s=r+1
#	st=0
#	while st<1:
#		if s<ki:
#			ne=str(swritelist[1][s]) #sheetinput.cell(row=s, column=2)	# Precursorname
#		else:
#			ne='stop loop'
#		if ne==e:
#			s=s+1
#			st=0
#		else:
#			s=s-1
#			st=1		# end determine s
#	kl=0
#	cut=0		
#	while kl<(len(keeplist)):
#		rw=keeplist[kl]
#		f=swritelist[6][rw] #sheetinput.cell(row=keeplist[kl], column=7)	# Productname
#		g=swritelist[6][r] #sheetinput.cell(row=r, column=7)	# Productname
#		if f==g:
#			rw=keeplist[kl]
#			h=swritelist[12][rw] #sheetinput.cell(row=keeplist[kl], column=13)	# retention time
#			i=swritelist[12][r] #sheetinput.cell(row=r, column=13)	# retention time
#			if h==i:
#				cut=0 #1 for remove duplicates, 0 for KEEP ALL DUPLICATES
#		kl=kl+1
#	if cut==1:
#		r=s+1
#	else:
#		while r<s+1:
#			keeplist.append(r)
#			r=r+1
	#next species
kl=0
t=0
while kl<(len(swritelist[0])): #(len(keeplist)):
	#t=keeplist[kl]
	e=swritelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving 
	mlistname.append(e)
	e=swritelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
	precname.append(e)
	e=swritelist[2][t] #sheetinput.cell(row=t, column=3)	# precformula	
	precformula.append(e)
	e=swritelist[3][t] #sheetinput.cell(row=t, column=4)	# precadduct	
	precadduct.append(e)
	e=swritelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
	precmz.append(e)
	e=swritelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
	precchrg.append(e)
	e=swritelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
	prodname.append(e)
	e=swritelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
	prodformula.append(e)
	e=swritelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
	prodadduct.append(e)
	e=swritelist[9][t] #sheetinput.cell(row=t, column=10)	# 
	prodmz.append(e)
	e=swritelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
	prodchrg.append(e)
	e=swritelist[11][t] #sheetinput.cell(row=t, column=12)	# 	
	mzerror.append(e)
	e=swritelist[12][t] #sheetinput.cell(row=t, column=13)	# 	
	rettime.append(e)
	e=swritelist[13][t] #sheetinput.cell(row=t, column=14)	# 	
	area.append(e)
	e=swritelist[14][t] #sheetinput.cell(row=t, column=15)	# 	
	areanormalpercent.append(e)
	e=swritelist[15][t] #sheetinput.cell(row=t, column=16)	# 	
	background.append(e)
	e=swritelist[16][t] #sheetinput.cell(row=t, column=17)	# 	
	fwhm.append(e)
	e=swritelist[17][t] #sheetinput.cell(row=t, column=18)	# 	
	explicitrt.append(e)
	e=swritelist[18][t] #sheetinput.cell(row=t, column=19)	# 	
	rtstart.append(e)
	e=swritelist[19][t] #sheetinput.cell(row=t, column=20)	# 	
	rtend.append(e)
	e=str(swritelist[1][t]) #sheetinput.cell(row=t, column=2)	# precname to get exrt
	ls=len(e)
	k=e[ls-1]
	go=1
	z=1
	while go==1:
		k=e[ls-z]
		if k=='_':
			go=0
		else:
			z=z+1
	x=z-2
	cexrt=e[ls-x]
	x=x-1
	while x>0:
		cexrt=cexrt+e[ls-x]
		x=x-1
	cexrt=str(cexrt)
	cexrt=float(cexrt)
	exrt.append(cexrt)	#exrt done
	exrtwindow.append(0.1)		################################# ENTER EXPLICIT RETENTION TIME WINDOW ##############################
	precoverlap.append('ok')
	kl=kl+1
	t=t+1

### begin add decoy precursor (-H transition of precursor)
if manualfilter==0:
	rd=0
	while rd<(len(mlistname)):
		if str(prodname[rd][len(prodname[rd])-1])=='r':
			mld=mlistname[rd]
			mlistname.insert(rd+1,mld)
			mld=precname[rd]
			precname.insert(rd+1,mld)
			mld=precformula[rd]
			precformula.insert(rd+1,mld)
			mld=precadduct[rd]
			precadduct.insert(rd+1,mld)
			mld=precmz[rd]
			precmz.insert(rd+1,mld)
			mld=precchrg[rd]
			precchrg.insert(rd+1,mld)
			#change prodname, prodMz and prodformula
			mld=prodname[rd].replace('precursor','decoy')
			prodname.insert(rd+1,mld)
			mld=float(prodmz[rd])-imass[0]
			prodmz.insert(rd+1,mld)
			# begin read precursor sum formula and # begin edit decoy sum formula 
			e=prodformula[rd]
			#print(e)
			#print(e[0])
			clist=[]
			hlist=[]
			dlist=[]
			nlist=[]
			olist=[]
			plist=[]
			ilist=[]
			i=0
			ca=0
			ha=0
			da=0
			na=0
			oa=0
			pa=0
			ia=0
			while i<len(e):
				if e[i]=='H':
					if e[i+1]=="'":
						ha=0
					else:
						ca=0
				#if e[i]=='D':
				#	ha=0		
				if e[i]=='N':
					ha=0
					da=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
				if e[i]=='P':
					ha=0
					da=0
					na=0
					oa=0
				if e[i]=='I':
					ha=0
					da=0
					na=0
					oa=0
					pa=0
					ia=0
				if ca==1:
					clist.append(e[i])
				if ha==1:
					hlist.append(e[i])
				if da==1:
					dlist.append(e[i])
				if na==1:
					nlist.append(e[i])
				if oa==1:
					olist.append(e[i])
				if pa==1:
					plist.append(e[i])
				if ia==1:
					ilist.append(e[i])
				if e[i]=='C':
					ca=1
				if e[i]=='H':
					if e[i+1]=="'":
						ca=0
						ha=0
						da=1
						i=i+1
					else:
						ca=0
						ha=1
				#if e[i]=='D':
				#	ca=0
				#	ha=0
				#	da=1		
				if e[i]=='N':
					ha=0
					da=0
					na=1
					if e[i+1]=='O':
						nlist.append('1')
						na=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
					oa=1
					if (i+1)<len(e):
						if e[i+1]=='P':
							olist.append('1')
							oa=0
					else:
						olist.append('1')
						oa=0					
				if e[i]=='P':
					da=0
					na=0
					oa=0
					pa=1
					if (i+1)<len(e):
						if e[i+1]=='I':
							plist.append('1')
							pa=0
					else:
						plist.append('1')
						pa=0
				if e[i]=='I':
					da=0
					na=0
					oa=0
					pa=0
					ia=1
					if i==(len(e)-1):
						ilist.append('1')
						ia=0
				i=i+1
			#print(clist)
			#print(hlist)
			#print(dlist)
			#print(nlist)
			#print(olist)
			#print(plist)
			if len(clist)==0:
				cn=0
			if len(hlist)==0:
				hn=0
			if len(dlist)==0:
				dn=0	
			if len(nlist)==0:
				nn=0
			if len(olist)==0:
				on=0
			if len(plist)==0:
				pn=0
			if len(ilist)==0:
				iodon=0
			if len(clist)==1:
				cn=int(clist[0])
			if len(clist)==2:
				cn=10*int(clist[0])+int(clist[1])
			if len(clist)==3:
				cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
			if len(hlist)==1:
				hn=int(hlist[0])
			if len(hlist)==2:
				hn=10*int(hlist[0])+int(hlist[1])
			if len(hlist)==3:
				hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
			if len(dlist)==1:
				dn=int(dlist[0])
			if len(dlist)==2:
				dn=10*int(dlist[0])+int(dlist[1])
			if len(dlist)==3:
				dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
			if len(nlist)==1:
				nn=int(nlist[0])
			if len(nlist)==2:
				nn=10*int(nlist[0])+int(nlist[1])
			if len(nlist)==3:
				nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
			if len(olist)==1:
				on=int(olist[0])
			if len(olist)==2:
				on=10*int(olist[0])+int(olist[1])
			if len(olist)==3:
				on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
			if len(plist)==1:
				pn=int(plist[0])
			if len(plist)==2:
				pn=10*int(plist[0])+int(plist[1])
			if len(plist)==3:
				pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])	
			if len(ilist)==1:
				iodon=int(ilist[0])
			if len(ilist)==2:
				iodon=10*int(ilist[0])+int(ilist[1])
			if len(ilist)==3:
				iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
			hndecoy=hn-1		# subtract H to create sum formula of decoy precursor
			decoy='C'+str(cn)+'H'+str(hndecoy)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
			prodformula.insert(rd+1,decoy)		# end edit sum formula of decoy and insert in list
			mld=prodadduct[rd]
			prodadduct.insert(rd+1,mld)
			mld=prodchrg[rd]
			prodchrg.insert(rd+1,mld)
			mld=explicitrt[rd]
			explicitrt.insert(rd+1,mld)
			mld=exrtwindow[rd]
			exrtwindow.insert(rd+1,mld)
			rd=rd+1
		rd=rd+1

### end add decoy precursor (-H transition of precursor)
nrowsfa=len(mlistname)
print('Number of rows before splitting into chunks: %d' % nrowsfa)
# begin save as csv in case of first filter step before manual filtering
if manualfilter==0:
	# begin save to csv file	# split results into multiple files, if ################################################################ SPLIT OR NOT TO SPLIT #######
	minrowcutoff=150000		############################################################################################################ SPLIT OR NOT TO SPLIT #######
	lnm=len(mlistname)
	if lnm%minrowcutoff==0:
		minrowcutoff=minrowcutoff+1
	if lnm<minrowcutoff:		#normal procedure, no split
		print('No splitting of the analysis is performed.')
		toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 
		'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']
		writelist=[]
		writelist.append(mlistname)
		writelist.append(precname)
		writelist.append(precformula)
		writelist.append(precadduct)
		writelist.append(precmz)
		writelist.append(precchrg)
		writelist.append(prodname)
		writelist.append(prodformula)
		writelist.append(prodadduct)
		writelist.append(prodmz)
		writelist.append(prodchrg)
		writelist.append(explicitrt)
		writelist.append(exrtwindow)
		#print('writelist created')
		transitionresultsdf=pd.DataFrame(writelist).transpose()
		#print('Transposed')
		transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
		#print('Transposed and DataFrame created')
		after=datetime.datetime.now()
		after=str(after)
		today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_5_'+fourlettcode+'_'
		filename='jpmlipidomics_vpw20_3_1st_filter.csv'
		transitionresultsdf.to_csv(filename, index=False)
		print('Transition list is saved as jpmlipidomics_vpw20_3_1st_filter.csv')
		afterall=datetime.datetime.now()
		dt=afterall-beforeall
		print('Calculation time (h:mm:ss) is:')
		print(dt)
		subprocess.call([r'C:\Users\menzel2\batchprogramming\OzFAD1\OzFAD1_black_box\Skyline_Analysis_First_Filter.bat'])		# 
	else:
		nchunks=int((lnm/minrowcutoff)-((lnm % minrowcutoff)/minrowcutoff)+1)		# number of files the list will be split into
		#begin save nchunks in Workflow parameters file
		ws.cell(row=16, column=2).value=nchunks		# transferlist[15]
		wb.save('jpmlipidomics_workflow_parameters.xlsx')
		#end save nchunks in Workflow parameters file
		print('Number of chunks that the analysis is split into: %d' % nchunks)
		filecount=0
		rowswritten=0
		toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 
			'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']
		while filecount<nchunks:
			fileindex=filecount+1
			cmlistname=[]
			cprecname=[]
			cprecformula=[]
			cprecadduct=[]
			cprecmz=[]
			cprecchrg=[]
			cprodname=[]
			cprodformula=[]
			cprodadduct=[]
			cprodmz=[]
			cprodchrg=[]
			cexplicitrt=[]
			cexrtwindow=[]
			writelist=[]
			go=1
			if rowswritten>(fileindex*minrowcutoff):
				go=0
			while go==1:
				if rowswritten<(fileindex*minrowcutoff):
					go=1
				else:
					lmln=(len(mlistname))-1 # lmln is the index of the last list element in the transition list
					if rowswritten==lmln:
						go=0
					else:
						if precname[rowswritten]==precname[rowswritten+1]:
							go=1
						else:
							go=0	#0 
				lmln=(len(mlistname))-1 # lmln is the index of the last list element in the transition list
				if rowswritten<(lmln+1):
					cmlistname.append(mlistname[rowswritten])
					cprecname.append(precname[rowswritten])
					cprecformula.append(precformula[rowswritten])
					cprecadduct.append(precadduct[rowswritten])
					cprecmz.append(precmz[rowswritten])
					cprecchrg.append(precchrg[rowswritten])
					cprodname.append(prodname[rowswritten])
					cprodformula.append(prodformula[rowswritten])
					cprodadduct.append(prodadduct[rowswritten])
					cprodmz.append(prodmz[rowswritten])
					cprodchrg.append(prodchrg[rowswritten])
					cexplicitrt.append(explicitrt[rowswritten])
					cexrtwindow.append(exrtwindow[rowswritten])
					rowswritten=rowswritten+1
				else:
					go=0
			
			writelist.append(cmlistname)
			writelist.append(cprecname)
			writelist.append(cprecformula)
			writelist.append(cprecadduct)
			writelist.append(cprecmz)
			writelist.append(cprecchrg)
			writelist.append(cprodname)
			writelist.append(cprodformula)
			writelist.append(cprodadduct)
			writelist.append(cprodmz)
			writelist.append(cprodchrg)
			writelist.append(cexplicitrt)
			writelist.append(cexrtwindow)

			if filecount>0:
				# begin merge old and new file !!
				mergedradf=pd.read_csv('jpmlipidomics_vpw20_4_rank1_2nd_filter.csv')
				mergedrbdf=pd.read_csv('jpmlipidomics_vpw20_4_rank2_2nd_filter.csv')
				toprowx=[mergedradf.columns.values.tolist()]
				toprowm=toprowx[0]
				mergedradf=mergedradf.transpose()
				mergedrbdf=mergedrbdf.transpose()
				mergedralist=mergedradf.values.tolist()
				mergedrblist=mergedrbdf.values.tolist()
				# old file is read, next process current chunk of data
				transitionresultsdf=pd.DataFrame(writelist).transpose()		#print('Transposed')
				transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
				#filename='jpmlipidomics_vpw19_3_1st_filter_file_'+fileindex+'.csv'
				filename='jpmlipidomics_vpw20_3_1st_filter_chunk.csv'
				transitionresultsdf.to_csv(filename, index=False)
				# xic results of current chunk are written in file, next, process these in Skyline (and analyse results in python)
				subprocess.call([r'C:\Users\menzel2\batchprogramming\OzFAD1\OzFAD1_black_box\Skyline_Analysis_First_Filter_chunk.bat'])		#
				# current chunk of data is processed
				chunkradf=pd.read_csv('jpmlipidomics_vpw20_4_rank1_2nd_filter.csv')
				chunkrbdf=pd.read_csv('jpmlipidomics_vpw20_4_rank2_2nd_filter.csv')
				chunkradf=chunkradf.transpose()
				chunkrbdf=chunkrbdf.transpose()
				chunkralist=chunkradf.values.tolist()
				chunkrblist=chunkrbdf.values.tolist()
				coli=0
				while coli<13:
					mergedralist[coli]=mergedralist[coli]+chunkralist[coli]
					mergedrblist[coli]=mergedrblist[coli]+chunkrblist[coli]
					coli=coli+1
				mergedradf=pd.DataFrame(mergedralist).transpose()		#print('Transposed')
				mergedradf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
				filename='jpmlipidomics_vpw20_4_rank1_2nd_filter.csv'
				mergedradf.to_csv(filename, index=False)
				mergedrbdf=pd.DataFrame(mergedrblist).transpose()		#print('Transposed')
				mergedrbdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
				filename='jpmlipidomics_vpw20_4_rank2_2nd_filter.csv'
				mergedrbdf.to_csv(filename, index=False)
			else:
				transitionresultsdf=pd.DataFrame(writelist).transpose()		#print('Transposed')
				transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
				filename='jpmlipidomics_vpw20_3_1st_filter_chunk.csv'
				transitionresultsdf.to_csv(filename, index=False)
				# current chunk is written in file, next, process these in Skyline (and analyse results in python)
				subprocess.call([r'C:\Users\menzel2\batchprogramming\OzFAD1\OzFAD1_black_box\Skyline_Analysis_First_Filter_chunk.bat'])		#

			filecount=filecount+1

	# end save to csv file 	#if number of rows larger than threshold, then transitions are split over multiple (nchunk) files (nchunk is saved in workflow parameters excel file)
# begin call batch script to run as many skyline runner instances in sequence, as number of files (filecount)
# 		for each skyline run, export results and chromatograms
#		for each skyline run, call python filter2 script to generate report
#if len(mlistname)<minrowcutoff:		#normal procedure, no split
	# begin call batch script (single run)
	#subprocess.call([r'C:\Users\menzel2\batchprogramming\Skyline_Analysis_First_Filter.bat'])		# 
	# end call batch script (single run)
#else:
	# begin call batch script (multi run)
	#filecount=0
	#while filecount<nchunks:
	#	subprocess.call([r'C:\Users\menzel2\batchprogramming\Skyline_Analysis_First_Filter_chunk.bat'])		# 
	#	filecount=filecount+1

	
	# end call batch script (multi run)
	# end call batch script 
	# begin merge transition lists into one transition list



	# end merge transition lists into one transition list

