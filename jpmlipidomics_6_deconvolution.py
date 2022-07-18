# -*- coding: UTF-8 -*-

# Jan Philipp Menzel 
#created: 09 07 2020
#last edit: 12 07 2022
# Goal: read csv file containing data from last skyline routine (manual curation of list), save data as transition list csv file and sorted xlsx file for plotting
## NOTES: STAGE 4 . 
##			flags overlap of precursor species, deconvolution of precursor
##			corrects areas for isotopic pattern using sum formula of species
##			saves data both in csv file and xlsx file
##	DONE ## 
import math
import openpyxl
import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.styles.colors import RGB
from openpyxl.chart.axis import ChartLines
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import PatternFill
import requests
import urllib.request
import time
from bs4 import BeautifulSoup
import statistics
from statistics import mean
from statistics import median
import copy
beforeall=datetime.datetime.now()

searchlmsd=0 	# switch to 1, if lookup of LIPIDMAPS IDs is required at this step (normally it should not be required)

################ DATABASE ## Source: Internetchemie.info
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.7, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
################
#print('Before proceeding, please make sure that the Skyline report file is named skyl_report_final_dia_manual_int.csv')

#selectiontype=eval(input('Generate Transition Results based on m/z error and retention time cutoff only (0) or based on strict selection criteria (1)? : '))
# begin determine derivatization group sum formula
fourlettcode=input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
if fourlettcode=='AMPP':
	cderiv=12
	hderiv=12
	dderiv=0
	nderiv=2
	oderiv=0
	pderiv=0
	ideriv=0
elif fourlettcode=='NMPA':
	cderiv=7
	hderiv=10
	dderiv=0
	nderiv=2
	oderiv=0
	pderiv=0
	ideriv=0
elif fourlettcode=='NMPE':
	cderiv=7
	hderiv=9
	dderiv=0
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='MDPE':
	cderiv=7
	hderiv=6
	dderiv=3
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='NEPE':
	cderiv=8
	hderiv=11
	dderiv=0
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='EDPE':
	cderiv=6
	hderiv=6
	dderiv=5
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='NPPE':
	cderiv=9
	hderiv=13
	dderiv=0
	nderiv=1
	oderiv=1
	pderiv=0
	ideriv=0
elif fourlettcode=='IAMP':
	cderiv=12
	hderiv=11
	dderiv=0
	nderiv=2
	oderiv=0
	pderiv=0
	ideriv=1
else:
	cderiv=eval(input('Number of C atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	hderiv=eval(input('Number of H atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	dderiv=eval(input('Number of D atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	nderiv=eval(input('Number of N atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	oderiv=eval(input('Number of O atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	pderiv=eval(input('Number of P atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
	ideriv=eval(input('Number of I atoms that derivatization group contains? (minus fatty acyl CO-CnHm) :'))
# end determine derivatization group sum formula
manualfilter=1 
selectiontype=1
#rettimecutoff=17.8
#precareathreshold=2000	#applies to precursor
#prodareathreshold=1000	#applies to products of loss
#abundance=[99.9885, 98.93, 99.636, 99.7, 94.93] not updated
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
#end create empty lists
#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists

trdf=pd.read_csv('skyl_report_dia_int.csv')
toprowx=[trdf.columns.values.tolist()]
toprow=toprowx[0]
trdf=trdf.transpose()
swritelist=trdf.values.tolist()
ki=len(swritelist[0])
print('Number of rows in skyl_report_dia_int.csv: %d' % ki)
##########################################################################################################
######################################################begin flag overlap#################################################
# begin create empty lists
#toprow=[]
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
exrt=[]
exrtwindow=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
precoverlap=[]
#end create empty lists
#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists

ki=len(swritelist[0])

keeplist=[]		# list with indexes r which are to be kept (the ones not included belong to lines that are part of duplicates)
r=0
e=str(swritelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname		# begin determine which row to start (r) and to end (s)
s=r+1
st=0
while st<1:
	if s>(len(swritelist[1])-1):
		st=1
		s=s-1
	else:
		ne=str(swritelist[1][s]) #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
#print('check0')
while r<s+1:
	keeplist.append(r)	# attach first species to keeplist
	r=r+1

#print('check1')
kl=0
t=0
while kl<(len(swritelist[0])): #(len(keeplist)):
	#t=keeplist[kl]
	e=swritelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving 
	mlistname.append(e)

	correctpname=0 #1 1 only when using MS DIA only workflow with messed up precnames
	if correctpname==1:
		#e=e[:len(e)-3]
		swritelist[1][t]=str(swritelist[1][t])
		swritelist[1][t]=swritelist[1][t][:len(swritelist[1][t])-3]

	e=swritelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
	precname.append(e)
	e=swritelist[2][t] #sheetinput.cell(row=t, column=3)	# precformula	
	precformula.append(e)
	e=swritelist[3][t] #sheetinput.cell(row=t, column=4)	# precadduct	
	precadduct.append(e)
	e=swritelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
	precmz.append(e)
	e=swritelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
	precchrg.append(e)
	e=swritelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
	prodname.append(e)
	e=swritelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
	prodformula.append(e)
	e=swritelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
	prodadduct.append(e)
	e=swritelist[9][t] #sheetinput.cell(row=t, column=10)	# 
	prodmz.append(e)
	e=swritelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
	prodchrg.append(e)
	e=swritelist[11][t] #sheetinput.cell(row=t, column=12)	# 	
	mzerror.append(e)
	e=swritelist[12][t] #sheetinput.cell(row=t, column=13)	# 	
	rettime.append(e)
	e=swritelist[13][t] #sheetinput.cell(row=t, column=14)	# 	
	area.append(e)
	e=swritelist[14][t] #sheetinput.cell(row=t, column=15)	# 	
	areanormalpercent.append(e)
	e=swritelist[15][t] #sheetinput.cell(row=t, column=16)	# 	
	background.append(e)
	e=swritelist[16][t] #sheetinput.cell(row=t, column=17)	# 	
	fwhm.append(e)
	e=swritelist[17][t] #sheetinput.cell(row=t, column=18)	# 	
	explicitrt.append(e)
	e=swritelist[18][t] #sheetinput.cell(row=t, column=19)	# 	
	rtstart.append(e)
	e=swritelist[19][t] #sheetinput.cell(row=t, column=20)	# 	
	rtend.append(e)
	e=str(swritelist[1][t]) #sheetinput.cell(row=t, column=2)	# precname to get exrt
	#print(e)
	ls=len(e)
	k=e[ls-1]
	go=1
	z=1
	while go==1:
		#print(e)
		k=e[ls-z]
		k=str(k)
		if k=='_':
			#print('check')
			go=0
		else:
			z=z+1
	x=z-2
	cexrt=e[ls-x]
	x=x-1
	while x>0:
		cexrt=cexrt+e[ls-x]
		x=x-1
	cexrt=str(cexrt)
	cexrt=float(cexrt)
	exrt.append(cexrt)	#exrt done
	exrtwindow.append(0.05)		################################# ENTER EXPLICIT RETENTION TIME WINDOW ##############################
	precoverlap.append('ok')
	kl=kl+1
	t=t+1
	if t<(len(swritelist[1])):
		if str(swritelist[1][t])=='nan':
			kl=1+(len(swritelist[1]))
#print('check2')
#### begin detect and flag overlap #######################################

if manualfilter==1:
	f=0
	while f<len(prodformula):
		g=f+1
		while g<len(prodformula):
			if prodformula[f]==prodformula[g]:
				if str(prodname[f][len(prodname[f])-1])=='r':
					if str(prodname[g][len(prodname[g])-1])=='r':
						if rtstart[f]<rtend[g]:
							if rtend[f]>rtstart[g]:
								#print('FLAG')
								precoverlap[f]='Overlapping precursor'  #column entry 'PrecursorOverlap'
								precoverlap[g]='Overlapping precursor'	#column entry 'PrecursorOverlap'
								pf=prodname[f]
								pg=prodname[g]
								pf=str(pf)
								pg=str(pg)
								if prodname[f][len(prodname[f])-10]=='_':
									pfn=pf+', OVERLAPPING precursor'
									prodname[f]=pfn
								if prodname[g][len(prodname[g])-10]=='_':
									pgn=pg+', OVERLAPPING precursor'
									prodname[g]=pgn
									
			g=g+1
		f=f+1
	#print('Overlapping precursors are detected and flagged.')
#### end detect and flag overlap #######################################

# begin apply correction for isotopic pattern to area
if manualfilter==1:
	uncorrectedarea=area
	area=[]
	cr=0
	while cr<(len(uncorrectedarea)):
		# begin read precursor sum formula and edit product sum formula
		e=prodformula[cr]
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=='D':
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=='D':
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])	
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		corra=(uncorrectedarea[cr])/((0.9893**cn)*(0.999885**hn)*(0.99636**nn)*(0.99757**on)) #correct area with factor to change from area of principal ion to area of whole isotopic pattern
		area.append(corra)
		#if cr==60:
		#	print(uncorrectedarea[cr])
		#	print(prodformula[cr])
		#	print(cn)
		#	print(hn)
		#	print(nn)
		#	print(on)
		#	print(corra)
		cr=cr+1

# end apply correction for isotopic pattern to area
# begin save as csv in case of first filter step before manual filtering
if manualfilter==1:
	# begin save to csv file
	toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 
	'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']
	writelist=[]
	writelist.append(mlistname)
	writelist.append(precname)
	writelist.append(precformula)
	writelist.append(precadduct)
	writelist.append(precmz)
	writelist.append(precchrg)
	writelist.append(prodname)
	writelist.append(prodformula)
	writelist.append(prodadduct)
	writelist.append(prodmz)
	writelist.append(prodchrg)
	writelist.append(explicitrt)
	writelist.append(exrtwindow)
	#print('writelist created')
	writefile=0
	if writefile==1:
		transitionresultsdf=pd.DataFrame(writelist).transpose()
		#print('Transposed')
		transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
		#print('Transposed and DataFrame created')
		transitionresultsdf.to_csv('jpmlipidomics_5_0_filtered_tr.csv', index=False)
		print('Final filtered transition list is saved as jpmlipidomics_5_0_filtered_tr.csv')
# end save as csv in case of first filter step before manual filtering
#begin save excel file #######################################################################################################################

#begin create excel sheet for analysis results
wb = Workbook(write_only=True)
ws = wb.create_sheet('transition_results')
#ws = wb.create_sheet('sorted_results')
ws = wb.create_sheet('precursor_results')
ws = wb.create_sheet('ozid_barchart')
ws = wb.create_sheet('final_barchart')
wb.save('OzFAD1_4_input_DIA_Q.xlsx')

wb=openpyxl.load_workbook('OzFAD1_4_input_DIA_Q.xlsx')
sheet=wb['transition_results']
#print('DONE.')
#quit()

#end create excel sheet for analysis results

toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 
'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow', 'mzErrorPPM', 'RT', 'Area',
 'AreaNormalized', 'Background', 'FWHM', 'PeakRT', 'RTStart', 'RTEnd', 'PrecursorOverlap','UncorrectedArea']
c=1
while c<(len(toprow)+1):
	sheet.cell(row=1, column=c).value=toprow[c-1]
	c=c+1

r=2
while r<(len(mlistname)+2):
	sheet.cell(row=r, column=1).value=mlistname[r-2]
	sheet.cell(row=r, column=2).value=precname[r-2]
	sheet.cell(row=r, column=3).value=precformula[r-2]
	sheet.cell(row=r, column=4).value=precadduct[r-2]
	sheet.cell(row=r, column=5).value=precmz[r-2]
	sheet.cell(row=r, column=6).value=precchrg[r-2]
	sheet.cell(row=r, column=7).value=prodname[r-2]
	sheet.cell(row=r, column=8).value=prodformula[r-2]
	sheet.cell(row=r, column=9).value=prodadduct[r-2]
	sheet.cell(row=r, column=10).value=prodmz[r-2]
	sheet.cell(row=r, column=11).value=prodchrg[r-2]
	sheet.cell(row=r, column=12).value=explicitrt[r-2]
	sheet.cell(row=r, column=13).value=exrtwindow[r-2]
	sheet.cell(row=r, column=14).value=mzerror[r-2]
	sheet.cell(row=r, column=15).value=rettime[r-2]
	sheet.cell(row=r, column=16).value=area[r-2]
	sheet.cell(row=r, column=17).value=areanormalpercent[r-2]
	sheet.cell(row=r, column=18).value=background[r-2]
	sheet.cell(row=r, column=19).value=fwhm[r-2]
	sheet.cell(row=r, column=20).value=explicitrt[r-2]
	sheet.cell(row=r, column=21).value=rtstart[r-2]
	sheet.cell(row=r, column=22).value=rtend[r-2]
	sheet.cell(row=r, column=23).value=precoverlap[r-2]
	if manualfilter==1:
		sheet.cell(row=r, column=24).value=uncorrectedarea[r-2]
	r=r+1
wb.save('OzFAD1_4_input_DIA_Q.xlsx')
#print('Duplicates are removed, potential overlap (similar retention times) is flagged.')
#quit() #check ok
# excel file saved now contains explicit retention time and explicit retention time window for use in final skyline analysis
#######################################################end remove duplicates, flag overlap##################################################
############################################################################################################################################

############################################################################################################################################
# begin write precursor_results
shortlistfa=[]	#list of unsaturated FA species in precursorresults
allshortlistfa=[]
shortlistarea=[]	#list of areas representing sum of product fragments associated to precursor
wb=openpyxl.load_workbook('OzFAD1_4_input_DIA_Q.xlsx')
sheet=wb['precursor_results']
toprow=['PrecursorName', 'ProductName', 'Fatty acid species', 'Retention time', 'Area', 'Sum of Product Areas']
c=1
while c<7:
	sheet.cell(row=1, column=c).value=toprow[c-1]
	c=c+1
print('Writing precursor results is in progress.')
r=2
rw=2
fragareasum=0
while r<(len(mlistname)+2):
	findloss=prodname[r-2].find('loss')
	if findloss>0:
		fragareasum=fragareasum+area[r-2]
	else:
		wfragareasum=fragareasum
		if str(prodname[r-2][(len(prodname[r-2])-1)])=='r':
			fragareasum=0
	if str(prodname[r-2][(len(prodname[r-2])-1)])=='r':		#unsaturated FA, precursor
		sheet.cell(row=rw, column=1).value=precname[r-2]
		sheet.cell(row=rw, column=2).value=prodname[r-2]
		k=5	# for AMPP k=5 
		label=str('')
		while k<(len(precname[r-2])):
			label=label+str(precname[r-2][k])
			k=k+1
		sheet.cell(row=rw, column=3).value=label	#column C in excel file, e.g. 16:1_n-5_6.7
		shortlistfa.append(label)
		allshortlistfa.append(label)
		sheet.cell(row=rw, column=4).value=rettime[r-2]
		sheet.cell(row=rw, column=5).value=area[r-2]
		sheet.cell(row=rw, column=6).value=wfragareasum		#Sum of product transition areas
		shortlistarea.append(wfragareasum)
		rw=rw+1
	elif len(prodname[r-2])==9:		#saturated FA
		sheet.cell(row=rw, column=1).value=precname[r-2]
		sheet.cell(row=rw, column=2).value=prodname[r-2]
		k=5	# for AMPP k=5 
		label=str('')
		while k<(len(precname[r-2])):
			label=label+str(precname[r-2][k])
			k=k+1
		sheet.cell(row=rw, column=3).value=label
		allshortlistfa.append(label)	
		sheet.cell(row=rw, column=4).value=rettime[r-2]
		sheet.cell(row=rw, column=5).value=area[r-2]
		rw=rw+1
	r=r+1
wb.save('OzFAD1_4_input_DIA_Q.xlsx')
# end write precursor_results
# begin generate systematic names and check precursor_results for LipidMAPS IDs
# begin generate systematic name from Fatty acid species identifier
#shortlistfa=['09:1_n-6_7.83', '16:1_n-5_6.68', '16:1_n-7_6.68', '18:2_n-6_n-9_7.13', '22:6_n-3_n-6_n-9_n-12_n-15_n-18_7.15']
sysnamelist=[]
zeropreflist=['', '', '', 'prop', 'but', 'pent', 'hex', 'hept', 'oct', 'non']
decpreflist=['', 'un', 'do', 'tri', 'tetra', 'penta', 'hexa', 'hepta', 'octa', 'nona']
cospreflist=['ei', 'henei', 'do', 'tri', 'tetra', 'penta', 'hexa', 'hepta', 'octa', 'nona']
triacontpreflist=['', 'hen', 'do', 'tri', 'tetra', 'penta', 'hexa', 'hepta', 'octa', 'nona']
tetracontpreflist=['', 'hen', 'do', 'tri', 'tetra', 'penta', 'hexa', 'hepta', 'octa', 'nona']
sli=0
while sli<len(allshortlistfa):
	sysname=str()
	if int(allshortlistfa[sli][3])>0:	# unsaturated FA, generate delta db position(s)
		numdb=int(allshortlistfa[sli][3])
		kdb=numdb
		slfi=len(allshortlistfa[sli])-1
		multi=0
		while kdb>0:	# add delta positions
			go=1
			while go==1:
				if allshortlistfa[sli][slfi]=='_':
					go=0
				slfi=slfi-1
			if allshortlistfa[sli][slfi-1]=='-':
				cdb=int(allshortlistfa[sli][slfi])
			elif allshortlistfa[sli][slfi-2]=='-':
				cdb=int(allshortlistfa[sli][slfi])+10*(int(allshortlistfa[sli][slfi-1]))
			cdbd=int(allshortlistfa[sli][1])+10*(int(allshortlistfa[sli][0]))-cdb
			if multi==1:
				sysname=sysname+','+str(cdbd)+'Z'
			else:
				sysname=sysname+str(cdbd)+'Z'
			multi=1
			slfi=slfi-1
			kdb=kdb-1
		sysname=sysname+'-'
		if int(allshortlistfa[sli][0])==0:
			pref=str(zeropreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref
		elif int(allshortlistfa[sli][0])==1:
			pref=str(decpreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref+'dec'
		elif int(allshortlistfa[sli][0])==2:
			pref=str(cospreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref+'cos'
		elif int(allshortlistfa[sli][0])==3:
			pref=str(triacontpreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref+'triacont'
		elif int(allshortlistfa[sli][0])==4:
			pref=str(tetracontpreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref+'tetracont'
		else:
			sysname=sysname
		if int(allshortlistfa[sli][3])==1:
			sysname=sysname+'en'
		elif int(allshortlistfa[sli][3])==2:
			sysname=sysname+'adien'
		elif int(allshortlistfa[sli][3])==3:
			sysname=sysname+'atrien'
		elif int(allshortlistfa[sli][3])==4:
			sysname=sysname+'atetraen'
		elif int(allshortlistfa[sli][3])==5:
			sysname=sysname+'apentaen'
		elif int(allshortlistfa[sli][3])==6:
			sysname=sysname+'ahexaen'
	else:
		if int(allshortlistfa[sli][0])==0:
			pref=str(zeropreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref
		elif int(allshortlistfa[sli][0])==1:
			pref=str(decpreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref+'dec'
		elif int(allshortlistfa[sli][0])==2:
			pref=str(cospreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref+'cos'
		elif int(allshortlistfa[sli][0])==3:
			pref=str(triacontpreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref+'triacont'
		elif int(allshortlistfa[sli][0])==4:
			pref=str(tetracontpreflist[int(allshortlistfa[sli][1])])
			sysname=sysname+pref+'tetracont'
		sysname=sysname+'an'
	sysname=sysname+'oic+acid'
	sysnamelist.append(sysname)
	sli=sli+1
#print(sysnamelist)
# end generate systematic name from Fatty acid species identifier
fullsearch=len(allshortlistfa)
if searchlmsd==0:
	fullsearch=0
lipidmapsidlist=[]
linklist=[]
fsi=0
while fsi<fullsearch:
	searchfor=sysnamelist[fsi]
	# begin extract LipidMAPS ID
	#searchfor='6Z,9Z-octadecadienoic+acid'
	#print(searchfor)
	#searchfor='6Z,9Z-tetradecadienoic+acid'
	#searchfor='8Z,11Z-eicosadienoic+acid'
	urlpart='https://www.lipidmaps.org/search/quicksearch.php?Name='
	url=urlpart+searchfor
	#url='https://www.lipidmaps.org/search/quicksearch.php?Name=9Z-octadecenoic+acid'
	response = requests.get(url)
	#print(response)
	soup = BeautifulSoup(response.text, 'html.parser')
	results = soup.find_all('a')
	#print(len(results))
	k=0
	idfound=0
	while k<len(results):
		one_a_tag = soup.findAll('a')[k]
		link = one_a_tag['href']
		#print(link)
		slink=str(link)
		finder=slink.find('LMID')
		if finder>0:
			idfound=1
			#print(link)
			#print(finder)
			extractid=str()
			eid=finder
			go=0
			while eid<(len(link)-1):
				if link[eid]=='=':
					go=1
				if go==1:	
					extractid=extractid+str(link[eid+1])
				eid=eid+1
			
		k=k+1
	if idfound==1:
		print('LipidMAPS ID is:')
		print(extractid)
		lipidmapsidlist.append(extractid)
		clink='https://www.lipidmaps.org/databases/lmsd/'
		clink=clink+extractid+'?LMID='+extractid
		linklist.append(clink)
	else:
		print('Fatty acid not found.')
		lipidmapsidlist.append('Not found in LIPID MAPS.')
		linklist.append('_')
	# end extract LipidMAPS ID
	time.sleep(0.1)
	fsi=fsi+1
#if searchlmsd==0:
#	ok=1
#else:
wb=openpyxl.load_workbook('OzFAD1_4_input_DIA_Q.xlsx')
sheet=wb['precursor_results']
toprowadd=['Systematic Name', 'LIPID MAPS ID']
c=7
while c<9:
	sheet.cell(row=1, column=c).value=toprowadd[c-7]
	c=c+1
r=2
while r<(len(allshortlistfa)+2):
	csn=str(sysnamelist[r-2])
	ncsn=csn.replace('+', ' ')
	sheet.cell(row=r, column=7).value=ncsn
	if searchlmsd==0:
		ok=1
	else:
		sheet.cell(row=r, column=8).value=lipidmapsidlist[r-2]
		if linklist[r-2]=='_':
			r=r
		else:
			sheet.cell(row=r, column=8).hyperlink=linklist[r-2]
	r=r+1
wb.save('OzFAD1_4_input_DIA_Q.xlsx')
# end generate systematic names and check precursor_results for LipidMAPS IDs
############################################################################################################################################

###########################################################################################################################################################################
###########################################################################################################################################################################
###########################################################################################################################################################################
###########################################################################################################################################################################
############################################################################################################################################
# begin write barchart_results
#print(shortlistfa)
#print(shortlistarea)
wb=openpyxl.load_workbook('OzFAD1_4_input_DIA_Q.xlsx')
sheetbc=wb['ozid_barchart']
#sheetprecr=wb['precursorresults']
toprow=['FA', 'n-2 (Me)', 'n-2 (Bu)', 'n-2 (Other)', 'n-3 (Me)', 'n-3 (Bu)', 'n-3 (Other)', 'n-4 (Me)', 'n-4 (Bu)', 'n-4 (Other)', 'n-5 (Me)', 'n-5 (Bu)', 'n-5 (Other)', 'n-6 (Me)', 'n-6 (Bu)', 'n-6 (Other)',
'n-7 (Me)', 'n-7 (Bu)', 'n-7 (Other)', 'n-8 (Me)', 'n-8 (Bu)', 'n-8 (Other)', 'n-9 (Me)', 'n-9 (Bu)', 'n-9 (Other)', 'n-10 (Me)', 'n-10 (Bu)', 'n-10 (Other)', 'n-11 (Me)', 
'n-11 (Bu)', 'n-11 (Other)', 'n-12 (Me)', 'n-12 (Bu)', 'n-12 (Other)', 'n-13 (Me)', 'n-13 (Bu)', 'n-13 (Other)', 'n-14 (Me)', 'n-14 (Bu)', 'n-14 (Other)', 'n-15 (Me)', 'n-15 (Bu)', 
'n-15 (Other)', 'n-16 (Me)', 'n-16 (Bu)', 'n-16 (Other)']
ebclist=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
ibclist=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
c=1
while c<(len(toprow)+1):
	sheetbc.cell(row=1, column=c).value=toprow[c-1]
	c=c+1
barchartlist=[]
#ibclist=ebclist
i=0

if len(shortlistfa)>0:
	if shortlistfa[i][8]=='_':		#determine n position (dbn) of first FA
		dbn=int(shortlistfa[i][7])
	elif shortlistfa[i][9]=='_':
		dbn=10*int(shortlistfa[i][7])+int(shortlistfa[i][8])
	else:
		dbn=1
	dbn=dbn-1 #		# ##############################################################################
	if dbn>0:
		ibci=(3*dbn)-2
		ibclist[ibci]=shortlistarea[i]
		cfa=str(shortlistfa[i][0])+str(shortlistfa[i][1])+str(shortlistfa[i][2])+str(shortlistfa[i][3])
		ibclist[0]=cfa
		barchartlist.append(ibclist)
		#print(barchartlist)
		#quit()
	satfaonly=0
else:
	dbn=0
	satfaonly=1

i=0
while i<(len(shortlistfa)):
	cdbs=[]	#determine if species Me or Bu spaced or monounsat
	use=0
	if int(shortlistfa[i][3])==1:
		use=1
	else:
		fi=4		#determine if polyunsaturated FA Me/Bu spaced
		while fi<(len(shortlistfa[i])):
			if str(shortlistfa[i][fi])=='n':
				if str(shortlistfa[i][fi+3])=='_':
					cdb=int(shortlistfa[i][fi+2])
					cdbs.append(cdb)
				elif str(shortlistfa[i][fi+4])=='_':
					cdb=10*int(shortlistfa[i][fi+2])+int(shortlistfa[i][fi+3])
					cdb=int(cdb)
					cdbs.append(cdb)
			fi=fi+1
		dbi=0
		check=1
		while dbi<(len(cdbs)-1):
			if int(abs((cdbs[dbi])-(cdbs[dbi+1])) % 3)==0:	# Me or Bu or Hept spaced
				check=check
			else:
				check=0
			dbi=dbi+1
		if check==1:
			use=1
			cat=0
	if use==1:
		cat=0
		cfa=str(shortlistfa[i][0])+str(shortlistfa[i][1])+str(shortlistfa[i][2])+str(shortlistfa[i][3])
		if (len(cdbs))>1:
			if (cdbs[1]-cdbs[0])==6:
				cfa=cfa #+str(' (Bu)')
				cat=1 #dbn=dbn+13
			elif (cdbs[1]-cdbs[0])==9:
				cfa=cfa #+str(' (He)')
				cat=2 #dbn=dbn+26
	else:
		cfa=str(shortlistfa[i][0])+str(shortlistfa[i][1])+str(shortlistfa[i][2])+str(shortlistfa[i][3])
		cat=2 #dbn=dbn+26
	use=1
	if use==1:
		#print(cfa)
		#print(shortlistfa[i])
		#print(shortlistfa)
		#print(shortlistarea[i])
		#print(shortlistarea)
		#print(barchartlist)
		bcli=0
		found=0
		while bcli<(len(barchartlist)):
			if barchartlist[bcli][0]==cfa:
				if shortlistfa[i][8]=='_':
					dbn=int(shortlistfa[i][7])
				else:
					dbn=10*int(shortlistfa[i][7])+int(shortlistfa[i][8])
					dbn=int(dbn)
				#print(dbn)
				dbn=dbn-1			# because list starts with n-2, dbn (originally first double bond position from Me end becomes index in list)
				if cat==0:
					dbn=(3*dbn)-2
				elif cat==1:
					dbn=(3*dbn)-1
				elif cat==2:
					dbn=3*dbn
				barchartlist[bcli][dbn]=float(barchartlist[bcli][dbn])+float(shortlistarea[i])		# WHAT HAPPENS HERE
				found=1
			else:
				found=found
			bcli=bcli+1
		if found==0:
			ibclist=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
			ibclist[0]=cfa
			if shortlistfa[i][8]=='_':
				dbn=int(shortlistfa[i][7])
			elif shortlistfa[i][9]=='_':
				dbn=10*int(shortlistfa[i][7])+int(shortlistfa[i][8])
			dbn=dbn-1
			if cat==0:
				dbn=(3*dbn)-2
			elif cat==1:
				dbn=(3*dbn)-1
			elif cat==2:
				dbn=3*dbn
			ibci=dbn
			ibclist[ibci]=shortlistarea[i]
			barchartlist.append(ibclist)
			#print(barchartlist)
	i=i+1
#convert results to percentages
#print(barchartlist)
pbarchartlist=[]		#results for barchart in percentages
bcl=0
while bcl<(len(barchartlist)):
	cfal=1
	csum=0
	while cfal<(len(barchartlist[bcl])):
		csum=csum+barchartlist[bcl][cfal]
		cfal=cfal+1
	#print(barchartlist[bcl])
	#print(csum)
	pbcl=[]
	pbcl.append(barchartlist[bcl][0])
	cpi=1
	while cpi<(len(barchartlist[bcl])):
		if csum==0:
			pbcl.append(0)
		else:
			cp=(barchartlist[bcl][cpi]/csum)*100
			pbcl.append(cp)
		cpi=cpi+1
	pbarchartlist.append(pbcl)
	bcl=bcl+1
#print('pbarchartlist')
#print(pbarchartlist)
#write results in barchartlist in excel file
r=2
while r<(len(pbarchartlist)+2):
	c=1
	while c<(len(pbarchartlist[r-2])+1):
		sheetbc.cell(row=r, column=c).value=pbarchartlist[r-2][c-1]
		c=c+1
	r=r+1
#begin create bar chart in excel sheet
assigned=[]
assigned.append(toprow)
aslist=['index', 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2] # 0 is Me; 1 is Bu; 2 is Other
ias=0
while ias<(len(pbarchartlist)):
	assigned.append(aslist)
	ias=ias+1

mr=len(pbarchartlist)+1
if satfaonly==1:
	print('Only saturated fatty acids were found, results are saved in jpm_lipidomics_vpw11_5_final_output.xlsx')
	quit()
else:
	mc=len(pbarchartlist[0])
chart1 = BarChart()
chart1.type = "col"
chart1.style = 12
chart1.grouping = "stacked"
chart1.overlap = 100
chart1.title = "Fatty acids: Methylene and butylene interrupted and unusual double bond positions"
chart1.y_axis.title = 'Percentage of FA'
#chart1.x_axis.title = 'FA'
chart1.y_axis.scaling.max = 100   
data = Reference(sheetbc, min_col=2, min_row=1, max_row=mr, max_col=mc)
cats = Reference(sheetbc, min_col=1, min_row=2, max_row=mr)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4

gridoff=0
if gridoff==1:
	# begin turn majorGridlines off (setting colour white: FFFFFF)
	chart1.y_axis.majorGridlines.spPr = GraphicalProperties(noFill = 'True')
	chart1.y_axis.majorGridlines.spPr.ln = LineProperties(solidFill = 'FFFFFF')
	chart1.x_axis.majorGridlines = ChartLines()
	chart1.x_axis.majorGridlines.spPr = GraphicalProperties(noFill = 'True')
	chart1.x_axis.majorGridlines.spPr.ln = LineProperties(solidFill = 'FFFFFF')
	# end turn majorGridlines off (setting colour white: FFFFFF)

#colours=["0066cc", "660066", "003366", "ccffff", "800080", "00ccff", "993300", "99cc00", "000080", "ff00ff", "ff0000", "808000"] 
#colours=["4f81bd", "7f7f7f", "ffc000", "c4bd97", "8064a2", "4bacc6", "c0504d", "9bbb59", "1f497d", "f05eb5", "ff0000", "808000", "003300", 
#"0066cc", "660066", "003366", "ccffff", "800080", "00ccff", "993300", "99cc00", "000080", "ff00ff", "ff0000", "707000", "002200", 
#"0055cc", "550066", "003355", "00ffff", "700070", "00cc00", "883300", "88cc00", "000070", "0000ff", "ff0011", "606000", "001100"]
#cv=0
#while cv<len(colours):
#	s = chart1.series[cv]
#	s.graphicalProperties.line.solidFill = colours[cv]
#	s.graphicalProperties.solidFill = colours[cv]
#	cv=cv+1

#begin test new barchart with patterns
colors=['black', 'cornflowerBlue', 'lightGray', 'magenta', 'gold', 'mediumPurple', 'deepSkyBlue', 'sienna', 'limeGreen', 'lightYellow', 'orange', 'red', 'dkOliveGreen', 'ltCyan', 'blue']
colorschemebarchart=[]
cl=0
while cl<(len(colors)):
	csbc=str(colors[cl])
	colorschemebarchart.append(csbc)
	colorschemebarchart.append(csbc)
	colorschemebarchart.append(csbc)
	cl=cl+1

stbr=0
while stbr<(len(assigned[0])-1):
    clm=0
    while clm<(len(assigned)-1):      # 6 is number of columns in barchart (16:1, 17:1, ...)
        if assigned[clm+1][stbr+1]==0:
            s=chart1.series[stbr]   #define datapoint in column ([0] is n-3; [1] is n-4 ...)
            pt=DataPoint(idx=clm)     #define which column (e.g. idx=0 is 16:1; idx=1 is 17:1 ...)
            pt.graphicalProperties.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            s.dPt.append(pt)
        elif assigned[clm+1][stbr+1]==1:
            s=chart1.series[stbr]   #define datapoint in column ([0] is n-3; [1] is n-4 ...)
            pt=DataPoint(idx=clm)     #define which column (e.g. idx=0 is 16:1; idx=1 is 17:1 ...)
            fill=PatternFillProperties(prst='dkUpDiag')
            fill.foreground=ColorChoice(prstClr='black')
            fill.background=ColorChoice(prstClr=colorschemebarchart[stbr])
            pt.graphicalProperties.pattFill=fill
            pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            s.dPt.append(pt)
        else:
            s=chart1.series[stbr]   #define datapoint in column ([0] is n-3; [1] is n-4 ...)
            pt=DataPoint(idx=clm)     #define which column (e.g. idx=0 is 16:1; idx=1 is 17:1 ...)
            fill=PatternFillProperties(prst='dkVert')
            fill.foreground=ColorChoice(prstClr='black')
            fill.background=ColorChoice(prstClr=colorschemebarchart[stbr])
            pt.graphicalProperties.pattFill=fill
            pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[stbr])
            s.dPt.append(pt)
        clm=clm+1
    stbr=stbr+1
chart1.legend=None
#end test new barchart with patterns
chartposition=str('A')+str(len(pbarchartlist)+3)
sheetbc.add_chart(chart1, chartposition)
#end create bar chart in excel sheet
#begin create chart with legend
legendrows = [] 
lralist=['Categories', 'n-2', 'n-3', 'n-4', 'n-5', 'n-6', 'n-7', 'n-8', 'n-9', 'n-10', 'n-11', 'n-12', 'n-13', 'n-14', 'n-15', 'n-16', 'Me', 'Bu', 'Other']
legendrows.append(lralist)
lrlist=['FA', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
i=0
while i<(len(pbarchartlist)):
	fa=pbarchartlist[i][0]
	#lrlist[0]=fa
	lrblist=[]
	lrblist.append(fa)
	k=0
	while k<(len(lrlist)-1):
		lrblist.append(0)
		k=k+1
	legendrows.append(lrblist)
	i=i+1

r=70+len(pbarchartlist)
rinit=r
while r<(len(legendrows)+rinit):
	c=1
	while c<(len(legendrows[r-rinit])+1):
		sheetbc.cell(row=r, column=c).value=legendrows[r-rinit][c-1]
		c=c+1
	r=r+1
minr=rinit #+1
mr=r-1
mc=c-1
chart2 = BarChart()
chart2.type = "col"
chart2.style = 12
chart2.grouping = "stacked"
chart2.overlap = 100
chart2.title = chart1.title
chart2.y_axis.title = chart1.y_axis.title
#chart1.x_axis.title = 'Percentage of FA'
chart2.y_axis.scaling.max = 100   #############
data = Reference(sheetbc, min_col=2, min_row=minr, max_row=mr, max_col=mc)
cats = Reference(sheetbc, min_col=1, min_row=minr+1, max_row=mr)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.shape = 4

#colorscheme=['cornflowerBlue', 'lightGray', 'magenta', 'gold', 'mediumPurple', 'deepSkyBlue', 'sienna', 'limeGreen', 'lightYellow', 'orange', 'red', 'dkOliveGreen', 'ltCyan', 'white']
cl=0
while cl<13:
    sa = chart2.series[cl]
    fill=PatternFillProperties(prst='dkUpDiag')
    fill.foreground=ColorChoice(prstClr=colors[cl])
    fill.background=ColorChoice(prstClr=colors[cl])
    sa.graphicalProperties.pattFill=fill
    sa.graphicalProperties.line.solidFill=ColorChoice(prstClr=colors[cl])
    cl=cl+1

sa = chart2.series[13]
fill=PatternFillProperties(prst='dkUpDiag')
fill.foreground=ColorChoice(prstClr='white')
fill.background=ColorChoice(prstClr='white')
sa.graphicalProperties.pattFill=fill
sa.graphicalProperties.line.solidFill=ColorChoice(prstClr='black')

sa = chart2.series[14]
fill=PatternFillProperties(prst='dkUpDiag')
fill.foreground=ColorChoice(prstClr='black')
fill.background=ColorChoice(prstClr='white')
sa.graphicalProperties.pattFill=fill
sa.graphicalProperties.line.solidFill=ColorChoice(prstClr='white')

sa = chart2.series[15]
fill=PatternFillProperties(prst='dkVert')
fill.foreground=ColorChoice(prstClr='black')
fill.background=ColorChoice(prstClr='white')
sa.graphicalProperties.pattFill=fill
sa.graphicalProperties.line.solidFill=ColorChoice(prstClr='white')

chartposition=str('A')+str(35+len(pbarchartlist))
sheetbc.add_chart(chart2, chartposition)
#end create chart with legend
after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_10_'+fourlettcode+'_'
#filename='jpmlipidomics_vpw20_9_selected_final.xlsx'	#today+
#wb.save(filename)
#altfilename='jpmlipidomics_vpw20_5_final_quantification'
#wb.save(altfilename)



#print('Congratulations! Final results are saved as yyyy_mm_dd_10_xxxx_jpmlipidomics_vpw20_5_final_quantification.xlsx')
#print('The excel file jpmlipidomics_vpw20_5_final_quantification contains four worksheets:')
#print(' - The first contains a transition list with the selected transitions in the format readable by Skyline.')
#print(' - The second contains the data in a sorted structure, allowing the results to be plotted with color coding in relation to the different transitions from the double bonds.')
#print(' - The third contains a summary of the fatty acids found. The data specific to the fatty acids can for example be displayed in barcharts.')
#print(' - The fourth contains the calculated areas for each fatty acid that are displayed in the barchart.')
#print('    To complete the barchart shown, copy the legend from the "empty" barchart below into the barchart showing the data.')
#end write bar chart data and save in worksheet
###########################################################################################################################################################################
###########################################################################################################################################################################
print('Deconvolution of precursor chromatograms in progress ...')
###########################################################################################################################################################################
## begin read chromatograms, deconvolute precursor XICs, calculate updated relative quantification and barchart
convertfile=0
if convertfile==0:
	xictimesdf=pd.read_csv('skyl_xic_report_vpw20_6_times.csv', header=None, skiprows=1, nrows=1)
	xictimeslistfromdf=xictimesdf.values.tolist()
	xictimeslist=xictimeslistfromdf[0]		# contains times of XICs # actual list starts at index 8

segtrdf=pd.read_csv('skyl_xic_report_vpw20_6_intensities.csv', skiprows=1, header=None, low_memory=False)	#  nrows=4,
allxiclist=segtrdf.values.tolist()	# contains intensities of XICs # actual list starts at index 8

### in scope are   xictimeslist		allxiclist		swritelist (skyl report) and precoverlap (same length as swritelist, either 'ok' or 'Overlapping precursor')
cprecoverlap=precoverlap	# cprecoverlap is list that will gradually become all 'ok', as respective precursor XICs are deconvoluted

xicl=len(xictimeslist)
xicil=len(allxiclist)
#print(xicl)
#print(xicil)
#print(precoverlap)

deconpreclist=[] 	# list with either precursor integral or deconvoluted precursor integral
r=0
while r<ki:
	carea=area[r]
	deconpreclist.append(carea)
	r=r+1

exp=2.71828182845904523536028747135266249775724709369995  # mathematical constant e for gaussian expressions

plotcount=0
coznumintlist=[]
cprecnumintlist=[]
r=0
ki=len(swritelist[0])
while r<ki:
	e=swritelist[1][r] ## Precursorname		# begin determine which row to start (r) and to end (s)
	#print(e)
	s=r+1
	st=0
	while st<1:
		if s>(len(swritelist[1])-1):
			ne='stop_loop'
		else:
			ne=swritelist[1][s] #	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
	# begin get OzID product XICs and determine numerical integral and save in excel file in column 6 ##############################################################
	coznumint=0
	crtstart=float(swritelist[18][s])
	crtend=float(swritelist[19][s])
	t=r
	while t<s:
		ccoznumint=0
		crti=8
		while float(xictimeslist[crti])<crtstart:
			crti=crti+1
		while float(xictimeslist[crti])<crtend:
			ccoznumint=ccoznumint+allxiclist[t][crti]
			crti=crti+1
		coznumint=coznumint+ccoznumint
		t=t+1
	coznumintlist.append(coznumint)

	cprecnumint=0
	crti=8
	while float(xictimeslist[crti])<crtstart:
		crti=crti+1
	while float(xictimeslist[crti])<crtend:
		cprecnumint=cprecnumint+allxiclist[s][crti]
		crti=crti+1
	cprecnumintlist.append(cprecnumint)
	# OzID XIC Numerical Integral (sum of ozid integrals)

	# end get OzID product XICs and determine numerical integral and save in excel file in column 6 ################################################################
	if str(cprecoverlap[s])=='ok':
		# attach prec integral after isotope correction to deconpreclist
		r=r
	elif str(cprecoverlap[s])=='ok2':
		# attach prec integral after isotope correction to deconpreclist
		r=r	
	elif str(cprecoverlap[s])=='Overlapping precursor':
		# begin get all species for deconvolution, get XIC in range, deconvolute and modify deconpreclist according to results and set respective cprecoverlap to 'ok'
		dodeconpnlist=[]	#lists of precname, fwhm and index of species that need to be deconvoluted here
		dodeconfwhmlist=[]
		dodeconindexlist=[]
		dodeconrtstartlist=[]
		dodeconrtendlist=[]
		oziddep=[]
		cprecn=str(swritelist[1][r][5])+str(swritelist[1][r][6])+str(swritelist[1][r][7])+str(swritelist[1][r][8])
		q=0
		while q<ki:
			e=swritelist[1][q] ## Precursorname		# begin determine which row to start (q) and to end (w)
			#print(e)
			w=q+1
			st=0
			while st<1:
				if w>(len(swritelist[1])-1):
					ne='stop_loop'
				else:
					ne=swritelist[1][w] #	# Precursorname
				if ne==e:
					w=w+1
					st=0
				else:
					w=w-1
					st=1		# end determine w
			toz=q	# build oziddep list (same index as dependencies, but contains sum of ozid areas for the species)
			cozsum=0
			while toz<w:
				cozsum=cozsum+float(swritelist[13][toz])
				toz=toz+1
			
			if w==s:
				dodeconpnlist.append(str(swritelist[1][w]))
				dodeconfwhmlist.append(float(swritelist[16][w]))
				dodeconrtstartlist.append(float(swritelist[18][w]))
				dodeconrtendlist.append(float(swritelist[19][w]))
				dodeconindexlist.append(w)
				oziddep.append(cozsum)
			else:
				qprecn=str(swritelist[1][w][5])+str(swritelist[1][w][6])+str(swritelist[1][w][7])+str(swritelist[1][w][8])
				if qprecn==cprecn:
					if str(cprecoverlap[w])=='Overlapping precursor':
						dodeconpnlist.append(str(swritelist[1][w]))
						dodeconfwhmlist.append(float(swritelist[16][w]))	# FWHM
						dodeconrtstartlist.append(float(swritelist[18][w]))
						dodeconrtendlist.append(float(swritelist[19][w]))
						dodeconindexlist.append(w)
						oziddep.append(cozsum)

			q=w+1
		# begin get XIC in range for species listed in dodecon lists
		#print('## dodeconpnlist ##')
		#print(dodeconpnlist)
		#print('## dodeconfwhmlist ##')
		#print(dodeconfwhmlist)
		#print('## dodeconrtstartlist ##')
		#print(dodeconrtstartlist)
		#print('## dodeconrtendlist ##')
		#print(dodeconrtendlist)
		#print('## dodeconindexlist ##')
		#print(dodeconindexlist)
		xicrangemin=min(dodeconrtstartlist)		# range of XIC for deconvolution is defined by integration limits set manually in Skyline
		xicrangemax=max(dodeconrtendlist)
		#print('## xicrangemin ##')
		#print(xicrangemin)
		#print('## xicrangemax ##')
		#print(xicrangemax)
		undeconxicx=[]
		undeconxicy=[]
		f=8
		while float(xictimeslist[f])<xicrangemin:
			f=f+1
		while float(xictimeslist[f])<xicrangemax:
			cy=float(allxiclist[int(dodeconindexlist[0])][f])
			cx=float(xictimeslist[f])
			undeconxicy.append(cy)
			undeconxicx.append(cx)
			f=f+1
		#print('## undeconxicx ##')
		#print(undeconxicx)
		#print('## undeconxicy ##')
		#print(undeconxicy)
		# begin build gausslist (parameters for gaussian sum expression) with starting parameters
		gausslist=[]
		dependencies=[]
		depratio=[]
		undeconxicozidy=[]
		glt=0
		while glt<(len(dodeconpnlist)):
			glist=[]
			gb=(dodeconrtstartlist[glt]+dodeconrtendlist[glt])/2		# RT at mid peak 
			glg=0
			while undeconxicx[glg]<gb:
				ga=undeconxicy[glg]
				glg=glg+1
			ga=(ga+undeconxicy[glg])/2
			#gc=0.5*mean([median(dodeconfwhmlist), mean(dodeconfwhmlist)])	retired calculation for gaussian width start parameter
			gc=0.0004*float(gb)+0.0119
			gc=round(gc, 4)

			glist.append(ga)
			glist.append(gb)
			glist.append(gc)
			gausslist.append(glist)
			dependencies.append(-1)

			gdsum=0
			di=1
			while di<((int(dodeconpnlist[glt][8])*2)+1):
				fg=8
				while float(xictimeslist[fg])<gb:	#get gd as ozid intensity according to di (ozid transition) and glt (fa species)
					fg=fg+1
				gd=allxiclist[int(dodeconindexlist[glt])-di][fg]
				gd=(gd+allxiclist[int(dodeconindexlist[glt])-di][fg+1])/2
				gdsum=gdsum+gd
				di=di+1
			cozidy=gdsum/(int(dodeconpnlist[glt][8])*2)		# mean of intensities at mid peak for all ozid product XICs
			undeconxicozidy.append(cozidy)

			glt=glt+1
		#print('## gausslist before correction of starting intensity ##')
		#print(gausslist)
		depratio=dependencies
		# begin correct start value of ga for species with OzID transition that are close in RT to other OzID species 
		gicorrlist=[]	# later contains indices of dodeconpnlist that need to be limited for deconvolution
		gicorrgrouplist=[]
		groupindex=0
		glt=0
		while glt<(len(dodeconpnlist)):
			gipos=0
			glk=0
			while glk<(len(dodeconpnlist)):
				if glk==glt:
					glk=glk
				else:
					if abs(gausslist[glt][1]-gausslist[glk][1])<0.015:
						# species glt is too close to glk for reliable deconvolution, need to rely on OzID ratio here to set limits for deconvolution
						gipos=1
						if glt in gicorrlist:
							glt=glt
						else:
							gicorrlist.append(glt)
							gicorrgrouplist.append(groupindex)
						if glk in gicorrlist:
							glk=glk
						else:
							gicorrlist.append(glk)
							if glt in gicorrlist:
								groupindex=gicorrgrouplist[len(gicorrgrouplist)-1]
								gicorrgrouplist.append(groupindex)
							else:
								gicorrgrouplist.append(groupindex)
				glk=glk+1
			if gipos==1:
				groupindex=groupindex+1
			glt=glt+1
		#print('## undeconxicozidy mean intensity of ozid transitions ##')
		#print(undeconxicozidy)
		#print('## gicorrlist ##')
		#print(gicorrlist)
		#print('## gicorrgrouplist ##')
		#print(gicorrgrouplist)

		gic=0
		while gic<(len(gicorrgrouplist)):
			ozidtotal=0
			groupindex=gicorrgrouplist[gic]
			cint=[]
			dic=gic
			while gicorrgrouplist[gic]==groupindex:
				ozidtotal=ozidtotal+undeconxicozidy[gicorrlist[gic]]
				cnt=float(gausslist[int(gicorrlist[gic])][0])
				cint.append(cnt)
				if gic<(len(gicorrgrouplist)-1):
					gic=gic+1
				else:
					gic=gic
					groupindex=-1

			gic=dic
			if groupindex==-1:
				groupindex=gicorrgrouplist[gic]
			while gicorrgrouplist[gic]==groupindex:
				gausslist[gicorrlist[gic]][0]=max(cint)*float(undeconxicozidy[gicorrlist[gic]])/float(ozidtotal)		# correction of intensity in gausslist

				dependencies[gicorrlist[gic]]=groupindex

				if gic<(len(gicorrgrouplist)-1):
					gic=gic+1
				else:
					gic=gic
					groupindex=-1
			if groupindex==-1:
				gic=gic+1
			if gic<(len(gicorrgrouplist)):
				groupindex=gicorrgrouplist[gic]
			else:
				groupindex=groupindex
			gic=gic
		#print('## gausslist after correction of starting intensity ##')
		#print(gausslist)
		# end correct start value of ga for species with low abundance OzID transition that are close in RT to high abundance OzID species 
		# begin save data for current deconvolution in excel file and generate plot for quality control
		displaydecon=0
		if displaydecon==1:
			if plotcount==0:
				#begin create excel sheet for analysis results
				wbd = Workbook(write_only=True)
				wbd.save('OzFAD1_4_deconv_raw.xlsx')
				wbd=openpyxl.load_workbook('OzFAD1_4_deconv_raw.xlsx')
				#wbd.save('jpmlipidomics_vpw20_5_quantification_deconvolution1.xlsx')
				# workbook created
			else:
				wbd=openpyxl.load_workbook('OzFAD1_4_deconv_raw.xlsx')
			plotcount=plotcount+1
			#now write current precursor XIC in range and gauss peaks, incl sum into it for plotting
			sheetname=str(dodeconpnlist[0][5])+str(dodeconpnlist[0][6])+'_'+str(dodeconpnlist[0][8])
			ws = wbd.create_sheet(sheetname)
			wbd.save('OzFAD1_4_deconv_raw.xlsx')
			wbd=openpyxl.load_workbook('OzFAD1_4_deconv_raw.xlsx')
			sheet=wbd[sheetname]

			toprow=['Start parameter name', 'Start parameter (deconvolution)', '_', 'RT / min', 'XIC precursor', 'Sum of gaussians']
			tp=0		# complete and write toprow
			while tp<(len(dodeconpnlist)):
				toprow.append(dodeconpnlist[tp])
				tp=tp+1
			c=1
			while c<(len(toprow)+1):
				sheet.cell(row=1, column=c).value=toprow[c-1]
				c=c+1
			glw=0		# write first and second column with parameter description and start parameters for gaussian deconvolution
			rww=2
			while glw<(len(gausslist)):
				gll=0
				while gll<3:
					param=str(int(gll+1))+'_('+str(dodeconpnlist[glw])+')'
					sheet.cell(row=rww, column=1).value=param
					paramv=float(gausslist[glw][gll])
					sheet.cell(row=rww, column=2).value=paramv
					rww=rww+1
					gll=gll+1
				glw=glw+1
			rtw=0		# write original precursor XIC
			while rtw<(len(undeconxicx)):
				param=float(undeconxicx[rtw])
				rww=rtw+2
				sheet.cell(row=rww, column=4).value=param
				paramv=float(undeconxicy[rtw])
				sheet.cell(row=rww, column=5).value=paramv
				vdecon=0
				pdg=0
				while pdg<(len(gausslist)):
					vdecon=vdecon+(gausslist[pdg][0]*exp**(-((param-gausslist[pdg][1])**2)/(2*(gausslist[pdg][2])**2)))
					pdg=pdg+1
				sheet.cell(row=rww, column=6).value=vdecon
				videcon=0
				pdi=0
				while pdi<(len(gausslist)):
					videcon=(gausslist[pdi][0]*exp**(-((param-gausslist[pdi][1])**2)/(2*(gausslist[pdi][2])**2)))
					cli=int(7+pdi)
					sheet.cell(row=rww, column=cli).value=videcon
					pdi=pdi+1
				rtw=rtw+1
			wbd.save('OzFAD1_4_deconv_raw.xlsx')
		# end save data for current deconvolution in excel file and generate plot for quality control
		#######################################################################################################################################
		#######################################################################################################################################
		#######################################################################################################################################
		# begin actual deconvolution based on starting values in gausslist

		# begin define upperboundlist and lowerboundlist with limits to parameters for gaussians; dependencies and depratio
		upperboundlist=[]
		lowerboundlist=[]
		gli=0
		while gli<(len(gausslist)):
			li=0.5*float(gausslist[gli][0])
			ui=3*float(gausslist[gli][0])
			lrt=float(gausslist[gli][1])-(float(gausslist[gli][2]))
			urt=float(gausslist[gli][1])+(float(gausslist[gli][2]))
			lfwhm=0.75*float(gausslist[gli][2])
			ufwhm=3*float(gausslist[gli][2])
			gllist=[]
			gulist=[]
			gllist.append(li)
			gllist.append(lrt)
			gllist.append(lfwhm)
			gulist.append(ui)
			gulist.append(urt)
			gulist.append(ufwhm)
			upperboundlist.append(gllist)
			lowerboundlist.append(gulist)
			gli=gli+1
		#print('upperboundlist and lowerboundlist:')
		#print(upperboundlist)
		#print(lowerboundlist)	
		#print(dependencies)
		ndependencies=[]
		ndp=0
		while ndp<(len(dependencies)):
			ndpi=int(dependencies[ndp])
			ndependencies.append(ndpi)
			ndp=ndp+1
		# correct list dependencies to set highest intensity values in each group to -1 and dependent values to index of associated independent value
		olddep=0
		if olddep==1:
			dd=0
			while dd<(len(dependencies)):			# works, independent species is the one with highest ozid transition within group ! (may not always be correct !?) BROKEN FOR 18:2
				dp=dd
				ilist=[]
				if dependencies[dd]>-1:
					while dp<(len(dependencies)):
						if dependencies[dp]==dependencies[dd]:
							ilist.append(gausslist[dp][0])
							if max(ilist)==gausslist[dp][0]:
								indep=dp
							dp=dp+1
						else:
							dp=len(dependencies)
					dp=dd
					print(ilist)
					while dp<(len(dependencies)):
						if dependencies[dp]==dependencies[dd]:
							ndependencies[dp]=indep
							if dp==indep:
								ndependencies[dp]=-1
							dp=dp+1
							dr=dp#-1
						else:
							dr=dp#+1
							dp=len(dependencies)
					if dr>(len(dependencies)-1):
						dr=dr-1
					if dependencies[dr]==-1:
						dd=dr+1
					else:
						dd=dr+1 #or dd=dr+1 ?
				else:
					dd=dd+1
		else:
			ndependencies=copy.deepcopy(dependencies)
			dd=0
			while dd<(len(dependencies)):
				if dependencies[dd]>-1:
					dp=0
					cgind=[]
					cgozid=[]
					while dp<(len(dependencies)):
						if dependencies[dd]==dependencies[dp]:
							if dd in cgind:
								dd=dd
							else:
								cgind.append(dd)
								cgozid.append(oziddep[dd])		# wrong index !!! check building of oziddep !
							if dp in cgind:
								dp=dp
							else:
								cgind.append(dp)
								cgozid.append(oziddep[dp])								
						dp=dp+1
					ozd=0
					while ozd<(len(cgozid)):
						if cgozid[ozd]==max(cgozid):
							ndependencies[cgind[ozd]]=-1
							maxind=cgind[ozd]
						ozd=ozd+1
					ozd=0
					while ozd<(len(cgozid)):
						if cgozid[ozd]==max(cgozid):
							ozd=ozd
						else:
							ndependencies[cgind[ozd]]=maxind
						ozd=ozd+1
				dd=dd+1	

		#print('ndependencies')
		#print(ndependencies)
		#print('oziddep')
		#print(oziddep)
		depratio=[]
		nd=0
		while nd<(len(ndependencies)):
			if ndependencies[nd]>-1:
				dpr=oziddep[nd]/oziddep[ndependencies[nd]]
				depratio.append(dpr)
			else:
				depratio.append(1)
			nd=nd+1
		#print('depratio')
		#print(depratio)
		# end define upperboundlist and lowerboundlist with limits to parameters for gaussians; dependencies and depratio
		initialgausslist=copy.deepcopy(gausslist)
		#print('Running deconvolution algorithm')
		def gaussresidual():
			# begin determine residual (sum of squared abs differences) between fitted curve and experimental XIC
			residual=0
			resx=0
			while resx<(len(undeconxicx)):		# go through XIC and add up residual
				cxicy=undeconxicy[resx]
				cgy=0		# 
				cgg=0
				while cgg<(len(gausslist)):
					cgy=cgy+float(gausslist[cgg][0]*exp**(-1*((undeconxicx[resx]-gausslist[cgg][1])**2)/(2*(gausslist[cgg][2])**2)))
					cgg=cgg+1
				residual=residual+float(((cxicy-cgy)**2)*cxicy**2)		# least squares times intensity (focus on close fit for range with high intensity)
				resx=resx+1
			# end determine residual (sum of squared abs differences) between fitted curve and experimental XIC
			return residual
		itercount=0
		go=1
		while go==1:
			cresidual=gaussresidual()
			#print(cresidual)
			itercount=itercount+1
			gof=0		#1 if fwhm can be optimized
			# begin vary fwhm and check
			cfwhm=float(gausslist[0][2])
			nfwhm=cfwhm+0.001
			if nfwhm<upperboundlist[0][2]:
				nfwhm=nfwhm
			else:
				nfwhm=cfwhm
			cgausslist=copy.deepcopy(gausslist)		#cgausslist is gausslist without varied parameters, if not verfied to be better // gausslist may contain varied parameters
			cgg=0
			while cgg<(len(gausslist)):
				gausslist[cgg][2]=nfwhm
				cgg=cgg+1
			nresidual=gaussresidual()
			if nresidual<cresidual:
				cgausslist=copy.deepcopy(gausslist)	
				gof=1
			else:
				nfwhm=cfwhm-0.001
				if nfwhm>lowerboundlist[0][2]:
					nfwhm=nfwhm
				else:
					nfwhm=cfwhm
				cgg=0
				while cgg<(len(gausslist)):
					gausslist[cgg][2]=nfwhm
					cgg=cgg+1
				nresidual=gaussresidual()
				if nresidual<cresidual:
					cgausslist=copy.deepcopy(gausslist)
					gof=1
				else:
					gausslist=copy.deepcopy(cgausslist)
					gof=0
			# end vary fwhm and check

			goi=0		#1 if intensity can be optimized
			gop=0		#1 if positions can be optimized
			igg=0
			while igg<(len(initialgausslist)):
				# begin vary amplitude and check
				cresidual=gaussresidual()
				cgausslist=copy.deepcopy(gausslist)	
				camp=float(gausslist[igg][0])
				namp=camp+(camp*0.05)
				if namp>upperboundlist[igg][0]:
					namp=camp
				gausslist[igg][0]=namp
				nresidual=gaussresidual()
				if nresidual<cresidual:
					cgausslist=copy.deepcopy(gausslist)	
					goi=1
				else:
					namp=camp-(camp*0.05)
					if namp<lowerboundlist[igg][0]:
						namp=camp
					gausslist[igg][0]=namp
					nresidual=gaussresidual()
					if nresidual<cresidual:
						cgausslist=copy.deepcopy(gausslist)
						goi=1
					else:
						gausslist=copy.deepcopy(cgausslist)
						goi=goi
				# end vary amplitude and check

				# begin vary rt position and check
				cresidual=gaussresidual()
				cgausslist=copy.deepcopy(gausslist)	
				crtp=float(gausslist[igg][1])
				nrtp=crtp+(0.01)
				if nrtp<upperboundlist[igg][1]:
					nrtp=nrtp
				else:
					nrtp=crtp
				gausslist[igg][0]=nrtp 
				nresidual=gaussresidual()
				if nresidual<cresidual:
					cgausslist=copy.deepcopy(gausslist)	
					gop=1
				else:
					nrtp=crtp-(0.01)
					if nrtp>lowerboundlist[igg][1]:
						nrtp=nrtp
					else:
						nrtp=crtp
					gausslist[igg][1]=nrtp
					nresidual=gaussresidual()
					if nresidual<cresidual:
						cgausslist=copy.deepcopy(gausslist)
						gop=1
					else:
						gausslist=copy.deepcopy(cgausslist)
						gop=gop
				# end vary rt position and check
				igg=igg+1
			go=0
			if gof==1:
				go=1
			if gop==1:
				go=1
			if goi==1:
				go=1
			#print(itercount)
		#end of optimization
		#print('Number of iterations in first pass deconvolution algorithm')
		#print(itercount)
		itercount=0
		# begin fine tuning of deconvolution with smaller steps
		go=1
		while go==1:
			cresidual=gaussresidual()
			#print(cresidual)
			itercount=itercount+1
			gof=0		#1 if fwhm can be optimized
			# begin vary fwhm and check
			cfwhm=float(gausslist[0][2])
			nfwhm=cfwhm+0.0005
			if nfwhm<upperboundlist[0][2]:
				nfwhm=nfwhm
			else:
				nfwhm=cfwhm
			cgausslist=copy.deepcopy(gausslist)		#cgausslist is gausslist without varied parameters, if not verfied to be better // gausslist may contain varied parameters
			cgg=0
			while cgg<(len(gausslist)):
				gausslist[cgg][2]=nfwhm
				cgg=cgg+1
			nresidual=gaussresidual()
			if nresidual<cresidual:
				cgausslist=copy.deepcopy(gausslist)	
				gof=1
			else:
				nfwhm=cfwhm-0.0005
				if nfwhm>lowerboundlist[0][2]:
					nfwhm=nfwhm
				else:
					nfwhm=cfwhm
				cgg=0
				while cgg<(len(gausslist)):
					gausslist[cgg][2]=nfwhm
					cgg=cgg+1
				nresidual=gaussresidual()
				if nresidual<cresidual:
					cgausslist=copy.deepcopy(gausslist)
					gof=1
				else:
					gausslist=copy.deepcopy(cgausslist)
					gof=gof
			# end vary fwhm and check

			goi=0		#1 if intensity can be optimized
			gop=0		#1 if positions can be optimized
			igg=0
			while igg<(len(initialgausslist)):
				# begin vary amplitude and check
				cresidual=gaussresidual()
				cgausslist=copy.deepcopy(gausslist)	
				camp=float(gausslist[igg][0])
				namp=camp+(camp*0.005)
				if namp<upperboundlist[igg][0]:
					namp=namp
				else:
					namp=camp
				gausslist[igg][0]=namp
				nresidual=gaussresidual()
				if nresidual<cresidual:
					cgausslist=copy.deepcopy(gausslist)	
					goi=1
				else:
					namp=camp-(camp*0.005)
					if namp>lowerboundlist[igg][0]:
						namp=namp
					else:
						namp=camp
					gausslist[igg][0]=namp
					nresidual=gaussresidual()
					if nresidual<cresidual:
						cgausslist=copy.deepcopy(gausslist)
						goi=1
					else:
						gausslist=copy.deepcopy(cgausslist)
						goi=goi
				# end vary amplitude and check

				# begin vary rt position and check
				cresidual=gaussresidual()
				cgausslist=copy.deepcopy(gausslist)	
				crtp=float(gausslist[igg][1])
				nrtp=crtp+(0.002)
				if nrtp<upperboundlist[igg][1]:
					nrtp=nrtp
				else:
					nrtp=crtp
				gausslist[igg][0]=nrtp 
				nresidual=gaussresidual()
				if nresidual<cresidual:
					cgausslist=copy.deepcopy(gausslist)	
					gop=1
				else:
					nrtp=crtp-(0.002)
					if nrtp>lowerboundlist[igg][1]:
						nrtp=nrtp
					else:
						nrtp=crtp
					gausslist[igg][1]=nrtp
					nresidual=gaussresidual()
					if nresidual<cresidual:
						cgausslist=copy.deepcopy(gausslist)
						gop=1
					else:
						gausslist=copy.deepcopy(cgausslist)
						gop=gop
				# end vary rt position and check
				igg=igg+1
			go=0
			if gof==1:
				go=1
			if gop==1:
				go=1
			if goi==1:
				go=1
			#print(itercount)
		#end of optimization
		#print('Number of iterations in fine tuning deconvolution algorithm')
		#print(itercount)

		#quit()
		# end actual deconvolution based on starting values

		# begin change amplitude to value derived from OzID integrals in relation to max OzID integral
		amg=0
		while amg<len(gausslist):
			if oziddep[amg]==max(oziddep):
				maxampl=gausslist[amg][0]
				maxozid=oziddep[amg]
			amg=amg+1
		amg=0
		while amg<len(gausslist):
			if oziddep[amg]==max(oziddep):
				ok=1
			else:
				gausslist[amg][0]=round(maxampl/maxozid*oziddep[amg],0)
			amg=amg+1

		# end change amplitude to value derived from OzID integrals in relation to max OzID integral
		#######################################################################################################################################
		# begin save data for current deconvolution in excel file and generate plot for quality control
		colindex=['Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']
		displaydecon=1
		if displaydecon==1:
			if plotcount==0:
				#begin create excel sheet for analysis results
				wbd = Workbook(write_only=True)
				wbd.save('OzFAD1_4_DIA_deconv_raw.xlsx')
				wbd=openpyxl.load_workbook('OzFAD1_4_DIA_deconv_raw.xlsx')
				# workbook created
			else:
				wbd=openpyxl.load_workbook('OzFAD1_4_DIA_deconv_raw.xlsx')
			plotcount=plotcount+1
			#now write current precursor XIC in range and gauss peaks, incl sum into it for plotting
			sheetname='DC_'+str(dodeconpnlist[0][5])+str(dodeconpnlist[0][6])+'_'+str(dodeconpnlist[0][8])
			ws = wbd.create_sheet(sheetname)
			wbd.save('OzFAD1_4_DIA_deconv_raw.xlsx')
			wbd=openpyxl.load_workbook('OzFAD1_4_DIA_deconv_raw.xlsx')
			sheet=wbd[sheetname]
			nortv=str(len(undeconxicx))
			toprow=['Start parameter name', 'Start parameter (deconvolution)', 'Target fraction', 'Current fraction', 'OzID Integral', 'Parameter (OzID ratio)', 'max_FA', '_', '_', '_', '# RT values:', nortv, '_', 'RT / min', 'XIC precursor', 'Sum of gaussians']
			tp=0		# complete and write toprow
			while tp<(len(dodeconpnlist)):
				toprow.append(dodeconpnlist[tp])
				tp=tp+1
			c=1
			while c<(len(toprow)+1):
				sheet.cell(row=1, column=c).value=toprow[c-1]
				c=c+1
			glw=0		# write first and second column with parameter description and start parameters for gaussian deconvolution
			rww=2
			while glw<(len(gausslist)):
				gll=0
				while gll<3:
					param=str(int(gll+1))+'_('+str(dodeconpnlist[glw])+')'
					sheet.cell(row=rww, column=1).value=param
					if gll==0:
						paramv=round(float(gausslist[glw][gll]),0)
					elif gll==1:
						paramv=round(float(gausslist[glw][gll]),3)
					else:
						paramv=round(float(gausslist[glw][gll]),4)
					if glw>0:
						if gll==2:
							paramv='=B4'
							sheet.cell(row=rww, column=2).value=paramv
							#fillcell='B'+str(int(glw*3+4))
							#sheet[fillcell].fill=PatternFill(patternType='solid', fgColor='00C0C0C0')
						else:
							sheet.cell(row=rww, column=2).value=paramv
							#fillcell='B'+str(int(glw*3+4))
							#sheet[fillcell].fill=PatternFill(patternType='solid', fgColor='00C0C0C0') #(start_color='00C0C0C0', end_color='00C0C0C0', pattern_type='solid') #(fgColor='00C0C0C0', fill_type='solid')
							#sheet.cell(row=rww, column=2).value=paramv
					else:
						sheet.cell(row=rww, column=2).value=paramv
						#sheet.fill=PatternFill(bgColor='00C0C0C0', fill_type='solid')
					rww=rww+1
					gll=gll+1
				glw=glw+1
			rtw=0		# write original precursor XIC
			while rtw<(len(undeconxicx)):
				param=float(undeconxicx[rtw])
				rww=rtw+2
				sheet.cell(row=rww, column=14).value=param
				paramv=float(undeconxicy[rtw])
				sheet.cell(row=rww, column=15).value=paramv

				wvalue=0
				if wvalue==1:
					vdecon=0
					pdg=0
					while pdg<(len(gausslist)):
						vdecon=vdecon+(gausslist[pdg][0]*exp**(-((param-gausslist[pdg][1])**2)/((gausslist[pdg][2])**2)))
						pdg=pdg+1
					sheet.cell(row=rww, column=16).value=vdecon
				# begin enter excel formula for sum of gaussians
				lastlett=colindex[len(gausslist)-1]
				vdecon='=SUM(Q'+str(rww)+':'+str(lastlett)+str(rww)+')'
				sheet.cell(row=rww, column=16).value=vdecon
				# end enter excel formula for sum of gaussians
				videcon=0
				pdi=0
				while pdi<(len(gausslist)):
					#videcon=(gausslist[pdi][0]*exp**(-((param-gausslist[pdi][1])**2)/((gausslist[pdi][2])**2)))
					cli=int(17+pdi)
					# begin enter excel formula for gaussians
					aval=(pdi*3)+2
					bval=(pdi*3)+3
					cval=(pdi*3)+4
					videcon='=B'+str(aval)+'*EXP((-1*(N'+str(rww)+'-B'+str(bval)+')^2)/(2*B'+str(cval)+'^2))'
					sheet.cell(row=rww, column=cli).value=videcon
					# end enter excel formula for gaussians
					pdi=pdi+1
				rtw=rtw+1
			# begin edit fields for numerical integration
			lux=len(undeconxicx)
			pdi=0
			while pdi<(len(gausslist)):
				clett=colindex[pdi]
				numint='=SUM('+str(clett)+'2'+':'+str(clett)+str(lux+1)+')'
				sheet.cell(row=lux+5, column=17+pdi).value=numint		# numerical integral of gaussian peak
				fadescr='='+str(clett)+'1'
				sheet.cell(row=lux+4, column=17+pdi).value=fadescr		# label of fatty acid
				sheet.cell(row=lux+6, column=17+pdi).value=str(pdi)		# Index
				sheet.cell(row=lux+7, column=17+pdi).value=str(ndependencies[pdi])		# ndependencies 
				sheet.cell(row=lux+8, column=17+pdi).value=str(depratio[pdi])		# depratio (Target fraction)

				sheet.cell(row=pdi*3+2, column=3).value=str(depratio[pdi])		# depratio (Target fraction)
				sheet.cell(row=pdi*3+2, column=5).value=str(oziddep[pdi])		# oziddep (sum of ozid integrals)
				if oziddep[pdi]==max(oziddep):
					sheet.cell(row=pdi*3+2, column=7).value='max'
				#if str(swritelist[2][r])==dodeconpnlist[pdi]:																	## old, do not activate
				#	sheet.cell(row=pdi*3+2, column=6).value=str(coznumint[pdi])		# OzID XIC Numerical Integral (sum of ozid integrals)

				if depratio[pdi]<1:
					tshift=ndependencies[pdi]-pdi
					tlett=colindex[pdi+tshift]
					cfrac='='+str(clett)+str(int(lux+5))+'/'+str(tlett)+str(int(lux+5))
					sheet.cell(row=lux+10, column=17+pdi).value=cfrac		# Current Fraction
					sheet.cell(row=pdi*3+2, column=4).value=cfrac		# Current Fraction

				pdi=pdi+1
			sheet.cell(row=lux+4, column=14).value='Fatty acid'
			sheet.cell(row=lux+5, column=14).value='Integral gaussian'
			sheet.cell(row=lux+6, column=14).value='Index'
			sheet.cell(row=lux+7, column=14).value='Dependency'
			sheet.cell(row=lux+8, column=14).value='Target Fraction'
			sheet.cell(row=lux+10, column=14).value='Current Fraction'

			# end edit fields for numerical integration
			

			# begin make plot for deconvolution
			chart1 = ScatterChart()
			#chart1.type = "col"
			chart1.style = 13
			#chart1.grouping = "stacked"
			#chart1.overlap = 100
			chart1.title = "Deconvolution of precursor chromatogram"
			chart1.y_axis.title = 'Intensity / a.u.'
			chart1.x_axis.title = 'RT / min'
			#chart1.y_axis.scaling.max = 100 
			mr=1+len(undeconxicx)
			mc=16+len(gausslist)  
			xvalues=Reference(sheet, min_col=14, min_row=2, max_row=mr)
			for i in range(14, mc+1):
				values=Reference(sheet, min_col=i, min_row=1, max_row=mr)
				series=Series(values, xvalues, title_from_data=True)
				chart1.series.append(series)
			#data = Reference(sheet, min_col=14, min_row=1, max_row=mr, max_col=mc)
			#cats = Reference(sheet, min_col=14, min_row=2, max_row=mr)
			#chart1.add_data(data, titles_from_data=True)
			#chart1.set_categories(cats)
			#chart1.shape = 4
			chartposition=str('H3')
			sheet.add_chart(chart1, chartposition)


			# end make plot for deconvolution
			wbd.save('OzFAD1_4_DIA_deconv_raw.xlsx')
		# end save data for current deconvolution in excel file and generate plot for quality control
		#######################################################################################################################################
		#######################################################################################################################################
	# set cprecoverlap to 'ok2', so that deconvolution does not need to be carried out multiple times
	if str(cprecoverlap[s])=='Overlapping precursor':
		rset=0
		while rset<ki:
			e=swritelist[1][rset] ## Precursorname		# begin determine which row to start (rset) and to end (sset)
			#print(e)
			sset=rset+1
			st=0
			while st<1:
				if sset>(len(swritelist[1])-1):
					ne='stop_loop'
				else:
					ne=swritelist[1][sset] #	# Precursorname
				if ne==e:
					sset=sset+1
					st=0
				else:
					sset=sset-1
					st=1		# end determine sset	
			if str(cprecoverlap[sset])=='Overlapping precursor':
				ddn=0
				while ddn<(len(dodeconpnlist)):
					if str(swritelist[1][sset])==str(dodeconpnlist[ddn]):
						cprecoverlap[sset]='ok2'
					ddn=ddn+1
			rset=sset+1
	#quit()
	r=s+1

# begin save numerical integral of ozid products in precursor_results in new sheet Summary_OzID_Integrals
wbd=openpyxl.load_workbook('OzFAD1_4_DIA_deconv_raw.xlsx')
sheetname='Summary_OzID_Integrals'
ws = wbd.create_sheet(sheetname)
#wbd.save('jpmlipidomics_vpw20_5_quantification_deconvolution1.xlsx')
#wbd=openpyxl.load_workbook('jpmlipidomics_vpw20_5_quantification_deconvolution1.xlsx')
#sheet=wbd[sheetname]
ws.cell(row=1, column=1).value='FA'
ws.cell(row=1, column=2).value='Sum of numerical integrals of OzID product XICs'
ws.cell(row=1, column=3).value='Numerical integral of precursor XIC'
r=0
c=0
while r<(len(swritelist[0])):
	e=swritelist[1][r] ## Precursorname		# begin determine which row to start (r) and to end (s)
	#print(e)
	s=r+1
	st=0
	while st<1:
		if s>(len(swritelist[1])-1):
			ne='stop_loop'
		else:
			ne=swritelist[1][s] #	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
	cvaln=str(swritelist[1][r])
	ws.cell(row=c+2, column=1).value=cvaln
	cvali=float(coznumintlist[c])
	ws.cell(row=c+2, column=2).value=cvali
	cvalj=float(cprecnumintlist[c])
	ws.cell(row=c+2, column=3).value=cvalj
	r=s+1
	c=c+1
wbd.save('OzFAD1_4_DIA_deconv_raw.xlsx')
# end save numerical integral of ozid products in precursor_results in new sheet Summary_OzID_Integrals
afterall=datetime.datetime.now()
dt=afterall-beforeall
print('Calculation time(h:mm:s) is:')
print(dt)
print('XICs for deconvolution are saved in OzFAD1_4_DIA_deconv_raw.xlsx')

quit()















# begin add precursor area / deconvoluted precursor area as a column to precursor_results



# end add precursor area / deconvoluted precursor area as a column to precursor_results
# begin calculate and generate new barchart in final_barchart



# begin calculate and generate new barchart in final_barchart

###########################################################################################################################################################################
## end read chromatograms, deconvolute precursor XICs, calculate updated relative quantification and barchart
###########################################################################################################################################################################\
###########################################################################################################################################################################



############################################################################################################################################
#begin sort data and save in worksheet for plotting in origin
wb=openpyxl.load_workbook('jpm_lipidomics_vpw20_5_final_output.xlsx')
sheet=wb['sortedresults']
toprow=['PrecursorName', 'ProductName', 'RTPrecursor', 'RTAldehyde1', 'RTCriegee1', 'RTAldehyde2', 'RTCriegee2', 
'RTAldehyde3', 'RTCriegee3', 'RTAldehyde4', 'RTCriegee4', 'RTAldehyde5', 'RTCriegee5', 
'RTAldehyde6', 'RTCriegee6', 'AreaPrecursor', 'AreaAldehyde1', 'AreaCriegee1', 'AreaAldehyde2', 'AreaCriegee2', 
'AreaAldehyde3', 'AreaCriegee3', 'AreaAldehyde4', 'AreaCriegee4', 'AreaAldehyde5', 'AreaCriegee5', 
'AreaAldehyde6', 'AreaCriegee6']
c=1
while c<(len(toprow)+1):
	sheet.cell(row=1, column=c).value=toprow[c-1]
	c=c+1
r=2
rclist=[]
rc=0
while r<(len(mlistname)+2):
	if prodname[r-2][len(prodname[r-2])-1]=='e':
		r=r+1
		rc=rc+1
		rclist.append(rc)
	else:
		rclist.append(rc)
		sheet.cell(row=r-rclist[r-2], column=1).value=precname[r-2]
		sheet.cell(row=r-rclist[r-2], column=2).value=prodname[r-2]
		r=r+1
d=0
while d<20:
	precname.append(d)
	d=d+1
r=2
while r<(len(mlistname)+2):	# 
	if precname[r-1]==precname[r]:
		if precname[r+1]==precname[r+2]:
			if precname[r+3]==precname[r+4]:
				if precname[r+5]==precname[r+6]:
					if precname[r+7]==precname[r+8]:
						if precname[r+9]==precname[r+10]:
							sheet.cell(row=r-rclist[r-2], column=4).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=17).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=6).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=19).value=area[r-2]
							r=r+1	
							sheet.cell(row=r-rclist[r-2], column=8).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=21).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=10).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=23).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=12).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=25).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=14).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=27).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=5).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=18).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=7).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=20).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=9).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=22).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=11).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=24).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=13).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=26).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=15).value=rettime[r-2]	#criegee
							sheet.cell(row=r, column=28).value=area[r-2]
							r=r+2 #1
							sheet.cell(row=r-rclist[r-2], column=3).value=rettime[r-2]	#precursor
							sheet.cell(row=r-rclist[r-2], column=16).value=area[r-2]	
							r=r+1
						else:
							sheet.cell(row=r-rclist[r-2], column=4).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=17).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=6).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=19).value=area[r-2]
							r=r+1	
							sheet.cell(row=r-rclist[r-2], column=8).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=21).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=10).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=23).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=12).value=rettime[r-2]	#aldehyde
							sheet.cell(row=r-rclist[r-2], column=25).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=5).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=18).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=7).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=20).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=9).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=22).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=11).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=24).value=area[r-2]
							r=r+1
							sheet.cell(row=r-rclist[r-2], column=13).value=rettime[r-2]	#criegee
							sheet.cell(row=r-rclist[r-2], column=26).value=area[r-2]
							r=r+2 #1
							sheet.cell(row=r-rclist[r-2], column=3).value=rettime[r-2]	#precursor
							sheet.cell(row=r-rclist[r-2], column=16).value=area[r-2]	
							r=r+1
					else:
						sheet.cell(row=r-rclist[r-2], column=4).value=rettime[r-2]	#aldehyde
						sheet.cell(row=r-rclist[r-2], column=17).value=area[r-2]
						r=r+1
						sheet.cell(row=r-rclist[r-2], column=6).value=rettime[r-2]	#aldehyde
						sheet.cell(row=r-rclist[r-2], column=19).value=area[r-2]
						r=r+1	
						sheet.cell(row=r-rclist[r-2], column=8).value=rettime[r-2]	#aldehyde
						sheet.cell(row=r-rclist[r-2], column=21).value=area[r-2]
						r=r+1
						sheet.cell(row=r-rclist[r-2], column=10).value=rettime[r-2]	#aldehyde
						sheet.cell(row=r-rclist[r-2], column=23).value=area[r-2]
						r=r+1
						sheet.cell(row=r-rclist[r-2], column=5).value=rettime[r-2]	#criegee
						sheet.cell(row=r-rclist[r-2], column=18).value=area[r-2]
						r=r+1
						sheet.cell(row=r-rclist[r-2], column=7).value=rettime[r-2]	#criegee
						sheet.cell(row=r-rclist[r-2], column=20).value=area[r-2]
						r=r+1
						sheet.cell(row=r-rclist[r-2], column=9).value=rettime[r-2]	#criegee
						sheet.cell(row=r-rclist[r-2], column=22).value=area[r-2]
						r=r+1
						sheet.cell(row=r-rclist[r-2], column=11).value=rettime[r-2]	#criegee
						sheet.cell(row=r-rclist[r-2], column=24).value=area[r-2]
						r=r+2 #1
						sheet.cell(row=r-rclist[r-2], column=3).value=rettime[r-2]	#precursor
						sheet.cell(row=r-rclist[r-2], column=16).value=area[r-2]	
						r=r+1	
				else:
					sheet.cell(row=r-rclist[r-2], column=4).value=rettime[r-2]	#aldehyde
					sheet.cell(row=r-rclist[r-2], column=17).value=area[r-2]
					r=r+1
					sheet.cell(row=r-rclist[r-2], column=6).value=rettime[r-2]	#aldehyde
					sheet.cell(row=r-rclist[r-2], column=19).value=area[r-2]
					r=r+1	
					sheet.cell(row=r-rclist[r-2], column=8).value=rettime[r-2]	#aldehyde
					sheet.cell(row=r-rclist[r-2], column=21).value=area[r-2]
					r=r+1
					sheet.cell(row=r-rclist[r-2], column=5).value=rettime[r-2]	#criegee
					sheet.cell(row=r-rclist[r-2], column=18).value=area[r-2]
					r=r+1
					sheet.cell(row=r-rclist[r-2], column=7).value=rettime[r-2]	#criegee
					sheet.cell(row=r-rclist[r-2], column=20).value=area[r-2]
					r=r+1
					sheet.cell(row=r-rclist[r-2], column=9).value=rettime[r-2]	#criegee
					sheet.cell(row=r-rclist[r-2], column=22).value=area[r-2]
					r=r+2 #1
					sheet.cell(row=r-rclist[r-2], column=3).value=rettime[r-2]	#precursor
					sheet.cell(row=r-rclist[r-2], column=16).value=area[r-2]	
					r=r+1
			else:
				sheet.cell(row=r-rclist[r-2], column=4).value=rettime[r-2]	#aldehyde
				sheet.cell(row=r-rclist[r-2], column=17).value=area[r-2]
				r=r+1
				sheet.cell(row=r-rclist[r-2], column=6).value=rettime[r-2]	#aldehyde
				sheet.cell(row=r-rclist[r-2], column=19).value=area[r-2]
				r=r+1	
				sheet.cell(row=r-rclist[r-2], column=5).value=rettime[r-2]	#crigee
				sheet.cell(row=r-rclist[r-2], column=18).value=area[r-2]
				r=r+1
				sheet.cell(row=r-rclist[r-2], column=7).value=rettime[r-2]	#criegee
				sheet.cell(row=r-rclist[r-2], column=20).value=area[r-2]
				r=r+2 #1
				sheet.cell(row=r-rclist[r-2], column=3).value=rettime[r-2]	#precursor
				sheet.cell(row=r-rclist[r-2], column=16).value=area[r-2]	
				r=r+1
		else:
			sheet.cell(row=r-rclist[r-2], column=4).value=rettime[r-2]	#aldehyde
			sheet.cell(row=r-rclist[r-2], column=17).value=area[r-2]
			r=r+1
			sheet.cell(row=r-rclist[r-2], column=5).value=rettime[r-2]	#criegee
			sheet.cell(row=r-rclist[r-2], column=18).value=area[r-2]
			r=r+2 #1		
			sheet.cell(row=r-rclist[r-2], column=3).value=rettime[r-2]	#precursor
			sheet.cell(row=r-rclist[r-2], column=16).value=area[r-2]	
			r=r+1
	else:
		#r=r+1
		sheet.cell(row=r-rclist[r-2], column=3).value=rettime[r-2]	#precursor
		sheet.cell(row=r-rclist[r-2], column=16).value=area[r-2]		
		r=r+1 #1	
if manualfilter==1:
	wb.save('jpm_lipidomics_vpw20_5_final_output.xlsx')
	#print('All calculations are done. Excel file is saved as jpm_lipidomics_vpw11_5_final_output.xlsx')
#afterall=datetime.datetime.now()
#dt=afterall-beforeall
#print('Calculation time:')
#print(dt)
#end sort data and save in worksheet for plotting in origin
terminate=0
if terminate==1:
	quit()

