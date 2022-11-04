# -*- coding: UTF-8 -*-

#Jan Philipp Menzel, venn diagram inspired colour-segmented barchart
#created: 2021
#Notes:
import math
import datetime
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.styles.colors import RGB
from openpyxl.chart import (PieChart, ProjectedPieChart, Reference)
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties		# works?
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList

#processing step to derive input data for plot as in rowsdata and assigned
print('This code will create an excel file showing a segmented bar chart and a pie chart for a comparison similar to a Venn Diagram.')
print('Please ensure that the data is entered correctly into the file vennbar_chart_maker_input.xlsx')
print('The relevant area must not contain empty fields (1 and 0 is allowed) and all neighboring fields must be empty.')
wb=openpyxl.load_workbook('vennbar_chart_maker_input.xlsx')
ws=wb.active
r=0
cl=4
go=1
while go==1:    # determine how many columns are in matrix
	tfe=ws.cell(row=r+5, column=cl)
	tfe=tfe.value
	if tfe is None:
		go=0
	else:
		cl=cl+1  
clmax=cl
go=1
matrixlist=[]
while go==1:
	tfe=ws.cell(row=r+5, column=4)
	tfe=tfe.value
	if tfe is None:
		go=0
	else:
		cl=4
		rowlist=[]
		while cl<clmax:
			tfe=ws.cell(row=r+5, column=cl)
			tfe=tfe.value
			rowlist.append(tfe)
			cl=cl+1
		matrixlist.append(rowlist)
	r=r+1
#print(matrixlist)
litunique=0
ltu=0
while ltu<(len(matrixlist)):
	if int(matrixlist[ltu][len(matrixlist[ltu])-1])==0:
		litunique=litunique+1
	ltu=ltu+1
#get methods
c=4
r=4
datalist=[['Method']]
sortedlist=[['Method_sorted']]
testlist=[]
while c<(len(matrixlist[0])+4):
	tfe=ws.cell(row=r, column=c)
	tfe=tfe.value
	tfe=str(tfe)
	mlist=[]
	mlist.append(tfe)
	datalist.append(mlist)
	alist=[]
	alist.append(tfe)
	sortedlist.append(alist)
	c=c+1
#print(datalist)
#print(assignedlist)
#print('...')
r=0
while r<(len(matrixlist)):
	t=0
	found=0
	while t<r:
		if matrixlist[r]==matrixlist[t]:
			found=1
			f=t
			t=r
		t=t+1
	if found==0:
		#make new category
		newcat=str()
		i=0
		while i<(len(matrixlist[r])):
			e=int(matrixlist[r][i])
			if e==1:
				newcat=newcat+str(i+1)
			i=i+1
		datalist[0].append(newcat)
		mi=1
		while mi<(len(matrixlist[0])+1):
			if matrixlist[r][mi-1]==1:
				datalist[mi].append(1)
			else:
				datalist[mi].append(0)
			mi=mi+1
	else:
		#add to existing category
		ccat=str()
		i=0
		while i<(len(matrixlist[r])):
			e=int(matrixlist[r][i])
			if e==1:
				ccat=ccat+str(i+1)
			i=i+1
		idl=1
		while idl<(len(datalist[0])):
			if ccat==datalist[0][idl]:
				fdl=idl
				idl=(len(datalist[0]))
			idl=idl+1
		mi=1
		while mi<(len(matrixlist[0])+1):
			if matrixlist[r][mi-1]==1:
				datalist[mi][fdl]=datalist[mi][fdl]+1
			else:
				mi=mi
			mi=mi+1
	r=r+1
#print(datalist)
#print(sortedlist)
#print('/////')
#quit()
# begin sorting datalist
mi=1
while mi<(len(datalist[0])):
	mt=1
	go=1
	while go==1:
		if mt<(len(sortedlist[0])):
			#compare and rank
			if (len(sortedlist[0][mt]))<(len(datalist[0][mi])):
				#insert
				k=0
				while k<(len(datalist)):
					sortedlist[k].insert(mt,datalist[k][mi])
					k=k+1	
				go=0
			else:
				mt=mt+1
				go=1
				#print(assignedlist)
				#quit()
		else:
			k=0
			while k<(len(datalist)):
				sortedlist[k].append(datalist[k][mi])
				k=k+1
			go=0
	mi=mi+1
#print(sortedlist)
#print(len(sortedlist))
# end sort datalist
#workflowunique=1
fnd=1
litandworkflow=0
while fnd<(len(sortedlist[0])):
	if str(sortedlist[0][fnd])==str(len(sortedlist)-1):
		workflowunique=int(sortedlist[len(sortedlist)-1][fnd])
	else:
		litandworkflow=litandworkflow+sortedlist[len(sortedlist)-1][fnd]
	fnd=fnd+1
#print(litunique)
#print(workflowunique)
#print(litandworkflow)			## data for pie chart
# begin make plots and write data
wbout = Workbook()#write_only=True)
wsout = wbout.active

cl=0
r=0
while cl<(len(sortedlist[0])):
	r=0
	while r<(len(sortedlist)):
		wsout.cell(row=r+1, column=cl+1).value=sortedlist[r][cl]
		r=r+1
	cl=cl+1
wbout.save('vennbar_chart_maker_output.xlsx')

#colorschemebarchart=['cornflowerBlue', 'darkCyan', 'darkSlateGrey', 'darkSlateBlue', 'navajoWhite', 'lightSteelBlue', 'white', 'lawnGreen', 'mediumVioletRed', 'wheat', 'snow', 'yellow', 'dkOrange', 'ivory', 'beige', 'mediumOrchid', 'seaShell', 'medSlateBlue', 'coral', 'darkOliveGreen', 'navy', 'dkOliveGreen', 'gray', 'ltYellow', 'chocolate', 'lightCoral', 'orangeRed', 'sienna', 'ltBlue', 'medSpringGreen', 'turquoise', 'khaki', 'yellowGreen', 'whiteSmoke', 'orange', 'sandyBrown', 'dkTurquoise', 'dkViolet', 'lightYellow']

fdrgblist=[[255, 0, 0], [255, 255, 0], [0, 255, 0], [0, 255, 255], [0, 0, 255], [255, 0, 255]]
# begin generate rgbcodelist
rgbcodelist=[]
sli=1
while sli<(len(sortedlist[0])):
	cmix=str(sortedlist[0][sli])
	if (len(cmix))==1:
		rgbcode=fdrgblist[int(cmix)-1]
	else:
		suma=0
		sumb=0
		sumc=0
		ii=0
		while ii<(len(cmix)):
			suma=suma+fdrgblist[int(cmix[ii])-1][0]
			sumb=sumb+fdrgblist[int(cmix[ii])-1][1]
			sumc=sumc+fdrgblist[int(cmix[ii])-1][2]
			ii=ii+1
		rgba=(int(suma/(len(cmix))))-(4*((len(cmix))-1))
		rgbb=(int(sumb/(len(cmix))))-(4*((len(cmix))-1))
		rgbc=(int(sumc/(len(cmix))))-(4*((len(cmix))-1))
		if rgba<0:
			#print('----------a')
			rgba=0
		if rgbb<0:
			#print('----------b')
			rgbb=0
		if rgbc<0:
			#print('----------c')
			rgbc=0
		rgbcode=[]
		rgbcode.append(rgba)
		rgbcode.append(rgbb)
		rgbcode.append(rgbc)
	rgbcodelist.append(rgbcode)
	sli=sli+1
# end generate rgbcodelist
# begin tranform rgbcodelist to hexcodelist
hexcodelist=[]
rcl=0
while rcl<(len(rgbcodelist)):
	rgbcode=rgbcodelist[rcl] #[193, 217, 78]
	hexcode=str()
	convertin=[10, 11, 12, 13, 14, 15]
	convertout=['A', 'B', 'C', 'D', 'E', 'F']
	gen=0
	while gen<3:
		bc=((int(rgbcode[gen]))/16)-(((int(rgbcode[gen]))%16)/16)
		ac=(((int(rgbcode[gen]))%16)/16)*16
		#print(bc)
		#print(ac)
		if int(bc)>9:
			ci=0
			while ci<(len(convertin)):
				if convertin[ci]==int(bc):
					hexcode=hexcode+str(convertout[ci])
				ci=ci+1
		else:
			hexcode=hexcode+str(int(bc))
		if int(ac)>9:
			ci=0
			while ci<(len(convertin)):
				if convertin[ci]==int(ac):
					hexcode=hexcode+str(convertout[ci])
				ci=ci+1
		else:
			hexcode=hexcode+str(int(ac))
		gen=gen+1
	#print(hexcode)
	hexcodelist.append(hexcode)
	rcl=rcl+1
# end tranform rgbcodelist to hexcodelist

chart1 = BarChart()
chart1.type = "col"
chart1.style = 12
chart1.grouping = "stacked"
chart1.overlap = 100
chart1.title = 'Comparison of numbers of fatty acids detected in Human Plasma'
chart1.y_axis.title = 'Number of found fatty acid species'
chart1.x_axis.title = 'Source / Method of detection'
#chart1.y_axis.scaling.max = 100   #############
maxr=len(sortedlist)
maxcol=len(sortedlist[0])
data = Reference(wsout, min_col=2, min_row=1, max_row=maxr, max_col=maxcol)
cats = Reference(wsout, min_col=1, min_row=2, max_row=maxr)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4

# begin turn majorGridlines off (setting colour white: FFFFFF)
chart1.y_axis.majorGridlines.spPr = GraphicalProperties(noFill = 'True')
chart1.y_axis.majorGridlines.spPr.ln = LineProperties(solidFill = 'FFFFFF')
chart1.x_axis.majorGridlines = ChartLines()
chart1.x_axis.majorGridlines.spPr = GraphicalProperties(noFill = 'True')
chart1.x_axis.majorGridlines.spPr.ln = LineProperties(solidFill = 'FFFFFF')
# end turn majorGridlines off (setting colour white: FFFFFF)

chart1.legend=None		# no legend

stbr=0
while stbr<(len(sortedlist[0])-1):
	clm=0
	while clm<(len(sortedlist)-1):      # 6 is number of columns in barchart (16:1, 17:1, ...)
		s=chart1.series[stbr]   #define datapoint in column ([0] is n-3; [1] is n-4 ...)
		pt=DataPoint(idx=clm)     #define which column (e.g. idx=0 is 16:1; idx=1 is 17:1 ...)
		clrbc=stbr #(assigned[clm+1][stbr+1])-1
		pt.graphicalProperties.solidFill=hexcodelist[stbr] #ColorChoice(prstClr=colorschemebarchart[clrbc])
		pt.graphicalProperties.line.solidFill=hexcodelist[stbr] #ColorChoice(prstClr=colorschemebarchart[clrbc])
		s.dPt.append(pt)
		clm=clm+1
	stbr=stbr+1
#chart1.legend=None
wsout.add_chart(chart1, "B9")
# begin add data for pie chart and add pie chart
cl=len(sortedlist[0])+3
wsout.cell(row=1, column=cl-1).value='Category'
wsout.cell(row=4, column=cl-1).value='Number of FA found both by the workflow and by at least one other method.'
wsout.cell(row=2, column=cl-1).value='Number of FA found uniquely by the workflow.'
wsout.cell(row=3, column=cl-1).value='Number of FA not found by the workflow, but found by at least one other method.'
wsout.cell(row=1, column=cl).value='Found fatty acids'
wsout.cell(row=4, column=cl).value=litandworkflow
wsout.cell(row=2, column=cl).value=workflowunique
wsout.cell(row=3, column=cl).value=litunique
pie = PieChart()
labels = Reference(wsout, min_col=(cl-1), min_row=2, max_row=4)
data = Reference(wsout, min_col=cl, min_row=1, max_row=4)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = 'Workflow performance compared to literature'
colorschemebarchart=['blue', 'orangeRed', 'green', 'lawnGreen', 'darkCyan', 'darkSlateGrey', 'navajoWhite', 'lightSteelBlue', 'white', 'mediumVioletRed', 'wheat', 'snow', 'yellow', 'dkOrange', 'ivory', 'beige', 'mediumOrchid', 'seaShell', 'medSlateBlue', 'coral', 'navy', 'dkOliveGreen', 'gray', 'ltYellow', 'chocolate', 'lightCoral', 'sienna', 'ltBlue', 'turquoise', 'khaki', 'yellowGreen', 'whiteSmoke', 'orange', 'sandyBrown', 'dkTurquoise', 'dkViolet', 'lightYellow']




clm=0
while clm<3:      # 6 is number of columns in barchart (16:1, 17:1, ...)
	s=pie.series[0]   #define datapoint 
	pt=DataPoint(idx=clm)     #define 
	pt.graphicalProperties.solidFill=ColorChoice(prstClr=colorschemebarchart[clm])
	pt.graphicalProperties.line.solidFill=ColorChoice(prstClr=colorschemebarchart[clm])
	s.dPt.append(pt)
	clm=clm+1

wsout.add_chart(pie, 'M9')
# end add data for pie chart and add pie chart
wbout.save("vennbar_chart_maker_output.xlsx")
print('The graph is saved in the file vennbar_chart_maker_output.xlsx.')


