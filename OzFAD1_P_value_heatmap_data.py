# -*- coding: UTF-8 -*-

#Jan Philipp Menzel 
# Program: Calculate data for p value heatmap from OzFAD rep plot input excel file
#created: 2022 05 12
#Notes: Reads excel files of 2 sets of 3 replicates (6 excel files) of input files for rep plot of OzFAD workflow
#Notes: Calculates p values for each comparison of species (or groups of species in each field), ready to be plotted as a heatmap in origin
import math
import datetime
import pandas as pd
import scipy
from scipy import stats
import openpyxl
from openpyxl import Workbook

print('This code will create an excel file containing data that can be displayed as the P values of comparison of')
print('  two fatty acid analysis datasets from the OzFAD1 workflow.')
print('Please ensure that the data is entered correctly into the files:')
print('Six input files with fatty acid data are required, which need to be named as follows:')
print('jpmlipidomics_vpw20_10_quantified_final_rep1_d1.xlsx')
print('jpmlipidomics_vpw20_10_quantified_final_rep2_d1.xlsx')
print('jpmlipidomics_vpw20_10_quantified_final_rep3_d1.xlsx')
print('jpmlipidomics_vpw20_10_quantified_final_rep1_d2.xlsx')
print('jpmlipidomics_vpw20_10_quantified_final_rep2_d2.xlsx')
print('jpmlipidomics_vpw20_10_quantified_final_rep3_d2.xlsx')
print('Each three replicates of the two analyses are compared to each other and P values as well as fold change values are calculated.')

#begin read all 6 excel files into lists of lists (read relative isomer abundance data by FA and db array from sheet final_barchart)
tnrep=3 # total number of replicates is 3
xlc=tnrep*2 # 6 excel files in total
superlist=[]    # list of list of lists (contains all 6 slists, which contain lists for rows of raw input data) ## superlist[dataset 0 to 5][row index][column index]

cds=1
while cds<(xlc+1):
    if cds==1:
        wb=openpyxl.load_workbook('jpmlipidomics_vpw20_10_quantified_final_rep1_d1.xlsx')
        ws=wb['final_barchart']
    elif cds==2:
        wb=openpyxl.load_workbook('jpmlipidomics_vpw20_10_quantified_final_rep2_d1.xlsx')
        ws=wb['final_barchart']
        #print('Check1')
        #quit()
    elif cds==3:
        wb=openpyxl.load_workbook('jpmlipidomics_vpw20_10_quantified_final_rep3_d1.xlsx')
        ws=wb['final_barchart']
    elif cds==4:
        wb=openpyxl.load_workbook('jpmlipidomics_vpw20_10_quantified_final_rep1_d2.xlsx')
        ws=wb['final_barchart']
    elif cds==5:
        wb=openpyxl.load_workbook('jpmlipidomics_vpw20_10_quantified_final_rep2_d2.xlsx')
        ws=wb['final_barchart']
    elif cds==6:
        wb=openpyxl.load_workbook('jpmlipidomics_vpw20_10_quantified_final_rep3_d2.xlsx')
        ws=wb['final_barchart']
    rowlist=[]  #list with row entries (contains different db information for one FA group, e.g., 15:1)
    slist=[]    #list of rowlists (whole dataset for one sample, one individual replicate)
    r=1
    c=1
    go=1
    while go==1:
        c=1
        v=ws.cell(row=r, column=c)
        v=v.value
        v=str(v)
        #print(v)
        #vn=ws.cell(row=r+1, column=c)
        #vn=vn.value
        if v is None:
            go=0
        elif v=='None':
            go=0
        else:
            rowlist=[]  #list with row entries (contains different db information for one FA group, e.g., 15:1)
            c=1
            gor=1
            while gor==1:
                vc=ws.cell(row=r, column=c)
                vc=vc.value
                if vc is None:
                    gor=0
                else:
                    rowlist.append(vc)
                c=c+1
            slist.append(rowlist)
            #print(rowlist)
        r=r+1
    superlist.append(slist)
    #print(slist)
    cds=cds+1
print('Reading of datasets complete, superlist is generated.')
#print(superlist)
#print(superlist[1][27][6])
#print(len(superlist[0]))
#print(len(superlist[1]))
#print(len(superlist[2]))
#print(len(superlist[3]))
#print(len(superlist[4]))
#print(len(superlist[5]))
#begin check datasets in superlist
ok=0
ok2=0
if len(superlist[0])==len(superlist[1]):
    if len(superlist[2])==len(superlist[1]):
        ok=1
    else:
        print('Check input datasets. Data may be inconsistent (different number of rows in excel files 2 3).')
else:
    print('Check input datasets. Data may be inconsistent (different number of rows in excel files 1 2).')
if len(superlist[3])==len(superlist[4]):
    if len(superlist[3])==len(superlist[5]):
        ok2=1
    else:
        print('Check input datasets. Data may be inconsistent (different number of rows in excel files 4 6).')
else:
    print('Check input datasets. Data may be inconsistent (different number of rows in excel files 4 5).')
if ok==0:
    quit()
elif ok2==0:
    quit()
#end check datasets in superlist
#begin insert empty rows in superlist to complete datasets and align rows of FA
print('Datasets must align, otherwise p values will be calculated wrongly!')
#create code to automate insertion of rows...
#end insert empty rows in superlist to complete datasets and align rows of FA
#end read all 6 excel files into lists of lists (read relative isomer abundance data by FA and db array from sheet final_barchart)

#begin calculate p values 
tlist=[]
ctlist=[]
plist=[]
cplist=[]
r=1
c=1
while r<len(superlist[0]):
    c=1
    ctlist=[]
    cplist=[]
    while c<len(superlist[0][0]):
        cd1list=[]
        cd2list=[]
        cds=0
        while cds<3:
            cd1=float(superlist[cds][r][c])
            cd1list.append(cd1)
            cds=cds+1
        while cds<6:
            cd2=float(superlist[cds][r][c])
            cd2list.append(cd2)
            cds=cds+1

        t,p=stats.ttest_ind(cd1list, cd2list, equal_var=False)
        t=str(t)
        p=str(p)
        if t=='nan':
            t=str(1.0)
        if p=='nan':
            p=str(1.0)
        ctlist.append(t)
        cplist.append(p)
        c=c+1
    tlist.append(ctlist)
    plist.append(cplist)
    r=r+1
#print(plist)
#print(plist[10][15])


    #if str(superlist[0][r][0])==str(superlist[1][r][0]):
    #    if str(superlist[0][r][0])==str(superlist[2][r][0]):
    #        if str(superlist[0][r][0])==str(superlist[3][r][0]):

# begin calculate mean fold changes between datasets
foldlist=[]
cfoldlist=[]
r=1
c=1
while r<len(superlist[0]):
    c=1
    cfoldlist=[]
    while c<len(superlist[0][0]):
        cablist=[]
        cds=0
        while cds<6:
            cab=float(superlist[cds][r][c])
            cablist.append(cab)
            cds=cds+1
        if 0 in cablist:
            cfold=0
        else:
            cfold=((cablist[3]/cablist[0])+(cablist[4]/cablist[1])+(cablist[5]/cablist[2]))/3   # calculate mean fold change between samples
        cfoldlist.append(cfold)
        c=c+1
    foldlist.append(cfoldlist)
    r=r+1
#print(foldlist)
#print(foldlist[10][15]) # [10][15] in MCF7 LNCaP LNCaP_SCD-1i is 14:1n-5cis
#print(foldlist[11][25]) # [11][25] in MCF7 LNCaP LNCaP_SCD-1i is 16:1n-7cis


# end calculate mean fold changes between datasets

#end calculate p values
writeoutput=1
if writeoutput==1:
    #begin write p values in output excel file 
    wb = Workbook(write_only=True)
    wb.save('jpmlipidomics_p_values_for_heatmap.xlsx')
    wb=openpyxl.load_workbook('jpmlipidomics_p_values_for_heatmap.xlsx')
    ws=wb.active

    c=0
    while c<len(superlist[0][0]):   #write top row with db assignment
        if 'NMI, E' in str(superlist[0][0][c]):     # correct entries for final plot
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx+1:
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'branched)'
                dx=dx+1
        elif 'NMI, Z' in str(superlist[0][0][c]):
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx+1:
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'NMI)'
                dx=dx+1
        elif 'Bu' in str(superlist[0][0][c]):
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx+1:
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'NMI (Bu))'
                dx=dx+1
        elif 'Me, E' in str(superlist[0][0][c]):
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx+1:
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'trans (E))'
                dx=dx+1
        elif 'Me, Z' in str(superlist[0][0][c]):
            dblabel=str(superlist[0][0][c])
            dx=0
            while dx<len(dblabel):
                if dblabel[dx]=='(':
                    dxn=0
                    ndblabel=str()
                    while dxn<dx+1:
                        ndblabel=ndblabel+dblabel[dxn]
                        dxn=dxn+1
                    ndblabel=ndblabel+'cis (Z))'
                dx=dx+1
        else:
            ndblabel=str(superlist[0][0][c])
        ws.cell(row=1, column=c+1).value=ndblabel  # write row label for p value heatmap   
        ws.cell(row=1+4+len(superlist[0]), column=c+1).value=ndblabel  # write row label for fold change heatmap    
        c=c+1
    r=0
    while r<len(superlist[0]):      
        ws.cell(row=r+1, column=1).value=superlist[0][r][0]     #write column assignment for p value heatmap (FA)
        ws.cell(row=r+5+len(superlist[0]), column=1).value=superlist[0][r][0]     #write column assignment for fold-change heatmap (FA)
        r=r+1
    r=0
    while r<len(plist):
        c=0
        while c<len(plist[r]):
            ws.cell(row=r+2, column=c+2).value=float(plist[r][c])   #write p values
            ws.cell(row=r+6+len(superlist[0]), column=c+2).value=float(foldlist[r][c])   #write fold change values
            c=c+1
        r=r+1
    ws.cell(row=2+len(superlist[0]), column=2).value='Above: P values for heatmap'   #write Note
    ws.cell(row=3+len(superlist[0]), column=2).value='Below: Fold change values for heatmap'   #write Note

    wb.save('jpmlipidomics_p_values_for_heatmap.xlsx')
    print('P values are saved in excel file jpmlipidomics_p_values_for_heatmap.xlsx')
    #end write p values in output excel file 
print('Done.')

