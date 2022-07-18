# -*- coding: UTF-8 -*-

# Jan Philipp Menzel jpm_lipidomics_vpw13_1_precursor_tr.py
#created: 09 07 2020
#modified: regularly until 07 04 2021 
# Goal: STEP 1. Generate transition list for Skyline containing precursors (intact derivatized fatty acids), either one of seven pre-defined derivatives or any new structure
## Notes: Derivative, positive fixed charge
## NOTES: VIRTUAL PRECURSOR - PrecursorName and PrecursorMz are artificially set +Xe (only column 3 and 5), fragment transitions correct including precursor
## NOTES: Virtual precursor forces Skyline to consider all transitions incl. real precursor (fragment in transition list), Skyline Setting: TransitionSettings-Filter-IonTypes-f 
import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
import statistics
from statistics import mean
from statistics import median
################ DATABASE ## Source: Internetchemie.info 
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.757, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
################
#default=eval(input('Run workflow with default parameters? (Yes: 1 / No: 0) (Derivatization agent: AMPP; Slow and full workflow including all FA; Apply retention time limitation; Use Fatty Acid Library to prevent filtering out important FA; Max RT = 17.5 min; FA: 12 - 24 C; Max m/z error: 10 ppm; Precursor peak area threshold: 3000; Product peak area threshold: 200.) :'))
default=0
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
###########

########### begin read workflow parameters
transferlist=[]
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
go=1
while go==1:
	tfe=ws.cell(row=tli+1, column=2)
	tfe=tfe.value
	if tli==21:
		go=0
	if tfe is None:
		go=0
	else:
		transferlist.append(tfe)
	tli=tli+1
########## 
fourlettcode=transferlist[0]		#input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
cderiv=int(transferlist[1])
hderiv=int(transferlist[2])
dderiv=int(transferlist[3])
nderiv=int(transferlist[4])
oderiv=int(transferlist[5])
pderiv=int(transferlist[6])
ideriv=int(transferlist[7])
minlenfa=int(transferlist[8])
maxlenfa=int(transferlist[9])
mzcutoff=int(transferlist[10]) 		#eval(input('What is max mz [ppm] for positive identification of species? (e.g. 30) :'))
rettimecutoff=float(transferlist[11]) 		#eval(input('What is the maximum retention time [min]? (e.g. 11.9 or 17.8) :'))
areathreshold=int(transferlist[12])		#eval(input('What is the threshold for detected precursor peak area? (compare Skyline report file, e.g. 3000) :'))
workflowmode=int(transferlist[13])
productareathreshold=int(transferlist[14])
nchunks=int(transferlist[15])
rtlimitation=int(transferlist[16])
mostwanted=int(transferlist[17])
transtest=int(transferlist[18])
runprecheck=int(transferlist[19])
workflowidentifier=str(transferlist[20])
########### end read workflow parameters
# begin adjust workflow parameters for DDA analysis

# end adjust workflow parameters for DDA analysis


beforeall=datetime.datetime.now()
print('Workflow is running ...')

# begin build mostwantedlist 
mostwantedlist=[]
wbfa=openpyxl.load_workbook('jpm_fa_lib.xlsx')
wsfa=wbfa.active
mwi=0
go=1
while go==1:
	tfe=wsfa.cell(row=mwi+3, column=1)
	tfe=tfe.value
	if tfe is None:
		go=0
	else:
		if int(tfe)<10:
			tfe='0'+str(tfe)
		mwfa=str(tfe)+':'
		tfe=wsfa.cell(row=mwi+3, column=2)
		tfe=tfe.value
		dbfa=int(tfe)
		if int(tfe)==0:
			mwfa=mwfa+str(tfe)
		else:
			mwfa=mwfa+str(tfe)+'_n-'
			dbfai=1
			while dbfai<(dbfa+1):
				tfe=wsfa.cell(row=mwi+3, column=dbfai+2)
				tfe=tfe.value
				if dbfai<(dbfa):
					mwfa=mwfa+str(tfe)+'_n-'
				else:
					mwfa=mwfa+str(tfe)
				dbfai=dbfai+1
		mostwantedlist.append(mwfa)
	mwi=mwi+1
# end build mostwantedlist
#print('Arrived here.#################################################################################################################################')
#quit()

#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists
trdf=pd.read_csv('skyl_report_dda_vpw20_3_rt_shift1.csv')
toprowx=[trdf.columns.values.tolist()]
toprow=toprowx[0]
trdf=trdf.transpose()
writelist=trdf.values.tolist()
ki=len(writelist[0])
print('Number of rows in skyl_report_dda_vpw20_3_rt_shift1.csv: %d' % ki)
#print(ki)
#####################################################################################################

# begin get rt of 16:1_n-7 cis, chromatograms of precursor and OzID transitions and derive RT for 16:1_n-7 cis in DIA dataset, calc RT shift and apply to exrt for transition list with DDA confirmed species
anc=str()
anc=anc+str(fourlettcode)+'_16:1_n-7_precursor'
if str(writelist[6][2])==anc:
	ddartanchor=float(writelist[17][2])		#is retention time from DDA analysis
else:
	print('Check workflow...')

convertfile=0
if convertfile==0:
	xictimesdf=pd.read_csv('skyl_xic_dda_report_vpw20_3_times.csv', header=None, skiprows=1, nrows=1)
	xictimeslistfromdf=xictimesdf.values.tolist()
	xictimeslist=xictimeslistfromdf[0]		# contains times of XICs # actual list starts at index 8

segtrdf=pd.read_csv('skyl_xic_dda_report_vpw20_3_intensities.csv', skiprows=1, header=None, nrows=4, low_memory=False)
allxiclist=segtrdf.values.tolist()

lowrt=ddartanchor-0.3
lrt=8
trt=xictimeslist[lrt]
highrt=ddartanchor+0.3
#print(lowrt)
#print(highrt)
#print('lowrt and highrt')
#print(trt)
pmtimes=[]
apmxic=[]
cpmxic=[]
ppmxic=[]
while trt<lowrt:
	trt=float(xictimeslist[lrt])
	lrt=lrt+1
while trt<highrt:
	trt=float(xictimeslist[lrt])
	art=float(allxiclist[0][lrt])
	crt=float(allxiclist[1][lrt])
	prt=float(allxiclist[2][lrt])
	#print(trt)
	#print(art)
	#print(crt)
	#print(prt)
	#print('#')
	if prt>200:
		if prt>art:
			if prt>crt:
				pmtimes.append(trt)
				apmxic.append(art)
				cpmxic.append(crt)
				ppmxic.append(prt)
	lrt=lrt+1
#print(apmxic)
ma=max(apmxic)
mc=max(cpmxic)
mt=0
while mt<(len(pmtimes)):
	if apmxic[mt]==ma:
		mart=mt
	mt=mt+1
mt=0
while mt<(len(pmtimes)):
	if cpmxic[mt]==mc:
		mcrt=mt
	mt=mt+1
diartanchor=(pmtimes[mart]+pmtimes[mcrt])/2
ddadiashift=diartanchor-ddartanchor		# shift for palmitoleic acid, for larger RT, needs to be shifted accordingly; add shift to values in list
relshift=ddadiashift/ddartanchor		# relative shift

#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists
trdf=pd.read_csv('skyl_report_dda_vpw20_2_filtered.csv')
toprowx=[trdf.columns.values.tolist()]
toprow=toprowx[0]
trdf=trdf.transpose()
swritelist=trdf.values.tolist()
ki=len(swritelist[0])
#print('Number of rows in skyl_report_dda_vpw20_3_rt_shift1.csv: %d' % ki)

mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
exrtwindow=[]
explicitrt=[]

t=0
while t<(len(swritelist[0])): 
	e=swritelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving 
	mlistname.append(e)
	e=swritelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
	precname.append(e)
	e=swritelist[2][t] #sheetinput.cell(row=t, column=3)	# precformula	
	precformula.append(e)
	e=swritelist[3][t] #sheetinput.cell(row=t, column=4)	# precadduct	
	precadduct.append(e)
	e=swritelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
	precmz.append(e)
	e=swritelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
	precchrg.append(e)
	e=swritelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
	prodname.append(e)
	e=swritelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
	prodformula.append(e)
	e=swritelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
	prodadduct.append(e)
	e=swritelist[9][t] #sheetinput.cell(row=t, column=10)	# 
	prodmz.append(e)
	e=swritelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
	prodchrg.append(e)
	e=swritelist[17][t] ##
	e=e+(e*relshift) 	
	explicitrt.append(e)
	exrtwindow.append(0.01)		################################# ENTER EXPLICIT RETENTION TIME WINDOW ##############################
	t=t+1


writelist=[]
writelist.append(mlistname)
writelist.append(precname)
writelist.append(precformula)
writelist.append(precadduct)
writelist.append(precmz)
writelist.append(precchrg)
writelist.append(prodname)
writelist.append(prodformula)
writelist.append(prodadduct)
writelist.append(prodmz)
writelist.append(prodchrg)
writelist.append(explicitrt)
writelist.append(exrtwindow)



toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 
		'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']
#print('swritelist created')
transitionresultsdf=pd.DataFrame(writelist).transpose()
#print('Transposed')
transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
#print('Transposed and DataFrame created')
after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_5_'+fourlettcode+'_'
filename='jpmlipidomics_dda_vpw20_4_rt_shifted.csv'
transitionresultsdf.to_csv(filename, index=False)
print('Transition list is saved as jpmlipidomics_dda_vpw20_4_rt_shifted.csv')
afterall=datetime.datetime.now()
dt=afterall-beforeall
print('Calculation time (h:mm:ss) is:')
print(dt)








