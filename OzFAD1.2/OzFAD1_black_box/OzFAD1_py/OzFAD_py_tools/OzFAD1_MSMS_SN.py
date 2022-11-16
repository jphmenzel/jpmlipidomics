# -*- coding: UTF-8 -*-

#Jan Philipp Menzel
#created: 2022 01 24, modified regularly until 21 Feb 2022
#Notes: Reads Excel file with selected MSMS spectra and combines the ones from same species with similar RT, saves combined MSMS as excel file.
#Notes: Then determines Signal to Noise for each species
import math
import datetime
import pandas as pd
import openpyxl
from openpyxl import Workbook
#import psutil
#import subprocess
#import brainpy
#from brainpy import isotopic_variants

isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]

devmode=0
# For normal full analysis, comment out next line: #
#devmode=eval(input('Development mode: 1-Y 0-N ::'))
trouble=0

if devmode==0:
    print('This script reads MS/MS spectra from excel input file OzFAD1_2_MSMS_input.xlsx (sheet MSMS_spectra) and calculates Signal-to-Noise values.')
    print('Insert mass spectral data from Skyline (Select MSMS spectra, scale to full view - m/z = 100 to 1200 and export data) into fields A1, D1, G1, ...')
    #fourlettcode=eval(input('What is the four letter code of the derivatization agent? (e.g. AMPP, IAMP...) ::'))
    #fourlettcode=str(fourlettcode)
    # begin read excel input file into lists of lists
    rawxlx=[]   # will contain m/z values etc from excel file
    rawxly=[]   # will contain Intensity values etc from excel file

    #wbraw=openpyxl.load_workbook('jpmlipidomics_msms_sn_calc_input.xlsx')
    wbraw=openpyxl.load_workbook('OzFAD1_2_MSMS_input.xlsx')   # input file with selected MSMS for testing
    wsraw=wbraw['MSMS_spectra']

    xlcol=1   #xlcol is number of column in input excel file
    xlrow=1   #xlrow is number of row in input excel file
    gocol=1
    while gocol==1:
        crawxlx=[]
        crawxly=[]
        tfe=wsraw.cell(row=4, column=xlcol+2)
        tfe=tfe.value
        tff=wsraw.cell(row=4, column=xlcol+3)
        tff=tff.value
        #print(tfe)
        #print(tff)
        if tfe is None:
            if tff is None:
                gocol=0
        if gocol<2: ## ensure to also read last spectrum
            xlrow=1 
            gorow=1
            while gorow==1:
                tfe=wsraw.cell(row=xlrow, column=xlcol)
                tfe=tfe.value
                if tfe is None:
                    tfg=wsraw.cell(row=xlrow+3, column=xlcol)
                    tfg=tfg.value
                    if tfg is None:
                        gorow=0
                    else:
                        crawxlx.append('_')
                        crawxly.append('_')
                        crawxlx.append('_')
                        crawxly.append('_')
                        crawxlx.append('_')
                        crawxly.append('_')
                        xlrow=xlrow+3
                else:
                    tfy=wsraw.cell(row=xlrow, column=xlcol+1)
                    tfy=tfy.value
                    if tfy is None:
                        tfy='_'
                    crawxlx.append(tfe)
                    crawxly.append(tfy)
                    xlrow=xlrow+1
            rawxlx.append(crawxlx)
            rawxly.append(crawxly)
            if crawxlx[len(crawxlx)-1]<1175:
                print('################################################   CAUTION   #################################################')
                print('####################   Make sure that MSMS from Skyline is in full MS range to 1200   ########################')
                print('####################  Check species:')
                print(crawxlx[0])
                print(crawxlx[1])
                trouble=1
            xlcol=xlcol+3
            #print(crawxlx)
            #print(crawxly)
    print('Excel file is read.')
    if trouble==1:
        print('Code interrupted. Correct errors in MSMS excel file, then run this script again.')
        quit()
    fourlettcode=str(rawxlx[0][1][0])+str(rawxlx[0][1][1])+str(rawxlx[0][1][2])+str(rawxlx[0][1][3])
    #print(fourlettcode)
    #print(len(rawxlx))
    #print(len(rawxly))
    #print(rawxlx[len(rawxlx)-1][len(rawxlx[len(rawxlx)-1])-1])
    #print(rawxly[len(rawxly)-1][len(rawxly[len(rawxly)-1])-1])
    # end read excel input file into lists of lists
    ###########################
    # begin combine spectra to one combined MSMS spectrum per species (distinguish between cases of MSMS from Skyline only and MSMS from MassLynx)
    # , then save in excel file
    mzraw=[]    #m/z values of processed (combined) MSMS spectra in Skyline matched format
    allmzraw=[]
    intraw=[]   #Intensity values of processed (combined) MSMS spectra in Skyline matched format
    allintraw=[]
    rttolerance=0.09     # [min] tolerance for two segments to be apart in RT to be considered from same species, not cis\trans isomers
    scount=0
    specn=0
    xlx=0
    while specn<(len(rawxlx)-1):
        #print(specn)
        currentFA=str()
        ccf=0
        while ccf<(len(rawxlx[specn][1])-23):
            currentFA=currentFA+rawxlx[specn][1][ccf]
            ccf=ccf+1
        #print(rawxlx[specn][1])
        print(currentFA)
        if str(rawxlx[specn+1][0])=='_':
            # Found MSMS from MassLynx
            cmzraw=[]
            cintraw=[]
            # Begin transform MassLynx combined MSMS into Skyline reported format with matched m/z values
            foundlist=[]    # list with as many entries, at first all 0, as specn+1
            kf=0
            z=0
            while kf<(len(rawxlx[specn+1])):
                if str(rawxlx[specn+1][kf])=='all':
                    kf=len(rawxlx[specn+1])
                else:
                    foundlist.append(z)
                kf=kf+1
            gomatch=1
            x=0
            cmzraw.append(str(rawxlx[specn][x]))
            cintraw.append('_')
            x=x+1
            while gomatch==1:
                if fourlettcode in str(rawxlx[specn][x]):
                    cmzraw.append(str(rawxlx[specn][x]))
                    cmzraw.append(str(rawxlx[specn][x+1]))
                    cintraw.append('_')
                    cintraw.append(str(rawxly[specn][x+1]))
                    x=x+2
                elif str(rawxlx[specn][x])=='unmatched':
                    cmzraw.append(str(rawxlx[specn][x]))
                    cmzraw.append(str(rawxlx[specn][x+1]))
                    cintraw.append('_')
                    cintraw.append(str(rawxly[specn][x+1]))
                    x=x+2
                elif str(rawxlx[specn][x])=='all':
                    #cmzraw.append(str(rawxlx[specn][x]))
                    #cmzraw.append(str(rawxlx[specn][x+1]))
                    gomatch=0
                    x=x+2
                elif rawxlx[specn][x] is None:
                    gomatch=0
                else:
                    if gomatch==1:
                        # begin replace value with MassLynx combined spectrum value and save in list
                        xs=3
                        while xs<len(rawxlx[specn+1]):
                            if (float(rawxlx[specn][x])-float(rawxlx[specn+1][xs]))<0.005:
                                foundlist[xs]=1
                                cmzraw.append(float(rawxlx[specn+1][xs]))
                                cintraw.append(float(rawxly[specn+1][xs]))
                                xs=len(rawxlx[specn+1])
                            xs=xs+1
                        ok=1
                        x=x+1
            

            # begin go through foundlist and insert values from specn+1 with associated 0 value in foundlist
            xf=3
            while xf<(len(foundlist)):
                if foundlist[xf]==0:
                    if str(rawxlx[specn+1][xf])=='unmatched':
                        xf=xf+2
                    elif fourlettcode in str(rawxlx[specn+1][xf]):
                        xf=xf+2
                    else:
                        xff=3
                        goinsert=1
                        while goinsert==1:
                            if fourlettcode in str(cmzraw[xff]):
                                xff=xff+2
                            elif str(cmzraw[xff])=='unmatched':
                                xff=xff+2
                            else:
                                if xff<(len(cmzraw)-1):
                                    if fourlettcode in str(cmzraw[xff+1]):
                                        xff=xff+3
                                    if str(cmzraw[xff+1])=='unmatched':
                                        xff=xff+3
                                    if str(cmzraw[xff+1])=='all':
                                        goinsert=0
                                    elif float(rawxlx[specn+1][xf])>float(cmzraw[xff]):     
                                        if float(rawxlx[specn+1][xf])<float(cmzraw[xff+1]):
                                            cmzraw.insert(xff+1, float(rawxlx[specn+1][xf]))
                                            cintraw.insert(xff+1, float(rawxly[specn+1][xf]))
                                            goinsert=0
                                    xff=xff+1
                                else:
                                    goinsert=0
                            if xff>(len(cmzraw)-1):
                                goinsert=0
                        xf=xf+1
                else:
                    xf=xf+1
            # end go through foundlist and insert values from specn+1 with associated 0 value in foundlist

            specn=specn+2
            allmzraw.append(cmzraw)
            allintraw.append(cintraw)
            scount=0
            #print(cmzraw)
            #print(cintraw)
            # calculate exclusion values (characteristic pattern) based on precursor m/z (all m/z with tolerance around all possible transitions and isotopes of all transitions)

            # append()  # save combined MS/MS in list of lists
        elif str(rawxlx[specn][1])==str(rawxlx[specn+1][1]):    # found MSMS from Skyline
            #scount=0
            crt=str(rawxlx[specn][0][len(rawxlx[specn][0])-9])+str(rawxlx[specn][0][len(rawxlx[specn][0])-8])+str(rawxlx[specn][0][len(rawxlx[specn][0])-7])+str(rawxlx[specn][0][len(rawxlx[specn][0])-6])
            nrt=str(rawxlx[specn+1][0][len(rawxlx[specn+1][0])-9])+str(rawxlx[specn+1][0][len(rawxlx[specn+1][0])-8])+str(rawxlx[specn+1][0][len(rawxlx[specn+1][0])-7])+str(rawxlx[specn+1][0][len(rawxlx[specn+1][0])-6])
            if abs(float(crt)-float(nrt))<rttolerance:  # compare RT of current and next MS/MS to determine, whether to combine
                cmzraw=[]
                cintraw=[]
                foundlist=[]    # list with as many entries, at first all 0, as specn+1
                kf=0
                z=0
                while kf<(len(rawxlx[specn+1])):
                    if str(rawxlx[specn+1][kf])=='all':
                        kf=len(rawxlx[specn+1])
                    else:
                        foundlist.append(z)
                    kf=kf+1
                # Found another MSMS from Skyline, same species, to be added
                #cmzraw=[]
                x=0
                cmzraw.append(str(rawxlx[specn][x]))
                cintraw.append('_')
                x=x+1
                gomatch=1
                while gomatch==1:
                    #print(rawxly[specn+1][x])
                    if fourlettcode in str(rawxlx[specn][x]):
                        cmzraw.append(str(rawxlx[specn][x]))
                        cmzraw.append(str(rawxlx[specn][x+1]))
                        cintraw.append('_')
                        cintraw.append(str(rawxly[specn][x+1]))
                        x=x+2
                    elif str(rawxlx[specn][x])=='unmatched':
                        cmzraw.append(str(rawxlx[specn][x]))
                        cmzraw.append(str(rawxlx[specn][x+1]))
                        cintraw.append('_')
                        cintraw.append(str(rawxly[specn][x+1]))
                        x=x+2
                    elif str(rawxlx[specn][x])=='all':
                        #cmzraw.append(str(rawxlx[specn][x]))
                        #cmzraw.append(str(rawxlx[specn][x+1]))
                        gomatch=0
                        x=x+2
                    elif rawxlx[specn][x] is None:
                        gomatch=0
                    else:
                        # begin combine value with Skyline matched or unmatched spectrum value and save in list
                        scount=1 ########## Note that an MSMS was found and another could be found next
                        xs=3
                        gofind=1
                        found=0
                        while gofind==1:
                            if str(rawxlx[specn+1][xs])=='all':
                                gofind=0
                            elif str(rawxlx[specn+1][xs])=='unmatched':
                                xs=xs+2
                            elif fourlettcode in str(rawxlx[specn+1][xs]):
                                xs=xs+2
                            elif abs(float(rawxlx[specn][x])-float(rawxlx[specn+1][xs]))<0.002:     # if m/z present in both MS/MS, add; if not use both values as is
                                found=1
                                if foundlist[xs]==1:
                                    print('Check code !!')
                                ccint=float(rawxly[specn][x])+float(rawxly[specn+1][xs])
                                ccmz=float(rawxlx[specn][x])
                                cmzraw.append(ccmz)
                                cintraw.append(ccint)
                                foundlist[xs]=1
                                xs=len(rawxlx[specn+1])
                            else:
                                xs=xs+1
                            if xs>(len(rawxlx[specn+1])-1):
                                gofind=0
                        if found==0:
                            cmzraw.append(rawxlx[specn][x])
                            cintraw.append(rawxly[specn][x])
                        x=x+1
                # begin go through foundlist and insert values from specn+1 with associated 0 value in foundlist
                xf=3
                while xf<(len(foundlist)):
                    if foundlist[xf]==0:
                        if str(rawxlx[specn+1][xf])=='unmatched':
                            xf=xf+2
                        elif fourlettcode in str(rawxlx[specn+1][xf]):
                            xf=xf+2
                        else:
                            xff=3
                            goinsert=1
                            while goinsert==1:
                                if fourlettcode in str(cmzraw[xff]):
                                    xff=xff+2
                                elif str(cmzraw[xff])=='unmatched':
                                    xff=xff+2
                                else:
                                    if xff<(len(cmzraw)-1):
                                        if fourlettcode in str(cmzraw[xff+1]):
                                            xff=xff+2
                                        elif str(cmzraw[xff+1])=='m/z':
                                            xff=xff+1
                                        elif str(cmzraw[xff+1])=='unmatched':
                                            xff=xff+2
                                        elif str(cmzraw[xff+1])=='all':
                                            goinsert=0
                                        elif float(rawxlx[specn+1][xf])>float(cmzraw[xff]):     
                                            if float(rawxlx[specn+1][xf])<float(cmzraw[xff+1]):
                                                cmzraw.insert(xff+1, float(rawxlx[specn+1][xf]))
                                                cintraw.insert(xff+1, float(rawxly[specn+1][xf]))
                                                goinsert=0
                                        xff=xff+1
                                    else:
                                        goinsert=0
                                if xff>(len(cmzraw)-1):
                                    goinsert=0
                            xf=xf+1
                    else:
                        xf=xf+1
                # end go through foundlist and insert values from specn+1 with associated 0 value in foundlist
                specn=specn+1
            elif scount==0:
                # take species as is, only one MS/MS
                cmzraw=rawxlx[specn]
                cintraw=rawxly[specn]
                specn=specn+1
                scount=0
            elif scount>1:
                ok=1
                specn=specn+1
                #scount=0
            else:
                specn=specn+1
                #scount=0

            # begin check for further spectra of same species from Skyline and add (update cmzraw and cintraw)
            if scount==1:
                gosky=1
                while gosky==1:
                    if (specn+1)>(len(rawxlx)-1):
                        gosky=0
                    else:
                        if str(rawxlx[specn][1])==str(rawxlx[specn+1][1]):    # found MSMS from Skyline
                            crt=str(rawxlx[specn][0][len(rawxlx[specn][0])-9])+str(rawxlx[specn][0][len(rawxlx[specn][0])-8])+str(rawxlx[specn][0][len(rawxlx[specn][0])-7])+str(rawxlx[specn][0][len(rawxlx[specn][0])-6])
                            nrt=str(rawxlx[specn+1][0][len(rawxlx[specn+1][0])-9])+str(rawxlx[specn+1][0][len(rawxlx[specn+1][0])-8])+str(rawxlx[specn+1][0][len(rawxlx[specn+1][0])-7])+str(rawxlx[specn+1][0][len(rawxlx[specn+1][0])-6])
                            if abs(float(crt)-float(nrt))<rttolerance:  # compare RT of current and next MS/MS to determine, whether to combine
                                foundlist=[]    # list with as many entries, at first all 0, as specn+1
                                kf=0
                                z=0
                                while kf<(len(rawxlx[specn+1])):
                                    if str(rawxlx[specn+1][kf])=='all':
                                        kf=len(rawxlx[specn+1])
                                    else:
                                        foundlist.append(z)
                                    kf=kf+1
                                x=1
                                gomatch=1
                                while gomatch==1:       # check cmzraw vs specn+1
                                    #print(rawxly[specn+1][x])
                                    if x>(len(cmzraw)-1):
                                        gomatch=0
                                    elif fourlettcode in str(cmzraw[x]):  #str(rawxlx[specn][x]):       # here check against cmzraw ??? (account for different numbers of matched m/z values in MSMS)
                                        x=x+2
                                    elif str(cmzraw[x])=='unmatched':  #str(rawxlx[specn][x])=='unmatched':
                                        x=x+2
                                    elif str(cmzraw[x])=='all': #str(rawxlx[specn][x])=='all':
                                        gomatch=0
                                        x=x+2
                                        #elif cmzraw[x] is None: #rawxlx[specn][x] is None:
                                        #    gomatch=0
                                    else:
                                        # begin combine value with Skyline matched or unmatched spectrum value and save in list
                                        scount=scount+1 ########## Note that another MSMS was found
                                        xs=3
                                        gofind=1
                                        #found=0
                                        while gofind==1:
                                            if str(rawxlx[specn+1][xs])=='all':
                                                gofind=0
                                            elif str(rawxlx[specn+1][xs])=='unmatched':
                                                xs=xs+2
                                            elif fourlettcode in str(rawxlx[specn+1][xs]):
                                                xs=xs+2
                                            elif xs>(len(rawxlx[specn+1])-1):
                                                gofind=0
                                            elif abs(float(cmzraw[x])-float(rawxlx[specn+1][xs]))<0.002:     # if m/z present in both MS/MS, add; if not use both values as is
                                                #found=1
                                                if foundlist[xs]==1:
                                                    print('Check code !!')
                                                ccint=float(cintraw[x])+float(rawxly[specn+1][xs])
                                                cintraw[x]=ccint
                                                foundlist[xs]=1
                                                xs=len(rawxlx[specn+1])
                                            else:
                                                xs=xs+1
                                            if xs>(len(rawxlx[specn+1])-1):
                                                gofind=0
                                        x=x+1
                                # begin go through foundlist and insert values from specn+1 with associated 0 value in foundlist
                                xf=3
                                while xf<(len(foundlist)):
                                    if foundlist[xf]==0:
                                        if str(rawxlx[specn+1][xf])=='unmatched':
                                            xf=xf+2
                                        elif fourlettcode in str(rawxlx[specn+1][xf]):
                                            xf=xf+2
                                        else:
                                            xff=3
                                            goinsert=1
                                            while goinsert==1:
                                                if fourlettcode in str(cmzraw[xff]):
                                                    xff=xff+2
                                                elif str(cmzraw[xff])=='unmatched':
                                                    xff=xff+2
                                                else:
                                                    if xff<(len(cmzraw)-1):
                                                        if fourlettcode in str(cmzraw[xff+1]):
                                                            xff=xff+2
                                                        elif str(cmzraw[xff+1])=='unmatched':
                                                            xff=xff+2
                                                        elif str(cmzraw[xff+1])=='m/z':
                                                            xff=xff+1
                                                        elif str(cmzraw[xff+1])=='all':
                                                            goinsert=0
                                                        elif float(rawxlx[specn+1][xf])>float(cmzraw[xff]):     
                                                            if float(rawxlx[specn+1][xf])<float(cmzraw[xff+1]):
                                                                cmzraw.insert(xff+1, float(rawxlx[specn+1][xf]))
                                                                cintraw.insert(xff+1, float(rawxly[specn+1][xf]))
                                                                goinsert=0
                                                        xff=xff+1
                                                    else:
                                                        goinsert=0
                                                if xff>(len(cmzraw)-1):
                                                    goinsert=0
                                            xf=xf+1
                                    else:
                                        xf=xf+1
                                # end go through foundlist and insert values from specn+1 with associated 0 value in foundlist
                                specn=specn+1
                            else:
                                gosky=0
                                specn=specn+1
                                scount=0
                        else:
                            gosky=0
                            specn=specn+1
                            scount=0
            # end check for further spectra of same species from Skyline and add (update cmzraw and cintraw)
            allmzraw.append(cmzraw)
            allintraw.append(cintraw)

        else:
            # Found single spectrum, save as is
            allmzraw.append(rawxlx[specn])
            allintraw.append(rawxly[specn])
            specn=specn+1
        #specn=specn+1
        #print('Round completed.')
        #specn=1000 #only calculate for first spectrum set

    # end combine spectra to one combined MSMS spectrum per species (distinguish between cases of MSMS from Skyline only or only from MassLynx)
    # begin save combined MS/MS spectra in excel file
    wb = Workbook()   #write_only=True)
    std=wb['Sheet']
    wb.remove(std)
    ws = wb.create_sheet('Combined_MSMS')
    r=0
    c=0
    while c<(len(allmzraw)):
        r=0
        while r<(len(allmzraw[c])):
            ws.cell(row=r+1, column=2*c+1).value=allmzraw[c][r]
            ws.cell(row=r+1, column=2*c+2).value=allintraw[c][r]
            r=r+1
        c=c+1

    wb.save('Combined_MSMS_spectra.xlsx')
    print('Combined MS/MS spectra are saved in excel file Combined_MSMS_spectra.xlsx.')
    # end save combined MS/MS spectra in excel file
###########################

# begin read excel file for development mode
if devmode==1:
    fourlettcode='AMPP'
    #read excel file
    wb=openpyxl.load_workbook('Combined_MSMS_spectra_development_input.xlsx')
    ws=wb['combined_MSMS']
    allmzraw=[]
    allintraw=[]
    r=1
    c=1
    goc=1
    while goc==1:
        r=1
        tfe=ws.cell(row=r, column=c)
        tfe=tfe.value
        if tfe is None:
            goc=0
        else:
            cmzraw=[]
            cintraw=[]
            r=1
            gor=1
            while gor==1:
                tfe=ws.cell(row=r, column=c)
                tfe=tfe.value
                if tfe is None:
                    gor=0
                else:
                    tfe=ws.cell(row=r, column=c)
                    tfe=tfe.value
                    tff=ws.cell(row=r, column=c+1)
                    tff=tff.value
                    cmzraw.append(tfe)
                    cintraw.append(tff)
                    r=r+1
            allmzraw.append(cmzraw)
            allintraw.append(cintraw)
        c=c+2
else:
    ok=1 # use allmzraw and allintraw as calculated above.
# end read excel file for development mode

print('Number of fatty acid isomers to be tested: %d' % len(allmzraw))
#print('Number of MSMS spectra: %d' % len(allintraw))
#print('Running calculations...')
#print('_________________________________________________')
#print(allmzraw[spec][1])

#spec=0
#currentFA=str()
#ccf=0
#while ccf<(len(allmzraw[spec][1])-23):
#    currentFA=currentFA+allmzraw[spec][1][ccf]
#    ccf=ccf+1
#print(currentFA)

# begin calculate characteristic pattern for each species
mztol=0.08  # for exclusion of characteristic pattern use each hypothetical m/z +- mz tolerance of 0.08
falist=[]   # list with FA species
snlist=[]   # list with signal to noise values
spec=0
while spec<(len(allmzraw)):
    charamz=[]                          # m/z values for characteristic pattern for exclusion
    cfa=str(allmzraw[spec][1])
    cchain=10*int(cfa[5])+int(cfa[6])   # number of C in FA
    # begin get precursor m/z
    sm=3
    gos=1
    while gos==1:
        if str(allmzraw[spec][sm])=='unmatched':
            gos=0
        elif fourlettcode in str(allmzraw[spec][sm]):
            sm=sm+2
            if 'precursor' in str(allmzraw[spec][sm-2]):
                goss=1
                precpeak=[]
                while goss==1:
                    if '_'==str(allintraw[spec][sm]):
                        goss=0
                    else:
                        precpeak.append(allintraw[spec][sm])
                        if float(allintraw[spec][sm])==max(precpeak):
                            cprecmz=allmzraw[spec][sm]
                            cprecsm=sm  
                        sm=sm+1
        else:
            sm=sm+1
    charamz.append(float(allmzraw[spec][cprecsm]))
    print('_________________________________________________')
    #print(allmzraw[spec][1])

    currentFA=str()
    ccf=0
    while ccf<(len(allmzraw[spec][1])-23):
        currentFA=currentFA+allmzraw[spec][1][ccf]
        ccf=ccf+1
    print(currentFA)

    #print(charamz[0])
    # end get precursor m/z
    #isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
    #imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
    #           0        1        2         3         4         
    charamz.append(float(allmzraw[spec][cprecsm])+imass[1]-imass[0])   # +1
    charamz.append(float(allmzraw[spec][cprecsm])+2*(imass[1]-imass[0])) # +2
    charamz.append(float(allmzraw[spec][cprecsm])+3*(imass[1]-imass[0])) # +3
    charamz.append(float(allmzraw[spec][cprecsm])+imass[4])   # +O
    charamz.append(float(allmzraw[spec][cprecsm])+3*imass[4]) # +O3
    ccc=cchain-2
    while ccc>1:
        # outermost dbs
        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4])   # ald
        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+(imass[1]-imass[0]))   # ald+1
        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4])   # cri
        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+(imass[1]-imass[0]))   # cri+1
        if int(allmzraw[spec][1][8])>1:
            if ccc<cchain-4:
                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+2*imass[0])   # ald
                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+(imass[1]-imass[0])+2*imass[0])   # ald+1
                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+2*imass[0])   # cri
                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+(imass[1]-imass[0])+2*imass[0])   # cri+1
                if int(allmzraw[spec][1][8])>2:
                    if ccc<cchain-6:
                        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+4*imass[0])   # ald
                        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+(imass[1]-imass[0])+4*imass[0])   # ald+1
                        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+4*imass[0])   # cri
                        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+(imass[1]-imass[0])+4*imass[0])   # cri+1
                        if int(allmzraw[spec][1][8])>3:
                            if ccc<cchain-8:
                                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+6*imass[0])   # ald
                                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+(imass[1]-imass[0])+6*imass[0])   # ald+1
                                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+6*imass[0])   # cri
                                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+(imass[1]-imass[0])+6*imass[0])   # cri+1
                                if int(allmzraw[spec][1][8])>4:
                                    if ccc<cchain-10:
                                        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+8*imass[0])   # ald
                                        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+(imass[1]-imass[0])+8*imass[0])   # ald+1
                                        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+8*imass[0])   # cri
                                        charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+(imass[1]-imass[0])+8*imass[0])   # cri+1
                                        if int(allmzraw[spec][1][8])>5:
                                            if ccc<cchain-12:
                                                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+10*imass[0])   # ald
                                                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+imass[4]+(imass[1]-imass[0])+10*imass[0])   # ald+1
                                                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+10*imass[0])   # cri
                                                charamz.append(float(allmzraw[spec][cprecsm])-(imass[2]*ccc+2*ccc*imass[0])+2*imass[4]+(imass[1]-imass[0])+10*imass[0])   # cri+1
        ccc=ccc-1

    # end calculate characteristic pattern for each species

    #print('Number of mz values in characteristic pattern: %d' % len(charamz))

    #print('Get signal integrals...')
    # begin find matched peaks for each species and calculate signal to noise values for each species
    signal=0     # signal is average integral of OzID fragment peaks 
    noise=0      # noise is average integral of noise peaks (integral of individual noise peaks / number of noise peaks)
    pkintlist=[]
    fs=1
    gofs=1
    while gofs==1:
        if 'precursor' in str(allmzraw[spec][fs]):
            gofs=0
        elif fourlettcode in str(allmzraw[spec][fs]):
            fs=fs+2
            pkint=0
            gopk=1
            while gopk==1:
                if '_'==str(allintraw[spec][fs]):
                    gopk=0
                else:
                    pkint=pkint+float(allintraw[spec][fs])
                    fs=fs+1
            pkintlist.append(pkint)
    signal=sum(pkintlist)/len(pkintlist)
    #print(signal)
    #print('Get noise integrals...')
    # next, calculate noise based on noise peaks outside of charamz +- mztol
    nointlist=[]    # integral of eligible noise peaks
    nnpk=0  # number of eligible noise peaks
    fs=3
    getn=0
    while fs<(len(allmzraw[spec])):
        if 'unmatched'==str(allmzraw[spec][fs]):
            getn=1
            fs=fs+2
        if getn==1:
            #print(allintraw[spec][fs])
            if str(allintraw[spec][fs])=='_':
                fs=len(allmzraw[spec])
            elif int(allintraw[spec][fs])==0:
                fs=fs+1
                #np=0
            else:
                characheck=0
                chx=0
                while chx<(len(charamz)):
                    if abs(float(allmzraw[spec][fs])-(charamz[chx]))<mztol:
                        characheck=1
                        chx=len(charamz)
                    chx=chx+1
                if characheck==0:
                    cnoint=float(allintraw[spec][fs])
                    goadd=1
                    while goadd==1:
                        fs=fs+1
                        if fs>(len(allintraw[spec])-1):
                            goadd=0
                        elif int(allintraw[spec][fs])==0:
                            goadd=0
                        else:
                            characheck=0
                            chx=0
                            while chx<(len(charamz)):
                                if abs(float(allmzraw[spec][fs])-(charamz[chx]))<mztol:
                                    characheck=1
                                    chx=len(charamz)
                                chx=chx+1
                            if characheck==0:
                                cnoint=cnoint+float(allintraw[spec][fs])
                    nointlist.append(cnoint)
                    nnpk=nnpk+1
                else:
                    fs=fs+1
        else:
            fs=fs+1
    noise=sum(nointlist)/len(nointlist)
    #print('Number of noise peaks is %d.' % len(nointlist))
    #print('Noise value is %d.' % noise)
    # end find matched peaks for each species and calculate signal to noise values for each species
    csn=round(signal/noise, 1)
    cfa=str(allmzraw[spec][1])
    cfaa=str()
    cgo=1
    cgi=0
    while cgo==1:
        if cgi>7:
            if cgi<(len(cfa)-2):
                if str(cfa[cgi+1])=='a':
                    cgo=0
            else:
                cgo=0
        if cgo==1:
            cfaa=cfaa+cfa[cgi]
        cgi=cgi+1
    falist.append(cfaa)
    snlist.append(csn)
    #print(cfa)
    print('SIGNAL to NOISE: %.1f' % csn)
    spec=spec+1
###########################
print('_________________________________________________')
print('Saving final S/N values...')
# begin write Signal to Noise values for each species in file
wb = Workbook()   #write_only=True)
std=wb['Sheet']
wb.remove(std)
ws = wb.create_sheet('SN_MSMS')
r=0
while r<(len(falist)):        
    ws.cell(row=r+1, column=1).value=falist[r]
    ws.cell(row=r+1, column=2).value=snlist[r]
    r=r+1

wb.save('Signal_to_noise_of_MSMS_spectra.xlsx')
# end write Signal to Noise values for each species in file
###########################

#print('Combined MS/MS spectra are saved in: Combined_MSMS_spectra.xlsx.')
print('S/N values are saved in Signal_to_noise_of_MSMS_spectra.xlsx in sheet SN_MSMS.')
print('Use this information to delete false positive identifications from Skyline file OzFAD1_2_DDA_found.sky')


