# -*- coding: UTF-8 -*-

#Jan Philipp Menzel 
#created: 2022
#Notes: Generate segmented bar chart with matplotlib of fatty acid isomer relative abundances as workflow output (replicate plot of 3 replicates)
import math
import datetime
import pandas as pd
import matplotlib.cm as cm
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
# ask for number of replicates

#defined otherwise in line 539:  #This colour scheme defines the colours in the plot:
#colourschemebarchart=['n-2  ', 'n-3           ', 'n-4   ', 'n-5    ', 'n-6 ', 'n-7         ', 'n-8        ', 'n-9   ', 'n-10     ', 'n-11  ', 'n-12  ', 'n-13', 'n-14   ', 'n-15', 'n-16']
#colourschemebarchart=['coral', 'cornflowerblue', 'silver', 'magenta', 'gold', 'mediumpurple', 'deepskyblue', 'sienna', 'limegreen', 'yellow', 'darkorange', 'red', 'seagreen', 'cyan', 'blue']

# Backup standard scheme:
#colourschemebarchart=['n-2  ', 'n-3           ', 'n-4   ', 'n-5    ', 'n-6 ', 'n-7         ', 'n-8        ', 'n-9   ', 'n-10     ', 'n-11  ', 'n-12  ', 'n-13', 'n-14   ', 'n-15', 'n-16']
colourschemebarchart=['salmon', 'cornflowerblue', 'silver', 'magenta', 'gold', 'mediumpurple', 'deepskyblue', 'sienna', 'limegreen', 'yellow', 'orange', 'red', 'seagreen', 'cyan', 'blue']


#defined otherwise in line 698:
#colourschemebarchartlegend=['white', 'white', 'white', 'white', 'white', 'blue', 'cyan', 'seagreen', 'red', 'orange', 'yellow', 'limegreen', 'sienna', 'deepskyblue', 'mediumpurple', 'gold', 'magenta', 'silver', 'cornflowerblue', 'salmon']

colourschemebarchartlegend=['white', 'white', 'white', 'white', 'white']        # build legend colour scheme based on plot colour scheme
csb=len(colourschemebarchart)-1
while csb>-1:
    cap=str(colourschemebarchart[csb])
    colourschemebarchartlegend.append(cap)
    csb=csb-1

#print(colourschemebarchart)
#print(colourschemebarchartlegend)

print('This program is part of the OzFAD1 workflow and creates a replicate plot from three fatty acid analysis excel files.')
print('The data needs to be in files with the following filenames:')
print('OzFAD1_5_plot_table_rep1.xlsx')
print('OzFAD1_5_plot_table_rep2.xlsx')
print('OzFAD1_5_plot_table_rep3.xlsx')
# begin read excel files into dataframes and calculate average values

#wb1=openpyxl.load_workbook('jpmlipidomics_vpw20_10_quantified_final_rep1.xlsx', data_only=True)
wb1=openpyxl.load_workbook('OzFAD1_5_plot_table_rep1.xlsx', data_only=True)
sheetr1=wb1['final_barchart']
#wb2=openpyxl.load_workbook('jpmlipidomics_vpw20_10_quantified_final_rep2.xlsx', data_only=True)
wb2=openpyxl.load_workbook('OzFAD1_5_plot_table_rep2.xlsx', data_only=True)
sheetr2=wb2['final_barchart']
#wb3=openpyxl.load_workbook('jpmlipidomics_vpw20_10_quantified_final_rep3.xlsx', data_only=True)
wb3=openpyxl.load_workbook('OzFAD1_5_plot_table_rep3.xlsx', data_only=True)
sheetr3=wb3['final_barchart']
# begin determine dfcolumns
print('Excel files with replicates are read. Calculating ...')
dfcolumns=[] # list with label for n position
cc=2
go=1
while go==1:
    e=sheetr1.cell(row=1, column=cc)
    e=e.value
    if e is None:
        go=0
        e=str(e)
    elif str(e)=='':
        go=0
    else:
        dfcolumns.append(e)
    cc=cc+1
# end determine dfcolumns
#print(len(dfcolumns))
#print(dfcolumns)
# begin determine dfindex and data for each index (16:1, ...)
rowsdf1=[]
rowsdf2=[]
rowsdf3=[]

datasize=int(len(dfcolumns))

dfindex1=[] # list with labels for FA species in rep1
cr=2
go=1
while go==1:
    e=sheetr1.cell(row=cr, column=1)
    e=e.value
    if e is None:
        go=0
    elif str(e)=='None':
        go=0
    elif str(e)=='':
        go=0
    else:
        dfindex1.append(e)
        crow=[]
        cri=2
        while cri<(datasize+2):
            crd=sheetr1.cell(row=cr, column=cri)
            crd=crd.value
            #print(crd)
            crd=float(crd)
            crow.append(crd)
            cri=cri+1
        rowsdf1.append(crow)
    cr=cr+1
dfindex2=[]
cr=2
go=1
while go==1:
    e=sheetr2.cell(row=cr, column=1)
    e=e.value
    if e is None:
        go=0
    elif str(e)=='':
        go=0
    else:
        dfindex2.append(e)
        crow=[]
        cri=2
        while cri<(datasize+2):
            crd=sheetr2.cell(row=cr, column=cri)
            crd=crd.value
            crd=float(crd)
            crow.append(crd)
            cri=cri+1
        rowsdf2.append(crow)
    cr=cr+1
cr=2
dfindex3=[]
go=1
while go==1:
    e=sheetr3.cell(row=cr, column=1)
    e=e.value
    if e is None:
        go=0
    elif str(e)=='':
        go=0
    else:
        dfindex3.append(e)
        crow=[]
        cri=2
        while cri<(datasize+2):                                   ############################################### adjusted to size of input data
            crd=sheetr3.cell(row=cr, column=cri)
            crd=crd.value
            crd=float(crd)
            crow.append(crd)
            cri=cri+1
        rowsdf3.append(crow)
    cr=cr+1
# check if index different for replicates
#print(dfindex1)
#print(dfindex2)
#print(dfindex3)

#print('###############################')
#print(len(dfindex1))
#print('###############################')

#print(rowsdf1)
#print(rowsdf2)
#print(rowsdf3)

#print('###############################')
#print(len(rowsdf1))
#print(len(rowsdf1[0]))
#print('###############################')
#quit()



# begin define function for adding empty row to data        ############################# NEW FUNCTION HERE - HOW TO APPEND LIST TO LIST OF LISTS ??
erow=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
#print('len(erow)')
#print(len(erow))

#def add_row1(addrowsdf1, pr):
#    rowsdf1.insert(pr, erow)
#    return rowsdf1
# end define function for adding empty row to data




# begin check species in rep 1 and add in rep2 and rep3 if required
ri=0
while ri<len(dfindex1):
    if dfindex1[ri] in dfindex2:
        ok=1
    else:
        # begin add an empty row to data associated to dfindex2
        # begin add species to dfindex2
        rib=0
        while rib<len(dfindex2):
            if int(dfindex1[ri][3])==int(dfindex2[rib][3]):
                if (10*int(dfindex1[ri][0])+int(dfindex1[ri][1]))<(10*int(dfindex2[rib][0])+int(dfindex2[rib][1])):
                    if rib>0:
                        if (10*int(dfindex1[ri][0])+int(dfindex1[ri][1]))>(10*int(dfindex2[rib-1][0])+int(dfindex2[rib-1][1])):
                            dfindex2.insert(rib, dfindex1[ri])
                            rowsdf2.insert(rib, erow)   # add empty row to dataset                            
                    else:
                        dfindex2.insert(rib, dfindex1[ri])
                        rowsdf2.insert(rib, erow)   # add empty row to dataset
                else:
                    if int(dfindex1[ri][3])<int(dfindex3[rib+1][3]):
                        dfindex2.insert(rib+1, dfindex1[ri])
                        rowsdf2.insert(rib+1, erow)   # add empty row to dataset
            else:
                if dfindex2[rib][3]<dfindex1[ri][3]:
                    if dfindex2[rib+1][3]>dfindex1[ri][3]:
                        dfindex2.insert(rib+1, dfindex1[ri])
                        rowsdf2.insert(rib+1, erow)   # add empty row to dataset
            if dfindex1[ri] in dfindex2:
                rib=len(dfindex2)+1
            else:
                rib=rib+1
        # end add an empty row to data associated to dfindex2
        # end add species to dfindex2
    if dfindex1[ri] in dfindex3:
        ok=1
    else:
        # begin add an empty row to data associated to dfindex3
        # begin add species to dfindex3
        rib=0
        while rib<len(dfindex3):
            if int(dfindex1[ri][3])==int(dfindex3[rib][3]):
                if (10*int(dfindex1[ri][0])+int(dfindex1[ri][1]))<(10*int(dfindex3[rib][0])+int(dfindex3[rib][1])):
                    if rib>0:
                        if (10*int(dfindex1[ri][0])+int(dfindex1[ri][1]))>(10*int(dfindex3[rib-1][0])+int(dfindex3[rib-1][1])):
                            dfindex3.insert(rib, dfindex1[ri])
                            rowsdf3.insert(rib, erow)   # add empty row to dataset
                    else:
                        dfindex3.insert(rib, dfindex1[ri])
                        rowsdf3.insert(rib, erow)   # add empty row to dataset
                else:
                    if int(dfindex1[ri][3])<int(dfindex3[rib+1][3]):
                        dfindex3.insert(rib+1, dfindex1[ri])
                        rowsdf3.insert(rib+1, erow)   # add empty row to dataset
                    else:
                        ok=1
                        # do nothing here, as species to be added soon
                        #print('MODIFY CODE IN LINE 129')
                        #print(dfindex1[ri])
                        #print(dfindex3[rib])
            else:
                if dfindex3[rib][3]<dfindex1[ri][3]:
                    if dfindex3[rib+1][3]>dfindex1[ri][3]:
                        dfindex3.insert(rib+1, dfindex1[ri])
                        rowsdf3.insert(rib+1, erow)   # add empty row to dataset
            if dfindex1[ri] in dfindex3:
                rib=len(dfindex3)+1
            else:
                rib=rib+1
        # end add an empty row to data associated to dfindex2
        # end add species to dfindex2
    ri=ri+1
# end check species in rep 1 and add in rep2 and rep3 if required         
# begin check species in rep 2 and add in rep1 and rep3 if required
ri=0
while ri<len(dfindex2):
    if dfindex2[ri] in dfindex1:
        ok=1
    else:
        # begin add an empty row to data associated to dfindex2
        # begin add species to dfindex2
        rib=0
        while rib<len(dfindex1):
            if int(dfindex2[ri][3])==int(dfindex1[rib][3]):
                if (10*int(dfindex2[ri][0])+int(dfindex2[ri][1]))<(10*int(dfindex1[rib][0])+int(dfindex1[rib][1])):
                    if rib>0:
                        if (10*int(dfindex2[ri][0])+int(dfindex2[ri][1]))>(10*int(dfindex1[rib-1][0])+int(dfindex1[rib-1][1])):
                            dfindex1.insert(rib, dfindex2[ri])
                            rowsdf1.insert(rib, erow)   # add empty row to dataset
                    else:
                        dfindex1.insert(rib, dfindex2[ri])
                        rowsdf1.insert(rib, erow)   # add empty row to dataset
                elif rib==len(dfindex1):
                    ok=1
                elif int(dfindex2[ri][3])<int(dfindex1[rib+1][3]):
                    dfindex1.insert(rib, dfindex2[ri])
                    rowsdf1.insert(rib, erow)   # add empty row to dataset
            elif int(dfindex2[ri][3])>int(dfindex1[rib][3]):
                if rib==len(dfindex1):
                    dfindex1.insert(rib, dfindex2[ri])
                    rowsdf1.insert(rib, erow)   # add empty row to dataset
                else:
                    if int(dfindex2[ri][3])<int(dfindex1[rib+1][3]):
                        dfindex1.insert(rib, dfindex2[ri])
                        rowsdf1.insert(rib, erow)   # add empty row to dataset    
            rib=rib+1
        # end add an empty row to data associated to dfindex1
        # end add species to dfindex1
    if dfindex2[ri] in dfindex3:
        ok=1
    else:
        # begin add an empty row to data associated to dfindex3
        # begin add species to dfindex3
        rib=0
        while rib<len(dfindex3):
            if int(dfindex2[ri][3])==int(dfindex3[rib][3]):
                if (10*int(dfindex2[ri][0])+int(dfindex2[ri][1]))<(10*int(dfindex3[rib][0])+int(dfindex3[rib][1])):
                    if rib>0:
                        if (10*int(dfindex2[ri][0])+int(dfindex2[ri][1]))>(10*int(dfindex3[rib-1][0])+int(dfindex3[rib-1][1])):
                            dfindex3.insert(rib, dfindex2[ri])
                            rowsdf3.insert(rib, erow)   # add empty row to dataset
                    else:
                        dfindex3.insert(rib, dfindex2[ri])
                        rowsdf3.insert(rib, erow)   # add empty row to dataset
                else:
                    if int(dfindex2[ri][3])<int(dfindex3[rib+1][3]):
                        dfindex3.insert(rib+1, dfindex2[ri])
                        rowsdf3.insert(rib+1, erow)   # add empty row to dataset
                    else:
                        ok=1
            else:
                if dfindex3[rib][3]<dfindex2[ri][3]:
                    if dfindex3[rib+1][3]>dfindex2[ri][3]:
                        dfindex3.insert(rib+1, dfindex2[ri])
                        rowsdf3.insert(rib+1, erow)   # add empty row to dataset
            if dfindex2[ri] in dfindex3:
                rib=len(dfindex3)+1
            else:
                rib=rib+1
        # end add an empty row to data associated to dfindex2
        # end add species to dfindex2
    ri=ri+1
# end check species in rep 2 and add in rep1 and rep3 if required
# begin check species in rep 3 and add in rep1 and rep2 if required
ri=0
while ri<len(dfindex3):
    if dfindex3[ri] in dfindex2:
        ok=1
    else:
        # begin add an empty row to data associated to dfindex2
        # begin add species to dfindex2
        rib=0
        while rib<len(dfindex2):
            if int(dfindex3[ri][3])==int(dfindex2[rib][3]):
                if (10*int(dfindex3[ri][0])+int(dfindex3[ri][1]))<(10*int(dfindex2[rib][0])+int(dfindex2[rib][1])):
                    if rib>0:
                        if (10*int(dfindex3[ri][0])+int(dfindex3[ri][1]))>(10*int(dfindex2[rib-1][0])+int(dfindex2[rib-1][1])):
                            dfindex2.insert(rib, dfindex3[ri])
                            rowsdf2.insert(rib, erow)   # add empty row to dataset
                    else:
                        dfindex2.insert(rib, dfindex3[ri])
                        rowsdf2.insert(rib, erow)   # add empty row to dataset
                else:
                    if int(dfindex3[ri][3])<int(dfindex2[rib+1][3]):
                        dfindex2.insert(rib+1, dfindex3[ri])
                        rowsdf2.insert(rib+1, erow)   # add empty row to dataset
            else:
                if dfindex2[rib][3]<dfindex3[ri][3]:
                    if dfindex2[rib+1][3]>dfindex3[ri][3]:
                        dfindex2.insert(rib+1, dfindex3[ri])
                        rowsdf2.insert(rib+1, erow)   # add empty row to dataset
            if dfindex3[ri] in dfindex2:
                rib=len(dfindex2)+1
            else:
                rib=rib+1
        # end add an empty row to data associated to dfindex2
        # end add species to dfindex2
    if dfindex3[ri] in dfindex1:
        ok=1
    else:
        # begin add an empty row to data associated to dfindex3
        # begin add species to dfindex3
        rib=0
        while rib<len(dfindex1):
            if int(dfindex3[ri][3])==int(dfindex1[rib][3]):
                if (10*int(dfindex3[ri][0])+int(dfindex3[ri][1]))<(10*int(dfindex1[rib][0])+int(dfindex1[rib][1])):
                    if rib>0:
                        if (10*int(dfindex3[ri][0])+int(dfindex3[ri][1]))>(10*int(dfindex1[rib-1][0])+int(dfindex1[rib-1][1])):
                            dfindex1.insert(rib, dfindex3[ri])
                            rowsdf1.insert(rib, erow)   # add empty row to dataset
                    else:
                        dfindex1.insert(rib, dfindex3[ri])
                        rowsdf1.insert(rib, erow)   # add empty row to dataset
                else:
                    if int(dfindex3[ri][3])<int(dfindex2[rib+1][3]):
                        dfindex1.insert(rib+1, dfindex3[ri])
                        rowsdf1.insert(rib+1, erow)   # add empty row to dataset
            else:
                if dfindex1[rib][3]<dfindex3[ri][3]:
                    if dfindex1[rib+1][3]>dfindex3[ri][3]:
                        dfindex1.insert(rib+1, dfindex3[ri])
                        rowsdf1.insert(rib+1, erow)   # add empty row to dataset
            if dfindex3[ri] in dfindex1:
                rib=len(dfindex1)+1
            else:
                rib=rib+1
        # end add an empty row to data associated to dfindex2
        # end add species to dfindex2
    ri=ri+1
# end check species in rep 3 and add in rep1 and rep2 if required



# end determine dfindex

#print(dfindex1)
#print(dfindex2)
#print(dfindex3)


# end read excel files into dataframes and calculate average values
#quit()

#print(" -     PASS 1")

def plot_clustered_stacked(dfall, labels=None, title='', H="/", **kwargs): #title="multiple stacked bar plot",  H="/", **kwargs):
    """Given a list of dataframes, with identical columns and index, create a clustered stacked bar plot. 
    labels is a list of the names of the dataframe, used for the legend
    title is a string for the title of the plot
    H is the hatch used for identification of the different dataframe"""
    #print(" -     PASS 5")
    n_df = len(dfall)               # number of dataframes in dfall (number of replicates)
    n_col = len(dfall[0].columns)   # number of columns (n-x positions)
    #print('n_col')
    #print(n_col)
    n_ind = len(dfall[0].index)     # number of rows    (FA species 16:1; 17:1 ...)
    #print('n_ind')
    #print(n_ind)
    axe = plt.subplot(111)

    for df in dfall : # for each data frame
        axe = df.plot(kind="bar",
                      linewidth=0,
                      stacked=True,
                      ax=axe,
                      legend=False,
                      grid=False,
                      **kwargs)  # make bar plots

    h,l = axe.get_legend_handles_labels() # get the handles we want to modify
    krep=0
    for i in range(0, n_df * n_col, n_col): # len(h) = n_col * n_df
        knx=0
        for j, pa in enumerate(h[i:i+n_col]):       # j (knx) is index for n-x position (columns)
            kfa=0    # count of FA species (16:1, 17:1 ...)
            for rect in pa.patches: # for each index
                rect.set_x(rect.get_x() + 1 / float(n_df + 1) * i / float(n_col))
                #rect.set_hatch(H * int(i / n_col)) #edited part     # diagonal pattern for distinguishing replicates
                #if j==1:    # disable later
                    #rect.set_color('r') #colourschemebarchart[j]) # set colour for specific field 
                    #rect.set(hatch='/', facecolor='r', edgecolor='black')   # 
                plt.rcParams['hatch.linewidth']=1.5
                if j==len(colourschemebarchart)-1:
                    rect.set(hatch=hatchschemebarchart[krep][kfa][knx], facecolor=colourschemebarchart[j], edgecolor='grey')   # set grey hatches for n-2 position (doesn't work)
                    #print('Well...')
                else:
                    rect.set(hatch=hatchschemebarchart[krep][kfa][knx], facecolor=colourschemebarchart[j], edgecolor='black')   # 
                #rect.set(hatch=hatchschemebarchart[krep][kfa][knx], facecolor=colourschemebarchart[j], edgecolor='black')   # 
                rect.set_width(1 / float(n_df + 1))
                kfa=kfa+1
            knx=knx+1
        krep=krep+1

    plt.rcParams['font.size']='14'  #OK

    plt.ylim(0,100) #OK

    axe.set_xticks((np.arange(0, 2 * n_ind, 2) + 1 / float(n_df + 1)) / 2.)
    axe.set_xticklabels(df.index, rotation = 90, font='Arial') #OK fontsize=28

    axe.set_yticks((0,10,20,30,40,50,60,70,80,90,100))
    axe.set_yticklabels((0,10,20,30,40,50,60,70,80,90,100),font='Arial')    #OK ,fontsize=28

    axe.set_title(title)

    # Add invisible data to add another legend
    n=[]        
    for i in range(n_df):
        n.append(axe.bar(0, 0, color="black", hatch=H * i))   #### n.append(axe.bar(0, 0, color="gray", hatch=H * i))

    l1 = axe.legend(h[:n_col], l[:n_col], loc=[1.01, 0.5])  #legend and defined position
    #if labels is not None:
    #    l2 = plt.legend(n, labels, loc=[1.01, 0.1])    # labed for replicates, only use if colours of replicates shaded
    axe.add_artist(l1)

    

    return axe
    
# begin take dataframes as lists of lists and then use mytuple=tuple(mylist) to convert lists into tuples and construct lists of tuples for assembly of dataframe
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
rf=0
rowsdf1b=[]
rowsdf2b=[]
rowsdf3b=[]
while rf<(len(rowsdf1)):
    ctuple1=tuple(rowsdf1[rf])
    rowsdf1b.append(ctuple1)
    ctuple2=tuple(rowsdf2[rf])
    rowsdf2b.append(ctuple2)
    ctuple3=tuple(rowsdf3[rf])
    rowsdf3b.append(ctuple3)
    rf=rf+1
# begin take dataframes as lists of lists and then use mytuple=tuple(mylist) to convert lists into tuples and construct lists of tuples for assembly of dataframe
#print(" -     PASS 2")
#rowsdf1 = [(10, 60, 30), (20, 50, 30)]  # each replicate in one dataframe, [(16:1), (18:1)]; rowsdf1 is one replicate / list of tuples; each tuple one species
#rowsdf2 = [(0, 69, 31), (21, 57, 22)]
#rowsdf3 = [(11, 63, 26), (14, 57, 29)]

#asrowsdf1 = [(1, 0, 3), (2, 1, 4)]  # assignment for each replicate in one dataframe, [(16:1), (18:1)], 0 is Me, 1 is Bu, 2 is trans or other NMI
#asrowsdf2 = [(1, 0, 3), (2, 1, 4)]
#asrowsdf3 = [(1, 0, 3), (2, 1, 4)]
casrowsdf1 = [0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4, 0, 1, 2, 3, 4]
#print(len(casrowsdf1))
asrowsdf1=[]
asrowsdf2=[]
asrowsdf3=[]
di=0
while di<(len(dfindex1)):
    asrowsdf1.append(casrowsdf1)
    asrowsdf2.append(casrowsdf1)
    asrowsdf3.append(casrowsdf1)
    di=di+1
#asrowsdf2=asrowsdf1
#asrowsdf3=asrowsdf1

# begin take dataframes as lists of lists and then use mytuple=tuple(mylist) to convert lists into tuples and construct lists of tuples for assembly of dataframe
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
rf=0
asrowsdf1b=[]
asrowsdf2b=[]
asrowsdf3b=[]
while rf<(len(asrowsdf1)):
    ctuple1=tuple(asrowsdf1[rf])
    asrowsdf1b.append(ctuple1)
    ctuple2=tuple(asrowsdf2[rf])
    asrowsdf2b.append(ctuple2)
    ctuple3=tuple(asrowsdf3[rf])
    asrowsdf3b.append(ctuple3)
    rf=rf+1
# begin take dataframes as lists of lists and then use mytuple=tuple(mylist) to convert lists into tuples and construct lists of tuples for assembly of dataframe


#print('len(dfindex1)')
#print(len(dfindex1))
#print(" -     PASS 3")

assignment=[]
assignment.append(asrowsdf1b)
assignment.append(asrowsdf2b)
assignment.append(asrowsdf3b)
#print(assignment)
#dfindex=["16:1", "18:1"]                # index and columns of dataframes need to be same for each replicates
#dfcolumns=["n-7", "n-9", "n-10"]

df1=pd.DataFrame(rowsdf1b, index=dfindex1, columns=dfcolumns)   # number of elements in index and columns need to match shape of lists of tuples
df2=pd.DataFrame(rowsdf2b, index=dfindex2, columns=dfcolumns)
df3=pd.DataFrame(rowsdf3b, index=dfindex3, columns=dfcolumns)

asdf1=pd.DataFrame(asrowsdf1b, index=dfindex1, columns=dfcolumns)   # number of elements in index and columns need to match shape of lists of tuples
asdf2=pd.DataFrame(asrowsdf2b, index=dfindex2, columns=dfcolumns)
asdf3=pd.DataFrame(asrowsdf3b, index=dfindex3, columns=dfcolumns)

#colourschemebarchart=['cornflowerblue', 'silver', 'magenta']
#colourschemebarchart=['salmon', 'cornflowerblue', 'silver', 'magenta', 'gold', 'mediumpurple', 'deepskyblue', 'sienna', 'limegreen', 'yellow', 'orange', 'red', 'seagreen', 'cyan', 'blue']

#backup standard colour scheme
#colourschemebarchart=['salmon', 'cornflowerblue', 'silver', 'magenta', 'gold', 'mediumpurple', 'deepskyblue', 'sienna', 'limegreen', 'yellow', 'orange', 'red', 'seagreen', 'cyan', 'blue']




#print(colourschemebarchart)    # begin extend colourschemebarchart
ncolcat=4       # higher than 0, if category is not defined by assignment list, but position in dataframes
ncb=len(colourschemebarchart)
while ncb>0:
    ccol=colourschemebarchart[ncb-1]
    ncc=0
    while ncc<ncolcat:
        colourschemebarchart.insert(ncb-1, str(colourschemebarchart[ncb-1]))
        ncc=ncc+1
    ncb=ncb-1
#print(colourschemebarchart)    # end extend colourschemebarchart

# begin make list hatchschemebarchart (replicate, FA_species, n-x position)
#hatchschemebarchart=[[['', '', '/'], ['', '', '']], [['', '', ''], ['', '|', '']], [['', '', ''], ['x', '', '']]]
hatchsymbols=['', '\\\\\\\\', '////', '||', 'xx']     # defines patterns via number/index   0=''= Me_cis   1='\\'=Me_trans   2='//'=Bu_cis  3='/'=NMI_cis  4='x'=NMI_trans 
hatchschemebarchart=[]
ilhatch=[]
ihatch=[]
asrep=0
while asrep<(len(assignment)):
    asfa=0
    ilhatch=[]
    while asfa<(len(assignment[0])):
        asnx=0
        ihatch=[]
        while asnx<(len(assignment[0][0])):
            ihatch.append(hatchsymbols[int(assignment[asrep][asfa][asnx])])
            asnx=asnx+1
        ilhatch.append(ihatch)
        asfa=asfa+1
    hatchschemebarchart.append(ilhatch)
    asrep=asrep+1
# end make list hatchschemebarchart (replicate, FA_species, n-x position)
#print(len(hatchschemebarchart[0]))
#print(hatchschemebarchart)





#print(" -     PASS 4")
# call function to plot: ##################################################################################################################################################
plot_clustered_stacked([df1, df2, df3],["rep1", "rep2", "rep3"])       # , cmap=plt.cm.viridis cmap argument allows colour scheme change

#print(df1)
#print(df2)
#print(df3)

#print(asdf1)
#print(asdf2)
#print(asdf3)

plt.show()      #display plot in separate window, copy from there as vector image into Adobe Illustrator to make figure for publication


print('Copy plot into power point or other program or save as svg / png.')
print('Legend will be generated separately (after plot was copied).')
check=1 #eval(input('Figure copied? YES:1 ::'))

# begin generate empty plot with legend

def plot_legend_clustered_stacked(dfall, labels=None, title='', H="/", **kwargs): #title="multiple stacked bar plot",  H="/", **kwargs):
    """Given a list of dataframes, with identical columns and index, create a clustered stacked bar plot. 
    labels is a list of the names of the dataframe, used for the legend
    title is a string for the title of the plot
    H is the hatch used for identification of the different dataframe"""

    n_df = len(dfall)               # number of dataframes in dfall (number of replicates)
    n_col = len(dfall[0].columns)   # number of columns (n-x positions)
    n_ind = len(dfall[0].index)     # number of rows    (FA species 16:1; 17:1 ...)
    axe = plt.subplot(111)

    for df in dfall : # for each data frame
        axe = df.plot(kind="bar",
                      linewidth=0,
                      stacked=True,
                      ax=axe,
                      legend=True, #False,
                      grid=False,
                      **kwargs)  # make bar plots

    h,l = axe.get_legend_handles_labels() # get the handles we want to modify
    krep=0
    for i in range(0, n_df * n_col, n_col): # len(h) = n_col * n_df
        knx=0
        for j, pa in enumerate(h[i:i+n_col]):       # j (knx) is index for n-x position (columns)
            kfa=0    # count of FA species (16:1, 17:1 ...)
            for rect in pa.patches: # for each index
                rect.set_x(rect.get_x() + 1 / float(n_df + 1) * i / float(n_col))
                #rect.set_hatch(H * int(i / n_col)) #edited part     # diagonal pattern for distinguishing replicates
                #if j==1:    # disable later
                    #rect.set_color('r') #colourschemebarchart[j]) # set colour for specific field 
                    #rect.set(hatch='/', facecolor='r', edgecolor='black')   # 
                plt.rcParams['hatch.linewidth']=1.5
                if j==len(colourschemebarchart)-1:
                    rect.set(hatch=hatchschemebarchart[krep][kfa][knx], facecolor=colourschemebarchartlegend[j], edgecolor='grey')   # set grey hatches for n-2 position
                    #print('Well...')
                else:
                    rect.set(hatch=hatchschemebarchart[krep][kfa][knx], facecolor=colourschemebarchartlegend[j], edgecolor='black')   # 
                rect.set_width(1 / float(n_df + 1))
                kfa=kfa+1
            knx=knx+1
        krep=krep+1

    axe.set_xticks((np.arange(0, 2 * n_ind, 2) + 1 / float(n_df + 1)) / 2.)
    axe.set_xticklabels(df.index, rotation = 0)
    axe.set_title(title)

    # Add invisible data to add another legend
    n=[]        
    for i in range(n_df):
        n.append(axe.bar(0, 0, color="black", hatch=H * i))  ### n.append(axe.bar(0, 0, color="gray", hatch=H * i))

    l1 = axe.legend(h[:n_col], l[:n_col], loc=[1.01, 0.05], prop={'family': 'Arial'})  #legend and defined position

    plt.rcParams['font.size']='14'  #OK
    plt.ylim(0,100) #OK

    axe.set_yticks((0,10,20,30,40,50,60,70,80,90,100))
    axe.set_yticklabels((0,10,20,30,40,50,60,70,80,90,100),font='Arial')    #OK ,fontsize=28

    axe.add_artist(l1)

    return axe


dfindex=["16:1"] #, "18:1"]                # index and columns of dataframes need to be same for each replicates
#dfcolumns=["n-3", "n-4", "n-5", "n-6", "n-7", "n-8", "n-9", "n-10", "n-11", "n-12", "n-13", "n-14", "n-15", "Me (Z)", "Me (E)", "NMI (Z)", "Bu (Z)", "NMI (E)"]
dfcolumns=["Branched", "NMI", "NMI (Bu)", "trans (E)", "cis (Z)", "n-16", "n-15", "n-14", "n-13", "n-12", "n-11", "n-10", "n-9", "n-8", "n-7", "n-6", "n-5", "n-4", "n-3", "n-2"]

rowsdf1 = [(5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5)] #, (20, 50, 30)]  # each replicate in one dataframe, [(16:1), (18:1)]; rowsdf1 is one replicate / list of tuples; each tuple one species
rowsdf2 = [(5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5)] #, (21, 57, 22)]
rowsdf3 = [(5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5)] #, (14, 57, 29)]

asrowsdf1 = [(4, 3, 2, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)] #, (2, 1, 4)]  # assignment for each replicate in one dataframe, [(16:1), (18:1)], 0 is Me, 1 is Bu, 2 is trans or other NMI
asrowsdf2 = [(4, 3, 2, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)] #, (2, 1, 4)]
asrowsdf3 = [(4, 3, 2, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)] #, (2, 1, 4)]

assignment=[]
assignment.append(asrowsdf1)
assignment.append(asrowsdf2)
assignment.append(asrowsdf3)

df4=pd.DataFrame(rowsdf1, index=dfindex, columns=dfcolumns)   # number of elements in index and columns need to match shape of lists of tuples
df5=pd.DataFrame(rowsdf2, index=dfindex, columns=dfcolumns)
df6=pd.DataFrame(rowsdf3, index=dfindex, columns=dfcolumns)


#colourschemebarchartlegend=['white', 'white', 'white', 'white', 'white', 'blue', 'cyan', 'seagreen', 'red', 'orange', 'yellow', 'limegreen', 'sienna', 'deepskyblue', 'mediumpurple', 'gold', 'magenta', 'silver', 'cornflowerblue', 'salmon']

#print(colourschemebarchartlegend)    # begin extend colourschemebarchartlegend
ncolcat=0       # higher than 0, if category is not defined by assignment list, but position in dataframes
ncb=len(colourschemebarchartlegend)
while ncb>0:
    ccol=colourschemebarchartlegend[ncb-1]
    ncc=0
    while ncc<ncolcat:
        colourschemebarchartlegend.insert(ncb-1, str(colourschemebarchartlegend[ncb-1]))
        ncc=ncc+1
    ncb=ncb-1
#print(colourschemebarchartlegend)    # end extend colourschemebarchartlegend

# begin make list hatchschemebarchart (replicate, FA_species, n-x position)
#hatchschemebarchart=[[['', '', '/'], ['', '', '']], [['', '', ''], ['', '|', '']], [['', '', ''], ['x', '', '']]]
hatchsymbols=['', '\\\\\\\\', '////', '||', 'xx']     # defines patterns via number/index   0=''= Me_cis   1='\\'=Me_trans   2='//'=Bu_cis  3='/'=NMI_cis  4='x'=NMI_trans 
hatchschemebarchart=[]
ilhatch=[]
ihatch=[]
asrep=0
while asrep<(len(assignment)):
    asfa=0
    ilhatch=[]
    while asfa<(len(assignment[0])):
        asnx=0
        ihatch=[]
        while asnx<(len(assignment[0][0])):
            ihatch.append(hatchsymbols[int(assignment[asrep][asfa][asnx])])
            asnx=asnx+1
        ilhatch.append(ihatch)
        asfa=asfa+1
    hatchschemebarchart.append(ilhatch)
    asrep=asrep+1
# end make list hatchschemebarchart (replicate, FA_species, n-x position)

plot_legend_clustered_stacked([df4, df5, df6],["rep1", "rep2", "rep3"])
plt.show()











