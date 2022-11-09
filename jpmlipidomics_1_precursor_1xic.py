# -*- coding: UTF-8 -*-

# Jan Philipp Menzel jpm_lipidomics_vpw13_1_precursor_tr.py
#created: 09 07 2020
#modified: regularly until 07 04 2021 
# Goal: STEP 1. Generate transition list for Skyline containing precursors (intact derivatized fatty acids), either one of seven pre-defined derivatives or any new structure
## Notes: Derivative, positive fixed charge
## NOTES: VIRTUAL PRECURSOR - PrecursorName and PrecursorMz are artificially set +Xe (only column 3 and 5), fragment transitions correct including precursor
## NOTES: Virtual precursor forces Skyline to consider all transitions incl. real precursor (fragment in transition list), Skyline Setting: TransitionSettings-Filter-IonTypes-f 
import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
################ DATABASE ## Source: Internetchemie.info 
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.757, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
################
#default=eval(input('Run workflow with default parameters? (Yes: 1 / No: 0) (Derivatization agent: AMPP; Slow and full workflow including all FA; Apply retention time limitation; Use Fatty Acid Library to prevent filtering out important FA; Max RT = 17.5 min; FA: 12 - 24 C; Max m/z error: 10 ppm; Precursor peak area threshold: 3000; Product peak area threshold: 200.) :'))
default=0
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
###########

# begin read workflow parameters
transferlist=[]
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
go=1
while go==1:
	tfe=ws.cell(row=tli+1, column=2)
	tfe=tfe.value
	if tli==21:
		go=0
	if tfe is None:
		go=0
	else:
		transferlist.append(tfe)
	tli=tli+1
########## 
fourlettcode=transferlist[0]		#input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
cderiv=int(transferlist[1])
hderiv=int(transferlist[2])
dderiv=int(transferlist[3])
nderiv=int(transferlist[4])
oderiv=int(transferlist[5])
pderiv=int(transferlist[6])
ideriv=int(transferlist[7])
minlenfa=transferlist[8] #eval(input('Enter number of C atoms in shortest FA chain (e.g. 4) :'))
maxlenfa=transferlist[9] #eval(input('Enter number of C atoms in longest FA chain (e.g. 24) :'))
mzcutoff=int(transferlist[10]) 		#eval(input('What is max mz [ppm] for positive identification of species? (e.g. 30) :'))
rettimecutoff=float(transferlist[11]) 		#eval(input('What is the maximum retention time [min]? (e.g. 11.9 or 17.8) :'))
areathreshold=int(transferlist[12])		#eval(input('What is the threshold for detected precursor peak area? (compare Skyline report file, e.g. 3000) :'))
workflowmode=int(transferlist[13])
rtlimitation=int(transferlist[16])
mostwanted=int(transferlist[17])
runprecheck=int(transferlist[19])
identifier=str(transferlist[20])
# end read workflow parameters
if runprecheck==1:
	# begin derive setpoint and slope from chromatograms of precheck analysis
	convertfile=0
	if convertfile==0:
		xictimesdf=pd.read_csv('skyl_xic_report_vpw20_0_times.csv', header=None, skiprows=1, nrows=1)
		xictimeslistfromdf=xictimesdf.values.tolist()
		xictimeslist=xictimeslistfromdf[0]		# contains times of XICs # actual list starts at index 8

	segtrdf=pd.read_csv('skyl_xic_report_vpw20_0_intensities.csv', skiprows=1, header=None, nrows=3, low_memory=False)
	allxiclist=segtrdf.values.tolist()
	#begin write trdf to csv file to check its contents 	# TROUBLESHOOTING
	#trdf.to_csv('skyl_xic_report_vpw16_3_troubleshooting.csv', index=False)
	#end write trdf to csv file to check its contents 	# TROUBLESHOOTING
	#trdf=trdf.transpose()#

	#print(allxiclist[0][1])		####################################### 
	kix=len(allxiclist)
	#print('Number of rows in segment in skyl_xic_report_vpw20_0_precheck.csv: %d' % kix)
	#begin determine columns in xic_results, determine length of XIC
	ci=len(allxiclist[0])
	xiclength=int((ci-8)) #/2
	#print('XIC length')
	#print(xiclength)
	#end determine columns in xic_results, determine length of XIC
	intalist=[]
	intblist=[]
	itl=8
	while itl<(len(allxiclist[0])):
		intalist.append(int(allxiclist[0][itl]))
		intblist.append(int(allxiclist[1][itl]))
		itl=itl+1
	maxinta=max(intalist)
	maxintb=max(intblist)
	#print('maxinta')
	#print(maxinta)
	#print('maxintb')
	#print(maxintb)
	itl=8
	while itl<(len(allxiclist[0])):
		if allxiclist[0][itl]==maxinta:
			ita=itl
		if allxiclist[1][itl]==maxintb:
			itb=itl
		itl=itl+1
	rtmaxa=xictimeslist[ita]
	rtmaxb=xictimeslist[itb]
	#print('rtmaxa')
	#print(rtmaxa)
	#print('rtmaxb')
	#print(rtmaxb)
	setpoint=(rtmaxa+rtmaxb)/2
	slope=(rtmaxb-rtmaxa)/2
	#print('Arrived here')
	# end derive setpoint and slope from chromatograms of precheck analysis
else:
	rtlimitation=2
	setpoint=8.14	#eval(input('Setpoint: '))
	slope=0.57		#eval(input('Slope: '))

predictedrtindex=[]	# list containing species identifier ['04:0', '05:0', ...'30:6', ... '40:6']
predictedrtexact=[]
predictedrtmin=[]	# list containing predicted RT range start for species identifier with same list index
predictedrtmax=[]	# list containing predicted RT range end for species identifier with same list index
predictionconstantsat=[-13.91, -13.8, -13.2, -11.9, -10.575, -9.2, -7.785, -6.6, -5.4, -4.208, -3.075, -2.02, -1, 0, 1, 1.94, 2.85, 3.68, 4.525, 5.34, 6.139, 6.92, 7.668, 8.4, 9.13, 9.84, 10.53, 11.2, 11.85, 12.48, 13.09, 13.716, 14.326, 14.92, 15.519, 16.06, 16.629]
predictionconstantmono=[1.43, 1.43, 1.43, 1.43, 1.43, 1.45, 1.47, 1.47, 1.46, 1.46, 1.47, 1.44, 1.43, 1.43, 1.43, 1.45, 1.45, 1.44, 1.42, 1.41, 1.4, 1.41, 1.4, 1.4, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39, 1.39]
predictionconstantbis=[1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.21, 1.2, 1.2, 1.21, 1.2, 1.21, 1.19, 1.17, 1.18, 1.18, 1.18, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17, 1.17]
predictionconstanttri=[1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1, 1.025, 1.05, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025, 1.025]
predictionconstanttetra=[0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72]
predictionconstantpenta=[1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04, 1.04]
predictionconstanthexa=[0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73, 0.73]
predictionconstant=[]
predictionconstant=predictionconstant+predictionconstantsat
predictionconstant=predictionconstant+predictionconstantmono
predictionconstant=predictionconstant+predictionconstantbis
predictionconstant=predictionconstant+predictionconstanttri
predictionconstant=predictionconstant+predictionconstanttetra
predictionconstant=predictionconstant+predictionconstantpenta
predictionconstant=predictionconstant+predictionconstanthexa
#print(predictionconstant[9])
pr=0
go=1
ndb=0
nc=4
while go==1:
	cfaid=str()
	if nc<10:
		cfaid=cfaid+'0'+str(nc)+':'+str(ndb)
	else:
		cfaid=cfaid+str(nc)+':'+str(ndb)
	predictedrtindex.append(cfaid)
	# begin make prediction
	if ndb==0:
		predicted=setpoint+(slope*(float(predictionconstant[pr])))
		if predicted<0.05:
			predicted=0.05
		predictedrtexact.append(predicted)
		if (0.15*predicted)>1.0:
			prtmin=predicted-(0.15*predicted)
			if prtmin<0.05:
				prtmin=0.05
			prtmax=predicted+(0.15*predicted)
		else:
			prtmin=predicted-(1.0)
			if prtmin<0.05:
				prtmin=0.05
			prtmax=predicted+(1.0)
		predictedrtmin.append(prtmin)
		predictedrtmax.append(prtmax)
	else:
		predicted=(float(predictedrtexact[pr-37]))-(slope*(float(predictionconstant[pr])))
		if predicted<0.05:
			predicted=0.05
		predictedrtexact.append(predicted)
		if (0.15*predicted)>1.0:
			prtmin=predicted-(0.15*predicted)
			if prtmin<0.05:
				prtmin=0.05
			prtmax=predicted+(0.15*predicted)
		else:
			prtmin=predicted-(1.0)
			if prtmin<0.05:
				prtmin=0.05
			prtmax=predicted+(1.0)
		predictedrtmin.append(prtmin)
		predictedrtmax.append(prtmax)
	# end make prediction
	if nc<40:
		go=1
	else:
		nc=3
		if ndb<6:
			ndb=ndb+1
		else:
			go=0
	nc=nc+1
	pr=pr+1
# end read info from workflow parameters and use, do not request new, make indexed RT prediction of range, calc +- 1 min or +- 15% of RT
print('Setpoint:')
print(setpoint)
print('Slope:')
print(slope)
###
#print(predictedrtindex)				### Write predicted range information in file to use in precursor analysis as RT limitation
#print(len(predictedrtindex))
#print(predictedrtexact)
#print(len(predictedrtexact))
#print(predictedrtmin)
#print(predictedrtmax)
#quit()

# begin write results of retention time range prediction in file
prpwritelist=[]
prpwritelist.append(predictedrtindex)
prpwritelist.append(predictedrtexact)
prpwritelist.append(predictedrtmin)
prpwritelist.append(predictedrtmax)
prpdf=pd.DataFrame(prpwritelist).transpose()
prpdf.columns=['FA', 'RT_predicted', 'RT_min_predicted', 'RT_max_predicted']
filename='Predicted_RT_range.csv'
prpdf.to_csv(filename, index=False)
print('List with predicted retention time ranges is saved as Predicted_RT_range.csv')

#prp=0
#while prp<len(predictedrtindex):
#	ok=1

# end write results of retention time range prediction in file



nchunks=1
# begin calculate monounsaturated precursors from input
beforeall=datetime.datetime.now()
print('Workflow is running ...')
lfa=maxlenfa-minlenfa+1
moleculegrouplist=[]
precursornamelist=[]
precursorformulalist=[]
precursoradductlist=[]
precursormzlist=[]
precursorchargelist=[]
productnamelist=[]
productformulalist=[]
productadductlist=[]
productmzlist=[]
productchargelist=[]
molg=fourlettcode+'_FA'
padd='[M]1+'
prdn='precursor'
li=0
while li<lfa:
	moleculegrouplist.append(molg)
	clfa=li+minlenfa
	currentlfa=str(clfa)
	if len(currentlfa)<2:
		currentlfa='0'+currentlfa
		currentlfa=str(currentlfa)
	pnm=fourlettcode+'_'+currentlfa+':1'
	precursornamelist.append(pnm)
	prf=''
	currentcderiv=0
	currenthderiv=0
	currentdderiv=0
	currentnderiv=0
	currentoderiv=0
	currentpderiv=0
	currentideriv=0
	if cderiv>0:
		currentcderiv=cderiv+int(currentlfa)
		prf=prf+'C'+str(currentcderiv)
	if hderiv>0:
		currenthderiv=hderiv+(2*int(currentlfa))-3
		prf=prf+'H'+str(currenthderiv)
	if dderiv>0:
		currentdderiv=dderiv
		prf=prf+"H'"+str(currentdderiv)
	if nderiv>0:
		currentnderiv=nderiv
		prf=prf+'N'+str(nderiv)
	if oderiv>(-1):
		currentoderiv=oderiv+1
		prf=prf+'O'+str(currentoderiv)
	if pderiv>0:
		currentpderiv=pderiv
		prf=prf+'P'+str(pderiv)
	if ideriv>0:
		currentideriv=ideriv
		prf=prf+'I'+str(ideriv)
	precursorformulalist.append(prf)
	productformulalist.append(prf)
	precursoradductlist.append(padd)
	prmz=imass[0]*currenthderiv+imass[1]*currentdderiv+imass[2]*currentcderiv+imass[3]*currentnderiv+imass[4]*currentoderiv+imass[5]*currentpderiv+imass[10]*currentideriv
	precursormzlist.append(prmz)
	productmzlist.append(prmz)
	precursorchargelist.append(1)
	productnamelist.append(prdn)
	productadductlist.append(padd)
	productchargelist.append(1)
	li=li+1
fareadlist=[]
fareadlist.append(moleculegrouplist)
fareadlist.append(precursornamelist)
fareadlist.append(precursorformulalist)
fareadlist.append(precursoradductlist)
fareadlist.append(precursormzlist)
fareadlist.append(precursorchargelist)
fareadlist.append(productnamelist)
fareadlist.append(productformulalist)
fareadlist.append(productadductlist)
fareadlist.append(productmzlist)
fareadlist.append(productchargelist)
# end calculate monounsaturated precursors from input
preconly=1
#minrt=0.0
#print(fareadlist[1])
#maxrt=eval(input('At which retention time [min] ends gradient? :'))
#explrt=(minrt+maxrt)/2
#explrtw=explrt-minrt
nspec=3	# number of species: precursor, aldehyde, criegee
# begin create empty lists
toprow=[]
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]		# double bond position list, contains strings with double bond position n-1, n-2 ...
dbindexlist=[]	# double bond index list, contains index for the double bond position that is closest to head group of FA
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
toprow=['Moleculegroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
ki=len(fareadlist[0])
satlist=fareadlist	#create lists for saturated FAs
#print(ki)
r=0 #2
#ki=ki+2
while r<ki:		#go through rows of excel file jpmlipidozidinput
	# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
	cchain=0
	e=str(fareadlist[1][r]) ## Precursorname
	i=len(e)-3
	add=int(e[i])
	cchain=cchain+add
	i=i-1
	add=int(e[i])
	if add>0:
		cchain=cchain+(10*add)
	else:
		cchain=cchain
		#print('Please check source code (determine number of C atoms in chain)')
	maxdbp=cchain-2
	# begin determine number of C atoms in chain, define highest possible double bond position
	kadd=nspec*maxdbp#+1		#(nspec=3 precursor and 2 products - aldehyde and crigee  - for each of maxdbp possible double bond positions)

	e=fareadlist[0][r] ## MoleculeGroup
	f=fareadlist[1][r] ## Precursorname
	g=fareadlist[2][r]  ##Precursorformula
	h=float(fareadlist[5][r])  ## PrecursorCharge
	k=0
	while k<kadd:
		mlistname.append(e)		# copied, no change
		precname.append(f)		# copied, no change
		precformula.append(g)	# copied, no change
		precchrg.append(h)		# copied, no change
		k=k+1

	e=fareadlist[3][r] ## Precursoradduct
	k=0
	while k<kadd:
		precadduct.append(e) 	# AMPP, precursor
		precadduct.append(e)	# AMPP, aldehyde product
		precadduct.append(e)	# AMPP, crigee product
		k=k+nspec

	e=float(fareadlist[4][r])  ## PrecursorMz
	k=0
	while k<kadd:
		precmz.append(e)	# precursor
		precmz.append(e)	# aldehyde
		precmz.append(e)	# crigee
		k=k+nspec

	e=fareadlist[1][r] ## Productname
	if e=='Chol':
		fragment='_'	
	else:	
		k=0
		csub=1
		while k<kadd:
			dbp='_n-'+str(csub)
			fragment='_precursor'
			ne=e+dbp+fragment
			prodname.append(ne)		# precursor
			dbl=[]	#begin save double bond position for later
			dbl.append(csub)
			dbindexlist.append(dbl)	#end save double bond position for later
			dbp='_n-'+str(csub)
			fragment='_aldehyde neutral loss'
			ne=e+dbp+fragment
			dbpi=0
			while dbpi<nspec:
				dbplist.append(dbp)
				dbpi=dbpi+1
			#
			if preconly==0:	
				prodname.append(ne)		# aldehyde
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				#prodname.append(ne)
				#prodname.append(ne)
				fragment='_criegee neutral loss'
				ne=e+dbp+fragment
				prodname.append(ne)		# crigee
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				#prodname.append(ne)
				#prodname.append(ne)
			if preconly==1:
				dbp='_n-'+str(csub)
				fragment='_dummy precursor a'
				ne=e+dbp+fragment
				prodname.append(ne)		# dummy precursor
				fragment='_dummy precursor b'
				ne=e+dbp+fragment
				prodname.append(ne)		# dummy precursor
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)
				dbindexlist.append(dbl)	#end save double bond position for later
			csub=csub+1
			k=k+nspec

	e=fareadlist[2][r] #
	# begin read precursor sum formula and edit product sum formula
	#print(e)
	#print(e[0])
	#print(e[1])
	#print(e[2])
	#print(e[3])
	#print(len(e))
	clist=[]
	hlist=[]
	dlist=[]
	nlist=[]
	olist=[]
	plist=[]
	ilist=[]
	i=0
	ca=0
	ha=0
	da=0
	na=0
	oa=0
	pa=0
	ia=0
	while i<len(e):
		if e[i]=='H':
			if e[i+1]=="'":
				ha=0
			else:
				ca=0
		#if e[i]=="'":
		#	ha=0		
		if e[i]=='N':
			ha=0
			da=0
		if e[i]=='O':
			ha=0
			da=0
			na=0
		if e[i]=='P':
			ha=0
			da=0
			na=0
			oa=0
		if e[i]=='I':
			ha=0
			da=0
			na=0
			oa=0
			pa=0
		if ca==1:
			clist.append(e[i])
		if ha==1:
			hlist.append(e[i])
		if da==1:
			dlist.append(e[i])
		if na==1:
			nlist.append(e[i])
		if oa==1:
			olist.append(e[i])
		if pa==1:
			plist.append(e[i])
		if ia==1:
			ilist.append(e[i])
		if e[i]=='C':
			ca=1
		if e[i]=='H':
			if e[i+1]=="'":
				ca=0
				ha=0
				da=1
				i=i+1
			else:
				ca=0
				ha=1
		#if e[i]=="'":
		#	ca=0
		#	ha=0
		#	da=1		
		if e[i]=='N':
			ha=0
			da=0
			na=1
			if e[i+1]=='O':
				nlist.append('1')
				na=0
		if e[i]=='O':
			ha=0
			da=0
			na=0
			oa=1
			if (i+1)<len(e):
				if e[i+1]=='P':
					olist.append('1')
					oa=0
			else:
				olist.append('1')
				oa=0					
		if e[i]=='P':
			da=0
			na=0
			oa=0
			pa=1
			if (i+1)<len(e):
				if e[i+1]=='I':
					plist.append('1')
					pa=0
			else:
				plist.append('1')
				pa=0
		if e[i]=='I':
			da=0
			na=0
			oa=0
			pa=0
			ia=1
			if i==(len(e)-1):
				ilist.append('1')
				ia=0
		i=i+1
	#print(clist)
	#print(hlist)
	#print(dlist)
	#print(nlist)
	#print(olist)
	#print(plist)
	if len(clist)==0:
		cn=0
	if len(hlist)==0:
		hn=0
	if len(dlist)==0:
		dn=0	
	if len(nlist)==0:
		nn=0
	if len(olist)==0:
		on=0
	if len(plist)==0:
		pn=0
	if len(ilist)==0:
		iodon=0
	if len(clist)==1:
		cn=int(clist[0])
	if len(clist)==2:
		cn=10*int(clist[0])+int(clist[1])
	if len(clist)==3:
		cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
	if len(hlist)==1:
		hn=int(hlist[0])
	if len(hlist)==2:
		hn=10*int(hlist[0])+int(hlist[1])
	if len(hlist)==3:
		hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
	if len(dlist)==1:
		dn=int(dlist[0])
	if len(dlist)==2:
		dn=10*int(dlist[0])+int(dlist[1])
	if len(dlist)==3:
		dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
	if len(nlist)==1:
		nn=int(nlist[0])
	if len(nlist)==2:
		nn=10*int(nlist[0])+int(nlist[1])
	if len(nlist)==3:
		nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
	if len(olist)==1:
		on=int(olist[0])
	if len(olist)==2:
		on=10*int(olist[0])+int(olist[1])
	if len(olist)==3:
		on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
	if len(plist)==1:
		pn=int(plist[0])
	if len(plist)==2:
		pn=10*int(plist[0])+int(plist[1])
	if len(plist)==3:
		pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])	
	if len(ilist)==1:
		iodon=int(ilist[0])
	if len(ilist)==2:
		iodon=10*int(ilist[0])+int(ilist[1])
	if len(ilist)==3:
		iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		
	# end read precursor sum formula
	#print(cn)
	e=fareadlist[0][r] ## begin calculate product sum formula
	if e=='SPLASH':
		e=e 		# begin calculate product sum formula for SPLASH
	else:
		k=0
		csub=1
		while k<kadd:
			cnp=cn-(csub)
			hnp=hn-(2*csub)
			onald=on+1
			oncrigee=on+2
			precursor='C'+str(cn)+'H'+str(hn)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
			productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
			productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
			prodformula.append(precursor)
			if preconly==0:
				prodformula.append(productaldehyde)
				#prodformula.append(productaldehyde)
				#prodformula.append(productaldehyde)
				prodformula.append(productcrigee)
				#prodformula.append(productcrigee)
				#prodformula.append(productcrigee)	# product formula is saved in list for current double bond position
			if preconly==1:
				prodformula.append(precursor)
				prodformula.append(precursor)
			precursormz=imass[0]*(hn)+imass[1]*(dn)+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
			prodmz.append(precursormz)
			if preconly==0:
				productmz=imass[0]*(hnp+0)+imass[1]*(dn)+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, AMPP
				prodmz.append(productmz)
				productmz=imass[0]*(hnp+0)+imass[1]*(dn)+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, AMPP
				prodmz.append(productmz)
			if preconly==1:
				prodmz.append(precursormz)
				prodmz.append(precursormz)
			csub=csub+1
			k=k+nspec			
	# end read precursor sum formula and edit product sum formula

	prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
	prodchrg=precchrg 	# ProductCharge

	r=r+1
	e=fareadlist[1][r-1] 	# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	if r>(len(fareadlist[0])-1):
		r=r+1
	else:
		e=fareadlist[1][r] 
		prec=e
		while prec==prevprec:
			r=r+1
			e=fareadlist[1][r-1] 
			prevprec=e
			e=fareadlist[1][r] 
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	
#end save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
writelist=[]
writelist.append(mlistname)
writelist.append(precname)
writelist.append(precformula)
writelist.append(precadduct)
writelist.append(precmz)
writelist.append(precchrg)
writelist.append(prodname)
writelist.append(prodformula)
writelist.append(prodadduct)
writelist.append(prodmz)
writelist.append(prodchrg)
#end save data in writelist

#print('All calculations for monounsaturated fatty acids are done.')
#end save excel file
#print ('odd : C_%d H_%d N_%d O_%d S_%d; DBE = %d; deviation: %.3f' % (formula[1], formula[0], formula[2], formula[3], formula[4], dbe, meandeviation))
###########################################################################################################################################################################
#############################################################################MONO##########################################################################################
#print('one done')
############################################################################DOUBLE#########################################################################################
###########################################################################################################################################################################
# begin add double unsaturated fatty acids
#ask=eval(input('Add fatty acids with two double bonds? Yes: 1; No: 0 | '))
ask=1
if ask==0:
	quit()

nspec=5	# number of species: precursor, aldehyde db1 and db2, crigee db1 and db2
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])

kmono=ki # List index in excel file to start writing FA transitions with two double bonds
kdoublestart=kmono
r=0
ki=ki
while r<ki:		#go through rows of excel file jpmlipidozidinput
	# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
	cchain=0
	e=str(writelist[1][r]) 	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	if len(e)==13:
		dbpmono=int(e[12])
	elif len(e)==14:
		dbpmono=(10*(int(e[12])))+(int(e[13]))
	else:
		print('Please check source code (determine double bond position of already located double bond)')
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the second double bond
	csubmono=dbpmono+2
	if nsecdbp>0:

		kadd=nspec*nsecdbp#+1		#(nspec=3 products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			k=k+1

		e=writelist[1][r] ## PrecursorName
		degunsat=int(e[8])		# change from monounsaturated to bisunsaturated FA in PrecursorName
		if degunsat==1:
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(2)
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# updated number of double bonds (bisunsaturated instead of monounsaturated)
			k=k+1

		e=writelist[3][r] ## Precursoradduct
		k=0
		while k<kadd:
			precadduct.append(e) 	# AMPP, precursor
			precadduct.append(e)	# AMPP, aldehyde product
			precadduct.append(e)	# AMPP, crigee product
			precadduct.append(e)	# AMPP, aldehyde product, outer db
			precadduct.append(e)	# AMPP, crigee product, outer db
			k=k+nspec

		e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added second double bond
		k=0
		while k<kadd:
			precmz.append(e)	# precursor
			precmz.append(e)	# aldehyde
			precmz.append(e)	# crigee
			precmz.append(e)	# aldehyde outer db
			precmz.append(e)	# crigee outer db
			k=k+nspec

		e=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			precchrg.append(e)
			k=k+1

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change from monounsaturated to bisunsaturated FA in PrecursorName
		if degunsat==1:
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(2)
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from monounsaturated to bisunsaturated FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond position for later
				firstdbp=dbindexlist[r][0]
				dbl.append(firstdbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					prodname.append(ne)		# aldehyde
					dbl=[]	#begin save double bond position for later
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	#end save double bond position for later
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde outer double bond
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee outer double bond
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor a'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor b'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor c'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor d'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #sheetinput.cell(row=r, column=3)
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=="'":
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=="'":
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with two double bonds
				hnpmono=hnprec-(2*dbpmono)	# applies to cleavage of outer double bond
				cnpmono=cn-(dbpmono)		# applies to cleavage of outer double bond
				cnp=cn-(csub)				# applies to cleavage of inner double bond
				hnp=hn-(2*csub)				# applies to cleavage of inner double bond
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydemono='C'+str(cnpmono)+'H'+str(hnpmono)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeemono='C'+str(cnpmono)+'H'+str(hnpmono)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydemono)
					prodformula.append(productcrigeemono)	
				if preconly==1:
					prodformula.append(precursor)	
					prodformula.append(precursor)
					prodformula.append(precursor)
					prodformula.append(precursor)
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, inner db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, inner db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpmono)+imass[1]*dn+imass[2]*cnpmono+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, inner db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpmono)+imass[1]*dn+imass[2]*cnpmono+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, inner db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data as writelist, merge new lists at end of each list in writelist
writelist[0]=writelist[0]+mlistname
writelist[1]=writelist[1]+precname
writelist[2]=writelist[2]+precformula
writelist[3]=writelist[3]+precadduct
writelist[4]=writelist[4]+precmz
writelist[5]=writelist[5]+precchrg
writelist[6]=writelist[6]+prodname
writelist[7]=writelist[7]+prodformula
writelist[8]=writelist[8]+prodadduct
writelist[9]=writelist[9]+prodmz
writelist[10]=writelist[10]+prodchrg
#end save data as writelist

#print('All calculations for fatty acids with two double bonds are done.')
#end save excel file
# end add double unsaturated fatty acids
###########################################################################################################################################################################
###########################################################################DOUBLE##########################################################################################
###########################################################################TRIPLE##########################################################################################
###########################################################################################################################################################################
# begin add triple unsaturated fatty acids
#print(len(dbindexlist))
#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
ask=1
if ask==0:
	quit()

nspec=7	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])

kmono=ki # List index in writelist to start writing FA transitions with three double bonds
ktriplestart=kmono
r=kdoublestart
ki=ki
while r<ki:		#go through rows of FAs with two double bonds in writelist
	# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
	cchain=0
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	#maxdbp=
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	#if len(e)==14:
	#	dbpmono=int(e[13])
	#elif len(e)==15:
	#	dbpmono=(10*(int(e[13])))+(int(e[14]))
	#else:
	#	print('Please check source code (determine double bond position of already located double bond)')
	dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond	
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
	csubmono=dbpmono+2
	if nsecdbp>0:

		kadd=nspec*nsecdbp#+1		#(nspec=3 products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		f=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			precchrg.append(f)	# copied, no change
			k=k+1	

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
		degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==2:
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(3)
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# write new precursorname, changed to FA with three db
			k=k+1

		e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precadduct.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added second double bond
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precmz.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==2:
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(3)
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from bisunsaturated to trisunsaturated FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond position for later 						
				firstdbp=dbindexlist[r][0]
				seconddbp=dbindexlist[r][1] ########################################## r instead of r-2 ##################
				dbl.append(firstdbp)
				dbl.append(seconddbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde last existing double bond
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee last existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor1'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor2'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor3'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor4'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor5'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor6'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]	########################################## r instead of r-2 ##################
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #PrecursorFormula
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=="'":
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=="'":
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])	
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
				cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
				hnp=hnprec-(2*csub-2*2)		# applies to cleavage of first (highest n) double bond
				hnptwo=hnprec-(2*dbpmono-2*1)	# applies to cleavage of second double bond
				cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
				hnpthree=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of third double bond
				cnpthree=cn-(dbindexlist[r][0])			# applies to cleavage of third double bond
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydetwo)
					prodformula.append(productcrigeetwo)
					prodformula.append(productaldehydethree)
					prodformula.append(productcrigeethree)	
				if preconly==1:
					prodformula.append(precursor)	
					prodformula.append(precursor)
					prodformula.append(precursor)
					prodformula.append(precursor)
					prodformula.append(precursor)
					prodformula.append(precursor)
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
writelist[0]=writelist[0]+mlistname
writelist[1]=writelist[1]+precname
writelist[2]=writelist[2]+precformula
writelist[3]=writelist[3]+precadduct
writelist[4]=writelist[4]+precmz
writelist[5]=writelist[5]+precchrg
writelist[6]=writelist[6]+prodname
writelist[7]=writelist[7]+prodformula
writelist[8]=writelist[8]+prodadduct
writelist[9]=writelist[9]+prodmz
writelist[10]=writelist[10]+prodchrg
#end save data in writelist

#print('All calculations for fatty acids with three double bonds are done.')
#end save excel file
# end add triple unsaturated fatty acids
######################################################################################################################################################
##############################################################TRIPLE##################################################################################
##############################################################FOUR DB#################################################################################
######################################################################################################################################################
# begin add triple unsaturated fatty acids
#print(len(dbindexlist))
#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
ask=1
if ask==0:
	quit()

nspec=9	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3  ############################ VARY ########################
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])

#print('Entries in excel file after bisunsaturated FA transitions are generated: %d' % ki)
kmono=ki # List index in excel file to start writing FA transitions with four double bonds
kfourstart=kmono
r=ktriplestart ################################################################################################### VARY ########################
ki=ki#+2
while r<ki:		#go through rows of FAs with three double bonds in excel file
	# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
	cchain=0
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	#maxdbp=
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	#if len(e)==14:
	#	dbpmono=int(e[13])
	#elif len(e)==15:
	#	dbpmono=(10*(int(e[13])))+(int(e[14]))
	#else:
	#	print('Please check source code (determine double bond position of already located double bond)')
	dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond        ######################################### CHECK ?
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
	csubmono=dbpmono+2
	if nsecdbp>0:
		kadd=nspec*nsecdbp#+1		#(nspec= number of products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		f=float(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			precchrg.append(f)	# copied, no change
			k=k+1	

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
		degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==3:			####################################### VARY ################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(4)		####################################### VARY ################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# write new precursorname, changed to FA with four db
			k=k+1

		e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precadduct.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=float(writelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added double bond
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precmz.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change FA in PrecursorName to generate ProductName
		if degunsat==3:			########################################## VARY ######################################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(4)	########################################## VARY ######################################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from degree of unsaturation of FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond position for later 						
				firstdbp=dbindexlist[r][0]
				seconddbp=dbindexlist[r][1]
				thirddbp=dbindexlist[r][2]
				dbl.append(firstdbp)
				dbl.append(seconddbp)
				dbl.append(thirddbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde last existing double bond
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee last existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][1])
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][1])
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor1'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor2'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor3'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor4'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor5'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor6'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor7'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor8'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #sheetinput.cell(row=r, column=3)
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=="'":
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=="'":
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
				cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
				hnp=hnprec-(2*csub-2*3)		# applies to cleavage of first (highest n) double bond
				hnptwo=hnprec-(2*dbpmono-2*2)	# applies to cleavage of second double bond
				cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
				hnpthree=hnprec-(2*(dbindexlist[r][1])-2*1)	# applies to cleavage of third double bond
				cnpthree=cn-(dbindexlist[r][1])			# applies to cleavage of third double bond
				hnpfour=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				cnpfour=cn-(dbindexlist[r][0])			# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydetwo)
					prodformula.append(productcrigeetwo)
					prodformula.append(productaldehydethree)
					prodformula.append(productcrigeethree)
					prodformula.append(productaldehydefour)
					prodformula.append(productcrigeefour)	
				if preconly==1:
					t=1
					while t<nspec:
						prodformula.append(precursor)
						t=t+1
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fourth db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
writelist[0]=writelist[0]+mlistname
writelist[1]=writelist[1]+precname
writelist[2]=writelist[2]+precformula
writelist[3]=writelist[3]+precadduct
writelist[4]=writelist[4]+precmz
writelist[5]=writelist[5]+precchrg
writelist[6]=writelist[6]+prodname
writelist[7]=writelist[7]+prodformula
writelist[8]=writelist[8]+prodadduct
writelist[9]=writelist[9]+prodmz
writelist[10]=writelist[10]+prodchrg
#end save data in writelist

#print('All calculations for fatty acids with four double bonds are done.')
#end save excel file
# end add triple unsaturated fatty acids
######################################################################################################################################################
##############################################################FOUR DB#################################################################################
##############################################################FIVE DB#################################################################################
######################################################################################################################################################
# begin add triple unsaturated fatty acids
#print(len(dbindexlist))
#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
ask=1
if ask==0:
	quit()

nspec=11	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3  ############################ VARY ########################
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])

#print('Entries in excel file after bisunsaturated FA transitions are generated: %d' % ki)
kmono=ki # List index in excel file to start writing FA transitions with four double bonds
kfivestart=kmono
r=kfourstart ################################################################################################### VARY ########################
ki=ki#+2
while r<ki:		#go through rows of FAs with three double bonds in excel file
	# begin determine number of C atoms in chain, define highest possible double bond position ###############
	cchain=0
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	#maxdbp=
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	#if len(e)==14:
	#	dbpmono=int(e[13])
	#elif len(e)==15:
	#	dbpmono=(10*(int(e[13])))+(int(e[14]))
	#else:
	#	print('Please check source code (determine double bond position of already located double bond)')
	dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond        ######################################### CHECK ?
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
	csubmono=dbpmono+2
	if nsecdbp>0:
		kadd=nspec*nsecdbp#+1		#(nspec= number of products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		f=float(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			precchrg.append(f)	# copied, no change
			k=k+1	

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
		degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==4:			####################################### VARY ################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(5)		####################################### VARY ################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# write new precursorname, changed to FA with four db
			k=k+1

		e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precadduct.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=float(writelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added double bond
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precmz.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change FA in PrecursorName to generate ProductName
		if degunsat==4:			########################################## VARY ######################################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(5)	########################################## VARY ######################################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from degree of unsaturation of FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond positions for later 	###################### CHECK !						
				firstdbp=dbindexlist[r][0]
				seconddbp=dbindexlist[r][1]
				thirddbp=dbindexlist[r][2]
				fourthdbp=dbindexlist[r][3]
				dbl.append(firstdbp)
				dbl.append(seconddbp)
				dbl.append(thirddbp)
				dbl.append(fourthdbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond positions for later
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde last existing double bond
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee last existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][2])
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][2])
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor1'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor2'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor3'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor4'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor5'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor6'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor7'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor8'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor9'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor10'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #sheetinput.cell(row=r, column=3)
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=="'":
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=="'":
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
				cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
				hnp=hnprec-(2*csub-2*4)		# applies to cleavage of first (highest n) double bond
				hnptwo=hnprec-(2*dbpmono-2*3)	# applies to cleavage of second double bond
				cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
				hnpthree=hnprec-(2*(dbindexlist[r][2])-2*2)	# applies to cleavage of third double bond
				cnpthree=cn-(dbindexlist[r][2])			# applies to cleavage of third double bond
				hnpfour=hnprec-(2*(dbindexlist[r][1])-2*1)	# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				cnpfour=cn-(dbindexlist[r][1])			# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				hnpfive=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				cnpfive=cn-(dbindexlist[r][0])			# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				productcrigeefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydetwo)
					prodformula.append(productcrigeetwo)
					prodformula.append(productaldehydethree)
					prodformula.append(productcrigeethree)
					prodformula.append(productaldehydefour)
					prodformula.append(productcrigeefour)	
					prodformula.append(productaldehydefive) ############################################# CHECK !!!!!!!!
					prodformula.append(productcrigeefive) ############################################# CHECK !!!!!!!!
				if preconly==1:
					t=1
					while t<nspec:
						prodformula.append(precursor)
						t=t+1
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fifth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fifth db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
if workflowmode==1:
	writelist[0]=writelist[0]+mlistname
	writelist[1]=writelist[1]+precname
	writelist[2]=writelist[2]+precformula
	writelist[3]=writelist[3]+precadduct
	writelist[4]=writelist[4]+precmz
	writelist[5]=writelist[5]+precchrg
	writelist[6]=writelist[6]+prodname
	writelist[7]=writelist[7]+prodformula
	writelist[8]=writelist[8]+prodadduct
	writelist[9]=writelist[9]+prodmz
	writelist[10]=writelist[10]+prodchrg
#end save data in writelist
#if workflow==1:
	#print('All calculations for fatty acids with five double bonds are done.')
#end save excel file
# end add triple unsaturated fatty acids
######################################################################################################################################################
##############################################################FIVE DB#################################################################################

##############################################################SIX DB#################################################################################
######################################################################################################################################################
# begin add triple unsaturated fatty acids
#print(len(dbindexlist))
#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
ask=0
if ask==1:
	quit()

nspec=13	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3  ############################ VARY ########################
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])

#print('Entries in excel file after bisunsaturated FA transitions are generated: %d' % ki)
kmono=ki # List index in excel file to start writing FA transitions with four double bonds
ksixstart=kmono
r=kfivestart ################################################################################################### VARY ########################
ki=ki#+2
while r<ki:		#go through rows of FAs with three double bonds in excel file
	# begin determine number of C atoms in chain, define highest possible double bond position ###############
	cchain=0
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	#maxdbp=
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	#if len(e)==14:
	#	dbpmono=int(e[13])
	#elif len(e)==15:
	#	dbpmono=(10*(int(e[13])))+(int(e[14]))
	#else:
	#	print('Please check source code (determine double bond position of already located double bond)')
	dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond, highest n        ######################################### CHECK ?
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
	csubmono=dbpmono+2
	if nsecdbp>0:
		kadd=nspec*nsecdbp#+1		#(nspec= number of products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		f=float(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			precchrg.append(f)	# copied, no change
			k=k+1	

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
		degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==5:			####################################### VARY ################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(6)		####################################### VARY ################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# write new precursorname, changed to FA with four db
			k=k+1

		e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precadduct.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=float(writelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added double bond
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precmz.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change FA in PrecursorName to generate ProductName
		if degunsat==5:			########################################## VARY ######################################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(6)	########################################## VARY ######################################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from degree of unsaturation of FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond positions for later 	###################### CHECK !						
				firstdbp=dbindexlist[r][0]
				seconddbp=dbindexlist[r][1]
				thirddbp=dbindexlist[r][2]
				fourthdbp=dbindexlist[r][3]
				fifthdbp=dbindexlist[r][4]
				dbl.append(firstdbp)
				dbl.append(seconddbp)
				dbl.append(thirddbp)
				dbl.append(fourthdbp)
				dbl.append(fifthdbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond positions for later
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde last existing double bond
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee last existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][3])
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][3])
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][2])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][2])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					fifthdbp=dbindexlist[r][4]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(fifthdbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor1'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor2'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor3'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor4'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor5'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor6'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor7'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor8'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor9'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor10'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor11'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor12'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					fifthdbp=dbindexlist[r][4]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(fifthdbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #sheetinput.cell(row=r, column=3)
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=="'":
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=="'":
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
				cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
				hnp=hnprec-(2*csub-2*5)		# applies to cleavage of first (highest n) double bond
				hnptwo=hnprec-(2*dbpmono-2*4)	# applies to cleavage of second double bond
				cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
				hnpthree=hnprec-(2*(dbindexlist[r][3])-2*3)	# applies to cleavage of third double bond
				cnpthree=cn-(dbindexlist[r][3])			# applies to cleavage of third double bond
				hnpfour=hnprec-(2*(dbindexlist[r][2])-2*2)	# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				cnpfour=cn-(dbindexlist[r][2])			# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				hnpfive=hnprec-(2*(dbindexlist[r][1])-2*1)	# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				cnpfive=cn-(dbindexlist[r][1])			# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				hnpsix=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				cnpsix=cn-(dbindexlist[r][0])			# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				productcrigeefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				productaldehydesix='C'+str(cnpsix)+'H'+str(hnpsix)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				productcrigeesix='C'+str(cnpsix)+'H'+str(hnpsix)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydetwo)
					prodformula.append(productcrigeetwo)
					prodformula.append(productaldehydethree)
					prodformula.append(productcrigeethree)
					prodformula.append(productaldehydefour)
					prodformula.append(productcrigeefour)	
					prodformula.append(productaldehydefive) ############################################# CHECK !!!!!!!!
					prodformula.append(productcrigeefive) ############################################# CHECK !!!!!!!!
					prodformula.append(productaldehydesix) ############################################# CHECK !!!!!!!!
					prodformula.append(productcrigeesix) ############################################# CHECK !!!!!!!!
				if preconly==1:
					t=1
					while t<nspec:
						prodformula.append(precursor)
						t=t+1
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fifth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fifth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpsix)+imass[1]*dn+imass[2]*cnpsix+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, sixth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpsix)+imass[1]*dn+imass[2]*cnpsix+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, sixth db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
if workflowmode==1:
	writelist[0]=writelist[0]+mlistname
	writelist[1]=writelist[1]+precname
	writelist[2]=writelist[2]+precformula
	writelist[3]=writelist[3]+precadduct
	writelist[4]=writelist[4]+precmz
	writelist[5]=writelist[5]+precchrg
	writelist[6]=writelist[6]+prodname
	writelist[7]=writelist[7]+prodformula
	writelist[8]=writelist[8]+prodadduct
	writelist[9]=writelist[9]+prodmz
	writelist[10]=writelist[10]+prodchrg
#end save data in writelist
#if workflow==1:
	#print('All calculations for fatty acids with six double bonds are done.')
#end save excel file
# end add triple unsaturated fatty acids
#print(len(writelist[0]))
#print(len(dbindexlist))
######################################################################################################################################################
##############################################################END ADD SIX DB##########################################################################
######################################################################################################################################################
#print('begin add saturated FAs')
r=0
while r<(len(satlist[1])):
	#satlist[1][r][len(satlist[1][r])-1]='0'	
	precn=''
	p=0
	while p<(len(satlist[1][r])-1):
		precn=precn+satlist[1][r][p]
		p=p+1
	precn=precn+'0'					#PrecursorName to saturated FA
	satlist[1][r]=precn
	satlist[4][r]=float(satlist[4][r])+(2*imass[0])	#PrecursorMz to sat. FA
	satlist[9][r]=float(satlist[9][r])+(2*imass[0])	#ProductMz to sat. FA
	hcurr=(10*(int(satlist[2][r][4])))+(int(satlist[2][r][5]))
	hcurr=str(hcurr+2)
	p=0
	precf=''
	while p<4:
		precf=precf+satlist[2][r][p]
		p=p+1
	precf=precf+hcurr
	p=p+2
	while p<(len(satlist[2][r])):
		precf=precf+satlist[2][r][p]
		p=p+1
	satlist[2][r]=str(precf)	#PrecursorFormula is edited
	satlist[7][r]=str(precf)	#ProductFormula is edited
	r=r+1
pnm=0
while pnm<len(satlist[0]):
	satlist[6][pnm]=str(satlist[1][pnm])+'_precursor'
	pnm=pnm+1
clmn=0
while clmn<(len(satlist)):
	writelist[clmn]=writelist[clmn]+satlist[clmn]
	clmn=clmn+1
# end add saturated FAs

##########################################################VIRTUAL#PRECURSOR###########################################################################
######################################################################################################################################################
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]		# content in lists is deleted to free up memory
# begin introduce virtual precursor
ki=len(writelist[0])

#print('Entries in excel file after all FA transitions are generated: %d' % ki)
virtualprecformula=[]
virtualprecmz=[]

vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]

r=0
ki=ki
while r<ki:
	e=str(writelist[2][r])#sheetinput.cell(row=r, column=3)	# PrecursorFormula
	ve=e+'Xe'
	virtualprecformula.append(ve) # virtual precursorFormula
	e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
	ve=e+(imass[9])		#add one Xe atom to generate virtual precursor
	virtualprecmz.append(ve)	# virtual precursormz
	e=str(writelist[0][r])#sheetinput.cell(row=r, column=1)
	vmlistname.append(e)
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)
	vprecname.append(e)
	e=str(writelist[3][r]) #sheetinput.cell(row=r, column=4)
	vprecadduct.append(e)
	e=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)
	vprecchrg.append(e)
	e=str(writelist[6][r]) #sheetinput.cell(row=r, column=7)
	vprodname.append(e)
	e=str(writelist[7][r]) #sheetinput.cell(row=r, column=8)
	vprodformula.append(e)
	e=str(writelist[8][r]) #sheetinput.cell(row=r, column=9)
	vprodadduct.append(e)
	e=float(writelist[9][r]) #sheetinput.cell(row=r, column=10)
	vprodmz.append(e)
	e=int(writelist[10][r]) #sheetinput.cell(row=r, column=11)
	vprodchrg.append(e)
	r=r+1
# begin save excel file with virtual precursor as csv file
toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
vwritelist=[]
vwritelist.append(vmlistname)
vwritelist.append(vprecname)
vwritelist.append(virtualprecformula)
vwritelist.append(vprecadduct)
vwritelist.append(virtualprecmz)
vwritelist.append(vprecchrg)
vwritelist.append(vprodname)
vwritelist.append(vprodformula)
vwritelist.append(vprodadduct)
vwritelist.append(vprodmz)
vwritelist.append(vprodchrg)
terminate=0
if terminate==1:
	print('writelist created')
	transitionresultsdf=pd.DataFrame(vwritelist).transpose()
	print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10]]
	print('Transposed and DataFrame created')
	transitionresultsdf.to_csv('jpmlipidomics_vpw20_virtual_precursor.csv', index=False)
	print('Transition list saved as jpmlipidomics_vpw20_virtual_precursor.csv')
# end save excel file with virtual precursor as csv file

#print('Saving data.')
#print('len(vmlistname)')
#print(len(vmlistname))
virtualprecformula=[]
virtualprecmz=[]
vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[] # test to delete lists to save up memory

#print('Transition list is modified with virtual precursor [M + Xe].')

#terminate=1
if terminate==1:
	quit()

#print('reduce and expand')
#################################################################################################################################################
############################# BEGIN REDUCE TO ONE ENTRY PER PRECURSOR AND EXPAND WITH VARIED EXPLICIT RETENTION TIME ############################
#################################################################################################################################################
#terminate=eval(input('Reduce transition list to one entry per precursor and expand with varied explicit retention time? Yes: 1 No: 0 ::'))
terminate=1
if terminate==0:
	quit()
else:
	#exrtstep=eval(input('Stepwidth for varied explicit retention time (e.g. 0.05 min)? ::'))
	#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
	normalexrtstep=0.027
	bigexrtstep=0.045
	#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
# begin reduce and expand
ki=len(vwritelist[0])
# end count entries in excel file
vmlistname=[]
vprecname=[]
virtualprecformula=[]
vprecadduct=[]
virtualprecmz=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]
precrt=[]
rtwindow=[]
r=0
ki=ki
pos=1
while r<ki:
	if pos==1:
		if rtlimitation==1:
			# begin redefine firstexrt as cfirstexrt and lastexrt as clastexrt depending on chainlength of current FA
			e=(10*int(vwritelist[1][r][5]))+int(vwritelist[1][r][6])
			e=int(e)
			if e>15:
				if e<30:
					cfirstexrt=firstexrt+(((lastexrt-firstexrt)*(1/4))*(((e-15)/15)))
				else:
					cfirstexrt=firstexrt+(((lastexrt-firstexrt)*(1/4))+((lastexrt-firstexrt)*(1/4))*(((e-30)/10)))
				cfirstexrt=round(cfirstexrt, 3)
			else:
				cfirstexrt=firstexrt
			if e<20:
				if e<8:
					clastexrt=lastexrt-(((lastexrt-firstexrt)*(1/8))+((lastexrt-firstexrt)*(1/8))*(((8-e)/4)))
				else:
					clastexrt=lastexrt-(((lastexrt-firstexrt)*(1/8))*(((20-e)/12)))
				clastexrt=round(clastexrt, 3)
			else:
				clastexrt=lastexrt
			# end redefine firstexrt as cfirstexrt and lastexrt as clastexrt depending on chainlength of current FA
		elif rtlimitation==2:
			# begin set cfirstexrt according to predicted RT range start (predictedrtmin) and clastexrt / predictedrtmax // predictedrtindex
			cfaspec=str(vwritelist[1][r][5])+str(vwritelist[1][r][6])+str(vwritelist[1][r][7])+str(vwritelist[1][r][8])
			pr=0
			while pr<(len(predictedrtindex)):
				if cfaspec==str(predictedrtindex[pr]):
					cfirstexrt=predictedrtmin[pr]
					clastexrt=predictedrtmax[pr]
				pr=pr+1
			# end set cfirstexrt according to predicted RT range start (predictedrtmin) and clastexrt / predictedrtmax // predictedrtindex
		else:
			cfirstexrt=firstexrt #float(firstexrt)
			clastexrt=lastexrt #float(lastexrt)
			clastexrt=round(clastexrt, 3)
			cfirstexrt=round(cfirstexrt, 3)
		#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
		if int(vwritelist[1][r][8])>2:
			nca=10*int(vwritelist[1][r][5])+int(vwritelist[1][r][6])
			if nca>17:
				exrt=cfirstexrt+bigexrtstep
			else:
				exrt=cfirstexrt+normalexrtstep
		else:	
			exrt=cfirstexrt+normalexrtstep		# firstexrt is the first explicit retention time that is set to look for species (e.g. 1.50 min)
		#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
		exrtstr=str(exrt)
		while exrt<clastexrt:		# lastexrt is the last explicit retention time that is set to look for species (e.g. 18.00 min)
			ve=str(vwritelist[2][r]) #sheetinput.cell(row=r, column=3)	# PrecursorFormula
			virtualprecformula.append(ve) # virtual precursorFormula
			ve=float(vwritelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz	
			virtualprecmz.append(ve)	# virtual precursormz
			e=str(vwritelist[0][r]) #sheetinput.cell(row=r, column=1)
			vmlistname.append(e)
			e=str(vwritelist[1][r]) #sheetinput.cell(row=r, column=2)
			cm='_'
			cm=str(cm)
			exrtstr=str(round(exrt, 2))		#
			ee=e+cm+exrtstr			#
			vprecname.append(ee)
			precrt.append(exrt)
			e=str(vwritelist[3][r]) #sheetinput.cell(row=r, column=4)
			vprecadduct.append(e)
			e=int(vwritelist[5][r]) #sheetinput.cell(row=r, column=6)
			vprecchrg.append(e)
			e=str(vwritelist[6][r]) #sheetinput.cell(row=r, column=7)
			vprodname.append(e)
			e=str(vwritelist[7][r]) #sheetinput.cell(row=r, column=8)
			vprodformula.append(e)
			e=str(vwritelist[8][r]) #sheetinput.cell(row=r, column=9)
			vprodadduct.append(e)
			e=float(vwritelist[9][r]) #sheetinput.cell(row=r, column=10)
			vprodmz.append(e)
			e=int(vwritelist[10][r]) #sheetinput.cell(row=r, column=11)
			vprodchrg.append(e)	
			#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
			if int(vwritelist[1][r][8])>2:
				nca=10*int(vwritelist[1][r][5])+int(vwritelist[1][r][6])
				if nca>17:
					exrt=exrt+bigexrtstep
					exrtstep=bigexrtstep
				else:
					exrt=exrt+normalexrtstep
					exrtstep=normalexrtstep	
			else:	
				exrt=exrt+normalexrtstep
				exrtstep=normalexrtstep
			#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain	
			e=exrtstep*2
			rtwindow.append(e)
			#print(exrt)
			#print(clastexrt)
			exrt=clastexrt		##### deactivate this line to revert to RT variation (expansion of transition list by varied RT within predicted range) ######## SWITCH #######
	vprecmz=vwritelist[4][r] #sheetinput.cell(row=r, column=5)	# PrecursorMz
	r=r+1
	corr=0
	if r==ki:
		r=r-1
		corr=1
	e=vwritelist[4][r] #sheetinput.cell(row=r, column=5)	# PrecursorMz
	if e==vprecmz:
		pos=0
	else:
		pos=1
	r=r+corr
	
# begin save to csv file
prt='PrecursorRT'
prt=str(prt)
toprow.append(prt)
rtw='PrecursorRTWindow'
rtw=str(rtw)
toprow.append(rtw)
ki=2+len(vmlistname)

vwritelist=[]
vwritelist.append(vmlistname)
vwritelist.append(vprecname)
vwritelist.append(virtualprecformula)
vwritelist.append(vprecadduct)
vwritelist.append(virtualprecmz)
vwritelist.append(vprecchrg)
vwritelist.append(vprodname)
vwritelist.append(vprodformula)
vwritelist.append(vprodadduct)
vwritelist.append(vprodmz)
vwritelist.append(vprodchrg)
vwritelist.append(precrt)
vwritelist.append(rtwindow)
#print('writelist created')
transitionresultsdf=pd.DataFrame(vwritelist).transpose()
#print('Transposed')
transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
#print('Transposed and DataFrame created')
after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_1_'+fourlettcode+'_'
#filename=today+'jpmlipidomics_vpw13_1_precursor.csv'
filename='jpmlipidomics_vpw20_1_precursor.csv'
transitionresultsdf.to_csv(filename, index=False)
afterall=datetime.datetime.now()
dt=afterall-beforeall
nrows=len(vmlistname)
#print('Transition list is saved as yyyy_mm_dd_1_xxxx_jpmlipidomics_vpw13_1_precursor.csv (%d rows)' % nrows)
print('Transition list is saved as jpmlipidomics_vpw20_1_precursor.csv (%d rows)' % nrows)
print('Calculation time (h:mm:ss) is:')
print(dt)
# end save to csv file
#################################################################################################################################################
############################### END REDUCE TO ONE ENTRY PER PRECURSOR AND EXPAND WITH VARIED EXPLICIT RETENTION TIME ############################
#################################################################################################################################################


