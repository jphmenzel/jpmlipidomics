# -*- coding: UTF-8 -*-

# Jan Philipp Menzel
# Goal: Analysis of DDA Skyline report using target list
## NOTES: VIRTUAL PRECURSOR - PrecursorName and PrecursorMz are artificially set +Xe (only column 3 and 5), fragment transitions correct including precursor
## NOTES: Virtual precursor forces Skyline to consider all transitions incl. real precursor (fragment in transition list), Skyline Setting: TransitionSettings-Filter-IonTypes-f 
import math
import openpyxl
import os
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
################ DATABASE ## Source: Internetchemie.info 
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.757, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
################
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
###########

discoverylevel=eval(input('Run full discovery workflow: 0; Streamlined discovery workflow (limited to FA_library for all FA > three db, otherwise full discovery): 1 or Library-based workflow (limited to FA_library): 2. Workflow: '))
dlevel=discoverylevel	

discoverylimitation=eval(input('Limit search to certain fatty acids by chain length? Yes: 1; No: 0. :'))
if discoverylimitation==1:
	cminlimit=eval(input('Search from minimum chain length: '))
	cmaxlimit=eval(input('to maximum chain length: '))


beforeall=datetime.datetime.now()
print('The targetlist is being analyzed to generate a transition list.')


# begin build mostwantedlist 
mostwantedlist=[]
wbfa=openpyxl.load_workbook('jpm_fa_lib.xlsx')
wsfa=wbfa.active
mwi=0
go=1
while go==1:
	tfe=wsfa.cell(row=mwi+3, column=1)
	tfe=tfe.value
	if tfe is None:
		go=0
	else:
		if int(tfe)<10:
			tfe='0'+str(tfe)
		mwfa=str(tfe)+':'
		tfe=wsfa.cell(row=mwi+3, column=2)
		tfe=tfe.value
		dbfa=int(tfe)
		if int(tfe)==0:
			mwfa=mwfa+str(tfe)
		else:
			mwfa=mwfa+str(tfe)+'_n-'
			dbfai=1
			while dbfai<(dbfa+1):
				tfe=wsfa.cell(row=mwi+3, column=dbfai+2)
				tfe=tfe.value
				if dbfai<(dbfa):
					mwfa=mwfa+str(tfe)+'_n-'
				else:
					mwfa=mwfa+str(tfe)
				dbfai=dbfai+1
		mostwantedlist.append(mwfa)
	mwi=mwi+1
# end build mostwantedlist

# begin get target list from txt file
rawtarget=open('jpmlipidomics_dda_targetlist.txt','r')
rawtargetlist=rawtarget.readlines()
#print(rawtargetlist)
mztargetlist=[]
rttargetlist=[]
rdi=0
while rdi<(len(rawtargetlist)):
	cstr=''
	rdii=0
	while rdii<10:
		stri=str(rawtargetlist[rdi][rdii])
		cstr=cstr+stri
		rdii=rdii+1
	cmz=float(cstr)
	mztargetlist.append(cmz)
	cstr=''
	rdii=33
	while rdii<43:
		stri=str(rawtargetlist[rdi][rdii])
		cstr=cstr+stri
		rdii=rdii+1
	if cstr[len(cstr)-1]==',':
		cstr=cstr[:-1]
	crt=float(cstr)/60
	rttargetlist.append(crt)
	rdi=rdi+1
lmzt=len(mztargetlist)
print('Number of targets in targetlist: %d' % lmzt)
#print(mztargetlist)
#print(rttargetlist)
#quit()
# end get target list from txt file
##############################################################################################################################################################################
agnostic=1		############ for using only identified species, set agnostic=0; for full de novo search ignoring outcome of AI workflow, set agnostic=1	############# !!! ####
if agnostic==0:
	# begin get data from csv file
	trdf=pd.read_csv('skyl_report_vpw20_4_rank1.csv')
	toprowx=[trdf.columns.values.tolist()]
	toprow=toprowx[0]
	trdf=trdf.transpose()
	writelist=trdf.values.tolist()
	ki=len(writelist[0])
	print('Number of rows in skyl_report_vpw20_4_rank1.csv (Skyline report) : %d' % ki)
	# end get data from csv file
	# begin make transition list with all found FA isomers for each target in targetlist
	# go through writelist, for each block of unsat FA (ignore duplicates), make transitions varied across all targetRT

	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	exrt=[]
	exrtwindow=[]

	specidlist=[] # list of species identities to identify duplicates
	twritelist=[]	# SETUP AS NEW twritelist with list elements and toprow ? ####################################################################
	r=0
	while r<ki:
		e=writelist[1][r] ## Precursorname		# begin determine which row to start (r) and to end (s)
		s=r+1
		st=0
		while st<1:
			if s>(len(writelist[1])-1):
				ne='stop_loop'
			else:
				ne=writelist[1][s] ## Precursorname
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		t=r
		
		if int(writelist[1][t][8])>0:
			cspecid=''
			csid=5
			if str(writelist[1][t][len(writelist[1][t])-4])=='_':
				csidstop=(len(writelist[1][t])-3)
			elif writelist[1][t][len(writelist[1][t])-6]=='_':
				csidstop=(len(writelist[1][t])-5)
			else:
				csidstop=(len(writelist[1][t])-4)
			while csid<csidstop:
				cspecid=cspecid+str(writelist[1][t][csid])
				csid=csid+1
			spidl=0	# test, whether current species was found before (duplicate)
			found=0
			while spidl<(len(specidlist)):
				if cspecid==specidlist[spidl]:
					found=1
				spidl=spidl+1
			if found==0:
				specidlist.append(cspecid)
				# begin make varied transitions across entries in targetlist
				csmz=float(writelist[9][s])
				crtlist=[]
				tgti=0
				while tgti<(len(rttargetlist)):
					if abs(csmz-(float(mztargetlist[tgti])))<0.02:
						crtlist.append(rttargetlist[tgti])
					tgti=tgti+1
				rtti=0
				while rtti<(len(crtlist)):
					t=r
					while t<(s+1):
						#append each list element from writelist with indexed target rt to new lists for twritelist
						e=writelist[0][t] ## mlistname	# begin append rows of suitable species to lists for later saving in excel output
						mlistname.append(e)
						e=writelist[1][t] ## precname
						if str(writelist[1][t][len(writelist[1][t])-4])=='_':
							cutn=3
						elif writelist[1][t][len(writelist[1][t])-6]=='_':
							cutn=5
						else:
							cutn=4
						e=e[:-cutn]	
						newrtid=str(format((crtlist[rtti]),'.3f'))
						e=e+newrtid
						precname.append(e)
						e=writelist[2][t] ## precformula	
						precformula.append(e)
						e=writelist[3][t] ## precadduct
						precadduct.append(e)
						e=writelist[4][t] ## precmz
						precmz.append(e)
						e=writelist[5][t] ## precchrg
						precchrg.append(e)
						e=writelist[6][t] ## prodname
						prodname.append(e)
						e=writelist[7][t] ## prodformula
						prodformula.append(e)
						e=writelist[8][t] ## prodadduct
						prodadduct.append(e)
						e=writelist[9][t] ## prodmz
						prodmz.append(e)
						e=writelist[10][t] ## prodchrg
						prodchrg.append(e)
						e=crtlist[rtti] ## exrt
						exrt.append(e)
						e=0.005 ## exrtwindow
						exrtwindow.append(e)
						t=t+1
					rtti=rtti+1
		r=s+1
	# end make transition list with all found FA isomers for each target in targetlist
	toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']
	vwritelist=[]
	vwritelist.append(mlistname)
	vwritelist.append(precname)
	vwritelist.append(precformula)
	vwritelist.append(precadduct)
	vwritelist.append(precmz)
	vwritelist.append(precchrg)
	vwritelist.append(prodname)
	vwritelist.append(prodformula)
	vwritelist.append(prodadduct)
	vwritelist.append(prodmz)
	vwritelist.append(prodchrg)
	vwritelist.append(exrt)
	vwritelist.append(exrtwindow)
	#print('writelist created')
	transitionresultsdf=pd.DataFrame(vwritelist).transpose()
	#print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
	#print('Transposed and DataFrame created')
	after=datetime.datetime.now()
	after=str(after)
	#today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_1_'+fourlettcode+'_'
	#filename=today+'jpmlipidomics_vpw13_1_precursor.csv'
	filename='jpmlipidomics_dda_vpw20_0.csv'
	transitionresultsdf.to_csv(filename, index=False)
	afterall=datetime.datetime.now()
	dt=afterall-beforeall
	nrows=len(mlistname)
	#print('Transition list is saved as yyyy_mm_dd_1_xxxx_jpmlipidomics_vpw13_1_precursor.csv (%d rows)' % nrows)
	print('Transition list is saved as jpmlipidomics_dda_vpw20_0.csv (%d rows)' % nrows)
	print('Calculation time (h:mm:ss) is:')
	print(dt)
	quit()
	# end save to csv file

###############################################################################################################################
###############################################################################################################################
###############################################################################################################################
# begin build full transition list and reduce and expand with reassigned targets (for neighboring targets, skip every second to reduce number of transitions leading to same peak)

# begin reassign targets in targetlist (skip neighboring targets - targets with two direct neighbors)
mztargettolerance=0.005
rttargettolerance=0.005
kt=0
while kt<(len(rttargetlist)-4):
	if abs(float(mztargetlist[kt])-float(mztargetlist[kt+1]))<mztargettolerance:
		if abs(float(mztargetlist[kt+1])-float(mztargetlist[kt+2]))<mztargettolerance:
			if abs(float(mztargetlist[kt+2])-float(mztargetlist[kt+3]))<mztargettolerance:
				if abs(float(mztargetlist[kt+3])-float(mztargetlist[kt+4]))<mztargettolerance:
					if abs((abs(float(rttargetlist[kt])-float(rttargetlist[kt+1])))-0.01)<rttargettolerance:
						if abs((abs(float(rttargetlist[kt+1])-float(rttargetlist[kt+2])))-0.01)<rttargettolerance:
							if abs((abs(float(rttargetlist[kt+2])-float(rttargetlist[kt+3])))-0.01)<rttargettolerance:
								if abs((abs(float(rttargetlist[kt+3])-float(rttargetlist[kt+4])))-0.01)<rttargettolerance:
									del mztargetlist[kt+1]
									del rttargetlist[kt+1]
									del mztargetlist[kt+2]
									del rttargetlist[kt+2]
									del mztargetlist[kt+3]
									del rttargetlist[kt+3]
				elif abs((abs(float(rttargetlist[kt])-float(rttargetlist[kt+1])))-0.01)<rttargettolerance:
					if abs((abs(float(rttargetlist[kt+1])-float(rttargetlist[kt+2])))-0.01)<rttargettolerance:
						if abs((abs(float(rttargetlist[kt+2])-float(rttargetlist[kt+3])))-0.01)<rttargettolerance:	
							del mztargetlist[kt+1]
							del rttargetlist[kt+1]	
							del mztargetlist[kt+2]
							del rttargetlist[kt+2]				
			elif abs((abs(float(rttargetlist[kt])-float(rttargetlist[kt+1])))-0.01)<rttargettolerance:
				if abs((abs(float(rttargetlist[kt+1])-float(rttargetlist[kt+2])))-0.01)<rttargettolerance:	
					del mztargetlist[kt+1]
					del rttargetlist[kt+1]	
	kt=kt+1
# end reassign targets in targetlist (skip neighboring targets - targets with two direct neighbors)
lmzt=len(mztargetlist)
#print('Number of targets in targetlist after first reassignment: %d' % lmzt)

# begin reassign targets in targetlist 

kt=0
while kt<(len(rttargetlist)-4):
	ku=0
	while ku<(len(rttargetlist)-4):
		if abs(float(mztargetlist[kt])-float(mztargetlist[ku]))<mztargettolerance:
			if abs((abs(float(rttargetlist[kt])-float(rttargetlist[ku])))-0.01)<rttargettolerance:
				kv=0
				while kv<(len(rttargetlist)-4):
					if abs(float(mztargetlist[kt])-float(mztargetlist[ku]))<mztargettolerance:
						if abs(float(mztargetlist[kt])-float(mztargetlist[kv]))<mztargettolerance:
							if abs((abs(float(rttargetlist[kt])-float(rttargetlist[ku])))-0.01)<rttargettolerance:
								if abs((abs(float(rttargetlist[kt])-float(rttargetlist[kv])))-0.01)<rttargettolerance:
									if kt<ku:
										if kt>kv:	
											del mztargetlist[kt]
											del rttargetlist[kt]
										else:
											if ku<kv:
												del mztargetlist[ku]
												del rttargetlist[ku]
											else:
												del mztargetlist[kv]
												del rttargetlist[kv]
									else:
										if kt<kv:
											del mztargetlist[kt]
											del rttargetlist[kt]
										else:
											if ku<kv:
												del mztargetlist[kv]
												del rttargetlist[kv]
											else:
												del mztargetlist[ku]
												del rttargetlist[ku]
					kv=kv+1
		ku=ku+1
	kt=kt+1
# end reassign targets in targetlist 
lmzt=len(mztargetlist)
#print('Number of targets in targetlist after second reassignment: %d' % lmzt)
print('Number of targets in processed targetlist after reassignment: %d. Transition list is being generated...' % lmzt)

########### begin read workflow parameters
transferlist=[]
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
go=1
while go==1:
	tfe=ws.cell(row=tli+1, column=2)
	tfe=tfe.value
	if tli==17:
		go=0
	if tfe is None:
		go=0
	else:
		transferlist.append(tfe)
	tli=tli+1
########## 
fourlettcode=transferlist[0]		#input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
cderiv=int(transferlist[1])
hderiv=int(transferlist[2])
dderiv=int(transferlist[3])
nderiv=int(transferlist[4])
oderiv=int(transferlist[5])
pderiv=int(transferlist[6])
ideriv=int(transferlist[7])
workflowmode=int(transferlist[13])
mzcutoff=int(transferlist[10]) 		#eval(input('What is max mz [ppm] for positive identification of species? (e.g. 30) :'))
rettimecutoff=float(transferlist[11]) 		#eval(input('What is the maximum retention time [min]? (e.g. 11.9 or 17.8) :'))
areathreshold=int(transferlist[12])		#eval(input('What is the threshold for detected precursor peak area? (compare Skyline report file, e.g. 3000) :'))
mostwanted=int(transferlist[17])
########### end read workflow parameters
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e        ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]


# begin calculate monounsaturated precursors from input
#maxlenfa=24 #longest expected fatty acid
#minlenfa=12	#shortest expected fatty acid
minlenfa=transferlist[8] #eval(input('Enter number of C atoms in shortest FA chain (e.g. 4) :'))
maxlenfa=transferlist[9] #eval(input('Enter number of C atoms in longest FA chain (e.g. 24) :'))

#print('Calculation of transition list is running ...')

if discoverylevel<2:
	lfa=maxlenfa-minlenfa+1
	moleculegrouplist=[]
	precursornamelist=[]
	precursorformulalist=[]
	precursoradductlist=[]
	precursormzlist=[]
	precursorchargelist=[]
	productnamelist=[]
	productformulalist=[]
	productadductlist=[]
	productmzlist=[]
	productchargelist=[]
	molg=fourlettcode+'_FA'
	padd='[M]1+'
	prdn='precursor'
	li=0
	while li<lfa:
		moleculegrouplist.append(molg)
		currentlfa=str(li+minlenfa)
		if len(currentlfa)<2:
			currentlfa='0'+currentlfa
			currentlfa=str(currentlfa)
		pnm=fourlettcode+'_'+currentlfa+':1'
		precursornamelist.append(pnm)
		prf=''
		currentcderiv=0
		currenthderiv=0
		currentdderiv=0
		currentnderiv=0
		currentoderiv=0
		currentpderiv=0
		currentideriv=0
		if cderiv>0:
			currentcderiv=cderiv+int(currentlfa)
			prf=prf+'C'+str(currentcderiv)
		if hderiv>0:
			currenthderiv=hderiv+(2*int(currentlfa))-3
			prf=prf+'H'+str(currenthderiv)
		if dderiv>0:
			currentdderiv=dderiv
			prf=prf+"H'"+str(currentdderiv)
		if nderiv>0:
			currentnderiv=nderiv
			prf=prf+'N'+str(nderiv)
		if oderiv>(-1):
			currentoderiv=oderiv+1
			prf=prf+'O'+str(currentoderiv)
		if pderiv>0:
			currentpderiv=pderiv
			prf=prf+'P'+str(pderiv)
		if ideriv>0:
			currentideriv=ideriv
			prf=prf+'I'+str(ideriv)
		precursorformulalist.append(prf)
		productformulalist.append(prf)
		precursoradductlist.append(padd)
		prmz=imass[0]*currenthderiv+imass[1]*currentdderiv+imass[2]*currentcderiv+imass[3]*currentnderiv+imass[4]*currentoderiv+imass[5]*currentpderiv+imass[10]*currentideriv
		precursormzlist.append(prmz)
		productmzlist.append(prmz)
		precursorchargelist.append(1)
		productnamelist.append(prdn)
		productadductlist.append(padd)
		productchargelist.append(1)
		li=li+1
	fareadlist=[]
	fareadlist.append(moleculegrouplist)
	fareadlist.append(precursornamelist)
	fareadlist.append(precursorformulalist)
	fareadlist.append(precursoradductlist)
	fareadlist.append(precursormzlist)
	fareadlist.append(precursorchargelist)
	fareadlist.append(productnamelist)
	fareadlist.append(productformulalist)
	fareadlist.append(productadductlist)
	fareadlist.append(productmzlist)
	fareadlist.append(productchargelist)
	# end calculate monounsaturated precursors from input

	preconly=0
	#minrt=0.0
	#maxrt=eval(input('At which retention time [min] ends gradient? :'))
	#explrt=(minrt+maxrt)/2
	#explrtw=explrt-minrt
	nspec=3	# number of species: precursor, aldehyde, criegee
	# begin create empty lists
	toprow=[]
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	dbplist=[]		# double bond position list, contains strings with double bond position n-1, n-2 ...
	dbindexlist=[]	# double bond index list, contains index for the double bond position that is closest to head group of FA
	#end create empty lists
	#create lists for saturated FAs
	toprow=['Moleculegroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
	ki=len(fareadlist[0])
	satlist=fareadlist
	#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

	r=0
	while r<ki:		#go through rows of fareadlist
		# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
		cchain=0
		e=str(fareadlist[1][r]) # Precursorname
		i=len(e)-3
		add=e[i]
		add=float(add)
		cchain=cchain+add
		i=i-1
		add=e[i]
		add=float(add)
		if add>0:
			cchain=cchain+(10*add)
		else:
			cchain=cchain
			#print('Please check source code (determine number of C atoms in chain)')
		maxdbp=cchain-2
		# begin determine number of C atoms in chain, define highest possible double bond position
		kadd=nspec*maxdbp#+1		#(nspec=3 precursor and 2 products - aldehyde and crigee  - for each of maxdbp possible double bond positions)

		e=fareadlist[0][r]  # MoleculeGroup
		f=fareadlist[1][r] 	# Precursorname
		g=fareadlist[2][r] 	# Precursorformula
		h=float(fareadlist[5][r]) # PrecursorCharge
		k=0
		while k<kadd:
			mlistname.append(e)		# copied, no change
			precname.append(f)		# copied, no change
			precformula.append(g)	# copied, no change
			precchrg.append(h)		# copied, no change
			k=k+1

		e=fareadlist[3][r] 	# Precursoradduct
		k=0
		while k<kadd:
			precadduct.append(e) 	# AMPP, precursor
			precadduct.append(e)	# AMPP, aldehyde product
			precadduct.append(e)	# AMPP, crigee product
			k=k+nspec

		e=float(fareadlist[4][r]) 	# PrecursorMz
		k=0
		while k<kadd:
			precmz.append(e)	# precursor
			precmz.append(e)	# aldehyde
			precmz.append(e)	# crigee
			k=k+nspec

		e=fareadlist[1][r] 	# Productname
		if e=='Cholesterol (+[2]H7)':
			fragment='_'	
		else:	
			k=0
			csub=1
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				dbp='_n-'+str(csub)
				fragment='_aldehyde neutral loss'
				ne=e+dbp+fragment
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				#
				if preconly==0:	
					prodname.append(ne)		# aldehyde
					dbl=[]	#begin save double bond position for later
					dbl.append(csub)
					dbindexlist.append(dbl)	#end save double bond position for later
					#prodname.append(ne)
					#prodname.append(ne)
					fragment='_criegee neutral loss'
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					dbl=[]	#begin save double bond position for later
					dbl.append(csub)
					dbindexlist.append(dbl)	#end save double bond position for later
					#prodname.append(ne)
					#prodname.append(ne)
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor a'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor b'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]	#begin save double bond position for later
					dbl.append(csub)
					dbindexlist.append(dbl)
					dbindexlist.append(dbl)	#end save double bond position for later
				csub=csub+1
				k=k+nspec

		e=fareadlist[2][r] 
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		#print(e[1])
		#print(e[2])
		#print(e[3])
		#print(len(e))
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=='D':
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=='D':
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)
		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=fareadlist[0][r] 	# begin calculate product sum formula
		if e=='SPLASH':
			e=e 
		else:
			k=0
			csub=1
			while k<kadd:
				cnp=cn-(csub)
				hnp=hn-(2*csub)
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hn)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
				if preconly==1:
					prodformula.append(precursor)
					prodformula.append(precursor)
				precursormz=imass[0]*(hn)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)

		e=fareadlist[0][r] 	# begin calculate product m/z
		if e=='SPLASH':
			print('SPLASH')	# end calculate product m/z for SPLASH
		else:
			e=e 	# no action, as product m/z was calculated from ProductFormula and saved respectively (see above)
		# end calculate product m/z
		prodchrg=precchrg 	# ProductCharge

		r=r+1
		e=fareadlist[1][r-1]	# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		prevprec=e
		if r>(len(fareadlist[0])-1):
			r=r+1
		else:
			e=fareadlist[1][r] 
			prec=e
			while prec==prevprec:
				r=r+1
				e=fareadlist[1][r-1] 
				prevprec=e
				e=fareadlist[1][r] 
				prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		
	#print(len(mlistname))
	#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

	#begin go through lists and add double bond position to PrecursorName
	k=0
	kt=len(precname)
	#print(precname)
	#print(dbplist)
	#if len(precname)==len(dbplist):
		#print('test correct')
	while k<kt:
		precname[k]=precname[k]+dbplist[k]
		k=k+1
	#end go through lists and add double bond position to PrecursorName

	#begin save data in writelist
	toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
	writelist=[]
	writelist.append(mlistname)
	writelist.append(precname)
	writelist.append(precformula)
	writelist.append(precadduct)
	writelist.append(precmz)
	writelist.append(precchrg)
	writelist.append(prodname)
	writelist.append(prodformula)
	writelist.append(prodadduct)
	writelist.append(prodmz)
	writelist.append(prodchrg)
	#end save data in writelist

	#print('All calculations for monounsaturated fatty acids are done.')
	#end save excel file
	#print ('odd : C_%d H_%d N_%d O_%d S_%d; DBE = %d; deviation: %.3f' % (formula[1], formula[0], formula[2], formula[3], formula[4], dbe, meandeviation))
	###########################################################################################################################################################################
	#############################################################################MONO##########################################################################################
	############################################################################DOUBLE#########################################################################################
	###########################################################################################################################################################################
	# begin add double unsaturated fatty acids
	#ask=eval(input('Add fatty acids with two double bonds? Yes: 1; No: 0 | '))
	ask=1
	if ask==0:
		quit()

	nspec=5	# number of species: precursor, aldehyde db1 and db2, crigee db1 and db2
	# begin create empty lists
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	dbplist=[]
	#end create empty lists
	#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
	#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
	#sheetinput=wb['transitionlist']
	ki=len(writelist[0])
	kmono=ki # List index in excel file to start writing FA transitions with two double bonds
	kdoublestart=kmono
	r=0
	ki=ki
	while r<ki:		#go through rows of excel file jpmlipidozidinput
		# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
		cchain=0
		e=str(writelist[1][r])#sheetinput.cell(row=r, column=2)	# Precursorname
		add=e[5]
		add=int(add)
		#print(add)
		cchain=cchain+(10*add)
		add=e[6]
		add=int(add)
		cchain=cchain+(1*add)
		maxdbp=cchain-2
		# end determine number of C atoms in chain, define highest possible double bond position
		# begin determine double bond position of already located double bond (dbpmono)
		if len(e)==13:
			dbpmono=int(e[12])
		elif len(e)==14:
			dbpmono=(10*(int(e[12])))+(int(e[13]))
		else:
			print('Please check source code (determine double bond position of already located double bond)')
		# end determine double bond position of already located double bond
		nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the second double bond
		csubmono=dbpmono+2
		if nsecdbp>0:
			kadd=nspec*nsecdbp#+1		#(nspec=3 products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)
			e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
			k=0
			while k<kadd:
				mlistname.append(e)	# copied, no change
				k=k+1

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
			degunsat=int(e[8])		# change from monounsaturated to bisunsaturated FA in PrecursorName
			if degunsat==1:
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(2)
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne
			k=0
			while k<kadd:
				precname.append(e)	# copied, no change
				k=k+1

			e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
			k=0
			while k<kadd:
				precadduct.append(e) 	# AMPP, precursor
				precadduct.append(e)	# AMPP, aldehyde product
				precadduct.append(e)	# AMPP, crigee product
				precadduct.append(e)	# AMPP, aldehyde product, outer db
				precadduct.append(e)	# AMPP, crigee product, outer db
				k=k+nspec

			e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
			e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added second double bond
			k=0
			while k<kadd:
				precmz.append(e)	# precursor
				precmz.append(e)	# aldehyde
				precmz.append(e)	# crigee
				precmz.append(e)	# aldehyde outer db
				precmz.append(e)	# crigee outer db
				k=k+nspec

			e=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
			k=0
			while k<kadd:
				precchrg.append(e)
				k=k+1

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
			degunsat=int(e[8])		# begin change from monounsaturated to bisunsaturated FA in PrecursorName
			if degunsat==1:
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(2)
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne		# end change from monounsaturated to bisunsaturated FA in PrecursorName
			if e=='Cholesterol (+[2]H7)':
				fragment='_ozone neutral gain' 		
			else:	
				k=0
				csub=csubmono
				while k<kadd:
					dbp='_n-'+str(csub)
					fragment='_precursor'
					ne=e+dbp+fragment
					prodname.append(ne)		# precursor
					dbl=[]	#begin save double bond position for later
					firstdbp=dbindexlist[r][0]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	#end save double bond position for later
					if preconly==0:
						dbp='_n-'+str(csub)
						fragment='_aldehyde neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
					dbpi=0
					while dbpi<nspec:
						dbplist.append(dbp)
						dbpi=dbpi+1
					if preconly==0:
						prodname.append(ne)		# aldehyde
						dbl=[]	#begin save double bond position for later
						dbl.append(firstdbp)
						dbl.append(csub)
						dbindexlist.append(dbl)	#end save double bond position for later
						fragment='_criegee neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee
						dbl=[]
						dbl.append(firstdbp)
						dbl.append(csub)
						dbindexlist.append(dbl)	# save double bond position for later
						fragment='_aldehyde neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde outer double bond
						dbl=[]
						dbl.append(firstdbp)
						dbl.append(csub)
						dbindexlist.append(dbl)	# save double bond position for later
						fragment='_criegee neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee outer double bond
						dbl=[]
						dbl.append(firstdbp)
						dbl.append(csub)
						dbindexlist.append(dbl)	# save double bond position for later
					if preconly==1:
						dbp='_n-'+str(csub)
						fragment='_dummy precursor a'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor b'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor c'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor d'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						dbl=[]
						dbl.append(firstdbp)
						dbl.append(csub)
						dbindexlist.append(dbl)	# save double bond position for later
						dbl=[]
						dbl.append(firstdbp)
						dbl.append(csub)
						dbindexlist.append(dbl)	# save double bond position for later
						dbl=[]
						dbl.append(firstdbp)
						dbl.append(csub)
						dbindexlist.append(dbl)	# save double bond position for later
						dbl=[]
						dbl.append(firstdbp)
						dbl.append(csub)
						dbindexlist.append(dbl)	# save double bond position for later
					csub=csub+1
					k=k+nspec

			e=writelist[2][r] #sheetinput.cell(row=r, column=3)
			# begin read precursor sum formula and edit product sum formula
			#print(e)
			#print(e[0])
			clist=[]
			hlist=[]
			dlist=[]
			nlist=[]
			olist=[]
			plist=[]
			ilist=[]
			i=0
			ca=0
			ha=0
			da=0
			na=0
			oa=0
			pa=0
			ia=0
			while i<len(e):
				if e[i]=='H':
					if e[i+1]=="'":
						ha=0
					else:
						ca=0
				#if e[i]=='D':
				#	ha=0		
				if e[i]=='N':
					ha=0
					da=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
				if e[i]=='P':
					ha=0
					da=0
					na=0
					oa=0
				if e[i]=='I':
					ha=0
					da=0
					na=0
					oa=0
					pa=0
				if ca==1:
					clist.append(e[i])
				if ha==1:
					hlist.append(e[i])
				if da==1:
					dlist.append(e[i])
				if na==1:
					nlist.append(e[i])
				if oa==1:
					olist.append(e[i])
				if pa==1:
					plist.append(e[i])
				if ia==1:
					ilist.append(e[i])
				if e[i]=='C':
					ca=1
				if e[i]=='H':
					if e[i+1]=="'":
						ca=0
						ha=0
						da=1
						i=i+1
					else:
						ca=0
						ha=1
				#if e[i]=='D':
				#	ca=0
				#	ha=0
				#	da=1		
				if e[i]=='N':
					ha=0
					da=0
					na=1
					if e[i+1]=='O':
						nlist.append('1')
						na=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
					oa=1
					if (i+1)<len(e):
						if e[i+1]=='P':
							olist.append('1')
							oa=0
					else:
						olist.append('1')
						oa=0					
				if e[i]=='P':
					da=0
					na=0
					oa=0
					pa=1
					if (i+1)<len(e):
						if e[i+1]=='I':
							plist.append('1')
							pa=0
					else:
						plist.append('1')
						pa=0
				if e[i]=='I':
					da=0
					na=0
					oa=0
					pa=0
					ia=1
					if i==(len(e)-1):
						ilist.append('1')
						ia=0
				i=i+1
			#print(clist)
			#print(hlist)
			#print(dlist)
			#print(nlist)
			#print(olist)
			#print(plist)

			if len(clist)==0:
				cn=0
			if len(hlist)==0:
				hn=0
			if len(dlist)==0:
				dn=0	
			if len(nlist)==0:
				nn=0
			if len(olist)==0:
				on=0
			if len(plist)==0:
				pn=0
			if len(ilist)==0:
				iodon=0
			if len(clist)==1:
				cn=int(clist[0])
			if len(clist)==2:
				cn=10*int(clist[0])+int(clist[1])
			if len(clist)==3:
				cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
			if len(hlist)==1:
				hn=int(hlist[0])
			if len(hlist)==2:
				hn=10*int(hlist[0])+int(hlist[1])
			if len(hlist)==3:
				hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
			if len(dlist)==1:
				dn=int(dlist[0])
			if len(dlist)==2:
				dn=10*int(dlist[0])+int(dlist[1])
			if len(dlist)==3:
				dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
			if len(nlist)==1:
				nn=int(nlist[0])
			if len(nlist)==2:
				nn=10*int(nlist[0])+int(nlist[1])
			if len(nlist)==3:
				nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
			if len(olist)==1:
				on=int(olist[0])
			if len(olist)==2:
				on=10*int(olist[0])+int(olist[1])
			if len(olist)==3:
				on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
			if len(plist)==1:
				pn=int(plist[0])
			if len(plist)==2:
				pn=10*int(plist[0])+int(plist[1])
			if len(plist)==3:
				pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
			if len(ilist)==1:
				iodon=int(ilist[0])
			if len(ilist)==2:
				iodon=10*int(ilist[0])+int(ilist[1])
			if len(ilist)==3:
				iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
			#print(cn)
			e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
			if e=='SPLASH':
				#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
				e=e
			else:
				k=0
				csub=csubmono
				while k<kadd:
					hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with two double bonds
					hnpmono=hnprec-(2*dbpmono)	# applies to cleavage of outer double bond
					cnpmono=cn-(dbpmono)		# applies to cleavage of outer double bond
					cnp=cn-(csub)				# applies to cleavage of inner double bond
					hnp=hn-(2*csub)				# applies to cleavage of inner double bond
					onald=on+1
					oncrigee=on+2
					precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
					productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydemono='C'+str(cnpmono)+'H'+str(hnpmono)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeemono='C'+str(cnpmono)+'H'+str(hnpmono)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					t=0
					while t<nspec:
						precformula.append(precursor)
						t=t+1
					prodformula.append(precursor)
					if preconly==0:
						prodformula.append(productaldehyde)
						prodformula.append(productcrigee)
						prodformula.append(productaldehydemono)
						prodformula.append(productcrigeemono)	
					if preconly==1:
						prodformula.append(precursor)	
						prodformula.append(precursor)
						prodformula.append(precursor)
						prodformula.append(precursor)
					# # product formula is saved in list for current double bond position
					precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
					prodmz.append(precursormz)
					if preconly==0:
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, inner db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, inner db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpmono)+imass[1]*dn+imass[2]*cnpmono+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, inner db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpmono)+imass[1]*dn+imass[2]*cnpmono+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, inner db, AMPP
						prodmz.append(productmz)
					if preconly==1:
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
					csub=csub+1
					k=k+nspec			
			# end read precursor sum formula and edit product sum formula

			prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
			prodchrg=precchrg 	# ProductCharge  #############

		r=r+1
		e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		prevprec=e
		e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
		prec=e
		while prec==prevprec:
			if r<(ki-1):
				r=r+1
				e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
				prevprec=e
				e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
				prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
			else:
				prec='stop_loop'
				r=r+1
		
	#print(len(mlistname))
	#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

	#begin go through lists and add double bond position to PrecursorName
	k=0
	kt=len(precname)
	#print(precname)
	#print(dbplist)
	#if len(precname)==len(dbplist):
		#print('test correct')
	while k<kt:
		precname[k]=precname[k]+dbplist[k]
		k=k+1
	#end go through lists and add double bond position to PrecursorName

	#begin save data as writelist, merge new lists at end of each list in writelist
	writelist[0]=writelist[0]+mlistname
	writelist[1]=writelist[1]+precname
	writelist[2]=writelist[2]+precformula
	writelist[3]=writelist[3]+precadduct
	writelist[4]=writelist[4]+precmz
	writelist[5]=writelist[5]+precchrg
	writelist[6]=writelist[6]+prodname
	writelist[7]=writelist[7]+prodformula
	writelist[8]=writelist[8]+prodadduct
	writelist[9]=writelist[9]+prodmz
	writelist[10]=writelist[10]+prodchrg
	#end save data as writelist

	#print('All calculations for fatty acids with two double bonds are done.')
	#end save excel file
	# end add double unsaturated fatty acids
	###########################################################################################################################################################################
	###########################################################################DOUBLE##########################################################################################
	###########################################################################TRIPLE##########################################################################################
	###########################################################################################################################################################################
	# begin add triple unsaturated fatty acids
	#print(len(dbindexlist))
	#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
	ask=1
	if ask==0:
		quit()

	nspec=7	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3
	# begin create empty lists
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	dbplist=[]
	#end create empty lists
	#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
	#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
	#sheetinput=wb['transitionlist']
	ki=len(writelist[0])

	kmono=ki # List index in writelist to start writing FA transitions with three double bonds
	ktriplestart=kmono
	r=kdoublestart
	ki=ki
	while r<ki:		#go through rows of FAs with two double bonds in excel file
		# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
		cchain=0
		e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
		add=e[5]
		add=int(add)
		#print(add)
		cchain=cchain+(10*add)
		add=e[6]
		add=int(add)
		cchain=cchain+(1*add)
		maxdbp=cchain-2
		#maxdbp=
		# end determine number of C atoms in chain, define highest possible double bond position
		# begin determine double bond position of already located double bond (dbpmono)
		#if len(e)==14:
		#	dbpmono=int(e[13])
		#elif len(e)==15:
		#	dbpmono=(10*(int(e[13])))+(int(e[14]))
		#else:
		#	print('Please check source code (determine double bond position of already located double bond)')
		dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond		########################################## r instead of r-2 ##################
		# end determine double bond position of already located double bond
		nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
		csubmono=dbpmono+2
		if nsecdbp>0:

			kadd=nspec*nsecdbp#+1		#(nspec=3 products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

			e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
			f=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
			k=0
			while k<kadd:
				mlistname.append(e)	# copied, no change
				precchrg.append(f)	# copied, no change
				k=k+1	

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
			degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
			if degunsat==2:
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(3)
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne
			k=0
			while k<kadd:
				precname.append(e)	# write new precursorname, changed to FA with three db
				k=k+1

			e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
			k=0
			while k<kadd:
				adding=0
				while adding<nspec:
					precadduct.append(e) 	# precursor and all transitions for each db
					adding=adding+1
				k=k+nspec

			e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
			e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added second double bond
			k=0
			while k<kadd:
				adding=0
				while adding<nspec:
					precmz.append(e) 	# precursor and all transitions for each db
					adding=adding+1
				k=k+nspec

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
			degunsat=int(e[8])		# begin change from bisunsaturated to trisunsaturated FA in PrecursorName
			if degunsat==2:
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(3)
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne		# end change from bisunsaturated to trisunsaturated FA in PrecursorName
			if e=='Cholesterol (+[2]H7)':
				fragment='_ozone neutral gain' 		
			else:	
				k=0
				csub=csubmono
				while k<kadd:
					dbp='_n-'+str(csub)
					fragment='_precursor'
					ne=e+dbp+fragment
					prodname.append(ne)		# precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1] ########################################## r instead of r-2 ##################
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	#end save double bond position for later
					dbpi=0
					while dbpi<nspec:
						dbplist.append(dbp)
						dbpi=dbpi+1
					if preconly==0:
						dbp='_n-'+str(csub)
						fragment='_aldehyde neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde
						fragment='_criegee neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee
						fragment='_aldehyde neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde last existing double bond
						fragment='_criegee neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee last existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						dbl=[]	#begin save double bond position for later 						
						firstdbp=dbindexlist[r][0]
						seconddbp=dbindexlist[r][1]
						dbl.append(firstdbp)
						dbl.append(seconddbp)
						dbl.append(csub)
						ti=0
						while ti<(nspec-1):
							dbindexlist.append(dbl)	#end save double bond position for later
							ti=ti+1
					if preconly==1:
						dbp='_n-'+str(csub)
						fragment='_dummy precursor1'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor2'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor3'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor4'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor5'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor6'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						dbl=[]	#begin save double bond position for later 						
						firstdbp=dbindexlist[r][0]
						seconddbp=dbindexlist[r][1]	########################################## r instead of r-2 ##################
						dbl.append(firstdbp)
						dbl.append(seconddbp)
						dbl.append(csub)
						ti=0
						while ti<(nspec-1):
							dbindexlist.append(dbl)	#end save double bond position for later
							ti=ti+1
					csub=csub+1
					k=k+nspec

			e=writelist[2][r] #PrecursorFormula
			# begin read precursor sum formula and edit product sum formula
			#print(e)
			#print(e[0])
			clist=[]
			hlist=[]
			dlist=[]
			nlist=[]
			olist=[]
			plist=[]
			ilist=[]
			i=0
			ca=0
			ha=0
			da=0
			na=0
			oa=0
			pa=0
			ia=0
			while i<len(e):
				if e[i]=='H':
					if e[i+1]=="'":
						ha=0
					else:
						ca=0
				#if e[i]=='D':
				#	ha=0		
				if e[i]=='N':
					ha=0
					da=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
				if e[i]=='P':
					ha=0
					da=0
					na=0
					oa=0
				if e[i]=='I':
					ha=0
					da=0
					na=0
					oa=0
					pa=0
				if ca==1:
					clist.append(e[i])
				if ha==1:
					hlist.append(e[i])
				if da==1:
					dlist.append(e[i])
				if na==1:
					nlist.append(e[i])
				if oa==1:
					olist.append(e[i])
				if pa==1:
					plist.append(e[i])
				if ia==1:
					ilist.append(e[i])
				if e[i]=='C':
					ca=1
				if e[i]=='H':
					if e[i+1]=="'":
						ca=0
						ha=0
						da=1
						i=i+1
					else:
						ca=0
						ha=1
				#if e[i]=='D':
				#	ca=0
				#	ha=0
				#	da=1		
				if e[i]=='N':
					ha=0
					da=0
					na=1
					if e[i+1]=='O':
						nlist.append('1')
						na=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
					oa=1
					if (i+1)<len(e):
						if e[i+1]=='P':
							olist.append('1')
							oa=0
					else:
						olist.append('1')
						oa=0					
				if e[i]=='P':
					da=0
					na=0
					oa=0
					pa=1
					if (i+1)<len(e):
						if e[i+1]=='I':
							plist.append('1')
							pa=0
					else:
						plist.append('1')
						pa=0
				if e[i]=='I':
					da=0
					na=0
					oa=0
					pa=0
					ia=1
					if i==(len(e)-1):
						ilist.append('1')
						ia=0
				i=i+1
			#print(clist)
			#print(hlist)
			#print(dlist)
			#print(nlist)
			#print(olist)
			#print(plist)

			if len(clist)==0:
				cn=0
			if len(hlist)==0:
				hn=0
			if len(dlist)==0:
				dn=0	
			if len(nlist)==0:
				nn=0
			if len(olist)==0:
				on=0
			if len(plist)==0:
				pn=0
			if len(ilist)==0:
				iodon=0
			if len(clist)==1:
				cn=int(clist[0])
			if len(clist)==2:
				cn=10*int(clist[0])+int(clist[1])
			if len(clist)==3:
				cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
			if len(hlist)==1:
				hn=int(hlist[0])
			if len(hlist)==2:
				hn=10*int(hlist[0])+int(hlist[1])
			if len(hlist)==3:
				hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
			if len(dlist)==1:
				dn=int(dlist[0])
			if len(dlist)==2:
				dn=10*int(dlist[0])+int(dlist[1])
			if len(dlist)==3:
				dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
			if len(nlist)==1:
				nn=int(nlist[0])
			if len(nlist)==2:
				nn=10*int(nlist[0])+int(nlist[1])
			if len(nlist)==3:
				nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
			if len(olist)==1:
				on=int(olist[0])
			if len(olist)==2:
				on=10*int(olist[0])+int(olist[1])
			if len(olist)==3:
				on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
			if len(plist)==1:
				pn=int(plist[0])
			if len(plist)==2:
				pn=10*int(plist[0])+int(plist[1])
			if len(plist)==3:
				pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
			if len(ilist)==1:
				iodon=int(ilist[0])
			if len(ilist)==2:
				iodon=10*int(ilist[0])+int(ilist[1])
			if len(ilist)==3:
				iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
			#print(cn)
			e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
			if e=='SPLASH':
				#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
				e=e
			else:
				k=0
				csub=csubmono
				while k<kadd:
					hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
					cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
					hnp=hnprec-(2*csub-2*2)		# applies to cleavage of first (highest n) double bond
					hnptwo=hnprec-(2*dbpmono-2*1)	# applies to cleavage of second double bond
					cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
					hnpthree=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of third double bond
					cnpthree=cn-(dbindexlist[r][0])			# applies to cleavage of third double bond
					onald=on+1
					oncrigee=on+2
					precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
					productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					t=0
					while t<nspec:
						precformula.append(precursor)
						t=t+1
					prodformula.append(precursor)
					if preconly==0:
						prodformula.append(productaldehyde)
						prodformula.append(productcrigee)
						prodformula.append(productaldehydetwo)
						prodformula.append(productcrigeetwo)
						prodformula.append(productaldehydethree)
						prodformula.append(productcrigeethree)	
					if preconly==1:
						prodformula.append(precursor)	
						prodformula.append(precursor)
						prodformula.append(precursor)
						prodformula.append(precursor)
						prodformula.append(precursor)
						prodformula.append(precursor)
					# # product formula is saved in list for current double bond position
					precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
					prodmz.append(precursormz)
					if preconly==0:
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
						prodmz.append(productmz)
					if preconly==1:
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
					csub=csub+1
					k=k+nspec			
			# end read precursor sum formula and edit product sum formula

			prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
			prodchrg=precchrg 	# ProductCharge  #############

		r=r+1
		e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		prevprec=e
		e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
		prec=e
		while prec==prevprec:
			if r<(ki-1):
				r=r+1
				e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
				prevprec=e
				e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
				prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
			else:
				prec='stop_loop'
				r=r+1
		
	#print(len(mlistname))
	#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

	#begin go through lists and add double bond position to PrecursorName
	k=0
	kt=len(precname)
	#print(precname)
	#print(dbplist)
	#if len(precname)==len(dbplist):
		#print('test correct')
	while k<kt:
		precname[k]=precname[k]+dbplist[k]
		k=k+1
	#end go through lists and add double bond position to PrecursorName

	#begin save data in writelist
	writelist[0]=writelist[0]+mlistname
	writelist[1]=writelist[1]+precname
	writelist[2]=writelist[2]+precformula
	writelist[3]=writelist[3]+precadduct
	writelist[4]=writelist[4]+precmz
	writelist[5]=writelist[5]+precchrg
	writelist[6]=writelist[6]+prodname
	writelist[7]=writelist[7]+prodformula
	writelist[8]=writelist[8]+prodadduct
	writelist[9]=writelist[9]+prodmz
	writelist[10]=writelist[10]+prodchrg
	#end save data in writelist

	#print('All calculations for fatty acids with three double bonds are done.')
	#end save excel file
	# end add triple unsaturated fatty acids
	######################################################################################################################################################
	##############################################################TRIPLE##################################################################################







if discoverylevel==0:



	##############################################################FOUR DB#################################################################################
	######################################################################################################################################################
	# begin a 4db unsaturated fatty acids
	#print(len(dbindexlist))
	#ask=eval(input('Add fatty acids with four double bonds? Yes: 1; No: 0 | '))
	ask=1
	if ask==0:
		quit()

	nspec=9	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3  ############################ VARY ########################
	# begin create empty lists
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	dbplist=[]
	#end create empty lists
	#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
	#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
	#sheetinput=wb['transitionlist']
	ki=len(writelist[0])

	#print('Entries in excel file after bisunsaturated FA transitions are generated: %d' % ki)
	kmono=ki # List index in excel file to start writing FA transitions with four double bonds
	kfourstart=kmono
	r=ktriplestart ################################################################################################### VARY ########################
	ki=ki#+2
	while r<ki:		#go through rows of FAs with three double bonds in excel file
		# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
		cchain=0
		e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
		add=e[5]
		add=int(add)
		#print(add)
		cchain=cchain+(10*add)
		add=e[6]
		add=int(add)
		cchain=cchain+(1*add)
		maxdbp=cchain-2
		#maxdbp=
		# end determine number of C atoms in chain, define highest possible double bond position
		# begin determine double bond position of already located double bond (dbpmono)
		#if len(e)==14:
		#	dbpmono=int(e[13])
		#elif len(e)==15:
		#	dbpmono=(10*(int(e[13])))+(int(e[14]))
		#else:
		#	print('Please check source code (determine double bond position of already located double bond)')
		dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond        ######################################### CHECK ?
		# end determine double bond position of already located double bond
		nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
		csubmono=dbpmono+2
		if nsecdbp>0:
			kadd=nspec*nsecdbp#+1		#(nspec= number of products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

			e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
			f=float(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
			k=0
			while k<kadd:
				mlistname.append(e)	# copied, no change
				precchrg.append(f)	# copied, no change
				k=k+1	

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
			degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
			if degunsat==3:			####################################### VARY ################################
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(4)		####################################### VARY ################################
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne
			k=0
			while k<kadd:
				precname.append(e)	# write new precursorname, changed to FA with four db
				k=k+1

			e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
			k=0
			while k<kadd:
				adding=0
				while adding<nspec:
					precadduct.append(e) 	# precursor and all transitions for each db
					adding=adding+1
				k=k+nspec

			e=float(writelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz
			e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added double bond
			k=0
			while k<kadd:
				adding=0
				while adding<nspec:
					precmz.append(e) 	# precursor and all transitions for each db
					adding=adding+1
				k=k+nspec

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
			degunsat=int(e[8])		# begin change FA in PrecursorName to generate ProductName
			if degunsat==3:			########################################## VARY ######################################################
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(4)	########################################## VARY ######################################################
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne		# end change from degree of unsaturation of FA in PrecursorName
			if e=='Cholesterol (+[2]H7)':
				fragment='_ozone neutral gain' 		
			else:	
				k=0
				csub=csubmono
				while k<kadd:
					dbp='_n-'+str(csub)
					fragment='_precursor'
					ne=e+dbp+fragment
					prodname.append(ne)		# precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	#end save double bond position for later
					dbpi=0
					while dbpi<nspec:
						dbplist.append(dbp)
						dbpi=dbpi+1
					if preconly==0:
						dbp='_n-'+str(csub)
						fragment='_aldehyde neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde
						fragment='_criegee neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee
						fragment='_aldehyde neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde last existing double bond
						fragment='_criegee neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee last existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][1])
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][1])
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						dbl=[]	#begin save double bond position for later 						
						firstdbp=dbindexlist[r][0]
						seconddbp=dbindexlist[r][1]
						thirddbp=dbindexlist[r][2]
						dbl.append(firstdbp)
						dbl.append(seconddbp)
						dbl.append(thirddbp)
						dbl.append(csub)
						ti=0
						while ti<(nspec-1):
							dbindexlist.append(dbl)	#end save double bond position for later
							ti=ti+1
					if preconly==1:
						dbp='_n-'+str(csub)
						fragment='_dummy precursor1'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor2'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor3'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor4'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor5'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor6'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor7'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor8'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						dbl=[]	#begin save double bond position for later 						
						firstdbp=dbindexlist[r][0]
						seconddbp=dbindexlist[r][1]
						thirddbp=dbindexlist[r][2]
						dbl.append(firstdbp)
						dbl.append(seconddbp)
						dbl.append(thirddbp)
						dbl.append(csub)
						ti=0
						while ti<(nspec-1):
							dbindexlist.append(dbl)	#end save double bond position for later
							ti=ti+1
					csub=csub+1
					k=k+nspec

			e=writelist[2][r] #sheetinput.cell(row=r, column=3)
			# begin read precursor sum formula and edit product sum formula
			#print(e)
			#print(e[0])
			clist=[]
			hlist=[]
			dlist=[]
			nlist=[]
			olist=[]
			plist=[]
			ilist=[]
			i=0
			ca=0
			ha=0
			da=0
			na=0
			oa=0
			pa=0
			ia=0
			while i<len(e):
				if e[i]=='H':
					if e[i+1]=="'":
						ha=0
					else:
						ca=0
				#if e[i]=='D':
				#	ha=0		
				if e[i]=='N':
					ha=0
					da=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
				if e[i]=='P':
					ha=0
					da=0
					na=0
					oa=0
				if e[i]=='I':
					ha=0
					da=0
					na=0
					oa=0
					pa=0
				if ca==1:
					clist.append(e[i])
				if ha==1:
					hlist.append(e[i])
				if da==1:
					dlist.append(e[i])
				if na==1:
					nlist.append(e[i])
				if oa==1:
					olist.append(e[i])
				if pa==1:
					plist.append(e[i])
				if ia==1:
					ilist.append(e[i])
				if e[i]=='C':
					ca=1
				if e[i]=='H':
					if e[i+1]=="'":
						ca=0
						ha=0
						da=1
						i=i+1
					else:
						ca=0
						ha=1
				#if e[i]=='D':
				#	ca=0
				#	ha=0
				#	da=1		
				if e[i]=='N':
					ha=0
					da=0
					na=1
					if e[i+1]=='O':
						nlist.append('1')
						na=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
					oa=1
					if (i+1)<len(e):
						if e[i+1]=='P':
							olist.append('1')
							oa=0
					else:
						olist.append('1')
						oa=0					
				if e[i]=='P':
					da=0
					na=0
					oa=0
					pa=1
					if (i+1)<len(e):
						if e[i+1]=='I':
							plist.append('1')
							pa=0
					else:
						plist.append('1')
						pa=0
				if e[i]=='I':
					da=0
					na=0
					oa=0
					pa=0
					ia=1
					if i==(len(e)-1):
						ilist.append('1')
						ia=0
				i=i+1
			#print(clist)
			#print(hlist)
			#print(dlist)
			#print(nlist)
			#print(olist)
			#print(plist)

			if len(clist)==0:
				cn=0
			if len(hlist)==0:
				hn=0
			if len(dlist)==0:
				dn=0	
			if len(nlist)==0:
				nn=0
			if len(olist)==0:
				on=0
			if len(plist)==0:
				pn=0
			if len(ilist)==0:
				iodon=0
			if len(clist)==1:
				cn=int(clist[0])
			if len(clist)==2:
				cn=10*int(clist[0])+int(clist[1])
			if len(clist)==3:
				cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
			if len(hlist)==1:
				hn=int(hlist[0])
			if len(hlist)==2:
				hn=10*int(hlist[0])+int(hlist[1])
			if len(hlist)==3:
				hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
			if len(dlist)==1:
				dn=int(dlist[0])
			if len(dlist)==2:
				dn=10*int(dlist[0])+int(dlist[1])
			if len(dlist)==3:
				dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
			if len(nlist)==1:
				nn=int(nlist[0])
			if len(nlist)==2:
				nn=10*int(nlist[0])+int(nlist[1])
			if len(nlist)==3:
				nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
			if len(olist)==1:
				on=int(olist[0])
			if len(olist)==2:
				on=10*int(olist[0])+int(olist[1])
			if len(olist)==3:
				on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
			if len(plist)==1:
				pn=int(plist[0])
			if len(plist)==2:
				pn=10*int(plist[0])+int(plist[1])
			if len(plist)==3:
				pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
			if len(ilist)==1:
				iodon=int(ilist[0])
			if len(ilist)==2:
				iodon=10*int(ilist[0])+int(ilist[1])
			if len(ilist)==3:
				iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
			#print(cn)
			e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
			if e=='SPLASH':
				#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
				e=e
			else:
				k=0
				csub=csubmono
				while k<kadd:
					hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
					cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
					hnp=hnprec-(2*csub-2*3)		# applies to cleavage of first (highest n) double bond
					hnptwo=hnprec-(2*dbpmono-2*2)	# applies to cleavage of second double bond
					cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
					hnpthree=hnprec-(2*(dbindexlist[r][1])-2*1)	# applies to cleavage of third double bond
					cnpthree=cn-(dbindexlist[r][1])			# applies to cleavage of third double bond
					hnpfour=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
					cnpfour=cn-(dbindexlist[r][0])			# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
					onald=on+1
					oncrigee=on+2
					precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
					productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					t=0
					while t<nspec:
						precformula.append(precursor)
						t=t+1
					prodformula.append(precursor)
					if preconly==0:
						prodformula.append(productaldehyde)
						prodformula.append(productcrigee)
						prodformula.append(productaldehydetwo)
						prodformula.append(productcrigeetwo)
						prodformula.append(productaldehydethree)
						prodformula.append(productcrigeethree)
						prodformula.append(productaldehydefour)
						prodformula.append(productcrigeefour)	
					if preconly==1:
						t=1
						while t<nspec:
							prodformula.append(precursor)
							t=t+1
					# # product formula is saved in list for current double bond position
					precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
					prodmz.append(precursormz)
					if preconly==0:
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fourth db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fourth db, AMPP
						prodmz.append(productmz)
					if preconly==1:
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
					csub=csub+1
					k=k+nspec			
			# end read precursor sum formula and edit product sum formula

			prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
			prodchrg=precchrg 	# ProductCharge  #############

		r=r+1
		e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		prevprec=e
		e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
		prec=e
		while prec==prevprec:
			if r<(ki-1):
				r=r+1
				e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
				prevprec=e
				e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
				prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
			else:
				prec='stop_loop'
				r=r+1
		# begin check whether next species is in Fa_library

		# begin reduce transition list to species in fa_lib if desired for certain fatty acids
		if str(prec)=='stop_loop':
			discoverylevel=0
		else:
			discoverylevel=0 #dlevel
		if discoverylevel>0:
			cfaisom=str()	#get next fatty acid isomer species in list
			kcf=len(writelist[1][r])-1
			go=1
			while go==1:
				if str(writelist[1][r][kcf])=='_':
					go=0
				kcf=kcf-1
			icf=5
			while icf<(kcf+1):		#
				cfaisom=cfaisom+str(writelist[1][r][icf])
				icf=icf+1

			cutfa=0
			if discoverylevel==2:
				cutfa=1
			elif discoverylevel==1:
				if int(cfaisom[3])>3:
					cutfa=1
			print(cfaisom)
			if cutfa==1:
				if cfaisom in mostwantedlist:
					ok=1
				else:
					s=r
					go=1
					while go==1:
						if s<len(writelist[1]):
							if str(writelist[1][r])==str(writelist[1][s]):
								s=s+1
							else:
								s=s
								go=0
						else:
							go=0
					r=s
						
		# end reduce transition list to species in fa_lib if desired for certain fatty acids













		
	#print(len(mlistname))
	#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

	#begin go through lists and add double bond position to PrecursorName
	k=0
	kt=len(precname)
	#print(precname)
	#print(dbplist)
	#if len(precname)==len(dbplist):
		#print('test correct')
	while k<kt:
		precname[k]=precname[k]+dbplist[k]
		k=k+1
	#end go through lists and add double bond position to PrecursorName

	#begin save data in writelist
	writelist[0]=writelist[0]+mlistname
	writelist[1]=writelist[1]+precname
	writelist[2]=writelist[2]+precformula
	writelist[3]=writelist[3]+precadduct
	writelist[4]=writelist[4]+precmz
	writelist[5]=writelist[5]+precchrg
	writelist[6]=writelist[6]+prodname
	writelist[7]=writelist[7]+prodformula
	writelist[8]=writelist[8]+prodadduct
	writelist[9]=writelist[9]+prodmz
	writelist[10]=writelist[10]+prodchrg
	#end save data in writelist

	#print('All calculations for fatty acids with four double bonds are done.')
	#end save excel file
	# end add triple unsaturated fatty acids
	######################################################################################################################################################
	##############################################################FOUR DB#################################################################################
	##############################################################FIVE DB#################################################################################
	######################################################################################################################################################
	# begin add triple unsaturated fatty acids
	#print(len(dbindexlist))
	#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
	ask=1
	if ask==0:
		quit()

	nspec=11	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3  ############################ VARY ########################
	# begin create empty lists
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	dbplist=[]
	#end create empty lists
	#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
	#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
	#sheetinput=wb['transitionlist']
	ki=len(writelist[0])

	#print('Entries in excel file after bisunsaturated FA transitions are generated: %d' % ki)
	kmono=ki # List index in excel file to start writing FA transitions with four double bonds
	kfivestart=kmono
	r=kfourstart ################################################################################################### VARY ########################
	if workflowmode==2:
		r=ki#+2
	while r<ki:		#go through rows of FAs with three double bonds in excel file
		# begin determine number of C atoms in chain, define highest possible double bond position ###############
		cchain=0
		e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
		add=e[5]
		add=int(add)
		#print(add)
		cchain=cchain+(10*add)
		add=e[6]
		add=int(add)
		cchain=cchain+(1*add)
		maxdbp=cchain-2
		#maxdbp=
		# end determine number of C atoms in chain, define highest possible double bond position
		# begin determine double bond position of already located double bond (dbpmono)
		#if len(e)==14:
		#	dbpmono=int(e[13])
		#elif len(e)==15:
		#	dbpmono=(10*(int(e[13])))+(int(e[14]))
		#else:
		#	print('Please check source code (determine double bond position of already located double bond)')
		dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond        ######################################### CHECK ?
		# end determine double bond position of already located double bond
		nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
		csubmono=dbpmono+2
		if nsecdbp>0:
			kadd=nspec*nsecdbp#+1		#(nspec= number of products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

			e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
			f=float(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
			k=0
			while k<kadd:
				mlistname.append(e)	# copied, no change
				precchrg.append(f)	# copied, no change
				k=k+1	

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
			degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
			if degunsat==4:			####################################### VARY ################################
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(5)		####################################### VARY ################################
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne
			k=0
			while k<kadd:
				precname.append(e)	# write new precursorname, changed to FA with four db
				k=k+1

			e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
			k=0
			while k<kadd:
				adding=0
				while adding<nspec:
					precadduct.append(e) 	# precursor and all transitions for each db
					adding=adding+1
				k=k+nspec

			e=float(writelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz
			e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added double bond
			k=0
			while k<kadd:
				adding=0
				while adding<nspec:
					precmz.append(e) 	# precursor and all transitions for each db
					adding=adding+1
				k=k+nspec

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
			degunsat=int(e[8])		# begin change FA in PrecursorName to generate ProductName
			if degunsat==4:			########################################## VARY ######################################################
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(5)	########################################## VARY ######################################################
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne		# end change from degree of unsaturation of FA in PrecursorName
			if e=='Cholesterol (+[2]H7)':
				fragment='_ozone neutral gain' 		
			else:	
				k=0
				csub=csubmono
				while k<kadd:
					dbp='_n-'+str(csub)
					fragment='_precursor'
					ne=e+dbp+fragment
					prodname.append(ne)		# precursor
					dbl=[]	#begin save double bond positions for later 	###################### CHECK !						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	#end save double bond positions for later
					dbpi=0
					while dbpi<nspec:
						dbplist.append(dbp)
						dbpi=dbpi+1
					if preconly==0:
						dbp='_n-'+str(csub)
						fragment='_aldehyde neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde
						fragment='_criegee neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee
						fragment='_aldehyde neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde last existing double bond
						fragment='_criegee neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee last existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][2])
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][2])
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						dbl=[]	#begin save double bond position for later 						
						firstdbp=dbindexlist[r][0]
						seconddbp=dbindexlist[r][1]
						thirddbp=dbindexlist[r][2]
						fourthdbp=dbindexlist[r][3]
						dbl.append(firstdbp)
						dbl.append(seconddbp)
						dbl.append(thirddbp)
						dbl.append(fourthdbp)
						dbl.append(csub)
						ti=0
						while ti<(nspec-1):
							dbindexlist.append(dbl)	#end save double bond position for later
							ti=ti+1
					if preconly==1:
						dbp='_n-'+str(csub)
						fragment='_dummy precursor1'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor2'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor3'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor4'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor5'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor6'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor7'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor8'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor9'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor10'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						dbl=[]	#begin save double bond position for later 						
						firstdbp=dbindexlist[r][0]
						seconddbp=dbindexlist[r][1]
						thirddbp=dbindexlist[r][2]
						fourthdbp=dbindexlist[r][3]
						dbl.append(firstdbp)
						dbl.append(seconddbp)
						dbl.append(thirddbp)
						dbl.append(fourthdbp)
						dbl.append(csub)
						ti=0
						while ti<(nspec-1):
							dbindexlist.append(dbl)	#end save double bond position for later
							ti=ti+1
					csub=csub+1
					k=k+nspec

			e=writelist[2][r] #sheetinput.cell(row=r, column=3)
			# begin read precursor sum formula and edit product sum formula
			#print(e)
			#print(e[0])
			clist=[]
			hlist=[]
			dlist=[]
			nlist=[]
			olist=[]
			plist=[]
			ilist=[]
			i=0
			ca=0
			ha=0
			da=0
			na=0
			oa=0
			pa=0
			ia=0
			while i<len(e):
				if e[i]=='H':
					if e[i+1]=="'":
						ha=0
					else:
						ca=0
				#if e[i]=='D':
				#	ha=0		
				if e[i]=='N':
					ha=0
					da=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
				if e[i]=='P':
					ha=0
					da=0
					na=0
					oa=0
				if e[i]=='I':
					ha=0
					da=0
					na=0
					oa=0
					pa=0
					ia=0
				if ca==1:
					clist.append(e[i])
				if ha==1:
					hlist.append(e[i])
				if da==1:
					dlist.append(e[i])
				if na==1:
					nlist.append(e[i])
				if oa==1:
					olist.append(e[i])
				if pa==1:
					plist.append(e[i])
				if ia==1:
					ilist.append(e[i])
				if e[i]=='C':
					ca=1
				if e[i]=='H':
					if e[i+1]=="'":
						ca=0
						ha=0
						da=1
						i=i+1
					else:
						ca=0
						ha=1
				#if e[i]=='D':
				#	ca=0
				#	ha=0
				#	da=1		
				if e[i]=='N':
					ha=0
					da=0
					na=1
					if e[i+1]=='O':
						nlist.append('1')
						na=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
					oa=1
					if (i+1)<len(e):
						if e[i+1]=='P':
							olist.append('1')
							oa=0
					else:
						olist.append('1')
						oa=0					
				if e[i]=='P':
					da=0
					na=0
					oa=0
					pa=1
					if (i+1)<len(e):
						if e[i+1]=='I':
							plist.append('1')
							pa=0
					else:
						plist.append('1')
						pa=0
				if e[i]=='I':
					da=0
					na=0
					oa=0
					pa=0
					ia=1
					if i==(len(e)-1):
						ilist.append('1')
						ia=0
				i=i+1
			#print(clist)
			#print(hlist)
			#print(dlist)
			#print(nlist)
			#print(olist)
			#print(plist)

			if len(clist)==0:
				cn=0
			if len(hlist)==0:
				hn=0
			if len(dlist)==0:
				dn=0	
			if len(nlist)==0:
				nn=0
			if len(olist)==0:
				on=0
			if len(plist)==0:
				pn=0
			if len(ilist)==0:
				iodon=0
			if len(clist)==1:
				cn=int(clist[0])
			if len(clist)==2:
				cn=10*int(clist[0])+int(clist[1])
			if len(clist)==3:
				cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
			if len(hlist)==1:
				hn=int(hlist[0])
			if len(hlist)==2:
				hn=10*int(hlist[0])+int(hlist[1])
			if len(hlist)==3:
				hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
			if len(dlist)==1:
				dn=int(dlist[0])
			if len(dlist)==2:
				dn=10*int(dlist[0])+int(dlist[1])
			if len(dlist)==3:
				dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
			if len(nlist)==1:
				nn=int(nlist[0])
			if len(nlist)==2:
				nn=10*int(nlist[0])+int(nlist[1])
			if len(nlist)==3:
				nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
			if len(olist)==1:
				on=int(olist[0])
			if len(olist)==2:
				on=10*int(olist[0])+int(olist[1])
			if len(olist)==3:
				on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
			if len(plist)==1:
				pn=int(plist[0])
			if len(plist)==2:
				pn=10*int(plist[0])+int(plist[1])
			if len(plist)==3:
				pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
			if len(ilist)==1:
				iodon=int(ilist[0])
			if len(ilist)==2:
				iodon=10*int(ilist[0])+int(ilist[1])
			if len(ilist)==3:
				iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
			#print(cn)
			e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
			if e=='SPLASH':
				#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
				e=e
			else:
				k=0
				csub=csubmono
				while k<kadd:
					hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
					cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
					hnp=hnprec-(2*csub-2*4)		# applies to cleavage of first (highest n) double bond
					hnptwo=hnprec-(2*dbpmono-2*3)	# applies to cleavage of second double bond
					cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
					hnpthree=hnprec-(2*(dbindexlist[r][2])-2*2)	# applies to cleavage of third double bond
					cnpthree=cn-(dbindexlist[r][2])			# applies to cleavage of third double bond
					hnpfour=hnprec-(2*(dbindexlist[r][1])-2*1)	# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
					cnpfour=cn-(dbindexlist[r][1])			# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
					hnpfive=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
					cnpfive=cn-(dbindexlist[r][0])			# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
					onald=on+1
					oncrigee=on+2
					precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
					productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
					productcrigeefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
					t=0
					while t<nspec:
						precformula.append(precursor)
						t=t+1
					prodformula.append(precursor)
					if preconly==0:
						prodformula.append(productaldehyde)
						prodformula.append(productcrigee)
						prodformula.append(productaldehydetwo)
						prodformula.append(productcrigeetwo)
						prodformula.append(productaldehydethree)
						prodformula.append(productcrigeethree)
						prodformula.append(productaldehydefour)
						prodformula.append(productcrigeefour)	
						prodformula.append(productaldehydefive) ############################################# CHECK !!!!!!!!
						prodformula.append(productcrigeefive) ############################################# CHECK !!!!!!!!
					if preconly==1:
						t=1
						while t<nspec:
							prodformula.append(precursor)
							t=t+1
					# # product formula is saved in list for current double bond position
					precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
					prodmz.append(precursormz)
					if preconly==0:
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fourth db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fourth db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fifth db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fifth db, AMPP
						prodmz.append(productmz)
					if preconly==1:
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
					csub=csub+1
					k=k+nspec			
			# end read precursor sum formula and edit product sum formula

			prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
			prodchrg=precchrg 	# ProductCharge  #############

		r=r+1
		e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		prevprec=e
		e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
		prec=e
		while prec==prevprec:
			if r<(ki-1):
				r=r+1
				e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
				prevprec=e
				e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
				prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
			else:
				prec='stop_loop'
				r=r+1
		
	#print(len(mlistname))
	#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

	#begin go through lists and add double bond position to PrecursorName
	k=0
	kt=len(precname)
	#print(precname)
	#print(dbplist)
	#if len(precname)==len(dbplist):
		#print('test correct')
	while k<kt:
		precname[k]=precname[k]+dbplist[k]
		k=k+1
	#end go through lists and add double bond position to PrecursorName

	#begin save data in writelist
	if workflowmode==1:
		writelist[0]=writelist[0]+mlistname
		writelist[1]=writelist[1]+precname
		writelist[2]=writelist[2]+precformula
		writelist[3]=writelist[3]+precadduct
		writelist[4]=writelist[4]+precmz
		writelist[5]=writelist[5]+precchrg
		writelist[6]=writelist[6]+prodname
		writelist[7]=writelist[7]+prodformula
		writelist[8]=writelist[8]+prodadduct
		writelist[9]=writelist[9]+prodmz
		writelist[10]=writelist[10]+prodchrg
	#end save data in writelist

	#print('All calculations for fatty acids with five double bonds are done.')
	#end save excel file
	# end add triple unsaturated fatty acids
	######################################################################################################################################################
	##############################################################FIVE DB#################################################################################

	##############################################################SIX DB#################################################################################
	######################################################################################################################################################
	# begin add triple unsaturated fatty acids
	#print(len(dbindexlist))
	#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
	ask=0
	if ask==1:
		quit()

	nspec=13	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3  ############################ VARY ########################
	# begin create empty lists
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	dbplist=[]
	#end create empty lists
	#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
	#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
	#sheetinput=wb['transitionlist']
	ki=len(writelist[0])

	#print('Entries in excel file after bisunsaturated FA transitions are generated: %d' % ki)
	kmono=ki # List index in excel file to start writing FA transitions with four double bonds
	ksixstart=kmono
	r=kfivestart ################################################################################################### VARY ########################
	if workflowmode==2:
		r=ki#
	while r<ki:		#go through rows of FAs with three double bonds in excel file
		# begin determine number of C atoms in chain, define highest possible double bond position ###############
		cchain=0
		e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
		add=e[5]
		add=int(add)
		#print(add)
		cchain=cchain+(10*add)
		add=e[6]
		add=int(add)
		cchain=cchain+(1*add)
		maxdbp=cchain-2
		#maxdbp=
		# end determine number of C atoms in chain, define highest possible double bond position
		# begin determine double bond position of already located double bond (dbpmono)
		#if len(e)==14:
		#	dbpmono=int(e[13])
		#elif len(e)==15:
		#	dbpmono=(10*(int(e[13])))+(int(e[14]))
		#else:
		#	print('Please check source code (determine double bond position of already located double bond)')
		dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond, highest n        ######################################### CHECK ?
		# end determine double bond position of already located double bond
		nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
		csubmono=dbpmono+2
		if nsecdbp>0:
			kadd=nspec*nsecdbp#+1		#(nspec= number of products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

			e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
			f=float(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
			k=0
			while k<kadd:
				mlistname.append(e)	# copied, no change
				precchrg.append(f)	# copied, no change
				k=k+1	

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
			degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
			if degunsat==5:			####################################### VARY ################################
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(6)		####################################### VARY ################################
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne
			k=0
			while k<kadd:
				precname.append(e)	# write new precursorname, changed to FA with six db
				k=k+1

			e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
			k=0
			while k<kadd:
				adding=0
				while adding<nspec:
					precadduct.append(e) 	# precursor and all transitions for each db
					adding=adding+1
				k=k+nspec

			e=float(writelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz
			e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added double bond
			k=0
			while k<kadd:
				adding=0
				while adding<nspec:
					precmz.append(e) 	# precursor and all transitions for each db
					adding=adding+1
				k=k+nspec

			e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
			degunsat=int(e[8])		# begin change FA in PrecursorName to generate ProductName
			if degunsat==5:			########################################## VARY ######################################################
				t=0
				ne=str()
				while t<8:
					l=e[t]
					ne=ne+str(l)
					t=t+1
				ne=ne+str(6)	########################################## VARY ######################################################
				t=9
				while t<len(e):
					l=e[t]
					ne=ne+str(l)
					t=t+1
				e=ne		# end change from degree of unsaturation of FA in PrecursorName
			if e=='Cholesterol (+[2]H7)':
				fragment='_ozone neutral gain' 		
			else:	
				k=0
				csub=csubmono
				while k<kadd:
					dbp='_n-'+str(csub)
					fragment='_precursor'
					ne=e+dbp+fragment
					prodname.append(ne)		# precursor
					dbl=[]	#begin save double bond positions for later 	###################### CHECK !						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					fifthdbp=dbindexlist[r][4]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(fifthdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	#end save double bond positions for later
					dbpi=0
					while dbpi<nspec:
						dbplist.append(dbp)
						dbpi=dbpi+1
					if preconly==0:
						dbp='_n-'+str(csub)
						fragment='_aldehyde neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde
						fragment='_criegee neutral loss from n-'+str(csub)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee
						fragment='_aldehyde neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde last existing double bond
						fragment='_criegee neutral loss from n-'+str(dbpmono)
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee last existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][3])
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][3])
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][2])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][2])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# aldehyde previous double bond
						fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
						ne=e+dbp+fragment
						prodname.append(ne)		# crigee previous existing double bond
						dbl=[]	#begin save double bond position for later 						
						firstdbp=dbindexlist[r][0]
						seconddbp=dbindexlist[r][1]
						thirddbp=dbindexlist[r][2]
						fourthdbp=dbindexlist[r][3]
						fifthdbp=dbindexlist[r][4]
						dbl.append(firstdbp)
						dbl.append(seconddbp)
						dbl.append(thirddbp)
						dbl.append(fourthdbp)
						dbl.append(fifthdbp)
						dbl.append(csub)
						ti=0
						while ti<(nspec-1):
							dbindexlist.append(dbl)	#end save double bond position for later
							ti=ti+1
					if preconly==1:
						dbp='_n-'+str(csub)
						fragment='_dummy precursor1'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor2'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor3'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor4'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor5'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor6'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor7'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor8'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor9'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor10'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor11'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						fragment='_dummy precursor12'
						ne=e+dbp+fragment
						prodname.append(ne)		# dummy precursor
						dbl=[]	#begin save double bond position for later 						
						firstdbp=dbindexlist[r][0]
						seconddbp=dbindexlist[r][1]
						thirddbp=dbindexlist[r][2]
						fourthdbp=dbindexlist[r][3]
						fifthdbp=dbindexlist[r][4]
						dbl.append(firstdbp)
						dbl.append(seconddbp)
						dbl.append(thirddbp)
						dbl.append(fourthdbp)
						dbl.append(fifthdbp)
						dbl.append(csub)
						ti=0
						while ti<(nspec-1):
							dbindexlist.append(dbl)	#end save double bond position for later
							ti=ti+1
					csub=csub+1
					k=k+nspec

			e=writelist[2][r] #sheetinput.cell(row=r, column=3)
			# begin read precursor sum formula and edit product sum formula
			#print(e)
			#print(e[0])
			clist=[]
			hlist=[]
			dlist=[]
			nlist=[]
			olist=[]
			plist=[]
			ilist=[]
			i=0
			ca=0
			ha=0
			da=0
			na=0
			oa=0
			pa=0
			ia=0
			while i<len(e):
				if e[i]=='H':
					if e[i+1]=="'":
						ha=0
					else:
						ca=0
				#if e[i]=='D':
				#	ha=0		
				if e[i]=='N':
					ha=0
					da=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
				if e[i]=='P':
					ha=0
					da=0
					na=0
					oa=0
				if e[i]=='I':
					ha=0
					da=0
					na=0
					oa=0
					pa=0
				if ca==1:
					clist.append(e[i])
				if ha==1:
					hlist.append(e[i])
				if da==1:
					dlist.append(e[i])
				if na==1:
					nlist.append(e[i])
				if oa==1:
					olist.append(e[i])
				if pa==1:
					plist.append(e[i])
				if ia==1:
					ilist.append(e[i])
				if e[i]=='C':
					ca=1
				if e[i]=='H':
					if e[i+1]=="'":
						ca=0
						ha=0
						da=1
						i=i+1
					else:
						ca=0
						ha=1
				#if e[i]=='D':
				#	ca=0
				#	ha=0
				#	da=1		
				if e[i]=='N':
					ha=0
					da=0
					na=1
					if e[i+1]=='O':
						nlist.append('1')
						na=0
				if e[i]=='O':
					ha=0
					da=0
					na=0
					oa=1
					if (i+1)<len(e):
						if e[i+1]=='P':
							olist.append('1')
							oa=0
					else:
						olist.append('1')
						oa=0					
				if e[i]=='P':
					da=0
					na=0
					oa=0
					pa=1
					if (i+1)<len(e):
						if e[i+1]=='I':
							plist.append('1')
							pa=0
					else:
						plist.append('1')
						pa=0
				if e[i]=='I':
					da=0
					na=0
					oa=0
					pa=0
					ia=1
					if i==(len(e)-1):
						ilist.append('1')
						ia=0
				i=i+1
			#print(clist)
			#print(hlist)
			#print(dlist)
			#print(nlist)
			#print(olist)
			#print(plist)

			if len(clist)==0:
				cn=0
			if len(hlist)==0:
				hn=0
			if len(dlist)==0:
				dn=0	
			if len(nlist)==0:
				nn=0
			if len(olist)==0:
				on=0
			if len(plist)==0:
				pn=0
			if len(ilist)==0:
				iodon=0
			if len(clist)==1:
				cn=int(clist[0])
			if len(clist)==2:
				cn=10*int(clist[0])+int(clist[1])
			if len(clist)==3:
				cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
			if len(hlist)==1:
				hn=int(hlist[0])
			if len(hlist)==2:
				hn=10*int(hlist[0])+int(hlist[1])
			if len(hlist)==3:
				hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
			if len(dlist)==1:
				dn=int(dlist[0])
			if len(dlist)==2:
				dn=10*int(dlist[0])+int(dlist[1])
			if len(dlist)==3:
				dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
			if len(nlist)==1:
				nn=int(nlist[0])
			if len(nlist)==2:
				nn=10*int(nlist[0])+int(nlist[1])
			if len(nlist)==3:
				nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
			if len(olist)==1:
				on=int(olist[0])
			if len(olist)==2:
				on=10*int(olist[0])+int(olist[1])
			if len(olist)==3:
				on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
			if len(plist)==1:
				pn=int(plist[0])
			if len(plist)==2:
				pn=10*int(plist[0])+int(plist[1])
			if len(plist)==3:
				pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
			if len(ilist)==1:
				iodon=int(ilist[0])
			if len(ilist)==2:
				iodon=10*int(ilist[0])+int(ilist[1])
			if len(ilist)==3:
				iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
			#print(cn)
			e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
			if e=='SPLASH':
				#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
				e=e
			else:
				k=0
				csub=csubmono
				while k<kadd:
					hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
					cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
					hnp=hnprec-(2*csub-2*5)		# applies to cleavage of first (highest n) double bond
					hnptwo=hnprec-(2*dbpmono-2*4)	# applies to cleavage of second double bond
					cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
					hnpthree=hnprec-(2*(dbindexlist[r][3])-2*3)	# applies to cleavage of third double bond
					cnpthree=cn-(dbindexlist[r][3])			# applies to cleavage of third double bond
					hnpfour=hnprec-(2*(dbindexlist[r][2])-2*2)	# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
					cnpfour=cn-(dbindexlist[r][2])			# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
					hnpfive=hnprec-(2*(dbindexlist[r][1])-2*1)	# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
					cnpfive=cn-(dbindexlist[r][1])			# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
					hnpsix=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
					cnpsix=cn-(dbindexlist[r][0])			# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
					onald=on+1
					oncrigee=on+2
					precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
					productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
					productcrigeefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
					productaldehydefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
					productcrigeefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
					productaldehydesix='C'+str(cnpsix)+'H'+str(hnpsix)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
					productcrigeesix='C'+str(cnpsix)+'H'+str(hnpsix)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
					t=0
					while t<nspec:
						precformula.append(precursor)
						t=t+1
					prodformula.append(precursor)
					if preconly==0:
						prodformula.append(productaldehyde)
						prodformula.append(productcrigee)
						prodformula.append(productaldehydetwo)
						prodformula.append(productcrigeetwo)
						prodformula.append(productaldehydethree)
						prodformula.append(productcrigeethree)
						prodformula.append(productaldehydefour)
						prodformula.append(productcrigeefour)	
						prodformula.append(productaldehydefive) ############################################# CHECK !!!!!!!!
						prodformula.append(productcrigeefive) ############################################# CHECK !!!!!!!!
						prodformula.append(productaldehydesix) ############################################# CHECK !!!!!!!!
						prodformula.append(productcrigeesix) ############################################# CHECK !!!!!!!!
					if preconly==1:
						t=1
						while t<nspec:
							prodformula.append(precursor)
							t=t+1
					# # product formula is saved in list for current double bond position
					precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
					prodmz.append(precursormz)
					if preconly==0:
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fourth db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fourth db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fifth db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fifth db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpsix)+imass[1]*dn+imass[2]*cnpsix+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, sixth db, AMPP
						prodmz.append(productmz)
						productmz=imass[0]*(hnpsix)+imass[1]*dn+imass[2]*cnpsix+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, sixth db, AMPP
						prodmz.append(productmz)
					if preconly==1:
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
						prodmz.append(precursormz)
					csub=csub+1
					k=k+nspec			
			# end read precursor sum formula and edit product sum formula

			prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
			prodchrg=precchrg 	# ProductCharge  #############

		r=r+1
		e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		prevprec=e
		e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
		prec=e
		while prec==prevprec:
			if r<(ki-1):
				r=r+1
				e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
				prevprec=e
				e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
				prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
			else:
				prec='stop_loop'
				r=r+1
		
	#print(len(mlistname))
	#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

	#begin go through lists and add double bond position to PrecursorName
	k=0
	kt=len(precname)
	#print(precname)
	#print(dbplist)
	#if len(precname)==len(dbplist):
		#print('test correct')
	while k<kt:
		precname[k]=precname[k]+dbplist[k]
		k=k+1
	#end go through lists and add double bond position to PrecursorName

	#begin save data in writelist
	if workflowmode==1:
		writelist[0]=writelist[0]+mlistname
		writelist[1]=writelist[1]+precname
		writelist[2]=writelist[2]+precformula
		writelist[3]=writelist[3]+precadduct
		writelist[4]=writelist[4]+precmz
		writelist[5]=writelist[5]+precchrg
		writelist[6]=writelist[6]+prodname
		writelist[7]=writelist[7]+prodformula
		writelist[8]=writelist[8]+prodadduct
		writelist[9]=writelist[9]+prodmz
		writelist[10]=writelist[10]+prodchrg
	#end save data in writelist

	#print('All calculations for fatty acids with six double bonds are done.')
	#end save excel file
	# end add triple unsaturated fatty acids
	#print(len(writelist[0]))
	#print(len(dbindexlist))
	######################################################################################################################################################
	##############################################################END ADD SIX DB##########################################################################
	######################################################################################################################################################
else:
	# begin add FA species in FA_library with more than 3 double bonds, if discoverylevel==1
	mlistname=[] # copied, no change
	precname=[]	# ok
	precformula=[] # ok
	precadduct=[] # copied, no change
	precmz=[]	# ok
	precchrg=[] # copied, no change
	prodname=[]	# ok
	prodformula=[] # 
	prodadduct=[] # copied, no change
	prodmz=[]	#
	prodchrg=[] # copied, no change

	if discoverylevel==1:
		dbcut=3
	else:
		dbcut=0
		toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
		writelist=[]
		writelist.append(mlistname)
		writelist.append(precname)
		writelist.append(precformula)
		writelist.append(precadduct)
		writelist.append(precmz)
		writelist.append(precchrg)
		writelist.append(prodname)
		writelist.append(prodformula)
		writelist.append(prodadduct)
		writelist.append(prodmz)
		writelist.append(prodchrg)
	# go through mostwantedlist
	mwi=0
	while mwi<len(mostwantedlist):
		if int(mostwantedlist[mwi][3])>dbcut:
			kadd=1+2*int(mostwantedlist[mwi][3])
			e=str(fourlettcode)+'_FA' #writelist[0][0]
			f=1 #writelist[5][0]
			g='[M]1+' #writelist[3][0]
			cpn=str(fourlettcode)+'_'+str(mostwantedlist[mwi]) #precursorname
			cnh=int(hderiv)+(((10*int(mostwantedlist[mwi][0])+int(mostwantedlist[mwi][1]))-2)*2)+3-(2*int(mostwantedlist[mwi][3])) # H in derivatized FA
			cnc=int(cderiv)+(10*int(mostwantedlist[mwi][0])+int(mostwantedlist[mwi][1])) # C in derivatized FA
			cnn=int(nderiv) # N in derivatized FA
			cno=int(oderiv)+1 # O in derivatized FA
			cpf='C'+str(cnc)+'H'+str(cnh)+'N'+str(cnn)+'O'+str(cno)+'I'+str(int(ideriv))
			cprecmz=imass[0]*(cnh)+imass[1]*dderiv+imass[2]*cnc+imass[3]*cnn+imass[4]*cno+imass[5]*pderiv+imass[10]*ideriv-imass[8]	#precursorMz
			# begin build dbindexlist
			#print(mostwantedlist[mwi])
			cdbindexlist=[]
			idb=5
			dbcount=0
			while idb<int(len(mostwantedlist[mwi])-2):
				if str(mostwantedlist[mwi][idb-2])=='n':
					if str(mostwantedlist[mwi][idb+2])=='_':
						cdbi=10*int(mostwantedlist[mwi][idb])+int(mostwantedlist[mwi][idb+1])
						cdbindexlist.append(cdbi)
						dbcount=dbcount+1
					elif str(mostwantedlist[mwi][idb+1])=='_':
						cdbi=int(mostwantedlist[mwi][idb])
						cdbindexlist.append(cdbi)
						dbcount=dbcount+1
					idb=idb+1
				elif idb==int(len(mostwantedlist[mwi])-4):
					if str(mostwantedlist[mwi][idb])=='n':
						cdbi=10*int(mostwantedlist[mwi][idb+2])+int(mostwantedlist[mwi][idb+3])
						cdbindexlist.append(cdbi)
						dbcount=dbcount+1
					idb=idb+1
				elif idb==int(len(mostwantedlist[mwi])-3):
					if str(mostwantedlist[mwi][idb])=='n':
						cdbi=int(mostwantedlist[mwi][idb+2])
						cdbindexlist.append(cdbi)
						dbcount=dbcount+1
					idb=idb+1
				else:
					idb=idb+1
			# end build cdbindexlist
			k=0
			tcount=0
			while k<kadd:
				mlistname.append(e)	# copied, no change
				precname.append(cpn)
				precformula.append(cpf)
				precadduct.append(g) # copied, no change
				precmz.append(cprecmz)
				if tcount==0:
					prodmz.append(cprecmz)
					cprodnm=str(cpn)+'_precursor'
					prodname.append(cprodnm)
					prodformula.append(cpf)
				else:
					chlist=[1,3,5,7,9,11,13,15]
					if tcount in chlist:
						#print(cpn)
						#print(kadd)
						#print(k)
						#print(tcount)
						#print(cdbindexlist)
						#print(cnh)
						cnpos=cdbindexlist[int(len(cdbindexlist)-(((tcount-1)/2)+1))]
						#print(cnpos)
						cfragnm=str(cpn)+'_aldehyde neutral loss from n-'+str(cnpos)
						cnhald=int(cnh-((cnpos*2)-(2*((len(cdbindexlist)-1)-((tcount-1)/2)))))
						#print(cnhald)
						cfragf='C'+str(cnc-cnpos)+'H'+str(cnhald)+'N'+str(cnn)+'O'+str(cno+1)+'I'+str(int(ideriv))
						cfragmz=imass[0]*(cnhald)+imass[1]*dderiv+imass[2]*(cnc-cnpos)+imass[3]*cnn+imass[4]*(cno+1)+imass[5]*pderiv+imass[10]*ideriv-imass[8]	#prodMz
						#print(cfragnm)
						#print(cfragf)
						#print(cfragmz)
						#if cpn=='AMPP_18:1_n-7':
						#	quit()
					else:
						cnpos=cdbindexlist[int(len(cdbindexlist)-((tcount)/2))]
						cfragnm=str(cpn)+'_criegee neutral loss from n-'+str(cnpos)
						cfragf='C'+str(cnc-cnpos)+'H'+str(int(cnh-((cnpos*2)-(2*(len(cdbindexlist)-((tcount)/2))))))+'N'+str(cnn)+'O'+str(cno+2)+'I'+str(int(ideriv))
						cfragmz=imass[0]*(cnh-((cnpos*2)-(2*(len(cdbindexlist)-((tcount)/2)))))+imass[1]*dderiv+imass[2]*(cnc-cnpos)+imass[3]*cnn+imass[4]*(cno+2)+imass[5]*pderiv+imass[10]*ideriv-imass[8]	#prodMz
					prodname.append(cfragnm)
					prodformula.append(cfragf)
					prodmz.append(cfragmz)
				precchrg.append(f)	# copied, no change
				prodadduct.append(g) # copied, no change
				prodchrg.append(f) # copied, no change
				k=k+1
				tcount=tcount+1
			
		mwi=mwi+1
		
	writelist[0]=writelist[0]+mlistname
	writelist[1]=writelist[1]+precname
	writelist[2]=writelist[2]+precformula
	writelist[3]=writelist[3]+precadduct
	writelist[4]=writelist[4]+precmz
	writelist[5]=writelist[5]+precchrg
	writelist[6]=writelist[6]+prodname
	writelist[7]=writelist[7]+prodformula
	writelist[8]=writelist[8]+prodadduct
	writelist[9]=writelist[9]+prodmz
	writelist[10]=writelist[10]+prodchrg


	# end add FA species in FA_library with more than 3 double bonds

##########################################################VIRTUAL#PRECURSOR###########################################################################
######################################################################################################################################################
# begin introduce virtual precursor
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]		# content in lists is deleted to free up memory
# begin introduce virtual precursor
ki=len(writelist[0])

#print('Entries in excel file after all FA transitions are generated: %d' % ki)
virtualprecformula=[]
virtualprecmz=[]

vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]

r=0
ki=ki
while r<ki:
	e=str(writelist[2][r])#sheetinput.cell(row=r, column=3)	# PrecursorFormula
	ve=e+'Xe'
	virtualprecformula.append(ve) # virtual precursorFormula
	e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
	ve=e+(imass[9])		#add one Xe atom to generate virtual precursor
	virtualprecmz.append(ve)	# virtual precursormz
	e=str(writelist[0][r])#sheetinput.cell(row=r, column=1)
	vmlistname.append(e)
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)
	vprecname.append(e)
	e=str(writelist[3][r]) #sheetinput.cell(row=r, column=4)
	vprecadduct.append(e)
	e=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)
	vprecchrg.append(e)
	e=str(writelist[6][r]) #sheetinput.cell(row=r, column=7)
	vprodname.append(e)
	e=str(writelist[7][r]) #sheetinput.cell(row=r, column=8)
	vprodformula.append(e)
	e=str(writelist[8][r]) #sheetinput.cell(row=r, column=9)
	vprodadduct.append(e)
	e=float(writelist[9][r]) #sheetinput.cell(row=r, column=10)
	vprodmz.append(e)
	e=int(writelist[10][r]) #sheetinput.cell(row=r, column=11)
	vprodchrg.append(e)
	r=r+1
# begin save excel file with virtual precursor as csv file
toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
vwritelist=[]
vwritelist.append(vmlistname)
vwritelist.append(vprecname)
vwritelist.append(virtualprecformula)
vwritelist.append(vprecadduct)
vwritelist.append(virtualprecmz)
vwritelist.append(vprecchrg)
vwritelist.append(vprodname)
vwritelist.append(vprodformula)
vwritelist.append(vprodadduct)
vwritelist.append(vprodmz)
vwritelist.append(vprodchrg)
terminate=0			############################################ check ok
if terminate==1:
	print('vwritelist created')
	transitionresultsdf=pd.DataFrame(vwritelist).transpose()
	print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10]]
	print('Transposed and DataFrame created')
	transitionresultsdf.to_csv('jpmlipidomics_vpw19_virtual_precursor.csv', index=False)
	print('Transition list saved as jpmlipidomics_vpw19_virtual_precursor.csv')  
# end save file with virtual precursor as csv file
#quit()
#lnn=len(vwritelist[0])
#print(lnn)
#print('Saving data.')
#print('len(vmlistname)')
#print(len(vmlistname))
virtualprecformula=[]
virtualprecmz=[]
vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[] # test to delete lists to save up memory

print('Transition list is modified with virtual precursor [M + Xe].')
#print(vwritelist[1])
#krows=len(vwritelist[1])
#print(krows)
#quit()
######################################################################################################################################################
######################################################################################################################################################
#################################################################################################################################################
############################# BEGIN REDUCE TO ENTRIES OF reassigned targets (PRECURSOR) AND EXPAND WITH PRECURSOR EXPLICIT RETENTION TIMES #######################
#################################################################################################################################################
#terminate=eval(input('Reduce transition list to entries for which precursor is found and expand with explicit retention time of the found precursor? Yes: 1 No: 0 ::'))
terminate=0 ############################################ check 
if terminate==1:
	quit()
ki=len(vwritelist[0])
#print('len(vwritelist[0])')
#print(ki)
kir=len(mztargetlist)
#print('len(mztargetlist)')
#print(kir)
vmlistname=[]
vprecname=[]
virtualprecformula=[]
vprecadduct=[]
virtualprecmz=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]
precrt=[]
rtwindow=[]
r=0
pos=1
while r<ki:
	rn=r
	precrlist=[]
	rexrtlist=[]
	pos=0
	while pos==0:
		# begin determine block length (start with r of block of species with same precursor mass, end with s)
		e=vwritelist[4][r] #(row=r, column=5)	# PrecursorMz		# begin determine which row to start (r) and to end (s)
		s=r+1
		st=0
		while st<1:
			if s<ki:
				ne=vwritelist[4][s] #(row=s, column=5)	# PrecursorMz
			else:
				ne='stop loop'
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		# begin determine if precursor of current species is found
		e=str(vwritelist[1][r]) #(row=r, column=2)
		precspecies=e[5]+e[6]+e[7]+e[8]
		precspecies=str(precspecies)	# precursor species in full list
		precspeciesmz=float(vwritelist[4][r])-float(imass[9])
		pos=0
		rr=0
		while rr<(kir+0):
			precrspeciesmz=mztargetlist[rr]	# precursorspecies in list with confirmed species from precursor only analysis
			if abs(float(precspeciesmz)-float(precrspeciesmz))<0.01:
				pos=1					#### the precursor of the current species in full list was found
				precrlist.append(rr)	#### list with precursorspecies that were identified 
				rexrt=float(rttargetlist[rr])
				rexrtlist.append(rexrt)	#### list with explicit retention times associated to precursorspecies that were identified
			rr=rr+1
		# end determine if precursor of current species is found
		if pos==0:
			if (s+1)<ki:
				r=s+1
				rn=r
			else:
				#print('(s+1)>(ki-1)')
				r=s
				pos=1
	# block identified for which precursor is found
	if (r+1)>ki:
		pos=0
		r=r+2
	
	if discoverylimitation==1:
		if int(str(vwritelist[1][r][5])+str(vwritelist[1][r][6]))<cminlimit:
			godi=0
		elif int(str(vwritelist[1][r][5])+str(vwritelist[1][r][6]))>cmaxlimit:
			godi=0
		else:
			godi=1
	else:
		godi=1
	if godi==1:
		pos=pos
	else:
		pos=0

	if pos==1:
		#print('Block found')
		#print(precspecies)
		#print(precrspecies)
		#print(r)
		#print(s)
		#print(len(vmlistname))
		#begin define expansion
		# begin determine block length (start with rr of block of species with same precursor mass, end with sr)

		#rr=precrlist[0]
		#sr=precrlist[(len(precrlist)-1)]

		#print(rr)
		#print(sr)
		#e=sheetresults.cell(row=rr, column=5)	# PrecursorMz		# begin determine which row to start (rr) and to end (sr)
		#e=e.value
		#sr=rr+1
		#st=0
		#while st<1:
		#	ne=sheetinput.cell(row=sr, column=5)	# PrecursorMz
		#	ne=ne.value
		#	if ne==e:
		#		sr=sr+1
		#		st=0
		#	else:
		#		sr=sr-1
		#		st=1		# end determine sr
		#end define expansion
		kr=0
		while kr<(len(rexrtlist)):
			rexrt=rexrtlist[kr]
			rexrt=str(rexrt)
			r=rn
			while r<(s+1):
				#e=writelist[0][r] ## mlistname	# begin append rows of suitable species to lists for later saving in excel output
				#vmlistname.append(e)
				#e=writelist[1][r] ## precname	
				#vprecname.append(e)
				#print(len(virtualprecformula))
				e=str(vwritelist[0][r]) #sheetinput.cell(row=r, column=1)
				vmlistname.append(e)
				e=str(vwritelist[1][r]) #sheetinput.cell(row=r, column=2)
				cm='_'
				cm=str(cm)
				rexrt=float(rexrt)
				rexrtstr=str(round(rexrt, 2))		#
				ee=e+cm+rexrtstr			#
				vprecname.append(ee)
				precrt.append(rexrt)	##################################
				ve=str(vwritelist[2][r]) #sheetinput.cell(row=r, column=3)	# PrecursorFormula
				virtualprecformula.append(ve) # virtual precursorFormula
				ve=float(vwritelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz		
				virtualprecmz.append(ve)	# virtual precursormz
				e=str(vwritelist[3][r]) #sheetinput.cell(row=r, column=4)
				vprecadduct.append(e)
				e=int(vwritelist[5][r]) #sheetinput.cell(row=r, column=6)
				vprecchrg.append(e)
				e=str(vwritelist[6][r]) #sheetinput.cell(row=r, column=7)
				vprodname.append(e)
				e=str(vwritelist[7][r]) #sheetinput.cell(row=r, column=8)
				vprodformula.append(e)
				e=str(vwritelist[8][r]) #sheetinput.cell(row=r, column=9)
				vprodadduct.append(e)
				e=float(vwritelist[9][r]) #sheetinput.cell(row=r, column=10)
				vprodmz.append(e)
				e=int(vwritelist[10][r]) #sheetinput.cell(row=r, column=11)
				vprodchrg.append(e)	
				#e=0.055 
				#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
				#if int(vwritelist[1][r][8])>2:
				#	nca=10*int(vwritelist[1][r][5])+int(vwritelist[1][r][6])
				#	if nca>17:
				#		exrtstep=0.045
				#else:	
				#	exrtstep=0.027
				#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
				e=0.1 #exrtstep*2
				rtwindow.append(e)	
				r=r+1
			kr=kr+1
	if (r+1)>ki:
		r=s+3
	else:
		r=s+1

#print('Main loop done')	
#krows=len(vmlistname)
#print(krows)
#print(ki)
#print(kir)
#print(pfwritelist[1])
#quit()
# begin save to csv file
#before = datetime.datetime.now()
prt='PrecursorRT'
prt=str(prt)
toprow.append(prt)
rtw='PrecursorRTWindow'
rtw=str(rtw)
toprow.append(rtw)
#ki=2+len(vmlistname)

writelist=[]
writelist.append(vmlistname)
writelist.append(vprecname)
writelist.append(virtualprecformula)
writelist.append(vprecadduct)
writelist.append(virtualprecmz)
writelist.append(vprecchrg)
writelist.append(vprodname)
writelist.append(vprodformula)
writelist.append(vprodadduct)
writelist.append(vprodmz)
writelist.append(vprodchrg)
writelist.append(precrt)
writelist.append(rtwindow)
#print('writelist created')
#print(len(precrt))
#print(len(vmlistname))



nrows=len(vmlistname)
#print('Full discovery transition list will contain %d transitions. If this is more than 1M, consider to run the streamlined discovery workflow.'% nrows)
#discoverylevel=eval(input('Run full discovery workflow (0); streamlined discovery workflow (1) (limited to FA_library for all FA > 3 db; no discovery workflow (2) (all limited to FA_library)? ::'))
discoverylevel=0
# begin reduce transition list to species in fa_lib if desired for certain fatty acids 
if discoverylevel>0:
	r=0
	while r<(len(writelist[0])):
		s=r
		go=1
		while go==1:
			if s<(len(writelist[0])):
				if s>r:
					if str(writelist[1][r])==str(writelist[1][s]):
						s=s+1
					else:
						go=0
						s=s-1
				else:
					s=s+1
			else:
				go=0
				s=s-1
		cfaisom=str()	#get current fatty acid isomer species in list
		kcf=len(writelist[1][r])-1
		go=1
		while go==1:
			if str(writelist[1][r][kcf])=='_':
				go=0
			kcf=kcf-1
		icf=5
		while icf<(kcf+1):		# faulty? build correctly
			cfaisom=cfaisom+str(writelist[1][r][icf])
			icf=icf+1

		cutfa=0
		if discoverylevel==2:
			cutfa=1
		elif discoverylevel==1:
			if int(cfaisom[3])>3:
				cutfa=1
		print(cfaisom)
		if cutfa==1:
			if cfaisom in mostwantedlist:
				r=s+1
			else:
				while s>(r-1):
					lii=0
					while lii<13:
						del writelist[lii][s]
						lii=lii+1
					s=s-1
				r=s+1
		else:
			r=s+1
# end reduce transition list to species in fa_lib if desired for certain fatty acids

mespacedrule=0											########## ACTIVATE (1) OR DEACTIVATE (0) METHYLENE (BUTYLENE) SPACED RULE ####### ACTIVATE DOES NOT WORK
if mespacedrule==0:			## MODIFY BEFORE ACTIVATE
	transitionresultsdf=pd.DataFrame(writelist).transpose()
	#print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
	#print('Transposed and DataFrame created')
	transitionresultsdf.to_csv('jpmlipidomics_dda_vpw20_0.csv', index=False)
	nrows=len(vmlistname)
	print('Transition list is saved as jpmlipidomics_dda_vpw20_0.csv (%d rows)' % nrows)
	afterall=datetime.datetime.now()
	dt=afterall-beforeall
	print('Calculation time (h:mm:ss) is: %s' % dt)
	#print('Calculation time:')
	#print(dt)
	quit()
	# end save to csv file

#################################################################################################################################################
############################### END REDUCE TO ENTRIES OF FOUND PRECURSOR AND EXPAND WITH PRECURSOR EXPLICIT RETENTION TIMES #####################
#################################################################################################################################################
nlrows=0
#################################################################################################################################################
## begin apply methylene (butylene) spacing rule to relevant species 
## (delete unrealistic species that can't be distinguished from realistic species as associated double bonds are non-diagnostic)
#################################################################################################################################################
if mespacedrule==1:
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	wprecrt=[]
	wrtwindow=[]
	ki=len(writelist[0])
	#print('Rows in jpmlipidomics_vpw10_2_full_tr_results.csv:')
	#print(ki)
	r=0
	while r<ki:		# go through rows of excel file 
		e=writelist[1][r] # Precursorname		# begin determine which row to start (r) and to end (s)
		s=r+1
		st=0
		while st<1:
			if s>(len(writelist[1])-1):
				ne='stop_loop'
			else:
				ne=writelist[1][s] #sheetinput.cell(row=s, column=2)	# Precursorname
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		t=r
		pos=0
		if int(writelist[1][t][8])>2:		# Only consider PUFA with more than 2 db
			go=1
			ch=len(writelist[1][t])-1
			while go==1:					# determine last and second last double bond position in PUFA
				if str(writelist[1][t][ch])=='n':	
					go=0
					if str(writelist[1][t][ch+4])=='_':
						lastdb=10*int(writelist[1][t][ch+2])+int(writelist[1][t][ch+3])
					else:
						lastdb=int(writelist[1][t][ch+2])
					if str(writelist[1][t][ch-4])=='-':	
						seclastdb=10*int(writelist[1][t][ch-3])+int(writelist[1][t][ch-2])
					else:
						seclastdb=int(writelist[1][t][ch-2])
				else:
					go=1
				ch=ch-1
			if lastdb>13:
				if lastdb-seclastdb==3:	# allow methylene interrupted species for last db
					pos=0
				elif lastdb-seclastdb==6:	# allow butylene interrupted species for last db
					pos=0
				else:
					pos=1
		if pos==1:
			r=s+1
		else:
			#check if compartmentalization of discovery process required
			if discoverylimitation==1:
				if int(str(writelist[1][t][5])+str(writelist[1][t][6]))<cminlimit:
					godi=0
				elif int(str(writelist[1][t][5])+str(writelist[1][t][6]))>cmaxlimit:
					godi=0
				else:
					godi=1
			else:
				godi=1
			if godi==1:
				t=r
				while t<(s+1):
					e=writelist[0][t] ## mlistname	# begin append rows of suitable species to lists for later saving in excel output
					mlistname.append(e)
					e=writelist[1][t] ## precname	
					precname.append(e)
					e=writelist[2][t] ## 	
					precformula.append(e)
					e=writelist[3][t] ## 	
					precadduct.append(e)
					e=writelist[4][t] ## 	
					precmz.append(e)
					e=writelist[5][t] ## 	
					precchrg.append(e)
					e=writelist[6][t] ## 	
					prodname.append(e)
					e=writelist[7][t] ## 	
					prodformula.append(e)
					e=writelist[8][t] ## 	
					prodadduct.append(e)
					e=writelist[9][t] ## 
					prodmz.append(e)
					e=writelist[10][t] ## 	
					prodchrg.append(e)
					e=str(writelist[11][t]) ## 
					wprecrt.append(e)
					e=str(writelist[12][t]) ## 	
					wrtwindow.append(e)
					t=t+1
					nlrows=nlrows+1
		r=s+1

# end build full transition list and reduce and expand with reassigned targets (for neighboring targets, skip every second to reduce number of transitions leading to same peak)
###############################################################################################################################
###############################################################################################################################
if discoverylimitation==1:
	print('Transitionlist limited to search criteria is saved containing %d rows.' % nlrows)
###############################################################################################################################

toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']

vwritelist=[]
vwritelist.append(vmlistname)
vwritelist.append(vprecname)
vwritelist.append(virtualprecformula)
vwritelist.append(vprecadduct)
vwritelist.append(virtualprecmz)
vwritelist.append(vprecchrg)
vwritelist.append(vprodname)
vwritelist.append(vprodformula)
vwritelist.append(vprodadduct)
vwritelist.append(vprodmz)
vwritelist.append(vprodchrg)
vwritelist.append(wprecrt)
vwritelist.append(wrtwindow)
#print('vwritelist created')
#print(len(wprecrt))
#print(len(vmlistname))



transitionresultsdf=pd.DataFrame(vwritelist).transpose()
#print('Transposed')
transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
#print('Transposed and DataFrame created')
after=datetime.datetime.now()
after=str(after)
#today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_1_'+fourlettcode+'_'
#filename=today+'jpmlipidomics_vpw13_1_precursor.csv'
filename='jpmlipidomics_dda_vpw20_0.csv'
transitionresultsdf.to_csv(filename, index=False)
afterall=datetime.datetime.now()
dt=afterall-beforeall
#print('nrows')
nrows=len(vmlistname)
#print('Transition list is saved as yyyy_mm_dd_1_xxxx_jpmlipidomics_vpw13_1_precursor.csv (%d rows)' % nrows)
print('Transition list is saved as jpmlipidomics_dda_vpw20_0.csv (%d rows)' % nrows)
#print('Calculation time (h:mm:ss) is: ')
print('Calculation time (h:mm:ss) is: %d' % dt)
#print(dt)
# end save to csv file
#################################################################################################################################################


