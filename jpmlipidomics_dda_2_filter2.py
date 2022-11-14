# -*- coding: UTF-8 -*-

# Jan Philipp Menzel 
# Goal: Second filter step of DDA-based analysis
## Notes: Derivative, positive fixed charge
## NOTES: VIRTUAL PRECURSOR - PrecursorName and PrecursorMz are artificially set +Xe (only column 3 and 5), fragment transitions correct including precursor
## NOTES: Virtual precursor forces Skyline to consider all transitions incl. real precursor (fragment in transition list), Skyline Setting: TransitionSettings-Filter-IonTypes-f 
import math
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
import statistics
from statistics import mean
from statistics import median
################ DATABASE ## Source: Internetchemie.info 
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.757, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
################
#default=eval(input('Run workflow with default parameters? (Yes: 1 / No: 0) (Derivatization agent: AMPP; Slow and full workflow including all FA; Apply retention time limitation; Use Fatty Acid Library to prevent filtering out important FA; Max RT = 17.5 min; FA: 12 - 24 C; Max m/z error: 10 ppm; Precursor peak area threshold: 3000; Product peak area threshold: 200.) :'))
default=0
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
###########

########### begin read workflow parameters
transferlist=[]
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
go=1
while go==1:
	tfe=ws.cell(row=tli+1, column=2)
	tfe=tfe.value
	if tli==21:
		go=0
	if tfe is None:
		go=0
	else:
		transferlist.append(tfe)
	tli=tli+1
########## 
fourlettcode=transferlist[0]		#input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
cderiv=int(transferlist[1])
hderiv=int(transferlist[2])
dderiv=int(transferlist[3])
nderiv=int(transferlist[4])
oderiv=int(transferlist[5])
pderiv=int(transferlist[6])
ideriv=int(transferlist[7])
minlenfa=int(transferlist[8])
maxlenfa=int(transferlist[9])
mzcutoff=int(transferlist[10]) 		#eval(input('What is max mz [ppm] for positive identification of species? (e.g. 30) :'))
rettimecutoff=float(transferlist[11]) 		#eval(input('What is the maximum retention time [min]? (e.g. 11.9 or 17.8) :'))
areathreshold=int(transferlist[12])		#eval(input('What is the threshold for detected precursor peak area? (compare Skyline report file, e.g. 3000) :'))
workflowmode=int(transferlist[13])
productareathreshold=int(transferlist[14])
nchunks=int(transferlist[15])
rtlimitation=int(transferlist[16])
mostwanted=int(transferlist[17])
transtest=int(transferlist[18])
runprecheck=int(transferlist[19])
workflowidentifier=str(transferlist[20])
########### end read workflow parameters
# begin adjust workflow parameters for DDA analysis
mzcutoff=60
precareathreshold=50
prodareathreshold=50
# end adjust workflow parameters for DDA analysis
# begin build DDA identifier (replace AI or DIA with DDA in identifier YYYY_MM_DD_DIA_SAMPLE_NAME)
wfidentifierdda=str()
wfi=0
while wfi<11:
	wfidentifierdda=wfidentifierdda+workflowidentifier[wfi]
	wfi=wfi+1
wfidentifierdda=wfidentifierdda+'DDA'
if str(workflowidentifier[13])=='_':
	wfi=13
elif str(workflowidentifier[14])=='_':
	wfi=14
while wfi<len(workflowidentifier):
	wfidentifierdda=wfidentifierdda+workflowidentifier[wfi]
	wfi=wfi+1
# end build DDA identifier (replace AI with DDA)

beforeall=datetime.datetime.now()
#print('Workflow is running ...')

# begin build mostwantedlist 
mostwantedlist=[]
wbfa=openpyxl.load_workbook('jpm_fa_lib.xlsx')
wsfa=wbfa.active
mwi=0
go=1
while go==1:
	tfe=wsfa.cell(row=mwi+3, column=1)
	tfe=tfe.value
	if tfe is None:
		go=0
	else:
		if int(tfe)<10:
			tfe='0'+str(tfe)
		mwfa=str(tfe)+':'
		tfe=wsfa.cell(row=mwi+3, column=2)
		tfe=tfe.value
		dbfa=int(tfe)
		if int(tfe)==0:
			mwfa=mwfa+str(tfe)
		else:
			mwfa=mwfa+str(tfe)+'_n-'
			dbfai=1
			while dbfai<(dbfa+1):
				tfe=wsfa.cell(row=mwi+3, column=dbfai+2)
				tfe=tfe.value
				if dbfai<(dbfa):
					mwfa=mwfa+str(tfe)+'_n-'
				else:
					mwfa=mwfa+str(tfe)
				dbfai=dbfai+1
		mostwantedlist.append(mwfa)
	mwi=mwi+1
# end build mostwantedlist
#print('Arrived here.#################################################################################################################################')
#quit()

# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
diagnostics=[]
#end create empty lists


#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists
trdf=pd.read_csv('skyl_report_dda_vpw20_1_filtered.csv')
toprowx=[trdf.columns.values.tolist()]
toprow=toprowx[0]
trdf=trdf.transpose()
writelist=trdf.values.tolist()
ki=len(writelist[0])
print('Number of rows in skyl_report_dda_vpw20_1_filtered.csv: %d' % ki)
#print(ki)
#####################################################################################################
# begin assign whether transitions diagnostic or non-diagnostic
r=0
while r<ki:
	t=r
	if int(writelist[1][t][8])>2:
		go=1
		ch=len(writelist[1][t])-1
		while go==1:
			if str(writelist[1][t][ch])=='n':
				go=0
				if str(writelist[1][t][ch+4])=='_':
					lastdb=10*int(writelist[1][t][ch+2])+int(writelist[1][t][ch+3])		# determine last db (closest to amide bond)
				else:
					lastdb=int(writelist[1][t][ch+2])
				if str(writelist[1][t][ch-4])=='-':	
					seclastdb=10*int(writelist[1][t][ch-3])+int(writelist[1][t][ch-2])	# determine second last db (second closest to amide bond)
				else:
					seclastdb=int(writelist[1][t][ch-2])
			else:
				go=1
			ch=ch-1
		if lastdb>13:
			if lastdb-seclastdb==3:
				if str(writelist[6][t][len(writelist[6][t])-1])=='r':
					diagnostics.append('diagnostic')
				else:
					if str(writelist[6][t][len(writelist[6][t])-4])=='n':
						if lastdb==10*int(writelist[6][t][len(writelist[6][t])-2])+int(writelist[6][t][len(writelist[6][t])-1]):
							diagnostics.append('non-diagnostic')
						else:
							diagnostics.append('diagnostic')
					else:
						if lastdb==int(writelist[6][t][len(writelist[6][t])-1]):
							diagnostics.append('non-diagnostic')
						else:
							diagnostics.append('diagnostic')
			else:
				diagnostics.append('diagnostic')
		else:
			diagnostics.append('diagnostic')
	else:
		diagnostics.append('diagnostic')
	r=r+1
# end assign whether transitions diagnostic or non-diagnostic
####################################################################################################
####################################################################################################
###### begin get chromatograms and filter out duplicates that have product XICs trending to same max

segmentsize=50000	# min number of entries in xic report and transitions report to be processed at once (once functional change to 500, then increase to test if advantageous)
segr=0
segs=segmentsize
go=0
if len(writelist[1])>segs:
	if str(writelist[1][segs])==str(writelist[1][segs+1]):
		go=1
	while go==1:
		if str(writelist[1][segs])==str(writelist[1][segs+1]):
			go=1
			segs=segs+1
		else:
			go=0
else:
	segs=len(writelist[1])-1
segrun=1

convertfile=0
if convertfile==0:
	xictimesdf=pd.read_csv('skyl_xic_dda_report_vpw20_1_times.csv', header=None, skiprows=1, nrows=1)
	xictimeslistfromdf=xictimesdf.values.tolist()
	xictimeslist=xictimeslistfromdf[0]		# contains times of XICs # actual list starts at index 8

if segr==0:
	spr=1
else:
	spr=segr+1  ##############################################
nrw=segs-segr+1
segtrdf=pd.read_csv('skyl_xic_dda_report_vpw20_1_intensities.csv', skiprows=spr, header=None, nrows=nrw, low_memory=False)
allxiclist=segtrdf.values.tolist()
#begin write trdf to csv file to check its contents 	# TROUBLESHOOTING
#trdf.to_csv('skyl_xic_report_vpw16_3_troubleshooting.csv', index=False)
#end write trdf to csv file to check its contents 	# TROUBLESHOOTING
#trdf=trdf.transpose()#

#print(allxiclist[0][1])		####################################### 
kix=len(allxiclist)
print('Number of rows in segment in skyl_xic_dda_report_vpw20_1_filtered.csv: %d' % kix)	#May include species that were excluded as false positives based on decoy filter
#begin determine columns in xic_results, determine length of XIC
ci=len(allxiclist[0])
xiclength=int((ci-8)) #/2
#end determine columns in xic_results, determine length of XIC
ki=len(writelist[0])
print('Number of rows in segment after filter: %d' % ki)	#May include species that were excluded as false positives based on decoy filter
#print(dwritelist[1]) 	# list after decoy filter

# go through list of species and determine for each the RT of the OzID peak. 
# Then find duplicates that have essentially the same OzID peak RT and keep only the one closest to it.

##### IN SCOPE ####		 writelist		xictimeslist		allxiclist
# begin build ozidrtpeaklist
checkup=0
ozidrtpeaklist=[]	# list with same index as lists in writelist, containing median RT of peak RTs of diagnostic OzID transitions
r=0
ki=ki
while r<ki:		# go through rows of list 
	e=writelist[1][r] ## Precursorname		# begin determine which row to start (r) and to end (s)
	if checkup==1:
		print(e)
	s=r+1
	st=0
	while st<1:
		if s>(len(writelist[1])-1):
			ne='stop_loop'
		else:
			ne=writelist[1][s] ## Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
	t=r
	pos=0	# delete block later if pos=1		
	#begin determine exactexrtdiff for current species
	currentfwhm=[]
	currentexrt=[]
	exactexrtlist=[]
	xxrtintlist=[]
	diagnosticxicrtlist=[]
	diagnosticxicintlist=[]
	while t<(s-1):	#only transitions, not decoy or precursor
		#print(dwritelist[16][t])
		cfwhm=writelist[16][t]
		currentfwhm.append(cfwhm)
		cexrt=writelist[17][t]
		currentexrt.append(cexrt)
		t=t+1
	t=r
	if t<(s-1):
		precnm=writelist[1][t]
		xk=0
		allgexrtlist=[]		# list of lists with retention times that were determined by fitting routine for each run (A, B, C) and each transition
		allgcxicint=[]		# list of lists with XICs of transitions (only intensity)
		allgcxicrt=[]		# list of lists with XICs of transitions (only retention time)
		#print('len(allxiclist)')
		#print(len(allxiclist))
		#print('len(allxiclist)')
		while xk<len(allxiclist): 
			if precnm==allxiclist[xk][1]:
				if str(allxiclist[xk][4][len(allxiclist[xk][4])-1])=='r':
					nrt=0
				elif str(allxiclist[xk][4][len(allxiclist[xk][4])-1])=='y':
					nrt=0
				else:	
					cxicrt=[]	# get current xic (part around +- xicrange (2*fwhm) of exrt) - each full xic made up of 4730 datapoints (always the case?) 
					cxicint=[]
					cindex=[]
					cfwhm=writelist[16][t]
					if cfwhm==0:
						cfwhm=0.004
					elif str(cfwhm)=='nan':
						cfwhm=0.004
					if checkup==1:
						print('cfwhm is:')
						print(cfwhm)
						print('exrt is:')
						print(float(writelist[17][t]))
					xicrange=cfwhm*3 # 2 or 2.2 # set range of used XIC for fitting routine
					if cfwhm<0.03:
						xicrange=0.09
					kxicrt=8												# begin get the XIC as measured (for each transition of the current species)
					while kxicrt<(xiclength+8):
						cxx=xictimeslist[kxicrt-8]	# is correct
						if cxx>(float(writelist[17][t])-xicrange):
							if cxx<(float(writelist[17][t])+xicrange):
								cxicrt.append(cxx)
								cindex.append(int(kxicrt-8))
						kxicrt=kxicrt+1
					kxicint=8
					while kxicint<((xiclength)+8):
						cxx=allxiclist[xk][(kxicint)]
						if kxicint>(cindex[0]-1):
							if kxicint<(cindex[len(cindex)-1]+1):
								cxicint.append(cxx)
						kxicint=kxicint+1					# end get the XIC as measured (for each transition of the current species)
					if len(cxicrt)<len(cxicint):
						print('CHECK XIC INDEXING, len(cxicrt)<len(cxicint)')
						#print(len(cxicrt))
						#print(len(cxicint))
						while len(cxicrt)<(len(cxicint)):
							#print('fixed indexing')
							fix=cxicrt[len(cxicrt)-1]
							cxicrt.append(fix)
					if len(cxicrt)<len(cxicint):
						print('NOT REPAIRED YET !')
					elif len(cxicrt)>len(cxicint):
						print('CHECK XIC INDEXING, len(cxicrt)>len(cxicint)')
						#print(len(cxicrt))
						#print(len(cxicint))
						while len(cxicrt)>(len(cxicint)):
							#print('fixed indexing')
							fix=cxicint[len(cxicint)-1]
							cxicint.append(fix)
					nrt=0
					while cxicrt[nrt]<writelist[17][t]: # dwritelist[17][t] is exrt
						intexrt=cxicint[nrt]			 # intexrt after this loop is intensity at exrt
						nrt=nrt+1
					#print('intensity at exrt is:')
					#print(intexrt)
					cfwhm=writelist[16][t]
					if cfwhm==0:
						cfwhm=0.004
					elif str(cfwhm)=='nan':
						cfwhm=0.004
					cexrt=writelist[17][t]
					if checkup==1:
						print('cxicint is:')
						print(cxicint)
						print('cxicrt is:')
						print(cxicrt)
						print('cindex is:')
						print(cindex)

					##### begin 7 point smoothing of XIC		##################################################################################################
					fpsxicrt=[]
					fpsxicint=[]
					nrt=3
					while nrt<(len(cxicrt)-3):
						fpsn=cxicrt[nrt]
						fpsxicrt.append(fpsn)
						nrt=nrt+1
					nrt=3
					while nrt<(len(cxicint)-3):
						fpsn=((cxicint[nrt-3]+cxicint[nrt-2]+cxicint[nrt-1]+cxicint[nrt]+cxicint[nrt+1]+cxicint[nrt+2]+cxicint[nrt+3])/7)
						fpsxicint.append(fpsn)
						nrt=nrt+1
					##### end 7 point smoothing of XIC				##################################################################################################

					##### begin find max near cexrt in fpsxicint/fpsxicrt (find true center of peak using smoothed xic)
					nrt=0
					if checkup==1:
						print('fpsxicint is:')
						print(fpsxicint)
						print('fpsxicrt is:')
						print(fpsxicrt)
						print('cexrt is:')
						print(cexrt)
					cfpsxicint=0
					while fpsxicrt[nrt]<cexrt:
						cfpsxicrt=fpsxicrt[nrt]
						nrt=nrt+1
					nrt=nrt-1
					if fpsxicint[nrt]<fpsxicint[nrt+1]:			#cexrt before peakmax
						stop=0
						while fpsxicint[nrt]<fpsxicint[nrt+1]:
							cfpsxicrt=fpsxicrt[nrt+1]
							cfpsxicint=fpsxicint[nrt+1]
							if (nrt+1)==(len(fpsxicint)-1):
								if stop==0:
									nrt=(-1)
									stop=1
								else:
									fpsxicint.append(0)
							nrt=nrt+1
					else:										#cexrt after peakmax
						stop=0
						while fpsxicint[nrt]>fpsxicint[nrt+1]:
							cfpsxicrt=fpsxicrt[nrt]
							cfpsxicint=fpsxicint[nrt]
							if (nrt)==0:
								if stop==0:
									nrt=(len(fpsxicint)-1)
									stop=1
								else:
									nrt=(len(fpsxicint)-1)
									fpsxicint.append(0)
							nrt=nrt-1
						nrt=nrt+1
					##### end find max near cexrt in fpsxicint (find true center of peak using smoothed xic)
					if str(diagnostics[xk])=='diagnostic':
						exactexrtlist.append(cfpsxicrt)
						xxrtintlist.append(cfpsxicint)
						diagnosticxicrtlist.append(fpsxicrt)
						diagnosticxicintlist.append(fpsxicint)
					#begin troubleshooting, display current XIC and index for 20:3_n-7_n-10_n-13_7.8
					#checkup=0
					if checkup==1:
						if str(allxiclist[xk][1])=='AMPP_16:1_n-10_6.87':
							if str(allxiclist[xk][4])=='AMPP_16:1_n-10_criegee neutral loss from n-10':
								print('AMPP_16:1_n-10_criegee neutral loss from n-10')
								print(cxicrt)
								print(cxicint)
								print(fpsxicrt)
								print(fpsxicint)
							elif str(allxiclist[xk][4])=='AMPP_16:1_n-10_aldehyde neutral loss from n-10':
								print('AMPP_16:1_n-10_aldehyde neutral loss from n-10')
								print(cxicrt)
								print(cxicint)
								print(fpsxicrt)
								print(fpsxicint)
						
					#end troubleshooting display current XIC and index for 20:3_n-7_n-10_n-13_7.8					
					#begin troubleshooting, display current XIC and index for 20:3_n-7_n-10_n-13_7.8
					#if str(allxiclist[xk][1])=='AMPP_20:3_n-6_n-9_n-12_7.6':
					#	if str(allxiclist[xk][4])=='AMPP_20:3_n-6_n-9_n-12_criegee neutral loss from n-12':
					#		print('AMPP_20:3_n-6_n-9_n-12_criegee neutral loss from n-12')
					#		print('cfpsxicrt')
					#		print(cfpsxicrt)
					#		#print(cindex)
					#end troubleshooting display current XIC and index for 20:3_n-7_n-10_n-13_7.8

			xk=xk+1
		# exactexrtlist contains peak RT for each OzID transition
		exactexrtdiff=(max(exactexrtlist))-(min(exactexrtlist))  #largest difference between exactexrt values of transitions # not used here
		avgexactexrt=mean([mean(exactexrtlist), median(exactexrtlist)])	
		to=r
		while to<(s+1): 
			ozidrtpeaklist.append(avgexactexrt)
			to=to+1
		#avgexactexrt=statistics.mean(exactexrtlist)
		#print(dwritelist[1][t])
		#print(exactexrtlist)
		#print(exactexrtdiff)
		#print(dwritelist[1][t])
		#print(avgexactexrt)
		#quit()
	else:
		print('Check code - this should not have happened.')
	r=s+1

if checkup==1:
	r=0
	while r<ki:
		print(writelist[1][r])	
		print(writelist[6][r])
		print(ozidrtpeaklist[r])
		r=r+1
# end read transition results and build ozidrtpeaklist
#print('Arrived at filter 2.')
#quit()
# begin use ozidpeaklist to decide, which duplicates to keep

valprecname=[]
valprodname=[]
valexplicitrt=[]
valmzerror=[]
ki=len(writelist[0])
#print('Number of entries after first filter step:')
#print(ki)
r=0
while r<ki:		# go through rows of list 
	e=writelist[1][r] # Precursorname		# begin determine which row to start (r) and to end (s)
	s=r+1
	st=0
	while st<1:
		if s>(len(writelist[1])-1):
			ne='stop_loop'
		else:
			ne=writelist[1][s] #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
	t=s
	valprecname.append(str(writelist[1][t]))
	valprodname.append(str(writelist[6][t]))
	valexplicitrt.append(float(writelist[17][t]))
	valmzerror.append(str(writelist[11][t]))
	r=s+1
#print(valmzerror)
# val lists are built as overview of species in list
# compare and delete duplicate species within max +- 0.07 min from swritelist
# determine duplicate with max prod area sum
droplist=[]		# species to be deleted later
rv=0
while rv<(len(valprecname)):
	kv=0
	comparelist=[]
	compareexplicitrt=[]
	compareozidrt=[]
	comparediffrt=[]
	comparemzerror=[]
	while kv<(len(valprecname)):
		if rv==kv:
			rv=rv
		else:
			if valprodname[rv]==valprodname[kv]:
				if ozidrtpeaklist[rv]==ozidrtpeaklist[kv]:
					comparelist.append(valprecname[kv])
					compareozidrt.append(ozidrtpeaklist[kv])
					compareexplicitrt.append(valexplicitrt[kv])
					diffrt=abs(valexplicitrt[kv]-ozidrtpeaklist[kv])
					rdiffrt=abs(valexplicitrt[rv]-ozidrtpeaklist[rv])
					comparediffrt.append(diffrt)
					comparemzerror.append(str(valmzerror[kv]))
		kv=kv+1
	if (len(comparelist))>0:
		if abs(valexplicitrt[rv]-ozidrtpeaklist[rv])<min(comparediffrt):  #check if current rv is closest to peak of smoothed ozid xic
			rv=rv
		else:
			cpl=0
			canceldrop=1
			while cpl<(len(comparelist)):
				if str(comparemzerror[cpl])=='nan':
					canceldrop=canceldrop
				else:
					canceldrop=0
				cpl=cpl+1
			if canceldrop==0:   # check if other duplicate will be dropped based on 'nan'
				droplist.append(str(valprecname[rv])) #mark this rv for deletion
	if str(valmzerror[rv])=='nan':
		droplist.append(str(valprecname[rv])) #mark this rv for deletion
	rv=rv+1
# droplist contains precnames of species to be deleted

# add species to droplist that have same prod area

#print(droplist)
ldl=len(droplist)
#print(ldl)
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
#end create empty lists
r=0
ki=len(writelist[0])
while r<ki:
	e=writelist[1][r] # Precursorname		# begin determine which row to start (r) and to end (s)
	s=r+1
	st=0
	while st<1:
		if s>(len(writelist[1])-1):
			ne='stop_loop'
		else:
			ne=writelist[1][s] #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
	t=r
	q=0
	dpos=0
	while q<(len(droplist)):
		if str(writelist[1][t])==str(droplist[q]):
			dpos=1#
		else:
			dpos=dpos
		q=q+1
	if dpos==1:
		r=s+1
	else:
		while t<s+1:
			e=writelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving in excel output
			mlistname.append(e)
			e=writelist[1][t] #sheetinput.cell(row=t, column=2)	# precname
			e=str(e)
			el=len(e)
			ce=int(el-1)
			go=1
			while go==1:
				if str(e[ce])=='_':
					go=0
				ce=ce-1
			cce=0
			ne=str()
			while cce<(ce+1):
				ne=ne+str(e[cce])
				cce=cce+1
			ne=ne+'_'+str(round(float(ozidrtpeaklist[t]),2))
			precname.append(ne)
			e=writelist[2][t] ## precname	
			precformula.append(e)
			e=writelist[3][t] ## precformula	
			precadduct.append(e)
			e=writelist[4][t]  	
			precmz.append(e)
			e=writelist[5][t]  	
			precchrg.append(e)
			e=writelist[6][t] 	
			prodname.append(e)
			e=writelist[7][t]
			prodformula.append(e)
			e=writelist[8][t]
			prodadduct.append(e)
			e=writelist[9][t] 
			prodmz.append(e)
			e=writelist[10][t] 	
			prodchrg.append(e)
			e=writelist[11][t] 	
			mzerror.append(e)
			e=writelist[12][t]  	
			rettime.append(e)
			e=writelist[13][t] 	
			area.append(e)
			e=writelist[14][t] 
			areanormalpercent.append(e)
			e=writelist[15][t] 	
			background.append(e)
			e=writelist[16][t]	
			fwhm.append(e)
			#e=writelist[17][t]  
			e=ozidrtpeaklist[t]	
			explicitrt.append(e)
			e=writelist[18][t]  	
			rtstart.append(e)
			e=writelist[19][t] 	
			rtend.append(e)
			t=t+1
	r=s+1

swritelist=[]
swritelist.append(mlistname)
swritelist.append(precname)
swritelist.append(precformula)
swritelist.append(precadduct)
swritelist.append(precmz)
swritelist.append(precchrg)
swritelist.append(prodname)
swritelist.append(prodformula)
swritelist.append(prodadduct)
swritelist.append(prodmz)
swritelist.append(prodchrg)
swritelist.append(mzerror)
swritelist.append(rettime)
swritelist.append(area)
swritelist.append(areanormalpercent)
swritelist.append(background)
swritelist.append(fwhm)
swritelist.append(explicitrt)
swritelist.append(rtstart)
swritelist.append(rtend)

# End delete duplicates that are close to the duplicate with max integral

# end use ozidpeaklist to decide, which duplicates to keep
# end filter out species, if exact retention time of transitions differ too much #####################

mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
exrt=[]
exrtwindow=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
precoverlap=[]
#end create empty lists
#begin read file and save data in lists, edit strings and calculate fragment masses, build output lists

ki=len(swritelist[0])
print('Number of rows after filter step: %d' % ki)
keeplist=[]		# list with indexes r which are to be kept (the ones not included belong to lines that are part of duplicates)
r=0
e=str(swritelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname		# begin determine which row to start (r) and to end (s)
s=r+1
st=0
while st<1:
	if s>(len(swritelist[1])-1):
		st=1
		s=s-1
	else:
		ne=str(swritelist[1][s]) #sheetinput.cell(row=s, column=2)	# Precursorname
		if ne==e:
			s=s+1
			st=0
		else:
			s=s-1
			st=1		# end determine s
while r<s+1:
	keeplist.append(r)	# attach first species to keeplist
	r=r+1

kl=0
t=0
while kl<(len(swritelist[0])): #(len(keeplist)):
	#t=keeplist[kl]
	e=swritelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving 
	mlistname.append(e)
	e=swritelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
	precname.append(e)
	e=swritelist[2][t] #sheetinput.cell(row=t, column=3)	# precformula	
	precformula.append(e)
	e=swritelist[3][t] #sheetinput.cell(row=t, column=4)	# precadduct	
	precadduct.append(e)
	e=swritelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
	precmz.append(e)
	e=swritelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
	precchrg.append(e)
	e=swritelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
	prodname.append(e)
	e=swritelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
	prodformula.append(e)
	e=swritelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
	prodadduct.append(e)
	e=swritelist[9][t] #sheetinput.cell(row=t, column=10)	# 
	prodmz.append(e)
	e=swritelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
	prodchrg.append(e)
	e=swritelist[11][t] #sheetinput.cell(row=t, column=12)	# 	
	mzerror.append(e)
	e=swritelist[12][t] #sheetinput.cell(row=t, column=13)	# 	
	rettime.append(e)
	e=swritelist[13][t] #sheetinput.cell(row=t, column=14)	# 	
	area.append(e)
	e=swritelist[14][t] #sheetinput.cell(row=t, column=15)	# 	
	areanormalpercent.append(e)
	e=swritelist[15][t] #sheetinput.cell(row=t, column=16)	# 	
	background.append(e)
	e=swritelist[16][t] ## 	
	fwhm.append(e)
	e=swritelist[17][t] ## 	
	explicitrt.append(e)
	e=swritelist[18][t] ## 	
	rtstart.append(e)
	e=swritelist[19][t] ## 	
	rtend.append(e)
	done=1
	if done==0:
		e=str(swritelist[1][t]) ## precname to get exrt
		ls=len(e)
		k=e[ls-1]
		go=1
		z=1
		while go==1:
			k=e[ls-z]
			if k=='_':
				go=0
			else:
				z=z+1
		x=z-2
		cexrt=e[ls-x]
		x=x-1
		while x>0:
			cexrt=cexrt+e[ls-x]
			x=x-1
		cexrt=str(cexrt)
		cexrt=float(cexrt)
		exrt.append(cexrt)	#exrt done
	exrtwindow.append(0.1)		################################# ENTER EXPLICIT RETENTION TIME WINDOW ##############################
	precoverlap.append('ok')
	kl=kl+1
	t=t+1


# begin remove H'0 from formula, where applicable
#print("Start removing H'0")
pfi=0
while pfi<len(precformula):
	if "H'0" in str(precformula[pfi]):
		#remove H'0 from formula
		ri=precformula[pfi].index("H'0")
		precformula[pfi]=precformula[pfi][0:ri:]+precformula[pfi][ri+3::]
	if "H'0" in str(prodformula[pfi]):
		#remove H'0 from formula
		ri=prodformula[pfi].index("H'0")
		prodformula[pfi]=prodformula[pfi][0:ri:]+prodformula[pfi][ri+3::]
	pfi=pfi+1
#print("End removing H'0")
# end remove H'0 from formula, where applicable


writelist=[]
writelist.append(mlistname)
writelist.append(precname)
writelist.append(precformula)
writelist.append(precadduct)
writelist.append(precmz)
writelist.append(precchrg)
writelist.append(prodname)
writelist.append(prodformula)
writelist.append(prodadduct)
writelist.append(prodmz)
writelist.append(prodchrg)
writelist.append(explicitrt)
writelist.append(exrtwindow)
#print('First filter step is complete.')
############################################################################################################################################################

toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 
		'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']
#print('swritelist created')
transitionresultsdf=pd.DataFrame(writelist).transpose()
#print('Transposed')
transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
#print('Transposed and DataFrame created')
after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_5_'+fourlettcode+'_'
filename='jpmlipidomics_dda_vpw20_2_filtered.csv'
transitionresultsdf.to_csv(filename, index=False)
print('Transition list is saved as jpmlipidomics_dda_vpw20_2_filtered.csv')
afterall=datetime.datetime.now()
dt=afterall-beforeall
print('Calculation time (h:mm:ss) is: %s' % dt)
#print('Calculation time (h:mm:ss) is:')
#print(dt)







