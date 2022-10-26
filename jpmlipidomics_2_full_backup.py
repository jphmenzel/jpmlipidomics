# -*- coding: UTF-8 -*-

# Jan Philipp Menzel jpm_lipidomics_vpw13_2_full.py
#created: 09 07 2020
#modified: regularly until 07 04 2021  
# Goal: read excel file containing data for monounsaturated lipids without double bond info, add rows for OzID product ions, save in excel file
## Notes: work in progress for fatty acids AMPP derivatives (no other ionization), double bond position added to Precursorname after calculations but before saving in excel file 
## Notes: addition for saturated FAs, bisunsaturated FAs, added line for precursor, option for precursor-only transition list with dummy percursor.
## NOTES: VIRTUAL PRECURSOR - PrecursorName and PrecursorMz are artificially set +Xe (only column 3 and 5), fragment transitions correct including precursor
## NOTES: Virtual precursor forces Skyline to consider all transitions incl. real precursor (fragment in transition list), Skyline Setting: TransitionSettings-Filter-IonTypes-f
## NOTES: Include Fatty acids with three double bonds EDIT WRITING OF TRANSITIONS ALDEHYDE AND CRIGEE INTROCUDE NEW VARIABLE FOR THIRD DBposition
## NOTES: Added option for reducing transition list to one entry per precursor and then expanding list with added varied explicit retention times 
#  NOTES: filters precursor results and builds large list with pandas. export to csv 
import math
import os
import openpyxl
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
beforeall=datetime.datetime.now()
############################################################################################################################################
##################################################begin skyline filter routine - precursor only transition report###########################
############################################################################################################################################
# automated file loading from skyl_report_vpw15.csv
#print('Before proceeding, please make sure that the Skyline report file is named jpmlipidomics_vpw13_precursor_report.csv')
#selectiontype=eval(input('Generate Transition Results based on m/z error and retention time cutoff only (0) or based on strict selection criteria (1)? : '))
selectiontype=1

# begin build mostwantedlist 
mostwantedlist=[]
wbfa=openpyxl.load_workbook('jpm_fa_lib.xlsx')
wsfa=wbfa.active
mwi=0
go=1
while go==1:
	tfe=wsfa.cell(row=mwi+3, column=1)
	tfe=tfe.value
	if tfe is None:
		go=0
	else:
		if int(tfe)<10:
			tfe='0'+str(tfe)
		mwfa=str(tfe)+':'
		tfe=wsfa.cell(row=mwi+3, column=2)
		tfe=tfe.value
		dbfa=int(tfe)
		if int(tfe)==0:
			mwfa=mwfa+str(tfe)
		else:
			mwfa=mwfa+str(tfe)+'_n-'
			dbfai=1
			while dbfai<(dbfa+1):
				tfe=wsfa.cell(row=mwi+3, column=dbfai+2)
				tfe=tfe.value
				if dbfai<(dbfa):
					mwfa=mwfa+str(tfe)+'_n-'
				else:
					mwfa=mwfa+str(tfe)
				dbfai=dbfai+1
		mostwantedlist.append(mwfa)
	mwi=mwi+1
# end build mostwantedlist
#print(mostwantedlist)
#print(mostwantedlist[len(mostwantedlist)-1])
########### begin read workflow parameters
transferlist=[]
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
go=1
while go==1:
	tfe=ws.cell(row=tli+1, column=2)
	tfe=tfe.value
	if tli==17:
		go=0
	if tfe is None:
		go=0
	else:
		transferlist.append(tfe)
	tli=tli+1
########## 
fourlettcode=transferlist[0]		#input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
cderiv=int(transferlist[1])
hderiv=int(transferlist[2])
dderiv=int(transferlist[3])
nderiv=int(transferlist[4])
oderiv=int(transferlist[5])
pderiv=int(transferlist[6])
ideriv=int(transferlist[7])
workflowmode=int(transferlist[13])
mzcutoff=int(transferlist[10]) 		#eval(input('What is max mz [ppm] for positive identification of species? (e.g. 30) :'))
rettimecutoff=float(transferlist[11]) 		#eval(input('What is the maximum retention time [min]? (e.g. 11.9 or 17.8) :'))
areathreshold=int(transferlist[12])		#eval(input('What is the threshold for detected precursor peak area? (compare Skyline report file, e.g. 3000) :'))
mostwanted=int(transferlist[17])
########### end read workflow parameters
#smallareathreshold=largeareathreshold # 
#smallareathreshold=eval(input('What is the threshold for precursor (<4db) peak areas? (e.g. 3000) :'))
#largeareathreshold=10000 # for sensitive analysis, use 750
#nspec=3	# number of species: precursor, aldehyde, criegee
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e        ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
#abundance=[99.9885, 98.93, 99.636, 99.7, 94.93] not updated
#ki=eval(input('Number of entries?'))
# begin create empty lists
toprow=[]
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
mzerror=[]
rettime=[]
area=[]
areanormalpercent=[]
background=[]
fwhm=[]
explicitrt=[]
rtstart=[]
rtend=[]
#end create empty lists
# begin read csv file (skyline report precursor only)

trdf=pd.read_csv('skyl_report_vpw20_1.csv')
toprowx=[trdf.columns.values.tolist()]
toprow=toprowx[0]
trdf=trdf.transpose()
writelist=trdf.values.tolist()
ki=len(writelist[0])
print('Number of rows in skyl_report_vpw20_1.csv (Skyline report on precursor retention times) : %d' % ki)

# end read csv file (skyline report precursor only)

r=0
while r<ki:		# go through rows of writelist
	currentmzerror=[]
	currentrettime=[]
	currentareas=[]
	currentprodmz=[]
	e=writelist[1][r] # Precursorname		# begin determine which row to start (r) and to end (s)
	s=r#+1
	#st=0
	#while st<1:
	#	ne=sheetinput.cell(row=s, column=2)	# Precursorname
	#	ne=ne.value
	#	if ne==e:
	#		s=s+1
	#		st=0
	#	else:
	#		s=s-1
	#		st=1		# end determine s
	t=r
	pos=0
	fpos=0
	gpos=0
	while t<(s+1):
		e=str(writelist[11][t]) ## mzerror
		if e=='nan':
			pos=1
		else:
			e=float(writelist[11][t]) # mzerror
			currentmzerror.append(e)
			if abs(e)>mzcutoff:
				if mostwanted==0:
					pos=1
				else:
					# begin test whether species to be cut is in the mostwantedlist (if it is, keep only if area higher than half the areacutoff)
					fashort=str()
					fas=5
					while fas<(len(writelist[1][t])-5):
						fashort=fashort+str(writelist[1][t][fas])
						fas=fas+1
					mwl=0
					gposcancel=0
					while mwl<(len(mostwantedlist)):
						if mostwantedlist[mwl]==fashort:
							gposcancel=1
							if abs(e)>(mzcutoff*1.25):		# use 1.25 times the m/z error threshold for mostwanted species
								pos=1
						mwl=mwl+1
					if gposcancel==0:
						pos=1
					# end test whether species to be cut is in the mostwantedlist
				#pos=1
			f=float(writelist[12][t]) ## retention time
			currentrettime.append(f)
			if f>rettimecutoff:
				fpos=1
			g=writelist[13][t] ## area is 13, areaNormalized is 14
			if g=='nan':
				gpos=1
			else:
				g=float(g)
				#print(g)
				dg=str(writelist[6][t]) #
				#if int(dg[8])<4:
				if g<areathreshold:
					if mostwanted==0:
						gpos=1
					else:
						# begin test whether species to be cut is in the mostwantedlist (if it is, keep only if area higher than half the areacutoff)
						fashort=str()
						fas=5
						while fas<(len(writelist[1][t])-5):
							fashort=fashort+str(writelist[1][t][fas])
							fas=fas+1
						mwl=0
						gposcancel=0
						while mwl<(len(mostwantedlist)):
							if mostwantedlist[mwl]==fashort:
								gposcancel=1
								if g<(areathreshold*0.5):		# use half the area threshold for mostwanted species
									gpos=1
							mwl=mwl+1
						if gposcancel==0:
							gpos=1
						# end test whether species to be cut is in the mostwantedlist
					#gpos=1
			# skip block
			#currentareas.append(g)
			h=float(writelist[9][t]) ## prodmz
			currentprodmz.append(h)
		t=t+1
	# begin use current area and prodmz to determine, whether current species has a precursor abundance that is lower than all fragments (criterium for exclusion)
	#if len(currentareas)>1:		#only exclude unsaturated FAs, keep saturated FAs and precursor only entries
	#	ci=0
	#	call=0
	#	while call<1:
	#		ciprodmz=currentprodmz[ci]
	#		cj=0
	#		cc=0
	#		while cj<len(currentprodmz):
	#			cjprodmz=currentprodmz[cj]
	#			if ciprodmz>(cjprodmz-0.001):
	#				cc=1
	#			else:
	#				ct=0
	#			cj=cj+1
	#		if ct==1:
	#			call=1
	#			cmax=ci
	#		else:
	#			ci=ci+1		# species with highest mz is identified
	#	cmax=len(currentareas)-1		# assume that last species is precursor (disable this line, if not the case!!!)
	#	# cmax is index for species with highest mass (precursor)
	#	ciareas=currentareas[cmax]
	#	cj=0
	#	cc=0
	#	ct=1
	#	while cj<len(currentareas):
	#		cjareas=currentareas[cj]
	#		if ciareas<(cjareas+0.001):
	#			cc=1
	#		else:
	#			ct=0
	#		cj=cj+1
	#	if ct==1:
	#		if selectiontype==1:
	#			gpos=1
	# end use current area and prodmz to determine, whether current species has a precursor abundance that is lower than all fragments (criterium for exclusion)
	if pos==1:
		r=s+1
	elif fpos==1:
		r=s+1
	elif gpos==1:
		r=s+1
	else:
		t=r
		while t<(s+1):
			e=writelist[18][t] #begin determine whether duplicate is found, duplicate not saved
			rtstartcurrent=e
			e=writelist[19][t]
			rtendcurrent=e
			if len(mlistname)==0:
				apos=1
			else:
				apos=1
				rtstartprevious=rtstart[(len(rtstart)-1)]
				rtendprevious=rtend[(len(rtend)-1)]
				if rtstartprevious==rtstartcurrent:
					if rtendprevious==rtendcurrent:
						apos=1			# 	duplicates not removed, kept for later stage of analysis with transitions of fragments
						#apos=0			# 	end determine whether duplicate is found, duplicate not saved in next step
			if apos==1:
				e=writelist[0][t] ## mlistname	# begin append rows of suitable species to lists for later saving in excel output
				mlistname.append(e)
				e=writelist[1][t] 	# precname	
				precname.append(e)
				e=writelist[2][t]
				precformula.append(e)
				e=writelist[3][t] # prodformula	
				precadduct.append(e)
				e=writelist[4][t] 
				precmz.append(e)
				e=writelist[5][t] 
				precchrg.append(e)
				e=writelist[6][t] 
				prodname.append(e)
				e=writelist[7][t] 
				prodformula.append(e)
				e=writelist[8][t] 
				prodadduct.append(e)
				e=writelist[9][t] 
				prodmz.append(e)
				e=writelist[10][t] 
				prodchrg.append(e)
				e=writelist[11][t] 
				mzerror.append(e)
				e=writelist[12][t] 
				rettime.append(e)
				e=writelist[13][t] 
				area.append(e)
				e=writelist[14][t]
				areanormalpercent.append(e)
				e=writelist[15][t] 
				background.append(e)
				e=writelist[16][t] 
				fwhm.append(e)
				e=writelist[17][t] 
				explicitrt.append(e)
				e=writelist[18][t] 
				rtstart.append(e)
				e=writelist[19][t] 
				rtend.append(e)
			t=t+1
		r=s+1
# end read transition results and append suitable species to lists
#begin save pfwritelist
toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'MassErrorPPM', 'RetentionTime', 'Area', 'AreaNormalized', 'Background', 'FWHM', 'ExplicitRT', 'StartRT', 'EndRT']
pfwritelist=[]
pfwritelist.append(mlistname)
pfwritelist.append(precname)
pfwritelist.append(precformula)
pfwritelist.append(precadduct)
pfwritelist.append(precmz)
pfwritelist.append(precchrg)
pfwritelist.append(prodname)
pfwritelist.append(prodformula)
pfwritelist.append(prodadduct)
pfwritelist.append(prodmz)
pfwritelist.append(prodchrg)
pfwritelist.append(mzerror)
pfwritelist.append(rettime)
pfwritelist.append(area)
pfwritelist.append(areanormalpercent)
pfwritelist.append(background)
pfwritelist.append(fwhm)
pfwritelist.append(explicitrt)
pfwritelist.append(rtstart)
pfwritelist.append(rtend)
#print('writelist created')
transitionresultsdf=pd.DataFrame(pfwritelist).transpose()
#print('Transposed')
transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12],toprow[13],toprow[14],toprow[15],toprow[16],toprow[17],toprow[18],toprow[19]]
#print('Transposed and DataFrame created')
transitionresultsdf.to_csv('jpmlipidomics_vpw20_2_precursor_analysis.csv', index=False)
print('Transition list saved as jpmlipidomics_vpw20_2_precursor_analysis.csv')
#print(pfwritelist[1])
#print(pfwritelist[2])
#print(pfwritelist[17])
print('Number of rows in transition list after filtering: %d' % len(mlistname))
#print('Precursor results are filtered. Full transition list is now being built...')
#quit() #################################################################################################################### check ok
#end save pfwritelist
############################################################################################################################################
####################################################end skyline filter routine - precursor only transition report###########################
############################################################################################################################################
# begin calculate target list 
# all filtered species are in pfwritelist; [9] is prodmz; [17] is exrt
tc=0
targetrtlist=[]
targetmzlist=[]
targetstepwidth=0.005 #min	# original value 0.01 min, now 0.005 min to be tested for more points across peaks for targeting (1 MS/MS scan per target)
tci=0
# 0.003 is 0.3% of RT as RT range extension to account for potential RT shift between DIA and DDA LC-MS run
while (float(pfwritelist[17][tc])-(2*targetstepwidth)+((tci+1)*targetstepwidth))>(float(pfwritelist[17][tc])-0.027-(0.003*(float(pfwritelist[17][tc])))): 
	tci=tci-1
while (float(pfwritelist[17][tc])-(2*targetstepwidth)+((tci+1)*targetstepwidth))<(float(pfwritelist[17][tc])+0.027):
	targetrtlist.append(float(pfwritelist[17][tc])-(2*targetstepwidth)+((tci+1)*targetstepwidth))
	targetmzlist.append(float(pfwritelist[9][tc]))
	tci=tci+1
tc=1
while tc<(len(pfwritelist[0])):
	if int(pfwritelist[6][tc][8])>0:		# unsaturated FA
		if pfwritelist[6][tc]==pfwritelist[6][tc-1]:
			if abs(float(pfwritelist[17][tc])-float(pfwritelist[17][tc-1]))>0.1:
				tci=0		# not neighboring
				# 0.003 is 0.3% of RT as RT range extension to account for potential RT shift between DIA and DDA LC-MS run
				while (float(pfwritelist[17][tc])-(2*targetstepwidth)+((tci+1)*targetstepwidth))>(float(pfwritelist[17][tc])-0.027-(0.003*(float(pfwritelist[17][tc])))): 
					tci=tci-1
				while (float(pfwritelist[17][tc])-(2*targetstepwidth)+((tci+1)*targetstepwidth))<(float(pfwritelist[17][tc])+0.027+(0.003*(float(pfwritelist[17][tc])))):
					targetrtlist.append(float(pfwritelist[17][tc])-(2*targetstepwidth)+((tci+1)*targetstepwidth))
					targetmzlist.append(float(pfwritelist[9][tc]))
					tci=tci+1
			else:
				ii=0	#neighboring
				prt=targetrtlist[(len(targetrtlist)-1)]
				while (float(pfwritelist[17][tc])+(4*targetstepwidth)+(0.003*(float(pfwritelist[17][tc]))))>prt:
					nrt=prt+targetstepwidth
					targetrtlist.append(nrt)
					nmz=float(pfwritelist[9][tc]) #+(ii*0.000001) variation disabled to allow comparison via mz
					targetmzlist.append(nmz)
					prt=targetrtlist[(len(targetrtlist)-1)]
					ii=ii+1
		else:
			tci=0	# new FA
			# 0.003 is 0.3% of RT as RT range extension to account for potential RT shift between DIA and DDA LC-MS run
			while (float(pfwritelist[17][tc])-(2*targetstepwidth)+((tci+1)*targetstepwidth))>(float(pfwritelist[17][tc])-0.027-(0.003*(float(pfwritelist[17][tc])))): 
				tci=tci-1
			while (float(pfwritelist[17][tc])-(2*targetstepwidth)+((tci+1)*targetstepwidth))<(float(pfwritelist[17][tc])+0.027+(0.003*(float(pfwritelist[17][tc])))):
				targetrtlist.append(float(pfwritelist[17][tc])-(2*targetstepwidth)+((tci+1)*targetstepwidth))
				targetmzlist.append(float(pfwritelist[9][tc]))			
				tci=tci+1
	tc=tc+1
# begin format m/z values to six digits after comma
rawtargetmzlist=[]
rawtargetrtlist=[]
fdg=0
while fdg<(len(targetmzlist)):
	targetmzlist[fdg]=format(targetmzlist[fdg],'.6f')
	targetrtlist[fdg]=float(format(targetrtlist[fdg],'.6f'))
	rawtargetmzlist.append(format(float(targetmzlist[fdg]),'.6f'))
	rawtargetrtlist.append(float(format(targetrtlist[fdg],'.6f')))
	fdg=fdg+1
# end format m/z values to six digits after comma

troubleshoot=1		##########################################################################################################################
if troubleshoot==1:
	# begin TROUBLESHOOTING prepare data for writing into target list file; create .txt file and write target list into it
	rawtxtwritelist=[]
	#targetrtlist=[9.11, 8.79, 7.2, 7.23]
	#targetmzlist=[477.383900, 477.383901, 447.3370, 495.3370]   
	cl=0
	while cl<(len(rawtargetrtlist)):
		rawtxtwritelist.append(['', '-1.000000', '-1.000000', '0', '', '-1.000000', '-1.000000','1'])
		cl=cl+1
	cl=0
	while cl<(len(rawtargetrtlist)):
		rawtxtwritelist[cl][0]=str(format(float(rawtargetmzlist[cl]),'.6f'))
		rawtxtwritelist[cl][4]=str(format((60*rawtargetrtlist[cl]),'.6f'))
		cl=cl+1
	#print(txtwritelist)
	after=datetime.datetime.now()
	after=str(after)
	today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_'+fourlettcode+'_'
	filename=today+'jpmlipidomics_dda_targetlist_raw.txt'
	path='C:/Users/menzel2/batchprogramming/OzFAD1' #'C:/Users/menzel2/pythonprogramming'
	file=filename
	with open(os.path.join(path, file), 'w') as fp:
		pass
		j=0
		while j<(len(rawtxtwritelist)):
			i=0
			while i<(len(rawtxtwritelist[j])):
				fp.write(str(rawtxtwritelist[j][i]))
				if i<(len(rawtxtwritelist[j])-1):
					fp.write(',')
				else:
					fp.write('\n')
				i=i+1
			j=j+1
	print('Raw targetlist is saved as a .txt file.')
	#quit()
	# end TROUBLESHOOTING prepare data for writing into target list file; create .txt file and write target list into it
print('Reassigning targets...')
# begin sort and delete and reassign species and RTs to prevent multiple species assignment at same RT range
check=0
go=1
while go==1:
	# begin test whether targets with RT overlap exist, if not, then go=0; if yes, find first overlap, identify whole range, redefine and restart search to test with go=1
	found=0
	i=0
	while i<(len(targetrtlist)):
		j=0
		while j<(len(targetrtlist)):
			if i==j:
				i=i
			else:
				if (abs(targetrtlist[i]-targetrtlist[j]))<(targetstepwidth-0.000001):
					if targetmzlist[i]==targetmzlist[j]:
						i=i
					else:
						found=1		# two or more entries at overlapping RT with different m/z # identify range, modify species within range
						#print('Two or more entries at overlapping RT with different m/z')
						#print(targetrtlist[i]*60)
						#print('tmzi')
						#print(targetmzlist[i])
						#print(targetrtlist[j]*60)
						#print('tmzj')
						#print(targetmzlist[j])
						thirdfound=0
						k=0
						while k<(len(targetrtlist)):
							if abs(targetrtlist[k]-targetrtlist[i])<(targetstepwidth-0.000001):
								if targetmzlist[k]==targetmzlist[i]:
									k=k
								elif targetmzlist[k]==targetmzlist[j]:
									k=k
								else:
									thirdfound=1
									tmzk=targetmzlist[k]
							else:
								thirdfound=thirdfound
							k=k+1
						#print('thirdfound')
						#print(thirdfound)
						#print('tmzk')
						#print(tmzk)
						if thirdfound==0:
							#print('two entries at overlapping RT with different m/z') # identify range, modify species within range
							#print('Two species overlap....................................-2-...........')
							rangestart=targetrtlist[i]
							tmzi=targetmzlist[i]
							tmzj=targetmzlist[j]
							crange=rangestart+targetstepwidth
							mi=0
							igo=0
							jgo=0
							zgo=1
							while mi<(len(targetrtlist)):
								if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
									if targetmzlist[mi]==tmzi:
										igo=1
									elif targetmzlist[mi]==tmzj:
										jgo=1
									else:
										zgo=0
								else:
									mi=mi
								mi=mi+1
							if igo==1:
								if jgo==1:
									if zgo==1:
										mgo=1
									else:
										mgo=0
								else:
									mgo=0
							else:
								mgo=0
							while mgo==1:
								crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
								mi=0
								igo=0
								jgo=0
								zgo=1
								while mi<(len(targetrtlist)):
									if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
										if targetmzlist[mi]==tmzi:
											igo=1
										elif targetmzlist[mi]==tmzj:
											jgo=1
										else:
											zgo=0
									else:
										#crange=crange-targetstepwidth
										mi=mi 
									mi=mi+1
								if igo==1:
									if jgo==1:
										if zgo==1:
											mgo=1
										else:
											mgo=0
											crange=crange-targetstepwidth
									else:
										mgo=0
										crange=crange-targetstepwidth
								else:
									mgo=0
									crange=crange-targetstepwidth
							rangeend=crange #-targetstepwidth		#end of this range is defined now
							#print('rangestart')
							#print(rangestart*60)
							#print('rangeend')
							#print(rangeend*60)
							#if rangestart<(365/60):
								#print('###############################################################')
							# begin delete all j species in range, reassign i species as alternating between j and i.
							md=0
							delgo=0
							while md<(len(targetrtlist)):
								if targetmzlist[md]==tmzj:
									if targetrtlist[md]>(rangestart-targetstepwidth):
										if targetrtlist[md]<(rangeend+targetstepwidth):
											#print('A species was deleted from overlap of 2:')
											#print(targetmzlist[md])
											#print(targetrtlist[md]*60)
											del targetrtlist[md]	# delete species at index md for rt list and mz list
											del targetmzlist[md]
											delgo=1
								if delgo==0:
									md=md+1
								else:
									delgo=0
							ma=0
							altern=0
							while ma<(len(targetrtlist)):
								if targetmzlist[ma]==tmzi:
									if targetrtlist[ma]>(rangestart-targetstepwidth):
										if targetrtlist[ma]<(rangeend+targetstepwidth):
											if altern==0:
												altern=altern
											elif altern==1:
												targetmzlist[ma]=tmzj
												#print('Newly assigned entry:')
												#print(targetmzlist[ma])
												#print(targetrtlist[ma]*60)
											altern=altern+1
											if altern==2:
												altern=0
								ma=ma+1
							j=len(targetrtlist)
							i=len(targetrtlist)
							#print('____________________________________')
						elif thirdfound==1:
							print('three or more entries at overlapping RT with different m/z # identify range, modify species within range')
							tmzi=targetmzlist[i]
							tmzj=targetmzlist[j]
							kf=0
							fourthfound=0
							while kf<(len(targetrtlist)):
								if abs(targetrtlist[kf]-targetrtlist[i])<(targetstepwidth-0.000001):
									if targetmzlist[kf]==tmzi:
										kf=kf
									elif targetmzlist[kf]==tmzj:
										kf=kf
									elif targetmzlist[kf]==tmzk:
										kf=kf
									else:
										fourthfound=1
										tmzkf=targetmzlist[kf]
								else:
									fourthfound=fourthfound
								kf=kf+1
							#print('fourthfound')
							#print(fourthfound)
							#print('tmzkf')
							#print(tmzkf)
							#quit()
							if fourthfound==0:
								k=k # three entries at overlapping RT with different m/z # identify range, modify species within range
								#print('Three species overlap....................................-3-...........')
								rangestart=targetrtlist[i]
								tmzi=targetmzlist[i]
								tmzj=targetmzlist[j]
								crange=rangestart+targetstepwidth
								mi=0
								igo=0
								jgo=0
								kgo=0
								zgo=1
								while mi<(len(targetrtlist)):
									if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
										if targetmzlist[mi]==tmzi:
											igo=1
										elif targetmzlist[mi]==tmzj:
											jgo=1
										elif targetmzlist[mi]==tmzk:
											kgo=1
										else:
											zgo=0
									else:
										mi=mi
									mi=mi+1
								if igo==1:
									if jgo==1:
										if kgo==1:
											if zgo==1:
												mgo=1
											else:
												mgo=0
										else:
											mgo=0
									else:
										mgo=0
								else:
									mgo=0
								while mgo==1:
									crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
									mi=0
									igo=0
									jgo=0
									kgo=0
									zgo=1
									while mi<(len(targetrtlist)):
										if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
											if targetmzlist[mi]==tmzi:
												igo=1
											elif targetmzlist[mi]==tmzj:
												jgo=1
											elif targetmzlist[mi]==tmzk:
												kgo=1
											else:
												zgo=0
										else:
											#crange=crange-targetstepwidth
											mi=mi
										mi=mi+1
									if igo==1:
										if jgo==1:
											if kgo==1:
												if zgo==1:
													mgo=1
												else:
													mgo=0
													crange=crange-targetstepwidth
											else:
												mgo=0
												crange=crange-targetstepwidth
										else:
											mgo=0
											crange=crange-targetstepwidth
									else:
										mgo=0
										crange=crange-targetstepwidth
								rangeend=crange		#end of this range is defined now
								# begin delete all j and k species in range, reassign i species as alternating between i, j and k.
								md=0
								delgo=0
								while md<(len(targetrtlist)):
									if targetmzlist[md]==tmzj:
										if targetrtlist[md]>(rangestart-targetstepwidth):
											if targetrtlist[md]<(rangeend+targetstepwidth):
												#print('A species was deleted from overlap of 3:')
												#print(targetmzlist[md])
												#print(targetrtlist[md]*60)
												del targetrtlist[md]	# delete j species at index md for rt list and mz list
												del targetmzlist[md]
												delgo=1
									if delgo==0:
										md=md+1
									else:
										delgo=0
								md=0
								delgo=0
								while md<(len(targetrtlist)):
									if targetmzlist[md]==tmzk:
										if targetrtlist[md]>(rangestart-targetstepwidth):
											if targetrtlist[md]<(rangeend+targetstepwidth):
												#print('A species was deleted from overlap of 3:')
												#print(targetmzlist[md])
												#print(targetrtlist[md]*60)
												del targetrtlist[md]	# delete k species at index md for rt list and mz list
												del targetmzlist[md]
												delgo=1
									if delgo==0:
										md=md+1
									else:
										delgo=0
								ma=0
								altern=0
								while ma<(len(targetrtlist)):
									if targetmzlist[ma]==tmzi:
										if targetrtlist[ma]>(rangestart-targetstepwidth):
											if targetrtlist[ma]<(rangeend+targetstepwidth):
												if altern==0:
													altern=altern
												elif altern==1:
													targetmzlist[ma]=tmzj
												else:
													targetmzlist[ma]=tmzk
												altern=altern+1
												if altern==3:
													altern=0
									ma=ma+1
								j=len(targetrtlist)
								i=len(targetrtlist)
							elif fourthfound==1:
								k=k # four or more entries at overlapping RT with different m/z # identify range, modify species within range
								tmzi=targetmzlist[i]
								tmzj=targetmzlist[j]
								kg=0
								fifthfound=0
								while kg<(len(targetrtlist)):
									if abs(targetrtlist[kg]-targetrtlist[i])<(targetstepwidth-0.000001):
										if targetmzlist[kg]==tmzi:
											kg=kg
										elif targetmzlist[kg]==tmzj:
											kg=kg
										elif targetmzlist[kg]==tmzk:
											kg=kg
										elif targetmzlist[kg]==tmzkf:
											kg=kg
										else:
											fifthfound=1
											tmzkg=targetmzlist[kg]
											#print('####################################################################################### OVERLAP OF 5 #######')
									else:
										fifthfound=fifthfound
									kg=kg+1
								#print('fifthfound')
								#print(fifthfound)
								#print('tmzkg')
								#print(tmzkg)
								#quit()
								if fifthfound==0:
									k=k # four entries at overlapping RT with different m/z # identify range, modify species within range
									#print('Four species overlap....................................-4-...........')
									rangestart=targetrtlist[i]
									tmzi=targetmzlist[i]
									tmzj=targetmzlist[j]
									crange=rangestart+targetstepwidth
									mi=0
									igo=0
									jgo=0
									kgo=0
									lgo=0
									zgo=1
									while mi<(len(targetrtlist)):
										if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
											if targetmzlist[mi]==tmzi:
												igo=1
											elif targetmzlist[mi]==tmzj:
												jgo=1
											elif targetmzlist[mi]==tmzk:
												kgo=1
											elif targetmzlist[mi]==tmzkf:
												lgo=1
											else:
												zgo=0
										else:
											mi=mi
										mi=mi+1
									if igo==1:
										if jgo==1:
											if kgo==1:
												if lgo==1:
													if zgo==1:
														mgo=1
													else:
														mgo=0
												else:
													mgo=0
											else:
												mgo=0
										else:
											mgo=0
									else:
										mgo=0
									while mgo==1:
										crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
										mi=0
										igo=0
										jgo=0
										kgo=0
										lgo=0
										zgo=1
										while mi<(len(targetrtlist)):
											if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
												if targetmzlist[mi]==tmzi:
													igo=1
												elif targetmzlist[mi]==tmzj:
													jgo=1
												elif targetmzlist[mi]==tmzk:
													kgo=1
												elif targetmzlist[mi]==tmzkf:
													lgo=1
												else:
													zgo=0
											else:
												mi=mi
											mi=mi+1
										if igo==1:
											if jgo==1:
												if kgo==1:
													if lgo==1:
														if zgo==1:
															mgo=1
														else:
															mgo=0
															crange=crange-targetstepwidth
													else:
														mgo=0
														crange=crange-targetstepwidth
												else:
													mgo=0
													crange=crange-targetstepwidth
											else:
												mgo=0
												crange=crange-targetstepwidth
										else:
											mgo=0
											crange=crange-targetstepwidth
									rangeend=crange		#end of this range is defined now
									# begin delete all j and k species in range, reassign i species as alternating between i, j and k.
									md=0
									delgo=0
									while md<(len(targetrtlist)):
										if targetmzlist[md]==tmzj:
											if targetrtlist[md]>(rangestart-targetstepwidth):
												if targetrtlist[md]<(rangeend+targetstepwidth):
													#print('A species was deleted from overlap of 4:')
													#print(targetmzlist[md])
													#print(targetrtlist[md]*60)
													del targetrtlist[md]	# delete j species at index md for rt list and mz list
													del targetmzlist[md]
													delgo=1
										if delgo==0:
											md=md+1
										else:
											delgo=0
									md=0
									delgo=0
									while md<(len(targetrtlist)):
										if targetmzlist[md]==tmzk:
											if targetrtlist[md]>(rangestart-targetstepwidth):
												if targetrtlist[md]<(rangeend+targetstepwidth):
													#print('A species was deleted from overlap of 4:')
													#print(targetmzlist[md])
													#print(targetrtlist[md]*60)
													del targetrtlist[md]	# delete k species at index md for rt list and mz list
													del targetmzlist[md]
													delgo=1
										if delgo==0:
											md=md+1
										else:
											delgo=0
									md=0
									delgo=0
									while md<(len(targetrtlist)):
										if targetmzlist[md]==tmzkf:
											if targetrtlist[md]>(rangestart-targetstepwidth):
												if targetrtlist[md]<(rangeend+targetstepwidth):
													#print('A species was deleted from overlap of 4:')
													#print(targetmzlist[md])
													#print(targetrtlist[md]*60)
													del targetrtlist[md]	# delete l species at index md for rt list and mz list
													del targetmzlist[md]
													delgo=1
										if delgo==0:
											md=md+1
										else:
											delgo=0
									ma=0
									altern=0
									while ma<(len(targetrtlist)):
										if targetmzlist[ma]==tmzi:
											if targetrtlist[ma]>(rangestart-targetstepwidth):
												if targetrtlist[ma]<(rangeend+targetstepwidth):
													if altern==0:
														altern=altern
													elif altern==1:
														targetmzlist[ma]=tmzj
													elif altern==2:
														targetmzlist[ma]=tmzk
													else:
														targetmzlist[ma]=tmzkf
													altern=altern+1
													if altern==4:
														altern=0
										ma=ma+1
									j=len(targetrtlist)
									i=len(targetrtlist)
								elif fifthfound==1:
									k=k # five or more entries at overlapping RT with different m/z # identify range, modify species within range
									tmzi=targetmzlist[i]
									tmzj=targetmzlist[j]
									kh=0
									sixthfound=0
									while kh<(len(targetrtlist)):
										if abs(targetrtlist[kh]-targetrtlist[i])<(targetstepwidth-0.000001):
											if targetmzlist[kh]==tmzi:
												kh=kh
											elif targetmzlist[kh]==tmzj:
												kh=kh
											elif targetmzlist[kh]==tmzk:
												kh=kh
											elif targetmzlist[kh]==tmzkf:
												kh=kh
											elif targetmzlist[kh]==tmzkg:
												kh=kh
											else:
												sixthfound=1
												tmzkh=targetmzlist[kh]
												#print('########################################################################### EDIT CODE FOR OVERLAP OF 6 #######')
										else:
											sixthfound=sixthfound
										kh=kh+1
									if sixthfound==0:
										k=k # five entries at overlapping RT with different m/z # identify range, modify species within range
										#print('Five species overlap....................................-5-...........')
										rangestart=targetrtlist[i]
										tmzi=targetmzlist[i]
										tmzj=targetmzlist[j]
										crange=rangestart+targetstepwidth
										mi=0
										igo=0
										jgo=0
										kgo=0
										lgo=0
										hgo=0
										zgo=1
										while mi<(len(targetrtlist)):
											if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
												if targetmzlist[mi]==tmzi:
													igo=1
												elif targetmzlist[mi]==tmzj:
													jgo=1
												elif targetmzlist[mi]==tmzk:
													kgo=1
												elif targetmzlist[mi]==tmzkf:
													lgo=1
												elif targetmzlist[mi]==tmzkg:
													hgo=1
												else:
													zgo=0
											else:
												mi=mi
											mi=mi+1
										if igo==1:
											if jgo==1:
												if kgo==1:
													if lgo==1:
														if hgo==1:
															if zgo==1:
																mgo=1
															else:
																mgo=0
														else:
															mgo=0
													else:
														mgo=0
												else:
													mgo=0
											else:
												mgo=0
										else:
											mgo=0
										while mgo==1:
											crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
											mi=0
											igo=0
											jgo=0
											kgo=0
											lgo=0
											hgo=0
											zgo=1
											while mi<(len(targetrtlist)):
												if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
													if targetmzlist[mi]==tmzi:
														igo=1
													elif targetmzlist[mi]==tmzj:
														jgo=1
													elif targetmzlist[mi]==tmzk:
														kgo=1
													elif targetmzlist[mi]==tmzkf:
														lgo=1
													elif targetmzlist[mi]==tmzkg:
														hgo=1
													else:
														zgo=0
												else:
													mi=mi
												mi=mi+1
											if igo==1:
												if jgo==1:
													if kgo==1:
														if lgo==1:
															if hgo==1:
																if zgo==1:
																	mgo=1
																else:
																	mgo=0
																	crange=crange-targetstepwidth
															else:
																mgo=0
																crange=crange-targetstepwidth
														else:
															mgo=0
															crange=crange-targetstepwidth
													else:
														mgo=0
														crange=crange-targetstepwidth
												else:
													mgo=0
													crange=crange-targetstepwidth
											else:
												mgo=0
												crange=crange-targetstepwidth
										rangeend=crange		#end of this range is defined now
										# begin delete all j and k species in range, reassign i species as alternating between i, j and k.
										md=0
										delgo=0
										while md<(len(targetrtlist)):
											if targetmzlist[md]==tmzj:
												if targetrtlist[md]>(rangestart-targetstepwidth):
													if targetrtlist[md]<(rangeend+targetstepwidth):
														#print('A species was deleted from overlap of 4:')
														#print(targetmzlist[md])
														#print(targetrtlist[md]*60)
														del targetrtlist[md]	# delete j species at index md for rt list and mz list
														del targetmzlist[md]
														delgo=1
											if delgo==0:
												md=md+1
											else:
												delgo=0
										md=0
										delgo=0
										while md<(len(targetrtlist)):
											if targetmzlist[md]==tmzk:
												if targetrtlist[md]>(rangestart-targetstepwidth):
													if targetrtlist[md]<(rangeend+targetstepwidth):
														#print('A species was deleted from overlap of 4:')
														#print(targetmzlist[md])
														#print(targetrtlist[md]*60)
														del targetrtlist[md]	# delete k species at index md for rt list and mz list
														del targetmzlist[md]
														delgo=1
											if delgo==0:
												md=md+1
											else:
												delgo=0
										md=0
										delgo=0
										while md<(len(targetrtlist)):
											if targetmzlist[md]==tmzkf:
												if targetrtlist[md]>(rangestart-targetstepwidth):
													if targetrtlist[md]<(rangeend+targetstepwidth):
														#print('A species was deleted from overlap of 4:')
														#print(targetmzlist[md])
														#print(targetrtlist[md]*60)
														del targetrtlist[md]	# delete l species at index md for rt list and mz list
														del targetmzlist[md]
														delgo=1
											if delgo==0:
												md=md+1
											else:
												delgo=0
										md=0
										delgo=0
										while md<(len(targetrtlist)):
											if targetmzlist[md]==tmzkg:
												if targetrtlist[md]>(rangestart-targetstepwidth):
													if targetrtlist[md]<(rangeend+targetstepwidth):
														#print('A species was deleted from overlap of 4:')
														#print(targetmzlist[md])
														#print(targetrtlist[md]*60)
														del targetrtlist[md]	# delete h species at index md for rt list and mz list
														del targetmzlist[md]
														delgo=1
											if delgo==0:
												md=md+1
											else:
												delgo=0
										ma=0
										altern=0
										while ma<(len(targetrtlist)):
											if targetmzlist[ma]==tmzi:
												if targetrtlist[ma]>(rangestart-targetstepwidth):
													if targetrtlist[ma]<(rangeend+targetstepwidth):
														if altern==0:
															altern=altern
														elif altern==1:
															targetmzlist[ma]=tmzj
														elif altern==2:
															targetmzlist[ma]=tmzk
														elif altern==3:
															targetmzlist[ma]=tmzkf
														else:
															targetmzlist[ma]=tmzkg
														altern=altern+1
														if altern==5:
															altern=0
											ma=ma+1
										j=len(targetrtlist)
										i=len(targetrtlist)
									elif sixthfound==1:
										k=k # five or more entries at overlapping RT with different m/z # identify range, modify species within range
										tmzi=targetmzlist[i]
										tmzj=targetmzlist[j]
										kh=0
										seventhfound=0
										while kh<(len(targetrtlist)):
											if abs(targetrtlist[kh]-targetrtlist[i])<(targetstepwidth-0.000001):
												if targetmzlist[kh]==tmzi:
													kh=kh
												elif targetmzlist[kh]==tmzj:
													kh=kh
												elif targetmzlist[kh]==tmzk:
													kh=kh
												elif targetmzlist[kh]==tmzkf:
													kh=kh
												elif targetmzlist[kh]==tmzkg:
													kh=kh
												elif targetmzlist[kh]==tmzkh:
													kh=kh
												else:
													seventhfound=1
													tmzki=targetmzlist[kh]
													#print('########################################################################### EDIT CODE FOR OVERLAP OF 7 #######')
											else:
												seventhfound=seventhfound
											kh=kh+1
										if seventhfound==0:
											k=k # six entries at overlapping RT with different m/z # identify range, modify species within range
											#print('Six species overlap....................................-6-...........')
											rangestart=targetrtlist[i]
											tmzi=targetmzlist[i]
											tmzj=targetmzlist[j]
											crange=rangestart+targetstepwidth
											mi=0
											igo=0
											jgo=0
											kgo=0
											lgo=0
											hgo=0
											kigo=0
											zgo=1
											while mi<(len(targetrtlist)):
												if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
													if targetmzlist[mi]==tmzi:
														igo=1
													elif targetmzlist[mi]==tmzj:
														jgo=1
													elif targetmzlist[mi]==tmzk:
														kgo=1
													elif targetmzlist[mi]==tmzkf:
														lgo=1
													elif targetmzlist[mi]==tmzkg:
														hgo=1
													elif targetmzlist[mi]==tmzkh:
														kigo=1
													else:
														zgo=0
												else:
													mi=mi
												mi=mi+1
											if igo==1:
												if jgo==1:
													if kgo==1:
														if lgo==1:
															if hgo==1:
																if kigo==1:
																	if zgo==1:
																		mgo=1
																	else:
																		mgo=0
																else:
																	mgo=0
															else:
																mgo=0
														else:
															mgo=0
													else:
														mgo=0
												else:
													mgo=0
											else:
												mgo=0
											while mgo==1:
												crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
												mi=0
												igo=0
												jgo=0
												kgo=0
												lgo=0
												hgo=0
												kigo=0
												zgo=1
												while mi<(len(targetrtlist)):
													if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
														if targetmzlist[mi]==tmzi:
															igo=1
														elif targetmzlist[mi]==tmzj:
															jgo=1
														elif targetmzlist[mi]==tmzk:
															kgo=1
														elif targetmzlist[mi]==tmzkf:
															lgo=1
														elif targetmzlist[mi]==tmzkg:
															hgo=1
														elif targetmzlist[mi]==tmzkh:
															kigo=1
														else:
															zgo=0
													else:
														mi=mi
													mi=mi+1
												if igo==1:
													if jgo==1:
														if kgo==1:
															if lgo==1:
																if hgo==1:
																	if kigo==1:
																		if zgo==1:
																			mgo=1
																		else:
																			mgo=0
																			crange=crange-targetstepwidth
																	else:
																		mgo=0
																		crange=crange-targetstepwidth	
																else:
																	mgo=0
																	crange=crange-targetstepwidth
															else:
																mgo=0
																crange=crange-targetstepwidth
														else:
															mgo=0
															crange=crange-targetstepwidth
													else:
														mgo=0
														crange=crange-targetstepwidth
												else:
													mgo=0
													crange=crange-targetstepwidth
											rangeend=crange		#end of this range is defined now
											# begin delete all j and k species in range, reassign i species as alternating between i, j and k.
											md=0
											delgo=0
											while md<(len(targetrtlist)):
												if targetmzlist[md]==tmzj:
													if targetrtlist[md]>(rangestart-targetstepwidth):
														if targetrtlist[md]<(rangeend+targetstepwidth):
															#print('A species was deleted from overlap of 4:')
															#print(targetmzlist[md])
															#print(targetrtlist[md]*60)
															del targetrtlist[md]	# delete j species at index md for rt list and mz list
															del targetmzlist[md]
															delgo=1
												if delgo==0:
													md=md+1
												else:
													delgo=0
											md=0
											delgo=0
											while md<(len(targetrtlist)):
												if targetmzlist[md]==tmzk:
													if targetrtlist[md]>(rangestart-targetstepwidth):
														if targetrtlist[md]<(rangeend+targetstepwidth):
															#print('A species was deleted from overlap of 4:')
															#print(targetmzlist[md])
															#print(targetrtlist[md]*60)
															del targetrtlist[md]	# delete k species at index md for rt list and mz list
															del targetmzlist[md]
															delgo=1
												if delgo==0:
													md=md+1
												else:
													delgo=0
											md=0
											delgo=0
											while md<(len(targetrtlist)):
												if targetmzlist[md]==tmzkf:
													if targetrtlist[md]>(rangestart-targetstepwidth):
														if targetrtlist[md]<(rangeend+targetstepwidth):
															#print('A species was deleted from overlap of 4:')
															#print(targetmzlist[md])
															#print(targetrtlist[md]*60)
															del targetrtlist[md]	# delete l species at index md for rt list and mz list
															del targetmzlist[md]
															delgo=1
												if delgo==0:
													md=md+1
												else:
													delgo=0
											md=0
											delgo=0
											while md<(len(targetrtlist)):
												if targetmzlist[md]==tmzkg:
													if targetrtlist[md]>(rangestart-targetstepwidth):
														if targetrtlist[md]<(rangeend+targetstepwidth):
															#print('A species was deleted from overlap of 4:')
															#print(targetmzlist[md])
															#print(targetrtlist[md]*60)
															del targetrtlist[md]	# delete h species at index md for rt list and mz list
															del targetmzlist[md]
															delgo=1
												if delgo==0:
													md=md+1
												else:
													delgo=0
											md=0
											delgo=0
											while md<(len(targetrtlist)):
												if targetmzlist[md]==tmzkh:
													if targetrtlist[md]>(rangestart-targetstepwidth):
														if targetrtlist[md]<(rangeend+targetstepwidth):
															#print('A species was deleted from overlap of 4:')
															#print(targetmzlist[md])
															#print(targetrtlist[md]*60)
															del targetrtlist[md]	# delete h species at index md for rt list and mz list
															del targetmzlist[md]
															delgo=1
												if delgo==0:
													md=md+1
												else:
													delgo=0
											ma=0
											altern=0
											while ma<(len(targetrtlist)):
												if targetmzlist[ma]==tmzi:
													if targetrtlist[ma]>(rangestart-targetstepwidth):
														if targetrtlist[ma]<(rangeend+targetstepwidth):
															if altern==0:
																altern=altern
															elif altern==1:
																targetmzlist[ma]=tmzj
															elif altern==2:
																targetmzlist[ma]=tmzk
															elif altern==3:
																targetmzlist[ma]=tmzkf
															elif altern==4:
																targetmzlist[ma]=tmzkg
															else:
																targetmzlist[ma]=tmzkh
															altern=altern+1
															if altern==6:
																altern=0
												ma=ma+1
											j=len(targetrtlist)
											i=len(targetrtlist)
										elif seventhfound==1:
											k=k # five or more entries at overlapping RT with different m/z # identify range, modify species within range
											tmzi=targetmzlist[i]
											tmzj=targetmzlist[j]
											kh=0
											eighthfound=0
											while kh<(len(targetrtlist)):
												if abs(targetrtlist[kh]-targetrtlist[i])<(targetstepwidth-0.000001):
													if targetmzlist[kh]==tmzi:
														kh=kh
													elif targetmzlist[kh]==tmzj:
														kh=kh
													elif targetmzlist[kh]==tmzk:
														kh=kh
													elif targetmzlist[kh]==tmzkf:
														kh=kh
													elif targetmzlist[kh]==tmzkg:
														kh=kh
													elif targetmzlist[kh]==tmzkh:
														kh=kh
													elif targetmzlist[kh]==tmzki:
														kh=kh
													else:
														eighthfound=1
														tmzkj=targetmzlist[kh]
														#print('########################################################################### EDIT CODE FOR OVERLAP OF 8 #######')
												else:
													eighthfound=eighthfound
												kh=kh+1
											if eighthfound==0:
												k=k # six entries at overlapping RT with different m/z # identify range, modify species within range
												#print('Six species overlap....................................-6-...........')
												rangestart=targetrtlist[i]
												tmzi=targetmzlist[i]
												tmzj=targetmzlist[j]
												crange=rangestart+targetstepwidth
												mi=0
												igo=0
												jgo=0
												kgo=0
												lgo=0
												hgo=0
												kigo=0
												kjgo=0
												zgo=1
												while mi<(len(targetrtlist)):
													if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
														if targetmzlist[mi]==tmzi:
															igo=1
														elif targetmzlist[mi]==tmzj:
															jgo=1
														elif targetmzlist[mi]==tmzk:
															kgo=1
														elif targetmzlist[mi]==tmzkf:
															lgo=1
														elif targetmzlist[mi]==tmzkg:
															hgo=1
														elif targetmzlist[mi]==tmzkh:
															kigo=1
														elif targetmzlist[mi]==tmzki:
															kjgo=1
														else:
															zgo=0
													else:
														mi=mi
													mi=mi+1
												if igo==1:
													if jgo==1:
														if kgo==1:
															if lgo==1:
																if hgo==1:
																	if kigo==1:
																		if kjgo==1:
																			if zgo==1:
																				mgo=1
																			else:
																				mgo=0
																		else:
																			mgo=0
																	else:
																		mgo=0
																else:
																	mgo=0
															else:
																mgo=0
														else:
															mgo=0
													else:
														mgo=0
												else:
													mgo=0
												while mgo==1:
													crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
													mi=0
													igo=0
													jgo=0
													kgo=0
													lgo=0
													hgo=0
													kigo=0
													kjgo=0
													zgo=1
													while mi<(len(targetrtlist)):
														if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
															if targetmzlist[mi]==tmzi:
																igo=1
															elif targetmzlist[mi]==tmzj:
																jgo=1
															elif targetmzlist[mi]==tmzk:
																kgo=1
															elif targetmzlist[mi]==tmzkf:
																lgo=1
															elif targetmzlist[mi]==tmzkg:
																hgo=1
															elif targetmzlist[mi]==tmzkh:
																kigo=1
															elif targetmzlist[mi]==tmzki:
																kjgo=1
															else:
																zgo=0
														else:
															mi=mi
														mi=mi+1
													if igo==1:
														if jgo==1:
															if kgo==1:
																if lgo==1:
																	if hgo==1:
																		if kigo==1:
																			if kjgo==1:
																				if zgo==1:
																					mgo=1
																				else:
																					mgo=0
																					crange=crange-targetstepwidth
																			else:
																				mgo=0
																				crange=crange-targetstepwidth
																		else:
																			mgo=0
																			crange=crange-targetstepwidth	
																	else:
																		mgo=0
																		crange=crange-targetstepwidth
																else:
																	mgo=0
																	crange=crange-targetstepwidth
															else:
																mgo=0
																crange=crange-targetstepwidth
														else:
															mgo=0
															crange=crange-targetstepwidth
													else:
														mgo=0
														crange=crange-targetstepwidth
												rangeend=crange		#end of this range is defined now
												# begin delete all j and k species in range, reassign i species as alternating between i, j and k.
												md=0
												delgo=0
												while md<(len(targetrtlist)):
													if targetmzlist[md]==tmzj:
														if targetrtlist[md]>(rangestart-targetstepwidth):
															if targetrtlist[md]<(rangeend+targetstepwidth):
																#print('A species was deleted from overlap of 4:')
																#print(targetmzlist[md])
																#print(targetrtlist[md]*60)
																del targetrtlist[md]	# delete j species at index md for rt list and mz list
																del targetmzlist[md]
																delgo=1
													if delgo==0:
														md=md+1
													else:
														delgo=0
												md=0
												delgo=0
												while md<(len(targetrtlist)):
													if targetmzlist[md]==tmzk:
														if targetrtlist[md]>(rangestart-targetstepwidth):
															if targetrtlist[md]<(rangeend+targetstepwidth):
																#print('A species was deleted from overlap of 4:')
																#print(targetmzlist[md])
																#print(targetrtlist[md]*60)
																del targetrtlist[md]	# delete k species at index md for rt list and mz list
																del targetmzlist[md]
																delgo=1
													if delgo==0:
														md=md+1
													else:
														delgo=0
												md=0
												delgo=0
												while md<(len(targetrtlist)):
													if targetmzlist[md]==tmzkf:
														if targetrtlist[md]>(rangestart-targetstepwidth):
															if targetrtlist[md]<(rangeend+targetstepwidth):
																#print('A species was deleted from overlap of 4:')
																#print(targetmzlist[md])
																#print(targetrtlist[md]*60)
																del targetrtlist[md]	# delete l species at index md for rt list and mz list
																del targetmzlist[md]
																delgo=1
													if delgo==0:
														md=md+1
													else:
														delgo=0
												md=0
												delgo=0
												while md<(len(targetrtlist)):
													if targetmzlist[md]==tmzkg:
														if targetrtlist[md]>(rangestart-targetstepwidth):
															if targetrtlist[md]<(rangeend+targetstepwidth):
																#print('A species was deleted from overlap of 4:')
																#print(targetmzlist[md])
																#print(targetrtlist[md]*60)
																del targetrtlist[md]	# delete h species at index md for rt list and mz list
																del targetmzlist[md]
																delgo=1
													if delgo==0:
														md=md+1
													else:
														delgo=0
												md=0
												delgo=0
												while md<(len(targetrtlist)):
													if targetmzlist[md]==tmzkh:
														if targetrtlist[md]>(rangestart-targetstepwidth):
															if targetrtlist[md]<(rangeend+targetstepwidth):
																#print('A species was deleted from overlap of 4:')
																#print(targetmzlist[md])
																#print(targetrtlist[md]*60)
																del targetrtlist[md]	# delete h species at index md for rt list and mz list
																del targetmzlist[md]
																delgo=1
													if delgo==0:
														md=md+1
													else:
														delgo=0
												md=0
												delgo=0
												while md<(len(targetrtlist)):
													if targetmzlist[md]==tmzki:
														if targetrtlist[md]>(rangestart-targetstepwidth):
															if targetrtlist[md]<(rangeend+targetstepwidth):
																#print('A species was deleted from overlap of 4:')
																#print(targetmzlist[md])
																#print(targetrtlist[md]*60)
																del targetrtlist[md]	# delete h species at index md for rt list and mz list
																del targetmzlist[md]
																delgo=1
													if delgo==0:
														md=md+1
													else:
														delgo=0
												ma=0
												altern=0
												while ma<(len(targetrtlist)):
													if targetmzlist[ma]==tmzi:
														if targetrtlist[ma]>(rangestart-targetstepwidth):
															if targetrtlist[ma]<(rangeend+targetstepwidth):
																if altern==0:
																	altern=altern
																elif altern==1:
																	targetmzlist[ma]=tmzj
																elif altern==2:
																	targetmzlist[ma]=tmzk
																elif altern==3:
																	targetmzlist[ma]=tmzkf
																elif altern==4:
																	targetmzlist[ma]=tmzkg
																elif altern==5:
																	targetmzlist[ma]=tmzkh
																else:
																	targetmzlist[ma]=tmzki
																altern=altern+1
																if altern==7:
																	altern=0
													ma=ma+1
												j=len(targetrtlist)
												i=len(targetrtlist)
											elif eighthfound==1:
												print('Checkpoint - Overlap of 8')
												k=k # eight or more entries at overlapping RT with different m/z # identify range, modify species within range
												tmzi=targetmzlist[i]
												tmzj=targetmzlist[j]
												kh=0
												ninthfound=0
												while kh<(len(targetrtlist)):
													if abs(targetrtlist[kh]-targetrtlist[i])<(targetstepwidth-0.000001):
														if targetmzlist[kh]==tmzi:
															kh=kh
														elif targetmzlist[kh]==tmzj:
															kh=kh
														elif targetmzlist[kh]==tmzk:
															kh=kh
														elif targetmzlist[kh]==tmzkf:
															kh=kh
														elif targetmzlist[kh]==tmzkg:
															kh=kh
														elif targetmzlist[kh]==tmzkh:
															kh=kh
														elif targetmzlist[kh]==tmzki:
															kh=kh
														elif targetmzlist[kh]==tmzkj:
															kh=kh
														else:
															ninthfound=1
															tmzkk=targetmzlist[kh]
															#print('############################################################## EDIT CODE FOR OVERLAP OF 9 #######')
													else:
														ninthfound=ninthfound
													kh=kh+1
												if ninthfound==0:
													k=k # six entries at overlapping RT with different m/z # identify range, modify species within range
													#print('Six species overlap....................................-6-...........')
													rangestart=targetrtlist[i]
													tmzi=targetmzlist[i]
													tmzj=targetmzlist[j]
													crange=rangestart+targetstepwidth
													mi=0
													igo=0
													jgo=0
													kgo=0
													lgo=0
													hgo=0
													kigo=0
													kjgo=0
													kkgo=0
													zgo=1
													while mi<(len(targetrtlist)):
														if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
															if targetmzlist[mi]==tmzi:
																igo=1
															elif targetmzlist[mi]==tmzj:
																jgo=1
															elif targetmzlist[mi]==tmzk:
																kgo=1
															elif targetmzlist[mi]==tmzkf:
																lgo=1
															elif targetmzlist[mi]==tmzkg:
																hgo=1
															elif targetmzlist[mi]==tmzkh:
																kigo=1
															elif targetmzlist[mi]==tmzki:
																kjgo=1
															elif targetmzlist[mi]==tmzkj:
																kkgo=1
															else:
																zgo=0
														else:
															mi=mi
														mi=mi+1
													if igo==1:
														if jgo==1:
															if kgo==1:
																if lgo==1:
																	if hgo==1:
																		if kigo==1:
																			if kjgo==1:
																				if kkgo==1:
																					if zgo==1:
																						mgo=1
																					else:
																						mgo=0
																				else:
																					mgo=0
																			else:
																				mgo=0
																		else:
																			mgo=0
																	else:
																		mgo=0
																else:
																	mgo=0
															else:
																mgo=0
														else:
															mgo=0
													else:
														mgo=0
													while mgo==1:
														crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
														mi=0
														igo=0
														jgo=0
														kgo=0
														lgo=0
														hgo=0
														kigo=0
														kjgo=0
														kkgo=0
														zgo=1
														while mi<(len(targetrtlist)):
															if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
																if targetmzlist[mi]==tmzi:
																	igo=1
																elif targetmzlist[mi]==tmzj:
																	jgo=1
																elif targetmzlist[mi]==tmzk:
																	kgo=1
																elif targetmzlist[mi]==tmzkf:
																	lgo=1
																elif targetmzlist[mi]==tmzkg:
																	hgo=1
																elif targetmzlist[mi]==tmzkh:
																	kigo=1
																elif targetmzlist[mi]==tmzki:
																	kjgo=1
																elif targetmzlist[mi]==tmzkj:
																	kkgo=1
																else:
																	zgo=0
															else:
																mi=mi
															mi=mi+1
														if igo==1:
															if jgo==1:
																if kgo==1:
																	if lgo==1:
																		if hgo==1:
																			if kigo==1:
																				if kjgo==1:
																					if kkgo==1:
																						if zgo==1:
																							mgo=1
																						else:
																							mgo=0
																							crange=crange-targetstepwidth
																					else:
																						mgo=0
																						crange=crange-targetstepwidth	
																				else:
																					mgo=0
																					crange=crange-targetstepwidth
																			else:
																				mgo=0
																				crange=crange-targetstepwidth	
																		else:
																			mgo=0
																			crange=crange-targetstepwidth
																	else:
																		mgo=0
																		crange=crange-targetstepwidth
																else:
																	mgo=0
																	crange=crange-targetstepwidth
															else:
																mgo=0
																crange=crange-targetstepwidth
														else:
															mgo=0
															crange=crange-targetstepwidth
													rangeend=crange		#end of this range is defined now
													# begin delete all j and k species in range, reassign i species as alternating between i, j and k.
													md=0
													delgo=0
													while md<(len(targetrtlist)):
														if targetmzlist[md]==tmzj:
															if targetrtlist[md]>(rangestart-targetstepwidth):
																if targetrtlist[md]<(rangeend+targetstepwidth):
																	#print('A species was deleted from overlap of 4:')
																	#print(targetmzlist[md])
																	#print(targetrtlist[md]*60)
																	del targetrtlist[md]	# delete j species at index md for rt list and mz list
																	del targetmzlist[md]
																	delgo=1
														if delgo==0:
															md=md+1
														else:
															delgo=0
													md=0
													delgo=0
													while md<(len(targetrtlist)):
														if targetmzlist[md]==tmzk:
															if targetrtlist[md]>(rangestart-targetstepwidth):
																if targetrtlist[md]<(rangeend+targetstepwidth):
																	#print('A species was deleted from overlap of 4:')
																	#print(targetmzlist[md])
																	#print(targetrtlist[md]*60)
																	del targetrtlist[md]	# delete k species at index md for rt list and mz list
																	del targetmzlist[md]
																	delgo=1
														if delgo==0:
															md=md+1
														else:
															delgo=0
													md=0
													delgo=0
													while md<(len(targetrtlist)):
														if targetmzlist[md]==tmzkf:
															if targetrtlist[md]>(rangestart-targetstepwidth):
																if targetrtlist[md]<(rangeend+targetstepwidth):
																	#print('A species was deleted from overlap of 4:')
																	#print(targetmzlist[md])
																	#print(targetrtlist[md]*60)
																	del targetrtlist[md]	# delete l species at index md for rt list and mz list
																	del targetmzlist[md]
																	delgo=1
														if delgo==0:
															md=md+1
														else:
															delgo=0
													md=0
													delgo=0
													while md<(len(targetrtlist)):
														if targetmzlist[md]==tmzkg:
															if targetrtlist[md]>(rangestart-targetstepwidth):
																if targetrtlist[md]<(rangeend+targetstepwidth):
																	#print('A species was deleted from overlap of 4:')
																	#print(targetmzlist[md])
																	#print(targetrtlist[md]*60)
																	del targetrtlist[md]	# delete h species at index md for rt list and mz list
																	del targetmzlist[md]
																	delgo=1
														if delgo==0:
															md=md+1
														else:
															delgo=0
													md=0
													delgo=0
													while md<(len(targetrtlist)):
														if targetmzlist[md]==tmzkh:
															if targetrtlist[md]>(rangestart-targetstepwidth):
																if targetrtlist[md]<(rangeend+targetstepwidth):
																	#print('A species was deleted from overlap of 4:')
																	#print(targetmzlist[md])
																	#print(targetrtlist[md]*60)
																	del targetrtlist[md]	# delete h species at index md for rt list and mz list
																	del targetmzlist[md]
																	delgo=1
														if delgo==0:
															md=md+1
														else:
															delgo=0
													md=0
													delgo=0
													while md<(len(targetrtlist)):
														if targetmzlist[md]==tmzki:
															if targetrtlist[md]>(rangestart-targetstepwidth):
																if targetrtlist[md]<(rangeend+targetstepwidth):
																	#print('A species was deleted from overlap of 4:')
																	#print(targetmzlist[md])
																	#print(targetrtlist[md]*60)
																	del targetrtlist[md]	# delete h species at index md for rt list and mz list
																	del targetmzlist[md]
																	delgo=1
														if delgo==0:
															md=md+1
														else:
															delgo=0
													md=0
													delgo=0
													while md<(len(targetrtlist)):
														if targetmzlist[md]==tmzkj:
															if targetrtlist[md]>(rangestart-targetstepwidth):
																if targetrtlist[md]<(rangeend+targetstepwidth):
																	#print('A species was deleted from overlap of 4:')
																	#print(targetmzlist[md])
																	#print(targetrtlist[md]*60)
																	del targetrtlist[md]	# delete h species at index md for rt list and mz list
																	del targetmzlist[md]
																	delgo=1
														if delgo==0:
															md=md+1
														else:
															delgo=0
													ma=0
													altern=0
													while ma<(len(targetrtlist)):
														if targetmzlist[ma]==tmzi:
															if targetrtlist[ma]>(rangestart-targetstepwidth):
																if targetrtlist[ma]<(rangeend+targetstepwidth):
																	if altern==0:
																		altern=altern
																	elif altern==1:
																		targetmzlist[ma]=tmzj
																	elif altern==2:
																		targetmzlist[ma]=tmzk
																	elif altern==3:
																		targetmzlist[ma]=tmzkf
																	elif altern==4:
																		targetmzlist[ma]=tmzkg
																	elif altern==5:
																		targetmzlist[ma]=tmzkh
																	else:
																		targetmzlist[ma]=tmzki
																	altern=altern+1
																	if altern==7:
																		altern=0
														ma=ma+1
													j=len(targetrtlist)
													i=len(targetrtlist)
												elif ninthfound==1:
													print('Checkpoint - Overlap of 9')
													k=k # nine or more entries at overlapping RT with different m/z # identify range, modify species within range
													tmzi=targetmzlist[i]
													tmzj=targetmzlist[j]
													kh=0
													tenthfound=0
													while kh<(len(targetrtlist)):
														if abs(targetrtlist[kh]-targetrtlist[i])<(targetstepwidth-0.000001):
															if targetmzlist[kh]==tmzi:
																kh=kh
															elif targetmzlist[kh]==tmzj:
																kh=kh
															elif targetmzlist[kh]==tmzk:
																kh=kh
															elif targetmzlist[kh]==tmzkf:
																kh=kh
															elif targetmzlist[kh]==tmzkg:
																kh=kh
															elif targetmzlist[kh]==tmzkh:
																kh=kh
															elif targetmzlist[kh]==tmzki:
																kh=kh
															elif targetmzlist[kh]==tmzkj:
																kh=kh
															elif targetmzlist[kh]==tmzkk:
																kh=kh
															else:
																tenthfound=1
																tmzkl=targetmzlist[kh]
																#print('############################ EDIT CODE FOR OVERLAP OF 10 #######')
														else:
															tenthfound=tenthfound
														kh=kh+1
													if tenthfound==0:
														k=k # six entries at overlapping RT with different m/z # identify range, modify species within range
														#print('Six species overlap....................................-6-...........')
														rangestart=targetrtlist[i]
														tmzi=targetmzlist[i]
														tmzj=targetmzlist[j]
														crange=rangestart+targetstepwidth
														mi=0
														igo=0
														jgo=0
														kgo=0
														lgo=0
														hgo=0
														kigo=0
														kjgo=0
														kkgo=0
														klgo=0
														zgo=1
														while mi<(len(targetrtlist)):
															if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
																if targetmzlist[mi]==tmzi:
																	igo=1
																elif targetmzlist[mi]==tmzj:
																	jgo=1
																elif targetmzlist[mi]==tmzk:
																	kgo=1
																elif targetmzlist[mi]==tmzkf:
																	lgo=1
																elif targetmzlist[mi]==tmzkg:
																	hgo=1
																elif targetmzlist[mi]==tmzkh:
																	kigo=1
																elif targetmzlist[mi]==tmzki:
																	kjgo=1
																elif targetmzlist[mi]==tmzkj:
																	kkgo=1
																elif targetmzlist[mi]==tmzkk:
																	klgo=1
																else:
																	zgo=0
															else:
																mi=mi
															mi=mi+1
														if igo==1:
															if jgo==1:
																if kgo==1:
																	if lgo==1:
																		if hgo==1:
																			if kigo==1:
																				if kjgo==1:
																					if kkgo==1:
																						if klgo==1:
																							if zgo==1:
																								mgo=1
																							else:
																								mgo=0
																						else:
																							mgo=0
																					else:
																						mgo=0
																				else:
																					mgo=0
																			else:
																				mgo=0
																		else:
																			mgo=0
																	else:
																		mgo=0
																else:
																	mgo=0
															else:
																mgo=0
														else:
															mgo=0
														while mgo==1:
															crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
															mi=0
															igo=0
															jgo=0
															kgo=0
															lgo=0
															hgo=0
															kigo=0
															kjgo=0
															kkgo=0
															klgo=0
															zgo=1
															while mi<(len(targetrtlist)):
																if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
																	if targetmzlist[mi]==tmzi:
																		igo=1
																	elif targetmzlist[mi]==tmzj:
																		jgo=1
																	elif targetmzlist[mi]==tmzk:
																		kgo=1
																	elif targetmzlist[mi]==tmzkf:
																		lgo=1
																	elif targetmzlist[mi]==tmzkg:
																		hgo=1
																	elif targetmzlist[mi]==tmzkh:
																		kigo=1
																	elif targetmzlist[mi]==tmzki:
																		kjgo=1
																	elif targetmzlist[mi]==tmzkj:
																		kkgo=1
																	elif targetmzlist[mi]==tmzkk:
																		klgo=1
																	else:
																		zgo=0
																else:
																	mi=mi
																mi=mi+1
															if igo==1:
																if jgo==1:
																	if kgo==1:
																		if lgo==1:
																			if hgo==1:
																				if kigo==1:
																					if kjgo==1:
																						if kkgo==1:
																							if klgo==1:
																								if zgo==1:
																									mgo=1
																								else:
																									mgo=0
																									crange=crange-targetstepwidth
																							else:
																								mgo=0
																								crange=crange-targetstepwidth
																						else:
																							mgo=0
																							crange=crange-targetstepwidth
																					else:
																						mgo=0
																						crange=crange-targetstepwidth
																				else:
																					mgo=0
																					crange=crange-targetstepwidth	
																			else:
																				mgo=0
																				crange=crange-targetstepwidth
																		else:
																			mgo=0
																			crange=crange-targetstepwidth
																	else:
																		mgo=0
																		crange=crange-targetstepwidth
																else:
																	mgo=0
																	crange=crange-targetstepwidth
															else:
																mgo=0
																crange=crange-targetstepwidth
														rangeend=crange		#end of this range is defined now
														# begin delete all j and k species in range, reassign i species as alternating between i, j and k.
														md=0
														delgo=0
														while md<(len(targetrtlist)):
															if targetmzlist[md]==tmzj:
																if targetrtlist[md]>(rangestart-targetstepwidth):
																	if targetrtlist[md]<(rangeend+targetstepwidth):
																		#print('A species was deleted from overlap of 4:')
																		#print(targetmzlist[md])
																		#print(targetrtlist[md]*60)
																		del targetrtlist[md]	# delete j species at index md for rt list and mz list
																		del targetmzlist[md]
																		delgo=1
															if delgo==0:
																md=md+1
															else:
																delgo=0
														md=0
														delgo=0
														while md<(len(targetrtlist)):
															if targetmzlist[md]==tmzk:
																if targetrtlist[md]>(rangestart-targetstepwidth):
																	if targetrtlist[md]<(rangeend+targetstepwidth):
																		#print('A species was deleted from overlap of 4:')
																		#print(targetmzlist[md])
																		#print(targetrtlist[md]*60)
																		del targetrtlist[md]	# delete k species at index md for rt list and mz list
																		del targetmzlist[md]
																		delgo=1
															if delgo==0:
																md=md+1
															else:
																delgo=0
														md=0
														delgo=0
														while md<(len(targetrtlist)):
															if targetmzlist[md]==tmzkf:
																if targetrtlist[md]>(rangestart-targetstepwidth):
																	if targetrtlist[md]<(rangeend+targetstepwidth):
																		#print('A species was deleted from overlap of 4:')
																		#print(targetmzlist[md])
																		#print(targetrtlist[md]*60)
																		del targetrtlist[md]	# delete l species at index md for rt list and mz list
																		del targetmzlist[md]
																		delgo=1
															if delgo==0:
																md=md+1
															else:
																delgo=0
														md=0
														delgo=0
														while md<(len(targetrtlist)):
															if targetmzlist[md]==tmzkg:
																if targetrtlist[md]>(rangestart-targetstepwidth):
																	if targetrtlist[md]<(rangeend+targetstepwidth):
																		#print('A species was deleted from overlap of 4:')
																		#print(targetmzlist[md])
																		#print(targetrtlist[md]*60)
																		del targetrtlist[md]	# delete h species at index md for rt list and mz list
																		del targetmzlist[md]
																		delgo=1
															if delgo==0:
																md=md+1
															else:
																delgo=0
														md=0
														delgo=0
														while md<(len(targetrtlist)):
															if targetmzlist[md]==tmzkh:
																if targetrtlist[md]>(rangestart-targetstepwidth):
																	if targetrtlist[md]<(rangeend+targetstepwidth):
																		#print('A species was deleted from overlap of 4:')
																		#print(targetmzlist[md])
																		#print(targetrtlist[md]*60)
																		del targetrtlist[md]	# delete h species at index md for rt list and mz list
																		del targetmzlist[md]
																		delgo=1
															if delgo==0:
																md=md+1
															else:
																delgo=0
														md=0
														delgo=0
														while md<(len(targetrtlist)):
															if targetmzlist[md]==tmzki:
																if targetrtlist[md]>(rangestart-targetstepwidth):
																	if targetrtlist[md]<(rangeend+targetstepwidth):
																		#print('A species was deleted from overlap of 4:')
																		#print(targetmzlist[md])
																		#print(targetrtlist[md]*60)
																		del targetrtlist[md]	# delete h species at index md for rt list and mz list
																		del targetmzlist[md]
																		delgo=1
															if delgo==0:
																md=md+1
															else:
																delgo=0
														md=0
														delgo=0
														while md<(len(targetrtlist)):
															if targetmzlist[md]==tmzkj:
																if targetrtlist[md]>(rangestart-targetstepwidth):
																	if targetrtlist[md]<(rangeend+targetstepwidth):
																		#print('A species was deleted from overlap of 4:')
																		#print(targetmzlist[md])
																		#print(targetrtlist[md]*60)
																		del targetrtlist[md]	# delete h species at index md for rt list and mz list
																		del targetmzlist[md]
																		delgo=1
															if delgo==0:
																md=md+1
															else:
																delgo=0
														md=0
														delgo=0
														while md<(len(targetrtlist)):
															if targetmzlist[md]==tmzkk:
																if targetrtlist[md]>(rangestart-targetstepwidth):
																	if targetrtlist[md]<(rangeend+targetstepwidth):
																		#print('A species was deleted from overlap of 4:')
																		#print(targetmzlist[md])
																		#print(targetrtlist[md]*60)
																		del targetrtlist[md]	# delete h species at index md for rt list and mz list
																		del targetmzlist[md]
																		delgo=1
															if delgo==0:
																md=md+1
															else:
																delgo=0
														ma=0
														altern=0
														while ma<(len(targetrtlist)):
															if targetmzlist[ma]==tmzi:
																if targetrtlist[ma]>(rangestart-targetstepwidth):
																	if targetrtlist[ma]<(rangeend+targetstepwidth):
																		if altern==0:
																			altern=altern
																		elif altern==1:
																			targetmzlist[ma]=tmzj
																		elif altern==2:
																			targetmzlist[ma]=tmzk
																		elif altern==3:
																			targetmzlist[ma]=tmzkf
																		elif altern==4:
																			targetmzlist[ma]=tmzkg
																		elif altern==5:
																			targetmzlist[ma]=tmzkh
																		else:
																			targetmzlist[ma]=tmzki
																		altern=altern+1
																		if altern==7:
																			altern=0
															ma=ma+1
														j=len(targetrtlist)
														i=len(targetrtlist)
													elif tenthfound==1:	###1        ##############################################################################################
														print('Checkpoint - Overlap of 10')
														k=k # nine or more entries at overlapping RT with different m/z # identify range, modify species within range ############
														tmzi=targetmzlist[i]
														tmzj=targetmzlist[j]
														kh=0
														eleventhfound=0
														while kh<(len(targetrtlist)):
															if abs(targetrtlist[kh]-targetrtlist[i])<(targetstepwidth-0.000001):
																if targetmzlist[kh]==tmzi:
																	kh=kh
																elif targetmzlist[kh]==tmzj:
																	kh=kh
																elif targetmzlist[kh]==tmzk:
																	kh=kh
																elif targetmzlist[kh]==tmzkf:
																	kh=kh
																elif targetmzlist[kh]==tmzkg:
																	kh=kh
																elif targetmzlist[kh]==tmzkh:
																	kh=kh
																elif targetmzlist[kh]==tmzki:
																	kh=kh
																elif targetmzlist[kh]==tmzkj:
																	kh=kh
																elif targetmzlist[kh]==tmzkk:
																	kh=kh
																elif targetmzlist[kh]==tmzkl:	###2
																	kh=kh
																else:
																	eleventhfound=1		###3
																	tmzkm=targetmzlist[kh]	###4
																	print('############################ EDIT CODE FOR OVERLAP OF 11 #######') ###5
															else:
																eleventhfound=eleventhfound 	###6
															kh=kh+1
														if eleventhfound==0:	###7
															k=k # six entries at overlapping RT with different m/z # identify range, modify species within range
															#print('Six species overlap....................................-6-...........')
															rangestart=targetrtlist[i]
															tmzi=targetmzlist[i]
															tmzj=targetmzlist[j]
															crange=rangestart+targetstepwidth
															mi=0
															igo=0
															jgo=0
															kgo=0
															lgo=0
															hgo=0
															kigo=0
															kjgo=0
															kkgo=0
															klgo=0
															kmgo=0	###8
															zgo=1
															while mi<(len(targetrtlist)):
																if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
																	if targetmzlist[mi]==tmzi:
																		igo=1
																	elif targetmzlist[mi]==tmzj:
																		jgo=1
																	elif targetmzlist[mi]==tmzk:
																		kgo=1
																	elif targetmzlist[mi]==tmzkf:
																		lgo=1
																	elif targetmzlist[mi]==tmzkg:
																		hgo=1
																	elif targetmzlist[mi]==tmzkh:
																		kigo=1
																	elif targetmzlist[mi]==tmzki:
																		kjgo=1
																	elif targetmzlist[mi]==tmzkj:
																		kkgo=1
																	elif targetmzlist[mi]==tmzkk:
																		klgo=1
																	elif targetmzlist[mi]==tmzkl:	###9
																		kmgo=1			###10
																	else:
																		zgo=0
																else:
																	mi=mi
																mi=mi+1
															if igo==1:
																if jgo==1:
																	if kgo==1:
																		if lgo==1:
																			if hgo==1:
																				if kigo==1:
																					if kjgo==1:
																						if kkgo==1:
																							if klgo==1:
																								if kmgo==1:
																									if zgo==1:		###11
																										mgo=1
																									else:
																										mgo=0
																								else:
																									mgo=0
																							else:
																								mgo=0
																						else:
																							mgo=0
																					else:
																						mgo=0
																				else:
																					mgo=0
																			else:
																				mgo=0
																		else:
																			mgo=0
																	else:
																		mgo=0
																else:
																	mgo=0
															else:
																mgo=0
															while mgo==1:
																crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
																mi=0
																igo=0
																jgo=0
																kgo=0
																lgo=0
																hgo=0
																kigo=0
																kjgo=0
																kkgo=0
																klgo=0
																kmgo=0	###12
																zgo=1
																while mi<(len(targetrtlist)):
																	if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
																		if targetmzlist[mi]==tmzi:
																			igo=1
																		elif targetmzlist[mi]==tmzj:
																			jgo=1
																		elif targetmzlist[mi]==tmzk:
																			kgo=1
																		elif targetmzlist[mi]==tmzkf:
																			lgo=1
																		elif targetmzlist[mi]==tmzkg:
																			hgo=1
																		elif targetmzlist[mi]==tmzkh:
																			kigo=1
																		elif targetmzlist[mi]==tmzki:
																			kjgo=1
																		elif targetmzlist[mi]==tmzkj:
																			kkgo=1
																		elif targetmzlist[mi]==tmzkk:
																			klgo=1
																		elif targetmzlist[mi]==tmzkl:	###13
																			kmgo=1
																		else:
																			zgo=0
																	else:
																		mi=mi
																	mi=mi+1
																if igo==1:
																	if jgo==1:
																		if kgo==1:
																			if lgo==1:
																				if hgo==1:
																					if kigo==1:
																						if kjgo==1:
																							if kkgo==1:
																								if klgo==1:
																									if kmgo==1:
																										if zgo==1:	###14
																											mgo=1
																										else:
																											mgo=0
																											crange=crange-targetstepwidth
																									else:
																										mgo=0
																										crange=crange-targetstepwidth
																								else:
																									mgo=0
																									crange=crange-targetstepwidth
																							else:
																								mgo=0
																								crange=crange-targetstepwidth
																						else:
																							mgo=0
																							crange=crange-targetstepwidth
																					else:
																						mgo=0
																						crange=crange-targetstepwidth	
																				else:
																					mgo=0
																					crange=crange-targetstepwidth
																			else:
																				mgo=0
																				crange=crange-targetstepwidth
																		else:
																			mgo=0
																			crange=crange-targetstepwidth
																	else:
																		mgo=0
																		crange=crange-targetstepwidth
																else:
																	mgo=0
																	crange=crange-targetstepwidth
															rangeend=crange		#end of this range is defined now
															# begin delete all j and k species in range, reassign i species as alternating between i, j and k.
															md=0
															delgo=0
															while md<(len(targetrtlist)):
																if targetmzlist[md]==tmzj:
																	if targetrtlist[md]>(rangestart-targetstepwidth):
																		if targetrtlist[md]<(rangeend+targetstepwidth):
																			#print('A species was deleted from overlap of 4:')
																			#print(targetmzlist[md])
																			#print(targetrtlist[md]*60)
																			del targetrtlist[md]	# delete j species at index md for rt list and mz list
																			del targetmzlist[md]
																			delgo=1
																if delgo==0:
																	md=md+1
																else:
																	delgo=0
															md=0
															delgo=0
															while md<(len(targetrtlist)):
																if targetmzlist[md]==tmzk:
																	if targetrtlist[md]>(rangestart-targetstepwidth):
																		if targetrtlist[md]<(rangeend+targetstepwidth):
																			#print('A species was deleted from overlap of 4:')
																			#print(targetmzlist[md])
																			#print(targetrtlist[md]*60)
																			del targetrtlist[md]	# delete k species at index md for rt list and mz list
																			del targetmzlist[md]
																			delgo=1
																if delgo==0:
																	md=md+1
																else:
																	delgo=0
															md=0
															delgo=0
															while md<(len(targetrtlist)):
																if targetmzlist[md]==tmzkf:
																	if targetrtlist[md]>(rangestart-targetstepwidth):
																		if targetrtlist[md]<(rangeend+targetstepwidth):
																			#print('A species was deleted from overlap of 4:')
																			#print(targetmzlist[md])
																			#print(targetrtlist[md]*60)
																			del targetrtlist[md]	# delete l species at index md for rt list and mz list
																			del targetmzlist[md]
																			delgo=1
																if delgo==0:
																	md=md+1
																else:
																	delgo=0
															md=0
															delgo=0
															while md<(len(targetrtlist)):
																if targetmzlist[md]==tmzkg:
																	if targetrtlist[md]>(rangestart-targetstepwidth):
																		if targetrtlist[md]<(rangeend+targetstepwidth):
																			#print('A species was deleted from overlap of 4:')
																			#print(targetmzlist[md])
																			#print(targetrtlist[md]*60)
																			del targetrtlist[md]	# delete h species at index md for rt list and mz list
																			del targetmzlist[md]
																			delgo=1
																if delgo==0:
																	md=md+1
																else:
																	delgo=0
															md=0
															delgo=0
															while md<(len(targetrtlist)):
																if targetmzlist[md]==tmzkh:
																	if targetrtlist[md]>(rangestart-targetstepwidth):
																		if targetrtlist[md]<(rangeend+targetstepwidth):
																			#print('A species was deleted from overlap of 4:')
																			#print(targetmzlist[md])
																			#print(targetrtlist[md]*60)
																			del targetrtlist[md]	# delete h species at index md for rt list and mz list
																			del targetmzlist[md]
																			delgo=1
																if delgo==0:
																	md=md+1
																else:
																	delgo=0
															md=0
															delgo=0
															while md<(len(targetrtlist)):
																if targetmzlist[md]==tmzki:
																	if targetrtlist[md]>(rangestart-targetstepwidth):
																		if targetrtlist[md]<(rangeend+targetstepwidth):
																			#print('A species was deleted from overlap of 4:')
																			#print(targetmzlist[md])
																			#print(targetrtlist[md]*60)
																			del targetrtlist[md]	# delete h species at index md for rt list and mz list
																			del targetmzlist[md]
																			delgo=1
																if delgo==0:
																	md=md+1
																else:
																	delgo=0
															md=0
															delgo=0
															while md<(len(targetrtlist)):
																if targetmzlist[md]==tmzkj:
																	if targetrtlist[md]>(rangestart-targetstepwidth):
																		if targetrtlist[md]<(rangeend+targetstepwidth):
																			#print('A species was deleted from overlap of 4:')
																			#print(targetmzlist[md])
																			#print(targetrtlist[md]*60)
																			del targetrtlist[md]	# delete h species at index md for rt list and mz list
																			del targetmzlist[md]
																			delgo=1
																if delgo==0:
																	md=md+1
																else:
																	delgo=0
															md=0
															delgo=0
															while md<(len(targetrtlist)):
																if targetmzlist[md]==tmzkk:
																	if targetrtlist[md]>(rangestart-targetstepwidth):
																		if targetrtlist[md]<(rangeend+targetstepwidth):
																			#print('A species was deleted from overlap of 4:')
																			#print(targetmzlist[md])
																			#print(targetrtlist[md]*60)
																			del targetrtlist[md]	# delete h species at index md for rt list and mz list
																			del targetmzlist[md]
																			delgo=1
																if delgo==0:
																	md=md+1
																else:
																	delgo=0
															md=0
															delgo=0
															while md<(len(targetrtlist)):
																if targetmzlist[md]==tmzkl:		###15
																	if targetrtlist[md]>(rangestart-targetstepwidth):
																		if targetrtlist[md]<(rangeend+targetstepwidth):
																			#print('A species was deleted from overlap of 4:')
																			#print(targetmzlist[md])
																			#print(targetrtlist[md]*60)
																			del targetrtlist[md]	# delete h species at index md for rt list and mz list
																			del targetmzlist[md]
																			delgo=1
																if delgo==0:
																	md=md+1
																else:
																	delgo=0
															ma=0
															altern=0
															while ma<(len(targetrtlist)):
																if targetmzlist[ma]==tmzi:
																	if targetrtlist[ma]>(rangestart-targetstepwidth):
																		if targetrtlist[ma]<(rangeend+targetstepwidth):
																			if altern==0:
																				altern=altern
																			elif altern==1:
																				targetmzlist[ma]=tmzj
																			elif altern==2:
																				targetmzlist[ma]=tmzk
																			elif altern==3:
																				targetmzlist[ma]=tmzkf
																			elif altern==4:
																				targetmzlist[ma]=tmzkg
																			elif altern==5:
																				targetmzlist[ma]=tmzkh
																			else:
																				targetmzlist[ma]=tmzki
																			altern=altern+1
																			if altern==7:
																				altern=0
																ma=ma+1
															j=len(targetrtlist)
															i=len(targetrtlist)
														elif eleventhfound==1:	###1        #########################################################################################
															print('Checkpoint - Overlap of 11')
															k=k # nine or more entries at overlapping RT with different m/z # identify range, modify species within range ############
															tmzi=targetmzlist[i]
															tmzj=targetmzlist[j]
															kh=0
															twelvethfound=0
															while kh<(len(targetrtlist)):
																if abs(targetrtlist[kh]-targetrtlist[i])<(targetstepwidth-0.000001):
																	if targetmzlist[kh]==tmzi:
																		kh=kh
																	elif targetmzlist[kh]==tmzj:
																		kh=kh
																	elif targetmzlist[kh]==tmzk:
																		kh=kh
																	elif targetmzlist[kh]==tmzkf:
																		kh=kh
																	elif targetmzlist[kh]==tmzkg:
																		kh=kh
																	elif targetmzlist[kh]==tmzkh:
																		kh=kh
																	elif targetmzlist[kh]==tmzki:
																		kh=kh
																	elif targetmzlist[kh]==tmzkj:
																		kh=kh
																	elif targetmzlist[kh]==tmzkk:
																		kh=kh
																	elif targetmzlist[kh]==tmzkl:	
																		kh=kh
																	elif targetmzlist[kh]==tmzkm:	###2
																		kh=kh
																	else:
																		twelvethfound=1		###3
																		tmzkn=targetmzlist[kh]	###4
																		print('############################ EDIT CODE FOR OVERLAP OF 12 #######') ###5
																else:
																	twelvethfound=twelvethfound 	###6
																kh=kh+1
															if twelvethfound==0:	###7
																k=k # six entries at overlapping RT with different m/z # identify range, modify species within range
																#print('Six species overlap....................................-6-...........')
																rangestart=targetrtlist[i]
																tmzi=targetmzlist[i]
																tmzj=targetmzlist[j]
																crange=rangestart+targetstepwidth
																mi=0
																igo=0
																jgo=0
																kgo=0
																lgo=0
																hgo=0
																kigo=0
																kjgo=0
																kkgo=0
																klgo=0
																kmgo=0	###8
																kngo=0
																zgo=1
																while mi<(len(targetrtlist)):
																	if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
																		if targetmzlist[mi]==tmzi:
																			igo=1
																		elif targetmzlist[mi]==tmzj:
																			jgo=1
																		elif targetmzlist[mi]==tmzk:
																			kgo=1
																		elif targetmzlist[mi]==tmzkf:
																			lgo=1
																		elif targetmzlist[mi]==tmzkg:
																			hgo=1
																		elif targetmzlist[mi]==tmzkh:
																			kigo=1
																		elif targetmzlist[mi]==tmzki:
																			kjgo=1
																		elif targetmzlist[mi]==tmzkj:
																			kkgo=1
																		elif targetmzlist[mi]==tmzkk:
																			klgo=1
																		elif targetmzlist[mi]==tmzkl:	###9
																			kmgo=1			###10
																		elif targetmzlist[mi]==tmzkm:	###9
																			kngo=1			###10
																		else:
																			zgo=0
																	else:
																		mi=mi
																	mi=mi+1
																if igo==1:
																	if jgo==1:
																		if kgo==1:
																			if lgo==1:
																				if hgo==1:
																					if kigo==1:
																						if kjgo==1:
																							if kkgo==1:
																								if klgo==1:
																									if kmgo==1:
																										if kngo==1:
																											if zgo==1:		###11
																												mgo=1
																											else:
																												mgo=0
																										else:
																											mgo=0
																									else:
																										mgo=0
																								else:
																									mgo=0
																							else:
																								mgo=0
																						else:
																							mgo=0
																					else:
																						mgo=0
																				else:
																					mgo=0
																			else:
																				mgo=0
																		else:
																			mgo=0
																	else:
																		mgo=0
																else:
																	mgo=0
																while mgo==1:
																	crange=crange+targetstepwidth # check if condition still met for this m, if not, then range is defined.
																	mi=0
																	igo=0
																	jgo=0
																	kgo=0
																	lgo=0
																	hgo=0
																	kigo=0
																	kjgo=0
																	kkgo=0
																	klgo=0
																	kmgo=0	###12
																	kngo=0
																	zgo=1
																	while mi<(len(targetrtlist)):
																		if abs(targetrtlist[mi]-crange)<(targetstepwidth-0.000001):
																			if targetmzlist[mi]==tmzi:
																				igo=1
																			elif targetmzlist[mi]==tmzj:
																				jgo=1
																			elif targetmzlist[mi]==tmzk:
																				kgo=1
																			elif targetmzlist[mi]==tmzkf:
																				lgo=1
																			elif targetmzlist[mi]==tmzkg:
																				hgo=1
																			elif targetmzlist[mi]==tmzkh:
																				kigo=1
																			elif targetmzlist[mi]==tmzki:
																				kjgo=1
																			elif targetmzlist[mi]==tmzkj:
																				kkgo=1
																			elif targetmzlist[mi]==tmzkk:
																				klgo=1
																			elif targetmzlist[mi]==tmzkl:	###13
																				kmgo=1
																			elif targetmzlist[mi]==tmzkm:	###13
																				kngo=1
																			else:
																				zgo=0
																		else:
																			mi=mi
																		mi=mi+1
																	if igo==1:
																		if jgo==1:
																			if kgo==1:
																				if lgo==1:
																					if hgo==1:
																						if kigo==1:
																							if kjgo==1:
																								if kkgo==1:
																									if klgo==1:
																										if kmgo==1:
																											if kngo==1:
																												if zgo==1:	###14
																													mgo=1
																												else:
																													mgo=0
																													crange=crange-targetstepwidth
																											else:
																												mgo=0
																												crange=crange-targetstepwidth
																										else:
																											mgo=0
																											crange=crange-targetstepwidth
																									else:
																										mgo=0
																										crange=crange-targetstepwidth
																								else:
																									mgo=0
																									crange=crange-targetstepwidth
																							else:
																								mgo=0
																								crange=crange-targetstepwidth
																						else:
																							mgo=0
																							crange=crange-targetstepwidth	
																					else:
																						mgo=0
																						crange=crange-targetstepwidth
																				else:
																					mgo=0
																					crange=crange-targetstepwidth
																			else:
																				mgo=0
																				crange=crange-targetstepwidth
																		else:
																			mgo=0
																			crange=crange-targetstepwidth
																	else:
																		mgo=0
																		crange=crange-targetstepwidth
																rangeend=crange		#end of this range is defined now
																# begin delete all j and k species in range, reassign i species as alternating between i, j and k.
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzj:
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete j species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzk:
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete k species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzkf:
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete l species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzkg:
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete h species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzkh:
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete h species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzki:
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete h species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzkj:
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete h species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzkk:
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete h species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzkl:		###15
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete h species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																md=0
																delgo=0
																while md<(len(targetrtlist)):
																	if targetmzlist[md]==tmzkm:		###15
																		if targetrtlist[md]>(rangestart-targetstepwidth):
																			if targetrtlist[md]<(rangeend+targetstepwidth):
																				#print('A species was deleted from overlap of 4:')
																				#print(targetmzlist[md])
																				#print(targetrtlist[md]*60)
																				del targetrtlist[md]	# delete h species at index md for rt list and mz list
																				del targetmzlist[md]
																				delgo=1
																	if delgo==0:
																		md=md+1
																	else:
																		delgo=0
																ma=0
																altern=0
																while ma<(len(targetrtlist)):
																	if targetmzlist[ma]==tmzi:
																		if targetrtlist[ma]>(rangestart-targetstepwidth):
																			if targetrtlist[ma]<(rangeend+targetstepwidth):
																				if altern==0:
																					altern=altern
																				elif altern==1:
																					targetmzlist[ma]=tmzj
																				elif altern==2:
																					targetmzlist[ma]=tmzk
																				elif altern==3:
																					targetmzlist[ma]=tmzkf
																				elif altern==4:
																					targetmzlist[ma]=tmzkg
																				elif altern==5:
																					targetmzlist[ma]=tmzkh
																				else:
																					targetmzlist[ma]=tmzki
																				altern=altern+1
																				if altern==7:
																					altern=0
																	ma=ma+1
																j=len(targetrtlist)
																i=len(targetrtlist)





			j=j+1
		i=i+1
	if found==1:
		go=1
	else:
		go=0

ltgt=len(targetrtlist)					
print('Target list after primary reassignment contains %d targets.' % ltgt)
#print(ltgt)
# end sort and delete and reassign species and RTs to prevent multiple species assignment at same RT range


# begin sort targetlist and vary m/z values at last digit (targetmzlist, targetrtlist)
ntargetmzlist=[]
ntargetrtlist=[]
t=0
while t<(len(targetmzlist)):
	found=0
	count=0
	i=0
	while i<(len(ntargetmzlist)):
		if abs(float(targetmzlist[t])-float(ntargetmzlist[i]))<0.0001:
			count=count+1
			if targetrtlist[t]==ntargetrtlist[i]:
				found=1
		i=i+1
	if found==0:
		if (len(ntargetrtlist))==0:
			ccrt=targetrtlist[t]
			ccmz=float(targetmzlist[t])+(0.000001*count)
			ntargetrtlist.append(ccrt)
			ntargetmzlist.append(ccmz)
		else:
			#print(count)
			i=0
			while i<(len(ntargetmzlist)):
				if targetrtlist[t]>ntargetrtlist[i]:
					if (i+1)==(len(ntargetrtlist)):
						ccrt=targetrtlist[t]
						ccmz=float(targetmzlist[t])+(0.000001*count)
						ntargetrtlist.append(ccrt)
						ntargetmzlist.append(ccmz)
						i=(len(ntargetmzlist))
					elif targetrtlist[t]<ntargetrtlist[i+1]:
						ccrt=targetrtlist[t]
						ccmz=float(targetmzlist[t])+(0.000001*count)
						ntargetrtlist.insert((i+1), ccrt)
						ntargetmzlist.insert((i+1), ccmz)
						i=(len(ntargetmzlist))
				else:
					ccrt=targetrtlist[t]
					ccmz=float(targetmzlist[t])+(0.000001*count)
					ntargetrtlist.insert((i), ccrt)
					ntargetmzlist.insert((i), ccmz)
					i=(len(ntargetmzlist))
				i=i+1
	else:
		print('*************')
	t=t+1

#print(targetrtlist)
#print(ntargetrtlist)
#print(targetmzlist)
#print(ntargetmzlist)
#print('Number of targets in sorted target list for DDA run:')
#ltgt=len(ntargetrtlist)
#print(ltgt)
# end sort targetlist and vary m/z values at last digit

# begin secondary reassignment of targetlist to balance target frequencies
# begin detect missing target regions
# go through raw targetlist and check, if closest nearby reassigned target further than maxdt sec away, 
# --> identified as missing target, initiate swap from highest frequency target
print('Begin secondary reassignment.')
# rawtargetmzlist
# rawtargetrtlist	(raw targetlist)

# targetmzlist
# targetrtlist	(primary reassigned targetlist)
maxdt=2.1 #2.1	# max delta t [sec]; maximum allowed time between raw and closest reassigned target		# 2.1 sec allows approx 7 different targets in a row
swapcount=0

rwtg=0
while rwtg<(len(rawtargetmzlist)):
	#print('__________________________________')
	#print(rawtargetmzlist[rwtg])
	#print(rawtargetrtlist[rwtg])
	#scan=1
	scantg=0
	cltg=0 #index for closest target in reassigned targetlist
	dtct=1000	# delta t [in sec] to closest reassigned target
	while scantg<(len(ntargetmzlist)):
		if abs(float(rawtargetmzlist[rwtg])-float(ntargetmzlist[scantg]))<0.01:
			if abs((60*float(rawtargetrtlist[rwtg]))-(60*float(ntargetrtlist[scantg])))<dtct:
				dtct=abs((60*float(rawtargetrtlist[rwtg]))-(60*float(ntargetrtlist[scantg])))
				cltg=scantg
		scantg=scantg+1
	#print(dtct)
	if dtct>maxdt:
		#print('__________________________________')
		#print(dtct)
		#print('Found a raw target that is not represented well in reassigned targetlist:')
		#print('In raw targetlist, not found or none close enough in reassigned targetlist:')
		#print(rawtargetmzlist[rwtg])
		#print(60*rawtargetrtlist[rwtg])
		#print('Closest target in reassigned list')
		#print(ntargetmzlist[cltg])
		#print(60*ntargetrtlist[cltg])
		#if dtct=100:
			# raw target not at all present in reassigned targetlist
		#else:
			# find target in reassigned targetlist that can be swapped
		scanrt=0 #index of target in reassigned targetlist closest to target in rawtargetlist that needs to be introduced
		drt=100
		while scanrt<(len(ntargetrtlist)):
			if abs((60*float(rawtargetrtlist[rwtg]))-(60*float(ntargetrtlist[scanrt])))<drt:
				drt=abs((60*float(rawtargetrtlist[rwtg]))-(60*float(ntargetrtlist[scanrt])))
				clrt=scanrt
			scanrt=scanrt+1
		#print(clrt)	# index of retention time of target closest to the one in raw targetlist that needs to be included
		ranktglist=[]	# contains number/rank of associated targets
		ranklist=[] # contains suitable m/z
		rtg=clrt-7
		while rtg<(clrt+7):
			rank=1
			if rtg<0:
				rtg=0
			if rtg>len(ntargetmzlist)-1:
				ok=1
			else:
				ntt=float(ntargetmzlist[rtg])
				ranktglist.append(ntt)
				ranklist.append(rank)
			rcheck=0
			while rcheck<(len(ranklist)-1):
				if abs(float(ranktglist[rcheck])-float(ranktglist[len(ranklist)-1]))<0.01:
					if float(ranklist[rcheck])==float(ranklist[len(ranklist)-1]):
						if int(ranklist[rcheck])==1:
							ranklist[rcheck]=int(ranklist[rcheck])+1
							ranklist[len(ranklist)-1]=int(ranklist[len(ranklist)-1])+1
					elif float(ranklist[rcheck])>float(ranklist[len(ranklist)-1]):
						ranklist[rcheck]=int(ranklist[rcheck])+1
						ranklist[len(ranklist)-1]=int(ranklist[rcheck])
					elif float(ranklist[rcheck])<float(ranklist[len(ranklist)-1]):
						ranklist[rcheck]=int(ranklist[len(ranklist)-1])
				rcheck=rcheck+1
				# need to extend this list building module?
			rtg=rtg+1
		#print(ranklist)
		#print(ranktglist)
		irank=7
		goir=1
		ci=1
		swapped=0
		while goir==1:
			# ENTER condition for irank out of range
			if ranklist[irank]==max(ranklist):
				goir=0
				# found best target in reassigned targetlist to be swapped out
				ntargetmzlist[(clrt+(irank-7))]=float(rawtargetmzlist[rwtg])+0.0001
				#print('Swapped one.')
				swapcount=swapcount+1
				swapped=1
			else:
				irank=irank+ci
				if irank<(len(ranklist)):
					if irank>0:
						ok=1
					else:
						goir=0
						if swapped==0:
							irank=irank-ci
							ntargetmzlist[(clrt+(irank-7))]=float(rawtargetmzlist[rwtg])+0.0001
							#print('Swapped one.')
							swapcount=swapcount+1
							swapped=1
				else:
					goir=0
					if swapped==0:
						irank=irank-ci
						ntargetmzlist[(clrt+(irank-7))]=float(rawtargetmzlist[rwtg])+0.0001
						#print('Swapped one.')
						swapcount=swapcount+1
						swapped=1
				ci=ci*(-1)
				if ci>0:
					ci=ci+1
				else:
					ci=ci-1
			
	rwtg=rwtg+1
# end secondary reassignment of targetlist to balance target frequencies
#print('Secondary reassignment complete.')
print('During secondary reassignment, %d targets were reassigned.' % swapcount)

## begin modify target list with 13C isotope labels									##########   ISOTOPE LABEL IN TARGETLIST
isotopelabelled=0
if isotopelabelled==1:
	ntm=0
	ntr=0
	while ntm<(len(ntargetmzlist)):
		if abs(ntargetmzlist[ntm]-451.3683)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		elif abs(ntargetmzlist[ntm]-423.337)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(16*1.00335)
		elif abs(ntargetmzlist[ntm]-395.3057)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(14*1.00335)
		elif abs(ntargetmzlist[ntm]-449.3526)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		elif abs(ntargetmzlist[ntm]-421.3213)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(16*1.00335)
		elif abs(ntargetmzlist[ntm]-393.29)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(14*1.00335)
		elif abs(ntargetmzlist[ntm]-447.337)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		elif abs(ntargetmzlist[ntm]-445.3213)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		elif abs(ntargetmzlist[ntm]-477.3839)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		elif abs(ntargetmzlist[ntm]-475.3683)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		elif abs(ntargetmzlist[ntm]-473.3526)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		elif abs(ntargetmzlist[ntm]-471.3370)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		elif abs(ntargetmzlist[ntm]-503.3996)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		elif abs(ntargetmzlist[ntm]-531.4309)<0.01:
			ntargetmzlist[ntm]=ntargetmzlist[ntm]+(18*1.00335)
		ntm=ntm+1

## end modify target list with 13C isotope labels

# begin prepare data for writing into target list file; create .txt file and write target list into it
txtwritelist=[]
#targetrtlist=[9.11, 8.79, 7.2, 7.23]
#targetmzlist=[477.383900, 477.383901, 447.3370, 495.3370]   
cl=0
while cl<(len(ntargetrtlist)):
    txtwritelist.append(['', '-1.000000', '-1.000000', '0', '', '-1.000000', '-1.000000','1'])
    cl=cl+1
cl=0
while cl<(len(ntargetrtlist)):
    txtwritelist[cl][0]=str(format(float(ntargetmzlist[cl]),'.6f'))
    txtwritelist[cl][4]=str(format((60*ntargetrtlist[cl]),'.6f'))
    cl=cl+1
#print(txtwritelist)
after=datetime.datetime.now()
after=str(after)
today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_'+fourlettcode+'_'
filename=today+'jpmlipidomics_dda_targetlist.txt'
path='C:/Users/menzel2/batchprogramming/OzFAD1' #'C:/Users/menzel2/pythonprogramming'
file=filename
with open(os.path.join(path, file), 'w') as fp:
    pass
    j=0
    while j<(len(txtwritelist)):
        i=0
        while i<(len(txtwritelist[j])):
            fp.write(str(txtwritelist[j][i]))
            if i<(len(txtwritelist[j])-1):
                fp.write(',')
            else:
                fp.write('\n')
            i=i+1
        j=j+1
genfilename='jpmlipidomics_dda_targetlist.txt'
path='C:/Users/menzel2/batchprogramming/OzFAD1' #'C:/Users/menzel2/pythonprogramming'
file=genfilename
with open(os.path.join(path, file), 'w') as fp:
    pass
    j=0
    while j<(len(txtwritelist)):
        i=0
        while i<(len(txtwritelist[j])):
            fp.write(str(txtwritelist[j][i]))
            if i<(len(txtwritelist[j])-1):
                fp.write(',')
            else:
                fp.write('\n')
            i=i+1
        j=j+1
print('Final targetlist for DDA LC-OzID-MS run is saved as a .txt file.')
#quit()
# end prepare data for writing into target list file; create .txt file and write target list into it


# end calculate target list 
############################################################################################################################################
############################################################################################################################################
################ DATABASE ## Source: Internetchemie.info
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.7, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
################
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]

# begin calculate monounsaturated precursors from input
#maxlenfa=24 #longest expected fatty acid
#minlenfa=12	#shortest expected fatty acid
minlenfa=transferlist[8] #eval(input('Enter number of C atoms in shortest FA chain (e.g. 4) :'))
maxlenfa=transferlist[9] #eval(input('Enter number of C atoms in longest FA chain (e.g. 24) :'))

print('Calculation of transition list is running ...')
lfa=maxlenfa-minlenfa+1
moleculegrouplist=[]
precursornamelist=[]
precursorformulalist=[]
precursoradductlist=[]
precursormzlist=[]
precursorchargelist=[]
productnamelist=[]
productformulalist=[]
productadductlist=[]
productmzlist=[]
productchargelist=[]
molg=fourlettcode+'_FA'
padd='[M]1+'
prdn='precursor'
li=0
while li<lfa:
	moleculegrouplist.append(molg)
	currentlfa=str(li+minlenfa)
	if len(currentlfa)<2:
		currentlfa='0'+currentlfa
		currentlfa=str(currentlfa)
	pnm=fourlettcode+'_'+currentlfa+':1'
	precursornamelist.append(pnm)
	prf=''
	currentcderiv=0
	currenthderiv=0
	currentdderiv=0
	currentnderiv=0
	currentoderiv=0
	currentpderiv=0
	currentideriv=0
	if cderiv>0:
		currentcderiv=cderiv+int(currentlfa)
		prf=prf+'C'+str(currentcderiv)
	if hderiv>0:
		currenthderiv=hderiv+(2*int(currentlfa))-3
		prf=prf+'H'+str(currenthderiv)
	if dderiv>0:
		currentdderiv=dderiv
		prf=prf+"H'"+str(currentdderiv)
	if nderiv>0:
		currentnderiv=nderiv
		prf=prf+'N'+str(nderiv)
	if oderiv>(-1):
		currentoderiv=oderiv+1
		prf=prf+'O'+str(currentoderiv)
	if pderiv>0:
		currentpderiv=pderiv
		prf=prf+'P'+str(pderiv)
	if ideriv>0:
		currentideriv=ideriv
		prf=prf+'I'+str(ideriv)
	precursorformulalist.append(prf)
	productformulalist.append(prf)
	precursoradductlist.append(padd)
	prmz=imass[0]*currenthderiv+imass[1]*currentdderiv+imass[2]*currentcderiv+imass[3]*currentnderiv+imass[4]*currentoderiv+imass[5]*currentpderiv+imass[10]*currentideriv
	precursormzlist.append(prmz)
	productmzlist.append(prmz)
	precursorchargelist.append(1)
	productnamelist.append(prdn)
	productadductlist.append(padd)
	productchargelist.append(1)
	li=li+1
fareadlist=[]
fareadlist.append(moleculegrouplist)
fareadlist.append(precursornamelist)
fareadlist.append(precursorformulalist)
fareadlist.append(precursoradductlist)
fareadlist.append(precursormzlist)
fareadlist.append(precursorchargelist)
fareadlist.append(productnamelist)
fareadlist.append(productformulalist)
fareadlist.append(productadductlist)
fareadlist.append(productmzlist)
fareadlist.append(productchargelist)
# end calculate monounsaturated precursors from input

preconly=0
#minrt=0.0
#maxrt=eval(input('At which retention time [min] ends gradient? :'))
#explrt=(minrt+maxrt)/2
#explrtw=explrt-minrt
nspec=3	# number of species: precursor, aldehyde, criegee
# begin create empty lists
toprow=[]
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]		# double bond position list, contains strings with double bond position n-1, n-2 ...
dbindexlist=[]	# double bond index list, contains index for the double bond position that is closest to head group of FA
#end create empty lists
#create lists for saturated FAs
toprow=['Moleculegroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
ki=len(fareadlist[0])
satlist=fareadlist
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

r=0
while r<ki:		#go through rows of fareadlist
	# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
	cchain=0
	e=str(fareadlist[1][r]) # Precursorname
	i=len(e)-3
	add=e[i]
	add=float(add)
	cchain=cchain+add
	i=i-1
	add=e[i]
	add=float(add)
	if add>0:
		cchain=cchain+(10*add)
	else:
		cchain=cchain
		#print('Please check source code (determine number of C atoms in chain)')
	maxdbp=cchain-2
	# begin determine number of C atoms in chain, define highest possible double bond position
	kadd=nspec*maxdbp#+1		#(nspec=3 precursor and 2 products - aldehyde and crigee  - for each of maxdbp possible double bond positions)

	e=fareadlist[0][r]  # MoleculeGroup
	f=fareadlist[1][r] 	# Precursorname
	g=fareadlist[2][r] 	# Precursorformula
	h=float(fareadlist[5][r]) # PrecursorCharge
	k=0
	while k<kadd:
		mlistname.append(e)		# copied, no change
		precname.append(f)		# copied, no change
		precformula.append(g)	# copied, no change
		precchrg.append(h)		# copied, no change
		k=k+1

	e=fareadlist[3][r] 	# Precursoradduct
	k=0
	while k<kadd:
		precadduct.append(e) 	# AMPP, precursor
		precadduct.append(e)	# AMPP, aldehyde product
		precadduct.append(e)	# AMPP, crigee product
		k=k+nspec

	e=float(fareadlist[4][r]) 	# PrecursorMz
	k=0
	while k<kadd:
		precmz.append(e)	# precursor
		precmz.append(e)	# aldehyde
		precmz.append(e)	# crigee
		k=k+nspec

	e=fareadlist[1][r] 	# Productname
	if e=='Cholesterol (+[2]H7)':
		fragment='_'	
	else:	
		k=0
		csub=1
		while k<kadd:
			dbp='_n-'+str(csub)
			fragment='_precursor'
			ne=e+dbp+fragment
			prodname.append(ne)		# precursor
			dbl=[]	#begin save double bond position for later
			dbl.append(csub)
			dbindexlist.append(dbl)	#end save double bond position for later
			dbp='_n-'+str(csub)
			fragment='_aldehyde neutral loss'
			ne=e+dbp+fragment
			dbpi=0
			while dbpi<nspec:
				dbplist.append(dbp)
				dbpi=dbpi+1
			#
			if preconly==0:	
				prodname.append(ne)		# aldehyde
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				#prodname.append(ne)
				#prodname.append(ne)
				fragment='_criegee neutral loss'
				ne=e+dbp+fragment
				prodname.append(ne)		# crigee
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				#prodname.append(ne)
				#prodname.append(ne)
			if preconly==1:
				dbp='_n-'+str(csub)
				fragment='_dummy precursor a'
				ne=e+dbp+fragment
				prodname.append(ne)		# dummy precursor
				fragment='_dummy precursor b'
				ne=e+dbp+fragment
				prodname.append(ne)		# dummy precursor
				dbl=[]	#begin save double bond position for later
				dbl.append(csub)
				dbindexlist.append(dbl)
				dbindexlist.append(dbl)	#end save double bond position for later
			csub=csub+1
			k=k+nspec

	e=fareadlist[2][r] 
	# begin read precursor sum formula and edit product sum formula
	#print(e)
	#print(e[0])
	#print(e[1])
	#print(e[2])
	#print(e[3])
	#print(len(e))
	clist=[]
	hlist=[]
	dlist=[]
	nlist=[]
	olist=[]
	plist=[]
	ilist=[]
	i=0
	ca=0
	ha=0
	da=0
	na=0
	oa=0
	pa=0
	ia=0
	while i<len(e):
		if e[i]=='H':
			if e[i+1]=="'":
				ha=0
			else:
				ca=0
		#if e[i]=='D':
		#	ha=0		
		if e[i]=='N':
			ha=0
			da=0
		if e[i]=='O':
			ha=0
			da=0
			na=0
		if e[i]=='P':
			ha=0
			da=0
			na=0
			oa=0
		if e[i]=='I':
			ha=0
			da=0
			na=0
			oa=0
			pa=0
		if ca==1:
			clist.append(e[i])
		if ha==1:
			hlist.append(e[i])
		if da==1:
			dlist.append(e[i])
		if na==1:
			nlist.append(e[i])
		if oa==1:
			olist.append(e[i])
		if pa==1:
			plist.append(e[i])
		if ia==1:
			ilist.append(e[i])
		if e[i]=='C':
			ca=1
		if e[i]=='H':
			if e[i+1]=="'":
				ca=0
				ha=0
				da=1
				i=i+1
			else:
				ca=0
				ha=1
		#if e[i]=='D':
		#	ca=0
		#	ha=0
		#	da=1		
		if e[i]=='N':
			ha=0
			da=0
			na=1
			if e[i+1]=='O':
				nlist.append('1')
				na=0
		if e[i]=='O':
			ha=0
			da=0
			na=0
			oa=1
			if (i+1)<len(e):
				if e[i+1]=='P':
					olist.append('1')
					oa=0
			else:
				olist.append('1')
				oa=0					
		if e[i]=='P':
			da=0
			na=0
			oa=0
			pa=1
			if (i+1)<len(e):
				if e[i+1]=='I':
					plist.append('1')
					pa=0
			else:
				plist.append('1')
				pa=0
		if e[i]=='I':
			da=0
			na=0
			oa=0
			pa=0
			ia=1
			if i==(len(e)-1):
				ilist.append('1')
				ia=0
		i=i+1
	#print(clist)
	#print(hlist)
	#print(dlist)
	#print(nlist)
	#print(olist)
	#print(plist)
	if len(clist)==0:
		cn=0
	if len(hlist)==0:
		hn=0
	if len(dlist)==0:
		dn=0	
	if len(nlist)==0:
		nn=0
	if len(olist)==0:
		on=0
	if len(plist)==0:
		pn=0
	if len(ilist)==0:
		iodon=0
	if len(clist)==1:
		cn=int(clist[0])
	if len(clist)==2:
		cn=10*int(clist[0])+int(clist[1])
	if len(clist)==3:
		cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
	if len(hlist)==1:
		hn=int(hlist[0])
	if len(hlist)==2:
		hn=10*int(hlist[0])+int(hlist[1])
	if len(hlist)==3:
		hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
	if len(dlist)==1:
		dn=int(dlist[0])
	if len(dlist)==2:
		dn=10*int(dlist[0])+int(dlist[1])
	if len(dlist)==3:
		dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
	if len(nlist)==1:
		nn=int(nlist[0])
	if len(nlist)==2:
		nn=10*int(nlist[0])+int(nlist[1])
	if len(nlist)==3:
		nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
	if len(olist)==1:
		on=int(olist[0])
	if len(olist)==2:
		on=10*int(olist[0])+int(olist[1])
	if len(olist)==3:
		on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
	if len(plist)==1:
		pn=int(plist[0])
	if len(plist)==2:
		pn=10*int(plist[0])+int(plist[1])
	if len(plist)==3:
		pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
	if len(ilist)==1:
		iodon=int(ilist[0])
	if len(ilist)==2:
		iodon=10*int(ilist[0])+int(ilist[1])
	if len(ilist)==3:
		iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
	#print(cn)
	e=fareadlist[0][r] 	# begin calculate product sum formula
	if e=='SPLASH':
		e=e 
	else:
		k=0
		csub=1
		while k<kadd:
			cnp=cn-(csub)
			hnp=hn-(2*csub)
			onald=on+1
			oncrigee=on+2
			precursor='C'+str(cn)+'H'+str(hn)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
			productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
			productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
			prodformula.append(precursor)
			if preconly==0:
				prodformula.append(productaldehyde)
				prodformula.append(productcrigee)
			if preconly==1:
				prodformula.append(precursor)
				prodformula.append(precursor)
			precursormz=imass[0]*(hn)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
			prodmz.append(precursormz)
			if preconly==0:
				productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, AMPP
				prodmz.append(productmz)
				productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, AMPP
				prodmz.append(productmz)
			if preconly==1:
				prodmz.append(precursormz)
				prodmz.append(precursormz)
			csub=csub+1
			k=k+nspec			
	# end read precursor sum formula and edit product sum formula

	prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)

	e=fareadlist[0][r] 	# begin calculate product m/z
	if e=='SPLASH':
		print('SPLASH')	# end calculate product m/z for SPLASH
	else:
		e=e 	# no action, as product m/z was calculated from ProductFormula and saved respectively (see above)
	# end calculate product m/z
	prodchrg=precchrg 	# ProductCharge

	r=r+1
	e=fareadlist[1][r-1]	# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	if r>(len(fareadlist[0])-1):
		r=r+1
	else:
		e=fareadlist[1][r] 
		prec=e
		while prec==prevprec:
			r=r+1
			e=fareadlist[1][r-1] 
			prevprec=e
			e=fareadlist[1][r] 
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
writelist=[]
writelist.append(mlistname)
writelist.append(precname)
writelist.append(precformula)
writelist.append(precadduct)
writelist.append(precmz)
writelist.append(precchrg)
writelist.append(prodname)
writelist.append(prodformula)
writelist.append(prodadduct)
writelist.append(prodmz)
writelist.append(prodchrg)
#end save data in writelist

#print('All calculations for monounsaturated fatty acids are done.')
#end save excel file
#print ('odd : C_%d H_%d N_%d O_%d S_%d; DBE = %d; deviation: %.3f' % (formula[1], formula[0], formula[2], formula[3], formula[4], dbe, meandeviation))
###########################################################################################################################################################################
#############################################################################MONO##########################################################################################
############################################################################DOUBLE#########################################################################################
###########################################################################################################################################################################
# begin add double unsaturated fatty acids
#ask=eval(input('Add fatty acids with two double bonds? Yes: 1; No: 0 | '))
ask=1
if ask==0:
	quit()

nspec=5	# number of species: precursor, aldehyde db1 and db2, crigee db1 and db2
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])
kmono=ki # List index in excel file to start writing FA transitions with two double bonds
kdoublestart=kmono
r=0
ki=ki
while r<ki:		#go through rows of excel file jpmlipidozidinput
	# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
	cchain=0
	e=str(writelist[1][r])#sheetinput.cell(row=r, column=2)	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	if len(e)==13:
		dbpmono=int(e[12])
	elif len(e)==14:
		dbpmono=(10*(int(e[12])))+(int(e[13]))
	else:
		print('Please check source code (determine double bond position of already located double bond)')
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the second double bond
	csubmono=dbpmono+2
	if nsecdbp>0:
		kadd=nspec*nsecdbp#+1		#(nspec=3 products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			k=k+1

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
		degunsat=int(e[8])		# change from monounsaturated to bisunsaturated FA in PrecursorName
		if degunsat==1:
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(2)
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# copied, no change
			k=k+1

		e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
		k=0
		while k<kadd:
			precadduct.append(e) 	# AMPP, precursor
			precadduct.append(e)	# AMPP, aldehyde product
			precadduct.append(e)	# AMPP, crigee product
			precadduct.append(e)	# AMPP, aldehyde product, outer db
			precadduct.append(e)	# AMPP, crigee product, outer db
			k=k+nspec

		e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added second double bond
		k=0
		while k<kadd:
			precmz.append(e)	# precursor
			precmz.append(e)	# aldehyde
			precmz.append(e)	# crigee
			precmz.append(e)	# aldehyde outer db
			precmz.append(e)	# crigee outer db
			k=k+nspec

		e=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			precchrg.append(e)
			k=k+1

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change from monounsaturated to bisunsaturated FA in PrecursorName
		if degunsat==1:
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(2)
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from monounsaturated to bisunsaturated FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond position for later
				firstdbp=dbindexlist[r][0]
				dbl.append(firstdbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					prodname.append(ne)		# aldehyde
					dbl=[]	#begin save double bond position for later
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	#end save double bond position for later
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde outer double bond
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee outer double bond
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor a'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor b'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor c'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor d'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
					dbl=[]
					dbl.append(firstdbp)
					dbl.append(csub)
					dbindexlist.append(dbl)	# save double bond position for later
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #sheetinput.cell(row=r, column=3)
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=='D':
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=='D':
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with two double bonds
				hnpmono=hnprec-(2*dbpmono)	# applies to cleavage of outer double bond
				cnpmono=cn-(dbpmono)		# applies to cleavage of outer double bond
				cnp=cn-(csub)				# applies to cleavage of inner double bond
				hnp=hn-(2*csub)				# applies to cleavage of inner double bond
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydemono='C'+str(cnpmono)+'H'+str(hnpmono)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeemono='C'+str(cnpmono)+'H'+str(hnpmono)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydemono)
					prodformula.append(productcrigeemono)	
				if preconly==1:
					prodformula.append(precursor)	
					prodformula.append(precursor)
					prodformula.append(precursor)
					prodformula.append(precursor)
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, inner db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, inner db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpmono)+imass[1]*dn+imass[2]*cnpmono+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, inner db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpmono)+imass[1]*dn+imass[2]*cnpmono+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, inner db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data as writelist, merge new lists at end of each list in writelist
writelist[0]=writelist[0]+mlistname
writelist[1]=writelist[1]+precname
writelist[2]=writelist[2]+precformula
writelist[3]=writelist[3]+precadduct
writelist[4]=writelist[4]+precmz
writelist[5]=writelist[5]+precchrg
writelist[6]=writelist[6]+prodname
writelist[7]=writelist[7]+prodformula
writelist[8]=writelist[8]+prodadduct
writelist[9]=writelist[9]+prodmz
writelist[10]=writelist[10]+prodchrg
#end save data as writelist

#print('All calculations for fatty acids with two double bonds are done.')
#end save excel file
# end add double unsaturated fatty acids
###########################################################################################################################################################################
###########################################################################DOUBLE##########################################################################################
###########################################################################TRIPLE##########################################################################################
###########################################################################################################################################################################
# begin add triple unsaturated fatty acids
#print(len(dbindexlist))
#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
ask=1
if ask==0:
	quit()

nspec=7	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])

kmono=ki # List index in writelist to start writing FA transitions with three double bonds
ktriplestart=kmono
r=kdoublestart
ki=ki
while r<ki:		#go through rows of FAs with two double bonds in excel file
	# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
	cchain=0
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	#maxdbp=
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	#if len(e)==14:
	#	dbpmono=int(e[13])
	#elif len(e)==15:
	#	dbpmono=(10*(int(e[13])))+(int(e[14]))
	#else:
	#	print('Please check source code (determine double bond position of already located double bond)')
	dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond		########################################## r instead of r-2 ##################
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
	csubmono=dbpmono+2
	if nsecdbp>0:

		kadd=nspec*nsecdbp#+1		#(nspec=3 products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		f=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			precchrg.append(f)	# copied, no change
			k=k+1	

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
		degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==2:
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(3)
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# write new precursorname, changed to FA with three db
			k=k+1

		e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precadduct.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added second double bond
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precmz.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==2:
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(3)
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from bisunsaturated to trisunsaturated FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond position for later 						
				firstdbp=dbindexlist[r][0]
				seconddbp=dbindexlist[r][1] ########################################## r instead of r-2 ##################
				dbl.append(firstdbp)
				dbl.append(seconddbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde last existing double bond
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee last existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor1'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor2'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor3'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor4'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor5'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor6'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]	########################################## r instead of r-2 ##################
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #PrecursorFormula
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=='D':
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=='D':
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
				cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
				hnp=hnprec-(2*csub-2*2)		# applies to cleavage of first (highest n) double bond
				hnptwo=hnprec-(2*dbpmono-2*1)	# applies to cleavage of second double bond
				cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
				hnpthree=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of third double bond
				cnpthree=cn-(dbindexlist[r][0])			# applies to cleavage of third double bond
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydetwo)
					prodformula.append(productcrigeetwo)
					prodformula.append(productaldehydethree)
					prodformula.append(productcrigeethree)	
				if preconly==1:
					prodformula.append(precursor)	
					prodformula.append(precursor)
					prodformula.append(precursor)
					prodformula.append(precursor)
					prodformula.append(precursor)
					prodformula.append(precursor)
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
writelist[0]=writelist[0]+mlistname
writelist[1]=writelist[1]+precname
writelist[2]=writelist[2]+precformula
writelist[3]=writelist[3]+precadduct
writelist[4]=writelist[4]+precmz
writelist[5]=writelist[5]+precchrg
writelist[6]=writelist[6]+prodname
writelist[7]=writelist[7]+prodformula
writelist[8]=writelist[8]+prodadduct
writelist[9]=writelist[9]+prodmz
writelist[10]=writelist[10]+prodchrg
#end save data in writelist

#print('All calculations for fatty acids with three double bonds are done.')
#end save excel file
# end add triple unsaturated fatty acids
######################################################################################################################################################
##############################################################TRIPLE##################################################################################

##############################################################FOUR DB#################################################################################
######################################################################################################################################################
# begin add triple unsaturated fatty acids
#print(len(dbindexlist))
#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
ask=1
if ask==0:
	quit()

nspec=9	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3  ############################ VARY ########################
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])

#print('Entries in excel file after bisunsaturated FA transitions are generated: %d' % ki)
kmono=ki # List index in excel file to start writing FA transitions with four double bonds
kfourstart=kmono
r=ktriplestart ################################################################################################### VARY ########################
ki=ki#+2
while r<ki:		#go through rows of FAs with three double bonds in excel file
	# begin determine number of C atoms in chain, define highest possible double bond position ###########################################################################
	cchain=0
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	#maxdbp=
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	#if len(e)==14:
	#	dbpmono=int(e[13])
	#elif len(e)==15:
	#	dbpmono=(10*(int(e[13])))+(int(e[14]))
	#else:
	#	print('Please check source code (determine double bond position of already located double bond)')
	dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond        ######################################### CHECK ?
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
	csubmono=dbpmono+2
	if nsecdbp>0:
		kadd=nspec*nsecdbp#+1		#(nspec= number of products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		f=float(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			precchrg.append(f)	# copied, no change
			k=k+1	

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
		degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==3:			####################################### VARY ################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(4)		####################################### VARY ################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# write new precursorname, changed to FA with four db
			k=k+1

		e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precadduct.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=float(writelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added double bond
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precmz.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change FA in PrecursorName to generate ProductName
		if degunsat==3:			########################################## VARY ######################################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(4)	########################################## VARY ######################################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from degree of unsaturation of FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond position for later 						
				firstdbp=dbindexlist[r][0]
				seconddbp=dbindexlist[r][1]
				thirddbp=dbindexlist[r][2]
				dbl.append(firstdbp)
				dbl.append(seconddbp)
				dbl.append(thirddbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond position for later
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde last existing double bond
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee last existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][1])
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][1])
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor1'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor2'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor3'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor4'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor5'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor6'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor7'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor8'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #sheetinput.cell(row=r, column=3)
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=='D':
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=='D':
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
				cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
				hnp=hnprec-(2*csub-2*3)		# applies to cleavage of first (highest n) double bond
				hnptwo=hnprec-(2*dbpmono-2*2)	# applies to cleavage of second double bond
				cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
				hnpthree=hnprec-(2*(dbindexlist[r][1])-2*1)	# applies to cleavage of third double bond
				cnpthree=cn-(dbindexlist[r][1])			# applies to cleavage of third double bond
				hnpfour=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				cnpfour=cn-(dbindexlist[r][0])			# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydetwo)
					prodformula.append(productcrigeetwo)
					prodformula.append(productaldehydethree)
					prodformula.append(productcrigeethree)
					prodformula.append(productaldehydefour)
					prodformula.append(productcrigeefour)	
				if preconly==1:
					t=1
					while t<nspec:
						prodformula.append(precursor)
						t=t+1
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fourth db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
writelist[0]=writelist[0]+mlistname
writelist[1]=writelist[1]+precname
writelist[2]=writelist[2]+precformula
writelist[3]=writelist[3]+precadduct
writelist[4]=writelist[4]+precmz
writelist[5]=writelist[5]+precchrg
writelist[6]=writelist[6]+prodname
writelist[7]=writelist[7]+prodformula
writelist[8]=writelist[8]+prodadduct
writelist[9]=writelist[9]+prodmz
writelist[10]=writelist[10]+prodchrg
#end save data in writelist

#print('All calculations for fatty acids with four double bonds are done.')
#end save excel file
# end add triple unsaturated fatty acids
######################################################################################################################################################
##############################################################FOUR DB#################################################################################
##############################################################FIVE DB#################################################################################
######################################################################################################################################################
# begin add triple unsaturated fatty acids
#print(len(dbindexlist))
#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
ask=1
if ask==0:
	quit()

nspec=11	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3  ############################ VARY ########################
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])

#print('Entries in excel file after bisunsaturated FA transitions are generated: %d' % ki)
kmono=ki # List index in excel file to start writing FA transitions with four double bonds
kfivestart=kmono
r=kfourstart ################################################################################################### VARY ########################
if workflowmode==2:
	r=ki#+2
while r<ki:		#go through rows of FAs with three double bonds in excel file
	# begin determine number of C atoms in chain, define highest possible double bond position ###############
	cchain=0
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	#maxdbp=
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	#if len(e)==14:
	#	dbpmono=int(e[13])
	#elif len(e)==15:
	#	dbpmono=(10*(int(e[13])))+(int(e[14]))
	#else:
	#	print('Please check source code (determine double bond position of already located double bond)')
	dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond        ######################################### CHECK ?
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
	csubmono=dbpmono+2
	if nsecdbp>0:
		kadd=nspec*nsecdbp#+1		#(nspec= number of products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		f=float(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			precchrg.append(f)	# copied, no change
			k=k+1	

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
		degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==4:			####################################### VARY ################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(5)		####################################### VARY ################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# write new precursorname, changed to FA with four db
			k=k+1

		e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precadduct.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=float(writelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added double bond
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precmz.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change FA in PrecursorName to generate ProductName
		if degunsat==4:			########################################## VARY ######################################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(5)	########################################## VARY ######################################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from degree of unsaturation of FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond positions for later 	###################### CHECK !						
				firstdbp=dbindexlist[r][0]
				seconddbp=dbindexlist[r][1]
				thirddbp=dbindexlist[r][2]
				fourthdbp=dbindexlist[r][3]
				dbl.append(firstdbp)
				dbl.append(seconddbp)
				dbl.append(thirddbp)
				dbl.append(fourthdbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond positions for later
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde last existing double bond
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee last existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][2])
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][2])
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor1'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor2'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor3'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor4'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor5'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor6'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor7'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor8'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor9'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor10'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #sheetinput.cell(row=r, column=3)
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=='D':
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
				ia=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=='D':
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
				cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
				hnp=hnprec-(2*csub-2*4)		# applies to cleavage of first (highest n) double bond
				hnptwo=hnprec-(2*dbpmono-2*3)	# applies to cleavage of second double bond
				cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
				hnpthree=hnprec-(2*(dbindexlist[r][2])-2*2)	# applies to cleavage of third double bond
				cnpthree=cn-(dbindexlist[r][2])			# applies to cleavage of third double bond
				hnpfour=hnprec-(2*(dbindexlist[r][1])-2*1)	# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				cnpfour=cn-(dbindexlist[r][1])			# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				hnpfive=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				cnpfive=cn-(dbindexlist[r][0])			# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				productcrigeefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydetwo)
					prodformula.append(productcrigeetwo)
					prodformula.append(productaldehydethree)
					prodformula.append(productcrigeethree)
					prodformula.append(productaldehydefour)
					prodformula.append(productcrigeefour)	
					prodformula.append(productaldehydefive) ############################################# CHECK !!!!!!!!
					prodformula.append(productcrigeefive) ############################################# CHECK !!!!!!!!
				if preconly==1:
					t=1
					while t<nspec:
						prodformula.append(precursor)
						t=t+1
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fifth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fifth db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
if workflowmode==1:
	writelist[0]=writelist[0]+mlistname
	writelist[1]=writelist[1]+precname
	writelist[2]=writelist[2]+precformula
	writelist[3]=writelist[3]+precadduct
	writelist[4]=writelist[4]+precmz
	writelist[5]=writelist[5]+precchrg
	writelist[6]=writelist[6]+prodname
	writelist[7]=writelist[7]+prodformula
	writelist[8]=writelist[8]+prodadduct
	writelist[9]=writelist[9]+prodmz
	writelist[10]=writelist[10]+prodchrg
#end save data in writelist

#print('All calculations for fatty acids with five double bonds are done.')
#end save excel file
# end add triple unsaturated fatty acids
######################################################################################################################################################
##############################################################FIVE DB#################################################################################

##############################################################SIX DB#################################################################################
######################################################################################################################################################
# begin add triple unsaturated fatty acids
#print(len(dbindexlist))
#ask=eval(input('Add fatty acids with three double bonds? Yes: 1; No: 0 | '))
ask=0
if ask==1:
	quit()

nspec=13	# number of species: precursor; aldehyde db1, db2 and db3; crigee db1, db2 and db3  ############################ VARY ########################
# begin create empty lists
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]
#end create empty lists
#begin read excel file and save data in lists, edit strings and calculate fragment masses, build output lists
#wb=openpyxl.load_workbook('jpmfaamppozidpolyoutputwrite.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
#sheetinput=wb['transitionlist']
ki=len(writelist[0])

#print('Entries in excel file after bisunsaturated FA transitions are generated: %d' % ki)
kmono=ki # List index in excel file to start writing FA transitions with four double bonds
ksixstart=kmono
r=kfivestart ################################################################################################### VARY ########################
if workflowmode==2:
	r=ki#
while r<ki:		#go through rows of FAs with three double bonds in excel file
	# begin determine number of C atoms in chain, define highest possible double bond position ###############
	cchain=0
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)	# Precursorname
	add=e[5]
	add=int(add)
	#print(add)
	cchain=cchain+(10*add)
	add=e[6]
	add=int(add)
	cchain=cchain+(1*add)
	maxdbp=cchain-2
	#maxdbp=
	# end determine number of C atoms in chain, define highest possible double bond position
	# begin determine double bond position of already located double bond (dbpmono)
	#if len(e)==14:
	#	dbpmono=int(e[13])
	#elif len(e)==15:
	#	dbpmono=(10*(int(e[13])))+(int(e[14]))
	#else:
	#	print('Please check source code (determine double bond position of already located double bond)')
	dbpmono=dbindexlist[r][(len(dbindexlist[r])-1)]	# Last existing double bond, highest n        ######################################### CHECK ?
	# end determine double bond position of already located double bond
	nsecdbp=cchain-dbpmono-3	# number of possible double bond positions for the double bond that is to be added
	csubmono=dbpmono+2
	if nsecdbp>0:
		kadd=nspec*nsecdbp#+1		#(nspec= number of products - precursor, aldehyde and crigee for AMPP - for each of nsecdbp possible double bond positions)

		e=writelist[0][r] #sheetinput.cell(row=r, column=1) # MoleculeGroup
		f=float(writelist[5][r]) #sheetinput.cell(row=r, column=6)	# PrecursorCharge
		k=0
		while k<kadd:
			mlistname.append(e)	# copied, no change
			precchrg.append(f)	# copied, no change
			k=k+1	

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# PrecursorName
		degunsat=int(e[8])		# change from bisunsaturated to trisunsaturated FA in PrecursorName
		if degunsat==5:			####################################### VARY ################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(6)		####################################### VARY ################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne
		k=0
		while k<kadd:
			precname.append(e)	# write new precursorname, changed to FA with four db
			k=k+1

		e=writelist[3][r] #sheetinput.cell(row=r, column=4)	# Precursoradduct
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precadduct.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=float(writelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz
		e=e-(2*imass[0])#-(2*imass[8])		#subtract H2 to account for added double bond
		k=0
		while k<kadd:
			adding=0
			while adding<nspec:
				precmz.append(e) 	# precursor and all transitions for each db
				adding=adding+1
			k=k+nspec

		e=writelist[1][r] #sheetinput.cell(row=r, column=2)	# read PrecursorName to generate Productname
		degunsat=int(e[8])		# begin change FA in PrecursorName to generate ProductName
		if degunsat==5:			########################################## VARY ######################################################
			t=0
			ne=str()
			while t<8:
				l=e[t]
				ne=ne+str(l)
				t=t+1
			ne=ne+str(6)	########################################## VARY ######################################################
			t=9
			while t<len(e):
				l=e[t]
				ne=ne+str(l)
				t=t+1
			e=ne		# end change from degree of unsaturation of FA in PrecursorName
		if e=='Cholesterol (+[2]H7)':
			fragment='_ozone neutral gain' 		
		else:	
			k=0
			csub=csubmono
			while k<kadd:
				dbp='_n-'+str(csub)
				fragment='_precursor'
				ne=e+dbp+fragment
				prodname.append(ne)		# precursor
				dbl=[]	#begin save double bond positions for later 	###################### CHECK !						
				firstdbp=dbindexlist[r][0]
				seconddbp=dbindexlist[r][1]
				thirddbp=dbindexlist[r][2]
				fourthdbp=dbindexlist[r][3]
				fifthdbp=dbindexlist[r][4]
				dbl.append(firstdbp)
				dbl.append(seconddbp)
				dbl.append(thirddbp)
				dbl.append(fourthdbp)
				dbl.append(fifthdbp)
				dbl.append(csub)
				dbindexlist.append(dbl)	#end save double bond positions for later
				dbpi=0
				while dbpi<nspec:
					dbplist.append(dbp)
					dbpi=dbpi+1
				if preconly==0:
					dbp='_n-'+str(csub)
					fragment='_aldehyde neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde
					fragment='_criegee neutral loss from n-'+str(csub)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee
					fragment='_aldehyde neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde last existing double bond
					fragment='_criegee neutral loss from n-'+str(dbpmono)
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee last existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][3])
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][3])
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][2])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][2])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][1])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					fragment='_aldehyde neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# aldehyde previous double bond
					fragment='_criegee neutral loss from n-'+str(dbindexlist[r][0])		######################################### CHECK !!!!!!!!!!!!!!!!!
					ne=e+dbp+fragment
					prodname.append(ne)		# crigee previous existing double bond
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					fifthdbp=dbindexlist[r][4]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(fifthdbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				if preconly==1:
					dbp='_n-'+str(csub)
					fragment='_dummy precursor1'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor2'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor3'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor4'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor5'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor6'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor7'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor8'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor9'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor10'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor11'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					fragment='_dummy precursor12'
					ne=e+dbp+fragment
					prodname.append(ne)		# dummy precursor
					dbl=[]	#begin save double bond position for later 						
					firstdbp=dbindexlist[r][0]
					seconddbp=dbindexlist[r][1]
					thirddbp=dbindexlist[r][2]
					fourthdbp=dbindexlist[r][3]
					fifthdbp=dbindexlist[r][4]
					dbl.append(firstdbp)
					dbl.append(seconddbp)
					dbl.append(thirddbp)
					dbl.append(fourthdbp)
					dbl.append(fifthdbp)
					dbl.append(csub)
					ti=0
					while ti<(nspec-1):
						dbindexlist.append(dbl)	#end save double bond position for later
						ti=ti+1
				csub=csub+1
				k=k+nspec

		e=writelist[2][r] #sheetinput.cell(row=r, column=3)
		# begin read precursor sum formula and edit product sum formula
		#print(e)
		#print(e[0])
		clist=[]
		hlist=[]
		dlist=[]
		nlist=[]
		olist=[]
		plist=[]
		ilist=[]
		i=0
		ca=0
		ha=0
		da=0
		na=0
		oa=0
		pa=0
		ia=0
		while i<len(e):
			if e[i]=='H':
				if e[i+1]=="'":
					ha=0
				else:
					ca=0
			#if e[i]=='D':
			#	ha=0		
			if e[i]=='N':
				ha=0
				da=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
			if e[i]=='P':
				ha=0
				da=0
				na=0
				oa=0
			if e[i]=='I':
				ha=0
				da=0
				na=0
				oa=0
				pa=0
			if ca==1:
				clist.append(e[i])
			if ha==1:
				hlist.append(e[i])
			if da==1:
				dlist.append(e[i])
			if na==1:
				nlist.append(e[i])
			if oa==1:
				olist.append(e[i])
			if pa==1:
				plist.append(e[i])
			if ia==1:
				ilist.append(e[i])
			if e[i]=='C':
				ca=1
			if e[i]=='H':
				if e[i+1]=="'":
					ca=0
					ha=0
					da=1
					i=i+1
				else:
					ca=0
					ha=1
			#if e[i]=='D':
			#	ca=0
			#	ha=0
			#	da=1		
			if e[i]=='N':
				ha=0
				da=0
				na=1
				if e[i+1]=='O':
					nlist.append('1')
					na=0
			if e[i]=='O':
				ha=0
				da=0
				na=0
				oa=1
				if (i+1)<len(e):
					if e[i+1]=='P':
						olist.append('1')
						oa=0
				else:
					olist.append('1')
					oa=0					
			if e[i]=='P':
				da=0
				na=0
				oa=0
				pa=1
				if (i+1)<len(e):
					if e[i+1]=='I':
						plist.append('1')
						pa=0
				else:
					plist.append('1')
					pa=0
			if e[i]=='I':
				da=0
				na=0
				oa=0
				pa=0
				ia=1
				if i==(len(e)-1):
					ilist.append('1')
					ia=0
			i=i+1
		#print(clist)
		#print(hlist)
		#print(dlist)
		#print(nlist)
		#print(olist)
		#print(plist)

		if len(clist)==0:
			cn=0
		if len(hlist)==0:
			hn=0
		if len(dlist)==0:
			dn=0	
		if len(nlist)==0:
			nn=0
		if len(olist)==0:
			on=0
		if len(plist)==0:
			pn=0
		if len(ilist)==0:
			iodon=0
		if len(clist)==1:
			cn=int(clist[0])
		if len(clist)==2:
			cn=10*int(clist[0])+int(clist[1])
		if len(clist)==3:
			cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
		if len(hlist)==1:
			hn=int(hlist[0])
		if len(hlist)==2:
			hn=10*int(hlist[0])+int(hlist[1])
		if len(hlist)==3:
			hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
		if len(dlist)==1:
			dn=int(dlist[0])
		if len(dlist)==2:
			dn=10*int(dlist[0])+int(dlist[1])
		if len(dlist)==3:
			dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
		if len(nlist)==1:
			nn=int(nlist[0])
		if len(nlist)==2:
			nn=10*int(nlist[0])+int(nlist[1])
		if len(nlist)==3:
			nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
		if len(olist)==1:
			on=int(olist[0])
		if len(olist)==2:
			on=10*int(olist[0])+int(olist[1])
		if len(olist)==3:
			on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
		if len(plist)==1:
			pn=int(plist[0])
		if len(plist)==2:
			pn=10*int(plist[0])+int(plist[1])
		if len(plist)==3:
			pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
		if len(ilist)==1:
			iodon=int(ilist[0])
		if len(ilist)==2:
			iodon=10*int(ilist[0])+int(ilist[1])
		if len(ilist)==3:
			iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
		#print(cn)
		e=writelist[0][r] #sheetinput.cell(row=r, column=1)		# begin calculate product sum formula
		if e=='SPLASH':
			#e=sheetinput.cell(row=r, column=2)		# begin calculate product sum formula for SPLASH
			e=e
		else:
			k=0
			csub=csubmono
			while k<kadd:
				hnprec=hn-2		 			# subtract 2 H to account for inner double bond, applies to precursor with three double bonds
				cnp=cn-(csub)				# applies to cleavage of first (highest n) double bond
				hnp=hnprec-(2*csub-2*5)		# applies to cleavage of first (highest n) double bond
				hnptwo=hnprec-(2*dbpmono-2*4)	# applies to cleavage of second double bond
				cnptwo=cn-(dbpmono)				# applies to cleavage of second double bond
				hnpthree=hnprec-(2*(dbindexlist[r][3])-2*3)	# applies to cleavage of third double bond
				cnpthree=cn-(dbindexlist[r][3])			# applies to cleavage of third double bond
				hnpfour=hnprec-(2*(dbindexlist[r][2])-2*2)	# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				cnpfour=cn-(dbindexlist[r][2])			# applies to cleavage of fourth double bond ############################################# CHECK !!!!!!!!
				hnpfive=hnprec-(2*(dbindexlist[r][1])-2*1)	# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				cnpfive=cn-(dbindexlist[r][1])			# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				hnpsix=hnprec-(2*(dbindexlist[r][0]))	# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				cnpsix=cn-(dbindexlist[r][0])			# applies to cleavage of fifth double bond ############################################# CHECK !!!!!!!!
				onald=on+1
				oncrigee=on+2
				precursor='C'+str(cn)+'H'+str(hnprec)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				productaldehyde='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigee='C'+str(cnp)+'H'+str(hnp)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeetwo='C'+str(cnptwo)+'H'+str(hnptwo)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeethree='C'+str(cnpthree)+'H'+str(hnpthree)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon)
				productcrigeefour='C'+str(cnpfour)+'H'+str(hnpfour)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon)
				productaldehydefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				productcrigeefive='C'+str(cnpfive)+'H'+str(hnpfive)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				productaldehydesix='C'+str(cnpsix)+'H'+str(hnpsix)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onald)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				productcrigeesix='C'+str(cnpsix)+'H'+str(hnpsix)+"H'"+str(dn)+'N'+str(nn)+'O'+str(oncrigee)+'P'+str(pn)+'I'+str(iodon) ############################################# CHECK !!!!!!!!
				t=0
				while t<nspec:
					precformula.append(precursor)
					t=t+1
				prodformula.append(precursor)
				if preconly==0:
					prodformula.append(productaldehyde)
					prodformula.append(productcrigee)
					prodformula.append(productaldehydetwo)
					prodformula.append(productcrigeetwo)
					prodformula.append(productaldehydethree)
					prodformula.append(productcrigeethree)
					prodformula.append(productaldehydefour)
					prodformula.append(productcrigeefour)	
					prodformula.append(productaldehydefive) ############################################# CHECK !!!!!!!!
					prodformula.append(productcrigeefive) ############################################# CHECK !!!!!!!!
					prodformula.append(productaldehydesix) ############################################# CHECK !!!!!!!!
					prodformula.append(productcrigeesix) ############################################# CHECK !!!!!!!!
				if preconly==1:
					t=1
					while t<nspec:
						prodformula.append(precursor)
						t=t+1
				# # product formula is saved in list for current double bond position
				precursormz=imass[0]*(hnprec)+imass[1]*dn+imass[2]*cn+imass[3]*nn+imass[4]*on+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for precursor, AMPP
				prodmz.append(precursormz)
				if preconly==0:
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnp+0)+imass[1]*dn+imass[2]*cnp+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, first db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnptwo)+imass[1]*dn+imass[2]*cnptwo+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, second db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpthree)+imass[1]*dn+imass[2]*cnpthree+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, third db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfour)+imass[1]*dn+imass[2]*cnpfour+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fourth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, fifth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpfive)+imass[1]*dn+imass[2]*cnpfive+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, fifth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpsix)+imass[1]*dn+imass[2]*cnpsix+imass[3]*nn+imass[4]*onald+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for aldehyde neutral loss, sixth db, AMPP
					prodmz.append(productmz)
					productmz=imass[0]*(hnpsix)+imass[1]*dn+imass[2]*cnpsix+imass[3]*nn+imass[4]*oncrigee+imass[5]*pn+imass[10]*iodon-imass[8] 		# ProductMz for crigee neutral loss, sixth db, AMPP
					prodmz.append(productmz)
				if preconly==1:
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
					prodmz.append(precursormz)
				csub=csub+1
				k=k+nspec			
		# end read precursor sum formula and edit product sum formula

		prodadduct=precadduct #define column ProductAdduct (same as precursor adduct)
		prodchrg=precchrg 	# ProductCharge  #############

	r=r+1
	e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)		# begin determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
	prevprec=e
	e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
	prec=e
	while prec==prevprec:
		if r<(ki-1):
			r=r+1
			e=writelist[1][r-1] #sheetinput.cell(row=(r-1), column=2)
			prevprec=e
			e=writelist[1][r] #sheetinput.cell(row=(r), column=2)
			prec=e		# end determine whether PrecursorName is the same as previous Precursorname, if yes: r=r+1 (go through rows without action)
		else:
			prec='stop_loop'
			r=r+1
	
#print(len(mlistname))
#end read excel file and save data in lists, edit strings and calculate fragment masses, build output lists

#begin go through lists and add double bond position to PrecursorName
k=0
kt=len(precname)
#print(precname)
#print(dbplist)
#if len(precname)==len(dbplist):
	#print('test correct')
while k<kt:
	precname[k]=precname[k]+dbplist[k]
	k=k+1
#end go through lists and add double bond position to PrecursorName

#begin save data in writelist
if workflowmode==1:
	writelist[0]=writelist[0]+mlistname
	writelist[1]=writelist[1]+precname
	writelist[2]=writelist[2]+precformula
	writelist[3]=writelist[3]+precadduct
	writelist[4]=writelist[4]+precmz
	writelist[5]=writelist[5]+precchrg
	writelist[6]=writelist[6]+prodname
	writelist[7]=writelist[7]+prodformula
	writelist[8]=writelist[8]+prodadduct
	writelist[9]=writelist[9]+prodmz
	writelist[10]=writelist[10]+prodchrg
#end save data in writelist

#print('All calculations for fatty acids with six double bonds are done.')
#end save excel file
# end add triple unsaturated fatty acids
#print(len(writelist[0]))
#print(len(dbindexlist))
######################################################################################################################################################
##############################################################END ADD SIX DB##########################################################################
######################################################################################################################################################
# begin add saturated FAs
r=0
while r<(len(satlist[1])):
	#satlist[1][r][len(satlist[1][r])-1]='0'	
	precn=''
	p=0
	while p<(len(satlist[1][r])-1):
		precn=precn+satlist[1][r][p]
		p=p+1
	precn=precn+'0'					#PrecursorName to saturated FA
	satlist[1][r]=precn
	satlist[6][r]=satlist[1][r]
	satlist[4][r]=float(satlist[4][r])+(2*imass[0])	#PrecursorMz to sat. FA
	satlist[9][r]=float(satlist[9][r])+(2*imass[0])	#ProductMz to sat. FA
	hcurr=(10*(int(satlist[2][r][4])))+(int(satlist[2][r][5]))
	hcurr=str(hcurr+2)
	p=0
	precf=''
	while p<4:
		precf=precf+satlist[2][r][p]
		p=p+1
	precf=precf+hcurr
	p=p+2
	while p<(len(satlist[2][r])):
		precf=precf+satlist[2][r][p]
		p=p+1
	satlist[2][r]=str(precf)	#PrecursorFormula is edited
	satlist[7][r]=str(precf)	#ProductFormula is edited
	r=r+1
clmn=0
while clmn<(len(satlist)):
	writelist[clmn]=writelist[clmn]+satlist[clmn]
	clmn=clmn+1
# end add saturated FAs

##########################################################VIRTUAL#PRECURSOR###########################################################################
######################################################################################################################################################
# begin introduce virtual precursor
mlistname=[]
precname=[]
precformula=[]
precadduct=[]
precmz=[]
precchrg=[]
prodname=[]
prodformula=[]
prodadduct=[]
prodmz=[]
prodchrg=[]
dbplist=[]		# content in lists is deleted to free up memory
# begin introduce virtual precursor
ki=len(writelist[0])

#print('Entries in excel file after all FA transitions are generated: %d' % ki)
virtualprecformula=[]
virtualprecmz=[]

vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]

r=0
ki=ki
while r<ki:
	e=str(writelist[2][r])#sheetinput.cell(row=r, column=3)	# PrecursorFormula
	ve=e+'Xe'
	virtualprecformula.append(ve) # virtual precursorFormula
	e=float(writelist[4][r])#sheetinput.cell(row=r, column=5)	# PrecursorMz
	ve=e+(imass[9])		#add one Xe atom to generate virtual precursor
	virtualprecmz.append(ve)	# virtual precursormz
	e=str(writelist[0][r])#sheetinput.cell(row=r, column=1)
	vmlistname.append(e)
	e=str(writelist[1][r]) #sheetinput.cell(row=r, column=2)
	vprecname.append(e)
	e=str(writelist[3][r]) #sheetinput.cell(row=r, column=4)
	vprecadduct.append(e)
	e=int(writelist[5][r]) #sheetinput.cell(row=r, column=6)
	vprecchrg.append(e)
	e=str(writelist[6][r]) #sheetinput.cell(row=r, column=7)
	vprodname.append(e)
	e=str(writelist[7][r]) #sheetinput.cell(row=r, column=8)
	vprodformula.append(e)
	e=str(writelist[8][r]) #sheetinput.cell(row=r, column=9)
	vprodadduct.append(e)
	e=float(writelist[9][r]) #sheetinput.cell(row=r, column=10)
	vprodmz.append(e)
	e=int(writelist[10][r]) #sheetinput.cell(row=r, column=11)
	vprodchrg.append(e)
	r=r+1
# begin save excel file with virtual precursor as csv file
toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge']
vwritelist=[]
vwritelist.append(vmlistname)
vwritelist.append(vprecname)
vwritelist.append(virtualprecformula)
vwritelist.append(vprecadduct)
vwritelist.append(virtualprecmz)
vwritelist.append(vprecchrg)
vwritelist.append(vprodname)
vwritelist.append(vprodformula)
vwritelist.append(vprodadduct)
vwritelist.append(vprodmz)
vwritelist.append(vprodchrg)
terminate=0			############################################ check ok
if terminate==1:
	print('vwritelist created')
	transitionresultsdf=pd.DataFrame(vwritelist).transpose()
	print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10]]
	print('Transposed and DataFrame created')
	transitionresultsdf.to_csv('jpmlipidomics_vpw19_virtual_precursor.csv', index=False)
	print('Transition list saved as jpmlipidomics_vpw19_virtual_precursor.csv')  
# end save excel file with virtual precursor as csv file
#quit()
#print('Saving data.')
#print('len(vmlistname)')
#print(len(vmlistname))
virtualprecformula=[]
virtualprecmz=[]
vmlistname=[]
vprecname=[]
vprecadduct=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[] # test to delete lists to save up memory

print('Transition list is modified with virtual precursor [M + Xe].')
#print(vwritelist[1])
#krows=len(vwritelist[1])
#print(krows)
#quit()
#################################################################################################################################################
############################# BEGIN REDUCE TO ENTRIES OF FOUND PRECURSOR AND EXPAND WITH PRECURSOR EXPLICIT RETENTION TIMES #######################
#################################################################################################################################################
#terminate=eval(input('Reduce transition list to entries for which precursor is found and expand with explicit retention time of the found precursor? Yes: 1 No: 0 ::'))
terminate=0 ############################################ check 
if terminate==1:
	quit()
ki=len(vwritelist[0])
#else:
	#exrtstep=eval(input('Stepwidth for varied explicit retention time (e.g. 0.05 min)? ::'))
	#exrtstep=0.05
# begin reduce and expand
#wb=openpyxl.load_workbook('jpm_lipidomics_vpw03_output.xlsx')			# load excel file from home folder 
#sheetinput=wb['transitionlist']
#r=2		# begin count entries in excel file
#ki=0
#test=1
#while test==1:
#	e=sheetinput.cell(row=r, column=5)
#	e=e.value
#	if e is None: #
#		test=0
#		ki=ki-1
#	r=r+1
#	ki=ki+1
# end count entries in excel file

# begin read transitionresults excel file
#wbr=openpyxl.load_workbook('jpm_lipidomics_vpw07_selected_output_redgreen.xlsx')			# load excel file from home folder, results from previous jpmlipidomics_vpw05_1_2_testredgreen.py
#sheetresults=wbr['transitionresults']
#rr=2		# begin count entries in excel file
#kir=0
#test=1
#while test==1:
#	e=sheetresults.cell(row=rr, column=5)
#	e=e.value
#	if e is None: #
#		test=0
#		kir=kir-1
#	rr=rr+1
#	kir=kir+1
# end count entries in excel file
kir=len(pfwritelist[0])
# end read transitionresults excel file
vmlistname=[]
vprecname=[]
virtualprecformula=[]
vprecadduct=[]
virtualprecmz=[]
vprecchrg=[]
vprodname=[]
vprodformula=[]
vprodadduct=[]
vprodmz=[]
vprodchrg=[]
precrt=[]
rtwindow=[]
r=0
pos=1
while r<ki:
	rn=r
	precrlist=[]
	rexrtlist=[]
	pos=0
	while pos==0:
		# begin determine block length (start with r of block of species with same precursor mass, end with s)
		e=vwritelist[4][r] #(row=r, column=5)	# PrecursorMz		# begin determine which row to start (r) and to end (s)
		s=r+1
		st=0
		while st<1:
			if s<ki:
				ne=vwritelist[4][s] #(row=s, column=5)	# PrecursorMz
			else:
				ne='stop loop'
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		# begin determine if precursor of current species is found
		e=str(vwritelist[1][r]) #(row=r, column=2)
		precspecies=e[5]+e[6]+e[7]+e[8]
		precspecies=str(precspecies)	# precursor species in full list
		pos=0
		rr=0
		while rr<(kir+0):
			e=pfwritelist[1][rr] #sheetresults.cell(row=rr, column=2)
			e=str(e)
			precrspecies=e[5]+e[6]+e[7]+e[8]
			precrspecies=str(precrspecies)	# precursorspecies in list with confirmed species from precursor only analysis
			if precspecies==precrspecies:
				pos=1					#### the precursor of the current species in full list was found
				precrlist.append(rr)	#### list with precursorspecies that were identified
				e=pfwritelist[17][rr]  #sheetresults.cell(row=rr, column=18)
				rexrt=float(e)
				rexrtlist.append(rexrt)	#### list with explicit retention times associated to precursorspecies that were identified
			rr=rr+1
		# end determine if precursor of current species is found
		if pos==0:
			if (s+1)<ki:
				r=s+1
				rn=r
			else:
				#print('(s+1)>(ki-1)')
				r=s
				pos=1
	# block identified for which precursor is found
	if (r+1)>ki:
		pos=0
		r=r+2
	if pos==1:
		#print('Block found')
		#print(precspecies)
		#print(precrspecies)
		#print(r)
		#print(s)
		#print(len(vmlistname))
		#begin define expansion
		# begin determine block length (start with rr of block of species with same precursor mass, end with sr)

		#rr=precrlist[0]
		#sr=precrlist[(len(precrlist)-1)]

		#print(rr)
		#print(sr)
		#e=sheetresults.cell(row=rr, column=5)	# PrecursorMz		# begin determine which row to start (rr) and to end (sr)
		#e=e.value
		#sr=rr+1
		#st=0
		#while st<1:
		#	ne=sheetinput.cell(row=sr, column=5)	# PrecursorMz
		#	ne=ne.value
		#	if ne==e:
		#		sr=sr+1
		#		st=0
		#	else:
		#		sr=sr-1
		#		st=1		# end determine sr
		#end define expansion
		kr=0
		while kr<(len(rexrtlist)):
			rexrt=rexrtlist[kr]
			rexrt=str(rexrt)
			r=rn
			while r<(s+1):
				#print(len(virtualprecformula))
				ve=str(vwritelist[2][r]) #sheetinput.cell(row=r, column=3)	# PrecursorFormula
				virtualprecformula.append(ve) # virtual precursorFormula
				ve=float(vwritelist[4][r]) #sheetinput.cell(row=r, column=5)	# PrecursorMz		
				virtualprecmz.append(ve)	# virtual precursormz
				e=str(vwritelist[0][r]) #sheetinput.cell(row=r, column=1)
				vmlistname.append(e)
				e=str(vwritelist[1][r]) #sheetinput.cell(row=r, column=2)
				cm='_'
				cm=str(cm)
				rexrt=float(rexrt)
				rexrtstr=str(round(rexrt, 2))		#
				ee=e+cm+rexrtstr			#
				vprecname.append(ee)
				precrt.append(rexrt)	##################################
				e=str(vwritelist[3][r]) #sheetinput.cell(row=r, column=4)
				vprecadduct.append(e)
				e=int(vwritelist[5][r]) #sheetinput.cell(row=r, column=6)
				vprecchrg.append(e)
				e=str(vwritelist[6][r]) #sheetinput.cell(row=r, column=7)
				vprodname.append(e)
				e=str(vwritelist[7][r]) #sheetinput.cell(row=r, column=8)
				vprodformula.append(e)
				e=str(vwritelist[8][r]) #sheetinput.cell(row=r, column=9)
				vprodadduct.append(e)
				e=float(vwritelist[9][r]) #sheetinput.cell(row=r, column=10)
				vprodmz.append(e)
				e=int(vwritelist[10][r]) #sheetinput.cell(row=r, column=11)
				vprodchrg.append(e)	
				#e=0.055 
				#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
				if int(vwritelist[1][r][8])>2:
					nca=10*int(vwritelist[1][r][5])+int(vwritelist[1][r][6])
					if nca>17:
						exrtstep=0.045
				else:	
					exrtstep=0.027
				#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
				e=exrtstep*2
				if float(e)<0.1:
					e=0.1
					rtwindow.append(e)	
				else:
					rtwindow.append(e)	
				r=r+1
			kr=kr+1
	if (r+1)>ki:
		r=s+3
	else:
		r=s+1

#print('Main loop done')	
#krows=len(vmlistname)
#print(krows)
#print(ki)
#print(kir)
#print(pfwritelist[1])
#quit()
# begin save to csv file
#before = datetime.datetime.now()
prt='PrecursorRT'
prt=str(prt)
toprow.append(prt)
rtw='PrecursorRTWindow'
rtw=str(rtw)
toprow.append(rtw)
#ki=2+len(vmlistname)

writelist=[]
writelist.append(vmlistname)
writelist.append(vprecname)
writelist.append(virtualprecformula)
writelist.append(vprecadduct)
writelist.append(virtualprecmz)
writelist.append(vprecchrg)
writelist.append(vprodname)
writelist.append(vprodformula)
writelist.append(vprodadduct)
writelist.append(vprodmz)
writelist.append(vprodchrg)
writelist.append(precrt)
writelist.append(rtwindow)
#print('writelist created')
mespacedrule=1											########## ACTIVATE (1) OR DEACTIVATE (0) METHYLENE (BUTYLENE) SPACED RULE ##############
if mespacedrule==0:
	transitionresultsdf=pd.DataFrame(writelist).transpose()
	#print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
	#print('Transposed and DataFrame created')
	transitionresultsdf.to_csv('jpmlipidomics_vpw20_2_full_list.csv', index=False)
	nrows=len(mlistname)
	print('Transition list is saved as jpmlipidomics_vpw20_2_full_list.csv (%d rows)' % nrows)
	afterall=datetime.datetime.now()
	dt=afterall-beforeall
	print('Calculation time:')
	print(dt)
	# end save to csv file

#################################################################################################################################################
############################### END REDUCE TO ENTRIES OF FOUND PRECURSOR AND EXPAND WITH PRECURSOR EXPLICIT RETENTION TIMES #####################
#################################################################################################################################################

#################################################################################################################################################
## begin apply methylene (butylene) spacing rule to relevant species 
## (delete unrealistic species that can't be distinguished from realistic species as associated double bonds are non-diagnostic)
#################################################################################################################################################
if mespacedrule==1:
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	precrt=[]
	rtwindow=[]
	ki=len(writelist[0])
	#print('Rows in jpmlipidomics_vpw10_2_full_tr_results.csv:')
	#print(ki)
	r=0
	while r<ki:		# go through rows of excel file 
		e=writelist[1][r] # Precursorname		# begin determine which row to start (r) and to end (s)
		s=r+1
		st=0
		while st<1:
			if s>(len(writelist[1])-1):
				ne='stop_loop'
			else:
				ne=writelist[1][s] #sheetinput.cell(row=s, column=2)	# Precursorname
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		t=r
		pos=0
		if int(writelist[1][t][8])>2:		# Only consider PUFA with more than 2 db
			go=1
			ch=len(writelist[1][t])-1
			while go==1:					# determine last and second last double bond position in PUFA
				if str(writelist[1][t][ch])=='n':	
					go=0
					if str(writelist[1][t][ch+4])=='_':
						lastdb=10*int(writelist[1][t][ch+2])+int(writelist[1][t][ch+3])
					else:
						lastdb=int(writelist[1][t][ch+2])
					if str(writelist[1][t][ch-4])=='-':	
						seclastdb=10*int(writelist[1][t][ch-3])+int(writelist[1][t][ch-2])
					else:
						seclastdb=int(writelist[1][t][ch-2])
				else:
					go=1
				ch=ch-1
			if lastdb>13:
				if lastdb-seclastdb==3:	# allow methylene interrupted species for last db
					pos=0
				elif lastdb-seclastdb==6:	# allow butylene interrupted species for last db
					pos=0
				else:
					pos=1
		if pos==1:
			r=s+1
		else:
			t=r
			while t<(s+1):
				e=writelist[0][t] ## mlistname	# begin append rows of suitable species to lists for later saving in excel output
				mlistname.append(e)
				e=writelist[1][t] ## precname	
				precname.append(e)
				e=writelist[2][t] ## 	
				precformula.append(e)
				e=writelist[3][t] ## 	
				precadduct.append(e)
				e=writelist[4][t] ## 	
				precmz.append(e)
				e=writelist[5][t] ## 	
				precchrg.append(e)
				e=writelist[6][t] ## 	
				prodname.append(e)
				e=writelist[7][t] ## 	
				prodformula.append(e)
				e=writelist[8][t] ## 	
				prodadduct.append(e)
				e=writelist[9][t] ## 
				prodmz.append(e)
				e=writelist[10][t] ## 	
				prodchrg.append(e)
				e=writelist[11][t] ## 
				precrt.append(e)
				e=writelist[12][t] ## 	
				rtwindow.append(e)
				t=t+1
		r=s+1

	writelist=[]
	writelist.append(mlistname)
	writelist.append(precname)
	writelist.append(precformula)
	writelist.append(precadduct)
	writelist.append(precmz)
	writelist.append(precchrg)
	writelist.append(prodname)
	writelist.append(prodformula)
	writelist.append(prodadduct)
	writelist.append(prodmz)
	writelist.append(prodchrg)
	writelist.append(precrt)
	writelist.append(rtwindow)
	#print('writelist created')
	transitionresultsdf=pd.DataFrame(writelist).transpose()
	#print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
	#print('Transposed and DataFrame created')
	after=datetime.datetime.now()
	after=str(after)
	today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_3_'+fourlettcode+'_'
	filename='jpmlipidomics_vpw20_2_full_list.csv'
	transitionresultsdf.to_csv(filename, index=False)
	nrows=len(mlistname)
	print('Transition list is saved as jpmlipidomics_vpw20_2_full_list.csv (%d rows)' % nrows)
	afterall=datetime.datetime.now()
	dt=afterall-beforeall
	print('Calculation time (h:mm:ss) is:')
	print(dt)
	# end save to csv file

#################################################################################################################################################
## end apply methylene spacing rule to relevant species 
## (delete unrealistic species that can't be distinguished from realistic species as associated double bonds are non-diagnostic)
#################################################################################################################################################


