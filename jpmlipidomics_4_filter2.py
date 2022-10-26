# -*- coding: UTF-8 -*-

# Philipp Menzel lipidomics oz id transition list generate from excel list, use for skyline
#created: 10 12 2020
#modified: 2021 02 02; 
# Goal: read excel file containing data for monounsaturated lipids without double bond info, add rows for OzID product ions, save in excel file
## Notes: work in progress for fatty acids AMPP derivatives (no other ionization), double bond position added to Precursorname after calculations but before saving in excel file 
## Notes: addition for saturated FAs, bisunsaturated FAs, added line for precursor, option for precursor-only transition list with dummy percursor.
## NOTES: STAGE 4 for LIPIDOMICS WORKFLOW AMPP VPW08. Here: delete species for which -H decoy is more than twice the area than precursor principal ion
## NOTES: then delete species for which exact retention times of relevant (excluding n-6) transitions (based on gaussion fit) differ too much (exactexrtthreshold)
##	DONE ## 
import math
import openpyxl
import pandas as pd
import datetime
import statistics
from statistics import mean
from statistics import median
import csv
beforeall=datetime.datetime.now()

segmentsize=50000	# min number of entries in xic report and transitions report to be processed at once (once functional change to 500, then increase to test if advantageous)

convertfile=0		# set 0 for troubleshooting (or run this python script on its own as jpmtsvtocsv), 1 is default value for running workflow through batch file
if convertfile==1:
	# begin convert tsv file generated from Skyline runner to csv file # BEGIN EXTRACT INTENSITIES
	try:
	    with open(r'skyl_xic_report_vpw20_3.tsv', 'r', newline='\n') as in_f, \
	         open(r'skyl_xic_report_vpw20_3_intensities.csv', 'w', newline='\n') as out_f:
	        reader = csv.reader(in_f, delimiter='\t')
	        writer = csv.writer(out_f, delimiter=',')
	        for li in reader:
	            try:
	                writer.writerow([li[0], li[1], li[2], li[3], li[4], li[5], li[6], li[7], li[9]])
	            except IndexError:  # Prevent errors on blank lines.
	                pass
	except IOError as err:
	    print(err)
	# end convert tsv file generated from Skyline runner to csv file
	# begin delete double quotes from generated csv file
	with open('skyl_xic_report_vpw20_3_intensities.csv', "r+", encoding="utf-8") as csv_file:
	    content = csv_file.read()
	with open('skyl_xic_report_vpw20_3_intensities.csv', "w+", encoding="utf-8") as csv_file:
	    csv_file.write(content.replace('"', ''))
	# end delete double quotes from generated csv file # END EXTRACT INTENSITIES
	# begin convert float values for intensities to integers to reduce file size
	tempdf=pd.read_csv('skyl_xic_report_vpw20_3_intensities.csv', header=None, skiprows=1)
	templist=tempdf.values.tolist()
	tcol=0
	trow=0
	while trow<(len(templist)):		# replaces content of first column (FileName) with int(0)
		templist[trow][tcol]=int(0)
		trow=trow+1
	tcol=2
	trow=0
	while trow<(len(templist)):		# replaces content of third column (Precursorcharge) with int(0)
		templist[trow][tcol]=int(0)
		trow=trow+1
	tcol=6
	trow=0
	while trow<(len(templist)):		# replaces content of seventh column (IsotopeLabel) with int(0)
		templist[trow][tcol]=int(0)
		trow=trow+1
	tcol=9
	trow=0
	while trow<(len(templist)):		# converts intensities to integers
		tcol=9
		while tcol<(len(templist[0])):
			templist[trow][tcol]=float(templist[trow][tcol])
			templist[trow][tcol]=round(templist[trow][tcol], 0)
			templist[trow][tcol]=int(templist[trow][tcol])
			tcol=tcol+1
		trow=trow+1
	tempconvdf=pd.DataFrame(templist)
	filename='skyl_xic_report_vpw20_3_intensities.csv'
	tempconvdf.to_csv(filename, index=False)
	templist=[]
	tempconvdf=pd.DataFrame(templist)
	tempdf=pd.DataFrame(templist)
	# end convert float values for intensities to integers to reduce file size
if convertfile==1:
	# begin convert tsv file generated from Skyline runner to csv file # BEGIN EXTRACT TIMES
	try:
	    with open(r'skyl_xic_report_vpw20_3.tsv', 'r', newline='\n') as in_f, \
	         open(r'skyl_xic_report_vpw20_3_times.csv', 'w', newline='\n') as out_f:
	        reader = csv.reader(in_f, delimiter='\t')
	        writer = csv.writer(out_f, delimiter=',')
	        for li in reader:
	            try:
	                writer.writerow([li[8]])
	            except IndexError:  # Prevent errors on blank lines.
	                pass
	except IOError as err:
	    print(err)
	# end convert tsv file generated from Skyline runner to csv file
	# begin delete double quotes from generated csv file
	with open('skyl_xic_report_vpw20_3_times.csv', "r+", encoding="utf-8") as csv_file:
	    content = csv_file.read()
	with open('skyl_xic_report_vpw20_3_times.csv', "w+", encoding="utf-8") as csv_file:
	    csv_file.write(content.replace('"', ''))
	# end delete double quotes from generated csv file
	xictimesdf=pd.read_csv('skyl_xic_report_vpw20_3_times.csv', header=None, skiprows=1, nrows=1)
	xictimeslistfromdf=xictimesdf.values.tolist()
	xictimeslist=xictimeslistfromdf[0]		# contains times of XICs
	# begin save times of XICs in csv file
	xictimesconvdf=pd.DataFrame(xictimeslist)
	filename='skyl_xic_report_vpw20_3_times.csv'
	xictimesconvdf.to_csv(filename, index=False)
	# end save times of XICs in csv file
	xictimesdf=pd.DataFrame(templist)
	xictimeslistfromdf=[]
	#print(xictimesdf)
	#print('list:')
	#print(xictimeslist)
	#print(len(xictimeslist))
	# END EXTRACT TIMES
#timedelay=eval(input('File saved completely? (1) :'))
if convertfile==0:
	xictimesdf=pd.read_csv('skyl_xic_report_vpw20_3_times.csv', header=None, skiprows=1, nrows=1)
	xictimeslistfromdf=xictimesdf.values.tolist()
	xictimeslist=xictimeslistfromdf[0]		# contains times of XICs # actual list starts at index 8


#xicintensitiesdf=pd.read_csv('skyl_xic_report_vpw16_3_intensities.csv', header=None, skiprows=1, nrows=1)
#xicintensitieslistfromdf=xicintensitiesdf.values.tolist()
#xicintensitieslist=xicintensitieslistfromdf[0]		# contains times of XICs
#print(xicintensitiesdf)
#print('list:')
#print(xicintensitieslistfromdf)
#print(len(xicintensitieslist))

#quit()

# begin build mostwantedlist 
mostwantedlist=[]
wbfa=openpyxl.load_workbook('jpm_fa_lib.xlsx')
wsfa=wbfa.active
mwi=0
go=1
while go==1:
	tfe=wsfa.cell(row=mwi+3, column=1)
	tfe=tfe.value
	if tfe is None:
		go=0
	else:
		if int(tfe)<10:
			tfe='0'+str(tfe)
		mwfa=str(tfe)+':'
		tfe=wsfa.cell(row=mwi+3, column=2)
		tfe=tfe.value
		dbfa=int(tfe)
		if int(tfe)==0:
			mwfa=mwfa+str(tfe)
		else:
			mwfa=mwfa+str(tfe)+'_n-'
			dbfai=1
			while dbfai<(dbfa+1):
				tfe=wsfa.cell(row=mwi+3, column=dbfai+2)
				tfe=tfe.value
				if dbfai<(dbfa):
					mwfa=mwfa+str(tfe)+'_n-'
				else:
					mwfa=mwfa+str(tfe)
				dbfai=dbfai+1
		mostwantedlist.append(mwfa)
	mwi=mwi+1
# end build mostwantedlist

################ DATABASE ## Source: Internetchemie.info
#isotope=["1H", "2H", "12C", "13C", "14N", "15N", "16O", "17O", "18O", "19F", "23Na", "28Si", "29Si", "30Si", "31P", "32S", "33S", "34S", "36S", "39K", "40K", "41K", "35Cl", "37Cl", "79Br", "81Br"]
#mass=[1.00783, 2.01410 , 12.00000, 13.00335, 14.00307, 15.00011, 15.99491, 16.99913, 17.99916, 18.99840, 22.97977, 27.97693, 28.97649, 29.97377, 30.97376, 31.97207, 32.97146, 33.96787, 35.96708, 38.96371, 39.96400, 40.96183, 34.96885, 36,96590, 78.91834, 80.91629]
#abundance=[99.9885, 0.0115, 98.93, 1.07, 99.636, 0.364, 99.7, 0.04, 0.2, 100, 100, 92.233, 4.685, 3.092, 100, 94.93, 0.76, 4.29, 0.02, 93.2581, 0.0117, 6.7302, 75.76, 24.24, 50.69, 49.31]
isotope=['1H   ', '2H  ', '12C   ', '14N   ', '16O    ', '31P   ', '32S    ' '23Na     ', 'e     ', '132Xe', '   127I']
imass=[1.007825, 2.0141, 12.00000, 14.00307, 15.99491, 30.973762, 31.97207, 22.98977, 0.000548585, 131.9041535, 126.904473]
################
########### begin read workflow parameters
transferlist=[]
wb=openpyxl.load_workbook('OzFAD1_workflow_parameters.xlsx')
ws=wb.active
tli=0
go=1
while go==1:
	tfe=ws.cell(row=tli+1, column=2)
	tfe=tfe.value
	if tli==18:
		go=0
	if tfe is None:
		go=0
	else:
		transferlist.append(tfe)
	tli=tli+1
########## 
fourlettcode=transferlist[0]		#input('Enter four letter code of derivatization agent (e.g. AMPP, NMPA, NMPE, MDPE, NEPE, EDPE, NPPE, IAMP) :')
fourlettcode=str(fourlettcode)
cderiv=transferlist[1]
hderiv=transferlist[2]
dderiv=transferlist[3]
nderiv=transferlist[4]
oderiv=transferlist[5]
pderiv=transferlist[6]
ideriv=transferlist[7]
mostwanted=int(transferlist[17])
transtest=int(transferlist[18])
########### end read workflow parameters 
#print('Before proceeding, please make sure that the Skyline report file is named jpmlipidomics_vpw13_3_tr_results.csv')
#print('Before proceeding, please make sure that the Skyline XIC report file is named jpmlipidomics_vpw13_3_xic_results.csv')
#abundance=[99.9885, 98.93, 99.636, 99.7, 94.93] not updated
trdf=pd.read_csv('skyl_report_vpw20_3.csv')
toprowx=[trdf.columns.values.tolist()]
toprow=toprowx[0]
trdf=pd.read_csv('skyl_report_vpw20_3.csv', skiprows=1, header=None, low_memory=False)	#########, dtype=str######################## read header or not ? , skiprows=1
trdf=trdf.transpose()
writelist=trdf.values.tolist()
ki=len(writelist[0])
print('Number of rows in skyl_report_vpw20_3.csv: %d' % ki)
# begin make seglists
segmlistname=[]
segprecname=[]
segprecformula=[]
segprecadduct=[]
segprecmz=[]
segprecchrg=[]
segprodname=[]
segprodformula=[]
segprodadduct=[]
segprodmz=[]
segprodchrg=[]
segexplicitrt=[]
segexrtwindow=[]
segfmlistname=[]
segfprecname=[]
segfprecformula=[]
segfprecadduct=[]
segfprecmz=[]
segfprecchrg=[]
segfprodname=[]
segfprodformula=[]
segfprodadduct=[]
segfprodmz=[]
segfprodchrg=[]
segfexplicitrt=[]
segfexrtwindow=[]
# end make seglists
segr=0
segs=segmentsize
go=0
if len(writelist[1])>segs:
	if str(writelist[1][segs])==str(writelist[1][segs+1]):
		go=1
	while go==1:
		if str(writelist[1][segs])==str(writelist[1][segs+1]):
			go=1
			segs=segs+1
		else:
			go=0
else:
	segs=len(writelist[1])-1
segrun=1
#begin loop for sequential analysis 
while segrun==1:

	# begin create empty lists
	cutlist=[]		# will contain information on whether to keep species based on gaussian fitting or not
	wcutlist=[]     # info in cutlist for species that are later written in file
	fwcutlist=[]
	avgexactexrtlist=[]	# contains average value of exactexrt of transitions (same value for all of one species), to be written in file for use in final code for barchart plotting
	wavgexactexrtlist=[]	# info in avgexactexrtlist for species that are later written in file
	fwavgexactexrtlist=[]
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	exrt=[]
	exrtwindow=[]
	mzerror=[]
	rettime=[]
	area=[]
	areanormalpercent=[]
	background=[]
	fwhm=[]
	explicitrt=[]
	rtstart=[]
	rtend=[]
	diagnostics=[]
	#end create empty lists
	#####################################################################################################
	# begin assign whether transitions diagnostic or non-diagnostic
	r=segr
	ki=segs+1
	while r<ki:
		t=r
		if int(writelist[1][t][8])>2:
			go=1
			ch=len(writelist[1][t])-1
			while go==1:
				if str(writelist[1][t][ch])=='n':
					go=0
					if str(writelist[1][t][ch+4])=='_':
						lastdb=10*int(writelist[1][t][ch+2])+int(writelist[1][t][ch+3])
					else:
						lastdb=int(writelist[1][t][ch+2])
					if str(writelist[1][t][ch-4])=='-':	
						seclastdb=10*int(writelist[1][t][ch-3])+int(writelist[1][t][ch-2])
					else:
						seclastdb=int(writelist[1][t][ch-2])
				else:
					go=1
				ch=ch-1
			if lastdb>13:
				if lastdb-seclastdb==3:
					if str(writelist[6][t][len(writelist[6][t])-1])=='r':
						diagnostics.append('diagnostic')
					elif str(writelist[6][t][len(writelist[6][t])-1])=='y':
						diagnostics.append('diagnostic')
					else:
						if str(writelist[6][t][len(writelist[6][t])-4])=='n':
							if lastdb==10*int(writelist[6][t][len(writelist[6][t])-2])+int(writelist[6][t][len(writelist[6][t])-1]):
								diagnostics.append('non-diagnostic')
							else:
								diagnostics.append('diagnostic')
						else:
							if lastdb==int(writelist[6][t][len(writelist[6][t])-1]):
								diagnostics.append('non-diagnostic')
							else:
								diagnostics.append('diagnostic')
				else:
					diagnostics.append('diagnostic')
			else:
				diagnostics.append('diagnostic')
		else:
			diagnostics.append('diagnostic')
		r=r+1
	# end assign whether transitions diagnostic or non-diagnostic
	####################################################################################################
	#print(diagnostics)

	r=segr
	ki=segs+1
	while r<ki:		# go through rows of list 
		decoyarea=0
		e=writelist[1][r] ## Precursorname		# begin determine which row to start (r) and to end (s)
		s=r+1
		st=0
		while st<1:
			if s>(len(writelist[1])-1):
				ne='stop_loop'
			else:
				ne=writelist[1][s] #	# Precursorname
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		t=r
		pos=0
		fpos=0
		gpos=0
		while t<(s+1):
			#print(writelist[6][t])
			if str(writelist[6][t])=='nan':
				#print('ok')
				t=s+1
			else:
				if str(writelist[6][t][len(str(writelist[6][t]))-1])=='y':
					decoyarea=float(writelist[13][t])
				if str(writelist[6][t][len(str(writelist[6][t]))-1])=='r':
					precarea=float(writelist[13][t])
				elif str(writelist[6][t][len(str(writelist[6][t]))-1])=='0':
					precarea=float(writelist[13][t])
			t=t+1
		if decoyarea>(2*precarea):		# exclude species for which decoy area is more than double precursor target area 
			if mostwanted==0:
				pos=1
			else:
				# begin test whether species to be cut is in the mostwantedlist (if it is, keep)
				fashort=str()
				fas=5
				while fas<(len(writelist[1][t])-5):
					fashort=fashort+str(writelist[1][t][fas])
					fas=fas+1
				mwl=0
				gposcancel=0
				while mwl<(len(mostwantedlist)):
					if mostwantedlist[mwl]==fashort:
						gposcancel=1
					mwl=mwl+1
				if gposcancel==0:
					pos=1
				# end test whether species to be cut is in the mostwantedlist
			#pos=1
		if pos==1:
			r=s+1
		else:
			t=r
			while t<(s+1):
				apos=0
				#print(writelist[6][t])
				if str(writelist[6][t])=='nan':
					#print('ok')
					apos=1
					t=s+1
				else:
					if str(writelist[6][t][len(writelist[6][t])-1])=='y':
						apos=1
				if apos==0:
					e=writelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving in excel output
					mlistname.append(e)
					e=writelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
					precname.append(e)
					e=writelist[2][t] #sheetinput.cell(row=t, column=3)	# prodname	
					precformula.append(e)
					e=writelist[3][t] #sheetinput.cell(row=t, column=4)	# prodformula	
					precadduct.append(e)
					e=writelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
					precmz.append(e)
					e=writelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
					precchrg.append(e)
					e=writelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
					prodname.append(e)
					e=writelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
					prodformula.append(e)
					e=writelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
					prodadduct.append(e)
					e=writelist[9][t] #sheetinput.cell(row=t, column=10)	# 
					prodmz.append(e)
					e=writelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
					prodchrg.append(e)
					e=writelist[11][t] #sheetinput.cell(row=t, column=12)	# 	
					mzerror.append(e)
					e=writelist[12][t] #sheetinput.cell(row=t, column=13)	# 	
					rettime.append(e)
					e=writelist[13][t] #sheetinput.cell(row=t, column=14)	# 	
					area.append(e)
					e=writelist[14][t] #sheetinput.cell(row=t, column=15)	# 	
					areanormalpercent.append(e)
					e=writelist[15][t] #sheetinput.cell(row=t, column=16)	# 	
					background.append(e)
					e=writelist[16][t] #sheetinput.cell(row=t, column=17)	# 	
					fwhm.append(e)
					e=writelist[17][t] #sheetinput.cell(row=t, column=18)	# 	
					explicitrt.append(e)
					e=writelist[18][t] #sheetinput.cell(row=t, column=19)	# 	
					rtstart.append(e)
					e=writelist[19][t] #sheetinput.cell(row=t, column=20)	# 	
					rtend.append(e)
					#begin define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
					if int(writelist[1][r][8])>2:
						nca=10*int(writelist[1][r][5])+int(writelist[1][r][6])
						if nca>17:
							exrtstep=0.045
					else:	
						exrtstep=0.027
					#end define small and large exrtstep dependent on degree of unsaturation and number of C atoms in chain
					e=exrtstep*2
					exrtwindow.append(e)
				t=t+1
		r=s+1
	# end read transition results and append suitable species to lists 
	exrt=explicitrt

	#################################################################################################################################################
	#################################################################################################################################################
	# begin filter out species, if exact retention time of transitions differ too much, based on 5 point average smoothing of XICs
	exactexrtthreshold=0.018 ######################################################################################################### THRESHOLD CUT ###
	dwritelist=[]
	dwritelist.append(mlistname)
	dwritelist.append(precname)
	dwritelist.append(precformula)
	dwritelist.append(precadduct)
	dwritelist.append(precmz)
	dwritelist.append(precchrg)
	dwritelist.append(prodname)
	dwritelist.append(prodformula)
	dwritelist.append(prodadduct)
	dwritelist.append(prodmz)
	dwritelist.append(prodchrg)
	dwritelist.append(mzerror)
	dwritelist.append(rettime)
	dwritelist.append(area)
	dwritelist.append(areanormalpercent)
	dwritelist.append(background)
	dwritelist.append(fwhm)
	dwritelist.append(explicitrt)
	dwritelist.append(rtstart)
	dwritelist.append(rtend)
	dwritelist.append(exrtwindow)


	# begin create empty lists
	mlistname=[]
	precname=[]
	precformula=[]
	precadduct=[]
	precmz=[]
	precchrg=[]
	prodname=[]
	prodformula=[]
	prodadduct=[]
	prodmz=[]
	prodchrg=[]
	exrt=[]
	exrtwindow=[]
	mzerror=[]
	rettime=[]
	area=[]
	areanormalpercent=[]
	background=[]
	fwhm=[]
	explicitrt=[]
	rtstart=[]
	rtend=[]
	#end create empty lists
	# begin create empty lists
	fmlistname=[]
	fprecname=[]
	fprecformula=[]
	fprecadduct=[]
	fprecmz=[]
	fprecchrg=[]
	fprodname=[]
	fprodformula=[]
	fprodadduct=[]
	fprodmz=[]
	fprodchrg=[]
	fexrt=[]
	fexrtwindow=[]
	fmzerror=[]
	frettime=[]
	farea=[]
	fareanormalpercent=[]
	fbackground=[]
	ffwhm=[]
	fexplicitrt=[]
	frtstart=[]
	frtend=[]
	#end create empty lists
	#trdf=pd.read_csv('skyl_xic_report_vpw16_3.csv', skiprows=1, header=None) 	###############################################################################################
	#trdfchunk=pd.read_csv('skyl_xic_report_vpw16_3.csv', skiprows=1, header=None, iterator=True, chunksize=1000)
	#trdf=pd.concat(trdfchunk, ignore_index=True)
	if segr==0:
		spr=1
	else:
		spr=segr+1  ##############################################
	nrw=segs-segr+1
	segtrdf=pd.read_csv('skyl_xic_report_vpw20_3_intensities.csv', skiprows=spr, header=None, nrows=nrw, low_memory=False)

	#begin write trdf to csv file to check its contents 	# TROUBLESHOOTING
	#trdf.to_csv('skyl_xic_report_vpw16_3_troubleshooting.csv', index=False)
	#end write trdf to csv file to check its contents 	# TROUBLESHOOTING

	#trdf=trdf.transpose()#
	allxiclist=segtrdf.values.tolist()

	print(allxiclist[0][1])		####################################### 
	kix=len(allxiclist)
	print('Number of rows in segment in skyl_xic_report_vpw20_3.csv: %d' % kix)	#May include species that were excluded as false positives based on decoy filter
	#begin determine columns in xic_results, determine length of XIC
	ci=len(allxiclist[0])
	xiclength=int((ci-8)) #/2
	#end determine columns in xic_results, determine length of XIC
	ki=len(dwritelist[0])
	print('Number of rows in segment after decoy filter: %d' % ki)	#May include species that were excluded as false positives based on decoy filter
	#print(dwritelist[1]) 	# list after decoy filter
	r=0
	ki=ki
	while r<ki:		# go through rows of list with positives after decoy filter 
		e=dwritelist[1][r] ## Precursorname		# begin determine which row to start (r) and to end (s)
		#print(e)
		s=r+1
		st=0
		while st<1:
			if s>(len(dwritelist[1])-1):
				ne='stop_loop'
			else:
				ne=dwritelist[1][s] #	# Precursorname
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		t=r
		pos=0	# delete block later if pos=1		
		#begin determine exactexrtdiff for current species
		currentfwhm=[]
		currentexrt=[]
		exactexrtlist=[]
		xxrtintlist=[]
		diagnosticxicrtlist=[]
		diagnosticxicintlist=[]
		while t<(s-1):	#only transitions, not decoy or precursor
			#print(dwritelist[16][t])
			cfwhm=dwritelist[16][t]
			currentfwhm.append(cfwhm)
			cexrt=dwritelist[17][t]
			currentexrt.append(cexrt)
			t=t+1
		t=r
		if t<(s-1):
			precnm=dwritelist[1][t]
			xk=0
			allgexrtlist=[]		# list of lists with retention times that were determined by fitting routine for each run (A, B, C) and each transition
			allgcxicint=[]		# list of lists with XICs of transitions (only intensity)
			allgcxicrt=[]		# list of lists with XICs of transitions (only retention time)
			#print('len(allxiclist)')
			#print(len(allxiclist))
			#print('len(allxiclist)')
			while xk<len(allxiclist): 
				if precnm==allxiclist[xk][1]:
					if str(allxiclist[xk][4][len(allxiclist[xk][4])-1])=='r':
						nrt=0
					elif str(allxiclist[xk][4][len(allxiclist[xk][4])-1])=='y':
						nrt=0
					else:	
						cxicrt=[]	# get current xic (part around +- xicrange (2*fwhm) of exrt) - each full xic made up of 4730 datapoints (always the case?) 
						cxicint=[]
						cindex=[]
						cfwhm=dwritelist[16][t]
						if cfwhm==0:
							cfwhm=0.004
						elif str(cfwhm)=='nan':
							cfwhm=0.004
						#print('cfwhm is:')
						#print(cfwhm)
						#print('exrt is:')
						#print(float(dwritelist[17][t]))
						xicrange=cfwhm*3 # 2 or 2.2 # set range of used XIC for fitting routine
						if cfwhm<0.03:
							xicrange=0.09
						kxicrt=8												# begin get the XIC as measured (for each transition of the current species)
						while kxicrt<(xiclength+8):
							cxx=xictimeslist[kxicrt-8]
							if cxx>(float(dwritelist[17][t])-xicrange):
								if cxx<(float(dwritelist[17][t])+xicrange):
									cxicrt.append(cxx)
									cindex.append((kxicrt-8))
							kxicrt=kxicrt+1
						kxicint=8
						while kxicint<((xiclength)+8):
							cxx=allxiclist[xk][(kxicint)]
							if kxicint>(cindex[0]-1):
								if kxicint<(cindex[len(cindex)-1]+1):
									cxicint.append(cxx)
							kxicint=kxicint+1					# end get the XIC as measured (for each transition of the current species)
						if len(cxicrt)<len(cxicint):
							print('CHECK XIC INDEXING, len(cxicrt)<len(cxicint)')
							#print(len(cxicrt))
							#print(len(cxicint))
							while len(cxicrt)<(len(cxicint)):
								#print('fixed indexing')
								fix=cxicrt[len(cxicrt)-1]
								cxicrt.append(fix)
						if len(cxicrt)<len(cxicint):
							print('NOT REPAIRED YET !')
						elif len(cxicrt)>len(cxicint):
							print('CHECK XIC INDEXING, len(cxicrt)>len(cxicint)')
							#print(len(cxicrt))
							#print(len(cxicint))
							while len(cxicrt)>(len(cxicint)):
								#print('fixed indexing')
								fix=cxicint[len(cxicint)-1]
								cxicint.append(fix)
						nrt=0
						while cxicrt[nrt]<dwritelist[17][t]: # dwritelist[17][t] is exrt
							intexrt=cxicint[nrt]			 # intexrt after this loop is intensity at exrt
							nrt=nrt+1
						#print('intensity at exrt is:')
						#print(intexrt)
						cfwhm=dwritelist[16][t]
						if cfwhm==0:
							cfwhm=0.004
						elif str(cfwhm)=='nan':
							cfwhm=0.004
						cexrt=dwritelist[17][t]

						#print('cxicint is:')
						#print(cxicint)
						#print('cxicrt is:')
						#print(cxicrt)
						#print('cindex is:')
						#print(cindex)

						##### begin 5 point smoothing of XIC
						#fpsxicrt=[]
						#fpsxicint=[]
						#nrt=2
						#while nrt<(len(cxicrt)-2):
						#	fpsn=cxicrt[nrt]
						#	fpsxicrt.append(fpsn)
						#	nrt=nrt+1
						#nrt=2
						#while nrt<(len(cxicint)-2):
						#	fpsn=((cxicint[nrt-2]+cxicint[nrt-1]+cxicint[nrt]+cxicint[nrt+1]+cxicint[nrt+2])/5)
						#	fpsxicint.append(fpsn)
						#	nrt=nrt+1
						##### end 5 point smoothing of XIC

						##### begin 7 point smoothing of XIC		##################################################################################################
						fpsxicrt=[]
						fpsxicint=[]
						nrt=3
						while nrt<(len(cxicrt)-3):
							fpsn=cxicrt[nrt]
							fpsxicrt.append(fpsn)
							nrt=nrt+1
						nrt=3
						while nrt<(len(cxicint)-3):
							fpsn=((cxicint[nrt-3]+cxicint[nrt-2]+cxicint[nrt-1]+cxicint[nrt]+cxicint[nrt+1]+cxicint[nrt+2]+cxicint[nrt+3])/7)
							fpsxicint.append(fpsn)
							nrt=nrt+1
						##### end 7 point smoothing of XIC				##################################################################################################

						##### begin find max near cexrt in fpsxicint/fpsxicrt (find true center of peak using smoothed xic)
						nrt=0
						#print('fpsxicint is:')
						#print(fpsxicint)
						#print('fpsxicrt is:')
						#print(fpsxicrt)
						#print(cexrt)
						while fpsxicrt[nrt]<cexrt:
							cfpsxicrt=fpsxicrt[nrt]
							nrt=nrt+1
						nrt=nrt-1
						if fpsxicint[nrt]<fpsxicint[nrt+1]:			#cexrt before peakmax
							stop=0
							while fpsxicint[nrt]<fpsxicint[nrt+1]:
								cfpsxicrt=fpsxicrt[nrt+1]
								cfpsxicint=fpsxicint[nrt+1]
								if (nrt+1)==(len(fpsxicint)-1):
									if stop==0:
										nrt=(-1)
										stop=1
									else:
										fpsxicint.append(0)
								nrt=nrt+1
						else:										#cexrt after peakmax
							stop=0
							while fpsxicint[nrt]>fpsxicint[nrt+1]:
								cfpsxicrt=fpsxicrt[nrt]
								cfpsxicint=fpsxicint[nrt]
								if (nrt)==0:
									if stop==0:
										nrt=(len(fpsxicint)-1)
										stop=1
									else:
										nrt=(len(fpsxicint)-1)
										fpsxicint.append(0)
								nrt=nrt-1
							nrt=nrt+1
						##### begin find max near cexrt in fpsxicint (find true center of peak using smoothed xic)
						if str(diagnostics[xk])=='diagnostic':
							exactexrtlist.append(cfpsxicrt)
							xxrtintlist.append(cfpsxicint)
							diagnosticxicrtlist.append(fpsxicrt)
							diagnosticxicintlist.append(fpsxicint)
						#begin troubleshooting, display current XIC and index for 20:3_n-7_n-10_n-13_7.8
						if str(allxiclist[xk][1])=='AMPP_20:3_n-1_n-10_n-13_7.853':
							if str(allxiclist[xk][4])=='AMPP_20:3_n-1_n-10_n-13_criegee neutral loss from n-10':
								print('AMPP_20:3_n-1_n-10_n-13_criegee neutral loss from n-10')
								print(cxicrt)
								print(cxicint)
								print(fpsxicrt)
								print(fpsxicint)
							elif str(allxiclist[xk][4])=='AMPP_18:1_n-9_aldehyde neutral loss from n-10':
								print('AMPP_18:1_n-9_aldehyde neutral loss from n-10')
								print(cxicrt)
								print(cxicint)
								print(fpsxicrt)
								print(fpsxicint)
						#if str(allxiclist[xk][1])=='AMPP_21:3_n-1_n-3_n-5_10.88':
							#if str(allxiclist[xk][4])=='AMPP_21:3_n-1_n-3_n-5_criegee neutral loss from n-1':
								#print('AMPP_18:1_n-9_criegee neutral loss from n-1')
								#print(cxicrt)
								#print(cxicint)
								#print(fpsxicrt)
								#print(fpsxicint)
							#elif str(allxiclist[xk][4])=='AMPP_21:3_n-1_n-3_n-5_aldehyde neutral loss from n-1':
								#print('AMPP_21:3_n-1_n-3_n-5_aldehyde neutral loss from n-1')
								#print(cxicrt)
								#print(cxicint)
								#print(fpsxicrt)
								#print(fpsxicint)
							
						#end troubleshooting display current XIC and index for 20:3_n-7_n-10_n-13_7.8					
						#begin troubleshooting, display current XIC and index for 20:3_n-7_n-10_n-13_7.8
						#if str(allxiclist[xk][1])=='AMPP_20:3_n-6_n-9_n-12_7.6':
						#	if str(allxiclist[xk][4])=='AMPP_20:3_n-6_n-9_n-12_criegee neutral loss from n-12':
						#		print('AMPP_20:3_n-6_n-9_n-12_criegee neutral loss from n-12')
						#		print('cfpsxicrt')
						#		print(cfpsxicrt)
						#		#print(cindex)
						#end troubleshooting display current XIC and index for 20:3_n-7_n-10_n-13_7.8

				xk=xk+1
			exactexrtdiff=(max(exactexrtlist))-(min(exactexrtlist))  #largest difference between exactexrt values of transitions
			avgexactexrt=mean([mean(exactexrtlist), median(exactexrtlist)])	

			#avgexactexrt=statistics.mean(exactexrtlist)
			#print(dwritelist[1][t])
			#print(exactexrtlist)
			#print(exactexrtdiff)
			#print(dwritelist[1][t])
			#print(avgexactexrt)
			#quit()


		else:
			exactexrtdiff=0
			#sat FA, need to generate avgexactexrt !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
			precnm=dwritelist[1][t]
			xk=0
			allgexrtlist=[]		# list of lists with retention times that were determined by fitting routine for each run (A, B, C) and each transition
			allgcxicint=[]		# list of lists with XICs of transitions (only intensity)
			allgcxicrt=[]		# list of lists with XICs of transitions (only retention time)
			while xk<len(allxiclist):
				if precnm==allxiclist[xk][1]:
					if len(allxiclist[xk][4])==9:	# detect saturated FA
						cxicrt=[]	# get current xic (part around +- xicrange (2*fwhm) of exrt) - each full xic made up of 4730 datapoints (always the case?) 
						cxicint=[]
						cindex=[]
						cfwhm=dwritelist[16][t]
						xicrange=cfwhm*3 # 2 or 2.2 # set range of used XIC for fitting routine
						if cfwhm<0.03:
							xicrange=0.09
						kxicrt=8
						while kxicrt<(xiclength+8):
							cxx=xictimeslist[kxicrt-8]
							if cxx>(float(dwritelist[17][t])-xicrange):
								if cxx<(float(dwritelist[17][t])+xicrange):
									cxicrt.append(cxx)
									cindex.append(kxicrt)
							kxicrt=kxicrt+1
						kxicint=8
						while kxicint<((xiclength)+8):
							cxx=allxiclist[xk][kxicint]
							if kxicint>(cindex[0]-1):
								if kxicint<(cindex[len(cindex)-1]+1):
									cxicint.append(cxx)
							kxicint=kxicint+1
						if len(cxicrt)<len(cxicint):
							print('CHECK XIC INDEXING, len(cxicrt)<len(cxicint)')
							#print(len(cxicrt))
							#print(len(cxicint))
							while len(cxicrt)<(len(cxicint)):
								#print('fixed indexing')
								fix=cxicrt[len(cxicrt)-1]
								cxicrt.append(fix)
						elif len(cxicrt)>len(cxicint):
							print('CHECK XIC INDEXING, len(cxicrt)>len(cxicint)')
							#print(len(cxicrt))
							#print(len(cxicint))
							while len(cxicrt)>(len(cxicint)):
								#print('fixed indexing')
								fix=cxicint[len(cxicint)-1]
								cxicint.append(fix)
						nrt=0
						while cxicrt[nrt]<dwritelist[17][t]: # dwritelist[17][t] is exrt
							intexrt=cxicint[nrt]			 # intexrt after this loop is intensity at exrt
							nrt=nrt+1
						#print(intexrt)
						cfwhm=dwritelist[16][t]
						if cfwhm==0:
							cfwhm=0.004
						cexrt=dwritelist[17][t]
						##### begin 5 point smoothing of XIC
						fpsxicrt=[]
						fpsxicint=[]
						nrt=2
						while nrt<(len(cxicrt)-2):
							fpsn=cxicrt[nrt]
							fpsxicrt.append(fpsn)
							nrt=nrt+1
						nrt=2
						while nrt<(len(cxicint)-2):
							fpsn=((cxicint[nrt-2]+cxicint[nrt-1]+cxicint[nrt]+cxicint[nrt+1]+cxicint[nrt+2])/5)
							fpsxicint.append(fpsn)
							nrt=nrt+1
						##### end 5 point smoothing of XIC
						##### begin find max near cexrt in fpsxicint/fpsxicrt (find true center of peak using smoothed xic)
						nrt=0
						#print(fpsxicint)
						#print(fpsxicrt)
						#print(cexrt)
						while fpsxicrt[nrt]<cexrt:
							cfpsxicrt=fpsxicrt[nrt]
							nrt=nrt+1
						nrt=nrt-1
						if fpsxicint[nrt]<fpsxicint[nrt+1]:			#cexrt before peakmax
							stop=0
							while fpsxicint[nrt]<fpsxicint[nrt+1]:
								cfpsxicrt=fpsxicrt[nrt+1]
								cfpsxicint=fpsxicint[nrt+1]
								if (nrt+1)==(len(fpsxicint)-1):
									if stop==0:
										nrt=(-1)
										stop=1
									else:
										fpsxicint.append(0)
								nrt=nrt+1
						else:										#cexrt after peakmax
							stop=0
							while fpsxicint[nrt]>fpsxicint[nrt+1]:
								cfpsxicrt=fpsxicrt[nrt]
								cfpsxicint=fpsxicint[nrt]
								if (nrt)==0:
									if stop==0:
										nrt=(len(fpsxicint)-1)
										stop=1
									else:
										nrt=(len(fpsxicint)-1)
										fpsxicint.append(0)
								nrt=nrt-1
							nrt=nrt+1
						##### begin find max near cexrt in fpsxicint (find true center of peak using smoothed xic)
						exactexrtlist.append(cfpsxicrt)
						xxrtintlist.append(cfpsxicint)
						#begin troubleshooting, display current XIC and index for 20:3_n-7_n-10_n-13_7.8
						#if str(allxiclist[xk][1])=='AMPP_18:1_n-9_7.88':
						#	if str(allxiclist[xk][4])=='AMPP_18:1_n-9_criegee neutral loss':
						#		print('AMPP_18:1_n-9_criegee neutral loss')
						#		print(cxicrt)
						#		print(fpsxicrt)
						#		print(fpsxicint)
						#	elif str(allxiclist[xk][4])=='AMPP_18:1_n-9_aldehyde neutral loss':
						#		print('AMPP_18:1_n-9_aldehyde neutral loss')
						#		print(cxicrt)
						#		print(cxicint)
						#		print(fpsxicrt)
						#		print(fpsxicint)	
						#end troubleshooting display current XIC and index for 20:3_n-7_n-10_n-13_7.8					
						#begin troubleshooting, display current XIC and index for 20:3_n-7_n-10_n-13_7.8
						#if str(allxiclist[xk][1])=='AMPP_20:3_n-6_n-9_n-12_7.6':
						#	if str(allxiclist[xk][4])=='AMPP_20:3_n-6_n-9_n-12_criegee neutral loss from n-12':
						#		print('AMPP_20:3_n-6_n-9_n-12_criegee neutral loss from n-12')
						#		print('cfpsxicrt')
						#		print(cfpsxicrt)
						#		#print(cindex)
						#end troubleshooting display current XIC and index for 20:3_n-7_n-10_n-13_7.8
				xk=xk+1
			exactexrtdiff=0 #exactexrtdiff=(max(exactexrtlist))-(min(exactexrtlist))  #largest difference between exactexrt values of transitions	
			avgexactexrt=exactexrtlist[0] #avgexactexrt=statistics.mean(exactexrtlist)

		if exactexrtdiff>exactexrtthreshold:		# exclude species for which transitions deviate too much in RT 

			# begin check for false negative: if sufficiently similar transitions are present, despite no matching local maximum, then keep species
			prt=str(dwritelist[1][t])
			#print(prt)
			#print('exactexrtlist')
			#print(exactexrtlist)
			refgrouplist=[]
			ipos=0
			i=0
			groupindex=1
			while i<(len(exactexrtlist)):	# find outliers, mark as 0 in refgrouplist, all others mark as 1
				j=0
				ipos=0
				while j<(len(exactexrtlist)):	
					if j==i:
						ipos=ipos
					elif abs(exactexrtlist[i]-exactexrtlist[j])<exactexrtthreshold:
						ipos=ipos+1 	
					j=j+1
				if ipos>0:
					refgrouplist.append(groupindex)
				else:
					refgrouplist.append(0)
				i=i+1
			i=0
			while i<(len(refgrouplist)):
				j=0
				while j<(len(refgrouplist)):
					if j==i:
						j=j
					else:
						if refgrouplist[i]>0:
							if refgrouplist[j]>0:
								if refgrouplist[i]==refgrouplist[j]:
									if (abs(exactexrtlist[i]-exactexrtlist[j]))<exactexrtthreshold:
										j=j
									else:
										if j>i:
											refgrouplist[j]=refgrouplist[j]+1 #elevate index to higher group
										else:
											refgrouplist[i]=refgrouplist[i]+1 #elevate index to higher group
								else:
									if (abs(exactexrtlist[i]-exactexrtlist[j]))<exactexrtthreshold:
										if refgrouplist[j]>refgrouplist[i]:
											refgrouplist[i]=refgrouplist[j] #elevate index to higher group
										else:
											refgrouplist[j]=refgrouplist[i] #elevate index to higher group
					j=j+1
				i=i+1
			#print('refgrouplist')
			#print(refgrouplist)
			#end build refgrouplist
			hg=max(refgrouplist)
			cch=1
			cut=0
			while cch<(hg+1):
				check=refgrouplist.count(cch)
				if check==1:
					cut=1
				cch=cch+1
			if hg>1:
				check=refgrouplist.count(1)
				if check==0:
					cut=1
			if cut==0:		
				avgxxrtlist=[]
				stdxxrtlist=[]
				meanxxrtlist=[]
				iref=0
				rgsuma=0
				rga=0
				alist=[]
				rgsumb=0
				rgb=0
				blist=[]
				rgsumc=0
				rgc=0
				clist=[]
				while iref<(len(refgrouplist)):
					if refgrouplist[iref]==1:
						rgsuma=rgsuma+exactexrtlist[iref]
						rga=rga+1
						aa=xxrtintlist[iref]
						alist.append(aa)
					if refgrouplist[iref]==2:
						rgsumb=rgsumb+exactexrtlist[iref]
						rgb=rgb+1
						aa=xxrtintlist[iref]
						blist.append(aa)
					if refgrouplist[iref]==3:
						rgsumc=rgsumc+exactexrtlist[iref]
						rgc=rgc+1
						aa=xxrtintlist[iref]
						clist.append(aa)
					iref=iref+1
				#print(alist)
				#print(blist)
				#print(clist)
				# begin troubleshooting
				#if rgb==0:
					#print('****************')
					#print(dwritelist[1][t])
					#print(rgsuma)
					#print(rga)
					#print(rgsumb)
					#print(rgb)
					#print(rgsumc)
					#print(rgc)
				# end troubleshooting
				hg=max(refgrouplist)
				if hg>1:
					if rgb==0:
						hg=0
					if rgc==0:
						hg=0
				if hg>0:
					avgxxrt=rgsuma/rga
					avgxxrtlist.append(avgxxrt)
					stdxxrt=statistics.stdev(alist)
					stdxxrtlist.append(stdxxrt)
					meanxxrt=statistics.mean(alist)
					meanxxrtlist.append(meanxxrt)
				if hg>1:
					avgxxrt=rgsumb/rgb
					avgxxrtlist.append(avgxxrt)
					stdxxrt=statistics.stdev(blist)
					stdxxrtlist.append(stdxxrt)
					meanxxrt=statistics.mean(blist)
					meanxxrtlist.append(meanxxrt)
				if hg>2:
					avgxxrt=rgsumc/rgc
					avgxxrtlist.append(avgxxrt)
					stdxxrt=statistics.stdev(clist)
					stdxxrtlist.append(stdxxrt)
					meanxxrt=statistics.mean(clist)
					meanxxrtlist.append(meanxxrt)
				# end build avgxxrtlist
				#print(avgxxrtlist)
				#print(stdxxrtlist)
				#print(meanxxrtlist)

				opos=1
				aopos=1
				iavg=0	#group is iavg+1 #go though groups and check if current main group fits criteria (all transition intensities align within stdev)
				while iavg<(len(avgxxrtlist)):
					iref=0
					opos=1
					cpos=0
					while iref<(len(refgrouplist)):
						if refgrouplist[iref]==(iavg+1):
							iref=iref
						else:
							# test if outlier within stdev 
							xi=0
							while diagnosticxicrtlist[iref][xi]<avgxxrtlist[iavg]:
								outlier=diagnosticxicintlist[iref][xi]
								xi=xi+1
							outlier=(outlier+diagnosticxicintlist[iref][xi])/2
							#print('outlier')
							#print(outlier)
							timesstddev=6 	#1.5 			###################################################### ADJUST STD DEV CUTOFF #####
							if outlier>(meanxxrtlist[iavg]-(timesstddev*stdxxrtlist[iavg])):
								if outlier<(meanxxrtlist[iavg]+(timesstddev*stdxxrtlist[iavg])):
									opos=0
								else:
									cpos=1
							else:
								cpos=1
						iref=iref+1
					if opos==0:
						if cpos==0:
							aopos=0
					iavg=iavg+1
			else:
				aopos=1
			# end check for false negative (keep species if aopos=0; cut is aopos=1)
			if aopos==1:
				pos=1 #1 put as 1 to delete this species
				#print('CUT')
				xcut=str('CUT')
				nct=s-r+1
				kct=0
				while kct<nct:
					cutlist.append(xcut)
					avgexactexrtlist.append(avgexactexrt)
					kct=kct+1
			else:
				if mostwanted==0:
					pos=2
				else:
					# begin test whether species to be cut is in the mostwantedlist
					fashort=str()
					fas=5
					while fas<(len(dwritelist[1][t])-5):
						fashort=fashort+str(dwritelist[1][t][fas])
						fas=fas+1
					mwl=0
					gposcancel=0
					while mwl<(len(mostwantedlist)):
						cmwl=str(mostwantedlist[mwl])
						if cmwl==fashort:
							gposcancel=1
							print('************** retained ***')
							print(cmwl)
						mwl=mwl+1
					if gposcancel==0:
						pos=2
					else:
						pos=0	# rank 1
					# end test whether species to be cut is in the mostwantedlist

				#pos=2		# divert species into backup file, rank2
				if pos==0:
					xcut=str('OK')
				else:
					xcut=str('CHECK')
				#print('check')
				nct=s-r+1
				kct=0
				while kct<nct:
					cutlist.append(xcut)
					avgexactexrtlist.append(avgexactexrt)
					kct=kct+1
		else:
			#print('OK')		# rank1
			pos=0
			xcut=str('OK')
			nct=s-r+1
			kct=0
			while kct<nct:
				cutlist.append(xcut)
				avgexactexrtlist.append(avgexactexrt)
				kct=kct+1
		#end determine exactexrtdiff for current species
		# begin send species with an n-1 transition to rank2 (delete is pos=1; rank1 is pos=0; rank2 is pos=2)
		if pos==0:
			testname=str(dwritelist[1][t])
			ttn=testname.find('n-1_')
			if ttn>0:
				pos=2
		# begin send species with an n-1 transition to rank2 (delete is pos=1; rank1 is pos=0; rank2 is pos=2)
		if pos==1:				# species deleted
			r=s+1
		elif pos==2:
			t=r					# rank 2
			while t<(s+1):
				apos=0
				if apos==0:
					e=dwritelist[0][t] #sheetinput.cell(row=t, column=1)	# mlistname	# begin append rows of suitable species to lists for later saving in excel output
					fmlistname.append(e)
					e=dwritelist[1][t] #sheetinput.cell(row=t, column=2)	# precname	
					fprecname.append(e)
					e=dwritelist[2][t] #sheetinput.cell(row=t, column=3)	# prodname	
					fprecformula.append(e)
					e=dwritelist[3][t] #sheetinput.cell(row=t, column=4)	# prodformula	
					fprecadduct.append(e)
					e=dwritelist[4][t] #sheetinput.cell(row=t, column=5)	# 	
					fprecmz.append(e)
					e=dwritelist[5][t] #sheetinput.cell(row=t, column=6)	# 	
					fprecchrg.append(e)
					e=dwritelist[6][t] #sheetinput.cell(row=t, column=7)	# 	
					fprodname.append(e)
					e=dwritelist[7][t] #sheetinput.cell(row=t, column=8)	# 	
					fprodformula.append(e)
					e=dwritelist[8][t] #sheetinput.cell(row=t, column=9)	# 	
					fprodadduct.append(e)
					e=dwritelist[9][t] #sheetinput.cell(row=t, column=10)	# 
					fprodmz.append(e)
					e=dwritelist[10][t] #sheetinput.cell(row=t, column=11)	# 	
					fprodchrg.append(e)
					e=dwritelist[11][t] #sheetinput.cell(row=t, column=12)	# 	
					fmzerror.append(e)
					e=dwritelist[12][t] #sheetinput.cell(row=t, column=13)	# 	
					frettime.append(e)
					e=dwritelist[13][t] #sheetinput.cell(row=t, column=14)	# 	
					farea.append(e)
					e=dwritelist[14][t] #sheetinput.cell(row=t, column=15)	# 	
					fareanormalpercent.append(e)
					e=dwritelist[15][t] #sheetinput.cell(row=t, column=16)	# 	
					fbackground.append(e)
					e=dwritelist[16][t] #sheetinput.cell(row=t, column=17)	# 	
					ffwhm.append(e)
					e=dwritelist[17][t] #sheetinput.cell(row=t, column=18)	# 	
					fexplicitrt.append(e)
					e=dwritelist[18][t] #sheetinput.cell(row=t, column=19)	# 	
					frtstart.append(e)
					e=dwritelist[19][t] #sheetinput.cell(row=t, column=20)	# 	
					frtend.append(e)
					fexrtwindow.append(0.05)
					e=cutlist[t]
					fwcutlist.append(e)
					e=avgexactexrtlist[t]
					fwavgexactexrtlist.append(e)
				t=t+1
		else:
			t=r				# rank 1
			while t<(s+1):
				apos=0
				if apos==0:
					e=dwritelist[0][t] ## mlistname	# begin append rows of suitable species to lists for later saving in excel output
					mlistname.append(e)
					e=str(dwritelist[1][t]) ## precname		# begin reassign RT label
					aee=str(format(avgexactexrt, '.2f'))
					#go=1
					#while go==1:
					#	if e[(len(e)-1)]=='_':
					#		go=0
					#	else:
					#		e=e[:-1:]
					e=e+aee						# end reassign RT label
					precname.append(e)
					e=dwritelist[2][t] ## precformula	
					precformula.append(e)
					e=dwritelist[3][t] ## precadduct	
					precadduct.append(e)
					e=dwritelist[4][t] ## 	
					precmz.append(e)
					e=dwritelist[5][t] ## 	
					precchrg.append(e)
					e=dwritelist[6][t] ## 	
					prodname.append(e)
					e=dwritelist[7][t] ## 	
					prodformula.append(e)
					e=dwritelist[8][t] ## 	
					prodadduct.append(e)
					e=dwritelist[9][t] ## 
					prodmz.append(e)
					e=dwritelist[10][t] ## 	
					prodchrg.append(e)
					e=dwritelist[11][t] ## 	
					mzerror.append(e)
					e=dwritelist[12][t] ## 	
					rettime.append(e)
					e=dwritelist[13][t] ## 	
					area.append(e)
					e=dwritelist[14][t] ## 	
					areanormalpercent.append(e)
					e=dwritelist[15][t] ## 	
					background.append(e)
					e=dwritelist[16][t] ## 	
					fwhm.append(e)
					#e=dwritelist[17][t] ## 	Explicit RT for final rank 1 transition list
					#if int(dwritelist[1][t][8])==0:
					e=float(avgexactexrt)				########## Test, if this is good #########################################################################
					explicitrt.append(e)
					e=dwritelist[18][t] ## 	
					rtstart.append(e)
					e=dwritelist[19][t] ## 	
					rtend.append(e)
					exrtwindow.append(0.1)
					e=cutlist[t]
					wcutlist.append(e)
					e=avgexactexrtlist[t]
					wavgexactexrtlist.append(e)
				t=t+1
		r=s+1
	# end read transition results and append suitable species to lists
	# end filter out species, if exact retention time of transitions differ too much #####################
	######################################################################################################
	# begin filter out duplicates with exact same precname
	fdp=len(precname) #0
	while fdp<(len(precname)):
		r=fdp
		e=precname[r] # Precursorname		# begin determine which row to start (r) and to end (s)
		s=r+1
		st=0
		while st<1:
			if s>(len(precname)-1):
				ne='stop_loop'
			elif str(prodname[s][(len(prodname[s])-1)])=='r':
				s=s+1
				ne='stop_loop'
			else:
				ne=precname[s] # Precursorname
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		fdf=0
		while fdf<(len(precname)):
			rf=fdp
			e=precname[rf] # Precursorname		# begin determine which row to start (rf) and to end (sf)
			sf=rf+1
			st=0
			while st<1:
				if sf>(len(precname)-1):
					ne='stop_loop'
				elif str(prodname[sf][(len(prodname[sf])-1)])=='r':
					sf=sf+1
					ne='stop_loop'
				else:
					ne=precname[sf] # Precursorname
				if ne==e:
					sf=sf+1
					st=0
				else:
					sf=sf-1
					st=1		# end determine sf

			if sf==s:
				sf=sf
			elif precname[sf]==precname[s]:
				if str(prodname[sf][len(prodname[sf])-1])=='r':
					if str(prodname[s][len(prodname[s])-1])=='r':
						if sf>s:
							delfd=rf #fdf-(0+(2*(int(precname[fdf][8]))))
							delfdt=sf
							countd=delfdt-delfd+1
							#fdf=0 #fdf-(1+(2*(int(precname[fdf][8]))))
							#fdp=0
						else:
							delfd=r #fdp-(0+(2*(int(precname[fdp][8]))))
							delfdt=s #fdp
							countd=delfdt-delfd+1
							#fdp=0 #fdp-(1+(2*(int(precname[fdp][8]))))
							#fdf=0
						# remove entries with index delfd
						idel=0
						while idel<countd:
							del mlistname[delfd]
							del precname[delfd]
							del precformula[delfd]
							del precadduct[delfd]
							del precmz[delfd]
							del precchrg[delfd]
							del prodname[delfd]
							del prodformula[delfd]
							del prodadduct[delfd]
							del prodmz[delfd]
							del prodchrg[delfd]
							del mzerror[delfd]
							del rettime[delfd]
							del area[delfd]
							del areanormalpercent[delfd]
							del background[delfd]
							del fwhm[delfd]
							del explicitrt[delfd]
							del rtstart[delfd]
							del rtend[delfd]
							del exrtwindow[delfd]
							del cutlist[delfd]
							del wavgexactexrtlist[delfd]
							idel=idel+1
			fdf=sf+1
		fdp=s+1
	# begin filter out duplicates with exact same precname
	######################################################################################################
	# begin detect and delete duplicates, keep best matching duplicate based on avgexactexrt information
	#go through lists, find block, / go through list, find another block with both same prodnames and same avgexactexrt - list all these blocks (first transition) 
	#in list, keep block with closest matching exactexrt (column L) to avgexactexrt, delete others
	keepindexlist=[]	# 0 for index that will finally be kept, 1 for index that will finally be deleted (confirmed duplicate for removal)
	dmlistname=[]
	dprecname=[]
	dprecformula=[]
	dprecadduct=[]
	dprecmz=[]
	dprecchrg=[]
	dprodname=[]
	dprodformula=[]
	dprodadduct=[]
	dprodmz=[]
	dprodchrg=[]
	#dexrt=[]
	dexrtwindow=[]
	dmzerror=[]
	drettime=[]
	darea=[]
	dareanormalpercent=[]
	dbackground=[]
	dfwhm=[]
	dexplicitrt=[]
	drtstart=[]
	drtend=[]
	dwcutlist=[]
	dwavgexactexrtlist=[]
	r=0
	ki=len(mlistname)
	while r<ki:
		keepindexlist.append(0)		# default keep all entries. Below look for duplicates and mark for removal
		r=r+1

	r=0
	ki=len(mlistname)
	while r<ki:		# go through rows of lists 
		e=precname[r] # Precursorname		# begin determine which row to start (r) and to end (s)
		s=r+1
		st=0
		while st<1:
			if s>(len(precname)-1):
				ne='stop_loop'
			else:
				ne=precname[s] # Precursorname
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		t=r
		testindexlist=[]	# 0 for index not marked for comparison (no duplicate), 1 for block is a duplicate of the one under consideration (block r)
		rtest=0
		while rtest<ki:
			testindexlist.append(0)		# default no entries marked as duplicates
			rtest=rtest+1
		pdupscore=[]
		pdupindex=[]
		rtest=0
		while rtest<ki:
			e=precname[rtest] # Precursorname		# begin determine which row to start (r) and to end (s)
			stest=rtest+1
			sttest=0
			while sttest<1:
				if stest>(len(precname)-1):
					ne='stop_loop'
				else:
					ne=precname[stest] # Precursorname
				if ne==e:
					stest=stest+1
					sttest=0
				else:
					stest=stest-1
					sttest=1		# end determine s
			ttest=rtest

			if prodname[t]==prodname[ttest]:
				if wavgexactexrtlist[t]==wavgexactexrtlist[ttest]:
					testindexlist[ttest]=1			# each first transition in current duplicate block marked
					pdupindex.append(ttest)
					rtestdiff=abs((float(wavgexactexrtlist[ttest]))-(float(explicitrt[ttest])))
					rtestdiff=float(rtestdiff)
					pdupscore.append(rtestdiff)
					#print('check2')
			rtest=stest+1
		rdup=0
		while rdup<(len(pdupscore)):
			if pdupscore[rdup]==(min(pdupscore)):
				keepindexlist[pdupindex[rdup]]=0
			else:
				keepindexlist[pdupindex[rdup]]=1 ######## 
				#print('check1')
			rdup=rdup+1
		# current duplicates are marked (1) and duplicates with not minimum difference between exrt and avgexactexrt are marked for removal ('CUT')
		r=s+1		# go to beginning of next block (next FA)
	# duplicates marked for removal in keepindexlist (only first transition of block is marked)
	r=0
	ki=len(mlistname)
	while r<ki:		# go through rows of lists
		e=precname[r] # Precursorname		# begin determine which row to start (r) and to end (s)
		s=r+1
		st=0
		while st<1:
			if s>(len(precname)-1):
				ne='stop_loop'
			else:
				ne=precname[s] # Precursorname
			if ne==e:
				s=s+1
				st=0
			else:
				s=s-1
				st=1		# end determine s
		t=r
		if keepindexlist[r]==1:
			t=s+1
		else:
			while t<s+1:
				dmlistname.append(mlistname[t])
				dprecname.append(precname[t])
				dprecformula.append(precformula[t])
				dprecadduct.append(precadduct[t])
				dprecmz.append(precmz[t])
				dprecchrg.append(precchrg[t])
				dprodname.append(prodname[t])
				dprodformula.append(prodformula[t])
				dprodadduct.append(prodadduct[t])
				dprodmz.append(prodmz[t])
				dprodchrg.append(prodchrg[t])
				#dexrt.append(exrt[t])
				dexrtwindow.append(exrtwindow[t])
				dmzerror.append(mzerror[t])
				drettime.append(rettime[t])
				darea.append(area[t])
				dareanormalpercent.append(areanormalpercent[t])
				dbackground.append(background[t])
				dfwhm.append(fwhm[t])
				dexplicitrt.append(explicitrt[t])
				drtstart.append(rtstart[t])
				drtend.append(rtend[t])
				dwcutlist.append(wcutlist[t])
				dwavgexactexrtlist.append(wavgexactexrtlist[t])
				t=t+1
		r=s+1

	mlistname=dmlistname
	precname=dprecname
	precformula=dprecformula
	precadduct=dprecadduct
	precmz=dprecmz
	precchrg=dprecchrg
	prodname=dprodname
	prodformula=dprodformula
	prodadduct=dprodadduct
	prodmz=dprodmz
	prodchrg=dprodchrg
	#exrt=dexrt
	exrtwindow=dexrtwindow
	mzerror=dmzerror
	rettime=drettime
	area=darea
	areanormalpercent=dareanormalpercent
	background=dbackground
	fwhm=dfwhm
	explicitrt=dexplicitrt
	rtstart=drtstart
	rtend=drtend
	wcutlist=dwcutlist
	wavgexactexrtlist=dwavgexactexrtlist
	# end detect and delete duplicates, keep best matching duplicate based on avgexactexrt information
	#################################################################################################################################################
	#################################################################################################################################################
	### begin add epoxide (+O transition of precursor) # begin add -H2 transition
	manualfilter=0
	if transtest==1:
		rd=0
		while rd<(len(mlistname)):
			if str(prodname[rd][len(prodname[rd])-1])=='r':
				mld=mlistname[rd]
				mlistname.insert(rd+1,mld)
				mlistname.insert(rd+2,mld)
				mld=precname[rd]
				precname.insert(rd+1,mld)
				precname.insert(rd+2,mld)
				mld=precformula[rd]
				precformula.insert(rd+1,mld)
				precformula.insert(rd+2,mld)
				mld=precadduct[rd]
				precadduct.insert(rd+1,mld)
				precadduct.insert(rd+2,mld)
				mld=precmz[rd]
				precmz.insert(rd+1,mld)
				precmz.insert(rd+2,mld)
				mld=precchrg[rd]
				precchrg.insert(rd+1,mld)
				precchrg.insert(rd+2,mld)
				#change prodname, prodMz and prodformula
				mld=prodname[rd].replace('precursor','epoxide')
				mldh=prodname[rd].replace('precursor','minus2h')
				prodname.insert(rd+1,mld)
				prodname.insert(rd+2,mldh)
				mld=float(prodmz[rd])+imass[4]
				mldh=float(prodmz[rd])-(2*imass[0])
				prodmz.insert(rd+1,mld)
				prodmz.insert(rd+2,mldh)
				# begin read precursor sum formula and # begin edit epoxide sum formula 
				e=prodformula[rd]
				#print(e)
				#print(e[0])
				clist=[]
				hlist=[]
				dlist=[]
				nlist=[]
				olist=[]
				plist=[]
				ilist=[]
				i=0
				ca=0
				ha=0
				da=0
				na=0
				oa=0
				pa=0
				ia=0
				while i<len(e):
					if e[i]=='H':
						if e[i+1]=="'":
							ha=0
						else:
							ca=0
					#if e[i]=='D':
					#	ha=0		
					if e[i]=='N':
						ha=0
						da=0
					if e[i]=='O':
						ha=0
						da=0
						na=0
					if e[i]=='P':
						ha=0
						da=0
						na=0
						oa=0
					if e[i]=='I':
						ha=0
						da=0
						na=0
						oa=0
						pa=0
					if ca==1:
						clist.append(e[i])
					if ha==1:
						hlist.append(e[i])
					if da==1:
						dlist.append(e[i])
					if na==1:
						nlist.append(e[i])
					if oa==1:
						olist.append(e[i])
					if pa==1:
						plist.append(e[i])
					if ia==1:
						ilist.append(e[i])
					if e[i]=='C':
						ca=1
					if e[i]=='H':
						if e[i+1]=="'":
							ca=0
							ha=0
							da=1
							i=i+1
						else:
							ca=0
							ha=1
					#if e[i]=='D':
					#	ca=0
					#	ha=0
					#	da=1		
					if e[i]=='N':
						ha=0
						da=0
						na=1
						if e[i+1]=='O':
							nlist.append('1')
							na=0
					if e[i]=='O':
						ha=0
						da=0
						na=0
						oa=1
						if (i+1)<len(e):
							if e[i+1]=='P':
								olist.append('1')
								oa=0
						else:
							olist.append('1')
							oa=0					
					if e[i]=='P':
						da=0
						na=0
						oa=0
						pa=1
						if (i+1)<len(e):
							if e[i+1]=='I':
								plist.append('1')
								pa=0
						else:
							plist.append('1')
							pa=0
					if e[i]=='I':
						da=0
						na=0
						oa=0
						pa=0
						ia=1
						if i==(len(e)-1):
							ilist.append('1')
							ia=0
					i=i+1
				#print(clist)
				#print(hlist)
				#print(dlist)
				#print(nlist)
				#print(olist)
				#print(plist)
				if len(clist)==0:
					cn=0
				if len(hlist)==0:
					hn=0
				if len(dlist)==0:
					dn=0	
				if len(nlist)==0:
					nn=0
				if len(olist)==0:
					on=0
				if len(plist)==0:
					pn=0
				if len(ilist)==0:
					iodon=0
				if len(clist)==1:
					cn=int(clist[0])
				if len(clist)==2:
					cn=10*int(clist[0])+int(clist[1])
				if len(clist)==3:
					cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
				if len(hlist)==1:
					hn=int(hlist[0])
				if len(hlist)==2:
					hn=10*int(hlist[0])+int(hlist[1])
				if len(hlist)==3:
					hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
				if len(dlist)==1:
					dn=int(dlist[0])
				if len(dlist)==2:
					dn=10*int(dlist[0])+int(dlist[1])
				if len(dlist)==3:
					dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
				if len(nlist)==1:
					nn=int(nlist[0])
				if len(nlist)==2:
					nn=10*int(nlist[0])+int(nlist[1])
				if len(nlist)==3:
					nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
				if len(olist)==1:
					on=int(olist[0])
				if len(olist)==2:
					on=10*int(olist[0])+int(olist[1])
				if len(olist)==3:
					on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
				if len(plist)==1:
					pn=int(plist[0])
				if len(plist)==2:
					pn=10*int(plist[0])+int(plist[1])
				if len(plist)==3:
					pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
				if len(ilist)==1:
					iodon=int(ilist[0])
				if len(ilist)==2:
					iodon=10*int(ilist[0])+int(ilist[1])
				if len(ilist)==3:
					iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
				onepoxide=on+1		# add O to create sum formula of epoxide precursor
				hnminustwoh=hn-2	# subtract 2 H to create sum formula of minus2h precursor
				decoy='C'+str(cn)+'H'+str(hn)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onepoxide)+'P'+str(pn)+'I'+str(iodon)
				minustwoh='C'+str(cn)+'H'+str(hnminustwoh)+"H'"+str(dn)+'N'+str(nn)+'O'+str(on)+'P'+str(pn)+'I'+str(iodon)
				prodformula.insert(rd+1,decoy)		# end edit sum formula of epoxide and insert in list
				prodformula.insert(rd+2,minustwoh)		# end edit sum formula of epoxide and insert in list
				mld=prodadduct[rd]
				prodadduct.insert(rd+1,mld)
				prodadduct.insert(rd+2,mld)
				mld=prodchrg[rd]
				prodchrg.insert(rd+1,mld)
				prodchrg.insert(rd+2,mld)
				mld=explicitrt[rd]
				explicitrt.insert(rd+1,mld)
				explicitrt.insert(rd+2,mld)
				mld=exrtwindow[rd]
				exrtwindow.insert(rd+1,mld)
				exrtwindow.insert(rd+2,mld)
				mld=wcutlist[rd]
				wcutlist.insert(rd+1,mld)
				wcutlist.insert(rd+2,mld)
				mld=wavgexactexrtlist[rd]
				wavgexactexrtlist.insert(rd+1,mld)
				wavgexactexrtlist.insert(rd+2,mld)
				rd=rd+1
			rd=rd+1
	else:
		rd=0
		while rd<(len(mlistname)):
			if str(prodname[rd][len(prodname[rd])-1])=='r':
				mld=mlistname[rd]
				mlistname.insert(rd+1,mld)
				mld=precname[rd]
				precname.insert(rd+1,mld)
				mld=precformula[rd]
				precformula.insert(rd+1,mld)
				mld=precadduct[rd]
				precadduct.insert(rd+1,mld)
				mld=precmz[rd]
				precmz.insert(rd+1,mld)
				mld=precchrg[rd]
				precchrg.insert(rd+1,mld)
				#change prodname, prodMz and prodformula
				mld=prodname[rd].replace('precursor','epoxide')
				prodname.insert(rd+1,mld)
				mld=float(prodmz[rd])+imass[4]
				prodmz.insert(rd+1,mld)
				# begin read precursor sum formula and # begin edit epoxide sum formula 
				e=prodformula[rd]
				#print(e)
				#print(e[0])
				clist=[]
				hlist=[]
				dlist=[]
				nlist=[]
				olist=[]
				plist=[]
				ilist=[]
				i=0
				ca=0
				ha=0
				da=0
				na=0
				oa=0
				pa=0
				ia=0
				while i<len(e):
					if e[i]=='H':
						if e[i+1]=="'":
							ha=0
						else:
							ca=0
					#if e[i]=='D':
					#	ha=0		
					if e[i]=='N':
						ha=0
						da=0
					if e[i]=='O':
						ha=0
						da=0
						na=0
					if e[i]=='P':
						ha=0
						da=0
						na=0
						oa=0
					if e[i]=='I':
						ha=0
						da=0
						na=0
						oa=0
						pa=0
					if ca==1:
						clist.append(e[i])
					if ha==1:
						hlist.append(e[i])
					if da==1:
						dlist.append(e[i])
					if na==1:
						nlist.append(e[i])
					if oa==1:
						olist.append(e[i])
					if pa==1:
						plist.append(e[i])
					if ia==1:
						ilist.append(e[i])
					if e[i]=='C':
						ca=1
					if e[i]=='H':
						if e[i+1]=="'":
							ca=0
							ha=0
							da=1
							i=i+1
						else:
							ca=0
							ha=1
					#if e[i]=='D':
					#	ca=0
					#	ha=0
					#	da=1		
					if e[i]=='N':
						ha=0
						da=0
						na=1
						if e[i+1]=='O':
							nlist.append('1')
							na=0
					if e[i]=='O':
						ha=0
						da=0
						na=0
						oa=1
						if (i+1)<len(e):
							if e[i+1]=='P':
								olist.append('1')
								oa=0
						else:
							olist.append('1')
							oa=0					
					if e[i]=='P':
						da=0
						na=0
						oa=0
						pa=1
						if (i+1)<len(e):
							if e[i+1]=='I':
								plist.append('1')
								pa=0
						else:
							plist.append('1')
							pa=0
					if e[i]=='I':
						da=0
						na=0
						oa=0
						pa=0
						ia=1
						if i==(len(e)-1):
							ilist.append('1')
							ia=0
					i=i+1
				#print(clist)
				#print(hlist)
				#print(dlist)
				#print(nlist)
				#print(olist)
				#print(plist)
				if len(clist)==0:
					cn=0
				if len(hlist)==0:
					hn=0
				if len(dlist)==0:
					dn=0	
				if len(nlist)==0:
					nn=0
				if len(olist)==0:
					on=0
				if len(plist)==0:
					pn=0
				if len(ilist)==0:
					iodon=0
				if len(clist)==1:
					cn=int(clist[0])
				if len(clist)==2:
					cn=10*int(clist[0])+int(clist[1])
				if len(clist)==3:
					cn=100*int(clist[0])+10*int(clist[1])+int(clist[2])
				if len(hlist)==1:
					hn=int(hlist[0])
				if len(hlist)==2:
					hn=10*int(hlist[0])+int(hlist[1])
				if len(hlist)==3:
					hn=100*int(hlist[0])+10*int(hlist[1])+int(hlist[2])
				if len(dlist)==1:
					dn=int(dlist[0])
				if len(dlist)==2:
					dn=10*int(dlist[0])+int(dlist[1])
				if len(dlist)==3:
					dn=100*int(dlist[0])+10*int(dlist[1])+int(dlist[2])
				if len(nlist)==1:
					nn=int(nlist[0])
				if len(nlist)==2:
					nn=10*int(nlist[0])+int(nlist[1])
				if len(nlist)==3:
					nn=100*int(nlist[0])+10*int(nlist[1])+int(nlist[2])
				if len(olist)==1:
					on=int(olist[0])
				if len(olist)==2:
					on=10*int(olist[0])+int(olist[1])
				if len(olist)==3:
					on=100*int(olist[0])+10*int(olist[1])+int(olist[2])
				if len(plist)==1:
					pn=int(plist[0])
				if len(plist)==2:
					pn=10*int(plist[0])+int(plist[1])
				if len(plist)==3:
					pn=100*int(plist[0])+10*int(plist[1])+int(plist[2])
				if len(ilist)==1:
					iodon=int(ilist[0])
				if len(ilist)==2:
					iodon=10*int(ilist[0])+int(ilist[1])
				if len(ilist)==3:
					iodon=100*int(ilist[0])+10*int(ilist[1])+int(ilist[2])		# end read precursor sum formula
				onepoxide=on+1		# add O to create sum formula of epoxide precursor
				decoy='C'+str(cn)+'H'+str(hn)+"H'"+str(dn)+'N'+str(nn)+'O'+str(onepoxide)+'P'+str(pn)+'I'+str(iodon)
				prodformula.insert(rd+1,decoy)		# end edit sum formula of epoxide and insert in list
				mld=prodadduct[rd]
				prodadduct.insert(rd+1,mld)
				mld=prodchrg[rd]
				prodchrg.insert(rd+1,mld)
				mld=explicitrt[rd]
				explicitrt.insert(rd+1,mld)
				mld=exrtwindow[rd]
				exrtwindow.insert(rd+1,mld)
				mld=wcutlist[rd]
				wcutlist.insert(rd+1,mld)
				mld=wavgexactexrtlist[rd]
				wavgexactexrtlist.insert(rd+1,mld)
				rd=rd+1
			rd=rd+1

	### end add epoxide (+O transition of precursor) 	# begin add -H2 transition
	#################################################################################################################################################
	# data for output is in mlistname, precname, ...
	# gather output rows sequentially in segmlistname, segprecname, ... 
	segmlistname=segmlistname+mlistname
	segprecname=segprecname+precname
	segprecformula=segprecformula+precformula
	segprecadduct=segprecadduct+precadduct
	segprecmz=segprecmz+precmz
	segprecchrg=segprecchrg+precchrg
	segprodname=segprodname+prodname
	segprodformula=segprodformula+prodformula
	segprodadduct=segprodadduct+prodadduct
	segprodmz=segprodmz+prodmz
	segprodchrg=segprodchrg+prodchrg
	segexplicitrt=segexplicitrt+explicitrt
	segexrtwindow=segexrtwindow+exrtwindow
	segfmlistname=segfmlistname+fmlistname
	segfprecname=segfprecname+fprecname
	segfprecformula=segfprecformula+fprecformula
	segfprecadduct=segfprecadduct+fprecadduct
	segfprecmz=segfprecmz+fprecmz
	segfprecchrg=segfprecchrg+fprecchrg
	segfprodname=segfprodname+fprodname
	segfprodformula=segfprodformula+fprodformula
	segfprodadduct=segfprodadduct+fprodadduct
	segfprodmz=segfprodmz+fprodmz
	segfprodchrg=segfprodchrg+fprodchrg
	segfexplicitrt=segfexplicitrt+fexplicitrt
	segfexrtwindow=segfexrtwindow+fexrtwindow
	# lists concatenated
	# redefine segr and segs, segrun=1 if there are more rows to process
	#segs=segmentsize
	go=0
	if len(writelist[1])>(segs+1):
		segrun=1
		segr=segs+1
		segs=segs+segmentsize
		if len(writelist[1])>(segs+1):
			if str(writelist[1][segs])==str(writelist[1][segs+1]):
				go=1
			while go==1:
				if str(writelist[1][segs])==str(writelist[1][segs+1]):
					go=1
					segs=segs+1
				else:
					go=0
		else:
			segs=len(writelist[1])-1
	else:
		segrun=0


# rename after loop for further processing
mlistname=segmlistname
precname=segprecname
precformula=segprecformula
precadduct=segprecadduct
precmz=segprecmz
precchrg=segprecchrg
prodname=segprodname
prodformula=segprodformula
prodadduct=segprodadduct
prodmz=segprodmz
prodchrg=segprodchrg
explicitrt=segexplicitrt
exrtwindow=segexrtwindow
fmlistname=segfmlistname
fprecname=segfprecname
fprecformula=segfprecformula
fprecadduct=segfprecadduct
fprecmz=segfprecmz
fprecchrg=segfprecchrg
fprodname=segfprodname
fprodformula=segfprodformula
fprodadduct=segfprodadduct
fprodmz=segfprodmz
fprodchrg=segfprodchrg
fexplicitrt=segfexplicitrt
fexrtwindow=segfexrtwindow
# begin save as csv 
aa=0
if aa==0:
	# begin save to csv file
	toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 
	'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']
	writelist=[]
	writelist.append(mlistname)
	writelist.append(precname)
	writelist.append(precformula)
	writelist.append(precadduct)
	writelist.append(precmz)
	writelist.append(precchrg)
	writelist.append(prodname)
	writelist.append(prodformula)
	writelist.append(prodadduct)
	writelist.append(prodmz)
	writelist.append(prodchrg)
	writelist.append(explicitrt) # or exrt?
	writelist.append(exrtwindow)
	#writelist.append(wcutlist)
	#writelist.append(wavgexactexrtlist)
	#print('writelist created')
	transitionresultsdf=pd.DataFrame(writelist).transpose()
	#print('Transposed')
	transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
	#print('Transposed and DataFrame created')
	after=datetime.datetime.now()
	after=str(after)
	#today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_8_'+fourlettcode+'_'
	filename='jpmlipidomics_vpw20_4_rank1_2nd_filter.csv'
	transitionresultsdf.to_csv(filename, index=False)
	nrows=len(mlistname)
	print('Transition list is saved as jpmlipidomics_vpw20_4_rank_1_2nd_filter.csv (%d rows)' % nrows)
	#afterall=datetime.datetime.now()
	#dt=afterall-beforeall
	#print('Calculation time:')
	#print(dt)
	# end save as csv 
	# begin save to csv file
	if len(fmlistname)>0:
		toprow=['MoleculeGroup', 'PrecursorName', 'PrecursorFormula', 'PrecursorAdduct', 'PrecursorMz', 'PrecursorCharge', 'ProductName', 
		'ProductFormula', 'ProductAdduct', 'ProductMz', 'ProductCharge', 'PrecursorRT', 'PrecursorRTWindow']
		writelist=[]
		writelist.append(fmlistname)
		writelist.append(fprecname)
		writelist.append(fprecformula)
		writelist.append(fprecadduct)
		writelist.append(fprecmz)
		writelist.append(fprecchrg)
		writelist.append(fprodname)
		writelist.append(fprodformula)
		writelist.append(fprodadduct)
		writelist.append(fprodmz)
		writelist.append(fprodchrg)
		writelist.append(fexplicitrt) # or exrt?
		writelist.append(fexrtwindow)
		#writelist.append(fwcutlist)
		#writelist.append(fwavgexactexrtlist)
		#print('writelist created')
		transitionresultsdf=pd.DataFrame(writelist).transpose()
		#print('Transposed')
		transitionresultsdf.columns=[toprow[0],toprow[1],toprow[2],toprow[3],toprow[4],toprow[5],toprow[6],toprow[7],toprow[8],toprow[9],toprow[10],toprow[11],toprow[12]]
		#print('Transposed and DataFrame created')
		after=datetime.datetime.now()
		after=str(after)
		#today=after[0]+after[1]+after[2]+after[3]+'_'+after[5]+after[6]+'_'+after[8]+after[9]+'_8_'+fourlettcode+'_'
		filename='jpmlipidomics_vpw20_4_rank2_2nd_filter.csv'
		transitionresultsdf.to_csv(filename, index=False)
		print('Backup unconfirmed species - transition list is saved as jpmlipidomics_vpw20_4_rank_2_2nd_filter.csv')
	afterall=datetime.datetime.now()
	dt=afterall-beforeall
	print('Calculation time (h:mm:ss) is:')
	print(dt)
	if len(fmlistname)>0:
		print('The next step requires a critical assessment of the species contained in the lists in Skyline.')
		print('Delete false positives from list rank 1 (incl. potential duplicates) and copy positives (falsely in rank 2 assigned correct species) from list rank 2 into list rank 1.')
	else:
		print('The next step requires a critical assessment of the species contained in the list in Skyline.')
	# end save as csv


	



